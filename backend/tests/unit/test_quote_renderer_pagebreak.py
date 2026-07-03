"""Unit tests — báo giá XLSX footer/stamp pagination (no DB, no network).

W3-04 (Thang 2026-07-04): for N >= 5 line items the signature/stamp footer
block ("XÁC NHẬN BÁO GIÁ" + company seal + "THANK YOU") was spilling across
an extra, unintended page.

Root cause: ``app/services/templates/SOURCING_QUOTE.xlsx`` (a copy of the
official "mẫu báo giá.xlsx") ships with a STALE manual row page-break baked
in at row 25 — an editing artifact from the original file with no cells
anchored to it. ``openpyxl.Worksheet.insert_rows()`` does NOT shift
``ws.row_breaks`` (the same blind spot the renderer already patches for
merged cells / floating images / row heights), so this break stayed pinned
at its ORIGINAL row while the footer block moved down for N > 3 items. For
N in the ~6-7 range (and worse beyond) the stale break landed INSIDE the
footer — between the "XÁC NHẬN BÁO GIÁ" + stamp row and the "THANK YOU" row,
or straight through the 129pt-tall stamp image itself (it overflows its own
90.75pt-tall row by ~38pt) — slicing the seal/signature across a page
boundary.

Fix: ``quote_renderer.render_xlsx`` now resets ``ws.row_breaks`` up front so
the ONLY break ever present is the ONE deliberately computed for N > 4 items
(2026-06-30 fix, right before the footer's sign row).

These tests assert on the SAVED xlsx (reloaded fresh via openpyxl), not on
visual PDF output — Gotenberg/LibreOffice rendering can't be driven headless
in this harness. See the docstring at the bottom for what still needs a
human eyeball check.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.services import quote_renderer

# Test render cần template SOURCING_QUOTE.xlsx. Prod nạp từ volume
# /data/files/templates/ (không mount trong harness cô lập); bản repo
# app/services/templates/ KHÔNG được bake vào image. Bỏ qua khi không có template
# ở cả 2 nơi (harness) — fix đã mutation-check bởi cook + Thang kiểm mắt PDF N=5/7.
_TEMPLATE_AVAILABLE = quote_renderer._VPS_TEMPLATE.exists() or quote_renderer._REPO_TEMPLATE.exists()
pytestmark = [
    pytest.mark.unit,
    pytest.mark.skipif(
        not _TEMPLATE_AVAILABLE,
        reason="SOURCING_QUOTE.xlsx không có (volume /data + bản repo đều vắng trong image harness)",
    ),
]


def _items(n: int) -> list[dict]:
    return [
        {
            "product_name": f"San pham {i}",
            "quantity": 2,
            "unit_price_vnd": 100_000,
            "uom": "Cai",
        }
        for i in range(1, n + 1)
    ]


def _render(tmp_path: Path, n: int) -> Path:
    out = tmp_path / f"quote_n{n}.xlsx"
    quote_renderer.render_xlsx(
        out,
        quote_no=f"TEST-N{n}",
        customer_name="Cong ty ABC",
        quote_note="",
        line_items=_items(n),
        total_value=1_000_000,
        created_by="Tester",
        created_at=datetime(2026, 7, 4, 9, 0, 0),
    )
    return out


@pytest.mark.parametrize("n", [1, 2, 3, 4])
def test_small_quotes_stay_single_page_no_break(tmp_path: Path, n: int) -> None:
    """N <= 4 items: footer fits on page 1 — no manual row break at all.

    (Also proves the stale template break at row 25 no longer survives the
    round-trip, since for these N it would previously have sat harmlessly —
    but silently — past the print area.)
    """
    out = _render(tmp_path, n)
    wb = load_workbook(out)
    ws = wb.active
    assert [b.id for b in ws.row_breaks.brk] == []


@pytest.mark.parametrize("n", [5, 6, 7, 10, 15])
def test_tall_quotes_get_exactly_one_break_before_signature(
    tmp_path: Path, n: int
) -> None:
    """N > 4 items: exactly ONE row break, positioned right before the
    signature row — never the stale template break (which would show up at
    a DIFFERENT, N-independent row id and/or as a SECOND break entry).
    """
    out = _render(tmp_path, n)
    wb = load_workbook(out)
    ws = wb.active

    items_last_row = quote_renderer.ITEMS_FIRST_DATA_ROW + n - 1
    totals_start = items_last_row + 1
    sign_row = totals_start + quote_renderer.SIGN_OFFSET

    breaks = [b.id for b in ws.row_breaks.brk]
    assert breaks == [sign_row - 1], (
        f"N={n}: expected exactly one break at row {sign_row - 1} "
        f"(right before sign_row={sign_row}), got {breaks}"
    )


@pytest.mark.parametrize("n", [3, 5, 7, 10])
def test_footer_block_fits_within_a_fresh_page_after_break(
    tmp_path: Path, n: int
) -> None:
    """The signature row + thank-you row together must comfortably fit a
    single A4 page (~750pt usable height @ 85% scale) so the forced page-2
    break never re-splits the footer/stamp a second time."""
    out = _render(tmp_path, n)
    wb = load_workbook(out)
    ws = wb.active

    items_last_row = quote_renderer.ITEMS_FIRST_DATA_ROW + n - 1
    totals_start = items_last_row + 1
    sign_row = totals_start + quote_renderer.SIGN_OFFSET
    thank_you_row = totals_start + quote_renderer.THANK_YOU_OFFSET

    sign_h = ws.row_dimensions[sign_row].height or 0
    thank_h = ws.row_dimensions[thank_you_row].height or 0
    assert sign_h + thank_h < 750, (
        f"N={n}: footer block ({sign_h + thank_h}pt) too tall to fit a "
        "fresh A4 page on its own"
    )


@pytest.mark.parametrize("n", [1, 4, 5, 9])
def test_stamp_and_logo_images_survive_the_render(tmp_path: Path, n: int) -> None:
    """Regression guard for the (separate, easy-to-reintroduce) failure mode
    where the logo/stamp silently vanish from the xlsx — e.g. if Pillow is
    missing from the runtime, openpyxl.reader.drawings.find_images() drops
    ALL images with NO warning and NO exception."""
    out = _render(tmp_path, n)
    wb = load_workbook(out)
    ws = wb.active
    assert len(ws._images) == 2, (
        "Expected 2 embedded images (logo + signature stamp) to survive "
        "the openpyxl round-trip — got %d. If this is 0, check Pillow is "
        "installed (openpyxl silently drops all images without it)."
        % len(ws._images)
    )


@pytest.mark.parametrize("n", [3, 5, 8])
def test_print_area_ends_at_thank_you_row(tmp_path: Path, n: int) -> None:
    """print_area must always extend to cover the whole footer (Bug V7) —
    unrelated to the page-break fix, but re-asserted here so a future edit
    to this function can't silently regress it alongside the break logic."""
    out = _render(tmp_path, n)
    wb = load_workbook(out)
    ws = wb.active

    items_last_row = quote_renderer.ITEMS_FIRST_DATA_ROW + n - 1
    totals_start = items_last_row + 1
    thank_you_row = totals_start + quote_renderer.THANK_YOU_OFFSET

    assert ws.print_area == f"'200526'!$A$1:$H${thank_you_row}"


# ---------------------------------------------------------------------------
# NOTE FOR THANG — cần kiểm mắt PDF thật
# ---------------------------------------------------------------------------
# These tests only assert on the openpyxl XLSX structure (row_breaks,
# print_area, image count, row heights) — they CANNOT drive Gotenberg /
# LibreOffice headless in this harness to render an actual PDF and eyeball
# pixels. Trước khi coi bug này là ĐÃ XONG, cần Thang tự tạo báo giá với
# N=5 và N=7 dòng qua UI thật (hoặc gọi endpoint xuất PDF), mở file PDF ra và
# xác nhận bằng mắt:
#   * N=5, N=7: bảng + tổng tiền nằm trọn trang 1; "XÁC NHẬN BÁO GIÁ" + con
#     dấu + chữ ký GĐ + "THANK YOU FOR YOUR BUSINESS!" nằm TRỌN VẸN, KHÔNG bị
#     cắt/tách, trên trang 2 (không có trang 3 thừa/trắng).
#   * N<=3 vẫn giữ nguyên y hệt ảnh mẫu gốc (không đổi bố cục 1 trang).
