#!/usr/bin/env python3
"""
Song Châu ERP — Import TOÀN BỘ dữ liệu từ 501 file Excel (OneDrive) vào PostgreSQL.

Script mở rộng đầy đủ 15 nhóm file → 15+ bảng database.
Mỗi entry trong IMPORT_MAP bao gồm: files, sheets, table, column_mapping, insert_sql (UPSERT).
Hỗ trợ multiline headers, Vietnamese date formats, VN number formats.
Tính source_hash (SHA256) cho change detection, log vào etl_sync_log.

Usage:
    python scripts/import_all_data.py --source /path/to/onedrive
    python scripts/import_all_data.py --source ./data --dry-run
    python scripts/import_all_data.py --source ./data --table bqms_rfq --verbose
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Logging — Vietnamese messages
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("import_all_data")

# ---------------------------------------------------------------------------
# Database DSN — override bằng biến môi trường DATABASE_URL
# ---------------------------------------------------------------------------

DSN = os.getenv(
    "DATABASE_URL",
    "postgresql://scadmin:SC2026_ERP_Pr0d_X9k2mQ7wR4@postgres:5432/songchau_erp",
).replace("+asyncpg", "")

# ---------------------------------------------------------------------------
# Hằng số
# ---------------------------------------------------------------------------

DATA_SOURCE = "onedrive_sync"

# ---------------------------------------------------------------------------
# IMPORT_MAP — 15 nhóm file đầy đủ
# ---------------------------------------------------------------------------

IMPORT_MAP: list[dict[str, Any]] = [
    # ── 1. BQMS RFQ — Thống kê hỏi hàng ──
    {
        "files": ["Thong ke hoi hang BQMS.xlsx"],
        "sheet": "TONG HOP BQMS",
        "table": "bqms_rfq",
        "columns": {
            "Ngày": "inquiry_date",
            "Người phụ trách": "person_in_charge_name",
            "RFQ No.": "rfq_number",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "Maker": "maker",
            "Số lượng dự kiến": "expected_qty",
            "Giá nhập\nRMB": "purchase_price_rmb",
            "Giá nhập\nVND": "purchase_price_vnd",
            "Giá báo cho AMA": "quoted_price_ama",
            "Giá báo cho BQMS V1": "quoted_price_bqms_v1",
            "V2": "quoted_price_bqms_v2",
            "V3": "quoted_price_bqms_v3",
            "Ghi chú": "notes",
            "NCC": "supplier_name",
            "Kết quả\n(Y/N)": "result",
            "Báo cáo": "report",
        },
        "column_order": [
            "inquiry_date", "person_in_charge_name", "rfq_number", "bqms_code",
            "specification", "maker", "expected_qty",
            "purchase_price_rmb", "purchase_price_vnd",
            "quoted_price_ama", "quoted_price_bqms_v1",
            "quoted_price_bqms_v2", "quoted_price_bqms_v3",
            "notes", "supplier_name", "result", "report",
        ],
        "insert_sql": """
            INSERT INTO bqms_rfq (
                inquiry_date, person_in_charge_name, rfq_number, bqms_code,
                specification, maker, expected_qty,
                purchase_price_rmb, purchase_price_vnd,
                quoted_price_ama, quoted_price_bqms_v1,
                quoted_price_bqms_v2, quoted_price_bqms_v3,
                notes, supplier_name, result, report,
                data_source
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13,
                $14, $15, COALESCE($16::rfq_result, 'pending'), $17,
                '{data_source}'
            )
            ON CONFLICT (rfq_number, bqms_code)
                DO UPDATE SET
                    inquiry_date = EXCLUDED.inquiry_date,
                    person_in_charge_name = EXCLUDED.person_in_charge_name,
                    specification = EXCLUDED.specification,
                    maker = EXCLUDED.maker,
                    expected_qty = EXCLUDED.expected_qty,
                    purchase_price_rmb = EXCLUDED.purchase_price_rmb,
                    purchase_price_vnd = EXCLUDED.purchase_price_vnd,
                    supplier_name = EXCLUDED.supplier_name,
                    result = EXCLUDED.result,
                    updated_at = NOW()
        """,
        "conflict_key": "(rfq_number, bqms_code)",
    },
    # ── 2. BQMS Deliveries — Giao hàng ──
    {
        "files": [
            "Thong ke giao hang 2026.xlsx",
            "Thong ke giao hang 2025.xlsx",
            "Thong ke giao hang 2023-2024.xlsx",
        ],
        "sheet": "THỐNG KÊ PO",
        "table": "bqms_deliveries",
        "columns": {
            "Ngày PO": "po_date",
            "Số PO": "po_number",
            "Shipping No": "shipping_no",
            "Số QT": "quotation_no",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "SL": "quantity",
            "Đơn vị": "unit",
            "Đơn giá": "unit_price",
            "Thành tiền": "amount",
            "SEV/T": "sev_type",
            "MAIL PUR": "buyer_email",
            "TÊN NGƯỜI NHẬN": "recipient_name",
            "KHO NHẬN": "receiving_warehouse",
            "SĐT PUR": "buyer_phone",
            "TÌNH TRẠNG": "delivery_status",
            "NGÀY GIAO HÀNG": "delivery_date",
            "SL GIAO THỰC TẾ": "actual_delivered_qty",
            "THÔNG TIN GIAO HÀNG": "delivery_info",
            "CÁCH THỨC GIAO HÀNG": "delivery_method",
            "XUẤT XỨ": "country_origin",
            "TỔNG GIÁ TRỊ ĐÃ GIAO\n(VND)": "total_delivered_value_vnd",
        },
        "column_order": [
            "po_date", "po_number", "shipping_no", "quotation_no",
            "bqms_code", "specification", "quantity", "unit",
            "unit_price", "amount", "sev_type", "buyer_email",
            "recipient_name", "receiving_warehouse", "buyer_phone",
            "delivery_status", "delivery_date", "actual_delivered_qty",
            "delivery_info", "delivery_method", "country_origin",
            "total_delivered_value_vnd",
        ],
        "insert_sql": """
            INSERT INTO bqms_deliveries (
                po_date, po_number, shipping_no, quotation_no,
                bqms_code, specification, quantity, unit,
                unit_price, amount, sev_type, buyer_email,
                recipient_name, receiving_warehouse, buyer_phone,
                delivery_status, delivery_date, actual_delivered_qty,
                delivery_info, delivery_method, country_origin,
                total_delivered_value_vnd,
                data_source
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10,
                $11, $12, $13, $14, $15,
                COALESCE($16::delivery_status, 'chua_giao'),
                $17, $18, $19, $20, $21, $22,
                '{data_source}'
            )
            ON CONFLICT (po_number, shipping_no, bqms_code)
                DO UPDATE SET
                    delivery_status = EXCLUDED.delivery_status,
                    delivery_date = EXCLUDED.delivery_date,
                    actual_delivered_qty = EXCLUDED.actual_delivered_qty,
                    total_delivered_value_vnd = EXCLUDED.total_delivered_value_vnd,
                    updated_at = NOW()
        """,
        "conflict_key": "(po_number, shipping_no, bqms_code)",
    },
    # ── 3. BQMS Orders — Đặt hàng ──
    {
        "files": ["Thong ke dat hang.xlsx"],
        "sheet": "Sheet1",
        "table": "bqms_orders",
        "columns": {
            "RFQ No.": "rfq_number",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "Khách hàng": "customer_name",
            "Số lượng dự kiến": "expected_qty",
            "SL đặt hàng": "order_qty",
            "ĐVT": "unit",
            "Ngày đặt hàng": "order_date",
            "Thời hạn hiệu lực": "validity_date",
            "Trạng thái": "status",
            "SL giao": "delivered_qty",
            "Ngày giao": "delivery_date",
            "Ghi chú": "notes",
        },
        "column_order": [
            "rfq_number", "bqms_code", "specification", "customer_name",
            "expected_qty", "order_qty", "unit", "order_date",
            "validity_date", "status", "delivered_qty", "delivery_date", "notes",
        ],
        "insert_sql": """
            INSERT INTO bqms_orders (
                rfq_number, bqms_code, specification, customer_name,
                expected_qty, order_qty, unit, order_date,
                validity_date, status, delivered_qty, delivery_date, notes
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9,
                COALESCE($10, 'pending'), $11, $12, $13
            )
            ON CONFLICT (rfq_number, bqms_code)
                DO UPDATE SET
                    order_qty = EXCLUDED.order_qty,
                    status = EXCLUDED.status,
                    delivered_qty = EXCLUDED.delivered_qty,
                    delivery_date = EXCLUDED.delivery_date,
                    updated_at = NOW()
        """,
        "conflict_key": "(rfq_number, bqms_code)",
    },
    # ── 4. BQMS Samsung PO — PO từ Samsung ──
    {
        "files": ["IMV/BQMS - 2024/BQMS - PO.xlsx"],
        "sheet": "Sheet2",
        "table": "bqms_samsung_po",
        "columns": {
            "No": "_row_no",
            "P/O Date": "po_date",
            "P/O No": "po_number",
            "P/O Seq": "po_seq",
            "Request No": "request_no",
            "Request Seq.": "request_seq",
            "Process Status": "process_status",
            "Status": "confirm_status",
            "Vendor": "vendor_code",
            "Buyer": "buyer_name",
            "Company": "company",
            "Plant": "plant",
            "Spec": "specification",
            "Maker": "maker",
            "Part No": "part_no",
            "BQMS Code": "bqms_code",
            "CIS Code": "cis_code",
            "Category": "category",
            "Order Qty.": "order_qty",
            "Unit Price": "unit_price",
            "Amount": "amount",
            "Currency": "currency",
            "Recipient": "recipient_name",
            "Address": "delivery_address",
            "Delivery preferred date": "preferred_delivery_date",
            "Shipping Qty.": "shipping_qty",
            "GR Qty.": "gr_qty",
            "Invoice Qty.": "invoice_qty",
            "Remark": "remark",
            "Shipping Type": "shipping_type",
        },
        "column_order": [
            "po_date", "po_number", "po_seq", "request_no", "request_seq",
            "process_status", "confirm_status", "vendor_code", "buyer_name",
            "company", "plant", "specification", "maker", "part_no",
            "bqms_code", "cis_code", "category",
            "order_qty", "unit_price", "amount", "currency",
            "recipient_name", "delivery_address", "preferred_delivery_date",
            "shipping_qty", "gr_qty", "invoice_qty", "remark", "shipping_type",
        ],
        "insert_sql": """
            INSERT INTO bqms_samsung_po (
                po_date, po_number, po_seq, request_no, request_seq,
                process_status, confirm_status, vendor_code, buyer_name,
                company, plant, specification, maker, part_no,
                bqms_code, cis_code, category,
                order_qty, unit_price, amount, currency,
                recipient_name, delivery_address, preferred_delivery_date,
                shipping_qty, gr_qty, invoice_qty, remark, shipping_type
            ) VALUES (
                $1, $2, $3, $4, $5,
                COALESCE($6::samsung_po_process_status, 'new'), $7, $8, $9,
                $10, $11, $12, $13, $14,
                $15, $16, $17,
                $18, $19, $20, COALESCE($21::currency_code, 'VND'),
                $22, $23, $24,
                $25, $26, $27, $28, $29
            )
            ON CONFLICT (po_number)
                DO UPDATE SET
                    process_status = EXCLUDED.process_status,
                    shipping_qty = EXCLUDED.shipping_qty,
                    gr_qty = EXCLUDED.gr_qty,
                    invoice_qty = EXCLUDED.invoice_qty,
                    updated_at = NOW()
        """,
        "conflict_key": "(po_number)",
    },
    # ── 5. BQMS Raw Material PO — PO phôi ──
    {
        "files": ["TONG HOP BQMS/THEO DOI PO PHOI.xlsx"],
        "sheet": ["PO PHOI 2025", "PO PHOI 2026"],
        "table": "bqms_raw_material_po",
        "columns": {
            "Ngày PO": "po_date",
            "Số PO": "po_number",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "SL PO": "po_qty",
            "Đơn vị": "unit",
            "HÀNG SẴN": "in_stock",
            "SL\nCÒN THIẾU": "remaining_qty",
            "SL\nĐÃ GIAO": "delivered_qty",
            "PENDING": "pending",
        },
        "column_order": [
            "po_date", "po_number", "bqms_code", "specification",
            "po_qty", "unit", "in_stock", "remaining_qty",
            "delivered_qty", "pending",
        ],
        "insert_sql": """
            INSERT INTO bqms_raw_material_po (
                po_date, po_number, bqms_code, specification,
                po_qty, unit, in_stock, remaining_qty,
                delivered_qty, pending
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10
            )
            ON CONFLICT (po_number, bqms_code)
                DO UPDATE SET
                    remaining_qty = EXCLUDED.remaining_qty,
                    delivered_qty = EXCLUDED.delivered_qty,
                    pending = EXCLUDED.pending,
                    updated_at = NOW()
        """,
        "conflict_key": "(po_number, bqms_code)",
    },
    # ── 6. BQMS Material Pricing — Giá vật liệu ──
    {
        "files": ["TONG HOP BQMS/KET QUA PHOI TRUOT.xlsx"],
        "sheet": None,
        "table": "bqms_material_pricing",
        "columns": {
            "STT": "_row_no",
            "RFQ No.": "rfq_number",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "Đơn giá (VND)": "unit_price_vnd",
            "Trọng lượng (KG)": "weight_kg",
            "L": "dimension_l",
            "W": "dimension_w",
            "H": "dimension_h",
            "Type": "material_type",
            "gr/m3": "density_g_m3",
        },
        "column_order": [
            "rfq_number", "bqms_code", "specification",
            "unit_price_vnd", "weight_kg",
            "dimension_l", "dimension_w", "dimension_h",
            "material_type", "density_g_m3",
        ],
        "insert_sql": """
            INSERT INTO bqms_material_pricing (
                rfq_number, bqms_code, specification,
                unit_price_vnd, weight_kg,
                dimension_l, dimension_w, dimension_h,
                material_type, density_g_m3
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10
            )
            ON CONFLICT (rfq_number, bqms_code)
                DO UPDATE SET
                    unit_price_vnd = EXCLUDED.unit_price_vnd,
                    weight_kg = EXCLUDED.weight_kg
        """,
        "conflict_key": "(rfq_number, bqms_code)",
    },
    # ── 7. Import/Export Tracking — XNK ──
    {
        "files": ["TT XNK BQMS 2023-2026.xlsx"],
        "sheet": ["APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"],
        "table": "import_export_tracking",
        "columns": {
            "Stt": "_row_no",
            "Ngày Tháng": "tracking_date",
            "Đơn hàng": "rfq_number",
            "BMSQ": "bqms_code",
            "Tên hàng hóa": "product_name",
            "Explain for detail?": "detail_explain",
            "Loại hàng": "goods_type",
            "Maker\n업체": "maker",
            "Ghi chú": "notes",
            "Đơn vị tính": "unit_calc",
            "Số lượng": "quantity_calc",
            "Quote Deadline": "quote_deadline",
            "Ngày GD": "transaction_date",
            "Miêu tả hàng hóa": "customs_description",
            "Mã HS": "hs_code",
            "ĐVT": "unit",
            "SL": "quantity",
            "Tổng cộng USD": "total_usd",
            "Đơn giá USD": "unit_price_usd",
            "Đơn giá VND": "unit_price_vnd",
            "Bên mua": "buyer_name",
            "Bên bán": "seller_name",
            "SL Đã mua": "purchased_qty",
            "Nhà cung cấp khác": "alt_supplier",
        },
        "column_order": [
            "tracking_date", "rfq_number", "bqms_code", "product_name",
            "detail_explain", "goods_type", "maker", "notes",
            "unit_calc", "quantity_calc", "quote_deadline", "transaction_date",
            "customs_description", "hs_code", "unit", "quantity",
            "total_usd", "unit_price_usd", "unit_price_vnd",
            "buyer_name", "seller_name", "purchased_qty", "alt_supplier",
        ],
        "insert_sql": """
            INSERT INTO import_export_tracking (
                tracking_date, rfq_number, bqms_code, product_name,
                detail_explain, goods_type, maker, notes,
                unit_calc, quantity_calc, quote_deadline, transaction_date,
                customs_description, hs_code, unit, quantity,
                total_usd, unit_price_usd, unit_price_vnd,
                buyer_name, seller_name, purchased_qty, alt_supplier,
                data_source
            ) VALUES (
                $1, $2, $3, $4, $5, $6::goods_type, $7, $8,
                $9, $10, $11, $12,
                $13, $14, $15, $16,
                $17, $18, $19,
                $20, $21, $22, $23,
                '{data_source}'
            )
            ON CONFLICT (tracking_date, bqms_code, rfq_number)
                DO UPDATE SET
                    quantity = EXCLUDED.quantity,
                    total_usd = EXCLUDED.total_usd,
                    updated_at = NOW()
        """,
        "conflict_key": "(tracking_date, bqms_code, rfq_number)",
    },
    # ── 8. IMV Inquiries — Hỏi hàng IMV ──
    {
        "files": ["IMV/Thong ke hoi hang - update 240424.xlsx"],
        "sheet": "TONG HOP",
        "table": "imv_inquiries",
        "columns": {
            "Tên KH": "customer_name",
            "Ng phụ\n trách": "person_in_charge_name",
            "Model": "model",
            "Tên sp": "product_name",
            "Maker": "maker",
            "Ngày hỏi giá\n(m/d/y)": "inquiry_date",
            "Giá nhập\nYên Nhật": "purchase_price_jpy",
            "Giá nhập\nUSD": "purchase_price_usd",
            "Giá nhập Won": "purchase_price_krw",
            "Giá nhập RMB": "purchase_price_rmb",
            "Giá nhập\nvnd": "purchase_price",
            "Giá bán": "selling_price",
            "Số lượng": "quantity",
            "Thuế xuất": "tax_rate",
            "HS Code": "hs_code",
            "Cân nặng": "weight_kg",
            "Ghi chú": "notes",
            "Hệ số": "coefficient",
            "Nhà cung cấp": "supplier_name",
            "Hình ảnh": "image_path",
            "Tỷ giá": "exchange_rate",
        },
        "column_order": [
            "customer_name", "person_in_charge_name", "model", "product_name",
            "maker", "inquiry_date",
            "purchase_price", "selling_price", "quantity",
            "tax_rate", "hs_code", "weight_kg", "notes",
            "coefficient", "supplier_name", "image_path", "exchange_rate",
        ],
        "insert_sql": """
            INSERT INTO imv_inquiries (
                customer_name, person_in_charge_name, model, product_name,
                maker, inquiry_date,
                purchase_price, selling_price, quantity,
                tax_rate, hs_code, weight_kg, notes,
                coefficient, supplier_name, image_path, exchange_rate,
                data_source
            ) VALUES (
                $1, $2, $3, $4, $5, $6,
                $7, $8, $9,
                $10, $11, $12, $13,
                $14, $15, $16, $17,
                '{data_source}'
            )
            ON CONFLICT (customer_name, product_name, maker, inquiry_date)
                DO UPDATE SET
                    purchase_price = EXCLUDED.purchase_price,
                    selling_price = EXCLUDED.selling_price,
                    updated_at = NOW()
        """,
        "conflict_key": "(customer_name, product_name, maker, inquiry_date)",
    },
    # ── 9. IMV Consolidated — Tổng hợp IMV ──
    {
        "files": ["IMV/Thong ke hoi hang - update 240424.xlsx"],
        "sheet": "Tổng hợp IMV",
        "table": "imv_consolidated",
        "columns": {
            "Báo Giá Số": "quotation_no",
            "Trạng Thái": "status",
            "Người Phụ Trách Mua Hàng": "purchaser_name",
            "Khách Hàng": "customer_name",
            "Cơ Sở Khách Hàng": "customer_branch",
            "Mã Hàng Khách Hàng": "customer_item_code",
            "Mã Hàng": "product_code",
            "Báo Giá Yêu Cầu Số": "rfq_number",
            "Tên Sản Phẩm": "product_name",
            "Kiểu Mẫu": "model",
            "Quy Cách": "specification",
            "Nhà Sản Xuất": "maker",
            "ĐVT": "unit",
            "Số Lượng Đặt Hàng Dự Kiến": "expected_order_qty",
            "Số PO Năm Trước": "prev_year_po_count",
            "Ngày Lên YCBG": "request_date",
            "Hạn BG": "quote_deadline",
            "MOQ": "moq",
            "Đơn giá báo": "quoted_price",
            "Đơn giá nhập": "purchase_price",
            "Chênh lệch": "price_diff",
            "Trạng thái P/O": "po_status",
            "SL P/O": "po_qty",
            "Thành tiền P/O": "po_amount",
            "Lợi nhuận": "profit",
        },
        "column_order": [
            "quotation_no", "status", "purchaser_name", "customer_name",
            "customer_branch", "customer_item_code", "product_code",
            "rfq_number", "product_name", "model", "specification", "maker",
            "unit", "expected_order_qty", "prev_year_po_count",
            "request_date", "quote_deadline", "moq",
            "quoted_price", "purchase_price", "price_diff",
            "po_status", "po_qty", "po_amount", "profit",
        ],
        "insert_sql": """
            INSERT INTO imv_consolidated (
                quotation_no, status, purchaser_name, customer_name,
                customer_branch, customer_item_code, product_code,
                rfq_number, product_name, model, specification, maker,
                unit, expected_order_qty, prev_year_po_count,
                request_date, quote_deadline, moq,
                quoted_price, purchase_price, price_diff,
                po_status, po_qty, po_amount, profit,
                data_source
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7,
                $8, $9, $10, $11, $12,
                $13, $14, $15,
                $16, $17, $18,
                $19, $20, $21,
                $22, $23, $24, $25,
                '{data_source}'
            )
            ON CONFLICT (quotation_no, product_code)
                DO UPDATE SET
                    status = EXCLUDED.status,
                    po_status = EXCLUDED.po_status,
                    po_qty = EXCLUDED.po_qty,
                    po_amount = EXCLUDED.po_amount,
                    profit = EXCLUDED.profit,
                    updated_at = NOW()
        """,
        "conflict_key": "(quotation_no, product_code)",
    },
    # ── 10. IMV Purchase Orders — PO IMV ──
    {
        "files": ["IMV/1.PO IMV 2025.xlsx"],
        "sheet": "SONG CHAU",
        "table": "imv_purchase_orders",
        "columns": {
            "STT": "_row_no",
            "Ngày PO": "po_date",
            "Số PO": "po_number",
            "Mã hàng": "product_code",
            "Tên hàng": "product_name",
            "ĐVT": "unit",
            "SL YC": "requested_qty",
            "Đơn giá": "unit_price",
            "Thành tiền": "amount",
            "VAT": "vat_amount",
            "Tổng": "total_amount",
            "Đơn vị mua hàng": "purchasing_dept",
            "SL đã giao": "delivered_qty",
            "Ngày giao thực tế": "actual_delivery_date",
            "Còn thiếu": "remaining_qty",
            "NCC": "supplier_name",
            "Chứng từ": "document_ref",
            "Ghi chú": "notes",
        },
        "column_order": [
            "po_date", "po_number", "product_code", "product_name",
            "unit", "requested_qty", "unit_price", "amount",
            "vat_amount", "total_amount", "purchasing_dept",
            "delivered_qty", "actual_delivery_date", "remaining_qty",
            "supplier_name", "document_ref", "notes",
        ],
        "insert_sql": """
            INSERT INTO imv_purchase_orders (
                po_date, po_number, product_code, product_name,
                unit, requested_qty, unit_price, amount,
                vat_amount, total_amount, purchasing_dept,
                delivered_qty, actual_delivery_date, remaining_qty,
                supplier_name, document_ref, notes,
                data_source
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8,
                $9, $10, $11,
                $12, $13, $14,
                $15, $16, $17,
                '{data_source}'
            )
            ON CONFLICT (po_number, product_code)
                DO UPDATE SET
                    delivered_qty = EXCLUDED.delivered_qty,
                    actual_delivery_date = EXCLUDED.actual_delivery_date,
                    remaining_qty = EXCLUDED.remaining_qty,
                    updated_at = NOW()
        """,
        "conflict_key": "(po_number, product_code)",
    },
    # ── 11. Customer Contacts — Danh bạ ──
    {
        "files": ["Thong ke giao hang 2026.xlsx"],
        "sheet": "DANH BẠ",
        "table": "customer_contacts",
        "columns": {
            "Mail": "email",
            "Tên": "full_name",
            "Thông tin giao hàng": "delivery_info",
            "SĐT": "phone",
        },
        "column_order": [
            "email", "full_name", "delivery_info", "phone",
        ],
        "insert_sql": """
            INSERT INTO customer_contacts (
                customer_id, email, full_name, delivery_info, phone
            ) VALUES (
                (SELECT id FROM customers LIMIT 1),
                $1, $2, $3, $4
            )
            ON CONFLICT (customer_id, email)
                DO UPDATE SET
                    full_name = EXCLUDED.full_name,
                    delivery_info = EXCLUDED.delivery_info,
                    phone = EXCLUDED.phone,
                    updated_at = NOW()
        """,
        "conflict_key": "(customer_id, email)",
    },
    # ── 12. Revenue Invoices — Doanh thu ──
    {
        "files": [
            "Bảng theo dõi doanh thu SC.2025.xlsx",
            "Bang theo doi doanh thu SC.2025.xlsx",
        ],
        "sheet": ["T1.25", "T2.25", "T3.25", "T4.25", "T5.25", "T6.25",
                   "T7.25", "T8.25", "T9.25", "T10.25", "T11.25", "T12.25"],
        "table": "revenue_invoices",
        "columns": {
            "STT": "_row_no",
            "Số hóa đơn": "invoice_number",
            "Ngày hóa đơn": "invoice_date",
            "Tên KH": "customer_name",
            "Tên hàng": "product_name",
            "ĐVT": "unit",
            "SL": "quantity",
            "Đơn giá": "unit_price",
            "TT": "amount",
            "Thuế suất": "tax_rate",
            "Thuế GTGT": "vat_amount",
            "Tổng": "total_amount",
            "Số PO": "po_number",
            "Giá mua": "purchase_price",
            "VAT": "purchase_vat",
            "Vận chuyển": "shipping_cost",
            "COM": "commission",
            "Mua HĐ": "invoice_buying",
            "Hải quan": "customs_fee",
            "Thuế XK": "export_tax",
            "CP khác": "other_costs",
            "Lợi nhuận": "profit",
        },
        "column_order": [
            "invoice_number", "invoice_date", "customer_name", "product_name",
            "unit", "quantity", "unit_price", "amount",
            "tax_rate", "vat_amount", "total_amount",
            "po_number", "purchase_price", "purchase_vat",
            "shipping_cost", "commission", "invoice_buying",
            "customs_fee", "export_tax", "other_costs", "profit",
        ],
        "insert_sql": """
            INSERT INTO revenue_invoices (
                invoice_number, invoice_date, customer_name, product_name,
                unit, quantity, unit_price, amount,
                tax_rate, vat_amount, total_amount,
                po_number, purchase_price, purchase_vat,
                shipping_cost, commission, invoice_buying,
                customs_fee, export_tax, other_costs, profit,
                data_source
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8,
                $9, $10, $11,
                $12, $13, $14,
                $15, $16, $17,
                $18, $19, $20, $21,
                '{data_source}'
            )
            ON CONFLICT (invoice_number, product_name)
                DO UPDATE SET
                    total_amount = EXCLUDED.total_amount,
                    profit = EXCLUDED.profit,
                    updated_at = NOW()
        """,
        "conflict_key": "(invoice_number, product_name)",
    },
    # ── 13. Products — Sản phẩm (Samsung categories + ITEM_SONG CHAU) ──
    {
        "files": [
            "Samsung - categories.xlsx",
            "ITEM_SONG CHAU.xlsx",
        ],
        "sheet": None,
        "table": "products",
        "columns": {
            "BQMS Code": "bqms_code",
            "BQMS code": "bqms_code",
            "Mã hàng": "bqms_code",
            "Item Code": "bqms_code",
            "Tên hàng": "product_name",
            "Product Name": "product_name",
            "Tên sản phẩm": "product_name",
            "Spec": "specification",
            "Specification": "specification",
            "Quy cách": "specification",
            "Maker": "maker",
            "Nhà sản xuất": "maker",
            "Category": "category",
            "Loại": "category",
            "ĐVT": "unit",
            "Unit": "unit",
            "Xuất xứ": "country_origin",
            "Country": "country_origin",
        },
        "column_order": [
            "bqms_code", "product_name", "specification", "maker",
            "category", "unit", "country_origin",
        ],
        "insert_sql": """
            INSERT INTO products (
                bqms_code, product_name, specification, maker,
                category, unit, country_origin, is_active
            ) VALUES (
                $1, $2, $3, $4, $5, COALESCE($6, 'EA'), $7, true
            )
            ON CONFLICT (bqms_code)
                DO UPDATE SET
                    product_name = COALESCE(EXCLUDED.product_name, products.product_name),
                    specification = COALESCE(EXCLUDED.specification, products.specification),
                    maker = COALESCE(EXCLUDED.maker, products.maker),
                    category = COALESCE(EXCLUDED.category, products.category),
                    updated_at = NOW()
        """,
        "conflict_key": "(bqms_code)",
    },
    # ── 14. Exchange Rates — Tỷ giá ──
    {
        "files": ["TT XNK BQMS 2023-2026.xlsx"],
        "sheet": "TGUSD",
        "table": "exchange_rates",
        "columns": {
            "date": "rate_date",
            "rate": "rate",
        },
        "column_order": [
            "rate_date", "rate",
        ],
        "insert_sql": """
            INSERT INTO exchange_rates (
                rate_date, from_currency, to_currency, rate, source
            ) VALUES (
                $1, 'USD', 'VND', $2, '{data_source}'
            )
            ON CONFLICT (rate_date, from_currency, to_currency, rate_type)
                DO UPDATE SET
                    rate = EXCLUDED.rate
        """,
        "conflict_key": "(rate_date, from_currency, to_currency, rate_type)",
    },
    # ── 15. BQMS Won Quotations — Trúng BG ──
    {
        "files": ["Thong ke hoi hang BQMS.xlsx"],
        "sheet": "TRUNG BG",
        "table": "bqms_won_quotations",
        "columns": {
            "Người phụ trách": "person_in_charge_name",
            "RFQ No.": "rfq_number",
            "BQMS code": "bqms_code",
            "Description": "description",
            "Spec": "specification",
            "Số lượng dự kiến": "quantity",
            "Unit": "unit",
            "giá PO": "po_price",
            "hạn PO": "po_deadline",
            "Ghi chú": "notes",
            "NCC": "supplier_name",
            "HS code": "hs_code",
            "Miêu tả hàng hóa": "goods_description",
        },
        "column_order": [
            "person_in_charge_name", "rfq_number", "bqms_code",
            "description", "specification", "quantity", "unit",
            "po_price", "po_deadline", "notes",
            "supplier_name", "hs_code", "goods_description",
        ],
        "insert_sql": """
            INSERT INTO bqms_won_quotations (
                person_in_charge_name, rfq_number, bqms_code,
                description, specification, quantity, unit,
                po_price, po_deadline, notes,
                supplier_name, hs_code, goods_description
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13
            )
            ON CONFLICT (rfq_number, bqms_code)
                DO UPDATE SET
                    po_price = EXCLUDED.po_price,
                    po_deadline = EXCLUDED.po_deadline,
                    supplier_name = EXCLUDED.supplier_name
        """,
        "conflict_key": "(rfq_number, bqms_code)",
    },
]


# ---------------------------------------------------------------------------
# Value parsing helpers
# ---------------------------------------------------------------------------

DATE_FIELDS = {
    "inquiry_date", "po_date", "delivery_date", "order_date",
    "preferred_delivery_date", "deadline", "validity_date",
    "rate_date", "transaction_date", "quote_deadline",
    "tracking_date", "request_date", "actual_delivery_date",
    "invoice_date", "po_deadline",
}

NUMERIC_FIELDS = {
    "expected_qty", "quantity", "order_qty", "unit_price", "amount",
    "purchase_price_rmb", "purchase_price_vnd", "purchase_price",
    "purchase_price_jpy", "purchase_price_usd", "purchase_price_krw",
    "quoted_price_ama", "quoted_price_bqms_v1",
    "quoted_price_bqms_v2", "quoted_price_bqms_v3",
    "actual_delivered_qty", "delivered_qty",
    "shipping_qty", "gr_qty", "remaining_qty",
    "weight_kg", "total_delivered_value_vnd",
    "po_qty", "requested_qty", "vat_amount", "total_amount",
    "invoice_qty", "unit_price_vnd", "unit_price_usd",
    "total_usd", "quantity_calc", "purchased_qty",
    "rate", "selling_price", "tax_rate",
    "coefficient", "exchange_rate",
    "expected_order_qty", "prev_year_po_count", "moq",
    "quoted_price", "price_diff", "po_qty", "po_amount", "profit",
    "purchase_vat", "shipping_cost", "commission", "invoice_buying",
    "customs_fee", "export_tax", "other_costs",
    "po_price",
    "dimension_l", "dimension_w", "dimension_h", "density_g_m3",
}

BOOLEAN_FIELDS = {"in_stock", "pending"}

RESULT_MAP = {
    "y": "won", "yes": "won", "trúng": "won", "trung": "won",
    "n": "lost", "no": "lost", "thua": "lost",
    "hủy": "cancelled", "huy": "cancelled",
    "cancel": "cancelled", "cancelled": "cancelled",
}

DELIVERY_STATUS_MAP = {
    "chưa giao": "chua_giao", "chua giao": "chua_giao",
    "đang giao": "dang_giao", "dang giao": "dang_giao",
    "đã giao": "da_giao", "da giao": "da_giao",
    "giao một phần": "giao_mot_phan",
    "giao 1 phan": "giao_mot_phan", "giao mot phan": "giao_mot_phan",
}

GOODS_TYPE_MAP = {
    "gia công": "gia_cong", "gia cong": "gia_cong",
    "thương mại": "thuong_mai", "thuong mai": "thuong_mai",
}

SAMSUNG_PO_STATUS_MAP = {
    "new": "new", "mới": "new",
    "confirmed": "confirmed", "đã xác nhận": "confirmed",
    "unconfirmed": "unconfirmed",
    "shipped": "shipped", "đã gửi hàng": "shipped",
    "received": "received", "đã nhận": "received",
    "invoiced": "invoiced",
    "closed": "closed", "đóng": "closed",
}


def parse_date(value: Any) -> date | None:
    """Chuyển đổi nhiều format ngày tháng từ Excel sang date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            base = date(1899, 12, 30)
            return base + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    s = str(value).strip()
    if not s or s == "-" or s.lower() == "n/a":
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d",
                "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value: Any) -> float | None:
    """Chuyển đổi số từ Excel, hỗ trợ format VN (dấu chấm phân cách hàng nghìn)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        import math
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    s = str(value).strip()
    # Xóa ký tự không phải số
    s = s.replace(",", "").replace(" ", "").replace("\xa0", "")
    if not s or s == "-" or s.lower() == "n/a":
        return None
    # Format VN: 1.234.567 → 1234567 (dấu chấm phân cách nghìn)
    # Nhưng 1.5 → 1.5 (dấu chấm thập phân)
    if re.match(r'^\d{1,3}(\.\d{3})+$', s):
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_boolean(value: Any) -> bool | None:
    """Chuyển đổi boolean từ Excel."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("x", "yes", "y", "có", "co", "true", "1"):
        return True
    if s in ("", "-", "no", "n", "không", "khong", "false", "0"):
        return False
    return None


def parse_result(value: Any) -> str | None:
    """Chuyển đổi kết quả Y/N sang rfq_result enum."""
    if value is None:
        return None
    s = str(value).strip().lower()
    return RESULT_MAP.get(s)


def parse_delivery_status(value: Any) -> str | None:
    """Chuyển đổi trạng thái giao hàng sang delivery_status enum."""
    if value is None:
        return None
    s = str(value).strip().lower()
    return DELIVERY_STATUS_MAP.get(s)


def parse_goods_type(value: Any) -> str | None:
    """Chuyển đổi loại hàng sang goods_type enum."""
    if value is None:
        return None
    s = str(value).strip().lower()
    return GOODS_TYPE_MAP.get(s)


def parse_samsung_po_status(value: Any) -> str | None:
    """Chuyển đổi process status Samsung PO."""
    if value is None:
        return None
    s = str(value).strip().lower()
    return SAMSUNG_PO_STATUS_MAP.get(s, "new")


def parse_currency(value: Any) -> str | None:
    """Chuyển đổi mã tiền tệ."""
    if value is None:
        return None
    s = str(value).strip().upper()
    valid = {"VND", "USD", "RMB", "KRW", "JPY", "EUR"}
    return s if s in valid else None


def parse_cell(value: Any, field_name: str) -> Any:
    """Parse giá trị ô Excel dựa vào tên trường đích."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    # Bỏ qua cột _row_no (STT, No.)
    if field_name.startswith("_"):
        return None

    if field_name in DATE_FIELDS:
        return parse_date(value)

    if field_name in NUMERIC_FIELDS:
        return parse_number(value)

    if field_name in BOOLEAN_FIELDS:
        return parse_boolean(value)

    if field_name == "result":
        return parse_result(value)

    if field_name == "delivery_status":
        return parse_delivery_status(value)

    if field_name == "goods_type":
        return parse_goods_type(value)

    if field_name == "process_status":
        return parse_samsung_po_status(value)

    if field_name == "currency":
        return parse_currency(value)

    # Mặc định: chuỗi
    return str(value).strip()


def compute_source_hash(row_dict: dict[str, Any]) -> str:
    """Tính SHA256 hash của dòng dữ liệu để phát hiện thay đổi."""
    serialized = json.dumps(row_dict, default=str, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Excel reader — dùng openpyxl
# ---------------------------------------------------------------------------

def read_excel_sheet(
    filepath: str,
    sheet_name: str | None,
) -> tuple[list[str], list[list[Any]]]:
    """
    Đọc 1 sheet Excel, trả về (header_row, data_rows).
    Nếu sheet_name là None, đọc sheet đầu tiên.
    Hỗ trợ multiline headers (merged cells).
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            # Tìm sheet case-insensitive
            for sn in wb.sheetnames:
                if sn.strip().lower() == sheet_name.strip().lower():
                    sheet_name = sn
                    break
            else:
                wb.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' không tồn tại. "
                    f"Các sheet có sẵn: {wb.sheetnames}"
                )
        ws = wb[sheet_name]
    else:
        ws = wb.active or wb[wb.sheetnames[0]]

    # Đọc tất cả rows
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    wb.close()

    if len(all_rows) < 2:
        return [], []

    # Tìm header row: dòng đầu tiên có >= 3 ô không rỗng
    header_idx = 0
    for i, row in enumerate(all_rows[:20]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 3:
            header_idx = i
            break

    # Xử lý multiline headers: merge row header_idx với row tiếp theo
    # nếu row tiếp theo cũng là header (ít ô hơn nhưng bổ sung)
    headers = [str(c).strip() if c else "" for c in all_rows[header_idx]]

    # Kiểm tra nếu dòng kế tiếp là phần tiếp nối header (merged cells)
    if header_idx + 1 < len(all_rows):
        next_row = all_rows[header_idx + 1]
        next_non_empty = sum(1 for c in next_row if c is not None and str(c).strip())
        # Nếu dòng tiếp theo có ít hơn 50% ô khác rỗng so với header
        header_non_empty = sum(1 for h in headers if h)
        if 0 < next_non_empty < header_non_empty * 0.5:
            # Merge: nối nội dung dòng kế vào header
            for i, c in enumerate(next_row):
                if c is not None and str(c).strip() and i < len(headers):
                    if headers[i]:
                        headers[i] = headers[i] + "\n" + str(c).strip()
                    else:
                        headers[i] = str(c).strip()
            data = all_rows[header_idx + 2:]
        else:
            data = all_rows[header_idx + 1:]
    else:
        data = all_rows[header_idx + 1:]

    return headers, data


def find_column_indices(
    excel_headers: list[str],
    column_map: dict[str, str],
) -> dict[int, str]:
    """
    Tìm index của các cột Excel, map sang tên trường DB.
    Hỗ trợ header có xuống dòng (\\n) và so sánh fuzzy.
    Returns: {excel_col_index: db_field_name}
    """
    result: dict[int, str] = {}

    # Chuẩn hóa header Excel
    normalized_headers = []
    for h in excel_headers:
        norm = h.strip().replace("\r\n", "\n").replace("\r", "\n")
        normalized_headers.append(norm)

    for excel_name, db_field in column_map.items():
        # Bỏ qua cột _row_no
        if db_field.startswith("_"):
            continue

        excel_name_norm = excel_name.strip().replace("\r\n", "\n").replace("\r", "\n")

        # Exact match
        for i, h in enumerate(normalized_headers):
            if h == excel_name_norm and i not in result:
                result[i] = db_field
                break
        else:
            # Case-insensitive match
            excel_lower = excel_name_norm.lower()
            for i, h in enumerate(normalized_headers):
                if h.lower() == excel_lower and i not in result:
                    result[i] = db_field
                    break
            else:
                # Match dòng đầu tiên (bỏ phần xuống dòng)
                excel_first_line = excel_name_norm.split("\n")[0].strip().lower()
                for i, h in enumerate(normalized_headers):
                    h_first_line = h.split("\n")[0].strip().lower()
                    if h_first_line == excel_first_line and i not in result:
                        result[i] = db_field
                        break
                else:
                    # Substring match (header chứa tên cột)
                    for i, h in enumerate(normalized_headers):
                        h_clean = h.replace("\n", " ").strip().lower()
                        excel_clean = excel_name_norm.replace("\n", " ").strip().lower()
                        if excel_clean in h_clean and i not in result:
                            result[i] = db_field
                            break

    return result


# ---------------------------------------------------------------------------
# File finder — tìm file trong thư mục (bao gồm thư mục con)
# ---------------------------------------------------------------------------

def find_file(source_path: Path, relative_path: str) -> Path | None:
    """
    Tìm file theo đường dẫn tương đối, hỗ trợ thư mục con.
    Thử nhiều cách: exact, case-insensitive, recursive glob.
    """
    # Thử exact path
    filepath = source_path / relative_path
    if filepath.exists():
        return filepath

    # Thử case-insensitive (cùng thư mục)
    parts = Path(relative_path).parts
    if len(parts) == 1:
        filename = parts[0]
        for f in source_path.iterdir():
            if f.is_file() and f.name.lower() == filename.lower():
                return f
    else:
        # Thử tìm trong thư mục con
        subdir = source_path
        for part in parts[:-1]:
            found_dir = None
            if subdir.exists():
                for d in subdir.iterdir():
                    if d.is_dir() and d.name.lower() == part.lower():
                        found_dir = d
                        break
            if found_dir:
                subdir = found_dir
            else:
                break

        target_file = parts[-1]
        if subdir.exists():
            for f in subdir.iterdir():
                if f.is_file() and f.name.lower() == target_file.lower():
                    return f

    # Recursive glob
    filename_only = Path(relative_path).name
    for f in source_path.rglob("*"):
        if f.is_file() and f.name.lower() == filename_only.lower():
            return f

    return None


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

async def import_file_sheet(
    conn,
    filepath: str,
    sheet_name: str | None,
    config: dict[str, Any],
    dry_run: bool = False,
    verbose: bool = False,
) -> dict[str, int]:
    """
    Import một file/sheet Excel vào bảng DB theo config.
    Returns: {"inserted": N, "updated": N, "skipped": N, "errors": N}
    """
    table = config["table"]
    column_map = config["columns"]
    insert_sql = config["insert_sql"].replace("{data_source}", DATA_SOURCE)
    column_order = config["column_order"]

    logger.info("  Đọc: %s (sheet: %s)", Path(filepath).name, sheet_name or "mặc định")

    try:
        headers, data_rows = read_excel_sheet(filepath, sheet_name)
    except Exception as e:
        logger.error("  Lỗi đọc file %s: %s", filepath, e)
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": 1}

    if not headers or not data_rows:
        logger.warning("  File trống hoặc không đủ dữ liệu")
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0}

    # Map cột Excel → cột DB
    col_indices = find_column_indices(headers, column_map)

    if len(col_indices) < 2:
        logger.warning(
            "  Chỉ tìm thấy %d/%d cột. Headers: %s",
            len(col_indices), len(column_map), headers[:15],
        )
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0}

    mapped_fields = {v for v in col_indices.values()}
    logger.info(
        "  Map %d/%d cột → '%s'. Tổng %d dòng.",
        len(col_indices), len(column_map), table, len(data_rows),
    )
    if verbose:
        logger.debug("  Mapped fields: %s", sorted(mapped_fields))

    inserted = 0
    updated = 0
    skipped = 0
    errors = 0

    for row_num, row in enumerate(data_rows, start=2):
        # Bỏ qua dòng trống
        if not row or all(c is None or str(c).strip() == "" for c in row):
            skipped += 1
            continue

        # Xây dựng dict từ dòng Excel
        row_dict: dict[str, Any] = {}
        for col_idx, db_field in col_indices.items():
            if col_idx < len(row):
                row_dict[db_field] = parse_cell(row[col_idx], db_field)

        # Bỏ qua dòng thiếu dữ liệu quan trọng (2 cột đầu đều null)
        key_cols = [c for c in column_order[:3] if not c.startswith("_")]
        if all(row_dict.get(col) is None for col in key_cols):
            skipped += 1
            continue

        # Xây dựng params theo thứ tự INSERT
        params = []
        for col in column_order:
            if col.startswith("_"):
                continue
            params.append(row_dict.get(col))

        if dry_run:
            if inserted < 3:
                logger.info(
                    "  [DRY-RUN] Dòng %d: %s",
                    row_num, dict(zip(column_order, params)),
                )
            inserted += 1
            continue

        try:
            result = await conn.execute(insert_sql, *params)
            # Kiểm tra INSERT hay UPDATE
            if "INSERT 0 1" in result:
                inserted += 1
            elif "INSERT 0 0" in result:
                # ON CONFLICT DO NOTHING — already exists
                skipped += 1
            else:
                # ON CONFLICT DO UPDATE
                updated += 1
        except Exception as e:
            errors += 1
            if errors <= 10:
                logger.warning(
                    "  Lỗi dòng %d: %s | Data: %s",
                    row_num, str(e)[:200], params[:5],
                )

        # Hiện tiến trình mỗi 500 dòng
        total_processed = inserted + updated + skipped + errors
        if total_processed % 500 == 0 and total_processed > 0:
            logger.info(
                "  ... đã xử lý %d/%d dòng (I:%d U:%d S:%d E:%d)",
                total_processed, len(data_rows),
                inserted, updated, skipped, errors,
            )

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def log_etl_sync(
    conn,
    sync_type: str,
    status: str,
    files_processed: int = 0,
    rows_inserted: int = 0,
    rows_updated: int = 0,
    rows_skipped: int = 0,
    error_message: str | None = None,
) -> None:
    """Ghi log import vào bảng etl_sync_log."""
    try:
        await conn.execute(
            """
            INSERT INTO etl_sync_log (
                sync_type, status, files_processed,
                rows_inserted, rows_updated, rows_skipped,
                error_message, completed_at
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, NOW()
            )
            """,
            sync_type, status, files_processed,
            rows_inserted, rows_updated, rows_skipped,
            error_message,
        )
    except Exception as e:
        logger.warning("Không thể ghi etl_sync_log: %s", e)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main(
    source_dir: str,
    dry_run: bool = False,
    table_filter: str | None = None,
    verbose: bool = False,
) -> None:
    """Chạy import toàn bộ theo IMPORT_MAP."""

    import asyncpg

    source_path = Path(source_dir)
    if not source_path.exists():
        logger.error("Thư mục nguồn không tồn tại: %s", source_dir)
        sys.exit(1)

    logger.info("=" * 70)
    logger.info("SONG CHÂU ERP — IMPORT TOÀN BỘ DỮ LIỆU (15 nhóm)")
    logger.info("=" * 70)
    logger.info("Thư mục nguồn : %s", source_dir)
    logger.info("DSN           : %s", DSN.split("@")[-1])
    logger.info("Dry run       : %s", dry_run)
    logger.info("Data source   : %s", DATA_SOURCE)
    if table_filter:
        logger.info("Lọc bảng      : %s", table_filter)
    logger.info("-" * 70)

    start_time = time.time()

    # Kết nối database
    if dry_run:
        conn = None
        logger.info("[DRY-RUN] Bỏ qua kết nối database.")
    else:
        try:
            conn = await asyncpg.connect(DSN)
            logger.info("Kết nối database thành công.")
        except Exception as e:
            logger.error("Không thể kết nối database: %s", e)
            sys.exit(1)

    total_inserted = 0
    total_updated = 0
    total_skipped = 0
    total_errors = 0
    files_processed = 0
    table_results: dict[str, dict] = {}

    try:
        for config_idx, config in enumerate(IMPORT_MAP, start=1):
            table = config["table"]

            # Lọc theo bảng nếu có
            if table_filter and table_filter not in table:
                continue

            logger.info("")
            logger.info(
                "━━━ [%d/%d] Bảng: %s ━━━",
                config_idx, len(IMPORT_MAP), table,
            )

            table_stats = {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0}

            for filename in config["files"]:
                filepath = find_file(source_path, filename)

                if not filepath:
                    logger.warning("  KHÔNG TÌM THẤY: %s", filename)
                    continue

                # Xác định sheets cần đọc
                sheet_config = config.get("sheet")
                if isinstance(sheet_config, list):
                    sheets_to_read = sheet_config
                elif sheet_config is None:
                    sheets_to_read = [None]
                else:
                    sheets_to_read = [sheet_config]

                for sheet_name in sheets_to_read:
                    result = await import_file_sheet(
                        conn, str(filepath), sheet_name,
                        config, dry_run, verbose,
                    )

                    for key in ("inserted", "updated", "skipped", "errors"):
                        table_stats[key] += result[key]

                    files_processed += 1

            # Tổng kết cho bảng
            logger.info(
                "  ▸ %s: I=%d U=%d S=%d E=%d",
                table,
                table_stats["inserted"], table_stats["updated"],
                table_stats["skipped"], table_stats["errors"],
            )
            table_results[table] = table_stats

            total_inserted += table_stats["inserted"]
            total_updated += table_stats["updated"]
            total_skipped += table_stats["skipped"]
            total_errors += table_stats["errors"]

        # Ghi log vào DB
        if conn and not dry_run:
            status = "success" if total_errors == 0 else "error"
            await log_etl_sync(
                conn,
                sync_type="full_import",
                status=status,
                files_processed=files_processed,
                rows_inserted=total_inserted,
                rows_updated=total_updated,
                rows_skipped=total_skipped,
                error_message=f"{total_errors} errors" if total_errors else None,
            )

    finally:
        if conn:
            await conn.close()
            logger.info("Đã đóng kết nối database.")

    elapsed = time.time() - start_time

    # Tổng kết
    logger.info("")
    logger.info("=" * 70)
    logger.info("TỔNG KẾT IMPORT TOÀN BỘ")
    logger.info("=" * 70)
    logger.info("Files đã xử lý  : %d", files_processed)
    logger.info("Tổng INSERT      : %d", total_inserted)
    logger.info("Tổng UPDATE      : %d", total_updated)
    logger.info("Tổng SKIP        : %d", total_skipped)
    logger.info("Tổng LỖI         : %d", total_errors)
    logger.info("Thời gian        : %.1f giây", elapsed)
    logger.info("-" * 70)

    for table, stats in table_results.items():
        logger.info(
            "  %-30s I=%-6d U=%-6d S=%-6d E=%-6d",
            table,
            stats["inserted"], stats["updated"],
            stats["skipped"], stats["errors"],
        )

    logger.info("=" * 70)

    if total_errors > 0:
        logger.warning(
            "Có %d lỗi — kiểm tra log phía trên để biết chi tiết.",
            total_errors,
        )


def cli() -> None:
    """Parse command-line arguments và chạy import."""
    parser = argparse.ArgumentParser(
        description="Song Châu ERP — Import TOÀN BỘ dữ liệu từ Excel (15 nhóm file)",
    )
    parser.add_argument(
        "--source",
        required=True,
        help="Đường dẫn thư mục chứa file Excel (OneDrive folder)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Chỉ đọc và hiển thị, không ghi vào database",
    )
    parser.add_argument(
        "--table",
        default=None,
        help="Lọc bảng (vd: bqms_rfq, imv, revenue). Hỗ trợ substring match.",
    )
    parser.add_argument(
        "--dsn",
        default=None,
        help="Override DSN kết nối database",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Hiển thị thêm thông tin debug",
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.dsn:
        global DSN
        DSN = args.dsn

    asyncio.run(main(
        source_dir=args.source,
        dry_run=args.dry_run,
        table_filter=args.table,
        verbose=args.verbose,
    ))


if __name__ == "__main__":
    cli()
