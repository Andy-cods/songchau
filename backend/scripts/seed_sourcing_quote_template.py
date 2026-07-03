"""Seed the Song Châu báo-giá template to the VPS-mounted templates dir.

Copies the in-repo ``SOURCING_QUOTE.xlsx`` (byte-identical to the official
"mẫu báo giá.xlsx") to ``/data/files/templates/SOURCING_QUOTE.xlsx`` so the
volume-mounted, mutable copy survives container restarts — exactly like
``seed_dossier_template.py`` seeds AMA_DELIVERY_DOSSIER.

IMPORTANT: this does a RAW BYTE COPY (``shutil.copyfile``), NOT an openpyxl
round-trip. Re-saving through openpyxl would strip/relocate the 2 embedded
images (logo @A1, signature stamp @B21) and the print area. A byte copy
preserves the file exactly.

Run inside sc-api container:
    docker exec sc-api python /app/scripts/seed_sourcing_quote_template.py
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

# backend/scripts/seed_sourcing_quote_template.py -> parents[1] == backend/
SRC = (
    Path(__file__).resolve().parents[1]
    / "app" / "services" / "templates" / "SOURCING_QUOTE.xlsx"
)
DST = Path("/data/files/templates/SOURCING_QUOTE.xlsx")


def main() -> int:
    print("=== Seed Song Châu SOURCING_QUOTE template ===")
    print(f"SRC: {SRC}")
    print(f"DST: {DST}")

    if not SRC.exists():
        print(f"ERROR: source template missing: {SRC}", file=sys.stderr)
        return 1

    DST.parent.mkdir(parents=True, exist_ok=True)
    # Raw byte copy — preserves embedded images + print area (no openpyxl round-trip).
    shutil.copyfile(SRC, DST)
    print(f"Copied {SRC.stat().st_size} bytes -> {DST}")

    # ── Verify the seeded file ─────────────────────────────────────────
    wb = openpyxl.load_workbook(DST)
    print(f"Sheets: {wb.sheetnames}")
    assert wb.sheetnames == ["200526"], (
        f"Unexpected sheet names: {wb.sheetnames} (expected ['200526'])"
    )
    ws = wb["200526"]
    n_images = len(ws._images)  # noqa: SLF001 — openpyxl exposes images here only
    print(f"Embedded images: {n_images}")
    assert n_images == 2, (
        f"Expected 2 embedded images (logo @A1 + signature @B21), got {n_images}"
    )

    print("\n=== DONE — SOURCING_QUOTE.xlsx seeded + verified ===")
    return 0


if __name__ == "__main__":
    sys.exit(main())
