#!/usr/bin/env python3
"""
Song Châu ERP — Import dữ liệu lịch sử từ Excel files vào PostgreSQL.

Chạy MỘT LẦN trước go-live để nạp data từ OneDrive xuống database.
Script đọc các file Excel theo IMPORT_MAP, map cột, rồi INSERT INTO bảng tương ứng.
Sử dụng ON CONFLICT DO NOTHING để chạy lại an toàn (idempotent).

Usage:
    python scripts/import_data.py --source /path/to/onedrive/folder
    python scripts/import_data.py --source "D:/OneDrive - Song Chau/Data" --dry-run
    python scripts/import_data.py --source ./data --table bqms_rfq
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import os
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Logging — Vietnamese messages
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("import_data")

# ---------------------------------------------------------------------------
# Database DSN — có thể override bằng biến môi trường
# ---------------------------------------------------------------------------

DSN = os.getenv(
    "DATABASE_URL",
    "postgresql://scadmin:SC2026_ERP_Pr0d_X9k2mQ7wR4@postgres:5432/songchau_erp",
)

# ---------------------------------------------------------------------------
# Import mapping: file Excel → bảng database
# ---------------------------------------------------------------------------

IMPORT_MAP: list[dict[str, Any]] = [
    # ── BQMS Hỏi Hàng (RFQ) ──
    {
        "files": ["Thong ke hoi hang BQMS.xlsx"],
        "sheet": "TONG HOP BQMS",
        "table": "bqms_rfq",
        "columns": {
            "Ngày": "inquiry_date",
            "Người phụ trách": "person_in_charge_name",
            "RFQ No.": "rfq_number",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "Maker": "maker",
            "Số lượng dự kiến": "expected_qty",
            "Giá nhập\nRMB": "purchase_price_rmb",
            "Giá nhập\nVND": "purchase_price_vnd",
            "Giá báo cho AMA": "quoted_price_ama",
            "Giá báo cho BQMS V1": "quoted_price_bqms_v1",
            "V2": "quoted_price_bqms_v2",
            "V3": "quoted_price_bqms_v3",
            "NCC": "supplier_name",
            "Kết quả\n(Y/N)": "result",
            "Ghi chú": "notes",
        },
        "insert_sql": """
            INSERT INTO bqms_rfq (
                inquiry_date, person_in_charge_name, rfq_number, bqms_code,
                specification, maker, expected_qty,
                purchase_price_rmb, purchase_price_vnd,
                quoted_price_ama, quoted_price_bqms_v1,
                quoted_price_bqms_v2, quoted_price_bqms_v3,
                supplier_name, result, notes
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13,
                $14, COALESCE($15::rfq_result, 'pending'), $16
            )
            ON CONFLICT DO NOTHING
        """,
        "column_order": [
            "inquiry_date", "person_in_charge_name", "rfq_number", "bqms_code",
            "specification", "maker", "expected_qty",
            "purchase_price_rmb", "purchase_price_vnd",
            "quoted_price_ama", "quoted_price_bqms_v1",
            "quoted_price_bqms_v2", "quoted_price_bqms_v3",
            "supplier_name", "result", "notes",
        ],
    },
    # ── BQMS Giao Hàng (Deliveries) — nhiều năm ──
    {
        "files": [
            "Thong ke giao hang 2026.xlsx",
            "Thong ke giao hang 2025.xlsx",
            "Thong ke giao hang 2023-2024.xlsx",
        ],
        "sheet": None,  # Sheet đầu tiên
        "table": "bqms_deliveries",
        "columns": {
            "Ngày PO": "po_date",
            "Số PO": "po_number",
            "Shipping No": "shipping_no",
            "Số QT": "quotation_no",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "SL": "quantity",
            "Đơn vị": "unit",
            "Đơn giá": "unit_price",
            "Thành tiền": "amount",
            "SEV/T": "sev_type",
            "MAIL PUR": "buyer_email",
            "TÊN NGƯỜI NHẬN": "recipient_name",
            "KHO NHẬN": "receiving_warehouse",
            "SĐT PUR": "buyer_phone",
            "TÌNH TRẠNG": "delivery_status",
            "NGÀY GIAO HÀNG": "delivery_date",
            "SL GIAO THỰC TẾ": "actual_delivered_qty",
            "CÁCH THỨC GIAO HÀNG": "delivery_method",
            "XUẤT XỨ": "country_origin",
        },
        "insert_sql": """
            INSERT INTO bqms_deliveries (
                po_date, po_number, shipping_no, quotation_no,
                bqms_code, specification, quantity, unit,
                unit_price, amount, sev_type, buyer_email,
                recipient_name, receiving_warehouse, buyer_phone,
                delivery_status, delivery_date, actual_delivered_qty,
                delivery_method, country_origin
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10,
                $11, $12, $13, $14, $15,
                COALESCE($16::delivery_status, 'chua_giao'),
                $17, $18, $19, $20
            )
        """,
        "column_order": [
            "po_date", "po_number", "shipping_no", "quotation_no",
            "bqms_code", "specification", "quantity", "unit",
            "unit_price", "amount", "sev_type", "buyer_email",
            "recipient_name", "receiving_warehouse", "buyer_phone",
            "delivery_status", "delivery_date", "actual_delivered_qty",
            "delivery_method", "country_origin",
        ],
    },
    # ── BQMS Đặt Hàng (Orders) ──
    {
        "files": ["Thong ke dat hang.xlsx"],
        "sheet": "Sheet1",
        "table": "bqms_orders",
        "columns": {
            "RFQ No.": "rfq_number",
            "BQMS code": "bqms_code",
            "Spec": "specification",
            "Khách hàng": "customer_name",
            "Số lượng dự kiến": "expected_qty",
            "SL đặt hàng": "order_qty",
            "ĐVT": "unit",
            "Ngày đặt hàng": "order_date",
            "Trạng thái": "status",
            "SL giao": "delivered_qty",
            "Ngày giao": "delivery_date",
            "Ghi chú": "notes",
        },
        "insert_sql": """
            INSERT INTO bqms_orders (
                rfq_number, bqms_code, specification, customer_name,
                expected_qty, order_qty, unit, order_date,
                status, delivered_qty, delivery_date, notes
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12
            )
        """,
        "column_order": [
            "rfq_number", "bqms_code", "specification", "customer_name",
            "expected_qty", "order_qty", "unit", "order_date",
            "status", "delivered_qty", "delivery_date", "notes",
        ],
    },
]


# ---------------------------------------------------------------------------
# Value parsing helpers
# ---------------------------------------------------------------------------

DATE_FIELDS = {
    "inquiry_date", "po_date", "delivery_date", "order_date",
    "preferred_delivery_date", "deadline",
}

NUMERIC_FIELDS = {
    "expected_qty", "quantity", "order_qty", "unit_price", "amount",
    "purchase_price_rmb", "purchase_price_vnd",
    "quoted_price_ama", "quoted_price_bqms_v1",
    "quoted_price_bqms_v2", "quoted_price_bqms_v3",
    "actual_delivered_qty", "delivered_qty",
    "shipping_qty", "gr_qty", "remaining_qty",
    "weight_kg", "total_delivered_value_vnd",
}

RESULT_MAP = {
    "y": "won",
    "yes": "won",
    "trúng": "won",
    "trung": "won",
    "n": "lost",
    "no": "lost",
    "thua": "lost",
    "hủy": "cancelled",
    "huy": "cancelled",
    "cancel": "cancelled",
    "cancelled": "cancelled",
}

DELIVERY_STATUS_MAP = {
    "chưa giao": "chua_giao",
    "chua giao": "chua_giao",
    "đang giao": "dang_giao",
    "dang giao": "dang_giao",
    "đã giao": "da_giao",
    "da giao": "da_giao",
    "giao một phần": "giao_mot_phan",
    "giao 1 phan": "giao_mot_phan",
    "giao mot phan": "giao_mot_phan",
}


def parse_date(value: Any) -> date | None:
    """Chuyển đổi nhiều format ngày tháng từ Excel sang date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from datetime import timedelta
            base = date(1899, 12, 30)
            return base + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    s = str(value).strip()
    if not s or s == "-":
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value: Any) -> float | None:
    """Chuyển đổi số từ Excel, hỗ trợ format VN (dấu chấm phân cách hàng nghìn)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace(" ", "")
    if not s or s == "-" or s == "N/A":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_result(value: Any) -> str | None:
    """Chuyển đổi kết quả Y/N sang rfq_result enum."""
    if value is None:
        return None
    s = str(value).strip().lower()
    return RESULT_MAP.get(s)


def parse_delivery_status(value: Any) -> str | None:
    """Chuyển đổi trạng thái giao hàng sang delivery_status enum."""
    if value is None:
        return None
    s = str(value).strip().lower()
    return DELIVERY_STATUS_MAP.get(s)


def parse_cell(value: Any, field_name: str) -> Any:
    """Parse giá trị ô Excel dựa vào tên trường đích."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    if field_name in DATE_FIELDS:
        return parse_date(value)

    if field_name in NUMERIC_FIELDS:
        return parse_number(value)

    if field_name == "result":
        return parse_result(value)

    if field_name == "delivery_status":
        return parse_delivery_status(value)

    # Mặc định: chuỗi
    return str(value).strip()


# ---------------------------------------------------------------------------
# Excel reader — dùng openpyxl (đã có trong requirements.txt)
# ---------------------------------------------------------------------------

def read_excel_sheet(filepath: str, sheet_name: str | None) -> tuple[list[str], list[list[Any]]]:
    """
    Đọc 1 sheet Excel, trả về (header_row, data_rows).
    Nếu sheet_name là None, đọc sheet đầu tiên.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            # Thử tìm sheet gần giống (case-insensitive)
            for sn in wb.sheetnames:
                if sn.strip().lower() == sheet_name.strip().lower():
                    sheet_name = sn
                    break
            else:
                wb.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' không tồn tại. "
                    f"Các sheet có sẵn: {wb.sheetnames}"
                )
        ws = wb[sheet_name]
    else:
        ws = wb.active or wb[wb.sheetnames[0]]

    # Đọc tất cả rows
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    wb.close()

    if len(all_rows) < 2:
        return [], []

    # Tìm header row: dòng đầu tiên có >= 3 ô không rỗng
    header_idx = 0
    for i, row in enumerate(all_rows[:15]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 3:
            header_idx = i
            break

    headers = [str(c).strip() if c else "" for c in all_rows[header_idx]]
    data = all_rows[header_idx + 1:]

    return headers, data


def find_column_indices(
    excel_headers: list[str],
    column_map: dict[str, str],
) -> dict[int, str]:
    """
    Tìm index của các cột Excel, map sang tên trường DB.
    Hỗ trợ header có xuống dòng (\\n) và so sánh fuzzy.
    Returns: {excel_col_index: db_field_name}
    """
    result: dict[int, str] = {}

    # Chuẩn hóa header Excel: lower, strip, thay \\n bằng \\n
    normalized_headers = []
    for h in excel_headers:
        norm = h.strip().replace("\r\n", "\n").replace("\r", "\n")
        normalized_headers.append(norm)

    for excel_name, db_field in column_map.items():
        excel_name_norm = excel_name.strip().replace("\r\n", "\n").replace("\r", "\n")

        # Tìm exact match trước
        for i, h in enumerate(normalized_headers):
            if h == excel_name_norm:
                result[i] = db_field
                break
        else:
            # Thử match không phân biệt hoa thường
            excel_lower = excel_name_norm.lower()
            for i, h in enumerate(normalized_headers):
                if h.lower() == excel_lower:
                    result[i] = db_field
                    break
            else:
                # Thử match chỉ dòng đầu tiên (bỏ phần xuống dòng)
                excel_first_line = excel_name_norm.split("\n")[0].strip().lower()
                for i, h in enumerate(normalized_headers):
                    h_first_line = h.split("\n")[0].strip().lower()
                    if h_first_line == excel_first_line and i not in result:
                        result[i] = db_field
                        break

    return result


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

async def import_file(
    conn,
    filepath: str,
    config: dict[str, Any],
    dry_run: bool = False,
) -> dict[str, int]:
    """
    Import một file Excel vào bảng DB theo config.

    Returns: {"inserted": N, "skipped": N, "errors": N}
    """
    table = config["table"]
    sheet_name = config.get("sheet")
    column_map = config["columns"]
    insert_sql = config["insert_sql"]
    column_order = config["column_order"]

    logger.info("  Đọc file: %s (sheet: %s)", filepath, sheet_name or "đầu tiên")

    try:
        headers, data_rows = read_excel_sheet(filepath, sheet_name)
    except Exception as e:
        logger.error("  Lỗi đọc file %s: %s", filepath, e)
        return {"inserted": 0, "skipped": 0, "errors": 1}

    if not headers or not data_rows:
        logger.warning("  File %s trống hoặc không đủ dữ liệu", filepath)
        return {"inserted": 0, "skipped": 0, "errors": 0}

    # Map cột Excel → cột DB
    col_indices = find_column_indices(headers, column_map)

    if len(col_indices) < 2:
        logger.warning(
            "  Chỉ tìm thấy %d/%d cột khớp trong %s. Headers: %s",
            len(col_indices), len(column_map), filepath, headers[:20],
        )
        return {"inserted": 0, "skipped": 0, "errors": 0}

    logger.info(
        "  Đã map %d/%d cột → bảng '%s'. Tổng %d dòng dữ liệu.",
        len(col_indices), len(column_map), table, len(data_rows),
    )

    inserted = 0
    skipped = 0
    errors = 0

    for row_num, row in enumerate(data_rows, start=2):
        # Bỏ qua dòng trống hoàn toàn
        if not row or all(c is None or str(c).strip() == "" for c in row):
            skipped += 1
            continue

        # Xây dựng dict từ dòng Excel
        row_dict: dict[str, Any] = {}
        for col_idx, db_field in col_indices.items():
            if col_idx < len(row):
                row_dict[db_field] = parse_cell(row[col_idx], db_field)

        # Bỏ qua dòng không có dữ liệu quan trọng (3 cột đầu đều null)
        first_three = [row_dict.get(col) for col in column_order[:3]]
        if all(v is None for v in first_three):
            skipped += 1
            continue

        # Xây dựng params theo thứ tự INSERT
        params = [row_dict.get(col) for col in column_order]

        if dry_run:
            if inserted < 3:
                logger.info("  [DRY-RUN] Dòng %d: %s", row_num, dict(zip(column_order, params)))
            inserted += 1
            continue

        try:
            await conn.execute(insert_sql, *params)
            inserted += 1
        except Exception as e:
            errors += 1
            if errors <= 10:
                logger.warning("  Lỗi dòng %d: %s | Data: %s", row_num, e, params[:5])

        # Hiện tiến trình mỗi 500 dòng
        if (row_num % 500) == 0:
            logger.info("  ... đã xử lý %d/%d dòng", row_num, len(data_rows))

    return {"inserted": inserted, "skipped": skipped, "errors": errors}


async def log_etl_sync(
    conn,
    source_file: str,
    status: str,
    rows_inserted: int = 0,
    rows_skipped: int = 0,
    error_message: str | None = None,
) -> None:
    """Ghi log import vào bảng etl_sync_log."""
    try:
        await conn.execute(
            """
            INSERT INTO etl_sync_log (
                sync_type, source_file, status,
                rows_inserted, rows_skipped, error_message, completed_at
            ) VALUES (
                'excel_import', $1, $2, $3, $4, $5, NOW()
            )
            """,
            source_file, status, rows_inserted, rows_skipped, error_message,
        )
    except Exception as e:
        logger.warning("Không thể ghi etl_sync_log: %s", e)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def main(
    source_dir: str,
    dry_run: bool = False,
    table_filter: str | None = None,
) -> None:
    """Chạy import toàn bộ theo IMPORT_MAP."""

    import asyncpg

    source_path = Path(source_dir)
    if not source_path.exists():
        logger.error("Thư mục nguồn không tồn tại: %s", source_dir)
        sys.exit(1)

    logger.info("=" * 70)
    logger.info("SONG CHÂU ERP — IMPORT DỮ LIỆU LỊCH SỬ")
    logger.info("=" * 70)
    logger.info("Thư mục nguồn: %s", source_dir)
    logger.info("DSN: %s", DSN.split("@")[-1])  # Ẩn password
    logger.info("Dry run: %s", dry_run)
    if table_filter:
        logger.info("Chỉ import bảng: %s", table_filter)
    logger.info("-" * 70)

    start_time = time.time()

    # Kết nối database
    if dry_run:
        conn = None
        logger.info("[DRY-RUN] Bỏ qua kết nối database.")
    else:
        try:
            conn = await asyncpg.connect(DSN)
            logger.info("Kết nối database thành công.")
        except Exception as e:
            logger.error("Không thể kết nối database: %s", e)
            sys.exit(1)

    total_inserted = 0
    total_skipped = 0
    total_errors = 0
    files_processed = 0

    try:
        for config in IMPORT_MAP:
            table = config["table"]

            # Lọc theo bảng nếu có
            if table_filter and table != table_filter:
                continue

            logger.info("")
            logger.info("━━━ Bảng: %s ━━━", table)

            for filename in config["files"]:
                filepath = source_path / filename

                # Thử tìm file (hỗ trợ có/không dấu tiếng Việt)
                if not filepath.exists():
                    # Thử tìm trong tất cả files cùng thư mục
                    found = False
                    for f in source_path.iterdir():
                        if f.name.lower() == filename.lower():
                            filepath = f
                            found = True
                            break
                    if not found:
                        logger.warning("  KHÔNG TÌM THẤY: %s", filename)
                        continue

                result = await import_file(conn, str(filepath), config, dry_run)

                total_inserted += result["inserted"]
                total_skipped += result["skipped"]
                total_errors += result["errors"]
                files_processed += 1

                status = "success" if result["errors"] == 0 else "partial"
                logger.info(
                    "  Kết quả: %d inserted, %d skipped, %d errors → %s",
                    result["inserted"], result["skipped"], result["errors"],
                    filename,
                )

                # Ghi log vào DB
                if conn and not dry_run:
                    await log_etl_sync(
                        conn,
                        source_file=filename,
                        status=status,
                        rows_inserted=result["inserted"],
                        rows_skipped=result["skipped"],
                        error_message=f"{result['errors']} errors" if result["errors"] else None,
                    )

    finally:
        if conn:
            await conn.close()
            logger.info("Đã đóng kết nối database.")

    elapsed = time.time() - start_time

    # Tổng kết
    logger.info("")
    logger.info("=" * 70)
    logger.info("TỔNG KẾT IMPORT")
    logger.info("=" * 70)
    logger.info("Files đã xử lý : %d", files_processed)
    logger.info("Tổng rows INSERT: %d", total_inserted)
    logger.info("Tổng rows SKIP  : %d", total_skipped)
    logger.info("Tổng lỗi        : %d", total_errors)
    logger.info("Thời gian        : %.1f giây", elapsed)
    logger.info("=" * 70)

    if total_errors > 0:
        logger.warning("Có %d lỗi — kiểm tra log phía trên để biết chi tiết.", total_errors)


def cli() -> None:
    """Parse command-line arguments và chạy import."""
    parser = argparse.ArgumentParser(
        description="Song Châu ERP — Import dữ liệu lịch sử từ Excel",
    )
    parser.add_argument(
        "--source",
        required=True,
        help="Đường dẫn thư mục chứa file Excel (OneDrive folder)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Chỉ đọc và hiển thị, không ghi vào database",
    )
    parser.add_argument(
        "--table",
        default=None,
        help="Chỉ import bảng cụ thể (vd: bqms_rfq, bqms_deliveries)",
    )
    parser.add_argument(
        "--dsn",
        default=None,
        help="Override DSN kết nối database",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Hiển thị thêm thông tin debug",
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.dsn:
        global DSN
        DSN = args.dsn

    asyncio.run(main(
        source_dir=args.source,
        dry_run=args.dry_run,
        table_filter=args.table,
    ))


if __name__ == "__main__":
    cli()
