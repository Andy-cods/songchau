"""Seed AMA delivery dossier template + dump cell map JSON.

One-shot script: read sample .xls from `/data/onedrive-staging/Puplic/BQMS/Giao hàng/BBGH 2026/SEV/po 2112520763, ...`,
convert to .xlsx, save to `/data/files/templates/AMA_DELIVERY_DOSSIER.xlsx`,
dump cell positions + sheet structure to sibling `AMA_DELIVERY_DOSSIER.cells.json`.

Run inside sc-api container:
    docker exec sc-api python /app/scripts/seed_dossier_template.py
"""
from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SAMPLE_DIR = "/data/onedrive-staging/Puplic/BQMS/Giao hàng/BBGH 2026/SEV/po 2112520763, 2112522933 roller (2000, 700)"
OUTPUT_TEMPLATE = "/data/files/templates/AMA_DELIVERY_DOSSIER.xlsx"
OUTPUT_CELLS_JSON = "/data/files/templates/AMA_DELIVERY_DOSSIER.cells.json"


def find_sample_xls() -> str:
    """Find sample .xls file in SAMPLE_DIR."""
    if not os.path.isdir(SAMPLE_DIR):
        # Try without diacritic
        parent = "/data/onedrive-staging/Puplic/BQMS"
        for d in os.listdir(parent):
            if "iao" in d:
                candidate = f"{parent}/{d}/BBGH 2026/SEV/po 2112520763, 2112522933 roller (2000, 700)"
                if os.path.isdir(candidate):
                    return _pick_xls(candidate)
        raise FileNotFoundError(f"Sample dir not found: {SAMPLE_DIR}")
    return _pick_xls(SAMPLE_DIR)


def _pick_xls(d: str) -> str:
    for f in os.listdir(d):
        if f.lower().endswith((".xls", ".xlsx")) and "AMA" in f:
            return os.path.join(d, f)
    raise FileNotFoundError(f"No AMA xls in {d}")


def convert_xls_to_xlsx(xls_path: str, xlsx_path: str) -> dict:
    """Read xls via xlrd, write to xlsx via openpyxl, preserve sheet names + values."""
    src = xlrd.open_workbook(xls_path, formatting_info=False)
    dst = openpyxl.Workbook()
    # remove default sheet
    dst.remove(dst.active)

    sheet_meta: dict = {}
    for sn in src.sheet_names():
        sh = src.sheet_by_name(sn)
        ws = dst.create_sheet(title=sn)
        sheet_meta[sn] = {"rows": sh.nrows, "cols": sh.ncols, "value_cells": []}
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                if v != "" and v is not None:
                    cell = ws.cell(row=r + 1, column=c + 1, value=v)
                    addr = cell.coordinate
                    sheet_meta[sn]["value_cells"].append({"addr": addr, "value": v})

    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    dst.save(xlsx_path)
    return sheet_meta


def detect_cell_map(meta: dict) -> dict:
    """Heuristic detect key field cells based on label text in adjacent cells.

    For each sheet, scan for known labels (Shipping No, PO No, BQMS Code, etc.)
    and record the cell where the *value* (right side / below) lives.
    """
    label_patterns = {
        # cam_ket sheet labels
        "customer":           [r"^Customer$"],
        "vendor_name":        [r"^Vendor Name$"],
        "department":         [r"^Department$"],
        "pr_pic":             [r"^PR PIC$"],
        "receiver":           [r"Receiver.*Ký"],
        "shipping_no":        [r"Shipping No.*Invoice No", r"^Shipping No$"],
        "po_no":              [r"^PO No$"],
        "bqms_code":          [r"^BQMS Code$"],
        "item_name":          [r"^Item Name$"],
        "specification":      [r"^Specification$"],
        "quantity":           [r"^Quantity$"],
        "shipping_date":      [r"^Shipping date$"],
        "hinh_anh":           [r"^Hình ảnh$"],
        # packing_list / list_detail headers
        "stt":                [r"^STT$"],
        "po_seq":             [r"P/O Seq", r"PO Seq"],
        "sev_po_no":          [r"SEV PO No"],
        "bqms_code_h":        [r"^BQMS CODE\s*$"],
        # label sheet
        "vendor_label":       [r"^Vendor:?$"],
        "pr_person_label":    [r"PR Person"],
    }

    cell_map = {}
    for sn, sd in meta.items():
        sheet_map = {"sheet_name": sn, "rows": sd["rows"], "cols": sd["cols"], "labels": {}}
        for entry in sd["value_cells"]:
            val = entry["value"]
            if not isinstance(val, str):
                continue
            for key, patterns in label_patterns.items():
                for pat in patterns:
                    if re.search(pat, val, re.IGNORECASE):
                        # Value cell = same row, next non-empty column to the right
                        sheet_map["labels"].setdefault(key, []).append({
                            "label_addr": entry["addr"],
                            "label_value": val,
                        })
                        break
        cell_map[sn] = sheet_map
    return cell_map


def main():
    print("=== Seed AMA Delivery Dossier template ===")
    print(f"Sample dir: {SAMPLE_DIR}")

    xls_path = find_sample_xls()
    print(f"Sample xls: {xls_path}")

    print(f"\nConverting → {OUTPUT_TEMPLATE} ...")
    meta = convert_xls_to_xlsx(xls_path, OUTPUT_TEMPLATE)
    print(f"  Sheets: {list(meta.keys())}")
    for sn, sd in meta.items():
        print(f"    [{sn}] {sd['rows']}x{sd['cols']} ({len(sd['value_cells'])} value cells)")

    print(f"\nDetecting cell map ...")
    cmap = detect_cell_map(meta)

    # Persist cells.json
    out = {
        "source_xls": xls_path,
        "template_xlsx": OUTPUT_TEMPLATE,
        "sheets": cmap,
        # Hand-curated field → cell positions (will be refined after Thang reviews)
        "field_map": {
            "tong_hop_sheet":          "Tổng hợp",
            "packing_list_sheet":      "packing list",
            "cam_ket_template_sheet":  "Cam kết hình ảnh (1)",   # clone this per PO
            "list_detail_sheet":       "List Detail",
            "label_sheet":             "label",
            "cam_ket": {
                # Position confirmed from xlrd inspect:
                #   R5: Customer | C5=value
                #   R6: Vendor Name | C6=value
                #   R7: Department | C7=value
                #   R8: PR PIC | C8=value
                #   R12: Shipping No/Invoice No | C12=value
                #   R13: PO No | C13=value
                #   R14: BQMS Code | C14=value
                #   R15: Item Name | C15=value
                #   R16: Specification | C16=value
                #   R17: Quantity | C17=value
                #   R18: Shipping date | C18=value
                "customer":      "C5",
                "vendor_name":   "C6",
                "department":    "C7",
                "pr_pic":        "C8",
                "shipping_no":   "C12",
                "po_no":         "C13",
                "bqms_code":     "C14",
                "item_name":     "C15",
                "specification": "C16",
                "quantity":      "C17",
                "shipping_date": "C18",
                # Image anchors (TBD: confirm visually):
                #   R10: 'Hình ảnh', 'Hệ thống', 'Thực tế' headers
                #   R11+: image cells
                "img_system_anchor": {"from": "C11", "to": "F12"},   # 4col × 2row block for left thumbnail
                "img_actual_anchor": {"from": "G11", "to": "J12"},   # 4col × 2row block for right thumbnail
            },
            "packing_list": {
                "vendor_name_cell":  "B1",
                "customer_cell":     "B2",
                "header_row":        7,    # STT | Dept. | PR Person | SEV PO No | BQMS CODE | ITEM NAME | SPEC | UNIT | Box WEIGHT | DIM L | DIM W | DIM H
                "first_data_row":    8,
                "total_row_offset":  -1,   # last row = TOTAL
                "cols": {
                    "stt":         "B",
                    "dept":        "C",
                    "pr_person":   "D",
                    "po_no":       "E",
                    "bqms_code":   "F",
                    "item_name":   "G",
                    "specification": "H",
                    "unit":        "I",
                    "box_weight":  "J",
                    "dim_l":       "K",
                    "dim_w":       "L",
                    "dim_h":       "M"
                }
            },
            "list_detail": {
                "header_row":     1,
                "first_data_row": 3,
                "total_row_offset": -1,
                "cols": {
                    "no":           "B",
                    "item_code":    "C",
                    "name_specs":   "D",
                    "image":        "G",   # embed image here
                    "unit":         "H",
                    "qty":          "I",
                    "remark":       "J"
                }
            },
            "label": {
                # Each PO = 1 label block of 6 rows: Vendor / PR Person / PO No / [blank header] / BQMS / [blank]
                "block_height":     6,
                "first_block_row":  1,
                "cells_in_block": {
                    "vendor":    "C+0",   # offset row 0
                    "pr_person": "C+1",
                    "po_no":     "C+2",
                    "bqms_code": "B+4",
                    "qty":       "C+4"
                }
            },
            "tong_hop": {
                # Static checklist sheet — no per-job fill needed. Could optionally
                # update Shipping date in some cells if desired.
                "static": True
            }
        }
    }

    Path(OUTPUT_CELLS_JSON).parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CELLS_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nCells JSON saved: {OUTPUT_CELLS_JSON}")

    # Verify reload
    wb = openpyxl.load_workbook(OUTPUT_TEMPLATE)
    print(f"\nVerify reload xlsx sheets: {wb.sheetnames}")
    print(f"Total value cells: {sum(len(s['value_cells']) for s in meta.values())}")
    print("\n=== DONE ===")
    print(f"\nNext: register template in DB:")
    print(f"  INSERT INTO quotation_templates (name, template_type, file_path, is_default)")
    print(f"  VALUES ('AMA Delivery Dossier', 'delivery_dossier',")
    print(f"          '{OUTPUT_TEMPLATE}', true);")


if __name__ == "__main__":
    main()
