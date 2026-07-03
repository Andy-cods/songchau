"""Migrate Google Sheet "PIM of Thong ke hoi hang" → sourcing_entries.

One-shot import script. Pattern:
1. Đọc sheet "TỔNG HỢP" (35 cột, ~14k rows)
2. Parse từng row: date heuristic, currency 2-mode, supplier regex
3. UPSERT vào sourcing_entries
4. Skip empty/generic rows (SPRING, CUTTER, GEAR, ...)
5. created_by_email='migrate_pim_rfq.py' để dễ rollback

Usage:
    docker exec sc-worker python /app/scripts/migrate_pim_rfq.py \\
        --source /data/onedrive-staging/Puplic/IMV/Thong\\ ke\\ hoi\\ hang\\ -\\ update\\ 240424.xlsx \\
        --dry-run --verbose

    docker exec sc-worker python /app/scripts/migrate_pim_rfq.py \\
        --source /data/onedrive-staging/Puplic/IMV/Thong\\ ke\\ hoi\\ hang\\ -\\ update\\ 240424.xlsx
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("migrate_pim_rfq")

# 35-cột positional layout (verified từ workflow sheet analysis)
COLS = [
    "customer_name",       # 0  Tên KH
    "person_in_charge",    # 1  Ng phụ trách
    "model",               # 2  Model
    "product_name",        # 3  Tên sp
    "maker",               # 4  Maker
    "inquiry_date",        # 5  Ngày hỏi giá (m/d/y)
    "cost_jpy",            # 6  Giá nhập Yên Nhật
    "cost_usd",            # 7  Giá nhập USD
    "cost_krw",            # 8  Giá nhập Won
    "cost_rmb",            # 9  Giá nhập RMB
    "cost_vnd",            # 10 Giá nhập VND
    "sale_vnd",            # 11 Giá bán
    "quantity",            # 12 Số lượng
    "tax_pct",             # 13 Thuế xuất
    "hs_code",             # 14 HS Code
    "weight_kg",           # 15 Cân nặng
    "notes",               # 16 Ghi chú
    "coefficient",         # 17 Hệ số
    "supplier_raw",        # 18 Nhà cung cấp (parse → name + phone + email)
    "image_raw",           # 19 Hình ảnh (Shopee junk → skip)
    "exchange_rate_raw",   # 20 Tỷ giá
    "notes_internal",      # 21 Note SLL
    "row_classification",  # 22 Row Classification
    "catalog_category",    # 23 Catalog Category
    "normalized_model",    # 24 Normalized Model
    "brand_canonical",     # 25 Brand (Canonical)
    "part_type",           # 26 Part Type
    "subcategory_slug",    # 27 Subcategory
    "machine_model",       # 28 Machine Model
    "catalog_status",      # 29 Catalog Status
    "stage_raw",           # 30 Stage
    "image_url",           # 31 Ảnh URL (clean)
    "missing_fields_raw",  # 32 Thiếu gì?
    "updated_at_raw",      # 33 Ngày cập nhật
    "missing_count_raw",   # 34 Số trường thiếu
]

# Generic placeholders để skip
GENERIC_MODELS = {
    "SPRING", "CUTTER", "GEAR", "SENSOR", "PCB", "APTOMAT",
    "SPROCKET", "BELT", "BEARING", "MOTOR", "VALVE", "FUSE",
    "CABLE", "WIRE", "SCREW", "BOLT", "NUT", "WASHER",
}

# Currency code → cost column mapping
CURRENCY_COLS = {
    "JPY": "cost_jpy",
    "USD": "cost_usd",
    "KRW": "cost_krw",
    "RMB": "cost_rmb",
    "CNY": "cost_rmb",
    "VND": "cost_vnd",
}


# ─── Parsers ──────────────────────────────────────────────────────


def safe_str(v: Any, max_len: int = 500) -> str | None:
    """Strip + truncate, None if empty."""
    if v is None:
        return None
    s = str(v).strip()
    return s[:max_len] if s else None


def parse_date(v: Any) -> date | None:
    """Parse 3 formats: datetime / Excel serial / d/m/Y / m/d/Y (heuristic day>12 → d/m/y)."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial: days since 1899-12-30
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(v))).date()
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    # Try ISO first
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # 2 forms with delimiter / or -
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$", s)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        # Heuristic: if a > 12 → must be day → d/m/y
        # else if b > 12 → must be day → m/d/y
        # else default to m/d/y (US format from Google Sheet)
        if a > 12:
            day, month = a, b
        elif b > 12:
            day, month = b, a
        else:
            # Ambiguous — default to d/m/y (Vietnam format)
            day, month = a, b
        try:
            return date(y, month, day)
        except ValueError:
            try:
                return date(y, day, month)  # fallback swap
            except ValueError:
                return None
    return None


_NUM_CLEAN = re.compile(r"[^\d,.\-]")


def parse_number(v: Any) -> float | None:
    """Parse VN '1.260.000' OR US '1,260,000.50' OR '550,00' heuristic."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = _NUM_CLEAN.sub("", s)
    if not s:
        return None
    has_dot, has_comma = "." in s, "," in s
    try:
        if has_dot and has_comma:
            # Use rightmost separator as decimal
            if s.rfind(".") > s.rfind(","):
                # US: 1,234.56
                s = s.replace(",", "")
            else:
                # VN: 1.234,56
                s = s.replace(".", "").replace(",", ".")
        elif has_comma:
            # Could be VN decimal (550,00) or US thousands (1,234)
            # Heuristic: if 2 digits after last comma → decimal
            after = s.split(",")[-1]
            if len(after) == 2:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        # else just dots — could be VN thousands (1.260.000) or US decimal (1.5)
        elif has_dot:
            parts = s.split(".")
            # If multiple dots → VN thousands separator
            if len(parts) > 2:
                s = s.replace(".", "")
            # else assume decimal
        return float(s)
    except Exception:
        return None


_PHONE_RE = re.compile(r"(0\d{9,10}|\+84\d{8,10})")
_EMAIL_RE = re.compile(r"[\w.\-]+@[\w.\-]+\.\w+")


def parse_supplier(v: Any) -> tuple[str | None, str | None, str | None]:
    """Parse '名前 0909xxx email@x.com' → (name, phone, email)."""
    if v is None:
        return None, None, None
    s = str(v).strip()
    if not s:
        return None, None, None
    phone = None
    email = None
    m = _EMAIL_RE.search(s)
    if m:
        email = m.group(0)
        s = s.replace(email, "").strip()
    m = _PHONE_RE.search(s)
    if m:
        phone = m.group(0)
        s = s.replace(phone, "").strip()
    name = re.sub(r"[,\s]+$", "", s).strip(" ,;")
    return (name or None), phone, email


def parse_stage(v: Any) -> int | None:
    """Stage = 1/2/3 only."""
    if v is None:
        return None
    try:
        n = int(float(str(v).strip()))
        if n in (1, 2, 3):
            return n
    except Exception:
        pass
    return None


def parse_catalog_status(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip().upper().replace(" ", "_")
    if s in {"OK", "NEEDS_BRAND", "NOT_IN_CATALOG", "PRODUCT_CANDIDATE"}:
        return s
    return None


def parse_missing_fields(v: Any) -> list[str] | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    parts = [p.strip() for p in re.split(r"[,;]", s) if p.strip()]
    return parts or None


def parse_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip().replace(",", "")))
    except Exception:
        return None


def cap_numeric(v: float | None, max_abs: float) -> float | None:
    """Cap value to fit DB NUMERIC precision. Return None if exceed (let user fix in UI)."""
    if v is None:
        return None
    if abs(v) >= max_abs:
        return None
    return v


# ─── Row mapper ────────────────────────────────────────────────


def map_row(raw: tuple) -> dict | None:
    """Map 35-col row → kwargs dict. Returns None if row should skip."""
    # Pad if row is short
    raw = raw + (None,) * (35 - len(raw))
    if len(raw) > 35:
        raw = raw[:35]

    customer_name = safe_str(raw[0], 200)
    model = safe_str(raw[2], 200)
    product_name = safe_str(raw[3], 500)

    # Skip if empty model AND empty product
    if not model and not product_name:
        return None

    # Skip generic placeholders
    if model and model.upper() in GENERIC_MODELS:
        return {"_skip_generic": True, "model": model}

    supplier_name, supplier_phone, supplier_email = parse_supplier(raw[18])

    return {
        "customer_name": customer_name,
        "person_in_charge": safe_str(raw[1], 50),
        "model": model,
        "product_name": product_name,
        "maker": safe_str(raw[4], 200),
        "inquiry_date": parse_date(raw[5]),
        # Cap về precision DB:
        # cost_jpy/usd/krw/rmb (18,2): max 10^16
        # cost_vnd/sale_vnd (18,0): max 10^18
        # quantity (18,3): max 10^15
        # tax_pct (6,2): max 10^4
        # weight_kg (12,3): max 10^9
        # coefficient (8,4): max 10^4
        "cost_jpy": cap_numeric(parse_number(raw[6]), 1e16),
        "cost_usd": cap_numeric(parse_number(raw[7]), 1e16),
        "cost_krw": cap_numeric(parse_number(raw[8]), 1e16),
        "cost_rmb": cap_numeric(parse_number(raw[9]), 1e16),
        "cost_vnd": cap_numeric(parse_number(raw[10]), 1e18),
        "sale_vnd": cap_numeric(parse_number(raw[11]), 1e18),
        "quantity": cap_numeric(parse_number(raw[12]), 1e15),
        "tax_pct": cap_numeric(parse_number(raw[13]), 1e4),
        "hs_code": safe_str(raw[14], 30),
        "weight_kg": cap_numeric(parse_number(raw[15]), 1e9),
        "notes": safe_str(raw[16], 2000),
        "coefficient": cap_numeric(parse_number(raw[17]), 1e4),
        "supplier_name": supplier_name,
        "supplier_phone": supplier_phone,
        "supplier_email": supplier_email,
        "image_url": safe_str(raw[31], 1000),
        "notes_internal": safe_str(raw[21], 2000),
        "row_classification": safe_str(raw[22], 100),
        "catalog_category": safe_str(raw[23], 100),
        "normalized_model": safe_str(raw[24], 200),
        "brand_canonical": safe_str(raw[25], 100),
        "part_type": safe_str(raw[26], 100),
        "subcategory_slug": safe_str(raw[27], 100),
        "machine_model": safe_str(raw[28], 100),
        "catalog_status": parse_catalog_status(raw[29]),
        "stage": parse_stage(raw[30]),
        "missing_fields": parse_missing_fields(raw[32]),
        "missing_count": parse_int(raw[34]),
    }


# ─── DB insert ────────────────────────────────────────────────


INSERT_SQL = """
INSERT INTO sourcing_entries (
    customer_name, person_in_charge, model, product_name, maker,
    inquiry_date, cost_jpy, cost_usd, cost_krw, cost_rmb, cost_vnd,
    sale_vnd, quantity, tax_pct, hs_code, weight_kg, notes, coefficient,
    supplier_name, supplier_phone, supplier_email, image_url, notes_internal,
    row_classification, catalog_category, normalized_model, brand_canonical,
    part_type, subcategory_slug, machine_model, catalog_status, stage,
    missing_fields, missing_count, created_by_email, created_at, updated_at
) VALUES (
    %(customer_name)s, %(person_in_charge)s, %(model)s, %(product_name)s, %(maker)s,
    %(inquiry_date)s, %(cost_jpy)s, %(cost_usd)s, %(cost_krw)s, %(cost_rmb)s, %(cost_vnd)s,
    %(sale_vnd)s, %(quantity)s, %(tax_pct)s, %(hs_code)s, %(weight_kg)s, %(notes)s, %(coefficient)s,
    %(supplier_name)s, %(supplier_phone)s, %(supplier_email)s, %(image_url)s, %(notes_internal)s,
    %(row_classification)s, %(catalog_category)s, %(normalized_model)s, %(brand_canonical)s,
    %(part_type)s, %(subcategory_slug)s, %(machine_model)s, %(catalog_status)s, %(stage)s,
    %(missing_fields)s, %(missing_count)s, %(created_by_email)s, NOW(), NOW()
)
"""


def main() -> int:
    ap = argparse.ArgumentParser(description="Migrate PIM Hỏi hàng sheet → sourcing_entries")
    ap.add_argument("--source", required=True, help="Path tới .xlsx")
    ap.add_argument("--sheet", default="TỔNG HỢP", help="Tên sheet (default: TỔNG HỢP)")
    ap.add_argument("--header-row", type=int, default=1, help="0-indexed header row (default 1)")
    ap.add_argument("--dry-run", action="store_true", help="Không commit, chỉ stats")
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--limit", type=int, default=0, help="Limit rows (0 = all)")
    args = ap.parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)

    src = Path(args.source)
    if not src.exists():
        logger.error("File not found: %s", src)
        return 1

    logger.info("Load workbook: %s", src)
    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    sheet_name = args.sheet
    if sheet_name not in wb.sheetnames:
        # Try variants
        for candidate in wb.sheetnames:
            cl = candidate.strip().upper().replace(" ", "")
            if cl in {"TỔNGHỢP", "TONGHOP", "TONGHOPMASTER", "TỔNGHỢPMASTER"}:
                sheet_name = candidate
                break
        else:
            logger.error("Sheet %r not found. Available: %s", args.sheet, wb.sheetnames)
            return 1

    ws = wb[sheet_name]
    logger.info("Sheet: %r, scanning rows...", sheet_name)

    rows_data = []
    skip_empty = 0
    skip_generic = 0
    total_seen = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < args.header_row:
            continue
        if args.limit and len(rows_data) >= args.limit:
            break
        total_seen += 1
        mapped = map_row(row)
        if mapped is None:
            skip_empty += 1
            continue
        if mapped.get("_skip_generic"):
            skip_generic += 1
            if args.verbose:
                logger.debug("Skip generic: %s", mapped.get("model"))
            continue
        mapped["created_by_email"] = "migrate_pim_rfq.py"
        rows_data.append(mapped)
        if (len(rows_data) % 1000) == 0:
            logger.info("  Parsed %d rows...", len(rows_data))

    wb.close()
    logger.info(
        "Parse done: seen=%d, parsed=%d, skip_empty=%d, skip_generic=%d",
        total_seen, len(rows_data), skip_empty, skip_generic,
    )

    if args.dry_run:
        # Sample 5 first + 5 last
        logger.info("=== DRY RUN — sample first 5 rows ===")
        for r in rows_data[:5]:
            logger.info("  %s | model=%r maker=%r date=%s sale=%s",
                        r.get("customer_name"), r.get("model"), r.get("maker"),
                        r.get("inquiry_date"), r.get("sale_vnd"))
        logger.info("=== sample last 5 rows ===")
        for r in rows_data[-5:]:
            logger.info("  %s | model=%r maker=%r date=%s sale=%s",
                        r.get("customer_name"), r.get("model"), r.get("maker"),
                        r.get("inquiry_date"), r.get("sale_vnd"))
        # Stats
        cur_present = sum(1 for r in rows_data if any(
            r.get(f"cost_{c}") for c in ("jpy", "usd", "krw", "rmb", "vnd")
        ))
        date_present = sum(1 for r in rows_data if r.get("inquiry_date"))
        cat_present = sum(1 for r in rows_data if r.get("catalog_category"))
        logger.info(
            "Stats: cost_present=%d (%.1f%%), date_present=%d (%.1f%%), catalog_cat=%d (%.1f%%)",
            cur_present, 100 * cur_present / max(1, len(rows_data)),
            date_present, 100 * date_present / max(1, len(rows_data)),
            cat_present, 100 * cat_present / max(1, len(rows_data)),
        )
        logger.info("DRY-RUN done. KHÔNG commit DB.")
        return 0

    # Real insert
    dsn = os.environ.get(
        "SYNC_DSN",
        "postgresql://scadmin:SC2026_ERP_Pr0d_X9k2mQ7wR4@postgres:5432/songchau_erp",
    )
    logger.info("Connect DB...")
    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    inserted = 0
    failed = 0
    failed_samples: list[str] = []
    try:
        for i, row in enumerate(rows_data):
            try:
                with conn.cursor() as cur:
                    cur.execute(INSERT_SQL, row)
                conn.commit()
                inserted += 1
            except Exception as exc:
                conn.rollback()
                failed += 1
                if len(failed_samples) < 5:
                    failed_samples.append(
                        f"row#{i} model={row.get('model')!r}: {str(exc)[:200]}"
                    )
            if (i + 1) % 1000 == 0:
                logger.info("Progress: %d/%d (inserted=%d, failed=%d)",
                            i + 1, len(rows_data), inserted, failed)
    finally:
        conn.close()

    logger.info("Done. inserted=%d, failed=%d, skip_empty=%d, skip_generic=%d",
                inserted, failed, skip_empty, skip_generic)
    if failed_samples:
        logger.warning("Failed sample (first 5):")
        for s in failed_samples:
            logger.warning("  %s", s)
    return 0


if __name__ == "__main__":
    sys.exit(main())
