"""Enrich existing sourcing_entries với enrichment data từ PIM workbook 35-col.

Idempotent UPDATE-only migration. KHÔNG overwrite cost/sale/quantity/supplier/notes.
Chỉ fill enrichment columns: catalog_category, brand_canonical, part_type,
subcategory_slug, machine_model, catalog_status, stage, row_classification,
normalized_model, missing_fields, missing_count, image_url (if NULL).

Match strategy:
  Primary key: (model_norm, customer_name_norm)
  Disambiguation: inquiry_date exact → cost_signature closest → broadcast

Usage:
    docker exec sc-worker python /app/scripts/enrich_pim_rfq.py \\
        --source /data/pim/PIM-of-Thong-ke-hoi-hang-update-240424.xlsm \\
        --dry-run --verbose

    docker exec sc-worker python /app/scripts/enrich_pim_rfq.py \\
        --source /data/pim/PIM-of-Thong-ke-hoi-hang-update-240424.xlsm \\
        --batch-size 500

Args:
    --source         Path to .xlsm / .xlsx PIM workbook (required)
    --dry-run        Compute matches + print stats, no UPDATEs
    --verbose        DEBUG-level logs (per-row decisions)
    --limit N        Process only first N existing DB rows (testing)
    --batch-size N   Commit interval (default 500)
    --sheet NAME     Override sheet name (default: "TONG HOP")
    --created-by E   Filter DB rows by created_by_email (default: migrate_pim_rfq.py)
    --no-audit       Skip writing pim_enrichment_audit rows (faster, less safe)
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import sys
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import psycopg2
import psycopg2.extras

# ─── Logging ──────────────────────────────────────────────────────

LOG_DIR = Path(os.getenv("PIM_ENRICH_LOG_DIR", "/tmp"))
LOG_DIR.mkdir(parents=True, exist_ok=True)
RUN_ID = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
LOG_FILE = LOG_DIR / f"enrich_pim_run_{RUN_ID}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
    ],
)
logger = logging.getLogger("enrich_pim_rfq")


# ─── Constants ────────────────────────────────────────────────────

DEFAULT_SHEET = "TONG HOP"
DEFAULT_CREATED_BY = "migrate_pim_rfq.py"
EXPECTED_COLS = 35

# Positional column indices (0-based) in PIM sheet
COL_CUSTOMER = 0
COL_OWNER = 1
COL_MODEL = 2
COL_PRODUCT_NAME = 3
COL_MAKER = 4
COL_INQUIRY_DATE = 5
COL_COST_JPY = 6
COL_COST_USD = 7
COL_COST_KRW = 8
COL_COST_RMB = 9
COL_COST_VND = 10
COL_SALE_VND = 11
COL_QUANTITY = 12
# 13 tax_pct (CORRUPTED — ignore)
# 14 hs_code (CORRUPTED — ignore)
# 15 weight, 16 notes, 17 coefficient (CORRUPTED), 18 supplier, 19 image (junk)
# 20 fx, 21 notes_internal
COL_ROW_CLASS = 22
COL_CATALOG_CATEGORY = 23
COL_NORMALIZED_MODEL = 24  # IGNORED — we re-compute (Finding #7)
COL_BRAND_CANONICAL = 25
COL_PART_TYPE = 26
COL_SUBCATEGORY = 27
COL_MACHINE_MODEL = 28
COL_CATALOG_STATUS = 29
COL_STAGE = 30
COL_IMAGE_URL = 31  # always empty in PIM but honor "only if NULL" rule
COL_MISSING_FIELDS = 32
# 33 updated_at (always empty)
COL_MISSING_COUNT = 34

# Closed enum sets — for validation logging only
VALID_ROW_CLASS = {
    "Product Candidate", "Empty Model", "RFQ / Repeat",
    "Empty Row", "Generic Model",
}
VALID_CATALOG_CATEGORIES = {
    "Uncategorized", "Electronics Components", "SMT Machine Parts",
    "Tools, Consumables & Safety", "Motors, Drives & Motion",
    "Sensors & Instrumentation", "Pneumatic & Hydraulic",
    "Power Supplies & Electrical", "Connectors, Cables & Wiring",
    "Industrial Automation",
}
VALID_CATALOG_STATUS = {"OK", "NOT IN CATALOG", "NEEDS TYPE", "NEEDS BRAND"}

# DB check constraint allows: OK, NEEDS_BRAND, NOT_IN_CATALOG, PRODUCT_CANDIDATE
# PIM "NEEDS TYPE" has no DB equivalent → coerce to None (leave NULL).
CATALOG_STATUS_DB_MAP: dict[str, str | None] = {
    "OK": "OK",
    "NOT IN CATALOG": "NOT_IN_CATALOG",
    "NEEDS BRAND": "NEEDS_BRAND",
    "NEEDS TYPE": None,  # no DB enum value — skip
}

# DB stage constraint: 1, 2, 3 only
VALID_STAGES_DB = {1, 2, 3}
VALID_SUBCATEGORIES = {
    "mechanical-parts", "resistors", "capacitors", "ics", "nozzles",
    "belts-chains", "feeders", "diodes-transistors", "smt-motors",
    "circuit-boards", "smt-sensors", "smt-filters", "leds", "inductors",
    "circuit-protection", "cameras-vision",
}
VALID_MACHINE_MODELS = {
    "YS/YSM", "CM402/CM602", "NXT/AIM", "CM/NPM", "NPM", "SM",
    "Reflow Oven", "YS24", "YSM20", "YS12", "NXT",
}

FORMULA_ERRORS = {"#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}

# Enrichment fields tracked for completeness scoring (canonical pick)
ENRICHMENT_FIELDS = [
    "row_classification", "catalog_category", "brand_canonical",
    "part_type", "subcategory_slug", "machine_model", "catalog_status",
    "stage", "missing_fields", "missing_count",
]

# Update precedence rules
ALWAYS_OVERWRITE = {
    "row_classification", "catalog_category", "brand_canonical",
    "part_type", "subcategory_slug", "machine_model", "catalog_status",
    "stage", "normalized_model", "missing_fields", "missing_count",
}
ONLY_IF_NULL = {"image_url"}


# ─── Normalization ────────────────────────────────────────────────

_MODEL_SEP_RE = re.compile(r"[/,;|]| - |\s{2,}")
_WS_RE = re.compile(r"\s+")


def normalize_model(raw: Any) -> str:
    """Canonical model key. Splits multi-models, takes first token, uppercases.

    Examples:
      "2EGTBC030200/ 2EGTBC0302" → "2EGTBC030200"
      "\\nFLASH DISK 2G KLA-M4255-001" → "FLASHDISK2GKLA-M4255-001"
      "  abc-123  " → "ABC-123"
    """
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    if not s:
        return ""
    # Replace newlines + tabs with spaces
    s = s.replace("\n", " ").replace("\t", " ")
    # Split on separators; take first non-empty token
    parts = _MODEL_SEP_RE.split(s)
    parts = [p.strip() for p in parts if p.strip()]
    if not parts:
        return ""
    token = parts[0]
    # Remove all internal whitespace (model codes don't have spaces)
    token = _WS_RE.sub("", token)
    return token


def normalize_customer(raw: Any) -> str:
    """Customer name → lowercase, collapsed whitespace."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    s = s.replace("\n", " ").replace("\t", " ")
    s = _WS_RE.sub(" ", s)
    return s.lower()


def safe_str(v: Any, max_len: int = 500) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in FORMULA_ERRORS:
        return None
    return s[:max_len]


def coerce_int(v: Any) -> int | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in FORMULA_ERRORS:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def coerce_missing_fields(v: Any) -> list[str] | None:
    """'Maker, Category, Brand, Part Type,' → ['Maker','Category','Brand','Part Type'].

    DB column is text[] (ARRAY). Trim trailing comma, split on ',', strip whitespace,
    drop empties. Returns None if input is null/empty/formula-error.
    """
    s = safe_str(v)
    if s is None:
        return None
    parts = [p.strip() for p in s.split(",")]
    parts = [p for p in parts if p]
    return parts or None


def parse_inquiry_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
        except Exception:
            return None
    s = str(v).strip()
    if not s or s in FORMULA_ERRORS:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in FORMULA_ERRORS:
        return None
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s:
        return None
    try:
        if "." in s and "," in s:
            if s.rfind(".") > s.rfind(","):
                s = s.replace(",", "")
            else:
                s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            after = s.split(",")[-1]
            if len(after) == 2:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        elif s.count(".") > 1:
            s = s.replace(".", "")
        return float(s)
    except Exception:
        return None


def cost_signature(jpy: Any, usd: Any, vnd: Any) -> float:
    """Stable scalar fingerprint for disambiguating multi-RFQ rows."""
    j = parse_number(jpy) or 0.0
    u = parse_number(usd) or 0.0
    v = parse_number(vnd) or 0.0
    return round(j + u * 100.0 + v / 1000.0, 2)


# ─── PIM Row ──────────────────────────────────────────────────────


class PimRow:
    """Parsed PIM workbook row."""

    __slots__ = (
        "row_idx", "raw", "model_norm", "customer_norm",
        "inquiry_date", "cost_sig",
        "row_classification", "catalog_category", "normalized_model_local",
        "brand_canonical", "part_type", "subcategory_slug",
        "machine_model", "catalog_status", "stage",
        "image_url", "missing_fields", "missing_count",
    )

    def __init__(self, row_idx: int, raw_cells: tuple):
        self.row_idx = row_idx
        self.raw = raw_cells

        model_raw = raw_cells[COL_MODEL] if len(raw_cells) > COL_MODEL else None
        customer_raw = raw_cells[COL_CUSTOMER] if len(raw_cells) > COL_CUSTOMER else None

        self.model_norm = normalize_model(model_raw)
        self.customer_norm = normalize_customer(customer_raw)
        self.normalized_model_local = self.model_norm  # we own this, ignore col 24

        self.inquiry_date = parse_inquiry_date(
            raw_cells[COL_INQUIRY_DATE] if len(raw_cells) > COL_INQUIRY_DATE else None
        )
        self.cost_sig = cost_signature(
            raw_cells[COL_COST_JPY] if len(raw_cells) > COL_COST_JPY else None,
            raw_cells[COL_COST_USD] if len(raw_cells) > COL_COST_USD else None,
            raw_cells[COL_COST_VND] if len(raw_cells) > COL_COST_VND else None,
        )

        self.row_classification = safe_str(raw_cells[COL_ROW_CLASS]) if len(raw_cells) > COL_ROW_CLASS else None
        self.catalog_category = safe_str(raw_cells[COL_CATALOG_CATEGORY]) if len(raw_cells) > COL_CATALOG_CATEGORY else None
        self.brand_canonical = safe_str(raw_cells[COL_BRAND_CANONICAL]) if len(raw_cells) > COL_BRAND_CANONICAL else None
        self.part_type = safe_str(raw_cells[COL_PART_TYPE]) if len(raw_cells) > COL_PART_TYPE else None
        self.subcategory_slug = safe_str(raw_cells[COL_SUBCATEGORY]) if len(raw_cells) > COL_SUBCATEGORY else None
        self.machine_model = safe_str(raw_cells[COL_MACHINE_MODEL]) if len(raw_cells) > COL_MACHINE_MODEL else None
        raw_status = safe_str(raw_cells[COL_CATALOG_STATUS]) if len(raw_cells) > COL_CATALOG_STATUS else None
        # Map PIM's spaced values → DB underscored enum. Unknown → None.
        self.catalog_status = (
            CATALOG_STATUS_DB_MAP.get(raw_status.upper(), None)
            if raw_status else None
        )
        raw_stage = coerce_int(raw_cells[COL_STAGE]) if len(raw_cells) > COL_STAGE else None
        # DB constraint: only 1/2/3 allowed. Coerce anything else to None.
        self.stage = raw_stage if raw_stage in VALID_STAGES_DB else None
        self.image_url = safe_str(raw_cells[COL_IMAGE_URL]) if len(raw_cells) > COL_IMAGE_URL else None
        self.missing_fields = coerce_missing_fields(
            raw_cells[COL_MISSING_FIELDS] if len(raw_cells) > COL_MISSING_FIELDS else None
        )
        self.missing_count = coerce_int(
            raw_cells[COL_MISSING_COUNT] if len(raw_cells) > COL_MISSING_COUNT else None
        )

    def is_enrichment_touched(self) -> bool:
        """At least one enrichment field populated?"""
        return self.row_classification is not None or self.catalog_category is not None

    def has_match_key(self) -> bool:
        return bool(self.model_norm) and bool(self.customer_norm)

    def completeness(self) -> int:
        """Count of non-null enrichment fields (for canonical picking)."""
        return sum(
            1 for f in ENRICHMENT_FIELDS
            if getattr(self, f) is not None
        )

    def enrichment_dict(self) -> dict[str, Any]:
        return {
            "row_classification": self.row_classification,
            "catalog_category": self.catalog_category,
            "normalized_model": self.normalized_model_local,
            "brand_canonical": self.brand_canonical,
            "part_type": self.part_type,
            "subcategory_slug": self.subcategory_slug,
            "machine_model": self.machine_model,
            "catalog_status": self.catalog_status,
            "stage": self.stage,
            "image_url": self.image_url,
            "missing_fields": self.missing_fields,
            "missing_count": self.missing_count,
        }


# ─── PIM Loader ──────────────────────────────────────────────────


def load_pim_rows(path: Path, sheet_name: str) -> list[PimRow]:
    logger.info("Loading PIM workbook: %s", path)
    wb = openpyxl.load_workbook(
        path, data_only=True, keep_vba=False, read_only=True
    )
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows: list[PimRow] = []
    header_seen = False
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not header_seen:
            header_seen = True
            if not row or not str(row[0] or "").strip().startswith("Tên KH"):
                logger.warning("Header row unexpected at row 1: %r", row[:3])
            continue
        if row is None:
            continue
        pim = PimRow(idx, row)
        rows.append(pim)
    wb.close()
    logger.info("Loaded %d PIM data rows", len(rows))
    return rows


def build_pim_lookup(
    pim_rows: Iterable[PimRow],
) -> tuple[dict[tuple[str, str], list[PimRow]], dict[str, int]]:
    """Build (model_norm, customer_norm) → list[PimRow] index.

    Only includes rows that:
      - have a match key (non-empty model + customer)
      - have at least one enrichment field populated
    """
    lookup: dict[tuple[str, str], list[PimRow]] = defaultdict(list)
    stats = {
        "total_pim_rows": 0,
        "no_match_key": 0,
        "no_enrichment": 0,
        "indexed_rows": 0,
    }
    for pim in pim_rows:
        stats["total_pim_rows"] += 1
        if not pim.has_match_key():
            stats["no_match_key"] += 1
            continue
        if not pim.is_enrichment_touched():
            stats["no_enrichment"] += 1
            continue
        key = (pim.model_norm, pim.customer_norm)
        lookup[key].append(pim)
        stats["indexed_rows"] += 1
    stats["distinct_keys"] = len(lookup)
    return lookup, stats


# ─── DB Layer ────────────────────────────────────────────────────


def get_dsn() -> str:
    dsn = os.getenv("DATABASE_URL") or os.getenv("POSTGRES_DSN")
    if dsn:
        # psycopg2 is sync — strip async drivers if present
        for prefix in ("postgresql+asyncpg://", "postgresql+psycopg://",
                       "postgres+asyncpg://"):
            if dsn.startswith(prefix):
                dsn = "postgresql://" + dsn[len(prefix):]
                break
        return dsn
    # Build from individual vars
    host = os.getenv("PGHOST", "localhost")
    port = os.getenv("PGPORT", "5432")
    user = os.getenv("PGUSER", "postgres")
    pwd = os.getenv("PGPASSWORD", "")
    db = os.getenv("PGDATABASE", "songchau")
    return f"postgresql://{user}:{pwd}@{host}:{port}/{db}"


def fetch_existing_rows(
    conn, created_by_email: str, limit: int | None
) -> list[dict]:
    """Fetch sourcing_entries rows that need enrichment."""
    sql = """
        SELECT
            id,
            model,
            customer_name,
            inquiry_date,
            cost_jpy, cost_usd, cost_vnd,
            row_classification,
            catalog_category, normalized_model, brand_canonical,
            part_type, subcategory_slug, machine_model,
            catalog_status, stage,
            image_url, missing_fields, missing_count
        FROM sourcing_entries
        WHERE created_by_email = %s
        ORDER BY id
    """
    params: tuple = (created_by_email,)
    if limit is not None:
        sql += " LIMIT %s"
        params = (created_by_email, limit)
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]
    logger.info("Fetched %d existing sourcing_entries (created_by=%s)",
                len(rows), created_by_email)
    return rows


def ensure_audit_table(conn) -> None:
    sql = """
    CREATE TABLE IF NOT EXISTS pim_enrichment_audit (
        id            BIGSERIAL PRIMARY KEY,
        run_id        TEXT NOT NULL,
        entry_id      BIGINT NOT NULL,
        pim_row_idx   INTEGER,
        match_tier    TEXT NOT NULL,
        before_json   JSONB NOT NULL,
        after_json    JSONB NOT NULL,
        applied_at    TIMESTAMPTZ NOT NULL DEFAULT now()
    );
    CREATE INDEX IF NOT EXISTS pim_enrich_audit_run_idx
        ON pim_enrichment_audit(run_id);
    CREATE INDEX IF NOT EXISTS pim_enrich_audit_entry_idx
        ON pim_enrichment_audit(entry_id);
    """
    with conn.cursor() as cur:
        cur.execute(sql)
    conn.commit()
    logger.info("Audit table ready")


def detect_schema_columns(conn) -> set[str]:
    """Return set of column names present on sourcing_entries."""
    sql = """
        SELECT column_name FROM information_schema.columns
        WHERE table_name = 'sourcing_entries'
    """
    with conn.cursor() as cur:
        cur.execute(sql)
        return {r[0] for r in cur.fetchall()}


# ─── Matching ────────────────────────────────────────────────────


def disambiguate(
    db_row: dict,
    candidates: list[PimRow],
) -> tuple[PimRow, str]:
    """Pick best PIM row from candidates. Returns (pim, match_tier)."""
    if len(candidates) == 1:
        return candidates[0], "unique"

    # Tier 1: exact inquiry_date match
    db_date = db_row.get("inquiry_date")
    if db_date is not None:
        date_hits = [p for p in candidates if p.inquiry_date == db_date]
        if len(date_hits) == 1:
            return date_hits[0], "date_exact"
        if date_hits:
            candidates = date_hits  # narrow

    # Tier 2: closest cost signature
    db_sig = cost_signature(
        db_row.get("cost_jpy"), db_row.get("cost_usd"), db_row.get("cost_vnd")
    )
    if db_sig > 0:
        candidates_with_cost = [
            (abs(p.cost_sig - db_sig), p) for p in candidates if p.cost_sig > 0
        ]
        if candidates_with_cost:
            candidates_with_cost.sort(key=lambda t: t[0])
            best_delta = candidates_with_cost[0][0]
            # Accept if delta < 1% of db_sig OR best is clearly closest
            if best_delta < max(db_sig * 0.01, 1.0):
                return candidates_with_cost[0][1], "cost_closest"

    # Tier 3: pick canonical (richest completeness, latest date, stable idx)
    canonical = max(
        candidates,
        key=lambda r: (
            r.completeness(),
            r.inquiry_date or date.min,
            -r.row_idx,
        ),
    )
    return canonical, "canonical_pick"


# ─── Update Execution ────────────────────────────────────────────


def build_update_plan(
    db_row: dict,
    pim: PimRow,
    schema_cols: set[str],
) -> dict[str, Any]:
    """Compute field deltas. Empty dict = no-op."""
    enrichment = pim.enrichment_dict()
    updates: dict[str, Any] = {}
    for field, new_val in enrichment.items():
        if field not in schema_cols:
            continue
        if new_val is None:
            continue
        current = db_row.get(field)
        if field in ONLY_IF_NULL:
            if current is None:
                updates[field] = new_val
            continue
        if field in ALWAYS_OVERWRITE:
            if current != new_val:
                updates[field] = new_val
            continue
    return updates


def execute_updates(
    conn,
    update_specs: list[tuple[int, dict[str, Any], dict[str, Any], PimRow, str]],
    schema_cols: set[str],
    run_id: str,
    write_audit: bool,
    batch_size: int,
) -> int:
    """Apply updates in batches. Each spec = (entry_id, before, updates, pim, tier).

    Returns rows actually updated.
    """
    if not update_specs:
        return 0
    updated = 0
    audit_rows: list[tuple] = []
    with conn.cursor() as cur:
        for entry_id, before, updates, pim, tier in update_specs:
            if not updates:
                continue
            set_frags = [f"{col} = %s" for col in updates.keys()]
            params = list(updates.values()) + [entry_id]
            sql = (
                f"UPDATE sourcing_entries SET "
                f"{', '.join(set_frags)}, updated_at = now() "
                f"WHERE id = %s"
            )
            cur.execute(sql, params)
            updated += 1
            if write_audit:
                audit_rows.append((
                    run_id, entry_id, pim.row_idx, tier,
                    psycopg2.extras.Json(before),
                    psycopg2.extras.Json(updates),
                ))
            if updated % batch_size == 0:
                if write_audit and audit_rows:
                    cur.executemany(
                        "INSERT INTO pim_enrichment_audit "
                        "(run_id, entry_id, pim_row_idx, match_tier, "
                        "before_json, after_json) VALUES (%s,%s,%s,%s,%s,%s)",
                        audit_rows,
                    )
                    audit_rows = []
                conn.commit()
                logger.info("Committed batch — %d rows updated so far", updated)
        if write_audit and audit_rows:
            cur.executemany(
                "INSERT INTO pim_enrichment_audit "
                "(run_id, entry_id, pim_row_idx, match_tier, "
                "before_json, after_json) VALUES (%s,%s,%s,%s,%s,%s)",
                audit_rows,
            )
    conn.commit()
    return updated


# ─── Main Pipeline ──────────────────────────────────────────────


def project_before(db_row: dict, fields: Iterable[str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for f in fields:
        val = db_row.get(f)
        if isinstance(val, (date, datetime)):
            val = val.isoformat()
        out[f] = val
    return out


def run(args: argparse.Namespace) -> int:
    if args.verbose:
        logger.setLevel(logging.DEBUG)

    source = Path(args.source).expanduser().resolve()
    if not source.exists():
        logger.error("Source not found: %s", source)
        return 2

    # ── PIM ingest ──
    pim_rows = load_pim_rows(source, args.sheet)
    lookup, idx_stats = build_pim_lookup(pim_rows)
    logger.info(
        "PIM index: total=%d, no_key=%d, no_enrichment=%d, indexed=%d, keys=%d",
        idx_stats["total_pim_rows"], idx_stats["no_match_key"],
        idx_stats["no_enrichment"], idx_stats["indexed_rows"],
        idx_stats["distinct_keys"],
    )

    # ── DB ingest ──
    dsn = get_dsn()
    logger.info("Connecting DSN: %s",
                re.sub(r":[^:@]*@", ":***@", dsn))
    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    try:
        schema_cols = detect_schema_columns(conn)
        missing_cols = (set(ENRICHMENT_FIELDS) | {"normalized_model", "image_url"}) - schema_cols
        if missing_cols:
            logger.warning("Schema missing columns (will skip): %s", missing_cols)

        if args.write_audit and not args.dry_run:
            ensure_audit_table(conn)

        db_rows = fetch_existing_rows(conn, args.created_by, args.limit)

        # ── Match + plan ──
        stats = {
            "db_total": len(db_rows),
            "db_unmatched": 0,
            "db_no_model": 0,
            "matched_unique": 0,
            "matched_date_exact": 0,
            "matched_cost_closest": 0,
            "matched_canonical_pick": 0,
            "rows_no_change": 0,
            "rows_planned": 0,
            "image_url_filled": 0,
            "formula_errors_coerced": 0,
        }

        pim_used: set[int] = set()
        update_specs: list[tuple[int, dict, dict, PimRow, str]] = []
        unmatched_db: list[dict] = []

        for db_row in db_rows:
            model_norm = normalize_model(db_row.get("model"))
            customer_norm = normalize_customer(db_row.get("customer_name"))
            if not model_norm or not customer_norm:
                stats["db_no_model"] += 1
                stats["db_unmatched"] += 1
                continue
            candidates = lookup.get((model_norm, customer_norm))
            if not candidates:
                stats["db_unmatched"] += 1
                unmatched_db.append({
                    "entry_id": db_row["id"],
                    "model": db_row.get("model"),
                    "model_norm": model_norm,
                    "customer": db_row.get("customer_name"),
                    "customer_norm": customer_norm,
                })
                continue
            pim, tier = disambiguate(db_row, candidates)
            pim_used.add(pim.row_idx)
            stats[f"matched_{tier}"] = stats.get(f"matched_{tier}", 0) + 1

            updates = build_update_plan(db_row, pim, schema_cols)
            if not updates:
                stats["rows_no_change"] += 1
                continue
            if "image_url" in updates:
                stats["image_url_filled"] += 1
            before = project_before(db_row, updates.keys())
            update_specs.append((db_row["id"], before, updates, pim, tier))
            stats["rows_planned"] += 1

            if logger.isEnabledFor(logging.DEBUG):
                logger.debug(
                    "id=%s tier=%s pim_row=%d updates=%s",
                    db_row["id"], tier, pim.row_idx,
                    {k: updates[k] for k in list(updates)[:4]},
                )

        # ── PIM unmatched (had enrichment, no DB target) ──
        pim_unmatched: list[PimRow] = []
        for key, rows in lookup.items():
            for p in rows:
                if p.row_idx not in pim_used:
                    pim_unmatched.append(p)

        # ── Report ──
        print_stats(stats, idx_stats, len(pim_unmatched))

        # ── Persist artifacts ──
        if unmatched_db:
            csv_path = LOG_DIR / f"unmatched_db_{RUN_ID}.csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as fh:
                w = csv.DictWriter(fh, fieldnames=list(unmatched_db[0].keys()))
                w.writeheader()
                w.writerows(unmatched_db)
            logger.info("Wrote %d unmatched DB rows → %s",
                        len(unmatched_db), csv_path)
        if pim_unmatched:
            csv_path = LOG_DIR / f"unmatched_pim_{RUN_ID}.csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow([
                    "pim_row_idx", "model_norm", "customer_norm",
                    "catalog_category", "brand_canonical", "stage",
                ])
                for p in pim_unmatched:
                    w.writerow([
                        p.row_idx, p.model_norm, p.customer_norm,
                        p.catalog_category or "", p.brand_canonical or "",
                        p.stage if p.stage is not None else "",
                    ])
            logger.info("Wrote %d unmatched PIM rows → %s",
                        len(pim_unmatched), csv_path)

        # ── Apply ──
        if args.dry_run:
            logger.info("DRY RUN — no UPDATEs executed. %d rows would change.",
                        len(update_specs))
            return 0

        if not update_specs:
            logger.info("Nothing to update.")
            return 0

        logger.info("Applying %d updates (run_id=%s)…",
                    len(update_specs), RUN_ID)
        updated = execute_updates(
            conn, update_specs, schema_cols, RUN_ID,
            args.write_audit, args.batch_size,
        )
        logger.info("DONE — %d rows updated. Log: %s", updated, LOG_FILE)
        return 0
    finally:
        conn.close()


def print_stats(stats: dict, idx_stats: dict, pim_unmatched_n: int) -> None:
    lines = [
        "──────────── PIM ENRICHMENT STATS ────────────",
        f"  PIM total rows           : {idx_stats['total_pim_rows']:>8}",
        f"  PIM with match key       : {idx_stats['total_pim_rows'] - idx_stats['no_match_key']:>8}",
        f"  PIM with enrichment      : {idx_stats['indexed_rows']:>8}",
        f"  PIM distinct keys        : {idx_stats['distinct_keys']:>8}",
        f"  PIM unmatched (no DB row): {pim_unmatched_n:>8}",
        "  ─────────────────────────────────────",
        f"  DB total fetched         : {stats['db_total']:>8}",
        f"  DB unmatched             : {stats['db_unmatched']:>8}",
        f"  DB no model/customer     : {stats['db_no_model']:>8}",
        "  ─────────────────────────────────────",
        f"  matched unique           : {stats.get('matched_unique', 0):>8}",
        f"  matched date_exact       : {stats.get('matched_date_exact', 0):>8}",
        f"  matched cost_closest     : {stats.get('matched_cost_closest', 0):>8}",
        f"  matched canonical_pick   : {stats.get('matched_canonical_pick', 0):>8}",
        "  ─────────────────────────────────────",
        f"  rows planned to update   : {stats['rows_planned']:>8}",
        f"  rows already up-to-date  : {stats['rows_no_change']:>8}",
        f"  image_url backfilled     : {stats['image_url_filled']:>8}",
        "──────────────────────────────────────────────",
    ]
    for ln in lines:
        logger.info(ln)


# ─── CLI ────────────────────────────────────────────────────────


def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Enrich sourcing_entries with PIM workbook enrichment cols",
    )
    p.add_argument("--source", required=True, help="Path to .xlsm/.xlsx")
    p.add_argument("--dry-run", action="store_true", help="No UPDATEs")
    p.add_argument("-v", "--verbose", action="store_true", help="DEBUG logs")
    p.add_argument("--limit", type=int, default=None,
                   help="Process first N DB rows (testing)")
    p.add_argument("--batch-size", type=int, default=500,
                   help="UPDATE commit interval")
    p.add_argument("--sheet", default=DEFAULT_SHEET,
                   help=f"Sheet name (default: {DEFAULT_SHEET})")
    p.add_argument("--created-by", default=DEFAULT_CREATED_BY,
                   help=f"Filter DB rows by created_by_email "
                        f"(default: {DEFAULT_CREATED_BY})")
    p.add_argument("--no-audit", dest="write_audit", action="store_false",
                   help="Skip pim_enrichment_audit writes (faster)")
    p.set_defaults(write_audit=True)
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    try:
        return run(args)
    except KeyboardInterrupt:
        logger.warning("Interrupted by user")
        return 130
    except Exception:
        logger.exception("Fatal error")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
