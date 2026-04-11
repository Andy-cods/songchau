#!/usr/bin/env python3
"""Rebuild xnk_price_lookup from the canonical TT XNK workbook."""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl

if TYPE_CHECKING:
    import asyncpg

DSN = os.getenv(
    "DATABASE_URL",
    "postgresql://scadmin:SC2026_ERP_Pr0d_X9k2mQ7wR4@postgres:5432/songchau_erp",
).replace("+asyncpg", "")

TARGET_TABLE = "xnk_price_lookup"
TARGET_SHEET = "DATA"
WORKBOOK_PATTERN = re.compile(r"^tt xnk bqms 2026\.xls[mx]$", re.IGNORECASE)
SCHEMA_SQL = (
    Path(__file__).resolve().parents[1] / "migrations" / "xnk_lookup.sql"
)

FIELD_INDEX = {
    "rfq_date": 1,
    "quotation_no": 2,
    "bqms_code": 3,
    "item_name": 4,
    "detail_note": 5,
    "item_type": 6,
    "maker": 7,
    "notes": 8,
    "notes2": 9,
    "unit_calc": 10,
    "quantity_calc": 11,
    "quote_deadline": 12,
    "quoted_date": 13,
    "bqms_code3": 14,
    "customs_description": 15,
    "hs_code": 16,
    "unit": 17,
    "quantity": 18,
    "total_usd": 19,
    "price_usd": 20,
    "price_vnd": 21,
    "buyer_name": 22,
    "seller_name": 23,
    "year_2022_or_before": 24,
    "year_2023": 25,
    "to_11_2024": 26,
    "alt_supplier_or_notes": 27,
}

INSERT_SQL = f"""
    INSERT INTO {TARGET_TABLE} (
        rfq_date, quotation_no, bqms_code, item_name, item_explain,
        item_type, maker, notes, notes2, unit, quantity,
        quote_deadline, quoted_date, bqms_code3, hs_code,
        price_usd, price_vnd, total_usd, buyer_name, seller_name,
        source, raw_data
    ) VALUES (
        $1, $2, $3, $4, $5,
        $6, $7, $8, $9, $10, $11,
        $12, $13, $14, $15,
        $16, $17, $18, $19, $20,
        $21, $22::jsonb
    )
"""


@dataclass(slots=True)
class ParsedRow:
    params: tuple[Any, ...]
    row_hash: str
    rfq_year: int | None


def matches_xnk_lookup_workbook(path: Path | str) -> bool:
    return bool(WORKBOOK_PATTERN.match(Path(path).name))


def resolve_workbook_path(path_like: str) -> Path:
    path = Path(path_like)
    if path.is_file():
        return path

    if path.is_dir():
        candidates = sorted(
            candidate for candidate in path.rglob("*") if candidate.is_file() and matches_xnk_lookup_workbook(candidate)
        )
        if candidates:
            return candidates[-1]

    raise FileNotFoundError(f"Could not find TT XNK BQMS 2026 workbook from '{path_like}'")


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").replace("\r", "\n").strip()
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    if not text or text.lower() == "none":
        return None
    return text


def normalize_hs_code(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace(" ", "")
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = normalize_text(value)
    if not text:
        return None

    stripped = text.replace(",", "").replace(" ", "")
    if stripped in {"", "-", "--"}:
        return None

    try:
        return Decimal(stripped)
    except InvalidOperation:
        try:
            return Decimal(stripped.replace(".", "").replace(",", "."))
        except InvalidOperation:
            return None


def parse_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return openpyxl.utils.datetime.from_excel(value).date()
        except (TypeError, ValueError, OverflowError):
            return None

    text = normalize_text(value)
    if not text:
        return None

    normalized = text.replace(".", "/").replace("-", "/")
    parts = normalized.split("/")
    if len(parts) == 3 and all(part.isdigit() for part in parts):
        first, second, third = (int(part) for part in parts)
        year_part = third if third > 100 else 2000 + third
        if first <= 12 and second <= 12:
            # The TT XNK 2024-2026 workbooks are authored in MM/DD/YY for ambiguous values.
            return date(year_part, first, second)
        if first <= 12 and second > 12:
            return date(year_part, first, second)
        if first > 12 and second <= 12:
            return date(year_part, second, first)

    for fmt in (
        "%Y/%m/%d",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%m-%d-%Y",
        "%m-%d-%y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def get_value(row: list[Any], field: str) -> Any:
    idx = FIELD_INDEX[field]
    return row[idx] if idx < len(row) else None


def combine_details(*values: Any) -> str | None:
    parts: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = normalize_text(value)
        if text and text not in seen:
            seen.add(text)
            parts.append(text)
    return "\n".join(parts) if parts else None


def should_keep_row(parsed: dict[str, Any]) -> bool:
    key_fields = (
        parsed["quotation_no"],
        parsed["bqms_code"],
        parsed["item_name"],
        parsed["item_explain"],
        parsed["seller_name"],
        parsed["hs_code"],
    )
    return any(key_fields)


def build_raw_data(
    headers: list[str | None],
    row: list[Any],
    workbook_name: str,
    excel_row_number: int,
) -> dict[str, Any]:
    raw_row: dict[str, Any] = {}
    for idx, value in enumerate(row):
        header = headers[idx] if idx < len(headers) else None
        key = normalize_text(header) or f"column_{idx}"
        if isinstance(value, (datetime, date)):
            raw_row[key] = value.isoformat()
        else:
            raw_row[key] = value

    raw_row["_workbook"] = workbook_name
    raw_row["_sheet"] = TARGET_SHEET
    raw_row["_excel_row_number"] = excel_row_number
    return raw_row


def compute_row_hash(raw_data: dict[str, Any]) -> str:
    payload = json.dumps(raw_data, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_data_sheet(workbook_path: Path) -> tuple[list[ParsedRow], dict[str, int]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if TARGET_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{TARGET_SHEET}' not found in {workbook_path.name}: {wb.sheetnames}")

    ws = wb[TARGET_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return [], {
            "rows_seen": 0,
            "rows_parsed": 0,
            "rows_skipped": 0,
            "rows_deduped": 0,
            "rows_invalid": 0,
        }

    headers = [normalize_text(value) for value in header_row]
    parsed_rows: list[ParsedRow] = []
    seen_hashes: set[str] = set()
    stats = {
        "rows_seen": 0,
        "rows_parsed": 0,
        "rows_skipped": 0,
        "rows_deduped": 0,
        "rows_invalid": 0,
    }

    for excel_row_number, raw_row in enumerate(rows_iter, start=2):
        row = list(raw_row)
        stats["rows_seen"] += 1

        if not any(normalize_text(value) for value in row):
            stats["rows_skipped"] += 1
            continue

        raw_data = build_raw_data(headers, row, workbook_path.name, excel_row_number)
        row_hash = compute_row_hash(raw_data)
        if row_hash in seen_hashes:
            stats["rows_deduped"] += 1
            continue

        detail_note = get_value(row, "detail_note")
        customs_description = get_value(row, "customs_description")
        parsed = {
            "rfq_date": parse_date_value(get_value(row, "rfq_date")),
            "quotation_no": normalize_text(get_value(row, "quotation_no")),
            "bqms_code": normalize_text(get_value(row, "bqms_code")),
            "item_name": normalize_text(get_value(row, "item_name")),
            "item_explain": combine_details(detail_note, customs_description),
            "item_type": normalize_text(get_value(row, "item_type")),
            "maker": normalize_text(get_value(row, "maker")),
            "notes": normalize_text(get_value(row, "notes")),
            "notes2": combine_details(
                get_value(row, "notes2"),
                get_value(row, "alt_supplier_or_notes"),
            ),
            "unit": normalize_text(get_value(row, "unit")) or normalize_text(get_value(row, "unit_calc")),
            "quantity": parse_decimal(get_value(row, "quantity")) or parse_decimal(get_value(row, "quantity_calc")),
            "quote_deadline": normalize_text(get_value(row, "quote_deadline")),
            "quoted_date": parse_date_value(get_value(row, "quoted_date")),
            "bqms_code3": normalize_text(get_value(row, "bqms_code3")),
            "hs_code": normalize_hs_code(get_value(row, "hs_code")),
            "price_usd": parse_decimal(get_value(row, "price_usd")),
            "price_vnd": parse_decimal(get_value(row, "price_vnd")),
            "total_usd": parse_decimal(get_value(row, "total_usd")),
            "buyer_name": normalize_text(get_value(row, "buyer_name")),
            "seller_name": normalize_text(get_value(row, "seller_name")),
        }

        if not should_keep_row(parsed):
            stats["rows_invalid"] += 1
            continue

        params = (
            parsed["rfq_date"],
            parsed["quotation_no"],
            parsed["bqms_code"],
            parsed["item_name"],
            parsed["item_explain"],
            parsed["item_type"],
            parsed["maker"],
            parsed["notes"],
            parsed["notes2"],
            parsed["unit"],
            parsed["quantity"],
            parsed["quote_deadline"],
            parsed["quoted_date"],
            parsed["bqms_code3"],
            parsed["hs_code"],
            parsed["price_usd"],
            parsed["price_vnd"],
            parsed["total_usd"],
            parsed["buyer_name"],
            parsed["seller_name"],
            "excel_import",
            json.dumps(raw_data, ensure_ascii=False, default=str),
        )

        parsed_rows.append(
            ParsedRow(
                params=params,
                row_hash=row_hash,
                rfq_year=parsed["rfq_date"].year if parsed["rfq_date"] else None,
            )
        )
        seen_hashes.add(row_hash)
        stats["rows_parsed"] += 1

    wb.close()
    return parsed_rows, stats


async def ensure_schema(conn: "asyncpg.Connection") -> None:
    await conn.execute(SCHEMA_SQL.read_text(encoding="utf-8"))


async def rebuild_xnk_price_lookup(
    conn: "asyncpg.Connection",
    workbook_path: str | Path,
    *,
    data_source: str = "excel_import",
    truncate_existing: bool = True,
    dry_run: bool = False,
) -> dict[str, Any]:
    workbook = resolve_workbook_path(str(workbook_path))
    parsed_rows, stats = parse_data_sheet(workbook)

    year_counts: dict[int, int] = {}
    for row in parsed_rows:
        if row.rfq_year is not None:
            year_counts[row.rfq_year] = year_counts.get(row.rfq_year, 0) + 1

    result: dict[str, Any] = {
        **stats,
        "workbook": str(workbook),
        "workbook_size_bytes": workbook.stat().st_size,
        "workbook_mtime": datetime.fromtimestamp(workbook.stat().st_mtime).isoformat(),
        "rows_inserted": 0,
        "years": year_counts,
    }

    if dry_run:
        return result

    await ensure_schema(conn)

    params_list = []
    for row in parsed_rows:
        params = list(row.params)
        params[20] = data_source
        params_list.append(tuple(params))

    async with conn.transaction():
        if truncate_existing:
            await conn.execute(f"TRUNCATE TABLE {TARGET_TABLE} RESTART IDENTITY")
        if params_list:
            await conn.executemany(INSERT_SQL, params_list)
        await conn.execute(f"ANALYZE {TARGET_TABLE}")

    result["rows_inserted"] = len(params_list)
    return result


async def _run_cli(args: argparse.Namespace) -> None:
    if args.dry_run:
        workbook = resolve_workbook_path(args.workbook)
        parsed_rows, stats = parse_data_sheet(workbook)
        year_counts: dict[int, int] = {}
        for row in parsed_rows:
            if row.rfq_year is not None:
                year_counts[row.rfq_year] = year_counts.get(row.rfq_year, 0) + 1

        result = {
            **stats,
            "workbook": str(workbook),
            "workbook_size_bytes": workbook.stat().st_size,
            "workbook_mtime": datetime.fromtimestamp(workbook.stat().st_mtime).isoformat(),
            "rows_inserted": 0,
            "years": year_counts,
        }
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return

    import asyncpg

    conn = await asyncpg.connect(DSN)
    try:
        result = await rebuild_xnk_price_lookup(
            conn,
            args.workbook,
            data_source=args.data_source,
            truncate_existing=not args.no_truncate,
            dry_run=args.dry_run,
        )
    finally:
        await conn.close()

    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


def main() -> None:
    parser = argparse.ArgumentParser(description="Rebuild xnk_price_lookup from TT XNK workbook")
    parser.add_argument(
        "--workbook",
        required=True,
        help="Path to workbook file or directory that contains TT XNK BQMS 2026.xlsm",
    )
    parser.add_argument(
        "--data-source",
        default="manual_rebuild",
        help="Value stored in xnk_price_lookup.source",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse workbook and print stats without writing to the database",
    )
    parser.add_argument(
        "--no-truncate",
        action="store_true",
        help="Append instead of truncating target table first",
    )
    args = parser.parse_args()
    asyncio.run(_run_cli(args))


if __name__ == "__main__":
    main()
