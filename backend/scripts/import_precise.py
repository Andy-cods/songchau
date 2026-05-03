#!/usr/bin/env python3
"""
Song Chau ERP -- PRECISE Import Script (v2)

Positional-column-based Excel-to-PostgreSQL importer.
Every column mapping is by Excel column INDEX, not header name.
Handles all 14 import targets with UPSERT (ON CONFLICT DO UPDATE).

Usage:
    python scripts/import_precise.py --source /path/to/onedrive
    python scripts/import_precise.py --source ./data --dry-run
    python scripts/import_precise.py --source ./data --table bqms_rfq --verbose
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import logging
import math
import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("import_precise")

# ---------------------------------------------------------------------------
# Database DSN
# ---------------------------------------------------------------------------
DSN = os.getenv(
    "DATABASE_URL",
    "postgresql://scadmin:SC2026_ERP_Pr0d_X9k2mQ7wR4@postgres:5432/songchau_erp",
).replace("+asyncpg", "")

DATA_SOURCE = "onedrive_sync"

# ---------------------------------------------------------------------------
# Schema migration -- add columns & unique indexes required by this script
# Each statement is executed individually (asyncpg does not support multi-statement).
# ---------------------------------------------------------------------------
SCHEMA_MIGRATION_STMTS: list[str] = [
    # -- Add source_hash + synced_at columns where missing --
    # bqms_rfq
    "ALTER TABLE bqms_rfq ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE bqms_rfq ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # bqms_deliveries
    "ALTER TABLE bqms_deliveries ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE bqms_deliveries ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # bqms_orders
    "ALTER TABLE bqms_orders ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE bqms_orders ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # bqms_raw_material_po
    "ALTER TABLE bqms_raw_material_po ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE bqms_raw_material_po ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # bqms_material_pricing
    "ALTER TABLE bqms_material_pricing ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE bqms_material_pricing ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # import_export_tracking
    "ALTER TABLE import_export_tracking ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE import_export_tracking ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # imv_inquiries
    "ALTER TABLE imv_inquiries ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE imv_inquiries ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # imv_consolidated
    "ALTER TABLE imv_consolidated ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE imv_consolidated ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # imv_purchase_orders
    "ALTER TABLE imv_purchase_orders ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE imv_purchase_orders ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # customer_contacts
    "ALTER TABLE customer_contacts ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE customer_contacts ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # revenue_invoices
    "ALTER TABLE revenue_invoices ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE revenue_invoices ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",
    # bqms_won_quotations
    "ALTER TABLE bqms_won_quotations ADD COLUMN IF NOT EXISTS source_hash TEXT",
    "ALTER TABLE bqms_won_quotations ADD COLUMN IF NOT EXISTS synced_at TIMESTAMPTZ",

    # -- Unique indexes for ON CONFLICT (idempotent: IF NOT EXISTS) --
    # uq_bqms_rfq_dedup created out-of-band 2026-05-04 after deleting 3877
    # content-duplicate rows. The (rfq_number, bqms_code) pair is NOT unique
    # for this dataset (e.g. 16 rows of QT23047153/'' with different specs).
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_bqms_rfq_dedup ON bqms_rfq (rfq_number, bqms_code, source_hash)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_bqms_del_po_ship_bqms ON bqms_deliveries (po_number, shipping_no, bqms_code)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_bqms_ord_rfq_bqms ON bqms_orders (rfq_number, bqms_code)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_bqms_rmp_po_bqms ON bqms_raw_material_po (po_number, bqms_code)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_bqms_mp_rfq_bqms ON bqms_material_pricing (rfq_number, bqms_code)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_xnk_rfq_bqms_date ON import_export_tracking (rfq_number, bqms_code, tracking_date)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_imviq_hash ON imv_inquiries (source_hash)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_imvcon_quot_prod ON imv_consolidated (quotation_no, product_code)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_imvpo_po_prod ON imv_purchase_orders (po_number, product_code)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_custcontact_name_phone ON customer_contacts (full_name, phone)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_revinv_num_date ON revenue_invoices (invoice_number, invoice_date)",
    "CREATE UNIQUE INDEX IF NOT EXISTS uq_bwq_rfq_bqms ON bqms_won_quotations (rfq_number, bqms_code)",
]


# ===========================================================================
# VALUE PARSING HELPERS
# ===========================================================================

def safe_str(value: Any) -> str | None:
    """Convert to stripped string, or None if empty."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("none", "nan"):
        return None
    return s


def parse_date(value: Any) -> date | None:
    """Parse Excel date: datetime objects, serial numbers, or strings."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        try:
            # Excel serial date (base = 1899-12-30)
            base = date(1899, 12, 30)
            return base + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    s = str(value).strip()
    if not s or s == "-" or s.lower() == "n/a":
        return None
    # Try multiple date formats
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%m/%d/%y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Try partial match "2025-09-08 ~ 2025-12-08" -> first date
    m = re.match(r'(\d{4}-\d{2}-\d{2})', s)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except ValueError:
            pass
    return None


def parse_date_mdy(value: Any) -> date | None:
    """Parse date in m/d/y format (US style used in IMV inquiries)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        try:
            base = date(1899, 12, 30)
            return base + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    s = str(value).strip()
    if not s or s == "-" or s.lower() == "n/a":
        return None
    # m/d/y first
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_date_dmy(value: Any) -> date | None:
    """Parse dd/mm/yyyy for exchange rates."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        try:
            base = date(1899, 12, 30)
            return base + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    s = str(value).strip()
    if not s or s == "-":
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value: Any) -> float | None:
    """Parse numeric value from Excel, handling Vietnamese formatting."""
    if value is None:
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    s = str(value).strip()
    # Strip currency symbols
    s = s.replace("¥", "").replace("₫", "").replace("$", "")
    s = s.replace(",", "").replace(" ", "").replace("\xa0", "")
    if not s or s == "-" or s.lower() == "n/a":
        return None
    # VN thousand separator: 1.234.567 -> 1234567
    # But 1.5 -> 1.5 (decimal point)
    if re.match(r'^\d{1,3}(\.\d{3})+$', s):
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_int_as_str(value: Any) -> str | None:
    """Convert Excel int/float to string (for PO numbers, codes)."""
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s or s.lower() in ("none", "nan", "-"):
        return None
    # If it looks like "1234.0", convert to "1234"
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s


def parse_boolean(value: Any) -> bool | None:
    """Parse boolean from Excel."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        return bool(value)
    s = str(value).strip().lower()
    if s in ("x", "yes", "y", "co", "true", "1"):
        return True
    if s in ("", "-", "no", "n", "khong", "false", "0"):
        return False
    return None


RESULT_MAP = {
    "y": "won", "yes": "won", "trung": "won",
    "n": "lost", "no": "lost", "thua": "lost",
    "huy": "cancelled", "cancel": "cancelled", "cancelled": "cancelled",
}

DELIVERY_STATUS_MAP = {
    "da giao": "da_giao",
    "chua giao": "chua_giao",
    "dang giao": "dang_giao",
    "giao mot phan": "giao_mot_phan",
    "giao 1 phan": "giao_mot_phan",
}


def parse_result(value: Any) -> str | None:
    """Map Y/N to rfq_result enum."""
    if value is None:
        return None
    s = str(value).strip().lower()
    # Remove Vietnamese diacritics for matching
    s_no_accent = (
        s.replace("\u00fa", "u").replace("\u1ee7", "u")
         .replace("\u1ecd", "o").replace("\u00f3", "o")
    )
    return RESULT_MAP.get(s) or RESULT_MAP.get(s_no_accent)


def parse_delivery_status(value: Any) -> str | None:
    """Map Vietnamese delivery status to enum."""
    if value is None:
        return None
    s = str(value).strip().lower()
    # Normalize Vietnamese characters
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return DELIVERY_STATUS_MAP.get(ascii_s) or DELIVERY_STATUS_MAP.get(s)


def compute_source_hash(row: list[Any]) -> str:
    """SHA256 hash of all cells in a row, pipe-delimited."""
    parts = [str(cell) if cell is not None else "" for cell in row]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


# ===========================================================================
# EXCEL READER -- openpyxl based, returns raw rows
# ===========================================================================

def read_excel_raw(
    filepath: str,
    sheet_name: str | None,
    header_row: int = 1,
) -> list[list[Any]]:
    """
    Read an Excel sheet, return ALL rows as list[list].
    header_row is 1-based. Returns rows AFTER the header row.
    If header_row=0, skips auto-detection and returns all rows.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            # Case-insensitive search
            for sn in wb.sheetnames:
                if sn.strip().lower() == sheet_name.strip().lower():
                    sheet_name = sn
                    break
            else:
                wb.close()
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
                )
        ws = wb[sheet_name]
    else:
        ws = wb.active or wb[wb.sheetnames[0]]

    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    wb.close()

    if header_row == 0:
        return all_rows

    if len(all_rows) <= header_row:
        return []

    # Return data rows after the header row (header_row is 1-based)
    return all_rows[header_row:]


def find_file(source_path: Path, filename: str) -> Path | None:
    """Find a file by name in the source directory, with fallback strategies."""
    # Exact match
    fp = source_path / filename
    if fp.exists():
        return fp

    # Case-insensitive in same dir
    for f in source_path.iterdir():
        if f.is_file() and f.name.lower() == filename.lower():
            return f

    # Recursive search by filename
    for f in source_path.rglob("*"):
        if f.is_file() and f.name.lower() == filename.lower():
            return f

    return None


# ===========================================================================
# TABLE-SPECIFIC IMPORT FUNCTIONS
# ===========================================================================

async def import_bqms_rfq(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """1. bqms_rfq <-- Thong ke hoi hang BQMS.xlsx, sheet TONG HOP BQMS

    Idempotency: ON CONFLICT (rfq_number, bqms_code, source_hash). Identical
    Excel row content (same source_hash) is a no-op; new content INSERTs.
    Requires UNIQUE INDEX uq_bqms_rfq_dedup created during 2026-05-04 dedup.

    Date fill-down: column A (inquiry_date) in Excel uses merged cells -- only
    the first row of each date group has the date filled in. We track the
    last seen date and forward-fill to subsequent rows with empty col A.
    """
    fp = find_file(source, "Thong ke hoi hang BQMS.xlsx")
    if not fp:
        logger.warning("  File not found: Thong ke hoi hang BQMS.xlsx")
        return _empty_stats()

    rows = read_excel_raw(str(fp), "TONG HOP BQMS", header_row=1)
    logger.info("  Read %d data rows from TONG HOP BQMS", len(rows))

    sql = """
        INSERT INTO bqms_rfq (
            inquiry_date, person_in_charge_name, rfq_number, bqms_code,
            specification, maker, expected_qty,
            purchase_price_rmb, purchase_price_vnd,
            quoted_price_ama, quoted_price_bqms_v1,
            quoted_price_bqms_v2, quoted_price_bqms_v3,
            notes, supplier_name, result, report,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13,
            $14, $15, COALESCE($16::rfq_result, 'pending'), $17,
            $18, $19, NOW()
        )
        ON CONFLICT (rfq_number, bqms_code, source_hash)
            DO UPDATE SET
                inquiry_date = COALESCE(EXCLUDED.inquiry_date, bqms_rfq.inquiry_date),
                person_in_charge_name = COALESCE(EXCLUDED.person_in_charge_name, bqms_rfq.person_in_charge_name),
                specification = COALESCE(EXCLUDED.specification, bqms_rfq.specification),
                maker = COALESCE(EXCLUDED.maker, bqms_rfq.maker),
                expected_qty = COALESCE(EXCLUDED.expected_qty, bqms_rfq.expected_qty),
                purchase_price_rmb = COALESCE(EXCLUDED.purchase_price_rmb, bqms_rfq.purchase_price_rmb),
                purchase_price_vnd = COALESCE(EXCLUDED.purchase_price_vnd, bqms_rfq.purchase_price_vnd),
                quoted_price_ama = COALESCE(EXCLUDED.quoted_price_ama, bqms_rfq.quoted_price_ama),
                quoted_price_bqms_v1 = COALESCE(EXCLUDED.quoted_price_bqms_v1, bqms_rfq.quoted_price_bqms_v1),
                quoted_price_bqms_v2 = COALESCE(EXCLUDED.quoted_price_bqms_v2, bqms_rfq.quoted_price_bqms_v2),
                quoted_price_bqms_v3 = COALESCE(EXCLUDED.quoted_price_bqms_v3, bqms_rfq.quoted_price_bqms_v3),
                supplier_name = COALESCE(EXCLUDED.supplier_name, bqms_rfq.supplier_name),
                result = COALESCE(EXCLUDED.result, bqms_rfq.result),
                report = COALESCE(EXCLUDED.report, bqms_rfq.report),
                notes = COALESCE(EXCLUDED.notes, bqms_rfq.notes),
                synced_at = NOW(),
                updated_at = NOW()
    """

    stats = _empty_stats()
    last_date = None  # forward-fill tracker for column A merged cells

    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        rfq_number = safe_str(row[2]) if len(row) > 2 else None
        bqms_code = safe_str(row[3]) if len(row) > 3 else None
        if not rfq_number and not bqms_code:
            stats["skip"] += 1
            continue

        # Forward-fill inquiry_date from the most recent non-empty col A.
        # Excel merges date cells visually but stores the value only in
        # the first row of each group; openpyxl returns None for the rest.
        cell_date = parse_date(_get(row, 0))
        if cell_date is not None:
            last_date = cell_date
        inquiry_date = cell_date if cell_date is not None else last_date

        # Parse purchase_price_rmb: strip "¥" prefix
        rmb_raw = row[7] if len(row) > 7 else None
        rmb_val = parse_number(rmb_raw)

        params = [
            inquiry_date,                       # inquiry_date (filled-down)
            safe_str(_get(row, 1)),             # person_in_charge_name
            rfq_number,                         # rfq_number
            bqms_code,                          # bqms_code
            safe_str(_get(row, 4)),             # specification
            safe_str(_get(row, 5)),             # maker
            parse_number(_get(row, 6)),         # expected_qty
            rmb_val,                            # purchase_price_rmb
            parse_number(_get(row, 8)),         # purchase_price_vnd
            parse_number(_get(row, 9)),         # quoted_price_ama
            parse_number(_get(row, 10)),        # quoted_price_bqms_v1
            parse_number(_get(row, 11)),        # quoted_price_bqms_v2
            parse_number(_get(row, 12)),        # quoted_price_bqms_v3
            safe_str(_get(row, 13)),            # notes
            safe_str(_get(row, 14)),            # supplier_name
            parse_result(_get(row, 15)),        # result
            safe_str(_get(row, 17)),            # report (col 17)
            DATA_SOURCE,                        # data_source
            compute_source_hash(row),           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_bqms_deliveries(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """2. bqms_deliveries <-- Thong ke giao hang 2026/2025/2023-2024.xlsx"""
    files_to_try = [
        ("Thong ke giao hang 2026.xlsx", None, 1),
        ("Thong ke giao hang 2025.xlsx", None, 3),   # header at row 3, positional
        ("Thong ke giao hang 2023-2024.xlsx", None, 1),
    ]

    sql = """
        INSERT INTO bqms_deliveries (
            po_date, po_number, shipping_no, quotation_no,
            bqms_code, specification, quantity, unit,
            unit_price, amount, sev_type, buyer_email,
            recipient_name, receiving_warehouse, buyer_phone,
            delivery_status, delivery_date, actual_delivered_qty,
            delivery_method, country_origin, total_delivered_value_vnd,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10,
            $11, $12, $13, $14, $15,
            COALESCE($16::delivery_status, 'chua_giao'),
            $17, $18, $19, $20, $21,
            $22, $23, NOW()
        )
        ON CONFLICT (po_number, shipping_no, bqms_code)
            DO UPDATE SET
                delivery_status = COALESCE(EXCLUDED.delivery_status, bqms_deliveries.delivery_status),
                delivery_date = COALESCE(EXCLUDED.delivery_date, bqms_deliveries.delivery_date),
                actual_delivered_qty = COALESCE(EXCLUDED.actual_delivered_qty, bqms_deliveries.actual_delivered_qty),
                total_delivered_value_vnd = COALESCE(EXCLUDED.total_delivered_value_vnd, bqms_deliveries.total_delivered_value_vnd),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    total_stats = _empty_stats()

    for filename, sheet_hint, header_row in files_to_try:
        fp = find_file(source, filename)
        if not fp:
            logger.warning("  File not found: %s", filename)
            continue

        # Determine sheet name: try common names
        import openpyxl
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        available_sheets = wb.sheetnames
        wb.close()

        # Pick the PO tracking sheet
        sheet_name = None
        for candidate in available_sheets:
            cl = candidate.strip().lower()
            if "po" in cl and ("ke" in cl or "kê" in cl or "th" in cl):
                sheet_name = candidate
                break
        if not sheet_name and available_sheets:
            sheet_name = available_sheets[0]

        rows = read_excel_raw(str(fp), sheet_name, header_row=header_row)
        logger.info("  Read %d rows from %s (sheet: %s, hdr: %d)",
                     len(rows), filename, sheet_name, header_row)

        stats = _empty_stats()
        for idx, row in enumerate(rows):
            if _is_empty_row(row):
                stats["skip"] += 1
                continue

            po_number = parse_int_as_str(_get(row, 1))
            shipping_no = parse_int_as_str(_get(row, 2))
            bqms_code = safe_str(_get(row, 4))
            if not po_number and not bqms_code:
                stats["skip"] += 1
                continue
            # Ensure we have non-null conflict keys
            if not po_number:
                po_number = ""
            if not shipping_no:
                shipping_no = ""
            if not bqms_code:
                bqms_code = ""

            params = [
                parse_date(_get(row, 0)),          # po_date
                po_number,                          # po_number (TEXT)
                shipping_no,                        # shipping_no (TEXT)
                safe_str(_get(row, 3)),             # quotation_no
                bqms_code,                          # bqms_code
                safe_str(_get(row, 5)),             # specification
                parse_number(_get(row, 6)),         # quantity
                safe_str(_get(row, 7)),             # unit
                parse_number(_get(row, 8)),         # unit_price
                parse_number(_get(row, 9)),         # amount
                safe_str(_get(row, 10)),            # sev_type
                safe_str(_get(row, 11)),            # buyer_email
                safe_str(_get(row, 12)),            # recipient_name
                safe_str(_get(row, 13)),            # receiving_warehouse
                safe_str(_get(row, 14)),            # buyer_phone
                parse_delivery_status(_get(row, 15)),  # delivery_status
                parse_date(_get(row, 16)),          # delivery_date
                parse_number(_get(row, 17)),        # actual_delivered_qty
                safe_str(_get(row, 19)),            # delivery_method (col 19)
                safe_str(_get(row, 20)),            # country_origin (col 20)
                parse_number(_get(row, 21)),        # total_delivered_value_vnd (col 21)
                DATA_SOURCE,                        # data_source
                compute_source_hash(row),           # source_hash
            ]

            await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
            _log_progress(stats, idx, len(rows))

        _merge_stats(total_stats, stats)

    return total_stats


async def import_bqms_orders(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """3. bqms_orders <-- Thong ke dat hang.xlsx, Sheet1"""
    fp = find_file(source, "Thong ke dat hang.xlsx")
    if not fp:
        logger.warning("  File not found: Thong ke dat hang.xlsx")
        return _empty_stats()

    rows = read_excel_raw(str(fp), "Sheet1", header_row=1)
    logger.info("  Read %d data rows from Sheet1", len(rows))

    sql = """
        INSERT INTO bqms_orders (
            rfq_number, bqms_code, specification, customer_name,
            expected_qty, order_qty, unit, order_date,
            validity_date, status, delivered_qty, delivery_date, notes,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8,
            $9, COALESCE($10, 'pending'), $11, $12, $13,
            $14, $15, NOW()
        )
        ON CONFLICT (rfq_number, bqms_code)
            DO UPDATE SET
                order_qty = COALESCE(EXCLUDED.order_qty, bqms_orders.order_qty),
                status = COALESCE(EXCLUDED.status, bqms_orders.status),
                delivered_qty = COALESCE(EXCLUDED.delivered_qty, bqms_orders.delivered_qty),
                delivery_date = COALESCE(EXCLUDED.delivery_date, bqms_orders.delivery_date),
                notes = COALESCE(EXCLUDED.notes, bqms_orders.notes),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        rfq_number = safe_str(_get(row, 1))
        bqms_code = safe_str(_get(row, 2))
        if not rfq_number and not bqms_code:
            stats["skip"] += 1
            continue

        # Parse validity_date: "2025-09-08 ~ 2025-12-08" -> parse first date
        validity_raw = _get(row, 9)
        validity_date = parse_date(validity_raw)

        # Map status text to valid enum
        status_raw = safe_str(_get(row, 10))
        status = _map_order_status(status_raw)

        params = [
            rfq_number,                         # rfq_number
            bqms_code,                          # bqms_code
            safe_str(_get(row, 3)),             # specification
            safe_str(_get(row, 4)),             # customer_name
            parse_number(_get(row, 5)),         # expected_qty
            parse_number(_get(row, 6)),         # order_qty
            safe_str(_get(row, 7)),             # unit
            parse_date(_get(row, 8)),           # order_date
            validity_date,                      # validity_date
            status,                             # status
            parse_number(_get(row, 11)),        # delivered_qty
            parse_date(_get(row, 12)),          # delivery_date
            safe_str(_get(row, 13)),            # notes
            DATA_SOURCE,                        # data_source
            compute_source_hash(row),           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_bqms_samsung_po(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """4. bqms_samsung_po <-- BQMS - PO.xlsx, Sheet2 -- ALREADY WORKS (skip)"""
    logger.info("  bqms_samsung_po: existing import works -- skipping.")
    return _empty_stats()


async def import_bqms_raw_material_po(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """5. bqms_raw_material_po <-- THEO DOI PO PHOI.xlsx, sheets PO PHOI 2025 + 2026"""
    fp = find_file(source, "THEO DOI PO PHOI.xlsx")
    if not fp:
        logger.warning("  File not found: THEO DOI PO PHOI.xlsx")
        return _empty_stats()

    sql = """
        INSERT INTO bqms_raw_material_po (
            po_date, po_number, bqms_code, specification,
            po_qty, unit, in_stock, remaining_qty,
            delivered_qty, pending,
            source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, NOW()
        )
        ON CONFLICT (po_number, bqms_code)
            DO UPDATE SET
                remaining_qty = COALESCE(EXCLUDED.remaining_qty, bqms_raw_material_po.remaining_qty),
                delivered_qty = COALESCE(EXCLUDED.delivered_qty, bqms_raw_material_po.delivered_qty),
                pending = COALESCE(EXCLUDED.pending, bqms_raw_material_po.pending),
                in_stock = COALESCE(EXCLUDED.in_stock, bqms_raw_material_po.in_stock),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    total_stats = _empty_stats()
    for sheet in ["PO PHOI 2025", "PO PHOI 2026"]:
        try:
            rows = read_excel_raw(str(fp), sheet, header_row=1)
        except ValueError:
            logger.warning("  Sheet '%s' not found, skipping", sheet)
            continue

        logger.info("  Read %d rows from %s", len(rows), sheet)
        stats = _empty_stats()
        for idx, row in enumerate(rows):
            if _is_empty_row(row):
                stats["skip"] += 1
                continue

            po_number = parse_int_as_str(_get(row, 1))
            bqms_code = safe_str(_get(row, 2))
            if not po_number and not bqms_code:
                stats["skip"] += 1
                continue
            if not po_number:
                po_number = ""
            if not bqms_code:
                bqms_code = ""

            params = [
                parse_date(_get(row, 0)),          # po_date
                po_number,                          # po_number
                bqms_code,                          # bqms_code
                safe_str(_get(row, 3)),             # specification
                parse_number(_get(row, 4)),         # po_qty
                safe_str(_get(row, 5)),             # unit
                parse_boolean(_get(row, 6)),        # in_stock
                parse_number(_get(row, 7)),         # remaining_qty
                parse_number(_get(row, 8)),         # delivered_qty
                parse_boolean(_get(row, 9)),        # pending
                compute_source_hash(row),           # source_hash
            ]

            await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
            _log_progress(stats, idx, len(rows))

        _merge_stats(total_stats, stats)

    return total_stats


async def import_bqms_material_pricing(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """6. bqms_material_pricing <-- KET QUA PHOI TRUOT.xlsx, Sheet1"""
    fp = find_file(source, "KET QUA PHOI TRUOT.xlsx")
    if not fp:
        logger.warning("  File not found: KET QUA PHOI TRUOT.xlsx")
        return _empty_stats()

    rows = read_excel_raw(str(fp), "Sheet1", header_row=1)
    logger.info("  Read %d data rows", len(rows))

    sql = """
        INSERT INTO bqms_material_pricing (
            rfq_number, bqms_code, specification,
            unit_price_vnd, weight_kg, notes,
            dimension_l, dimension_w, dimension_h,
            material_type, density_g_m3,
            source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, NOW()
        )
        ON CONFLICT (rfq_number, bqms_code)
            DO UPDATE SET
                unit_price_vnd = COALESCE(EXCLUDED.unit_price_vnd, bqms_material_pricing.unit_price_vnd),
                weight_kg = COALESCE(EXCLUDED.weight_kg, bqms_material_pricing.weight_kg),
                dimension_l = COALESCE(EXCLUDED.dimension_l, bqms_material_pricing.dimension_l),
                dimension_w = COALESCE(EXCLUDED.dimension_w, bqms_material_pricing.dimension_w),
                dimension_h = COALESCE(EXCLUDED.dimension_h, bqms_material_pricing.dimension_h),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW()
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        rfq_number = safe_str(_get(row, 1))
        bqms_code = safe_str(_get(row, 2))
        if not rfq_number and not bqms_code:
            stats["skip"] += 1
            continue

        params = [
            rfq_number,                         # rfq_number
            bqms_code,                          # bqms_code
            safe_str(_get(row, 3)),             # specification
            parse_number(_get(row, 4)),         # unit_price_vnd
            parse_number(_get(row, 5)),         # weight_kg
            safe_str(_get(row, 6)),             # notes
            parse_number(_get(row, 9)),         # dimension_l (col 9)
            parse_number(_get(row, 10)),        # dimension_w (col 10)
            parse_number(_get(row, 11)),        # dimension_h (col 11)
            safe_str(_get(row, 14)),            # material_type (col 14)
            parse_number(_get(row, 15)),        # density_g_m3 (col 15)
            compute_source_hash(row),           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_export_tracking_fn(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """7. import_export_tracking <-- TT XNK BQMS 2023.xlsx, sheet TONG HOP"""
    # Try multiple file name variants
    fp = None
    for name in [
        "TT XNK BQMS 2023.xlsx",
        "TT XNK BQMS 2023-2026.xlsx",
        "TT XNK BQMS 2024.xlsx",
    ]:
        fp = find_file(source, name)
        if fp:
            break
    if not fp:
        logger.warning("  File not found: TT XNK BQMS 2023*.xlsx")
        return _empty_stats()

    # Try to find the summary sheet
    import openpyxl
    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    available = wb.sheetnames
    wb.close()

    # Look for TONG HOP / TỔNG HỢP sheet, or iterate monthly sheets
    target_sheets = []
    for s in available:
        sl = s.strip().upper()
        # Match "TONG HOP" or "TỔNG HỢP" (Vietnamese diacritics)
        if "TONG" in sl or "T\u1ed4NG" in sl:
            target_sheets = [s]
            break
    if not target_sheets:
        # Use all non-utility sheets
        target_sheets = [s for s in available if s.strip().upper() not in ("TGUSD",)]

    sql = """
        INSERT INTO import_export_tracking (
            tracking_date, rfq_number, bqms_code, product_name,
            detail_explain, maker, unit_calc, quantity_calc,
            transaction_date, customs_description, hs_code,
            quantity, total_usd, unit_price_usd,
            buyer_name, seller_name, purchased_qty, alt_supplier,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11,
            $12, $13, $14, $15, $16, $17, $18,
            $19, $20, NOW()
        )
        ON CONFLICT (rfq_number, bqms_code, tracking_date)
            DO UPDATE SET
                quantity = COALESCE(EXCLUDED.quantity, import_export_tracking.quantity),
                total_usd = COALESCE(EXCLUDED.total_usd, import_export_tracking.total_usd),
                unit_price_usd = COALESCE(EXCLUDED.unit_price_usd, import_export_tracking.unit_price_usd),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    total_stats = _empty_stats()
    for sheet in target_sheets:
        try:
            rows = read_excel_raw(str(fp), sheet, header_row=1)
        except ValueError:
            continue

        logger.info("  Read %d rows from sheet '%s'", len(rows), sheet)
        stats = _empty_stats()
        for idx, row in enumerate(rows):
            if _is_empty_row(row):
                stats["skip"] += 1
                continue

            tracking_date = parse_date(_get(row, 1))
            rfq_number = safe_str(_get(row, 2))
            bqms_code = safe_str(_get(row, 3))

            # Need at least tracking_date + one of rfq/bqms for conflict key
            if not tracking_date or (not rfq_number and not bqms_code):
                stats["skip"] += 1
                continue
            if not rfq_number:
                rfq_number = ""
            if not bqms_code:
                bqms_code = ""

            hs_code_raw = _get(row, 12)
            hs_code = parse_int_as_str(hs_code_raw)

            params = [
                tracking_date,                      # tracking_date
                rfq_number,                         # rfq_number
                bqms_code,                          # bqms_code
                safe_str(_get(row, 4)),             # product_name
                safe_str(_get(row, 5)),             # detail_explain
                safe_str(_get(row, 6)),             # maker (col 6)
                safe_str(_get(row, 8)),             # unit_calc (col 8)
                parse_number(_get(row, 9)),         # quantity_calc (col 9)
                parse_date(_get(row, 10)),          # transaction_date (col 10)
                safe_str(_get(row, 11)),            # customs_description (col 11)
                hs_code,                            # hs_code (col 12)
                parse_number(_get(row, 13)),        # quantity (col 13)
                parse_number(_get(row, 14)),        # total_usd (col 14)
                parse_number(_get(row, 15)),        # unit_price_usd (col 15)
                safe_str(_get(row, 16)),            # buyer_name (col 16)
                safe_str(_get(row, 17)),            # seller_name (col 17)
                parse_number(_get(row, 18)),        # purchased_qty (col 18)
                safe_str(_get(row, 19)),            # alt_supplier (col 19)
                DATA_SOURCE,                        # data_source
                compute_source_hash(row),           # source_hash
            ]

            await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
            _log_progress(stats, idx, len(rows))

        _merge_stats(total_stats, stats)

    return total_stats


async def import_imv_inquiries(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """8. imv_inquiries <-- Thong ke hoi hang - update 240424.xlsx, sheet TONG HOP"""
    fp = find_file(source, "Thong ke hoi hang - update 240424.xlsx")
    if not fp:
        logger.warning("  File not found: Thong ke hoi hang - update 240424.xlsx")
        return _empty_stats()

    rows = read_excel_raw(str(fp), "TONG HOP", header_row=1)
    logger.info("  Read %d data rows from TONG HOP", len(rows))

    sql = """
        INSERT INTO imv_inquiries (
            customer_name, person_in_charge_name, model, product_name,
            maker, inquiry_date,
            purchase_price, purchase_currency,
            selling_price, quantity, tax_rate, hs_code,
            weight_kg, notes, coefficient, supplier_name,
            exchange_rate,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6,
            $7, $8::currency_code,
            $9, $10, $11, $12,
            $13, $14, $15, $16,
            $17,
            $18, $19, NOW()
        )
        ON CONFLICT (source_hash) DO NOTHING
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        # Must have at least customer_name or product_name
        customer = safe_str(_get(row, 0))
        product_name = safe_str(_get(row, 3))
        if not customer and not product_name:
            stats["skip"] += 1
            continue

        # Purchase price: pick first non-null from cols [6,7,8,9,10]
        # with corresponding currency
        purchase_price = None
        purchase_currency = None
        currency_cols = [
            (6, "JPY"), (7, "USD"), (8, "KRW"), (9, "RMB"), (10, "VND"),
        ]
        for col_idx, curr in currency_cols:
            val = parse_number(_get(row, col_idx))
            if val is not None:
                purchase_price = val
                purchase_currency = curr
                break

        src_hash = compute_source_hash(row)

        params = [
            customer,                           # customer_name
            safe_str(_get(row, 1)),             # person_in_charge_name
            safe_str(_get(row, 2)),             # model
            product_name,                       # product_name
            safe_str(_get(row, 4)),             # maker
            parse_date_mdy(_get(row, 5)),       # inquiry_date (m/d/y)
            purchase_price,                     # purchase_price
            purchase_currency,                  # purchase_currency
            parse_number(_get(row, 11)),        # selling_price
            parse_number(_get(row, 12)),        # quantity
            parse_number(_get(row, 13)),        # tax_rate
            safe_str(_get(row, 14)),            # hs_code
            parse_number(_get(row, 15)),        # weight_kg
            safe_str(_get(row, 16)),            # notes
            parse_number(_get(row, 17)),        # coefficient
            safe_str(_get(row, 18)),            # supplier_name
            parse_number(_get(row, 20)),        # exchange_rate (col 20)
            DATA_SOURCE,                        # data_source
            src_hash,                           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_imv_consolidated(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """9. imv_consolidated <-- same file, sheet Tong hop IMV"""
    fp = find_file(source, "Thong ke hoi hang - update 240424.xlsx")
    if not fp:
        logger.warning("  File not found")
        return _empty_stats()

    # Try multiple sheet name variants
    sheet = None
    for name in ["Tổng hợp IMV", "Tong hop IMV", "TONG HOP IMV"]:
        try:
            rows = read_excel_raw(str(fp), name, header_row=1)
            sheet = name
            break
        except ValueError:
            continue

    if sheet is None:
        logger.warning("  Sheet 'Tong hop IMV' not found")
        return _empty_stats()

    logger.info("  Read %d data rows from '%s'", len(rows), sheet)

    sql = """
        INSERT INTO imv_consolidated (
            quotation_no, status, purchaser_name, customer_name,
            customer_branch, customer_item_code, product_code,
            rfq_number, product_name, model, specification, maker,
            unit, expected_order_qty, prev_year_po_count,
            request_date, quote_deadline, end_date, moq,
            sales_person_name, quoted_price, purchase_price, price_diff,
            po_status, po_qty, po_amount, profit, notes,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7,
            $8, $9, $10, $11, $12,
            $13, $14, $15,
            $16, $17, $18, $19,
            $20, $21, $22, $23,
            $24, $25, $26, $27, $28,
            $29, $30, NOW()
        )
        ON CONFLICT (quotation_no, product_code)
            DO UPDATE SET
                status = COALESCE(EXCLUDED.status, imv_consolidated.status),
                po_status = COALESCE(EXCLUDED.po_status, imv_consolidated.po_status),
                po_qty = COALESCE(EXCLUDED.po_qty, imv_consolidated.po_qty),
                po_amount = COALESCE(EXCLUDED.po_amount, imv_consolidated.po_amount),
                profit = COALESCE(EXCLUDED.profit, imv_consolidated.profit),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        quotation_no = parse_int_as_str(_get(row, 0))
        product_code = parse_int_as_str(_get(row, 6))
        if not quotation_no and not product_code:
            stats["skip"] += 1
            continue

        params = [
            quotation_no,                       # quotation_no
            safe_str(_get(row, 1)),             # status
            safe_str(_get(row, 2)),             # purchaser_name
            safe_str(_get(row, 3)),             # customer_name
            safe_str(_get(row, 4)),             # customer_branch
            safe_str(_get(row, 5)),             # customer_item_code
            product_code,                       # product_code
            safe_str(_get(row, 7)),             # rfq_number
            safe_str(_get(row, 8)),             # product_name
            safe_str(_get(row, 9)),             # model
            safe_str(_get(row, 10)),            # specification
            safe_str(_get(row, 11)),            # maker
            safe_str(_get(row, 12)),            # unit
            parse_number(_get(row, 13)),        # expected_order_qty
            parse_number(_get(row, 14)),        # prev_year_po_count
            parse_date(_get(row, 15)),          # request_date
            parse_date(_get(row, 16)),          # quote_deadline
            parse_date(_get(row, 17)),          # end_date
            parse_number(_get(row, 18)),        # moq
            safe_str(_get(row, 19)),            # sales_person_name
            parse_number(_get(row, 20)),        # quoted_price
            parse_number(_get(row, 21)),        # purchase_price
            parse_number(_get(row, 22)),        # price_diff
            safe_str(_get(row, 23)),            # po_status
            parse_number(_get(row, 24)),        # po_qty
            parse_number(_get(row, 25)),        # po_amount
            parse_number(_get(row, 26)),        # profit
            safe_str(_get(row, 27)),            # notes
            DATA_SOURCE,                        # data_source
            compute_source_hash(row),           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_imv_purchase_orders(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """10. imv_purchase_orders <-- 1.PO IMV 2025.xlsx, sheet SONG CHAU, header row 2"""
    fp = find_file(source, "1.PO IMV 2025.xlsx")
    if not fp:
        logger.warning("  File not found: 1.PO IMV 2025.xlsx")
        return _empty_stats()

    rows = read_excel_raw(str(fp), "SONG CHAU", header_row=2)
    logger.info("  Read %d data rows from SONG CHAU (hdr=2)", len(rows))

    sql = """
        INSERT INTO imv_purchase_orders (
            po_date, po_number, product_code, product_name,
            unit, requested_qty, unit_price, amount,
            vat_amount, total_amount, purchasing_dept,
            delivered_qty, actual_delivery_date, invoice_date, remaining_qty,
            supplier_name, document_ref, notes,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8,
            $9, $10, $11,
            $12, $13, $14, $15,
            $16, $17, $18,
            $19, $20, NOW()
        )
        ON CONFLICT (po_number, product_code)
            DO UPDATE SET
                delivered_qty = COALESCE(EXCLUDED.delivered_qty, imv_purchase_orders.delivered_qty),
                actual_delivery_date = COALESCE(EXCLUDED.actual_delivery_date, imv_purchase_orders.actual_delivery_date),
                remaining_qty = COALESCE(EXCLUDED.remaining_qty, imv_purchase_orders.remaining_qty),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        po_number = parse_int_as_str(_get(row, 2))
        product_code = parse_int_as_str(_get(row, 3))
        if not po_number:
            stats["skip"] += 1
            continue
        if not product_code:
            product_code = ""

        params = [
            parse_date(_get(row, 1)),           # po_date (col 1)
            po_number,                          # po_number (col 2)
            product_code,                       # product_code (col 3)
            safe_str(_get(row, 4)),             # product_name (col 4)
            safe_str(_get(row, 5)),             # unit (col 5)
            parse_number(_get(row, 6)),         # requested_qty (col 6)
            parse_number(_get(row, 7)),         # unit_price (col 7)
            parse_number(_get(row, 8)),         # amount (col 8)
            parse_number(_get(row, 9)),         # vat_amount (col 9)
            parse_number(_get(row, 10)),        # total_amount (col 10)
            safe_str(_get(row, 11)),            # purchasing_dept (col 11)
            parse_number(_get(row, 12)),        # delivered_qty (col 12)
            parse_date(_get(row, 13)),          # actual_delivery_date (col 13)
            parse_date(_get(row, 14)),          # invoice_date (col 14)
            parse_number(_get(row, 15)),        # remaining_qty (col 15)
            safe_str(_get(row, 27)),            # supplier_name (col 27)
            safe_str(_get(row, 28)),            # document_ref (col 28)
            safe_str(_get(row, 29)),            # notes (col 29)
            DATA_SOURCE,                        # data_source
            compute_source_hash(row),           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_customer_contacts(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """11. customer_contacts <-- Thong ke giao hang 2026.xlsx, sheet DANH BA, header row 3"""
    fp = find_file(source, "Thong ke giao hang 2026.xlsx")
    if not fp:
        logger.warning("  File not found: Thong ke giao hang 2026.xlsx")
        return _empty_stats()

    # Try multiple sheet name variants (Vietnamese diacritics)
    sheet_found = None
    for sheet_candidate in ["DANH B\u1EA0", "DANH BA", "Danh b\u1ea1"]:
        try:
            rows = read_excel_raw(str(fp), sheet_candidate, header_row=3)
            sheet_found = sheet_candidate
            break
        except ValueError:
            continue
    if sheet_found is None:
        # Last resort: list sheets and find one containing "DANH"
        import openpyxl
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if "DANH" in sn.upper() or "danh" in sn.lower():
                sheet_found = sn
                break
        wb.close()
        if sheet_found:
            rows = read_excel_raw(str(fp), sheet_found, header_row=3)
        else:
            logger.warning("  Sheet DANH BA not found in %s", fp.name)
            return _empty_stats()
    logger.info("  Read %d data rows from '%s' (hdr=3)", len(rows), sheet_found)

    # customer_contacts requires customer_id NOT NULL.
    # We use a default Samsung customer or the first available customer.
    if conn and not dry_run:
        default_cust_id = await conn.fetchval(
            "SELECT id FROM customers ORDER BY id LIMIT 1"
        )
        if not default_cust_id:
            logger.error("  No customers in DB -- cannot import contacts.")
            return _empty_stats()
    else:
        default_cust_id = 1  # placeholder for dry-run

    sql = """
        INSERT INTO customer_contacts (
            customer_id, email, full_name, delivery_info, phone,
            source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, NOW()
        )
        ON CONFLICT (full_name, phone)
            DO UPDATE SET
                email = COALESCE(EXCLUDED.email, customer_contacts.email),
                delivery_info = COALESCE(EXCLUDED.delivery_info, customer_contacts.delivery_info),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        full_name = safe_str(_get(row, 1))
        phone = safe_str(_get(row, 3))
        if not full_name:
            stats["skip"] += 1
            continue
        # full_name NOT NULL is required, phone needed for conflict key
        if not phone:
            phone = ""

        params = [
            default_cust_id,                    # customer_id
            safe_str(_get(row, 0)),             # email (col 0)
            full_name,                          # full_name (col 1)
            safe_str(_get(row, 2)),             # delivery_info (col 2)
            phone,                              # phone (col 3)
            compute_source_hash(row),           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_revenue_invoices(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """12. revenue_invoices <-- Bang theo doi doanh thu SC.2025.xlsx, multiple month sheets, header row 3"""
    fp = None
    for name in [
        "Bang theo doi doanh thu SC.2025.xlsx",
        "Bảng theo dõi doanh thu SC.2025.xlsx",
    ]:
        fp = find_file(source, name)
        if fp:
            break
    if not fp:
        logger.warning("  File not found: Bang theo doi doanh thu SC.2025.xlsx")
        return _empty_stats()

    # Enumerate month sheets
    import openpyxl
    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    available = wb.sheetnames
    wb.close()

    month_sheets = []
    for s in available:
        # Match T1.25, T2.25, ... T12.25 or similar patterns
        if re.match(r'T\d+\.', s.strip()):
            month_sheets.append(s)

    if not month_sheets:
        logger.warning("  No month sheets found in revenue file. Available: %s", available)
        return _empty_stats()

    sql = """
        INSERT INTO revenue_invoices (
            invoice_number, invoice_date, customer_name, product_name,
            unit, quantity, unit_price, amount,
            tax_rate, vat_amount, total_amount,
            po_number, purchase_price, purchase_vat,
            shipping_cost, commission, customer_quoted, invoice_buying,
            customs_fee, export_tax, other_costs, profit,
            data_source, source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8,
            $9, $10, $11,
            $12, $13, $14,
            $15, $16, $17, $18,
            $19, $20, $21, $22,
            $23, $24, NOW()
        )
        ON CONFLICT (invoice_number, invoice_date)
            DO UPDATE SET
                total_amount = COALESCE(EXCLUDED.total_amount, revenue_invoices.total_amount),
                profit = COALESCE(EXCLUDED.profit, revenue_invoices.profit),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW(),
                updated_at = NOW()
    """

    total_stats = _empty_stats()
    for sheet in month_sheets:
        try:
            rows = read_excel_raw(str(fp), sheet, header_row=3)
        except ValueError:
            logger.warning("  Sheet '%s' not found", sheet)
            continue

        logger.info("  Read %d rows from sheet '%s'", len(rows), sheet)

        # Extract month/year from sheet name (e.g. "T1.25" -> month=1, year=2025)
        m = re.match(r'T(\d+)\.(\d+)', sheet.strip())
        invoice_month = int(m.group(1)) if m else None
        invoice_year = 2000 + int(m.group(2)) if m else None

        stats = _empty_stats()
        for idx, row in enumerate(rows):
            if _is_empty_row(row):
                stats["skip"] += 1
                continue

            invoice_number = parse_int_as_str(_get(row, 1))
            invoice_date = parse_date(_get(row, 2))
            if not invoice_number and not invoice_date:
                stats["skip"] += 1
                continue
            # We need both for conflict key
            if not invoice_number:
                invoice_number = ""
            if not invoice_date and invoice_month and invoice_year:
                # Fallback: use first of month
                try:
                    invoice_date = date(invoice_year, invoice_month, 1)
                except ValueError:
                    pass

            params = [
                invoice_number,                     # invoice_number
                invoice_date,                       # invoice_date
                safe_str(_get(row, 3)),             # customer_name
                safe_str(_get(row, 4)),             # product_name
                safe_str(_get(row, 5)),             # unit
                parse_number(_get(row, 6)),         # quantity
                parse_number(_get(row, 7)),         # unit_price
                parse_number(_get(row, 8)),         # amount (TT)
                parse_number(_get(row, 9)),         # tax_rate
                parse_number(_get(row, 10)),        # vat_amount
                parse_number(_get(row, 11)),        # total_amount
                safe_str(_get(row, 12)),            # po_number
                parse_number(_get(row, 13)),        # purchase_price
                parse_number(_get(row, 14)),        # purchase_vat
                parse_number(_get(row, 15)),        # shipping_cost
                parse_number(_get(row, 18)),        # commission (col 18)
                parse_number(_get(row, 19)),        # customer_quoted (col 19)
                parse_number(_get(row, 20)),        # invoice_buying (col 20)
                parse_number(_get(row, 21)),        # customs_fee (col 21)
                parse_number(_get(row, 22)),        # export_tax (col 22)
                parse_number(_get(row, 23)),        # other_costs (col 23)
                parse_number(_get(row, 25)),        # profit (col 25)
                DATA_SOURCE,                        # data_source
                compute_source_hash(row),           # source_hash
            ]

            await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
            _log_progress(stats, idx, len(rows))

        _merge_stats(total_stats, stats)

    return total_stats


async def import_exchange_rates(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """13. exchange_rates <-- TT XNK BQMS 2023.xlsx, sheet TGUSD"""
    fp = None
    for name in [
        "TT XNK BQMS 2023.xlsx",
        "TT XNK BQMS 2023-2026.xlsx",
    ]:
        fp = find_file(source, name)
        if fp:
            break
    if not fp:
        logger.warning("  File not found for exchange rates")
        return _empty_stats()

    # Read ALL rows (no header)
    rows = read_excel_raw(str(fp), "TGUSD", header_row=0)
    logger.info("  Read %d raw rows from TGUSD", len(rows))

    sql = """
        INSERT INTO exchange_rates (
            rate_date, from_currency, to_currency, rate, rate_type, source
        ) VALUES (
            $1, 'USD'::currency_code, 'VND'::currency_code, $2, 'transfer', $3
        )
        ON CONFLICT (rate_date, from_currency, to_currency, rate_type)
            DO UPDATE SET
                rate = EXCLUDED.rate
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row) or len(row) < 2:
            stats["skip"] += 1
            continue

        rate_date = parse_date_dmy(_get(row, 0))
        rate = parse_number(_get(row, 1))

        if not rate_date or not rate:
            stats["skip"] += 1
            continue

        params = [rate_date, rate, DATA_SOURCE]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


async def import_bqms_won_quotations(conn, source: Path, dry_run: bool, verbose: bool) -> dict:
    """14. bqms_won_quotations <-- Thong ke hoi hang BQMS.xlsx, sheet TRUNG BG"""
    fp = find_file(source, "Thong ke hoi hang BQMS.xlsx")
    if not fp:
        logger.warning("  File not found: Thong ke hoi hang BQMS.xlsx")
        return _empty_stats()

    rows = read_excel_raw(str(fp), "TRUNG BG", header_row=1)
    logger.info("  Read %d data rows from TRUNG BG", len(rows))

    sql = """
        INSERT INTO bqms_won_quotations (
            person_in_charge_name, rfq_number, bqms_code,
            specification, quantity,
            po_price, po_deadline, notes,
            supplier_name, hs_code, goods_description,
            source_hash, synced_at
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, NOW()
        )
        ON CONFLICT (rfq_number, bqms_code)
            DO UPDATE SET
                po_price = COALESCE(EXCLUDED.po_price, bqms_won_quotations.po_price),
                po_deadline = COALESCE(EXCLUDED.po_deadline, bqms_won_quotations.po_deadline),
                supplier_name = COALESCE(EXCLUDED.supplier_name, bqms_won_quotations.supplier_name),
                hs_code = COALESCE(EXCLUDED.hs_code, bqms_won_quotations.hs_code),
                goods_description = COALESCE(EXCLUDED.goods_description, bqms_won_quotations.goods_description),
                source_hash = EXCLUDED.source_hash,
                synced_at = NOW()
    """

    stats = _empty_stats()
    for idx, row in enumerate(rows):
        if _is_empty_row(row):
            stats["skip"] += 1
            continue

        rfq_number = safe_str(_get(row, 1))
        bqms_code = safe_str(_get(row, 2))
        if not rfq_number and not bqms_code:
            stats["skip"] += 1
            continue

        params = [
            safe_str(_get(row, 0)),             # person_in_charge_name (col 0)
            rfq_number,                         # rfq_number (col 1)
            bqms_code,                          # bqms_code (col 2)
            safe_str(_get(row, 3)),             # specification (col 3)
            parse_number(_get(row, 5)),         # quantity (col 5)
            parse_number(_get(row, 10)),        # po_price (col 10)
            parse_date(_get(row, 11)),          # po_deadline (col 11)
            safe_str(_get(row, 12)),            # notes (col 12)
            safe_str(_get(row, 13)),            # supplier_name (col 13)
            safe_str(_get(row, 14)),            # hs_code (col 14)
            safe_str(_get(row, 15)),            # goods_description (col 15)
            compute_source_hash(row),           # source_hash
        ]

        await _exec_row(conn, sql, params, stats, idx, dry_run, verbose)
        _log_progress(stats, idx, len(rows))

    return stats


# ===========================================================================
# HELPER UTILITIES
# ===========================================================================

def _get(row: list, idx: int) -> Any:
    """Safe get from list by index."""
    if idx < len(row):
        return row[idx]
    return None


def _is_empty_row(row: list | None) -> bool:
    """Check if a row is entirely empty/None."""
    if not row:
        return True
    return all(
        c is None or (isinstance(c, str) and not c.strip())
        for c in row
    )


def _empty_stats() -> dict:
    return {"insert": 0, "update": 0, "skip": 0, "error": 0}


def _merge_stats(total: dict, part: dict) -> None:
    for k in ("insert", "update", "skip", "error"):
        total[k] += part[k]


ORDER_STATUS_MAP = {
    "pending": "pending",
    "confirmed": "confirmed",
    "in_production": "in_production",
    "shipped": "shipped",
    "delivered": "delivered",
    "closed": "closed",
    "cancelled": "cancelled",
}


def _map_order_status(raw: str | None) -> str:
    """Map Vietnamese status text to bqms_orders.status enum."""
    if not raw:
        return "pending"
    s = raw.strip().lower()
    # Direct match
    if s in ORDER_STATUS_MAP:
        return s
    # Vietnamese mappings
    mapping = {
        "da giao": "delivered",
        "dang san xuat": "in_production",
        "da xac nhan": "confirmed",
        "da dong": "closed",
        "da huy": "cancelled",
        "cho": "pending",
        "dang giao": "shipped",
    }
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return mapping.get(ascii_s, "pending")


async def _exec_row(
    conn,
    sql: str,
    params: list,
    stats: dict,
    row_idx: int,
    dry_run: bool,
    verbose: bool,
) -> None:
    """Execute a single INSERT/UPSERT row."""
    if dry_run:
        if stats["insert"] < 3 and verbose:
            logger.info("  [DRY-RUN] Row %d: %s", row_idx + 2, params[:6])
        stats["insert"] += 1
        return

    try:
        result = await conn.execute(sql, *params)
        if "INSERT 0 1" in result:
            stats["insert"] += 1
        elif "INSERT 0 0" in result:
            stats["skip"] += 1
        else:
            # UPDATE path from ON CONFLICT
            stats["update"] += 1
    except Exception as e:
        stats["error"] += 1
        if stats["error"] <= 10:
            err_msg = str(e)[:300]
            logger.warning(
                "  Error row %d: %s | params[:5]=%s",
                row_idx + 2, err_msg, params[:5],
            )


def _log_progress(stats: dict, idx: int, total: int) -> None:
    """Log progress every 500 rows."""
    processed = stats["insert"] + stats["update"] + stats["skip"] + stats["error"]
    if processed > 0 and processed % 500 == 0:
        logger.info(
            "  ... processed %d/%d (I:%d U:%d S:%d E:%d)",
            processed, total,
            stats["insert"], stats["update"], stats["skip"], stats["error"],
        )


# ===========================================================================
# IMPORT REGISTRY -- ordered list of all importers
# ===========================================================================

IMPORT_REGISTRY: list[tuple[str, Any]] = [
    ("bqms_rfq",               import_bqms_rfq),
    ("bqms_deliveries",        import_bqms_deliveries),
    ("bqms_orders",            import_bqms_orders),
    ("bqms_samsung_po",        import_bqms_samsung_po),
    ("bqms_raw_material_po",   import_bqms_raw_material_po),
    ("bqms_material_pricing",  import_bqms_material_pricing),
    ("import_export_tracking", import_export_tracking_fn),
    ("imv_inquiries",          import_imv_inquiries),
    ("imv_consolidated",       import_imv_consolidated),
    ("imv_purchase_orders",    import_imv_purchase_orders),
    ("customer_contacts",      import_customer_contacts),
    ("revenue_invoices",       import_revenue_invoices),
    ("exchange_rates",         import_exchange_rates),
    ("bqms_won_quotations",    import_bqms_won_quotations),
]


# ===========================================================================
# ETL SYNC LOG
# ===========================================================================

async def log_etl_sync(
    conn,
    sync_type: str,
    status: str,
    files_processed: int = 0,
    rows_inserted: int = 0,
    rows_updated: int = 0,
    rows_skipped: int = 0,
    error_message: str | None = None,
) -> None:
    """Record import run in etl_sync_log."""
    try:
        await conn.execute(
            """
            INSERT INTO etl_sync_log (
                sync_type, status, files_processed,
                rows_inserted, rows_updated, rows_skipped,
                error_message, completed_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, NOW())
            """,
            sync_type, status, files_processed,
            rows_inserted, rows_updated, rows_skipped,
            error_message,
        )
    except Exception as e:
        logger.warning("  Could not write etl_sync_log: %s", e)


# ===========================================================================
# MAIN
# ===========================================================================

async def main(
    source_dir: str,
    dry_run: bool = False,
    table_filter: str | None = None,
    verbose: bool = False,
) -> None:
    import asyncpg

    source_path = Path(source_dir)
    if not source_path.exists():
        logger.error("Source directory does not exist: %s", source_dir)
        sys.exit(1)

    logger.info("=" * 72)
    logger.info("SONG CHAU ERP -- PRECISE IMPORT (14 targets)")
    logger.info("=" * 72)
    logger.info("Source     : %s", source_dir)
    logger.info("DSN        : %s", DSN.split("@")[-1] if "@" in DSN else "***")
    logger.info("Dry run    : %s", dry_run)
    logger.info("Data source: %s", DATA_SOURCE)
    if table_filter:
        logger.info("Table filter: %s", table_filter)
    logger.info("-" * 72)

    start_time = time.time()

    # Connect to database
    conn = None
    if not dry_run:
        try:
            conn = await asyncpg.connect(DSN)
            logger.info("Database connected.")
        except Exception as e:
            logger.error("Cannot connect to database: %s", e)
            sys.exit(1)

        # Run schema migration (add columns + unique indexes)
        logger.info("Running schema migration (%d statements)...", len(SCHEMA_MIGRATION_STMTS))
        migration_errors = 0
        for stmt in SCHEMA_MIGRATION_STMTS:
            try:
                await conn.execute(stmt)
            except Exception as e:
                migration_errors += 1
                logger.warning("  Migration stmt failed: %s -- %s", stmt[:80], e)
        if migration_errors:
            logger.warning("Schema migration: %d/%d statements had errors (may be OK if already applied).",
                           migration_errors, len(SCHEMA_MIGRATION_STMTS))
        else:
            logger.info("Schema migration complete (all %d statements OK).", len(SCHEMA_MIGRATION_STMTS))
    else:
        logger.info("[DRY-RUN] Skipping database connection.")

    # Run imports
    grand_total = _empty_stats()
    table_results: dict[str, dict] = {}

    for idx, (table_name, import_fn) in enumerate(IMPORT_REGISTRY, start=1):
        if table_filter and table_filter not in table_name:
            continue

        logger.info("")
        logger.info(
            "=== [%d/%d] %s ===",
            idx, len(IMPORT_REGISTRY), table_name,
        )

        try:
            stats = await import_fn(conn, source_path, dry_run, verbose)
        except Exception as e:
            logger.error("  FATAL error importing %s: %s", table_name, e)
            import traceback
            traceback.print_exc()
            stats = _empty_stats()
            stats["error"] = 1

        table_results[table_name] = stats
        _merge_stats(grand_total, stats)

        logger.info(
            "  Result: I=%d U=%d S=%d E=%d",
            stats["insert"], stats["update"], stats["skip"], stats["error"],
        )

    # Log to etl_sync_log
    if conn and not dry_run:
        status = "success" if grand_total["error"] == 0 else "partial"
        await log_etl_sync(
            conn,
            sync_type="precise_import",
            status=status,
            files_processed=len([t for t in table_results if table_results[t]["insert"] > 0]),
            rows_inserted=grand_total["insert"],
            rows_updated=grand_total["update"],
            rows_skipped=grand_total["skip"],
            error_message=f"{grand_total['error']} errors" if grand_total["error"] else None,
        )

    # Close connection
    if conn:
        await conn.close()
        logger.info("Database connection closed.")

    elapsed = time.time() - start_time

    # Summary
    logger.info("")
    logger.info("=" * 72)
    logger.info("IMPORT SUMMARY")
    logger.info("=" * 72)
    logger.info("%-30s %8s %8s %8s %8s", "TABLE", "INSERT", "UPDATE", "SKIP", "ERROR")
    logger.info("-" * 72)
    for table_name, stats in table_results.items():
        logger.info(
            "%-30s %8d %8d %8d %8d",
            table_name, stats["insert"], stats["update"], stats["skip"], stats["error"],
        )
    logger.info("-" * 72)
    logger.info(
        "%-30s %8d %8d %8d %8d",
        "TOTAL",
        grand_total["insert"], grand_total["update"],
        grand_total["skip"], grand_total["error"],
    )
    logger.info("=" * 72)
    logger.info("Elapsed: %.1f seconds", elapsed)

    if grand_total["error"] > 0:
        logger.warning("There were %d errors -- review log above.", grand_total["error"])


def cli() -> None:
    parser = argparse.ArgumentParser(
        description="Song Chau ERP -- Precise Excel Import (positional columns)",
    )
    parser.add_argument(
        "--source", required=True,
        help="Path to directory containing Excel files",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Read & parse only, do not write to database",
    )
    parser.add_argument(
        "--table", default=None,
        help="Import only tables matching this substring (e.g. bqms_rfq)",
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Show debug-level information",
    )
    parser.add_argument(
        "--dsn", default=None,
        help="Override database connection string",
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.dsn:
        global DSN
        DSN = args.dsn.replace("+asyncpg", "")

    asyncio.run(main(
        source_dir=args.source,
        dry_run=args.dry_run,
        table_filter=args.table,
        verbose=args.verbose,
    ))


if __name__ == "__main__":
    cli()
