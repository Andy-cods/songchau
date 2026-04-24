"""Import DANH BẠ contacts from Excel into bqms_contacts table.

Usage (on VPS):
    python -m scripts.import_contacts /path/to/Thong_ke_giao_hang_2026.xlsx

Or locally for generating SQL:
    python scripts/import_contacts.py --sql-only
"""

import sys
import openpyxl


def read_contacts(xlsx_path: str) -> list[dict]:
    """Read DANH BẠ sheet from the delivery tracking Excel file."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["DANH BẠ"]

    contacts = []
    for row_idx in range(4, ws.max_row + 1):  # Data starts at row 4 (row 3 is header)
        email = ws.cell(row=row_idx, column=1).value
        name = ws.cell(row=row_idx, column=2).value
        delivery_info = ws.cell(row=row_idx, column=3).value
        phone = ws.cell(row=row_idx, column=4).value

        if not email or not name:
            continue

        contacts.append({
            "email_username": str(email).strip(),
            "full_name": str(name).strip(),
            "delivery_info": str(delivery_info).strip() if delivery_info else None,
            "phone": str(phone).strip() if phone else None,
        })

    return contacts


def generate_sql(contacts: list[dict]) -> str:
    """Generate UPSERT SQL for all contacts."""
    lines = [
        "-- Auto-generated from DANH BẠ sheet",
        "-- Run: psql -U erp_user -d erp_db -f import_contacts.sql\n",
    ]
    for c in contacts:
        email = c["email_username"].replace("'", "''")
        name = c["full_name"].replace("'", "''")
        info = c["delivery_info"].replace("'", "''") if c["delivery_info"] else "NULL"
        phone = c["phone"].replace("'", "''") if c["phone"] else "NULL"

        info_val = f"'{info}'" if info != "NULL" else "NULL"
        phone_val = f"'{phone}'" if phone != "NULL" else "NULL"

        lines.append(
            f"INSERT INTO bqms_contacts (email_username, full_name, delivery_info, phone) "
            f"VALUES ('{email}', '{name}', {info_val}, {phone_val}) "
            f"ON CONFLICT (email_username) DO UPDATE SET "
            f"full_name = EXCLUDED.full_name, delivery_info = EXCLUDED.delivery_info, "
            f"phone = EXCLUDED.phone, updated_at = NOW();"
        )

    return "\n".join(lines)


if __name__ == "__main__":
    xlsx_path = "C:/Users/ASUS/OneDrive - SONG CHAU CO., LTD/Puplic/BQMS/Thong ke giao hang/Thong ke giao hang 2026.xlsx"

    if len(sys.argv) > 1 and sys.argv[1] != "--sql-only":
        xlsx_path = sys.argv[1]

    contacts = read_contacts(xlsx_path)
    print(f"Read {len(contacts)} contacts from DANH BẠ sheet")

    sql = generate_sql(contacts)

    output_path = "backend/migrations/import_contacts_data.sql"
    if "--sql-only" in sys.argv:
        print(sql)
    else:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(sql)
        print(f"Written to {output_path}")
