"""Render báo giá XLSX/TSV cho quote_batches.

XLSX strategy: TEMPLATE-DRIVEN (Option A).
- Loads ``templates/SOURCING_QUOTE.xlsx`` (copy of "mẫu báo giá.xlsx") so the
  output matches Song Châu's official layout 100% (logo, fonts, borders,
  merged cells, column widths, row heights, page-setup, etc. are preserved
  by openpyxl).
- Fills header + customer placeholders, expands the items table to ``N``
  rows (clones row 14 styling for every line), repositions the totals/
  footer block right below the items table, and writes formula-driven
  amount/subtotal/VAT/grand cells so the workbook stays editable.

Placeholder ↔ cell mapping lives in ``templates/SOURCING_QUOTE.cells.json``
(same convention as ``AMA_DELIVERY_DOSSIER.cells.json``). Update both
files together when the template evolves.

The on-disk template path used by the production VPS lives at
``/data/files/templates/SOURCING_QUOTE.xlsx`` (preferred, mutable) with the
in-repo copy under ``app/services/templates/SOURCING_QUOTE.xlsx`` as
fallback so dev/test environments still work.
"""
from __future__ import annotations

import csv
import json
import logging
import re
from copy import copy
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template locations
# ---------------------------------------------------------------------------
_HERE = Path(__file__).parent
_REPO_TEMPLATE = _HERE / "templates" / "SOURCING_QUOTE.xlsx"
_REPO_CELLS_MAP = _HERE / "templates" / "SOURCING_QUOTE.cells.json"
_VPS_TEMPLATE = Path("/data/files/templates/SOURCING_QUOTE.xlsx")

# Layout constants (mirror SOURCING_QUOTE.cells.json — duplicated here so the
# renderer never silently breaks when the JSON file is missing in a deploy).
ITEMS_HEADER_ROW = 13
ITEMS_FIRST_DATA_ROW = 14
ITEMS_SAMPLE_COUNT = 3            # template ships with 3 sample rows (14-16)
ITEMS_DEFAULT_ROW_HEIGHT = 63.0
TOTALS_BLOCK_ROWS = 3             # subtotal / VAT / grand
AMOUNT_WORDS_OFFSET = 3           # rows after totals_start
SIGN_OFFSET = 4
THANK_YOU_OFFSET = 5
# A4 portrait holds the header + ~4 fixed-height (63pt) item rows + the full
# footer/stamp block on one page. Beyond this the footer is pushed to a fresh
# page so the company seal never straddles a page break (Thang 2026-06-30).
FOOTER_FITS_ITEMS_PER_PAGE = 4
DEFAULT_VAT_RATE = Decimal("0.08")
DEFAULT_VALIDITY = "10 ngày"
DEFAULT_INTRO = "Chúng tôi xin gửi tới Quý công ty bảng báo giá với nội dung như sau:"
DEFAULT_TERMS = (
    "Điều khoản:\n"
    "* Giao hàng tại kho bên mua\n"
    "* Thời gian giao hàng: 5-7 ngày kể từ ngày xác nhận đặt hàng\n"
    "* Điều khoản thanh toán: Theo thỏa thuận hợp đồng.\n"
    "* Đơn giá có thể thay đổi nếu số lượng đặt hàng thay đổi."
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _fmt_vnd(v: float | None) -> str:
    """Format số VND với dấu '.' thousands (VN convention)."""
    if v is None:
        return ""
    try:
        return f"{int(round(float(v))):,}".replace(",", ".")
    except Exception:
        return ""


# ICE security #2 (Thang 2026-06-13): defuse XLSX/CSV formula injection.
# Excel/LibreOffice interpret any cell starting with =, +, -, @, |, %, \t, \r
# as a formula — turning user-controlled text (customer_name, quote_note,
# product_name …) into a vector for DDE exfiltration / RCE in older Office.
# OWASP-recommended fix: prefix a single quote so the spreadsheet renders the
# literal text. Numbers / None / formulas this module writes itself stay
# untouched — only strings flowing in from user input get sanitised.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "|", "%", "\t", "\r")


def _defuse_formula(value: Any) -> Any:
    """Strip leading formula triggers from str inputs; pass non-str through."""
    if not isinstance(value, str) or not value:
        return value
    if value[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _num2words_vn(n: Any) -> str:
    """Đọc số tiền VND -> chữ tiếng Việt. Tối đa 999 tỉ. Lifted from sourcing_pdf_renderer."""
    if n is None or n == 0:
        return "không"
    try:
        n = int(round(float(n)))
    except Exception:
        return str(n)
    if n < 0:
        return "âm " + _num2words_vn(-n)

    don_vi = ["", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]

    def _read_3(num: int, full: bool) -> str:
        tram, du = divmod(num, 100)
        chuc, dv = divmod(du, 10)
        out = []
        if full or tram > 0:
            out.append(don_vi[tram] + " trăm")
            if chuc == 0 and dv > 0:
                out.append("lẻ")
        if chuc > 1:
            out.append(don_vi[chuc] + " mươi")
            if dv == 1:
                out.append("mốt")
            elif dv == 5:
                out.append("lăm")
            elif dv > 0:
                out.append(don_vi[dv])
        elif chuc == 1:
            out.append("mười")
            if dv == 5:
                out.append("lăm")
            elif dv > 0:
                out.append(don_vi[dv])
        elif chuc == 0 and dv > 0:
            out.append(don_vi[dv])
        return " ".join(s for s in out if s)

    units = ["", "nghìn", "triệu", "tỉ"]
    groups: list[int] = []
    while n > 0:
        groups.append(n % 1000)
        n //= 1000
    parts: list[str] = []
    last_idx = len(groups) - 1
    for idx in range(last_idx, -1, -1):
        g = groups[idx]
        if g == 0:
            # Skip empty groups entirely; we don't want stray "không trăm" or
            # an empty unit suffix in the output.
            continue
        s = _read_3(g, full=(idx != last_idx))
        if s:
            parts.append(s + (" " + units[idx] if units[idx] else ""))
    out = " ".join(parts).strip()
    return out[:1].upper() + out[1:] if out else "không"


def _fix_drawing_rels_paths(xlsx_path: str | Path) -> int:
    """Rewrite absolute image paths in drawing rels to relative paths.

    openpyxl's save produces:
        <Relationship Target="/xl/media/image2.jpeg" .../>
    LibreOffice/Gotenberg silently drops images with absolute targets, so the
    stamp/seal vanishes from the converted PDF (the xlsx still opens fine in
    Excel). Standard xlsx files use relative targets:
        <Relationship Target="../media/image2.jpeg" .../>
    Both are valid per OOXML spec, but LibreOffice is strict. Ported verbatim
    from ``tools/gc_template_quotation._fix_drawing_rels_paths``.

    Returns count of rels rewritten.
    """
    import zipfile, shutil, tempfile  # noqa: F401  (tempfile kept to mirror GC import)
    n = 0
    src = Path(xlsx_path)
    tmp_path = src.with_suffix(".fix.tmp.xlsx")
    try:
        with zipfile.ZipFile(src, "r") as zin:
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    # Only rewrite drawing rels files (xl/drawings/_rels/*.xml.rels)
                    if "drawings/_rels" in item.filename and item.filename.endswith(".rels"):
                        original = data
                        # Fix absolute /xl/media/... → ../media/...
                        data = data.replace(
                            b'Target="/xl/media/',
                            b'Target="../media/',
                        )
                        # Same for absolute /xl/drawings/... rare but defensive
                        data = data.replace(
                            b'Target="/xl/drawings/',
                            b'Target="../drawings/',
                        )
                        if data != original:
                            n += 1
                            logger.info("Fixed drawing rel path in %s", item.filename)
                    zout.writestr(item, data)
        shutil.move(str(tmp_path), str(src))
    except Exception as exc:
        logger.warning("_fix_drawing_rels_paths failed: %s", exc)
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass
    return n


def _resolve_template() -> Path:
    """Prefer the VPS-mounted mutable template; fall back to in-repo copy."""
    if _VPS_TEMPLATE.exists():
        return _VPS_TEMPLATE
    if _REPO_TEMPLATE.exists():
        return _REPO_TEMPLATE
    raise FileNotFoundError(
        f"SOURCING_QUOTE template not found. Looked at {_VPS_TEMPLATE} and {_REPO_TEMPLATE}"
    )


def _copy_cell_style(src: Cell, dst: Cell) -> None:
    """Copy openpyxl cell style attributes (font/fill/border/alignment/number_format/protection)."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _unmerge_in_range(ws: Worksheet, min_row: int, max_row: int) -> list[str]:
    """Unmerge any merged-cell range fully inside [min_row, max_row]. Returns the bounds for re-merge per-row."""
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= min_row and rng.max_row <= max_row:
            to_unmerge.append(str(rng))
    for s in to_unmerge:
        ws.unmerge_cells(s)
    return to_unmerge


def _capture_merges_at_or_below(ws: Worksheet, pivot_row: int) -> list[tuple[int, int, int, int]]:
    """Return [(min_col, min_row, max_col, max_row), ...] for merges at-or-below ``pivot_row`` and unmerge them."""
    captured: list[tuple[int, int, int, int]] = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= pivot_row:
            captured.append((rng.min_col, rng.min_row, rng.max_col, rng.max_row))
    # Unmerge after the snapshot so we don't mutate while iterating.
    for (c0, r0, c1, r1) in captured:
        ws.merged_cells.ranges.remove(
            next(r for r in ws.merged_cells.ranges
                 if r.min_col == c0 and r.min_row == r0
                 and r.max_col == c1 and r.max_row == r1)
        )
    return captured


def _reapply_shifted_merges(
    ws: Worksheet,
    captured: list[tuple[int, int, int, int]],
    delta: int,
) -> None:
    """Re-merge previously captured ranges shifted by ``delta`` rows."""
    for (c0, r0, c1, r1) in captured:
        new_r0 = r0 + delta
        new_r1 = r1 + delta
        if new_r0 < 1 or new_r1 < 1:
            continue
        ws.merge_cells(
            start_row=new_r0, start_column=c0,
            end_row=new_r1, end_column=c1,
        )


def _shift_images_at_or_below(ws: Worksheet, pivot_row: int, delta: int) -> int:
    """Shift floating picture anchors at-or-below ``pivot_row`` by ``delta`` rows.

    openpyxl's ``insert_rows``/``delete_rows`` shift cell VALUES but NOT the
    floating drawing anchors (``ws._images``). The template's stamp/seal image
    lives below the items table; without this re-anchor it stays pinned to its
    original row and ends up overlapping the resized table. Anchor rows are
    0-indexed (``_from.row``), so a worksheet ``pivot_row`` (1-indexed) maps to
    anchor row ``pivot_row - 1``. Mirrors the merge-shift logic precisely.

    Returns count of images shifted. No-op (returns 0) when ws._images is empty.
    """
    images = getattr(ws, "_images", None)
    if not images:
        return 0
    pivot0 = pivot_row - 1  # anchor rows are 0-indexed
    shifted = 0
    for img in images:
        anchor = getattr(img, "anchor", None)
        if anchor is None:
            continue
        _from = getattr(anchor, "_from", None)
        if _from is None or getattr(_from, "row", None) is None:
            continue
        if _from.row >= pivot0:
            _from.row += delta
            # TwoCellAnchor also carries a ``to`` marker that must move in lockstep
            # so the image keeps its height; OneCellAnchor has no ``to``.
            _to = getattr(anchor, "to", None)
            if _to is not None and getattr(_to, "row", None) is not None:
                _to.row += delta
            shifted += 1
    return shifted


def _capture_row_heights_at_or_below(ws: Worksheet, pivot_row: int) -> dict[int, float]:
    """Snapshot explicit row heights at/below ``pivot_row``, then clear them.

    Same openpyxl gap as merges/images: ``insert_rows``/``delete_rows`` shift
    cell VALUES but NOT ``row_dimensions`` heights. The template makes the
    totals/footer rows TALL (≈127pt each) so the 129pt company stamp fits inside
    them; without moving those heights, the footer content lands on default-short
    rows and the seal overflows the print area → the "GIÁM ĐỐC … " part gets
    clipped. Capture-then-reapply (shifted) keeps the footer rows tall at their
    new positions, so the stamp always fits exactly like the N<=3 reference.
    """
    captured: dict[int, float] = {}
    for r in list(ws.row_dimensions.keys()):
        if r >= pivot_row:
            h = ws.row_dimensions[r].height
            if h is not None:
                captured[r] = h
    for r in captured:
        ws.row_dimensions[r].height = None
    return captured


def _reapply_shifted_row_heights(ws: Worksheet, captured: dict[int, float], delta: int) -> None:
    """Re-apply captured row heights shifted by ``delta`` rows."""
    for r, h in captured.items():
        nr = r + delta
        if nr >= 1:
            ws.row_dimensions[nr].height = h


def _clone_row(ws: Worksheet, src_row: int, dst_row: int, col_count: int = 8) -> None:
    """Clone src_row styling + row height into dst_row. Values NOT copied (caller sets them)."""
    src_h = ws.row_dimensions[src_row].height
    if src_h:
        ws.row_dimensions[dst_row].height = src_h
    for col in range(1, col_count + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        _copy_cell_style(src, dst)
        dst.value = None


def _clear_row(ws: Worksheet, row: int, col_count: int = 8) -> None:
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).value = None


def _merge_b_c(ws: Worksheet, row: int) -> None:
    """Re-create the B{row}:C{row} merge that the template uses for item description."""
    ref = f"B{row}:C{row}"
    if ref not in (str(r) for r in ws.merged_cells.ranges):
        ws.merge_cells(ref)


# ---------------------------------------------------------------------------
# Main renderer
# ---------------------------------------------------------------------------
def render_xlsx(
    out_path: Path,
    *,
    quote_no: str,
    customer_name: str,
    quote_note: str,
    line_items: list[dict],
    total_value: float,
    created_by: str,
    created_at: datetime,
    # Optional richer fields used when the API caller has them (sourcing_orders
    # already provides them — quote_batches passes ``None``):
    customer_contact: str | None = None,
    customer_address: str | None = None,
    customer_mst: str | None = None,
    quote_owner: str | None = None,
    validity: str | None = None,
    vat_rate: float | Decimal | None = None,
    terms_text: str | None = None,
) -> None:
    """Render báo giá XLSX dùng template mẫu báo giá.xlsx.

    The template is loaded fresh on every call so concurrent renders do not
    share mutable state. Items table is expanded (or contracted) to exactly
    ``len(line_items)`` rows; totals/footer block follows immediately.
    """
    template_path = _resolve_template()
    wb = load_workbook(template_path)
    ws = wb.active  # template ships with the báo-giá sheet as the only/active sheet

    n_items = max(1, len(line_items))
    vat = Decimal(str(vat_rate)) if vat_rate is not None else DEFAULT_VAT_RATE

    # ── Header (right side meta block) ────────────────────────────────────
    # ICE security #2: defuse XLSX formula injection on every user-controlled
    # string we write into a cell. quote_no/created_by/owner can all come from
    # editable fields; only created_at (datetime) and validity (admin-managed
    # constant) are exempt.
    ws["G4"] = _defuse_formula(quote_no)
    # G5 (Ngày báo giá): write a REAL date object and PRESERVE the template's
    # native number_format ([$-409]mmmm d, yyyy). Writing a pre-formatted
    # dd/mm/yyyy string destroyed the localized date display, so we capture the
    # cloned nf first, then restore it after assigning a date value.
    g5 = ws["G5"]
    g5_nf = g5.number_format
    g5.value = created_at.date() if hasattr(created_at, "date") else created_at
    g5.number_format = g5_nf
    ws["G6"] = _defuse_formula(quote_owner or created_by or "")
    ws["G7"] = validity or DEFAULT_VALIDITY

    # ── Customer block ───────────────────────────────────────────────────
    # NOTE: row 12 (A12) holds the intro line which is merged across the row,
    # so B12 is a read-only MergedCell — writing to it raises and breaks the
    # whole render. The MST therefore rides on the company line (B10) instead
    # of its own cell. (Bug found by E2E 2026-06-22.)
    ws["B9"] = _defuse_formula(customer_contact or "")
    company_line = customer_name or ""
    if customer_mst:
        company_line = (f"{company_line} - MST: {customer_mst}" if company_line
                        else f"MST: {customer_mst}")
    ws["B10"] = _defuse_formula(company_line)
    ws["B11"] = _defuse_formula(customer_address or "")
    ws["A12"] = DEFAULT_INTRO

    # ── Items table: expand / shrink to N rows ───────────────────────────
    # Template ships with 3 sample rows (14, 15, 16). We need ``n_items``
    # consecutive rows starting at row 14 that share the same styling.
    delta = n_items - ITEMS_SAMPLE_COUNT
    last_sample_row = ITEMS_FIRST_DATA_ROW + ITEMS_SAMPLE_COUNT - 1  # 16

    pivot = last_sample_row + 1  # row 17 — first row AFTER the items table
    if delta > 0:
        # Capture & detach merges sitting at-or-below the pivot BEFORE the
        # insert. openpyxl's ``insert_rows`` shifts cell values but not merge
        # ranges, and once values have moved we can no longer cleanly
        # unmerge by coordinate. So: snapshot → unmerge → insert_rows → re-
        # merge at shifted coordinates.
        captured = _capture_merges_at_or_below(ws, pivot)
        captured_heights = _capture_row_heights_at_or_below(ws, pivot)
        ws.insert_rows(pivot, amount=delta)
        _reapply_shifted_merges(ws, captured, delta)
        # Move the (tall) totals/footer row heights down with their content so the
        # stamp keeps fitting; insert_rows does NOT move row_dimensions heights.
        _reapply_shifted_row_heights(ws, captured_heights, delta)
        # Re-anchor floating images (stamp/seal) sitting at-or-below the pivot
        # so they follow the totals/footer down instead of overlapping the
        # now-larger items table. insert_rows does NOT move drawing anchors.
        _shift_images_at_or_below(ws, pivot, delta)
        # Clone the styling of the last sample row into the newly inserted ones.
        for offset in range(delta):
            new_row = last_sample_row + 1 + offset
            _clone_row(ws, last_sample_row, new_row)
            _merge_b_c(ws, new_row)
    elif delta < 0:
        # Fewer items than template samples — delete extras and shift the
        # downstream merges up by the same amount.
        rows_to_drop = abs(delta)
        first_drop = last_sample_row - rows_to_drop + 1
        _unmerge_in_range(ws, first_drop, last_sample_row)
        captured = _capture_merges_at_or_below(ws, pivot)
        captured_heights = _capture_row_heights_at_or_below(ws, pivot)
        ws.delete_rows(first_drop, rows_to_drop)
        _reapply_shifted_merges(ws, captured, delta)
        # Move the (tall) totals/footer row heights up with their content so the
        # stamp keeps fitting; delete_rows does NOT move row_dimensions heights.
        _reapply_shifted_row_heights(ws, captured_heights, delta)
        # Re-anchor floating images (stamp/seal) up by the same (negative) delta
        # so they stay snug below the shrunken totals/footer block.
        _shift_images_at_or_below(ws, pivot, delta)

    items_last_row = ITEMS_FIRST_DATA_ROW + n_items - 1

    # ── Fill item rows ────────────────────────────────────────────────────
    for idx, item in enumerate(line_items, start=1):
        row = ITEMS_FIRST_DATA_ROW + idx - 1
        _merge_b_c(ws, row)
        ws.cell(row=row, column=1, value=idx)
        desc = _build_item_description(item)
        # ICE security #2: desc / uom / note all come from user-editable
        # sourcing entries — defuse formula triggers before writing.
        ws.cell(row=row, column=2, value=_defuse_formula(desc))
        ws.cell(row=row, column=2).alignment = Alignment(
            wrap_text=True,
            vertical=ws.cell(row=row, column=2).alignment.vertical or "center",
            horizontal=ws.cell(row=row, column=2).alignment.horizontal or "left",
        )
        qty = item.get("quantity") or 0
        ws.cell(row=row, column=4, value=float(qty) if qty else 0)
        ws.cell(row=row, column=5, value=_defuse_formula(item.get("uom") or "Cái"))
        unit = item.get("unit_price_vnd") or 0
        # F (Đơn giá) keeps the cloned accounting nf from the template sample
        # rows — do NOT override with "#,##0" (that wiped the accounting format).
        ws.cell(row=row, column=6, value=float(unit) if unit else 0)
        # Formula keeps the file editable in Excel — internal, NOT user input.
        # G (Thành tiền) also keeps the cloned accounting nf.
        ws.cell(row=row, column=7, value=f"=D{row}*F{row}")
        # H (Ghi chú): base note (or hs_code fallback) + optional delivery_time.
        # Append "Giao: {delivery_time}" without overwriting an existing note —
        # separate by " · " when a note already exists, else stand alone.
        note_text = item.get("note") or item.get("hs_code") or ""
        delivery_time = item.get("delivery_time")
        if delivery_time:
            giao = f"Giao: {delivery_time}"
            note_text = f"{note_text} · {giao}" if note_text else giao
        ws.cell(
            row=row,
            column=8,
            value=_defuse_formula(note_text),
        )

    # ── Totals block (3 rows, immediately after items) ────────────────────
    totals_start = items_last_row + 1
    subtotal_row = totals_start
    vat_row = totals_start + 1
    grand_row = totals_start + 2

    # The template already laid out the totals block when we loaded it; after
    # insert/delete_rows openpyxl has shifted it so subtotal_row/vat_row/
    # grand_row hold the correct cells. We only need to re-fill the formulas
    # so they reference the right item range.
    sum_range = f"G{ITEMS_FIRST_DATA_ROW}:G{items_last_row}"
    # H totals keep the cloned accounting nf from the template totals block —
    # do NOT override with "#,##0".
    ws[f"H{subtotal_row}"] = f"=SUM({sum_range})"
    # VAT label uses the configured rate (display as percent in the label).
    vat_pct_label = f"VAT {int(round(float(vat) * 100))}%"
    ws[f"F{vat_row}"] = vat_pct_label
    ws[f"H{vat_row}"] = f"=H{subtotal_row}*{float(vat)}"
    ws[f"H{grand_row}"] = f"=H{subtotal_row}+H{vat_row}"

    # Terms text (left side of the totals block, merged A:E across all 3 rows
    # in the original template). After row insertions the merge has shifted
    # to the new totals_start automatically.
    # ICE security #2: quote_note and terms_text are user-supplied (textarea).
    final_terms = terms_text or (quote_note.strip() if quote_note else "") or DEFAULT_TERMS
    ws[f"A{subtotal_row}"] = _defuse_formula(final_terms)

    # ── Footer ────────────────────────────────────────────────────────────
    words_row = totals_start + AMOUNT_WORDS_OFFSET
    # Compute display total (formula not yet evaluated by openpyxl) for the
    # "bằng chữ" line. Use the snapshot we already have on hand.
    try:
        grand_total = float(total_value or 0) * (1.0 + float(vat))
    except Exception:
        grand_total = 0.0
    ws[f"A{words_row}"] = (
        f"Thành tiền bằng chữ:  {_num2words_vn(grand_total)} đồng."
    )

    # The static "XÁC NHẬN BÁO GIÁ" / "XÁC NHẬN ĐẶT HÀNG" / thank-you cells
    # are part of the template — no need to rewrite them; openpyxl preserves
    # them through the insert_rows shift.

    # ── Print area (Bug V7) ───────────────────────────────────────────────
    # The template ships a hard-coded print_area of A1:H22 sized for exactly 3
    # sample rows. With N != 3 items the totals / signature / thank-you block
    # shifts but the print_area does NOT — so the footer (chữ ký, "XÁC NHẬN
    # BÁO GIÁ", thank-you) gets clipped off the printed/PDF page. Recompute it
    # to end at the thank-you row so the whole quote always prints.
    # NOTE: columns are unchanged — still A:H (the column mapping is correct).
    thank_you_row = totals_start + THANK_YOU_OFFSET
    ws.print_area = f"A1:H{thank_you_row}"

    # Bug (Thang 2026-06-30): for tall quotes (>= ~5 items) the footer block —
    # "XÁC NHẬN BÁO GIÁ" + the company stamp/seal + chữ ký GĐ — straddled the
    # A4 page break and the seal got CLIPPED. Force the whole footer block onto a
    # fresh page (a clean signature page, the norm for multi-page quotes) so the
    # stamp is always intact and positioned under the totals exactly like the
    # one-page reference. The template fits ~4 items + the full footer on one A4
    # page (items are a fixed 63pt each); only break for more than that, so small
    # quotes (<=4 items) stay a single page and are visually unchanged.
    #
    # Bug (Thang 2026-07-04, W3-04): the RAW template ships with a STALE manual
    # row-break baked in at row 25 of "mẫu báo giá.xlsx" — an editing artifact
    # with no cells anchored to it (rows 23-29 are all empty). Just like merges
    # / images / row-heights above, openpyxl's insert_rows() does NOT shift
    # ws.row_breaks, so this stray break stays pinned at its ORIGINAL row while
    # the footer block moves down for N > 3. For N ≈ 6-7 items (and worse for
    # more) the stale break lands INSIDE the footer — between the "XÁC NHẬN BÁO
    # GIÁ" + stamp row and the "THANK YOU" row, or straight through the 129pt
    # stamp image itself (it overflows its own 90.75pt-tall row by ~38pt) —
    # slicing the seal/signature across an extra, unintended page. Reset all
    # row breaks up front so the ONLY break present is the one we compute
    # intentionally below.
    ws.row_breaks.brk = []

    sign_row = totals_start + SIGN_OFFSET
    if n_items > FOOTER_FITS_ITEMS_PER_PAGE:
        from openpyxl.worksheet.pagebreak import Break

        # Break AFTER (sign_row - 1) → the footer block starts a fresh page.
        ws.row_breaks.append(Break(id=sign_row - 1))

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    # Post-process: openpyxl writes the stamp image rel with an ABSOLUTE path
    # (Target="/xl/media/image2.jpeg") which LibreOffice/Gotenberg reject when
    # converting to PDF — the xlsx opens fine in Excel but the PDF loses the
    # stamp. Rewrite to a relative target so the seal survives in the PDF.
    _fix_drawing_rels_paths(out_path)


def _build_item_description(item: dict) -> str:
    """Compose the B-column description used by the template's items table.

    Template puts product name + optional brand/model on multi-line cells.
    Mirror that here so the output looks like a real Song Châu báo giá."""
    name = (item.get("product_name") or "").strip()
    model = (item.get("model") or "").strip()
    brand = (item.get("brand") or item.get("maker") or "").strip()
    bits: list[str] = []
    if name:
        bits.append(name)
    detail = " ".join(b for b in (brand, model) if b).strip()
    if detail:
        bits.append(detail)
    return "\n".join(bits) if bits else (model or brand or "")


# ---------------------------------------------------------------------------
# TSV renderer — unchanged (raw export remains stdlib / fast)
# ---------------------------------------------------------------------------
def render_tsv(out_path: Path, line_items: list[dict]) -> None:
    """Render TSV — đơn giản, paste sang Excel/Sheets nhanh."""
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        w.writerow([
            "Model", "Tên sản phẩm", "Maker", "Brand", "Supplier",
            "HS Code", "Quantity", "Unit price (VND)", "Line total (VND)",
        ])
        # ICE security #2: TSV is the same formula-injection target as XLSX
        # the moment a user pastes it into Excel/Sheets — defuse all text cols.
        for item in line_items:
            w.writerow([
                _defuse_formula(item.get("model") or ""),
                _defuse_formula(item.get("product_name") or ""),
                _defuse_formula(item.get("maker") or ""),
                _defuse_formula(item.get("brand") or ""),
                _defuse_formula(item.get("supplier") or ""),
                _defuse_formula(item.get("hs_code") or ""),
                item.get("quantity") or 0,
                _fmt_vnd(item.get("unit_price_vnd")),
                _fmt_vnd(item.get("line_total_vnd")),
            ])


# ---------------------------------------------------------------------------
# Debug / introspection helper — kept private. Useful when iterating on the
# template layout: ``python -m app.services.quote_renderer dump``.
# ---------------------------------------------------------------------------
def _load_cells_map() -> dict:
    """Load the JSON placeholder map shipped alongside the template."""
    if _REPO_CELLS_MAP.exists():
        return json.loads(_REPO_CELLS_MAP.read_text(encoding="utf-8"))
    return {}


__all__ = ["render_xlsx", "render_tsv"]
