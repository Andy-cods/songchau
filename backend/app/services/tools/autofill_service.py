"""
Auto-Fill Quotation Service (M01).

Ported from tool1_autofill/engine.py — core business logic only.
Replaces: SQLite → asyncpg, LibreOffice → Gotenberg, filesystem → /data/files.

Core workflow:
  1. parse_bc_bqms()  — Parse uploaded BC BQMS Excel → list of order items
  2. classify_loai_hang() — Classify items as GC/TM
  3. lookup_prices()   — Match prices from bqms_rfq history
  4. fill_cam_ket()    — Fill CAM KET template Excel
  5. fill_quotation()  — Fill Commercial Quotation template Excel
  6. convert to PDF    — Via Gotenberg container
"""

from __future__ import annotations

import io
import logging
import re
import uuid
from datetime import date as _date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string, get_column_letter

try:
    from zoneinfo import ZoneInfo
except ImportError:  # py<3.9 fallback (project pins 3.11+, kept for safety)
    ZoneInfo = None  # type: ignore[assignment]

logger = logging.getLogger(__name__)

# Thang 2026-06-13 (Bug fix T3): canonical VN tz so quotation dates never
# drift when ops moves the worker process between regions.
_VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh") if ZoneInfo else None


def _vn_now() -> datetime:
    """Today in Asia/Ho_Chi_Minh tz, naive (no tzinfo) for openpyxl writes."""
    if _VN_TZ is None:
        return datetime.now()
    return datetime.now(tz=_VN_TZ).replace(tzinfo=None)

# ─── Constants ────────────────────────────────────────────────
FILES_BASE = Path("/data/files/quotations")

COL_KEYWORDS: dict[str, list[str]] = {
    "don_hang":   ["đơn hàng", "don hang", "rfq"],
    "bqms":       ["bqms"],
    "spec":       ["tên hàng", "ten hang", "spec"],
    "short_name": ["explain", "tên ngắn", "ten ngan"],
    "loai_hang":  ["loại hàng", "loai hang"],
    "maker":      ["maker"],
    "mark":       ["mark"],
    "don_vi":     ["đơn vị", "don vi", "unit"],
    "so_luong":   ["số lượng", "so luong", "qty", "quantity"],
    "han_bg":     ["hạn bg", "han bg", "deadline", "hạn"],
    "ghi_chu":    ["ghi chú", "ghi chu", "note"],
}


# ─── Utility Functions ────────────────────────────────────────

def clean_cell(val: Any) -> str:
    """Strip non-breaking spaces and whitespace."""
    if val is None:
        return ""
    return str(val).replace("\xa0", " ").strip()


def fuzzy_match_col(header_str: str, keywords: list[str]) -> bool:
    """Check if header string matches any keyword (case-insensitive substring)."""
    h = header_str.lower().replace("\xa0", " ").strip()
    return any(kw.lower() in h for kw in keywords)


def classify_loai_hang(val: str) -> str:
    """Classify product type: 'gc' (gia công), 'tm' (thương mại), or 'unknown'."""
    if not val:
        return "unknown"
    lh = val.lower().replace("\xa0", " ").strip()
    if any(x in lh for x in ("gia công", "gia cong", "gc")):
        return "gc"
    if any(x in lh for x in ("thương mại", "thuong mai", "tm")):
        return "tm"
    abbr = lh[:2]
    if abbr == "gc":
        return "gc"
    if abbr == "tm":
        return "tm"
    return "unknown"


def _parse_deadline(raw: str) -> datetime | None:
    """Parse Vietnamese deadline strings like '2/4 17h' → datetime."""
    if not raw:
        return None
    raw = clean_cell(raw)
    # Pattern: d/m HHh or d/m
    m = re.match(r"(\d{1,2})/(\d{1,2})\s*(\d{1,2})?h?", raw)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        hour = int(m.group(3)) if m.group(3) else 17
        year = datetime.now().year
        try:
            return datetime(year, month, day, hour)
        except ValueError:
            return None
    return None


def _is_deadline_urgent(dt: datetime | None) -> bool:
    """Check if deadline is within 48 hours."""
    if not dt:
        return False
    return dt - datetime.now() < timedelta(hours=48)


# ─── BC BQMS Excel Parsing ───────────────────────────────────

def parse_bc_bqms(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse a BC BQMS Excel file into a list of order items.

    Args:
        file_bytes: Raw bytes of the uploaded Excel file.

    Returns:
        List of order dicts with keys: id, don_hang, bqms, spec, short_name,
        loai_hang, maker, mark, don_vi, so_luong, han_bg, deadline_dt,
        is_urgent, ghi_chu.
    """
    import python_calamine
    from python_calamine import CalamineWorkbook

    wb = CalamineWorkbook.from_buffer(file_bytes)
    sheet_names = wb.sheet_names

    # Find best sheet + header row via scoring
    best_sheet, best_header = sheet_names[0], 0
    best_score = 0

    for name in sheet_names:
        rows = wb.get_sheet_by_name(name).to_python()
        for i, row in enumerate(rows[:15]):
            row_text = " ".join(str(v) for v in row if v is not None).lower()
            score = (
                (2 if "đơn hàng" in row_text or "don hang" in row_text else 0)
                + (2 if "bqms" in row_text else 0)
                + (1 if "maker" in row_text else 0)
                + (1 if "loại hàng" in row_text or "loai hang" in row_text else 0)
            )
            if score > best_score:
                best_score, best_sheet, best_header = score, name, i

    # Read data rows
    all_rows = wb.get_sheet_by_name(best_sheet).to_python()
    if not all_rows or best_header >= len(all_rows):
        return []

    headers = [clean_cell(h) for h in all_rows[best_header]]

    # Map columns
    col_map: dict[str, int] = {}
    for col_idx, header in enumerate(headers):
        for field, keywords in COL_KEYWORDS.items():
            if field not in col_map and fuzzy_match_col(header, keywords):
                col_map[field] = col_idx

    def get_field(row: list, field: str) -> str:
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return ""
        return clean_cell(row[idx])

    orders: list[dict[str, Any]] = []
    for row in all_rows[best_header + 1:]:
        don_hang = get_field(row, "don_hang")
        bqms = get_field(row, "bqms")
        if not don_hang and not bqms:
            continue
        if don_hang.lower() in ("đơn hàng", "don hang", "nan", ""):
            if not bqms:
                continue

        loai_raw = get_field(row, "loai_hang")
        loai_norm = classify_loai_hang(loai_raw).upper()

        so_luong_raw = get_field(row, "so_luong")
        try:
            so_luong = int(float(so_luong_raw)) if so_luong_raw else 0
        except (ValueError, TypeError):
            so_luong = 0

        han_bg = get_field(row, "han_bg")
        deadline_dt = _parse_deadline(han_bg)

        orders.append({
            "id": f"{don_hang}_{bqms}",
            "don_hang": don_hang,
            "bqms": bqms,
            "spec": get_field(row, "spec"),
            "short_name": get_field(row, "short_name"),
            "loai_hang": loai_norm,
            "maker": get_field(row, "maker"),
            "mark": get_field(row, "mark"),
            "don_vi": get_field(row, "don_vi") or "EA",
            "so_luong": so_luong,
            "han_bg": han_bg,
            "deadline_dt": deadline_dt.isoformat() if deadline_dt else None,
            "is_urgent": _is_deadline_urgent(deadline_dt),
            "ghi_chu": get_field(row, "ghi_chu"),
        })

    logger.info("Parsed BC BQMS: %d items from sheet '%s'", len(orders), best_sheet)
    return orders


# ─── Price Lookup ─────────────────────────────────────────────

async def lookup_prices(conn, items: list[dict]) -> list[dict]:
    """Lookup latest prices from bqms_rfq for each item.

    Adds 'price_history' list and 'suggested_price' to each item dict.
    """
    for item in items:
        bqms_code = item.get("bqms", "")
        if not bqms_code:
            item["price_history"] = []
            item["suggested_price"] = None
            continue

        rows = await conn.fetch(
            """
            SELECT bqms_code, quoted_price_bqms_v1, quoted_price_bqms_v2,
                   quoted_price_bqms_v3, quoted_price_bqms_v4,
                   result, maker, specification, created_at
            FROM bqms_rfq
            WHERE bqms_code = $1
            ORDER BY created_at DESC
            LIMIT 10
            """,
            bqms_code,
        )

        history = []
        for r in rows:
            prices = [
                r["quoted_price_bqms_v1"],
                r["quoted_price_bqms_v2"],
                r["quoted_price_bqms_v3"],
                r["quoted_price_bqms_v4"],
            ]
            latest_price = next((p for p in reversed(prices) if p), None)
            history.append({
                "prices": [float(p) if p else None for p in prices],
                "result": r["result"],
                "latest_price": float(latest_price) if latest_price else None,
                "date": r["created_at"].isoformat() if r["created_at"] else None,
            })

        # Suggested price = latest winning price, or latest v1
        won_prices = [h["latest_price"] for h in history if h["result"] and "won" in str(h["result"]).lower() and h["latest_price"]]
        if won_prices:
            suggested = won_prices[0]
        elif history and history[0]["latest_price"]:
            suggested = history[0]["latest_price"]
        else:
            suggested = None

        item["price_history"] = history
        item["suggested_price"] = suggested

    return items


# ─── Image Matching ───────────────────────────────────────────

_RFQ_STAGING_ROOT = Path('/data/onedrive-staging/Puplic/BQMS/RFQ')
_IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp'}


def discover_rfq_images(rfq_no: str) -> list[tuple[str, bytes]]:
    """Find images on disk for a given RFQ.

    Walks /data/onedrive-staging/Puplic/BQMS/RFQ/RFQ {year}/THANG */{rfq_no}*/
    and returns (filename, bytes) for every PNG/JPG.

    Layout (per bqms_bidding_scraper.download_files_for_rfq):
      <rfq_dir>/
        ├── raw/      <- original .xlsx + attachments
        └── images/   <- extracted images, named after item codes
                        (e.g. Z0000002-544469_1.png, _2.png, _3.png)

    Strategy: scan recursively (rglob) so we catch images in any nested
    subdir — was a bug pre-2026-05-10 that only iterdir()'d the rfq_dir
    root and missed the entire images/ subdir.

    Returns empty list if no folder found -- silent fallback so the rest of
    the pipeline still runs.
    """
    if not rfq_no or rfq_no == 'UNKNOWN' or not _RFQ_STAGING_ROOT.exists():
        return []

    out: list[tuple[str, bytes]] = []
    try:
        for year_dir in _RFQ_STAGING_ROOT.glob('RFQ */'):
            for month_dir in year_dir.glob('THANG */'):
                for rfq_dir in month_dir.glob(f'{rfq_no}*/'):
                    # Recursive — catches both <rfq>/images/*.png and any
                    # legacy layout that drops images at <rfq>/*.png.
                    for img_path in rfq_dir.rglob('*'):
                        if img_path.is_file() and img_path.suffix.lower() in _IMAGE_EXTS:
                            try:
                                out.append((img_path.name, img_path.read_bytes()))
                            except OSError as exc:
                                logger.debug('discover_rfq_images: skip %s: %s', img_path, exc)
    except OSError as exc:
        logger.warning('discover_rfq_images: walk failed: %s', exc)

    if out:
        logger.info('discover_rfq_images: %s -> %d images', rfq_no, len(out))
    else:
        logger.info('discover_rfq_images: %s -> 0 images (folder may not exist or no .png/.jpg)', rfq_no)
    return out


def match_images_to_orders(
    uploaded_files: list[tuple[str, bytes]],
    orders: list[dict],
) -> dict[str, bytes]:
    """Match uploaded image files to BQMS codes by filename.

    Strict matching policy (2026-05-10): the FULL bqms code must appear
    inside the filename (case-insensitive). Longest codes win first to
    avoid mis-mapping when codes share a prefix
    (e.g. Z0000002-508700 vs Z0000002-508701: filename containing
    "Z0000002-508700" matches only the first, not both). When multiple
    image files map to the same code, the largest payload wins (real
    product photos are usually larger than placeholder/icon variants).

    Per Thang 2026-05-10: 1 RFQ may have N items, user only báo giá
    a subset → mapping must be precise so a wrong item never gets
    another item's image.

    Args:
        uploaded_files: List of (filename, file_bytes) tuples.
        orders: List of order dicts (must have 'bqms' key).

    Returns:
        Dict mapping bqms_code → image bytes (only keys for orders we
        actually have a confident image for; missing ones simply absent).
    """
    images_map: dict[str, bytes] = {}
    # Distinct codes only, sorted longest-first so specific codes match
    # before any shorter prefix that might also be a substring.
    bqms_codes = sorted(
        {(o["bqms"] or "").strip() for o in orders if o.get("bqms")},
        key=len, reverse=True,
    )
    bqms_codes = [c for c in bqms_codes if c]
    if not bqms_codes:
        return images_map

    unmatched_unmapped: list[str] = []
    for fname, img_bytes in uploaded_files:
        fname_upper = Path(fname).stem.upper()
        matched = False
        for code in bqms_codes:
            if code.upper() in fname_upper:
                # Strict full-code presence; first (longest) wins.
                existing = images_map.get(code)
                if existing is None or len(img_bytes) > len(existing):
                    images_map[code] = img_bytes
                    logger.info(
                        "Image '%s' (%d bytes) matched to BQMS %s",
                        fname, len(img_bytes), code,
                    )
                matched = True
                break  # one filename → one code
        # Defensive (Thang 2026-06-03): if a `_unmapped_*` file from the
        # bidding scraper falls through without matching any code, that's
        # the smoking gun for "image missing on báo giá" — call it out
        # individually so the operator sees exactly which file needs
        # renaming.
        if not matched and "_unmapped_" in fname.lower():
            unmatched_unmapped.append(fname)

    if unmatched_unmapped:
        logger.warning(
            "match_images_to_orders: %d scraper-unmapped file(s) did not "
            "match any bqms code via substring (candidate codes: %s). "
            "Files: %s. Rename to '{full_bqms_code}_1.ext' to fix.",
            len(unmatched_unmapped), bqms_codes[:10], unmatched_unmapped[:10],
        )

    return images_map


# ─── Excel Cell Helpers ───────────────────────────────────────

def _get_top_left_of_merged(ws, row: int, col: int) -> tuple[int, int]:
    """If (row, col) is in a merged range, return its top-left cell."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_row, mr.min_col
    return row, col


def _safe_set_cell(ws, row: int, col: int, value: Any) -> None:
    """Set cell value, handling merged cells."""
    r, c = _get_top_left_of_merged(ws, row, col)
    ws.cell(row=r, column=c).value = value


# ─── Defensive template fixups (Thang 2026-05-21) ────────────────────
#
# These run AFTER `openpyxl.load_workbook(template_path)` and BEFORE the
# row-specific writes. They scan the entire sheet for problem patterns
# inherited from the template binary (which the user can't easily edit):
#   1. Stale date strings that don't get updated by per-cell writes when
#      regenerating V2/V3 — defensively replace ANY cell matching
#      "Ngày X Tháng Y năm Z" / "dd/mm/yyyy" with today.
#   2. Bullet text "-Cam kết..." missing space after dash → fix to "- Cam kết".

import re as _re

_DATE_VN_PAT = _re.compile(
    r"Ngày\s*\d{1,2}\s*Tháng\s*\d{1,2}\s*năm\s*\d{4}",
    flags=_re.IGNORECASE,
)
_DATE_SLASH_PAT = _re.compile(r"\b\d{1,2}/\d{1,2}/\d{4}\b")
_DATE_DMY_PAT = _re.compile(r"\b\d{1,2}-\d{1,2}-\d{4}\b")
_DASH_BULLET_PAT = _re.compile(r"^(-)(?=[A-ZĐÀÁẢÃẠÂẦẤẨẪẬĂẰẮẲẴẶÈÉẺẼẸÊỀẾỂỄỆÌÍỈĨỊÒÓỎÕỌÔỒỐỔỖỘƠỜỚỞỠỢÙÚỦŨỤƯỪỨỬỮỰỲÝỶỸỴ])")


def _refresh_dates_in_sheet(ws, now) -> int:
    """Scan ALL cells; replace stale date strings with `now`'s date.

    Returns count of cells updated.
    Catches the V2/V3 bug where template had a hardcoded date that didn't
    get overwritten by the explicit per-cell date write (e.g. there are
    MULTIPLE date cells but code only touches one).
    """
    vn_today = f"Ngày {now.day} Tháng {now.month} năm {now.year}"
    slash_today = now.strftime("%d/%m/%Y")
    updated = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v:
                continue
            new_v = v
            # Pattern 1: VN "Ngày X Tháng Y năm Z" — replace anywhere in cell
            if _DATE_VN_PAT.search(new_v):
                new_v = _DATE_VN_PAT.sub(vn_today, new_v)
            # Pattern 2: "dd/mm/yyyy" — only replace if cell is a pure date
            # (avoid corrupting e.g. "Order received on 12/05/2026 by ABC").
            stripped = new_v.strip()
            if _DATE_SLASH_PAT.fullmatch(stripped):
                new_v = slash_today
            if _DATE_DMY_PAT.fullmatch(stripped):
                new_v = now.strftime("%d-%m-%Y")
            if new_v != v:
                try:
                    cell.value = new_v
                    updated += 1
                except Exception:
                    pass
    return updated


def _fix_dash_bullets_in_sheet(ws) -> int:
    """Normalize leading dash bullets so they render cleanly in PDF.

    Templates (often Word-pasted) carry inconsistent bullet patterns:
        "-Cam kết..."     (no space)              → "- Cam kết..."
        "-\\tCam kết..."   (dash + tab)            → "- Cam kết..."
        "-  Cam kết..."   (dash + multi-space)   → "- Cam kết..."
        "- Cam kết..."    (already OK)            → unchanged
    Plus normalize `\\r\\n` / `\\r` → `\\n` so cell line breaks render
    consistently in LibreOffice/Gotenberg.

    Thang 2026-05-22: original regex skipped tab → "-\\tCam kết..." in
    CAM_KET template was untouched and rendered as "-Cam kết" in PDF.
    Now treats `\\t` and multi-space as separator → forces canonical
    "- <text>" with exactly one space after the dash.
    """
    updated = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or "-" not in v:
                continue
            normalized = v.replace("\r\n", "\n").replace("\r", "\n")
            new_v = normalized
            lines = new_v.split("\n")
            new_lines = []
            line_changed = False
            for line in lines:
                stripped_left = line.lstrip()
                if (stripped_left.startswith("-")
                        and len(stripped_left) > 1
                        and stripped_left[1] not in ("-", ">")):
                    # Strip ALL whitespace right after the dash, then re-add
                    # exactly one space. Handles: "-X" (no sep), "-\tX"
                    # (tab), "-  X" (multi-space). Skip if the next char
                    # is "-" (em-dash style) or ">" (e.g. "->", arrow).
                    rest = stripped_left[1:].lstrip(" \t")
                    if not rest:
                        new_lines.append(line)
                        continue
                    indent = line[:len(line) - len(stripped_left)]
                    fixed = indent + "- " + rest
                    if fixed != line:
                        new_lines.append(fixed)
                        line_changed = True
                    else:
                        new_lines.append(line)
                else:
                    new_lines.append(line)
            if line_changed or normalized != v:
                new_v = "\n".join(new_lines)
                try:
                    cell.value = new_v
                    updated += 1
                except Exception:
                    pass
    return updated


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 20) -> None:
    """Copy font, fill, alignment, border, number_format from src to dst."""
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst.font = src.font.copy()
            dst.fill = src.fill.copy()
            dst.alignment = src.alignment.copy()
            dst.border = src.border.copy()
            dst.number_format = src.number_format


def _copy_merged_ranges_for_row(ws, src_row: int, dst_row: int) -> None:
    """Copy merged cell ranges from src_row to dst_row."""
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == src_row and mr.max_row == src_row:
            ws.merge_cells(
                start_row=dst_row, start_column=mr.min_col,
                end_row=dst_row, end_column=mr.max_col,
            )


def _pil_image_to_xl(img_bytes: bytes, max_w: int = 110, max_h: int = 75) -> XLImage | None:
    """Convert image bytes → openpyxl Image, resized with PIL.

    Default 110×75 px fits a single template cell (col ~13 chars wide ≈ 95 px,
    row height ≈ 60-90 px for the 3-row CAM KET blocks). Override max_w/max_h
    for bigger renders.
    """
    try:
        from PIL import Image as PILImage

        pil_img = PILImage.open(io.BytesIO(img_bytes))
        if pil_img.mode not in ("RGB", "RGBA"):
            pil_img = pil_img.convert("RGBA")
        # Pillow 12: prefer Resampling enum; falls back to module-level
        # alias on older Pillow installs.
        _lanczos = getattr(PILImage, "Resampling", PILImage).LANCZOS
        pil_img.thumbnail((max_w, max_h), _lanczos)
        w, h = pil_img.size
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.width = w
        xl_img.height = h
        return xl_img
    except Exception as exc:
        logger.warning("Image conversion error: %s", exc)
        return None


def _cell_pixel_size(
    ws, row: int, col: int, n_cols: int = 1, n_rows: int = 1,
) -> tuple[int, int]:
    """Approximate cell pixel size from openpyxl col widths + row heights.

    Excel units conversion (rough): 1 col-width-unit ≈ 7 px, 1 row-height-pt
    ≈ 1.333 px (96 dpi). Default col width = 8.43 units, default row = 15 pt.
    Subtract ~6 px margin so the image doesn't kiss the cell border.
    """
    width_units = 0.0
    for c in range(col, col + n_cols):
        w = ws.column_dimensions[get_column_letter(c)].width
        width_units += (w if w and w > 0 else 8.43)
    height_pts = 0.0
    for r in range(row, row + n_rows):
        h = ws.row_dimensions[r].height
        height_pts += (h if h and h > 0 else 15.0)
    width_px = max(40, int(width_units * 7.0) - 6)
    height_px = max(30, int(height_pts * 1.333) - 6)
    return width_px, height_px


def fix_drawing_rels_in_xlsx(xlsx_path: str) -> int:
    """Rewrite ABSOLUTE paths in ALL .rels files to relative paths AND
    strip openpyxl artifacts that crash OnlyOffice's spreadsheet renderer.

    OnlyOffice issues fixed here (Thang 2026-05-21):
    A) Absolute Target paths in rels → LibreOffice/x2t drops them. Excel
       accepts them. Fix: rewrite "/xl/X" to relative path from the rels
       file's owner directory.
    B) `<definedName name="_xlnm.Print_Area" localSheetId="0">'Sheet'!$A$1:
       $K$40</definedName>` carried over from openpyxl template → when
       OnlyOffice's sdkjs/cell parses it, the sheet reference resolution
       hits null → "Cannot read properties of null (reading 'mb')" →
       editor crashes with code -82. Fix: strip <definedNames> block.
    C) Stale formula `<f>SUM(G17:G17)</f>` over a single cell carried from
       template (`fitToPage`'d but the row contains insert_rows shifted
       boundaries) → OnlyOffice mis-resolves → mb null. Same fix as B —
       safer to strip these too.

    Per Thang 2026-05-11/21: openpyxl writes rels with absolute Target paths
    (`Target="/xl/worksheets/sheet1.xml"`) which:
      - LibreOffice/Gotenberg silently drops the resource → blank PDF
      - OnlyOffice's x2t converter rejects the file → exit code 88 →
        Editor.bin never generated → editor shows error -4 "Tải về không
        thành công" (the bug user has been hitting on Cam kết files).

    Standard xlsx (and Excel-saved files) use RELATIVE paths
    (`Target="../worksheets/sheet1.xml"` from a file in xl/_rels/, or
    `Target="worksheets/sheet1.xml"` depending on the source folder).

    Fix: for each `<Relationship Target="/xl/<rest>" />`, compute the proper
    relative path from the .rels file's directory and substitute. The rels
    file lives at e.g. `xl/_rels/workbook.xml.rels` → directory `xl/_rels/`
    → relative to xl/ is `..`, so the target becomes `../<rest>`. For
    `xl/drawings/_rels/drawing1.xml.rels` → directory `xl/drawings/_rels/`
    → ../media/image1.png.

    Returns count of cells rewritten. Idempotent — safe to call multiple
    times. Should be called after wb.save() but before PDF conversion.
    """
    import zipfile, shutil, os, re
    n = 0
    src = Path(xlsx_path)
    if not src.exists():
        return 0
    tmp_path = src.with_suffix(src.suffix + ".rels-fix.tmp")

    # Match ANY Target="..." attribute (not just absolute) — we'll re-resolve
    # every target to a canonical relative path. This handles ALL the broken
    # formats my previous code has produced over time:
    #   Target="/xl/media/X"           (openpyxl absolute)
    #   Target="xl/media/X"            (buggy partial fix v1, no leading /)
    #   Target="media/X"               (buggy partial fix v2, no ../)
    #   Target="../media/X"            (already correct)
    target_pat = re.compile(rb'Target="([^"]+)"')

    def _build_known_paths(zip_names: set[str]) -> set[str]:
        """Return the set of all internal paths in the zip (forward-slash,
        no leading slash). Used to validate Target resolution."""
        return {n.replace("\\", "/").lstrip("/") for n in zip_names}

    def _resolve_to_zip_path(target: str, owner_dir: str, known: set[str]) -> str | None:
        """Turn a Target attribute (any of the broken formats) into the
        actual file path inside the zip. Returns None if target is external
        (http/mailto) or already a valid relative path that points outside
        the zip and we shouldn't touch.

        Strategy: try every candidate interpretation, return the first one
        that matches a real file in the zip.
        """
        if target.startswith(("http://", "https://", "mailto:", "//")):
            return None  # external — leave alone
        cleaned = target.replace("\\", "/")
        candidates: list[str] = []
        # 1. Already absolute (`/xl/...` or `xl/...`)
        candidates.append(cleaned.lstrip("/"))
        # 2. Resolve as relative from owner_dir
        try:
            joined = os.path.normpath(os.path.join(owner_dir, cleaned)).replace("\\", "/")
            candidates.append(joined.lstrip("/"))
        except Exception:
            pass
        # 3. If target looks like "media/X" or "drawings/X" / "worksheets/X"
        #    and owner_dir is inside xl/, the user PROBABLY meant xl/<target>
        #    (which is what my buggy fix produced for drawing rels).
        if owner_dir.startswith("xl") and not cleaned.startswith(("xl/", "../")):
            candidates.append("xl/" + cleaned)
        # 4. Same but for root rels file (_rels/.rels)
        if not owner_dir:
            candidates.append(cleaned)

        for c in candidates:
            if c in known:
                return c
        return None

    def _rewrite_rels_data(rels_path: str, data: bytes, known: set[str]) -> tuple[bytes, int]:
        """Rewrite Target attributes in a .rels file to canonical relative
        paths from the rels file's owner directory."""
        # rels_path example: "xl/drawings/_rels/drawing1.xml.rels"
        rels_dir = os.path.dirname(rels_path)  # "xl/drawings/_rels"
        owner_dir = rels_dir.rstrip("/").rsplit("/_rels", 1)[0]  # "xl/drawings"
        # Special case: rels file at root level ("_rels/.rels")
        if owner_dir == "_rels":
            owner_dir = ""
        local_n = 0

        def _sub(m: "re.Match[bytes]") -> bytes:
            nonlocal local_n
            raw = m.group(1).decode("utf-8", errors="replace")
            zip_path = _resolve_to_zip_path(raw, owner_dir, known)
            if zip_path is None:
                # External or unresolvable — leave alone
                return m.group(0)
            # Compute canonical relative path from owner_dir to zip_path
            try:
                rel = os.path.relpath(zip_path, owner_dir or ".").replace("\\", "/")
            except ValueError:
                rel = zip_path
            # If unchanged, no need to log
            if rel == raw:
                return m.group(0)
            local_n += 1
            return f'Target="{rel}"'.encode("utf-8")

        new_data = target_pat.sub(_sub, data)
        return new_data, local_n

    try:
        with zipfile.ZipFile(src, "r") as zin:
            known_paths = _build_known_paths(set(zin.namelist()))
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.endswith(".rels") and "_rels" in item.filename:
                        new_data, count = _rewrite_rels_data(
                            item.filename, data, known_paths,
                        )
                        if count > 0:
                            n += count
                            logger.info("rels fix: %s — %d Target(s) rewritten",
                                        item.filename, count)
                        data = new_data
                    # Strip <definedNames> from workbook.xml (OnlyOffice -82 crash)
                    elif item.filename == "xl/workbook.xml":
                        try:
                            text = data.decode("utf-8")
                            new_text = re.sub(
                                r'<definedNames>.*?</definedNames>',
                                '<definedNames/>',
                                text,
                                flags=re.DOTALL,
                            )
                            if new_text != text:
                                data = new_text.encode("utf-8")
                                n += 1
                                logger.info("Stripped <definedNames> from %s (OO -82 fix)",
                                            item.filename)
                        except UnicodeDecodeError:
                            pass
                    zout.writestr(item, data)
        shutil.move(str(tmp_path), str(src))
    except Exception as exc:
        logger.warning("fix_drawing_rels_in_xlsx(%s) failed: %s", xlsx_path, exc)
        if tmp_path.exists():
            try: tmp_path.unlink()
            except Exception: pass
    if n:
        logger.info("fix_drawing_rels_in_xlsx: rewrote %d absolute Target path(s) in %s",
                    n, xlsx_path)
    return n


def _attach_image_fit(
    ws,
    img_bytes: bytes,
    anchor_row: int,
    anchor_col: int,
    *,
    span_cols: int = 1,
    span_rows: int = 1,
) -> bool:
    """Attach image CENTERED + ASPECT-PRESERVED inside (span_rows × span_cols).

    Thang 2026-05-22: switched from OneCellAnchor → TwoCellAnchor with the
    `to` marker computed precisely from the image's bottom-right pixel
    position inside the merge range. This pins the image WITHIN the cell
    range bounds in both LibreOffice + Excel + OnlyOffice — OneCellAnchor
    only specifies the size and top-left, so renderers that resize cells
    (Gotenberg fitToPage) may misposition the bottom edge.

    Aspect ratio is preserved (PIL thumbnail keeps ratio). Image is centered.
    fill_ratio 0.85 → 7.5% safety on each side. editAs="oneCell" keeps the
    image fixed at its computed size when user resizes cells in Excel.
    """
    from openpyxl.drawing.spreadsheet_drawing import (
        OneCellAnchor, TwoCellAnchor, AnchorMarker,
    )
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    target_w, target_h = _cell_pixel_size(
        ws, anchor_row, anchor_col, n_cols=span_cols, n_rows=span_rows,
    )
    fill_ratio = 0.85
    safety = 6
    img_max_w = max(20, int(target_w * fill_ratio))
    img_max_h = max(20, int(target_h * fill_ratio))
    xl_img = _pil_image_to_xl(img_bytes, max_w=img_max_w, max_h=img_max_h)
    if not xl_img:
        return False

    img_w = int(xl_img.width or img_max_w)
    img_h = int(xl_img.height or img_max_h)
    if img_w > target_w - safety * 2:
        scale = (target_w - safety * 2) / img_w
        img_w = max(20, int(img_w * scale))
        img_h = max(20, int(img_h * scale))
    if img_h > target_h - safety * 2:
        scale = (target_h - safety * 2) / img_h
        img_w = max(20, int(img_w * scale))
        img_h = max(20, int(img_h * scale))
    xl_img.width = img_w
    xl_img.height = img_h
    off_x = max(safety, (target_w - img_w) // 2)
    off_y = max(safety, (target_h - img_h) // 2)

    # ── Compute TwoCellAnchor `to` corner ─────────────────────
    # Image bottom-right pixel offset from anchor cell's top-left:
    target_br_x = off_x + img_w
    target_br_y = off_y + img_h

    # Walk through span_cols, finding which col contains target_br_x and the
    # remaining pixel offset within that col.
    to_col_idx = anchor_col - 1  # 0-indexed
    to_col_off_px = target_br_x
    cum = 0
    for offset in range(span_cols):
        c_letter = get_column_letter(anchor_col + offset)
        w_units = ws.column_dimensions[c_letter].width or 8.43
        w_px = max(1, int(w_units * 7.0))
        if cum + w_px >= target_br_x:
            to_col_idx = (anchor_col + offset) - 1
            to_col_off_px = target_br_x - cum
            break
        cum += w_px
    else:
        # target_br_x exceeds total span — clamp to last col's full width
        to_col_idx = (anchor_col + span_cols - 1) - 1
        last_w_units = (
            ws.column_dimensions[get_column_letter(anchor_col + span_cols - 1)].width
            or 8.43
        )
        to_col_off_px = max(1, int(last_w_units * 7.0))

    to_row_idx = anchor_row - 1
    to_row_off_px = target_br_y
    cum = 0
    for offset in range(span_rows):
        r = anchor_row + offset
        h_pts = ws.row_dimensions[r].height or 15.0
        h_px = max(1, int(h_pts * 1.333))
        if cum + h_px >= target_br_y:
            to_row_idx = r - 1
            to_row_off_px = target_br_y - cum
            break
        cum += h_px
    else:
        to_row_idx = (anchor_row + span_rows - 1) - 1
        last_h_pts = (
            ws.row_dimensions[anchor_row + span_rows - 1].height or 15.0
        )
        to_row_off_px = max(1, int(last_h_pts * 1.333))

    try:
        _from = AnchorMarker(
            col=anchor_col - 1, colOff=pixels_to_EMU(off_x),
            row=anchor_row - 1, rowOff=pixels_to_EMU(off_y),
        )
        to = AnchorMarker(
            col=to_col_idx, colOff=pixels_to_EMU(to_col_off_px),
            row=to_row_idx, rowOff=pixels_to_EMU(to_row_off_px),
        )
        xl_img.anchor = TwoCellAnchor(_from=_from, to=to, editAs="oneCell")
        ws._images.append(xl_img)
        return True
    except Exception as exc:
        logger.warning("TwoCellAnchor failed (%s) — fallback OneCellAnchor", exc)
        try:
            marker = AnchorMarker(
                col=anchor_col - 1, colOff=pixels_to_EMU(off_x),
                row=anchor_row - 1, rowOff=pixels_to_EMU(off_y),
            )
            ext = XDRPositiveSize2D(
                cx=pixels_to_EMU(img_w), cy=pixels_to_EMU(img_h),
            )
            xl_img.anchor = OneCellAnchor(_from=marker, ext=ext)
            ws._images.append(xl_img)
            return True
        except Exception:
            try:
                ws.add_image(xl_img, f"{get_column_letter(anchor_col)}{anchor_row}")
                return True
            except Exception:
                return False


def _attach_image_fit_DEAD(  # kept for diff-clarity, never called
    ws,
    img_bytes: bytes,
    anchor_row: int,
    anchor_col: int,
    *,
    span_cols: int = 1,
    span_rows: int = 1,
) -> bool:
    """Old TwoCellAnchor implementation — kept only for diff visibility."""
    target_w, target_h = _cell_pixel_size(
        ws, anchor_row, anchor_col, n_cols=span_cols, n_rows=span_rows,
    )
    img_max_w = max(20, int(target_w * 0.88))
    img_max_h = max(20, int(target_h * 0.88))
    xl_img = _pil_image_to_xl(img_bytes, max_w=img_max_w, max_h=img_max_h)
    if not xl_img:
        return False
    try:
        from openpyxl.drawing.spreadsheet_drawing import (
            TwoCellAnchor, AnchorMarker,
        )
        from openpyxl.utils.units import pixels_to_EMU

        img_w = int(xl_img.width or img_max_w)
        img_h = int(xl_img.height or img_max_h)
        # Center horizontally + vertically inside the span.
        off_x = max(0, (target_w - img_w) // 2)
        off_y = max(0, (target_h - img_h) // 2)
        # Last cell of the span (1-indexed → 0-indexed).
        end_col0 = anchor_col - 1 + max(0, span_cols - 1)
        end_row0 = anchor_row - 1 + max(0, span_rows - 1)
        # End offset within the LAST cell = total cell area minus image
        # tail margin. We compute the LAST cell's own width/height,
        # then derive how far from its top-left the image's bottom-right lands.
        last_col_w_units = (
            ws.column_dimensions[get_column_letter(anchor_col + span_cols - 1)].width
            or 8.43
        )
        last_row_h_pts = (
            ws.row_dimensions[anchor_row + span_rows - 1].height or 15.0
        )
        last_col_w_px = int(last_col_w_units * 7.0)
        last_row_h_px = int(last_row_h_pts * 1.333)
        # End offsets = last cell width/height minus tail margin (off_x/off_y).
        end_col_off_px = max(2, last_col_w_px - off_x)
        end_row_off_px = max(2, last_row_h_px - off_y)

        _from = AnchorMarker(
            col=anchor_col - 1, colOff=pixels_to_EMU(off_x),
            row=anchor_row - 1, rowOff=pixels_to_EMU(off_y),
        )
        to = AnchorMarker(
            col=end_col0, colOff=pixels_to_EMU(end_col_off_px),
            row=end_row0, rowOff=pixels_to_EMU(end_row_off_px),
        )
        xl_img.anchor = TwoCellAnchor(_from=_from, to=to, editAs="oneCell")
        ws._images.append(xl_img)
        return True
    except Exception as exc:
        # Fallback to plain top-left anchor if TwoCellAnchor API fails.
        logger.warning("TwoCellAnchor failed (%s) — falling back to top-left", exc)
        ws.add_image(xl_img, f"{get_column_letter(anchor_col)}{anchor_row}")
        return True


def _cleanup_data_area_merges(ws, data_start: int, scan_rows: int = 30) -> None:
    """Drop wide merge ranges (4+ cols) inside the item-data band.

    The CAM KET / Quotation templates carry a "Grand Total" row at the
    bottom of the placeholder area, merged A:E. Even after delete_rows()
    removes the ROW, openpyxl can leave the merge range pinned at the same
    coordinates — which then bleeds into row 18 (item 2) and centers our
    Maker text across A:E with empty Items + Image cells. Per Thang
    2026-05-10 screenshot: that's the "MICRO MOTION TECHNOLOGY centered
    across the row" bug.

    We only kill WIDE merges (≥4 cols) since narrow header/footer merges
    (e.g. A21:B21 for Terms labels) are intentional.
    """
    from openpyxl.utils.cell import range_boundaries
    to_drop: list[str] = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        if (data_start <= min_row <= data_start + scan_rows
                and max_row <= data_start + scan_rows
                and (max_col - min_col) >= 3):  # 4+ cols wide
            to_drop.append(str(mr))
    for mr in to_drop:
        try:
            ws.unmerge_cells(mr)
            logger.info("cleanup: unmerged stale wide merge %s", mr)
        except Exception:
            pass


# ─── Quotation fixups (post-load template cleanup) ───────────

def _fixup_quotation_template(ws) -> None:
    """Apply known fixes to the loaded Quotation template that the user
    asked for (Thang 2026-05-10):

    - Hide cols L..S (internal cost-tracking cols, not for vendor view)
    - Find rows containing "Validity" / terms text → bump row height so
      multi-line wrap_text actually shows both lines
    - Set page to LANDSCAPE + fit-to-page so the wide table (10 cols,
      ~980 px) fits on one PDF page horizontally and the items table
      doesn't get pushed onto page 2 vertically (Thang reported "vỡ
      bảng bị thụt xuống 2 trang").
    - Wrap-text on the Items col (C) and dynamic row height per item
      are handled where items are written, not here.

    Signature row fixup is handled separately AFTER all item inserts
    finish (since rows shift)."""
    # 1. Hide internal cost-tracking cols
    for col_letter in ("L", "M", "N", "O", "P", "Q", "R", "S"):
        ws.column_dimensions[col_letter].hidden = True

    # 2. Bump row height for terms-block rows (Price/Payment/Delivery/
    #    Packing/Validity) — each contains Vietnamese + English on 2 lines.
    #    Scan col A for known prefixes.
    TERM_PREFIXES = ("price", "payment", "delivery", "packing", "validity")
    for r in range(1, min(ws.max_row + 1, 100)):
        a_val = ws.cell(row=r, column=1).value
        if not isinstance(a_val, str):
            continue
        a_low = a_val.lower().strip()
        if any(a_low.startswith(p) for p in TERM_PREFIXES):
            current = ws.row_dimensions[r].height or 15
            if current < 32:
                ws.row_dimensions[r].height = 32

    # 3. Page setup — LANDSCAPE + shrink-to-fit. Table is 980+ px wide,
    #    portrait page is ~816 px. Without this, LibreOffice/Gotenberg
    #    splits the table across 2 pages.
    try:
        # Thang 2026-05-20: switched to A4 portrait per user request.
        # fitToWidth=1 scales wide content to fit portrait width — image
        # column + price columns will shrink but stay on one page.
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0 = unlimited rows, just fit width
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        # Tighter print margins so we get more usable area.
        from openpyxl.worksheet.page import PageMargins
        ws.page_margins = PageMargins(
            left=0.4, right=0.4, top=0.5, bottom=0.5,
            header=0.2, footer=0.2,
        )
    except Exception as exc:
        logger.warning("page setup failed: %s", exc)

    # 4. Vendor info rows (Name / Address / Tel / Email / Bank info).
    #    Per Thang 2026-05-11: previous 18pt cap was too tight — let rows
    #    breathe with min 22pt. Address rows (multi-line text) need more
    #    height to avoid wrapping into next row.
    # Scan col A for labels to set per-row heights correctly.
    LABEL_HEIGHT_MAP = {
        "name": 24,
        "address": 32,      # 2-line address needs more space
        "tel": 22,
        "tel(mobile)": 22,
        "e-mail": 22,
        "email": 22,
        "bank accounts": 28,  # "Bank accounts / Tai khoan NH" two-line label
        "tai khoan": 28,
        "product description": 26,
        "tên hàng": 26,
    }
    for r in range(1, min(ws.max_row + 1, 40)):
        a_val = ws.cell(row=r, column=1).value
        if not isinstance(a_val, str):
            continue
        a_low = a_val.lower().strip().rstrip(":").rstrip("/").strip()
        for prefix, h_target in LABEL_HEIGHT_MAP.items():
            if a_low.startswith(prefix):
                current = ws.row_dimensions[r].height or 0
                if current < h_target:
                    ws.row_dimensions[r].height = h_target
                break


def _fixup_signatures(ws) -> None:
    """Locate the signature row (rows whose col A contains 'AMA' AND
    col F contains a vendor counterparty marker), overwrite both cells
    with canonical text, AND re-merge A:E + F:I so the long company
    names render across the merged span instead of wrapping inside
    the narrow col A / col F.

    Defensive against row shifts AND merge loss caused by repeated
    insert_rows()/delete_rows() calls above."""
    from openpyxl.utils.cell import range_boundaries

    AMA_TEXT = "AMA BẮC NINH JSC"
    SEC_TEXT = "SAMSUNG ELECTRONICS VIETNAM CO., LTD."

    target_row: int | None = None
    for r in range(1, min(ws.max_row + 1, 200)):
        a_val = str(ws.cell(row=r, column=1).value or "").upper()
        f_val = str(ws.cell(row=r, column=6).value or "").upper()
        if "AMA" in a_val and (
            "SAMSUNG" in f_val or "SEC" in f_val or "ONIC" in f_val or "ELECTR" in f_val
        ):
            target_row = r
            break

    if target_row is None:
        # Fallback: append at the bottom so the document never ships unsigned.
        target_row = ws.max_row + 2

    # Drop any merge ranges that overlap the target row in cols A:I —
    # we'll re-create the two we want immediately after.
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        if min_row <= target_row <= max_row and min_col <= 9:
            to_unmerge.append(str(mr))
    for mr in to_unmerge:
        try:
            ws.unmerge_cells(mr)
        except Exception:
            pass

    # Write canonical text + center align
    from openpyxl.styles import Alignment as _Al, Font as _F
    sig_font = _F(name="Arial", size=11, bold=True)
    sig_align = _Al(horizontal="center", vertical="center", wrap_text=False)

    a_cell = ws.cell(row=target_row, column=1)
    a_cell.value = AMA_TEXT
    a_cell.font = sig_font
    a_cell.alignment = sig_align

    f_cell = ws.cell(row=target_row, column=6)
    f_cell.value = SEC_TEXT
    f_cell.font = sig_font
    f_cell.alignment = sig_align

    ws.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=5)
    ws.merge_cells(start_row=target_row, start_column=6, end_row=target_row, end_column=10)
    # Make signature row tall enough for the seal/signature underneath
    ws.row_dimensions[target_row].height = max(ws.row_dimensions[target_row].height or 0, 28)

    # Re-anchor the embedded stamp/signature image (CÔNG TY CỔ PHẦN AMA
    # BẮC NINH seal) so it sits 1 row below the AMA text and is centered
    # under the A:E merge — Thang 2026-05-10: "ảnh của phần kí tên căn
    # giữa thẳng với ama bắc ninh và lùi xuống 1 ô".
    _relocate_stamp_image(ws, target_row + 2)


def _relocate_stamp_image(ws, target_first_row: int) -> None:
    """Shift the stamp/signature image to the right by ~2cm.

    Per Thang 2026-05-11: PRESERVE the stamp's original size + aspect ratio
    (previous version scaled it to 85% which warped the seal). Only adjust
    its horizontal position by adding ~2cm (≈75 EMU pixels = 715000 EMU)
    to colOff on its existing anchor. Vertical position untouched.

    Heuristic: largest image whose original anchor row > 20 = stamp.
    The template's small logo (anchored near row 2) is left untouched.
    """
    from openpyxl.utils.units import pixels_to_EMU

    stamp = None
    best_score = -1
    for im in list(ws._images):
        try:
            anc = im.anchor
            from_row = getattr(getattr(anc, "_from", None), "row", -1) or -1
            w = int(im.width or 0)
            h = int(im.height or 0)
            score = (w * h) if from_row > 20 else 0
            if score > best_score:
                best_score = score
                stamp = im
        except Exception:
            continue
    if not stamp:
        return

    # Shift right by ~2cm = ~75 px
    shift_emu = pixels_to_EMU(75)

    try:
        anc = stamp.anchor
        _from = getattr(anc, "_from", None)
        if _from is not None:
            cur_off = int(_from.colOff or 0)
            _from.colOff = cur_off + shift_emu
        # If anchor is TwoCellAnchor, also shift `to` so the image doesn't stretch
        to = getattr(anc, "to", None)
        if to is not None:
            cur_to_off = int(to.colOff or 0)
            to.colOff = cur_to_off + shift_emu
        logger.info(
            "stamp shifted right by ~2cm (75px / %d EMU), preserved original size",
            shift_emu,
        )
    except Exception as exc:
        logger.warning("stamp shift failed: %s", exc)


def _enrich_maker(product: dict) -> str:
    """Best-effort 'maker' lookup with sensible fallbacks.
    Order: explicit 'maker' → 'brand' → 'thuong_hieu' → '_detail.items[0].maker'
    → empty string. Per Thang 2026-05-10: Quotation must show brand."""
    for key in ("maker", "brand", "thuong_hieu", "manufacturer"):
        v = product.get(key)
        if v and str(v).strip():
            return str(v).strip()
    detail = product.get("_detail") or {}
    items = detail.get("items") or []
    for it in items:
        m = it.get("maker") or it.get("brand")
        if m and str(m).strip():
            return str(m).strip()
    return ""


# ─── Fill CAM KET Template ───────────────────────────────────

def fill_cam_ket(
    template_path: str | None,
    products: list[dict],
    images_map: dict[str, bytes],
    output_path: str,
) -> bool:
    """Fill the CAM KET (commitment) Excel template.

    If template_path is None or not found, creates from scratch.
    Template layout:
      Row 16-18: Product 1 (3-row block)
      Row 19-21: Product 2 (3-row block)
      Row 22+:   Additional products (inserted rows)
      Column C: Index, D: BQMS, F: Spec, J: Maker, L: Unit Price, N: Image
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    if template_path and Path(template_path).exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        SP_START = 16
        ROWS_PER_SP = 3
        COL_C, COL_D, COL_F, COL_J, COL_L, COL_N = 3, 4, 6, 10, 12, 14

        now = datetime.now()
        # Thang 2026-05-21: defensive template fixups — refresh ALL date cells
        # to today (V2/V3 was inheriting V1's date because template had multiple
        # hardcoded date cells that single per-cell writes didn't touch).
        n_dates = _refresh_dates_in_sheet(ws, now)
        n_dash = _fix_dash_bullets_in_sheet(ws)
        if n_dates or n_dash:
            logger.info("CAM KET fixups: %d date cells refreshed, %d dash bullets fixed",
                        n_dates, n_dash)
        extra_rows = 0

        for i, product in enumerate(products):
            if i < 2:
                start_row = SP_START + i * ROWS_PER_SP
            else:
                insert_after = SP_START + 2 * ROWS_PER_SP + extra_rows - 1
                ws.insert_rows(insert_after + 1, ROWS_PER_SP)
                _copy_row_style(ws, SP_START + ROWS_PER_SP, insert_after + 1)
                start_row = insert_after + 1
                extra_rows += ROWS_PER_SP

            _safe_set_cell(ws, start_row, COL_C, i + 1)
            _safe_set_cell(ws, start_row, COL_D, product.get("bqms", ""))
            _safe_set_cell(ws, start_row, COL_F, product.get("spec", ""))
            _safe_set_cell(ws, start_row, COL_J, _enrich_maker(product))
            # Center the Spec cell (col F is the merged F:I "Spec" col).
            # Per Thang 2026-05-10: tên item căn giữa ô.
            try:
                from openpyxl.styles import Alignment as _Al
                ws.cell(row=start_row, column=COL_F).alignment = _Al(
                    wrap_text=True, vertical="center", horizontal="center",
                )
                ws.cell(row=start_row, column=COL_J).alignment = _Al(
                    wrap_text=True, vertical="center", horizontal="center",
                )
                ws.cell(row=start_row, column=COL_D).alignment = _Al(
                    vertical="center", horizontal="center",
                )
            except Exception:
                pass
            price = product.get("suggested_price") or product.get("unit_price")
            _safe_set_cell(ws, start_row, COL_L, price if price else "")
            # Image — strict per-item via images_map keyed by bqms code.
            # Auto-fit the merged N{r}:O{r+2} block (2 cols × 3 rows) so
            # the image co dãn theo cell. Per Thang 2026-05-10.
            bqms_code = (product.get("bqms") or "").strip()
            if bqms_code and bqms_code in images_map:
                ok = _attach_image_fit(
                    ws, images_map[bqms_code],
                    anchor_row=start_row, anchor_col=COL_N,
                    span_cols=2, span_rows=3,
                )
                if ok:
                    logger.info("CAM KET: image fit-attached for %s at row %d",
                                bqms_code, start_row)
            else:
                # Thang 2026-06-03: WARNING (was info) so missing-image
                # gaps surface in standard log filters next to the
                # QUOTATION sibling line — single noisy line beats a
                # silent blank in the PDF.
                logger.warning(
                    "CAM KET: NO image for bqms=%r at row %d. "
                    "images_map has %d keys (sample: %s). Cell will be blank.",
                    bqms_code, start_row, len(images_map),
                    list(images_map.keys())[:5],
                )
            # NOTE: rows 2-3 of each 3-row block may carry stale sample text
            # ("FINGER RF" etc.) from the template. Fix at template level
            # rather than guessing here. See plans/bqms-bugs.md.

        # FIX (Thang 2026-05-15): template có 2 block pre-existing (rows 16-18,
        # 19-21). Khi user báo giá < 2 sản phẩm, Block 2 vẫn hiện → dư dòng trống.
        # openpyxl `delete_rows()` không shift merged cells (B28:P29 stale)
        # gây mất text "Ngày…" + "Đại diện" — workaround: HIDE rows + clear content.
        # Hidden rows không hiện trong Excel UI lẫn PDF render qua Gotenberg.
        n_products = len(products)
        for unused_block_idx in range(n_products, 2):
            block_start = SP_START + unused_block_idx * ROWS_PER_SP
            for r in range(block_start, block_start + ROWS_PER_SP):
                ws.row_dimensions[r].hidden = True
                # Clear stale template text ("FINGER RF" etc.) defensively
                for c in [COL_C, COL_D, COL_F, COL_J, COL_L]:
                    try:
                        _safe_set_cell(ws, r, c, None)
                    except Exception:
                        pass
            logger.info("CAM KET: hidden unused Block %d (rows %d-%d)",
                        unused_block_idx + 1, block_start, block_start + ROWS_PER_SP - 1)

        # date_row stays at 31 — template position untouched
        date_row = 31 + extra_rows
        _safe_set_cell(ws, date_row, COL_L, f"Ngày {now.day} Tháng {now.month} năm {now.year}")

        # Per Thang 2026-05-21: "khoảng cách giữa các dòng chưa đều nhau"
        # — apply uniform heights so the page reads cleanly.
        # Header rows 1-13: uniform 24pt (was 22, mixed). Bullet text inside
        # header rows (multi-line "Đã gửi báo giá..." etc.) needs more space.
        for r in range(1, 14):
            cur = ws.row_dimensions[r].height
            if cur is None or cur < 24 or cur > 40:
                ws.row_dimensions[r].height = 24
            # If cell carries multi-line content, bump it more
            cell_val = ws.cell(row=r, column=1).value or ws.cell(row=r, column=2).value
            if isinstance(cell_val, str) and "\n" in cell_val:
                n_lines = cell_val.count("\n") + 1
                ws.row_dimensions[r].height = max(24, n_lines * 18)
            # Apply wrap_text + top alignment on bullet rows so dash content
            # reflows cleanly instead of overflowing right.
            try:
                from openpyxl.styles import Alignment as _Al
                c1 = ws.cell(row=r, column=1)
                if isinstance(c1.value, str) and ("-" in c1.value or "•" in c1.value):
                    c1.alignment = _Al(wrap_text=True, vertical="top", horizontal="left")
            except Exception:
                pass

        # Item blocks (rows 16-18, 19-21, then dynamic): each row in a block
        # carries spec text + image — bump to 36pt for visible image headroom.
        last_block_row = SP_START + len(products) * ROWS_PER_SP - 1 + extra_rows
        for r in range(SP_START, last_block_row + 1):
            cur = ws.row_dimensions[r].height
            if cur is None or cur < 32:
                ws.row_dimensions[r].height = 36
        # Date + signature footer rows
        for r in range(date_row, date_row + 5):
            cur = ws.row_dimensions[r].height
            if cur is None or cur < 24:
                ws.row_dimensions[r].height = 24
    else:
        # Create from scratch
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CAM KET"

        now = datetime.now()
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_font = Font(name='Arial', size=11, bold=True)
        data_font = Font(name='Arial', size=10)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(name='Arial', size=10, bold=True, color='FFFFFF')

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = 'CAM KET BAO GIA - AMA BAC NINH JSC'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:H2')
        ws['A2'] = f'Ngay: {now.strftime("%d/%m/%Y")}'
        ws['A2'].font = data_font
        ws['A2'].alignment = Alignment(horizontal='center')

        # RFQ info
        rfq_no = products[0].get("don_hang", "N/A") if products else "N/A"
        ws['A4'] = 'RFQ No:'
        ws['A4'].font = Font(name='Arial', size=10, bold=True)
        ws['B4'] = rfq_no
        ws['B4'].font = data_font

        # Headers row 6
        headers = ['#', 'BQMS Code', 'Specification', 'Maker', 'Qty', 'Unit', 'Unit Price (VND)', 'Image']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12

        # Data rows
        for i, product in enumerate(products):
            row = 7 + i
            price = product.get("suggested_price") or product.get("unit_price")
            data = [
                i + 1,
                product.get("bqms", ""),
                product.get("spec", ""),
                product.get("maker", ""),
                product.get("so_luong", 0),
                product.get("don_vi", "EA"),
                price if price else "",
                "",
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in (1, 5, 6):
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx == 7:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0'

            # Image
            bqms_code = product.get("bqms", "")
            if bqms_code in images_map:
                xl_img = _pil_image_to_xl(images_map[bqms_code])
                if xl_img:
                    ws.add_image(xl_img, f"H{row}")

        # Footer
        footer_row = 7 + len(products) + 2
        ws.merge_cells(f'A{footer_row}:H{footer_row}')
        ws[f'A{footer_row}'] = f'Ngày {now.day} Tháng {now.month} năm {now.year}'
        ws[f'A{footer_row}'].font = data_font
        ws[f'A{footer_row}'].alignment = Alignment(horizontal='right')

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    # Thang 2026-05-21: CRITICAL — openpyxl writes absolute drawing rels
    # (Target="/xl/media/image1.png") which LibreOffice/x2t silently reject.
    # Symptom: OnlyOffice editor opens, immediately throws error code=-4
    # "Tải về không thành công" because the xlsx → Editor.bin conversion
    # fails. PDF rendering via Gotenberg also drops images. Fix rewrites
    # paths to relative ("../media/image1.png") in the saved zip.
    # (Function existed at line 554 but was never called — dead code.)
    fix_drawing_rels_in_xlsx(output_path)
    logger.info("CAM KET filled: %d products -> %s", len(products), output_path)
    return True


# ─── Fill Quotation Template ─────────────────────────────────

def fill_quotation(
    template_path: str | None,
    products: list[dict],
    images_map: dict[str, bytes],
    rfq_no: str,
    output_path: str,
    round_n: int = 1,
    quote_date=None,
) -> bool:
    """Fill the Commercial Quotation Excel template.

    If template_path is None or not found, creates from scratch.
    Columns: A=idx, B=RFQ, C=BQMS+spec, D=maker, E=image,
             F=unit, G=qty, H=price, I=amount, J=notes

    Args:
        quote_date: optional date/datetime to stamp on the quotation header
                    (cell C4 + defensive date sweep). Defaults to today.
                    Per Thang 2026-06-13: passing a date pins it; default
                    today still wins for fresh round-1 quotes.
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Thang 2026-06-13: caller may pass an explicit quote_date (date,
    # datetime, or yyyy-mm-dd string). Default = today in VN tz (forced
    # via _vn_now) so server tz can't drift the stamped date. Always
    # converted to a datetime so `_refresh_dates_in_sheet` + strftime work
    # uniformly.
    if quote_date is None:
        now = _vn_now()
    elif isinstance(quote_date, datetime):
        now = quote_date
    elif isinstance(quote_date, _date):
        # `date` → datetime at midnight (time component unused for dd/mm/yyyy)
        now = datetime(quote_date.year, quote_date.month, quote_date.day)
    elif isinstance(quote_date, str) and quote_date.strip():
        try:
            now = datetime.fromisoformat(quote_date.strip())
        except ValueError:
            now = _vn_now()
    else:
        now = _vn_now()

    if template_path and Path(template_path).exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        for sheet_name in wb.sheetnames:
            if any(kw in sheet_name.lower() for kw in ("bqms", "code", "quotation")):
                ws = wb[sheet_name]
                break

        # Apply post-load template fixups: hide internal cols L..S,
        # bump terms-row heights so 2-line VN/EN text wraps properly.
        # Per Thang 2026-05-10.
        _fixup_quotation_template(ws)

        # Thang 2026-05-21: defensive — refresh ALL date cells in the
        # Quotation sheet so V2/V3 don't inherit V1's date from template
        # cells that aren't on the explicit per-cell write path.
        n_dates = _refresh_dates_in_sheet(ws, now)
        # Thang 2026-05-22: same dash-bullet fix as CAM_KET — template's
        # "-Cam kết..." (missing space after dash) renders ugly in PDF.
        # Normalize to "- Cam kết..." so QUOTATION matches the template
        # form 100%.
        n_dash = _fix_dash_bullets_in_sheet(ws)
        if n_dates or n_dash:
            logger.info("QUOTATION fixups: %d date cells refreshed, "
                        "%d dash bullets fixed", n_dates, n_dash)

        COL_A, COL_B, COL_C, COL_D = 1, 2, 3, 4
        COL_E, COL_F, COL_G, COL_H, COL_I, COL_J = 5, 6, 7, 8, 9, 10

        _safe_set_cell(ws, 4, COL_C, now.strftime("%d/%m/%Y"))
        _safe_set_cell(ws, 6, COL_H, rfq_no)
        # Quotation No format (Thang 2026-05-22, separator = underscore):
        #   1 mã   → "{RFQ_NO}_{BQMS_NO}_AMABACNINH_L{round}"
        #   2+ mã  → "{RFQ_NO}_AMABACNINH_L{round}"
        # Underscore between SEGMENTS so the dash inside BQMS code
        # (e.g. Z0000002-389973) stays intact and visually distinct.
        distinct_codes = []
        for p in products:
            code = (p.get("bqms") or "").strip()
            if code and code not in distinct_codes:
                distinct_codes.append(code)
        if len(distinct_codes) == 1:
            quotation_no = f"{rfq_no}_{distinct_codes[0]}_AMABACNINH_L{round_n}"
        else:
            quotation_no = f"{rfq_no}_AMABACNINH_L{round_n}"
        _safe_set_cell(ws, 7, COL_C, quotation_no)
        # Product description row REMOVED (Thang 2026-06-15): user requested it
        # be hidden — the items table below already lists every product, so the
        # aggregated header line just duplicates info and crowds the page.
        # Clear A14 so any placeholder text "Product description/ Tên hàng: "
        # left in the template doesn't render. Keep row 14 with modest height
        # so the spacing above the items table doesn't collapse.
        try:
            from openpyxl.utils.cell import range_boundaries as _rb
            for mr in list(ws.merged_cells.ranges):
                a, b, c, d = _rb(str(mr))
                if b <= 14 <= d and a >= 1 and c <= 10:
                    ws.unmerge_cells(str(mr))
        except Exception:
            pass
        _safe_set_cell(ws, 14, COL_A, "")
        ws.row_dimensions[14].height = 8

        # Widen B (BQMS code col) so item codes like Z0000002-544469 fit.
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 18)

        DATA_START = 17
        TEMPLATE_ROW = 17
        last_data_row = DATA_START

        # Bug fix 2026-05-07: clear any pre-existing "Grand Total" rows in the
        # template body so we don't end up with two totals (one stale = sample
        # data 360,000,000 from the template, one fresh from the SUM formula).
        # Scan rows DATA_START+1 .. DATA_START+30 for a cell whose value
        # contains "Grand Total" (case-insensitive) and wipe its row.
        _stale_total_rows: list[int] = []
        for r in range(DATA_START + 1, DATA_START + 30):
            v = ws.cell(row=r, column=COL_A).value
            if isinstance(v, str) and "grand total" in v.lower():
                _stale_total_rows.append(r)
        for r in sorted(_stale_total_rows, reverse=True):
            # Clear values + remove the row entirely
            ws.delete_rows(r, 1)

        # Drop any leftover wide merges inside the data band — fixes the
        # "MICRO MOTION TECHNOLOGY centered across row" bug where the
        # template's Grand Total merge A18:E18 survives delete_rows() and
        # bleeds into the next item's row.
        _cleanup_data_area_merges(ws, DATA_START)

        from openpyxl.styles import Alignment as _Alignment

        for i, product in enumerate(products):
            if i == 0:
                row = DATA_START
            else:
                row = last_data_row + 1
                ws.insert_rows(row)
                _copy_row_style(ws, TEMPLATE_ROW, row)
                _copy_merged_ranges_for_row(ws, TEMPLATE_ROW, row)

            # NO sequential — overwrite any stale "1" the template carried.
            _safe_set_cell(ws, row, COL_A, i + 1)
            # B col = BQMS item code (e.g. Z0000002-544469) per template
            # header "BQMS code". Per Thang 2026-05-10: previously this
            # mistakenly held don_hang (RFQ no like QT26061473).
            _safe_set_cell(ws, row, COL_B, product.get("bqms", ""))
            # C col = Items / Hạng mục — description only (no bqms duplicate).
            spec_text = product.get("spec", "") or product.get("short_name", "")
            _safe_set_cell(ws, row, COL_C, spec_text)
            # Force wrap_text + CENTER alignment (both axes) on Items col.
            # Per Thang 2026-05-10: tên item căn giữa ô.
            try:
                ws.cell(row=row, column=COL_C).alignment = _Alignment(
                    wrap_text=True, vertical="center", horizontal="center",
                )
            except Exception:
                pass
            # Brand — enriched fallback (maker → brand → _detail.items[0].maker).
            _safe_set_cell(ws, row, COL_D, _enrich_maker(product))
            _safe_set_cell(ws, row, COL_F, product.get("don_vi", "EA"))
            _safe_set_cell(ws, row, COL_G, product.get("so_luong", 0))
            price = product.get("suggested_price") or product.get("unit_price")
            _safe_set_cell(ws, row, COL_H, price if price else "")
            # Currency = VND đồng — ensure unit price + amount use thousands sep.
            ws.cell(row=row, column=COL_H).number_format = '#,##0'
            ws.cell(row=row, column=COL_I).value = f"=G{row}*H{row}"
            ws.cell(row=row, column=COL_I).number_format = '#,##0'
            # Per Thang 2026-05-10: Ghi chú để trống. Don't write
            # deadline/timestamp text into col J even if input has it.
            _safe_set_cell(ws, row, COL_J, "")
            # Auto row height FIRST — image fit calc reads row_dimensions.
            # Accommodate image (≥85px) + wrapped spec text.
            # Col C ≈ 39 chars wide; ~14px per wrapped line.
            spec_len = len(spec_text)
            estimated_lines = 1 + (spec_len // 38) + spec_text.count("\n")
            ws.row_dimensions[row].height = max(85, min(180, estimated_lines * 16 + 30))

            # Image (strict mapping) — auto-fit cell E{row}.
            # Thang 2026-05-21: row was already bumped to ≥85pt for image
            # headroom. Pass span_rows=1 (single-row cell, no merge), but
            # cell pixel size now includes the bumped row height so image
            # fills properly with new fill_ratio=0.92.
            bqms_code = (product.get("bqms") or "").strip()
            if bqms_code and bqms_code in images_map:
                ok = _attach_image_fit(
                    ws, images_map[bqms_code],
                    anchor_row=row, anchor_col=COL_E,
                    span_cols=1, span_rows=1,
                )
                if ok:
                    logger.info("QUOTATION: image fit-attached for %s at row %d", bqms_code, row)
            else:
                # Thang 2026-06-03: was info-level, easy to miss. Upgrade
                # to WARNING with all the context an operator needs to
                # fix it (rename a scraper file, picker-pin, or upload an
                # override). Matches CAM_KET sibling log.
                logger.warning(
                    "QUOTATION: NO image for bqms=%r at row %d. "
                    "images_map has %d keys (sample: %s). Cell will be blank.",
                    bqms_code, row, len(images_map), list(images_map.keys())[:5],
                )

            last_data_row = row

        total_row = last_data_row + 1
        ws.insert_rows(total_row)
        _copy_row_style(ws, TEMPLATE_ROW, total_row)
        _safe_set_cell(ws, total_row, COL_A, "Grand Total (VND)")
        # Grand Total label needs to span A:F (5 cols) so it doesn't get
        # squished into narrow col A. Image col E is included but stays
        # empty — fine since merge centers the label.
        try:
            ws.merge_cells(
                start_row=total_row, start_column=COL_A,
                end_row=total_row, end_column=COL_F,
            )
        except Exception:
            pass
        from openpyxl.styles import Alignment as _Al, Font as _F
        ws.cell(row=total_row, column=COL_A).alignment = _Al(
            horizontal="right", vertical="center",
        )
        ws.cell(row=total_row, column=COL_A).font = _F(name="Arial", size=11, bold=True)
        ws.cell(row=total_row, column=COL_G).value = f"=SUM(G{DATA_START}:G{last_data_row})"
        ws.cell(row=total_row, column=COL_G).number_format = '#,##0'
        ws.cell(row=total_row, column=COL_G).font = _F(name="Arial", size=11, bold=True)
        ws.cell(row=total_row, column=COL_I).value = f"=SUM(I{DATA_START}:I{last_data_row})"
        ws.cell(row=total_row, column=COL_I).number_format = '#,##0'
        ws.cell(row=total_row, column=COL_I).font = _F(name="Arial", size=11, bold=True)
        # Grand Total row needs height ≥ 28 so the bold label is readable.
        # Per Thang 2026-05-10: dòng dãn cách chưa đều — fix Grand Total
        # squished to 18pt + Validity row dropping to 15pt default.
        ws.row_dimensions[total_row].height = max(
            ws.row_dimensions[total_row].height or 0, 32,
        )

        # Re-run terms-row + signature fixups AFTER all inserts/deletes
        # finish (rows have shifted). Without this, my earlier fixup at
        # template-load time set the Validity row's height in template
        # coords (row 25), but after delete-stale + insert-items the
        # actual Validity ended up at a different row that we never
        # bumped → defaulted to ~15pt and looked squished in PDF.
        _fixup_quotation_template(ws)

        # Signature row — defensive fixup AFTER all item inserts have shifted
        # the original row 27 to its new position AND re-merge A:E + F:I
        # so long company names render across the merged span (not wrapped
        # vertically inside the narrow col A / col F).
        _fixup_signatures(ws)
    else:
        # Create from scratch — professional Commercial Quotation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QUOTATION"

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=10)
        title_font = Font(name='Arial', size=14, bold=True, color='2F5496')
        total_font = Font(name='Arial', size=11, bold=True)
        total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

        # Company header
        ws.merge_cells('A1:J1')
        ws['A1'] = 'AMA BAC NINH JSC - COMMERCIAL QUOTATION'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Metadata
        ws['A3'] = 'Date:'
        ws['A3'].font = Font(name='Arial', size=10, bold=True)
        ws['B3'] = now.strftime('%d/%m/%Y')
        ws['B3'].font = data_font

        ws['A4'] = 'Quotation No:'
        ws['A4'].font = Font(name='Arial', size=10, bold=True)
        _scratch_codes = []
        for p in products:
            c = (p.get("bqms") or "").strip()
            if c and c not in _scratch_codes:
                _scratch_codes.append(c)
        if len(_scratch_codes) == 1:
            ws['B4'] = f"{rfq_no}_{_scratch_codes[0]}_AMABACNINH_L{round_n}"
        else:
            ws['B4'] = f"{rfq_no}_AMABACNINH_L{round_n}"
        ws['B4'].font = data_font

        ws['F3'] = 'RFQ No:'
        ws['F3'].font = Font(name='Arial', size=10, bold=True)
        ws['G3'] = rfq_no
        ws['G3'].font = Font(name='Arial', size=10, bold=True, color='2F5496')

        # Column headers row 6
        headers = ['#', 'RFQ No', 'BQMS Code / Spec', 'Maker', 'Image', 'Unit', 'Qty', 'Unit Price (VND)', 'Amount (VND)', 'Notes']
        col_widths = [5, 14, 40, 15, 10, 8, 10, 18, 18, 15]

        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=6, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.row_dimensions[6].height = 30

        # Data rows
        DATA_START = 7
        last_data_row = DATA_START

        for i, product in enumerate(products):
            row = DATA_START + i
            price = product.get("suggested_price") or product.get("unit_price")
            qty = product.get("so_luong", 0)

            data = [
                i + 1,
                product.get("don_hang", ""),
                f"{product.get('bqms', '')}\n{product.get('spec', '')}",
                product.get("maker", ""),
                "",  # image placeholder
                product.get("don_vi", "EA"),
                qty,
                price if price else "",
                None,  # formula
                product.get("ghi_chu", ""),
            ]

            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx == 3:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                elif col_idx in (7, 8, 9):
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0'

            # Amount formula
            ws.cell(row=row, column=9).value = f"=G{row}*H{row}"
            ws.cell(row=row, column=9).number_format = '#,##0'

            # Image
            bqms_code = product.get("bqms", "")
            if bqms_code in images_map:
                xl_img = _pil_image_to_xl(images_map[bqms_code])
                if xl_img:
                    ws.add_image(xl_img, f"E{row}")

            last_data_row = row

        # Grand Total row
        total_row = last_data_row + 1
        ws.merge_cells(f'A{total_row}:F{total_row}')
        ws[f'A{total_row}'] = 'GRAND TOTAL (VND)'
        ws[f'A{total_row}'].font = total_font
        ws[f'A{total_row}'].fill = total_fill
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
        ws[f'A{total_row}'].border = thin_border

        ws.cell(row=total_row, column=7).value = f"=SUM(G{DATA_START}:G{last_data_row})"
        ws.cell(row=total_row, column=7).font = total_font
        ws.cell(row=total_row, column=7).fill = total_fill
        ws.cell(row=total_row, column=7).border = thin_border
        ws.cell(row=total_row, column=7).number_format = '#,##0'

        ws.cell(row=total_row, column=9).value = f"=SUM(I{DATA_START}:I{last_data_row})"
        ws.cell(row=total_row, column=9).font = total_font
        ws.cell(row=total_row, column=9).fill = total_fill
        ws.cell(row=total_row, column=9).border = thin_border
        ws.cell(row=total_row, column=9).number_format = '#,##0'

        for col in range(1, 11):
            if col not in (1, 7, 9):
                ws.cell(row=total_row, column=col).fill = total_fill
                ws.cell(row=total_row, column=col).border = thin_border

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    # Thang 2026-05-21: same drawing-rels fix as Cam kết — required for
    # OnlyOffice converter (x2t) to not crash on absolute image paths.
    fix_drawing_rels_in_xlsx(output_path)
    logger.info("Quotation filled: %d products -> %s", len(products), output_path)
    return True


# ─── Run Auto-Fill Job (orchestrator) ────────────────────────

async def run_autofill_job(
    conn,
    quotation_id: int,
    items: list[dict],
    images: list[tuple[str, bytes]] | None = None,
    cam_ket_template: str | None = None,
    commercial_template: str | None = None,
    flow_type: str = "tm",
    round_n: int = 1,
    quote_date: _date | datetime | str | None = None,
) -> dict[str, Any]:
    """Execute the full auto-fill pipeline for a quotation.

    Args:
        flow_type: "tm" (Thương Mại) or "gc" (Gia Công).
            TM: CAM KET (all items) + Commercial Quotation (TM items only).
            GC: CAM KET (all items) + Commercial Quotation (GC items with cost breakdown).
        quote_date: optional date (date | datetime | yyyy-mm-dd string) to
            stamp on output files. None → today in VN tz (Asia/Ho_Chi_Minh)
            so server tz can't drift the stamped date. Threaded into
            fill_quotation() + GC builder so all paths share the same date.
            Per Thang 2026-06-13 (Bug fix T3).

    Steps:
      1. Apply user-edited prices (unit_price field overrides suggested_price)
      2. Lookup prices from DB for items without user price
      3. Match images to orders
      4. Fill CAM KET template (all items)
      5. Fill Commercial Quotation template (filtered by flow_type)
      6. Convert to PDF via Gotenberg
      7. Update quotation record in DB
    """
    from app.services.gotenberg_service import convert_xlsx_to_pdf

    result: dict[str, Any] = {
        "success": False,
        "quotation_id": quotation_id,
        "files": [],
        "errors": [],
    }

    try:
        # Update status
        await conn.execute(
            "UPDATE quotations SET status = 'processing', updated_at = NOW() WHERE id = $1",
            quotation_id,
        )

        # 1. Apply user-edited prices — if item has 'unit_price', use it as suggested_price
        for item in items:
            user_price = item.get("unit_price")
            if user_price is not None and user_price != "" and user_price != 0:
                try:
                    item["suggested_price"] = float(user_price)
                except (ValueError, TypeError):
                    pass

        # 2. Lookup prices from DB for items without a user price
        items_needing_lookup = [i for i in items if not i.get("suggested_price")]
        if items_needing_lookup:
            items_needing_lookup = await lookup_prices(conn, items_needing_lookup)
            # Merge back
            lookup_map = {i["bqms"]: i for i in items_needing_lookup}
            for item in items:
                if not item.get("suggested_price") and item.get("bqms") in lookup_map:
                    looked = lookup_map[item["bqms"]]
                    item["suggested_price"] = looked.get("suggested_price")
                    item["price_history"] = looked.get("price_history", [])

        # Get RFQ no — used to locate the parent QT folder + name the L1 subfolder.
        rfq_no = items[0].get("don_hang", "UNKNOWN") if items else "UNKNOWN"
        # Thang 2026-06-13 (Bug fix T3): use caller-supplied quote_date if
        # provided; otherwise FORCE VN tz so the year/month folder naming
        # below matches Hanoi office calendar. Resolved into `quote_dt` so
        # we pass the exact same value down to fill_quotation().
        if quote_date is None:
            quote_dt: datetime = _vn_now()
        elif isinstance(quote_date, datetime):
            quote_dt = quote_date
        elif isinstance(quote_date, _date):
            quote_dt = datetime(quote_date.year, quote_date.month, quote_date.day)
        elif isinstance(quote_date, str) and quote_date.strip():
            try:
                quote_dt = datetime.fromisoformat(quote_date.strip())
            except ValueError:
                quote_dt = _vn_now()
        else:
            quote_dt = _vn_now()
        now = quote_dt

        # Per Thang 2026-05-10: output goes into the parent QT folder's L1
        # subfolder so cam ket + quotation files live together with the
        # downloaded raw/ + images/ from scrape time.
        # Path: <RFQ_ROOT>/RFQ <year>/THANG <month>/<pretty-QT-folder>/<QT>_AMA BAC NINH_L1/
        # If the parent QT folder doesn't exist yet (e.g. quote was triggered
        # without a prior scrape), fall back to the legacy /data/files/quotations
        # location so we never block the user.
        output_dir: Path
        try:
            from app.etl.bqms_bidding_scraper import (
                find_existing_rfq_folder,
                quote_round_subfolder,
            )
            parent = find_existing_rfq_folder(rfq_no, now)
            if parent is not None:
                # Thang 2026-05-21: pass actual round_n so files land directly
                # in L{round_n}/. Previously hardcoded round_n=1 and caller
                # post-moved files L1→L{n} — that move overwrote existing L{n}.
                output_dir = quote_round_subfolder(parent, rfq_no, round_n=round_n)
                logger.info(
                    "autofill: using L%d subfolder under parent QT folder: %s",
                    round_n, output_dir,
                )
            else:
                # Fallback: old layout under /data/files/quotations/
                year_str = f"RFQ {now.year}"
                month_str = f"THANG {now.month}"
                short_spec = ""
                if items:
                    spec = items[0].get("spec", "") or items[0].get("bqms", "")
                    short_spec = re.sub(r'[\\/:*?"<>|]', '', spec)[:40].strip()
                folder_name = f"{rfq_no} {short_spec} {now.day}-{now.month}".strip()
                output_dir = FILES_BASE / year_str / month_str / folder_name
                logger.info(
                    "autofill: parent QT folder not found, using fallback: %s",
                    output_dir,
                )
            output_dir.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            logger.warning("L1 subfolder logic failed (%s) — using fallback", exc)
            year_str = f"RFQ {now.year}"
            month_str = f"THANG {now.month}"
            folder_name = f"{rfq_no} {now.day}-{now.month}".strip()
            output_dir = FILES_BASE / year_str / month_str / folder_name
            output_dir.mkdir(parents=True, exist_ok=True)

        # 3. Match images: combine user-uploaded with auto-discovered from
        # /data/onedrive-staging/Puplic/BQMS/RFQ/.../{rfq_no}*/
        # Phase E1 (Thang 2026-05-12): /data/quote-overrides/{rfq}/{code}__product_photo.{ext}
        # has HIGHEST priority — user-uploaded override applies to BOTH TM + GC.
        all_image_files: list[tuple[str, bytes]] = list(images or [])

        # Override pass — load files from /data/quote-overrides/{rfq_no}/
        override_map: dict[str, bytes] = {}  # bqms_code → bytes
        try:
            ovr_dir = Path(f"/data/quote-overrides/{rfq_no}")
            if ovr_dir.exists():
                for ovr_path in ovr_dir.glob("*"):
                    if not ovr_path.is_file() or ovr_path.suffix.lower() not in _IMAGE_EXTS:
                        continue
                    # Filename: {bqms_code}__product_photo.{ext}
                    stem = ovr_path.stem
                    if "__" in stem:
                        code = stem.split("__", 1)[0]
                        if code and code not in override_map:
                            try:
                                override_map[code] = ovr_path.read_bytes()
                                logger.info("autofill: image OVERRIDE for %s ← %s",
                                            code, ovr_path.name)
                            except OSError as exc:
                                logger.warning("autofill: override read failed %s: %s", ovr_path, exc)
        except Exception as exc:
            logger.warning("autofill: override scan failed: %s", exc)

        discovered = discover_rfq_images(rfq_no)
        seen_names = {n for n, _ in all_image_files}
        for fname, fbytes in discovered:
            if fname not in seen_names:
                all_image_files.append((fname, fbytes))
        images_map: dict[str, bytes] = match_images_to_orders(all_image_files, items)

        # Override has highest priority — overlay on top of matched images
        for code, ovr_bytes in override_map.items():
            images_map[code] = ovr_bytes

        # PRIORITY 0 (Thang 2026-05-22): user-pinned primary image from picker
        # modal. Stored in `bqms_code_primary_image.image_path` (DB) — set via
        # POST /bqms/code/{code}/primary-image after user crops+pins via UI.
        # Same priority order /bqms-images uses for the gallery → keeps báo giá
        # in sync with what user sees in the BQMS table image column.
        # Overlay LAST so this beats both filesystem matches + per-rfq overrides.
        pinned_map: dict[str, bytes] = {}
        try:
            codes = list({(it.get("bqms") or "").strip() for it in items if it.get("bqms")})
            if codes:
                rows = await conn.fetch(
                    "SELECT bqms_code, image_path FROM bqms_code_primary_image "
                    "WHERE bqms_code = ANY($1::text[])",
                    codes,
                )
                for r in rows:
                    p = Path(r["image_path"]) if r["image_path"] else None
                    if p and p.exists() and p.is_file():
                        try:
                            pinned_map[r["bqms_code"]] = p.read_bytes()
                        except OSError as exc:
                            logger.warning("pinned image read failed %s: %s", p, exc)
                    elif p:
                        logger.warning(
                            "pinned image disappeared from disk: %s (code=%s)",
                            p, r["bqms_code"],
                        )
        except Exception as exc:
            logger.warning("pinned image DB lookup failed (continuing): %s", exc)

        for code, pin_bytes in pinned_map.items():
            images_map[code] = pin_bytes
            logger.info("autofill: image PINNED (primary) for %s (%d bytes)",
                        code, len(pin_bytes))

        if images_map:
            logger.info('autofill: %d products matched to images '
                        '(%d overrides, %d pinned)',
                        len(images_map), len(override_map), len(pinned_map))

        # Defensive diagnostic (Thang 2026-06-03): when a priced item's
        # bqms_code falls through ALL fallback layers (filesystem substring
        # match → /data/quote-overrides/{rfq}/ → bqms_code_primary_image),
        # the cell is left blank in CAM_KET + QUOTATION and the user has
        # no obvious clue why. Surface it now so the next /admin pass +
        # log review immediately spots the gap.
        # Also flag any "_unmapped_*" files we got from the bidding scraper
        # — those indicate multi-code workbooks the scraper couldn't
        # attribute (e.g. RFQ_Tool_Box.xlsx for QT26071059), and the
        # candidate codes are exactly the ones that will end up blank.
        try:
            requested_codes = sorted({
                (it.get("bqms") or "").strip()
                for it in items if it.get("bqms")
            })
            missing_codes = [c for c in requested_codes if c and c not in images_map]
            if missing_codes:
                logger.warning(
                    "autofill: %d/%d items have NO image after all fallbacks "
                    "(filesystem + override + pinned). Codes without image: %s. "
                    "RFQ=%s. Action: rename matching file under "
                    "/data/onedrive-staging/.../%s/images/ to '{code}_1.png', "
                    "OR picker-pin via /bqms/code/{code}/primary-image.",
                    len(missing_codes), len(requested_codes),
                    missing_codes[:20], rfq_no, rfq_no,
                )
            unmapped_files = [
                n for n, _ in all_image_files if "_unmapped_" in n.lower()
            ]
            if unmapped_files:
                logger.warning(
                    "autofill: bidding scraper left %d UNMAPPED image file(s) "
                    "for RFQ=%s: %s. These cannot match any bqms_code via "
                    "substring. Candidate codes from this RFQ's items: %s. "
                    "If the count matches, rename them in order: "
                    "'_unmapped_N.ext' -> '{candidate_code}_1.ext'.",
                    len(unmapped_files), rfq_no, unmapped_files[:10],
                    missing_codes[:20],
                )
        except Exception as _diag_exc:  # never let diag crash the pipeline
            logger.debug("autofill: missing-image diagnostic failed: %s", _diag_exc)

        # Filter items by flow_type
        if flow_type == "gc":
            target_items = [i for i in items if i.get("loai_hang") == "GC"] or items
        else:
            target_items = [i for i in items if i.get("loai_hang") == "TM"] or items

        # Per Thang 2026-05-11: chỉ tạo file cho mã được điền giá.
        # Items có unit_price hoặc suggested_price > 0 mới include vào
        # Quotation/CAM_KET/GC. Items không điền giá → bỏ qua hoàn toàn.
        # Empty / 0 / "" / None all mean "user didn't fill price".
        def _has_price(it: dict) -> bool:
            for k in ("unit_price", "suggested_price"):
                v = it.get(k)
                if v in (None, "", 0, "0"):
                    continue
                try:
                    return float(v) > 0
                except (TypeError, ValueError):
                    pass
            return False

        priced_items = [it for it in target_items if _has_price(it)]
        if priced_items:
            logger.info(
                "Price filter: %d/%d items have a price → only those go into output",
                len(priced_items), len(target_items),
            )
            target_items = priced_items
        else:
            logger.warning(
                "Price filter: NO items have a price. Skipping output file generation."
            )
            target_items = []

        all_items = target_items  # CAM KET also follows the price filter

        files: list[dict] = []
        is_gc_flow = (flow_type == "gc")

        # ─── GC (Gia Công) round-1 dispatch ───────────────────────────
        # Per Thang 2026-05-10: hàng gia công dùng template QUOTATION_GC.xlsx
        # (1 sheet/item, preserve formulas, editable, ảnh per sheet).
        # When flow_type='gc', we skip the TM CAM_KET + QUOTATION branches
        # and run the GC builder instead. OneDrive sync + DB update still run.
        if is_gc_flow:
            try:
                from app.services.tools.gc_template_quotation import run_gc_quote_round1
                gc_result = await run_gc_quote_round1(
                    rfq_no=rfq_no,
                    items=target_items,
                    images_map=images_map,
                    output_dir=output_dir,
                    gc_template_path="/data/files/templates/QUOTATION_GC.xlsx",
                    quote_date=quote_dt,
                )
                files.extend(gc_result.get("files", []))
                if gc_result.get("errors"):
                    result["errors"].extend(gc_result["errors"])
                logger.info(
                    "GC round 1 done: %d sheets, %d images, %d formula rewrites",
                    gc_result.get("sheets_created", 0),
                    gc_result.get("images_attached", 0),
                    gc_result.get("formula_rewrites", 0),
                )
            except Exception as exc:
                logger.exception("GC round 1 failed")
                result["errors"].append(f"GC round 1: {exc}")

        # 4. Fill CAM KET — ONLY for items that have a maker (per Thang
        # 2026-05-10: nếu mã linh kiện đó có maker thì mới tạo file cam
        # kết, không có thì thôi). Skip the file entirely if no item has
        # a maker — the cam kết is a "genuine sales commitment" certifying
        # the manufacturer, so without a maker name there's nothing to
        # commit to.
        # GC flow skips both CAM_KET + commercial QUOTATION (uses GC template).
        cam_ket_items = [] if is_gc_flow else [
            it for it in all_items if _enrich_maker(it).strip()
        ]
        if cam_ket_items:
            ck_xlsx = str(output_dir / f"CAM_KET_{rfq_no}.xlsx")
            fill_cam_ket(cam_ket_template, cam_ket_items, images_map, ck_xlsx)
            files.append({"type": "cam_ket_xlsx", "path": ck_xlsx})
            logger.info(
                "CAM KET generated for %d/%d items with maker",
                len(cam_ket_items), len(all_items),
            )

            try:
                ck_pdf = str(output_dir / f"CAM_KET_{rfq_no}.pdf")
                await convert_xlsx_to_pdf(ck_xlsx, ck_pdf)
                files.append({"type": "cam_ket_pdf", "path": ck_pdf})
            except Exception as exc:
                result["errors"].append(f"CAM KET PDF conversion failed: {exc}")
        else:
            logger.info(
                "CAM KET skipped — no items have maker (%d items total)",
                len(all_items),
            )

        # 5. Fill Commercial Quotation (TM only — GC handled above with
        # its own multi-sheet template builder).
        # Per Thang 2026-05-11: skip if no items have a price (price filter
        # already applied above; target_items would be []).
        if not is_gc_flow and target_items:
            # Quotation XLSX filename = parent folder name + .xlsx (matches PDF
            # naming convention). User requirement 2026-06-04: copy exactly to
            # make it easy to identify the round (L1/L2/L3/L4). Happy path
            # produces e.g. `QT26071059_AMABACNINH_L1.xlsx/.pdf`; fallback
            # folder shape (no AMABACNINH suffix) is honored automatically
            # because we derive from output_dir.name rather than hardcoding.
            base_name = output_dir.name
            qt_xlsx = str(output_dir / f"{base_name}.xlsx")
            fill_quotation(
                commercial_template, target_items, images_map, rfq_no, qt_xlsx,
                round_n=round_n, quote_date=quote_dt,
            )
            files.append({"type": "quotation_xlsx", "path": qt_xlsx})

            try:
                qt_pdf = str(output_dir / f"{base_name}.pdf")
                await convert_xlsx_to_pdf(qt_xlsx, qt_pdf)
                files.append({"type": "quotation_pdf", "path": qt_pdf})
            except Exception as exc:
                result["errors"].append(f"Quotation PDF conversion failed: {exc}")
        elif not is_gc_flow and not target_items:
            logger.info(
                "Skipping QUOTATION generation — 0 items have a price filled."
            )
            result["errors"].append(
                "Không có item nào được điền giá → không tạo file báo giá."
            )

        # 6. Sync to OneDrive (best-effort; failures don't fail the whole job).
        # Per Thang 2026-05-07: store under /Bao_Gia_BQMS/RFQ {year}/THANG {month}/
        onedrive_folder_id: str | None = None
        onedrive_url: str | None = None
        onedrive_share_url: str | None = None
        onedrive_error: str | None = None
        try:
            from app.services.quotation_onedrive import sync_quotation_to_onedrive
            od_result = await sync_quotation_to_onedrive(
                rfq_no=rfq_no,
                local_files=files,
                year=now.year,
                month=now.month,
                create_share_links=True,
            )
            onedrive_folder_id = od_result.get("folder_id")
            onedrive_url = od_result.get("primary_url")
            onedrive_share_url = od_result.get("primary_share")
            if od_result.get("errors"):
                onedrive_error = "; ".join(od_result["errors"])[:500]
                result["errors"].extend([f"OneDrive: {e}" for e in od_result["errors"]])
            # Enrich files list with OneDrive URLs so the API response carries them
            url_by_path = {it["path"]: it for it in od_result.get("items", [])}
            for f in files:
                if f["path"] in url_by_path:
                    od_item = url_by_path[f["path"]]
                    f["onedrive_url"] = od_item.get("web_url")
                    f["onedrive_share_url"] = od_item.get("share_url")
                    f["onedrive_item_id"] = od_item.get("item_id")
        except Exception as exc:
            logger.warning("OneDrive sync skipped (will retry via /sync-onedrive): %s", exc)
            onedrive_error = str(exc)[:500]

        # 7. Update DB
        output_xlsx = next((f["path"] for f in files if "xlsx" in f["type"]), None)
        output_pdf = next((f["path"] for f in files if "pdf" in f["type"]), None)

        await conn.execute(
            """
            UPDATE quotations
            SET status              = 'completed',
                items               = $2::jsonb,
                output_xlsx         = $3,
                output_pdf          = $4,
                filled_items        = $5,
                total_items         = $6,
                onedrive_folder_id  = $7,
                onedrive_url        = $8,
                onedrive_share_url  = $9,
                onedrive_synced_at  = CASE WHEN $10::text IS NULL THEN NOW() ELSE NULL END,
                onedrive_sync_error = $10,
                updated_at          = NOW()
            WHERE id = $1
            """,
            quotation_id,
            __import__("json").dumps(items, default=str, ensure_ascii=False),
            output_xlsx,
            output_pdf,
            len([i for i in items if i.get("suggested_price")]),
            len(items),
            onedrive_folder_id,
            onedrive_url,
            onedrive_share_url,
            onedrive_error,
        )

        result["success"] = True
        result["files"] = files
        result["rfq_no"] = rfq_no
        result["total_items"] = len(items)
        result["filled_items"] = len([i for i in items if i.get("suggested_price")])
        result["onedrive_url"] = onedrive_url
        result["onedrive_share_url"] = onedrive_share_url

    except Exception as exc:
        logger.exception("Auto-fill job failed for quotation %d", quotation_id)
        result["errors"].append(str(exc))
        await conn.execute(
            "UPDATE quotations SET status = 'failed', error_message = $2, updated_at = NOW() WHERE id = $1",
            quotation_id, str(exc),
        )

    return result
