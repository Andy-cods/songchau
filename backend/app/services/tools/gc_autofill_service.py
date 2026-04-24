"""
GC (Gia Công) Auto-Fill Service.

Ported from tool1_autofill/engine.py for Linux/Docker:
  - Fills EXISTING Samsung Excel files (not a template)
  - Targets column K at marker rows containing 절사 (Korean: truncation)
  - Clones L1→L2 folder structure for price adjustments
  - Uses openpyxl for read and write (no COM on Linux)
  - PDF via OnlyOffice x2t (per-sheet)
"""

from __future__ import annotations

import json
import logging
import os
import re
import shutil
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)

# ─── Constants ────────────────────────────────────────────────
ONEDRIVE_STAGING = Path("/data/onedrive-staging")
GC_OUTPUT_BASE = Path("/data/files/quotations/gc")
MARKER_PATTERN = "\uc808\uc0ac"  # 절사 (Korean: truncation/rounding)
COL_K = column_index_from_string("K")


# ─── Text Utilities (ported from engine.py) ───────────────────

def _clean(val: Any) -> str:
    if val is None:
        return ""
    return str(val).replace("\xa0", " ").strip()


def _norm_key(v: Any) -> str:
    s = _clean(v).lower().replace(" ", "")
    return re.sub(r"[^a-z0-9]", "", s)


def _to_decimal(v: Any) -> Decimal | None:
    s = _clean(v).replace(" ", "").replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
        if not m:
            return None
        try:
            return Decimal(m.group(0))
        except (InvalidOperation, ValueError):
            return None


# ─── Excel Scanning Helpers ───────────────────────────────────

def _is_marker_cell(val: Any) -> bool:
    s = _clean(val)
    if not s:
        return False
    compact = re.sub(r"\s+", "", s)
    return MARKER_PATTERN in compact


def _find_row_by_text(ws, keywords: list[str]) -> int:
    max_r = min(ws.max_row or 1, 350)
    max_c = min(ws.max_column or 1, 12)
    for r in range(1, max_r + 1):
        row_text = " ".join(
            _clean(ws.cell(row=r, column=c).value).lower()
            for c in range(1, max_c + 1)
        )
        if all(k.lower() in row_text for k in keywords):
            return r
    return 0


def _find_rows_by_text(ws, keywords: list[str]) -> list[int]:
    rows = []
    max_r = min(ws.max_row or 1, 350)
    max_c = min(ws.max_column or 1, 12)
    for r in range(1, max_r + 1):
        row_text = " ".join(
            _clean(ws.cell(row=r, column=c).value).lower()
            for c in range(1, max_c + 1)
        )
        if all(k.lower() in row_text for k in keywords):
            rows.append(r)
    return rows


def _marker_rows_in_sheet(ws) -> list[int]:
    rows = []
    max_r = min(ws.max_row or 1, 350)
    max_c = min(ws.max_column or 1, 14)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if _is_marker_cell(ws.cell(row=r, column=c).value):
                rows.append(r)
                break
    return rows


def _fallback_cut_row(ws) -> int:
    r_profit = _find_row_by_text(ws, ["profit"])
    r_result = _find_row_by_text(ws, ["result", "total", "amount"])
    if r_profit and r_result and (r_profit + 1) < r_result:
        return r_profit + 1
    if r_profit:
        return r_profit + 1
    return 0


def _is_summary_sheet(ws) -> bool:
    t = _clean(ws.title).lower()
    fold = unicodedata.normalize("NFKD", t)
    fold = "".join(ch for ch in fold if not unicodedata.combining(ch))
    fold = fold.replace("đ", "d")
    if any(k in fold for k in ["material", "process", "summary", "tong hop", "total"]):
        return True
    max_r = min(ws.max_row or 1, 80)
    max_c = min(ws.max_column or 1, 12)
    code_pat = re.compile(r"Z\d{6,}-\d{3,}", re.IGNORECASE)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = _clean(ws.cell(row=r, column=c).value)
            if code_pat.search(v):
                return False
    return True


def _find_code_for_sheet(ws, price_map_keys: set[str]) -> str:
    title_key = _norm_key(ws.title)
    for code_key in price_map_keys:
        if code_key and code_key in title_key:
            return code_key
    max_r = min(ws.max_row or 1, 220)
    max_c = min(ws.max_column or 1, 26)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell_key = _norm_key(ws.cell(row=r, column=c).value)
            if cell_key in price_map_keys:
                return cell_key
    return ""


def _calc_new_k_value(current_value: Any, delta_value: Any) -> int | float:
    cur = _to_decimal(current_value)
    delta = _to_decimal(delta_value)
    cur = abs(cur) if cur is not None else Decimal(0)
    delta = abs(delta) if delta is not None else Decimal(0)
    result = cur + delta
    return int(result) if result == result.to_integral_value() else float(result)


def _find_material_process_row(ws, r_profit: int) -> int:
    rows = _find_rows_by_text(ws, ["material", "process", "total"])
    if not rows:
        return 0
    above = [r for r in rows if r <= r_profit] if r_profit else rows
    return max(above) if above else rows[-1]


def _find_management_row(ws, r_profit: int) -> int:
    if not r_profit:
        return 0
    r = _find_row_by_text(ws, ["management", "expenses"])
    if r:
        return r
    for rr in range(max(1, r_profit - 6), r_profit):
        txt = " ".join(
            _clean(ws.cell(row=rr, column=c).value).lower() for c in range(1, 8)
        )
        if txt.strip():
            return rr
    return max(1, r_profit - 1)


def _find_result_row(ws, r_cut: int) -> int:
    result_rows = _find_rows_by_text(ws, ["result", "total", "amount"])
    if result_rows:
        return result_rows[-1]
    if not r_cut:
        return 0
    max_r = min(ws.max_row or 1, r_cut + 12)
    for rr in range(r_cut + 1, max_r + 1):
        txt = " ".join(
            _clean(ws.cell(row=rr, column=c).value).lower() for c in range(1, 8)
        )
        if "result" in txt or "total" in txt:
            return rr
    return min(ws.max_row or 1, r_cut + 1)


def _verify_or_fix_formula(ws) -> tuple[str, str, int]:
    """Build expected Result Total formula. Returns (status, formula, result_row)."""
    r_profit = _find_row_by_text(ws, ["profit"])
    r_cut = _fallback_cut_row(ws)
    r_mat = _find_material_process_row(ws, r_profit)
    r_mgmt = _find_management_row(ws, r_profit)
    r_result = _find_result_row(ws, r_cut)

    if not all([r_profit, r_cut, r_result]):
        return "skip", "missing anchors", 0

    sum_start = r_mat if r_mat else r_mgmt
    if not sum_start:
        return "skip", "missing sum-start anchors", 0

    expected = f"=ROUNDUP((SUM(K{sum_start}:K{r_profit})-K{r_cut}),-3)"
    return "fixed", expected, r_result


def _extract_level(name: str) -> int:
    n = _clean(name).lower()
    fold = unicodedata.normalize("NFKD", n)
    fold = "".join(ch for ch in fold if not unicodedata.combining(ch))
    fold = fold.replace("đ", "d")
    m = re.search(r"\bl\s*([0-9]{1,2})\b", fold)
    if m:
        return int(m.group(1))
    m = re.search(r"lan[^0-9]*([0-9]{1,2})", fold)
    if m:
        return int(m.group(1))
    return -1


# ─── Public API Functions ─────────────────────────────────────

async def detect_gc_files(
    rfq_no: str,
    year: int | None = None,
    month: int | None = None,
) -> dict[str, Any]:
    """Scan OneDrive staging to find the RFQ folder and its L-subfolders."""
    rfq_key = _norm_key(rfq_no)
    if not rfq_key:
        return {"rfq_folder": "", "levels": [], "max_level": 0}

    base = ONEDRIVE_STAGING / "Puplic" / "BQMS" / "RFQ"
    if not base.exists():
        # Fallback: try "Public"
        base = ONEDRIVE_STAGING / "Public" / "BQMS" / "RFQ"
    if not base.exists():
        return {"rfq_folder": "", "levels": [], "max_level": 0}

    # Narrow search by year/month if provided
    year_dirs = []
    if year:
        yd = base / f"RFQ {year}"
        if yd.exists():
            year_dirs = [yd]
    if not year_dirs:
        year_dirs = sorted(base.iterdir()) if base.exists() else []
        year_dirs = [d for d in year_dirs if d.is_dir()]

    month_dirs = []
    for yd in year_dirs:
        if month:
            md = yd / f"THANG {month}"
            if md.exists():
                month_dirs.append(md)
        else:
            month_dirs.extend(d for d in sorted(yd.iterdir()) if d.is_dir())

    # Find matching RFQ folder
    candidates = []
    for md in month_dirs:
        try:
            entries = list(md.iterdir())
        except PermissionError:
            continue
        for entry in entries:
            if not entry.is_dir():
                continue
            entry_key = _norm_key(entry.name)
            if entry_key.startswith(rfq_key):
                try:
                    mt = entry.stat().st_mtime
                except OSError:
                    mt = 0
                candidates.append((mt, entry))

    if not candidates:
        return {"rfq_folder": "", "levels": [], "max_level": 0}

    candidates.sort(key=lambda x: x[0], reverse=True)
    rfq_folder = candidates[0][1]

    # Find L-subfolders and list Excel files
    levels = []
    for sub in sorted(rfq_folder.iterdir()):
        if not sub.is_dir():
            continue
        lv = _extract_level(sub.name)
        if lv < 0:
            continue
        excel_files = [
            f.name for f in sorted(sub.iterdir())
            if f.suffix.lower() in (".xlsx", ".xlsm")
            and not f.name.startswith("~$")
        ]
        if excel_files:
            levels.append({
                "level": lv,
                "folder": str(sub),
                "excel_files": excel_files,
            })

    levels.sort(key=lambda x: x["level"])
    max_level = max((lv["level"] for lv in levels), default=0)

    return {
        "rfq_folder": str(rfq_folder),
        "levels": levels,
        "max_level": max_level,
    }


async def scan_markers(
    excel_path: str,
    price_map: dict[str, float],
) -> list[dict[str, Any]]:
    """Open Excel read-only and scan each sheet for GC markers."""
    resolved = str(Path(excel_path).resolve())
    if not resolved.startswith("/data/onedrive-staging/"):
        raise PermissionError(f"Path ngoài staging: {excel_path}")
    if not os.path.isfile(resolved):
        raise FileNotFoundError(f"Excel không tồn tại: {excel_path}")

    # Normalize price_map keys
    norm_map: dict[str, dict[str, Any]] = {}
    for code, price in price_map.items():
        nk = _norm_key(code)
        if nk:
            norm_map[nk] = {"code": code, "price": price}

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheets_info = []

    try:
        for ws in wb.worksheets:
            marker_rows = _marker_rows_in_sheet(ws)
            fstat, fval, f_row = _verify_or_fix_formula(ws)

            if _is_summary_sheet(ws):
                sheets_info.append({
                    "sheet_name": ws.title,
                    "sheet_type": "summary",
                    "bqms_code": "",
                    "marker_rows": marker_rows[:8],
                    "target_row": 0,
                    "current_k_value": None,
                    "suggested_price": None,
                    "new_k_value": None,
                    "formula_status": fstat,
                    "formula": fval,
                    "formula_row": f_row,
                    "status": "summary_skip",
                })
                continue

            code = _find_code_for_sheet(ws, set(norm_map.keys()))
            if not code:
                sheets_info.append({
                    "sheet_name": ws.title,
                    "sheet_type": "product",
                    "bqms_code": "",
                    "marker_rows": marker_rows[:8],
                    "target_row": 0,
                    "current_k_value": None,
                    "suggested_price": None,
                    "new_k_value": None,
                    "formula_status": fstat,
                    "formula": fval,
                    "formula_row": f_row,
                    "status": "no_code",
                })
                continue

            info = norm_map.get(code)
            price = info["price"] if info else None
            original_code = info["code"] if info else ""

            # Find target row (below Profit, fallback to marker rows)
            target_row = _fallback_cut_row(ws)
            if not target_row and marker_rows:
                target_row = marker_rows[0]

            current_k = None
            new_k = None
            status = "no_marker"

            if target_row:
                current_k_raw = ws.cell(row=target_row, column=COL_K).value
                _cur_dec = _to_decimal(current_k_raw)
                current_k = float(_cur_dec) if _cur_dec is not None else 0
                if price is not None:
                    new_k = _calc_new_k_value(current_k_raw, price)
                    status = "ready"
                else:
                    status = "no_price"
            elif price is None:
                status = "no_price"

            sheets_info.append({
                "sheet_name": ws.title,
                "sheet_type": "product",
                "bqms_code": original_code,
                "marker_rows": marker_rows[:8],
                "target_row": target_row,
                "current_k_value": current_k,
                "suggested_price": price,
                "new_k_value": new_k,
                "formula_status": fstat,
                "formula": fval,
                "formula_row": f_row,
                "status": status,
            })
    finally:
        wb.close()

    return sheets_info


async def clone_l_folder(
    rfq_folder: str,
    target_level: int,
) -> tuple[str, str]:
    """Clone source L-folder to create target L-folder.

    Returns (source_folder, cloned_folder).
    For L1: returns (L1_path, L1_path) — no clone.
    """
    rfq_path = Path(rfq_folder).resolve()
    if not str(rfq_path).startswith("/data/onedrive-staging/"):
        raise PermissionError(f"RFQ folder ngoài staging: {rfq_folder}")
    if not rfq_path.is_dir():
        raise RuntimeError(f"RFQ folder không tồn tại: {rfq_folder}")

    # List L-subfolders
    subdirs = []
    for entry in rfq_path.iterdir():
        if entry.is_dir():
            lv = _extract_level(entry.name)
            if lv > 0:
                subdirs.append((lv, entry))

    subdirs.sort(key=lambda x: x[0])

    if target_level <= 1:
        # L1: use directly, no clone
        l1 = next((p for lv, p in subdirs if lv == 1), None)
        if not l1:
            raise RuntimeError("Không tìm thấy thư mục L1")
        return str(l1), str(l1)

    # Find source: L{target-1}
    source_level = target_level - 1
    source = next((p for lv, p in subdirs if lv == source_level), None)
    if not source:
        # Fallback to highest available
        if subdirs:
            source = subdirs[-1][1]
        else:
            raise RuntimeError(f"Không tìm thấy thư mục L{source_level}")

    # Build destination name
    dst_name = re.sub(
        r"[Ll]\s*\d+",
        f"L{target_level}",
        source.name,
        count=1,
    )
    if dst_name == source.name:
        dst_name = f"L{target_level}"

    dst_path = rfq_path / dst_name
    if dst_path.exists():
        for v in range(2, 100):
            cand = rfq_path / f"{dst_name}_v{v}"
            if not cand.exists():
                dst_path = cand
                break

    shutil.copytree(str(source), str(dst_path))
    logger.info("Cloned %s -> %s", source, dst_path)

    return str(source), str(dst_path)


async def apply_gc_edits(
    excel_path: str,
    sheet_edits: list[dict[str, Any]],
) -> dict[str, Any]:
    """Apply price edits to GC Excel via openpyxl.

    Creates a .original backup before editing for PDF repair.
    """
    if not os.path.isfile(excel_path):
        return {"success": False, "edited_sheets": 0, "errors": [f"File không tồn tại: {excel_path}"]}

    # Backup original for _repair_xlsx_for_x2t
    original_backup = excel_path + ".original"
    if not os.path.exists(original_backup):
        shutil.copy2(excel_path, original_backup)

    errors = []
    edited = 0

    try:
        wb = openpyxl.load_workbook(excel_path, keep_vba=True, keep_links=True)

        for edit in sheet_edits:
            sheet_name = edit.get("sheet_name", "")
            target_row = edit.get("target_row", 0)
            new_k = edit.get("new_k_value")

            if not sheet_name or not target_row or new_k is None:
                continue

            if sheet_name not in wb.sheetnames:
                errors.append(f"Sheet '{sheet_name}' không tồn tại")
                continue

            ws = wb[sheet_name]

            # Write the new K value
            ws.cell(row=target_row, column=COL_K).value = new_k

            # Fix Result Total formula if needed
            formula_row = edit.get("formula_row", 0)
            formula = edit.get("formula", "")
            if formula_row and formula:
                ws.cell(row=formula_row, column=COL_K).value = formula

            edited += 1

        wb.save(excel_path)
        wb.close()
    except Exception as exc:
        errors.append(f"Lỗi khi sửa Excel: {exc}")
        return {"success": False, "edited_sheets": edited, "errors": errors}

    return {"success": True, "edited_sheets": edited, "errors": errors}


async def run_gc_autofill_job(
    conn,
    quotation_id: int,
    rfq_no: str,
    quote_level: int,
    sheet_edits: list[dict[str, Any]],
    gc_source_folder: str,
) -> dict[str, Any]:
    """Full GC auto-fill pipeline: clone → edit → PDF → save."""
    from app.services.gotenberg_service import convert_xlsx_to_pdf

    result: dict[str, Any] = {
        "success": False,
        "quotation_id": quotation_id,
        "files": [],
        "edit_report": [],
        "cloned_folder": "",
        "total_sheets": len(sheet_edits),
        "edited_sheets": 0,
        "errors": [],
    }

    rfq_folder = str(Path(gc_source_folder).parent)
    if not os.path.isdir(rfq_folder):
        result["errors"].append(f"RFQ folder không tồn tại: {rfq_folder}")
        await _update_quotation_status(conn, quotation_id, "failed", result)
        return result

    # Step 1: Clone L-folder (or use L1 directly)
    try:
        source_folder, cloned_folder = await clone_l_folder(rfq_folder, quote_level)
    except Exception as exc:
        result["errors"].append(f"Clone folder thất bại: {exc}")
        await _update_quotation_status(conn, quotation_id, "failed", result)
        return result
    result["cloned_folder"] = cloned_folder

    # Step 2: Find the Excel file in cloned folder
    cloned_path = Path(cloned_folder)
    excel_files = [
        f for f in sorted(cloned_path.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True)
        if f.suffix.lower() in (".xlsx", ".xlsm") and not f.name.startswith("~$")
    ]
    if not excel_files:
        result["errors"].append("Không tìm thấy file Excel trong thư mục clone")
        await _update_quotation_status(conn, quotation_id, "failed", result)
        return result
    excel_path = str(excel_files[0])

    # Step 3: Apply edits
    ready_edits = [s for s in sheet_edits if s.get("status") == "ready"]
    edit_result = await apply_gc_edits(excel_path, ready_edits)
    result["edited_sheets"] = edit_result["edited_sheets"]
    result["edit_report"] = sheet_edits
    if edit_result["errors"]:
        result["errors"].extend(edit_result["errors"])

    # Step 4: Create output directory
    now_str = f"{rfq_no}_L{quote_level}"
    output_dir = GC_OUTPUT_BASE / rfq_no / now_str
    output_dir.mkdir(parents=True, exist_ok=True)

    # Copy edited Excel to output
    output_xlsx = output_dir / excel_files[0].name
    shutil.copy2(excel_path, str(output_xlsx))
    result["files"].append({
        "type": "gc_xlsx",
        "path": str(output_xlsx),
        "name": excel_files[0].name,
    })

    # Step 5: Convert each sheet to individual PDF
    pdf_errors = []
    for edit in ready_edits:
        sheet_name = edit.get("sheet_name", "")
        bqms_code = edit.get("bqms_code", sheet_name)
        if not sheet_name:
            continue

        try:
            # Create single-sheet workbook for PDF conversion
            temp_xlsx = str(output_dir / f"_temp_{_norm_key(sheet_name)}.xlsx")
            wb = openpyxl.load_workbook(excel_path, keep_vba=True, keep_links=True)

            # Remove all sheets except the target
            for sn in list(wb.sheetnames):
                if sn != sheet_name:
                    del wb[sn]
            wb.save(temp_xlsx)
            wb.close()

            # Convert to PDF
            safe_code = re.sub(r'[\\/:*?"<>|]', '_', bqms_code)
            pdf_name = f"{rfq_no}_{safe_code}_AMABACNINH_L{quote_level}.pdf"
            pdf_path = str(output_dir / pdf_name)

            await convert_xlsx_to_pdf(temp_xlsx, pdf_path)

            if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                result["files"].append({
                    "type": f"gc_pdf_{bqms_code}",
                    "path": pdf_path,
                    "name": pdf_name,
                })
            else:
                pdf_errors.append(f"PDF rỗng cho sheet {sheet_name}")

            # Cleanup temp
            try:
                os.unlink(temp_xlsx)
            except OSError:
                pass

        except Exception as exc:
            pdf_errors.append(f"PDF lỗi sheet '{sheet_name}': {exc}")
            logger.warning("PDF conversion failed for sheet %s: %s", sheet_name, exc)

    if pdf_errors:
        result["errors"].extend(pdf_errors)

    # Cleanup .original backup
    original_backup = excel_path + ".original"
    try:
        if os.path.exists(original_backup):
            os.unlink(original_backup)
    except OSError:
        pass

    # Step 6: Update quotation record
    result["success"] = edit_result["success"]
    status = "completed" if result["success"] else "failed"
    await _update_quotation_status(conn, quotation_id, status, result)

    return result


async def _update_quotation_status(
    conn, quotation_id: int, status: str, result: dict[str, Any]
) -> None:
    """Update quotation record with GC results."""
    output_pdf = ""
    for f in result.get("files", []):
        if "pdf" in f.get("type", ""):
            output_pdf = f.get("path", "")
            break

    output_xlsx = ""
    for f in result.get("files", []):
        if f.get("type") == "gc_xlsx":
            output_xlsx = f.get("path", "")
            break

    try:
        await conn.execute(
            """
            UPDATE quotations SET
                status = $2,
                output_pdf = $3,
                output_xlsx = $4,
                gc_cloned_folder = $5,
                gc_sheet_report = $6::jsonb,
                updated_at = NOW()
            WHERE id = $1
            """,
            quotation_id,
            status,
            output_pdf,
            output_xlsx,
            result.get("cloned_folder", ""),
            json.dumps(result.get("edit_report", []), default=str, ensure_ascii=False),
        )
    except Exception as exc:
        logger.error("Failed to update quotation %d: %s", quotation_id, exc)
