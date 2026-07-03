"""GC (Gia Cong) Template-Based Quotation Builder — NEW flow per Thang 2026-05-10.

Distinct from `gc_autofill_service.py` (the OLD 절사-marker flow that edits
existing Samsung Excel files). This module builds a FRESH quotation
workbook from the AMA-controlled template `QUOTATION_GC.xlsx`:

  - 1 sheet per item code (mã linh kiện). Sheet name = bqms_code.
  - Each sheet has the item's image embedded at L2 (TwoCellAnchor).
  - PRESERVE all formulas (cross-sheet refs to Material&Process are
    rewritten so each new sheet references its own data, not the template).
  - File must be EDITABLE (no sheet protection, no read-only flags).
  - Template carries 1 sample item sheet (e.g. 'Z0000002-037802') + 1
    Material&Process lookup sheet. We clone the sample sheet per item,
    rewrite cross-sheet formula refs (sample-name → bqms-code), then
    delete the original sample sheet so output only has real items.

Output: 1 multi-sheet workbook (xlsx) + 1 PDF (Gotenberg). Saved into
the L1 subfolder built by `quote_round_subfolder()`.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date as _date, datetime
from pathlib import Path
from typing import Any

import openpyxl

try:
    from zoneinfo import ZoneInfo
except ImportError:  # py<3.9 fallback (project pins 3.11+, kept for safety)
    ZoneInfo = None  # type: ignore[assignment]

logger = logging.getLogger(__name__)

# Thang 2026-06-13 (Bug fix T3): VN tz is the canonical "today" for all
# quotations regardless of server tz. Hardcoded so quote dates never drift
# when ops moves the worker between regions.
_VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh") if ZoneInfo else None


# Thang 2026-06-13: defensive date sweep — same patterns as
# autofill_service._refresh_dates_in_sheet. Templates often carry hardcoded
# dates from the original author that survive cell-by-cell writes. We scan
# every cell and replace VN/EN date strings with the caller's `quote_date`.
_DATE_VN_PAT = re.compile(
    r"Ngày\s*\d{1,2}\s*Tháng\s*\d{1,2}\s*năm\s*\d{4}",
    flags=re.IGNORECASE,
)
_DATE_SLASH_PAT = re.compile(r"\b\d{1,2}/\d{1,2}/\d{4}\b")
_DATE_DMY_PAT = re.compile(r"\b\d{1,2}-\d{1,2}-\d{4}\b")


def _refresh_gc_dates_in_sheet(ws, now: datetime) -> int:
    """Replace stale dates in a GC quotation sheet with `now`.

    Mirrors autofill_service._refresh_dates_in_sheet but kept local to avoid
    cross-module coupling. Returns count of cells updated.
    """
    vn_today = f"Ngày {now.day} Tháng {now.month} năm {now.year}"
    slash_today = now.strftime("%d/%m/%Y")
    updated = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v:
                continue
            new_v = v
            if _DATE_VN_PAT.search(new_v):
                new_v = _DATE_VN_PAT.sub(vn_today, new_v)
            stripped = new_v.strip()
            if _DATE_SLASH_PAT.fullmatch(stripped):
                new_v = slash_today
            if _DATE_DMY_PAT.fullmatch(stripped):
                new_v = now.strftime("%d-%m-%Y")
            if new_v != v:
                try:
                    cell.value = new_v
                    updated += 1
                except Exception:
                    pass
    return updated


def _vn_now() -> datetime:
    """Today in Asia/Ho_Chi_Minh tz, naive (no tzinfo) for openpyxl writes."""
    if _VN_TZ is None:
        return datetime.now()
    return datetime.now(tz=_VN_TZ).replace(tzinfo=None)


def _coerce_quote_date(quote_date) -> datetime:
    """Normalize caller-supplied quote_date → datetime. None = today (VN tz).

    Thang 2026-06-13 (Bug fix T3): default FORCED to VN tz so the date
    stamped on quotations matches Hanoi office regardless of server tz.
    """
    if quote_date is None:
        return _vn_now()
    if isinstance(quote_date, datetime):
        return quote_date
    if isinstance(quote_date, _date):
        return datetime(quote_date.year, quote_date.month, quote_date.day)
    # ISO string fallback (yyyy-mm-dd from frontend)
    if isinstance(quote_date, str):
        s = quote_date.strip()
        if s:
            try:
                return datetime.fromisoformat(s)
            except ValueError:
                pass
    return _vn_now()


# Regex for the "sample item sheet" — Samsung BQMS item codes.
# Examples: Z0000002-037802, RC01H00I-000413
_BQMS_CODE_RE = re.compile(r"^[A-Z]{1,3}[0-9A-Z]{4,8}-[0-9]{4,8}$", re.IGNORECASE)

# Excel sheet-name forbidden chars: : \ / ? * [ ]   (max length = 31 chars)
_SHEET_FORBIDDEN_RE = re.compile(r"[:\\/?\*\[\]]")


def _safe_sheet_name(name: str, taken: set[str]) -> str:
    """Sanitize a name for use as an Excel sheet title (≤31 chars, unique)."""
    base = _SHEET_FORBIDDEN_RE.sub("-", (name or "Sheet").strip())[:31] or "Sheet"
    if base not in taken:
        return base
    for i in range(2, 100):
        cand = f"{base[:28]}_{i}"[:31]
        if cand not in taken:
            return cand
    return base[:31]


def _find_sample_sheet(wb) -> str | None:
    """Find the per-item template sheet by name pattern (BQMS code)."""
    for name in wb.sheetnames:
        if _BQMS_CODE_RE.match(name.strip()):
            return name
    # Fallback: first sheet not named 'Material&Process' / 'Process_list' / 'Material_list'
    for name in wb.sheetnames:
        low = name.lower().replace(" ", "")
        if "material" in low or "process" in low:
            continue
        return name
    return None


def _rewrite_self_refs(ws, old_name: str, new_name: str) -> int:
    """Walk every cell on `ws` and rewrite `'<old_name>'!` → `'<new_name>'!`
    inside formulas. openpyxl's copy_worksheet does NOT auto-update self-refs
    — we have to do it manually.

    Returns number of cells rewritten."""
    n = 0
    needle_quoted = f"'{old_name}'"
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            new_v = v
            if needle_quoted in new_v:
                new_v = new_v.replace(needle_quoted, f"'{new_name}'")
            # Also handle unquoted refs (rare — sheet names without spaces)
            if (old_name + "!") in new_v and needle_quoted not in v:
                new_v = new_v.replace(old_name + "!", f"'{new_name}'!")
            if new_v != v:
                cell.value = new_v
                n += 1
    return n


def _range_pixel_size(
    ws, from_col: int, from_row: int, to_col: int, to_row: int,
) -> tuple[int, int]:
    """Approximate pixel dims of cell range (from..to inclusive).
    Excel defaults: col-width-unit ≈ 7 px, row-height-pt ≈ 1.333 px.
    Default col width 8.43, default row height 15 pt."""
    from openpyxl.utils import get_column_letter
    width_units = 0.0
    for c in range(from_col, to_col + 1):
        w = ws.column_dimensions[get_column_letter(c)].width
        width_units += (w if w and w > 0 else 8.43)
    height_pts = 0.0
    for r in range(from_row, to_row + 1):
        h = ws.row_dimensions[r].height
        height_pts += (h if h and h > 0 else 15.0)
    return max(40, int(width_units * 7.0)), max(30, int(height_pts * 1.333))


def _attach_image_two_cell(
    ws,
    img_bytes: bytes,
    *,
    from_col: int,
    from_row: int,
    to_col: int,
    to_row: int,
    max_w_px: int = 700,
    max_h_px: int = 380,
    fill_ratio: float = 0.96,
) -> bool:
    """Embed an image CENTERED and SCALED-TO-FIT inside the (from..to) range.

    Thang 2026-05-15: ảnh xuất ra PDF nhỏ vì trước đây dùng `PIL.thumbnail()`
    chỉ thu nhỏ về `max_w_px×max_h_px`. Khi ảnh gốc < cell range → ảnh giữ
    nguyên size nhỏ → khoảng trống lớn trong PDF. Giờ:
      1. PIL.resize() proportional (fit-contain) tới ~96% cell range để chừa
         viền mỏng đẹp mắt.
      2. ENLARGE nếu ảnh gốc nhỏ hơn cell.
      3. Cap upscale tại 4× original để tránh blur quá mức.
      4. Center theo cả 2 trục bằng offsets.
    """
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    try:
        from PIL import Image as PILImage
        pil = PILImage.open(io.BytesIO(img_bytes))
        if pil.mode not in ("RGB", "RGBA"):
            pil = pil.convert("RGBA")

        range_w, range_h = _range_pixel_size(ws, from_col, from_row, to_col, to_row)
        target_w = max(40, int(range_w * fill_ratio))
        target_h = max(40, int(range_h * fill_ratio))

        orig_w, orig_h = pil.size
        # Fit-contain scale (preserve aspect)
        scale = min(target_w / orig_w, target_h / orig_h)
        # Upscale cap: don't blow images up more than 4× their natural size
        # (LibreOffice/Gotenberg renders enlarged images blurry past this).
        scale = min(scale, 4.0)
        # But always at least scale 1.0 if the image is already smaller than cap
        # — let it fill the cell as much as the cap allows.
        new_w = max(40, int(orig_w * scale))
        new_h = max(40, int(orig_h * scale))

        if (new_w, new_h) != (orig_w, orig_h):
            # Pillow 12: prefer Resampling enum; falls back to module-level
            # alias on older Pillow installs.
            _lanczos = getattr(PILImage, "Resampling", PILImage).LANCZOS
            pil = pil.resize((new_w, new_h), _lanczos)

        buf = io.BytesIO()
        pil.save(buf, format="PNG")
        buf.seek(0)
        xl_img = XLImage(buf)
        img_w, img_h = pil.size
        xl_img.width = img_w
        xl_img.height = img_h

        off_x = max(0, (range_w - img_w) // 2)
        off_y = max(0, (range_h - img_h) // 2)

        marker = AnchorMarker(
            col=from_col - 1, colOff=pixels_to_EMU(off_x),
            row=from_row - 1, rowOff=pixels_to_EMU(off_y),
        )
        ext = XDRPositiveSize2D(
            cx=pixels_to_EMU(img_w), cy=pixels_to_EMU(img_h),
        )
        xl_img.anchor = OneCellAnchor(_from=marker, ext=ext)
        ws._images.append(xl_img)
        logger.info(
            "GC image fit-attached: range=%dx%dpx, img=%dx%dpx (orig %dx%d, scale=%.2f), "
            "offset=(%d,%d) → cells %s%d:%s%d",
            range_w, range_h, img_w, img_h, orig_w, orig_h, scale,
            off_x, off_y,
            chr(64 + from_col), from_row, chr(64 + to_col), to_row,
        )
        return True
    except Exception as exc:
        logger.warning("GC centered image attach failed: %s", exc)
        # Fallback: top-left anchor with raw image
        try:
            from openpyxl.utils import get_column_letter
            buf = io.BytesIO(img_bytes)
            xl_img = XLImage(buf)
            ws.add_image(xl_img, f"{get_column_letter(from_col)}{from_row}")
            return True
        except Exception:
            return False


def fill_gc_quotation_from_template(
    template_path: str,
    items: list[dict[str, Any]],
    images_map: dict[str, bytes],
    rfq_no: str,
    output_path: str,
    quote_date=None,
) -> dict[str, Any]:
    """Build a multi-sheet GC quotation workbook from the AMA template.

    Args:
        template_path: path to QUOTATION_GC.xlsx (with sample item sheet
                       + Material&Process lookup sheet).
        items: list of dicts. Required key: bqms (item code).
               Optional: spec/description, short_name, so_luong, don_vi, maker.
        images_map: {bqms_code: image_bytes} from match_images_to_orders.
        rfq_no: RFQ number (kept for header context, not yet wired to a cell).
        output_path: where to write the resulting .xlsx.
        quote_date: optional date/datetime stamped on every sheet (G4 cell)
                    + defensive sweep over template-residual date strings.
                    Defaults to today. Per Thang 2026-06-13.

    Returns:
        {sheets_created, images_attached, formula_rewrites, errors[]}
    """
    if not Path(template_path).exists():
        raise FileNotFoundError(f"GC template not found: {template_path}")
    if not items:
        raise ValueError("GC quotation needs at least one item")

    quote_dt = _coerce_quote_date(quote_date)

    wb = openpyxl.load_workbook(template_path)
    sample_name = _find_sample_sheet(wb)
    if not sample_name:
        raise RuntimeError(
            f"Could not find sample item sheet in template (sheets: {wb.sheetnames})"
        )
    sample_sheet = wb[sample_name]

    sheets_created = 0
    images_attached = 0
    formula_rewrites = 0
    errors: list[str] = []

    taken: set[str] = set(wb.sheetnames)

    # Only items WITH a bqms code can become a sheet (sheet name must be non-empty).
    valid_items = [it for it in items if (it.get("bqms") or "").strip()]
    if not valid_items:
        raise ValueError("No items have a bqms code — GC needs at least one")

    for item in valid_items:
        bqms = (item.get("bqms") or "").strip()
        sheet_name = _safe_sheet_name(bqms, taken)
        taken.add(sheet_name)

        # Clone the sample sheet → fresh copy at the END of the workbook.
        new_sheet = wb.copy_worksheet(sample_sheet)
        new_sheet.title = sheet_name

        # Rewrite formulas that reference the OLD sample sheet → new sheet name.
        formula_rewrites += _rewrite_self_refs(new_sheet, sample_name, sheet_name)

        # Fill header cells per template layout:
        #   C4 = Quotation No. (= bqms code)
        #   G4 = Date (dd/mm/yyyy, Thang 2026-06-13)
        #   D9 = JIG NAME (= spec/description)
        new_sheet["C4"] = bqms
        try:
            new_sheet["G4"] = quote_dt.strftime("%d/%m/%Y")
        except Exception as exc:
            logger.warning("GC: stamping date on G4 failed for %s: %s", bqms, exc)
        spec_text = (
            item.get("spec") or item.get("description")
            or item.get("short_name") or ""
        )
        if spec_text:
            new_sheet["D9"] = spec_text

        # Defensive: sweep stale date strings (template carries hardcoded
        # author dates that survive cell-by-cell writes when they live in
        # unrelated header/footer cells).
        try:
            _refresh_gc_dates_in_sheet(new_sheet, quote_dt)
        except Exception as exc:
            logger.warning("GC: date sweep failed for %s: %s", bqms, exc)

        # Embed item image at L2 (matches template's anchor for stamp/photo).
        # Span L2:N8 — image fills the visible area without overflowing into
        # the items table on the left.
        img_bytes = images_map.get(bqms) if images_map else None
        if img_bytes:
            ok = _attach_image_two_cell(
                new_sheet, img_bytes,
                from_col=12, from_row=2,   # L2
                to_col=14, to_row=8,
            )
            if ok:
                images_attached += 1
            else:
                errors.append(f"image attach failed for {bqms}")
        else:
            logger.info("GC: no image for bqms=%r", bqms)

        sheets_created += 1

    # Delete the original sample sheet (we kept it intact while cloning).
    if sample_name in wb.sheetnames:
        del wb[sample_name]

    # Re-order: item sheets first (in input order), Material&Process last.
    desired: list[str] = []
    for it in valid_items:
        bqms = (it.get("bqms") or "").strip()
        for sn in wb.sheetnames:
            if sn == bqms or (bqms and sn.startswith(bqms[:28])):
                if sn not in desired:
                    desired.append(sn)
                break
    for sn in wb.sheetnames:
        if sn not in desired:
            desired.append(sn)
    wb._sheets = [wb[name] for name in desired]

    # NO sheet protection — file fully editable per Thang.
    # Also force landscape + fit-to-page so the image at col L (~1340px
    # from left) fits inside the printable PDF area. Without this,
    # LibreOffice/Gotenberg conversion silently drops images that fall
    # outside the default portrait-A4 printable region.
    # Per Thang 2026-05-11.
    from openpyxl.worksheet.page import PageMargins
    for sn in wb.sheetnames:
        try:
            ws = wb[sn]
            ws.protection.disable()
            # Thang 2026-05-20: switched to A4 portrait per user request.
            # fitToWidth=1 ensures wide content (image column + price columns)
            # scales down to fit portrait width instead of getting clipped.
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins = PageMargins(
                left=0.3, right=0.3, top=0.4, bottom=0.4,
                header=0.2, footer=0.2,
            )
        except Exception as exc:
            logger.warning("page setup for %r failed: %s", sn, exc)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Post-process: openpyxl writes image rels with ABSOLUTE paths
    # (Target="/xl/media/image1.png") which LibreOffice/Gotenberg silently
    # reject when converting to PDF (xlsx opens fine in Excel, but PDF
    # comes out without any images). Fix by rewriting absolute paths to
    # relative (../media/image1.png) inside the saved xlsx zip.
    # Per Thang 2026-05-11.
    _fix_drawing_rels_paths(output_path)

    logger.info(
        "GC quotation built: %d sheets, %d images, %d formula rewrites → %s",
        sheets_created, images_attached, formula_rewrites, output_path,
    )

    return {
        "sheets_created": sheets_created,
        "images_attached": images_attached,
        "formula_rewrites": formula_rewrites,
        "errors": errors,
    }


def fill_gc_quotation_from_wizard(
    template_path: str,
    wizard_items: list[dict[str, Any]],
    rfq_no: str,
    output_path: str,
    images_map: dict[str, bytes] | None = None,
    quote_date=None,
) -> dict[str, Any]:
    """Build GC quotation xlsx with materials + processes injected from wizard.

    Per Thang 2026-05-11: GC wizard collects material/process rows + nego
    per item. We clone the sample sheet per item, then OVERRIDE specific
    cells in the Material section (B13:K27) and Process section (B66:K84)
    with wizard values. K90 (nego) is also set per item.

    Per Thang 2026-05-11 (round 2):
      - Only items in `wizard_items` get a sheet (caller pre-filters by selected).
      - If `images_map[bqms_code]` provided, attach image CENTERED into
        Product photo cell range B93:H98.

    wizard_items shape (from frontend):
      [{
        bqms_code, jig_name, spec, qty,
        materials/parts/others/processes: [...],
        nego: number,
      }, ...]
    """
    if not Path(template_path).exists():
        raise FileNotFoundError(f"GC template not found: {template_path}")
    if not wizard_items:
        raise ValueError("GC wizard needs at least one item")
    images_map = images_map or {}

    # Thang 2026-06-13: default quote_date = today (matches autofill_service).
    quote_dt = _coerce_quote_date(quote_date)

    wb = openpyxl.load_workbook(template_path)
    sample_name = _find_sample_sheet(wb)
    if not sample_name:
        raise RuntimeError(
            f"Could not find sample item sheet in template (sheets: {wb.sheetnames})"
        )
    sample_sheet = wb[sample_name]

    sheets_created = 0
    formula_rewrites = 0
    errors: list[str] = []
    taken: set[str] = set(wb.sheetnames)

    # Material section (rows 13-27, max 15): C=name, D-F=W/L/H, G=Qty,
    #   H=weight, J=unit_price, K=J*H (formula in template).
    MAT_FIRST_ROW = 13
    MAT_LAST_ROW = 27
    # Parts section (rows 30-52, max 23): C=name, H=Qty, J=unit_price, K=J*H.
    PARTS_FIRST_ROW = 30
    PARTS_LAST_ROW = 52
    # Other section (rows 55-61, max 7): C=description, H=Qty, J=unit_price, K=J*H.
    OTHER_FIRST_ROW = 55
    OTHER_LAST_ROW = 61
    # Process section (rows 66-84, max 19): C=name, H=time_hr, J=unit_price, K=J*H.
    PROC_FIRST_ROW = 66
    PROC_LAST_ROW = 84

    for item in wizard_items:
        bqms = (item.get("bqms_code") or "").strip()
        if not bqms:
            errors.append(f"item missing bqms_code, skipped: {item}")
            continue

        sheet_name = _safe_sheet_name(bqms, taken)
        taken.add(sheet_name)

        new_sheet = wb.copy_worksheet(sample_sheet)
        new_sheet.title = sheet_name

        formula_rewrites += _rewrite_self_refs(new_sheet, sample_name, sheet_name)

        # Header
        new_sheet["C4"] = bqms
        # Thang 2026-06-13: stamp date in G4 + defensive sweep.
        try:
            new_sheet["G4"] = quote_dt.strftime("%d/%m/%Y")
        except Exception as exc:
            logger.warning("GC wizard: G4 date stamp failed for %s: %s", bqms, exc)
        try:
            _refresh_gc_dates_in_sheet(new_sheet, quote_dt)
        except Exception as exc:
            logger.warning("GC wizard: date sweep failed for %s: %s", bqms, exc)
        if item.get("jig_name"):
            new_sheet["D9"] = item["jig_name"]

        # Materials
        materials = item.get("materials") or []
        for i, mat in enumerate(materials):
            row = MAT_FIRST_ROW + i
            if row > MAT_LAST_ROW:
                errors.append(
                    f"{bqms}: too many materials ({len(materials)}), "
                    f"only first {MAT_LAST_ROW - MAT_FIRST_ROW + 1} kept"
                )
                break
            new_sheet[f"C{row}"] = mat.get("name") or ""
            if mat.get("w") is not None:
                new_sheet[f"D{row}"] = mat["w"]
            if mat.get("l") is not None:
                new_sheet[f"E{row}"] = mat["l"]
            if mat.get("h") is not None:
                new_sheet[f"F{row}"] = mat["h"]
            new_sheet[f"G{row}"] = mat.get("qty", 1)
            # H{row} stays as the template's =PRODUCT(D:F)*density formula
            # but we override J with our user-entered unit_price (replacing
            # the VLOOKUP formula that does table lookup).
            new_sheet[f"J{row}"] = float(mat.get("unit_price") or 0)
            # Replace H to a simple value (D*E*F*qty / 10^6 in mm³ → kg-ish)
            # to avoid the template's density-based formula breaking when
            # the material isn't in Material&Process lookup table.
            d, l, h, q = (
                float(mat.get("w") or 0), float(mat.get("l") or 0),
                float(mat.get("h") or 0), float(mat.get("qty") or 1),
            )
            new_sheet[f"H{row}"] = round(d * l * h * q / 1_000_000, 4) or q
            # Phase F (Thang 2026-05-13): Amount = price × Weight (NOT × Qty).
            # Template's K{row} formula was =J*G (price*qty) → đổi thành =J*H
            # (price*weight) cho phù hợp với cách Material được tính theo cân
            # nặng vật liệu thay vì số lượng phôi.
            new_sheet[f"K{row}"] = f"=J{row}*H{row}"

        # Clear remaining material rows so leftover sample data doesn't
        # corrupt the sub-total.
        for row in range(MAT_FIRST_ROW + len(materials), MAT_LAST_ROW + 1):
            for col in ("C", "D", "E", "F", "G", "H", "J"):
                if new_sheet[f"{col}{row}"].value is not None:
                    new_sheet[f"{col}{row}"] = None

        # Parts (rows 30-52)
        parts = item.get("parts") or []
        for i, part in enumerate(parts):
            row = PARTS_FIRST_ROW + i
            if row > PARTS_LAST_ROW:
                errors.append(
                    f"{bqms}: too many parts ({len(parts)}), kept "
                    f"first {PARTS_LAST_ROW - PARTS_FIRST_ROW + 1}"
                )
                break
            new_sheet[f"C{row}"] = part.get("name") or ""
            new_sheet[f"H{row}"] = float(part.get("qty") or 0)
            new_sheet[f"J{row}"] = float(part.get("unit_price") or 0)
        # Clear leftover parts — Thang 2026-05-15: also clear C (preset names
        # like BOLT/SPRING/etc) when user didn't add any parts, so PDF doesn't
        # list them as ghost rows.
        for row in range(PARTS_FIRST_ROW + len(parts), PARTS_LAST_ROW + 1):
            for col in ("C", "H", "J"):
                if new_sheet[f"{col}{row}"].value is not None:
                    new_sheet[f"{col}{row}"] = None

        # Others (rows 55-61)
        others = item.get("others") or []
        for i, oth in enumerate(others):
            row = OTHER_FIRST_ROW + i
            if row > OTHER_LAST_ROW:
                errors.append(
                    f"{bqms}: too many other rows ({len(others)}), kept "
                    f"first {OTHER_LAST_ROW - OTHER_FIRST_ROW + 1}"
                )
                break
            new_sheet[f"C{row}"] = oth.get("description") or ""
            new_sheet[f"H{row}"] = float(oth.get("qty") or 0)
            new_sheet[f"J{row}"] = float(oth.get("unit_price") or 0)
        # Clear leftover others
        for row in range(OTHER_FIRST_ROW + len(others), OTHER_LAST_ROW + 1):
            for col in ("C", "H", "J"):
                if new_sheet[f"{col}{row}"].value is not None:
                    new_sheet[f"{col}{row}"] = None

        # Processes
        # Thang 2026-05-15: filter ra các quy trình có giá tiền > 0.
        # User mặc định nhận 8 quy trình preloaded (frontend); những quy trình
        # không điền giá sẽ bị bỏ qua khi xuất file để PDF/Excel sạch sẽ.
        raw_processes = item.get("processes") or []
        processes = []
        skipped = 0
        for proc in raw_processes:
            try:
                price = float(proc.get("unit_price") or 0)
                hours = float(proc.get("time_hr") or 0)
            except (TypeError, ValueError):
                price, hours = 0.0, 0.0
            if price > 0 and hours > 0:
                processes.append(proc)
            else:
                skipped += 1
        if skipped:
            logger.info(
                "GC %s: filtered out %d zero-price processes (kept %d)",
                bqms, skipped, len(processes),
            )

        for i, proc in enumerate(processes):
            row = PROC_FIRST_ROW + i
            if row > PROC_LAST_ROW:
                errors.append(
                    f"{bqms}: too many processes ({len(processes)}), "
                    f"only first {PROC_LAST_ROW - PROC_FIRST_ROW + 1} kept"
                )
                break
            new_sheet[f"C{row}"] = proc.get("name") or ""
            new_sheet[f"H{row}"] = float(proc.get("time_hr") or 0)
            new_sheet[f"J{row}"] = float(proc.get("unit_price") or 0)

        # Clear leftover process rows (including those left by filtered-out processes)
        # Thang 2026-05-15: also clear I ("Hr" unit text) + K (amount) so empty
        # rows don't show ghost "Hr" units after filter-by-price drops processes.
        for row in range(PROC_FIRST_ROW + len(processes), PROC_LAST_ROW + 1):
            for col in ("C", "H", "I", "J", "K"):
                if new_sheet[f"{col}{row}"].value is not None:
                    new_sheet[f"{col}{row}"] = None

        # Nego (negative adjustment) at K90
        if item.get("nego"):
            new_sheet["K90"] = float(item["nego"])

        # Product photo: attach image into B93:H98 (merged in template),
        # centered horizontally + vertically per Thang 2026-05-11.
        img_bytes = images_map.get(bqms)
        if img_bytes:
            ok = _attach_image_two_cell(
                new_sheet, img_bytes,
                from_col=2, from_row=93,    # B93
                to_col=8, to_row=98,        # H98
                max_w_px=720, max_h_px=320,  # fit within ~796x349 cell area
            )
            if not ok:
                errors.append(f"{bqms}: product photo attach failed")

        sheets_created += 1

    # Drop the original sample sheet
    if sample_name in wb.sheetnames:
        del wb[sample_name]

    # Reorder: item sheets first, Material&Process last
    desired: list[str] = []
    for it in wizard_items:
        bqms = (it.get("bqms_code") or "").strip()
        for sn in wb.sheetnames:
            if sn == bqms or (bqms and sn.startswith(bqms[:28])):
                if sn not in desired:
                    desired.append(sn)
                break
    for sn in wb.sheetnames:
        if sn not in desired:
            desired.append(sn)
    wb._sheets = [wb[name] for name in desired]

    # Editable + landscape (same setup as template-from-bqms flow)
    from openpyxl.worksheet.page import PageMargins
    for sn in wb.sheetnames:
        try:
            ws = wb[sn]
            ws.protection.disable()
            # Thang 2026-05-20: switched to A4 portrait per user request.
            # fitToWidth=1 ensures wide content (image column + price columns)
            # scales down to fit portrait width instead of getting clipped.
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins = PageMargins(
                left=0.3, right=0.3, top=0.4, bottom=0.4,
                header=0.2, footer=0.2,
            )
        except Exception as exc:
            logger.warning("page setup for %r failed: %s", sn, exc)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    _fix_drawing_rels_paths(output_path)

    # Per Thang 2026-05-11: split the combined workbook → 1 xlsx per item
    # (each contains the item's sheet + Material&Process lookup so formulas
    # still resolve). Filename: <stem>_<bqms_code>.xlsx
    per_item_files: list[dict[str, str]] = []
    out_dir = Path(output_path).parent
    out_stem = Path(output_path).stem    # "QUOTATION_GC_QT26061473"
    item_sheet_names = [n for n in wb.sheetnames if n != "Material&Process"]
    if len(item_sheet_names) >= 1:
        for item_sn in item_sheet_names:
            try:
                wb_one = openpyxl.load_workbook(output_path)
                for n in list(wb_one.sheetnames):
                    if n != item_sn and n != "Material&Process":
                        del wb_one[n]
                # Ensure item sheet is active
                if item_sn in wb_one.sheetnames:
                    wb_one.active = wb_one.sheetnames.index(item_sn)
                # File-safe filename: replace forbidden chars
                safe_bqms = re.sub(r"[^A-Za-z0-9_-]", "_", item_sn)
                one_path = out_dir / f"{out_stem}_{safe_bqms}.xlsx"
                wb_one.save(one_path)
                _fix_drawing_rels_paths(str(one_path))
                per_item_files.append({
                    "bqms_code": item_sn,
                    "xlsx": str(one_path),
                })
            except Exception as exc:
                logger.warning("per-item split failed for %s: %s", item_sn, exc)
                errors.append(f"split {item_sn}: {exc}")

    logger.info(
        "GC wizard quotation built: %d sheets, %d formula rewrites, %d per-item files → %s",
        sheets_created, formula_rewrites, len(per_item_files), output_path,
    )
    return {
        "sheets_created": sheets_created,
        "formula_rewrites": formula_rewrites,
        "per_item_files": per_item_files,
        "errors": errors,
    }


def _fix_drawing_rels_paths(xlsx_path: str) -> int:
    """Rewrite absolute image paths in drawing rels to relative paths.

    openpyxl's copy_worksheet → save produces:
        <Relationship Target="/xl/media/image1.png" .../>
    LibreOffice silently drops images with absolute targets. Standard
    xlsx files (and Excel) use:
        <Relationship Target="../media/image1.png" .../>
    Both are technically valid per OOXML spec, but LibreOffice is strict.

    Returns count of rels rewritten.
    """
    import zipfile, shutil, tempfile
    n = 0
    src = Path(xlsx_path)
    tmp_path = src.with_suffix(".fix.tmp.xlsx")
    try:
        with zipfile.ZipFile(src, "r") as zin:
            with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    # Only rewrite drawing rels files (xl/drawings/_rels/*.xml.rels)
                    if "drawings/_rels" in item.filename and item.filename.endswith(".rels"):
                        original = data
                        # Fix absolute /xl/media/... → ../media/...
                        data = data.replace(
                            b'Target="/xl/media/',
                            b'Target="../media/',
                        )
                        # Same for absolute /xl/drawings/... rare but defensive
                        data = data.replace(
                            b'Target="/xl/drawings/',
                            b'Target="../drawings/',
                        )
                        if data != original:
                            n += 1
                            logger.info("Fixed drawing rel path in %s", item.filename)
                    zout.writestr(item, data)
        shutil.move(str(tmp_path), str(src))
    except Exception as exc:
        logger.warning("_fix_drawing_rels_paths failed: %s", exc)
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass
    return n


async def _generate_per_item_pdfs(
    xlsx_path: str,
    items: list[dict[str, Any]],
    output_dir: Path,
    rfq_no: str,
    round_n: int,
) -> list[dict[str, str]]:
    """Split multi-sheet xlsx → N PDFs (1 per item) — Thang 2026-05-19.

    Naming: `{rfq_no}_{bqms_code}_AMABACNINH_L{round_n}.pdf`

    Approach: load main xlsx → del all sheets except target → save tmp →
    Gotenberg/LibreOffice convert → cleanup tmp. Main xlsx + main PDF untouched.
    """
    from openpyxl import load_workbook
    from app.services.gotenberg_service import convert_xlsx_to_pdf

    out_files: list[dict[str, str]] = []
    tmp_dir = output_dir / "_tmp_per_item"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    for it in items:
        bqms_code = (it.get("bqms_code") or it.get("item_code")
                     or it.get("bqms") or "").strip()
        if not bqms_code:
            continue
        try:
            wb = load_workbook(xlsx_path)
            # Find sheet matching this item (sheet name == bqms_code, sanitized)
            target = None
            for sheet_name in wb.sheetnames:
                clean = _SHEET_FORBIDDEN_RE.sub("-", bqms_code)[:31]
                if sheet_name == clean or sheet_name == bqms_code:
                    target = sheet_name
                    break
            if target is None:
                # Fallback: any sheet matching first 20 chars
                for sheet_name in wb.sheetnames:
                    if bqms_code[:20] in sheet_name:
                        target = sheet_name
                        break
            if target is None:
                logger.warning("Per-item PDF: no sheet found for %s in %s",
                               bqms_code, xlsx_path)
                wb.close()
                continue

            # Remove all OTHER sheets — keep target only
            for sheet_name in list(wb.sheetnames):
                if sheet_name != target:
                    del wb[sheet_name]

            safe_code = bqms_code.replace("/", "-").replace("\\", "-")[:50]
            tmp_xlsx = tmp_dir / f"{safe_code}.xlsx"
            wb.save(str(tmp_xlsx))
            wb.close()

            pdf_name = f"{rfq_no}_{safe_code}_AMABACNINH_L{round_n}.pdf"
            pdf_path = output_dir / pdf_name
            await convert_xlsx_to_pdf(str(tmp_xlsx), str(pdf_path))
            out_files.append({
                "type": "gc_item_pdf",
                "path": str(pdf_path),
                "bqms_code": bqms_code,
            })
            try:
                tmp_xlsx.unlink()
            except Exception:
                pass
        except Exception as exc:
            logger.warning("Per-item PDF failed for %s: %s", bqms_code, exc)

    # Cleanup tmp dir
    try:
        if tmp_dir.exists() and not any(tmp_dir.iterdir()):
            tmp_dir.rmdir()
    except Exception:
        pass

    return out_files


async def run_gc_quote_round1(
    rfq_no: str,
    items: list[dict[str, Any]],
    images_map: dict[str, bytes],
    output_dir: Path,
    gc_template_path: str = "/data/files/templates/QUOTATION_GC.xlsx",
    round_n: int = 1,
    quote_date=None,
) -> dict[str, Any]:
    """Build GC quotation xlsx + PDFs for round N into output_dir.

    NAMING (Thang 2026-05-19):
      - Main Excel:    `{rfq}_AMABACNINH_L{round}.xlsx` (all sheets)
      - Main PDF:      `{rfq}_AMABACNINH_L{round}.pdf` (all sheets)
      - Per-item PDF:  `{rfq}_{bqms_code}_AMABACNINH_L{round}.pdf` (1 sheet only)

    Returns: {success, files: [{type, path}, ...], sheets_created, ...}
    """
    files: list[dict[str, str]] = []
    output_dir.mkdir(parents=True, exist_ok=True)

    # Quotation XLSX filename = parent folder name + .xlsx (matches PDF
    # naming convention). User requirement 2026-06-04: copy exactly to make
    # it easy to identify the round (L1/L2/L3/L4). Happy path produces e.g.
    # `QT26071059_AMABACNINH_L1.xlsx/.pdf` (same as before); fallback folder
    # shape (no AMABACNINH suffix) is honored automatically. Keeps TM + GC
    # symmetric — both flows produce `<folder>.xlsx/.pdf`.
    base_name = output_dir.name
    xlsx_path = str(output_dir / f"{base_name}.xlsx")
    fill_result = fill_gc_quotation_from_template(
        template_path=gc_template_path,
        items=items, images_map=images_map,
        rfq_no=rfq_no, output_path=xlsx_path,
        quote_date=quote_date,
    )
    files.append({"type": "gc_quotation_xlsx", "path": xlsx_path})

    # Main PDF (all sheets in 1 file)
    pdf_path = str(output_dir / f"{base_name}.pdf")
    try:
        from app.services.gotenberg_service import convert_xlsx_to_pdf
        await convert_xlsx_to_pdf(xlsx_path, pdf_path)
        files.append({"type": "gc_quotation_pdf", "path": pdf_path})
    except Exception as exc:
        logger.warning("GC main PDF conversion failed: %s", exc)
        fill_result.setdefault("errors", []).append(f"pdf: {exc}")

    # Per-item PDFs (1 PDF per BQMS code)
    try:
        per_item = await _generate_per_item_pdfs(
            xlsx_path, items, output_dir, rfq_no, round_n,
        )
        files.extend(per_item)
        logger.info("GC: generated %d per-item PDFs (round=%d)",
                    len(per_item), round_n)
    except Exception as exc:
        logger.warning("GC per-item PDFs failed: %s", exc)
        fill_result.setdefault("errors", []).append(f"per_item_pdfs: {exc}")

    return {
        "success": fill_result["sheets_created"] > 0,
        "files": files,
        **fill_result,
    }
