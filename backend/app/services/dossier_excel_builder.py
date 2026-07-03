"""Build delivery dossier Excel from template AMA_DELIVERY_DOSSIER.xlsx.

Template has 6 sheets:
  - Tổng hợp                  (static checklist; no per-job fill)
  - packing list              (1 row per PO×item)
  - Cam kết hình ảnh (1)      (per-PO; clone N times)
  - Cam kết hình ảnh (2)      (per-PO; clone N times — sample has 2)
  - List Detail               (per-bqms_code rows)
  - label                     (per-PO label block)

Per Thang 2026-05-16:
  - Clone "Cam kết hình ảnh (1)" sheet N times for N PO numbers
    (delete extra sheets, rename "(1)..(N)")
  - Each Cam kết sheet has 2 image anchors: "Hệ thống" (left) + "Thực tế" (right)
  - Skip "Thực tế" image if user didn't upload
  - Save as .xlsx

Pattern mirrors gc_template_quotation.py:
  - _attach_image_two_cell (centered, fit-contain, anchor with EMU offsets)
  - openpyxl wb.copy_worksheet for clone
"""
from __future__ import annotations

import io
import json
import logging
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

logger = logging.getLogger(__name__)

TEMPLATE_PATH = "/data/files/templates/AMA_DELIVERY_DOSSIER.xlsx"
CELLS_JSON_PATH = "/data/files/templates/AMA_DELIVERY_DOSSIER.cells.json"

# Sheet name patterns (template sample uses these exact names with diacritics)
SHEET_TONG_HOP = "Tổng hợp"
SHEET_PACKING = "packing list"
SHEET_CAM_KET_RE = re.compile(r"^Cam kết hình ảnh\s*\(\d+\)$")
SHEET_LIST_DETAIL = "List Detail"
SHEET_LABEL = "label"


# ---------------------------------------------------------------------------
# Image embed (port of gc_template_quotation._attach_image_two_cell)
# ---------------------------------------------------------------------------

def _range_pixel_size(ws, from_col: int, from_row: int, to_col: int, to_row: int) -> tuple[int, int]:
    """Approximate (width_px, height_px) of a cell range."""
    width_units = 0.0
    for c in range(from_col, to_col + 1):
        w = ws.column_dimensions[get_column_letter(c)].width
        width_units += (w if w and w > 0 else 8.43)
    height_pts = 0.0
    for r in range(from_row, to_row + 1):
        h = ws.row_dimensions[r].height
        height_pts += (h if h and h > 0 else 15.0)
    return max(40, int(width_units * 7.0)), max(30, int(height_pts * 1.333))


def _attach_image_centered(
    ws,
    img_bytes: bytes,
    *,
    from_col: int,
    from_row: int,
    to_col: int,
    to_row: int,
    fill_ratio: float = 0.96,
) -> bool:
    """Embed image FIT-CONTAIN (giữ tỷ lệ, không méo) + CĂN GIỮA trong ô/range.

    fill_ratio=0.96 → ảnh lấp ~96% chiều giới hạn của ô (chừa lề mỏng), giữ nguyên
    tỷ lệ ảnh; offset EMU căn giữa cả ngang lẫn dọc trong ô."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    try:
        from PIL import Image as PILImage
        pil = PILImage.open(io.BytesIO(img_bytes))
        if pil.mode not in ("RGB", "RGBA"):
            pil = pil.convert("RGBA")

        range_w, range_h = _range_pixel_size(ws, from_col, from_row, to_col, to_row)
        target_w = max(40, int(range_w * fill_ratio))
        target_h = max(40, int(range_h * fill_ratio))
        orig_w, orig_h = pil.size
        scale = min(target_w / orig_w, target_h / orig_h, 4.0)
        new_w = max(40, int(orig_w * scale))
        new_h = max(40, int(orig_h * scale))
        if (new_w, new_h) != (orig_w, orig_h):
            # Pillow 12: legacy Image.LANCZOS is kept as an alias, but the
            # Resampling enum is the documented API and won't trigger
            # DeprecationWarnings.
            _lanczos = getattr(PILImage, "Resampling", PILImage).LANCZOS
            pil = pil.resize((new_w, new_h), _lanczos)

        buf = io.BytesIO()
        pil.save(buf, format="PNG")
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.width = new_w
        xl_img.height = new_h
        off_x = max(0, (range_w - new_w) // 2)
        off_y = max(0, (range_h - new_h) // 2)
        marker = AnchorMarker(
            col=from_col - 1, colOff=pixels_to_EMU(off_x),
            row=from_row - 1, rowOff=pixels_to_EMU(off_y),
        )
        ext = XDRPositiveSize2D(cx=pixels_to_EMU(new_w), cy=pixels_to_EMU(new_h))
        xl_img.anchor = OneCellAnchor(_from=marker, ext=ext)
        ws._images.append(xl_img)
        return True
    except Exception as exc:
        logger.warning("attach image failed: %s", exc)
        return False


# ---------------------------------------------------------------------------
# Cell address helpers
# ---------------------------------------------------------------------------

def _parse_anchor(anchor: dict) -> tuple[int, int, int, int]:
    """Parse {from: 'C11', to: 'F12'} → (from_col, from_row, to_col, to_row)."""
    fc, fr = coordinate_from_string(anchor["from"])
    tc, tr = coordinate_from_string(anchor["to"])
    return column_index_from_string(fc), fr, column_index_from_string(tc), tr


def _anchor_cell(addr: str) -> tuple[int, int]:
    """Parse a single-cell anchor like 'C12' → (col_idx, row_idx) (1-based)."""
    col, row = coordinate_from_string(addr)
    return column_index_from_string(col), row


def _load_cells_map(path: str = CELLS_JSON_PATH) -> dict:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"cells.json not found: {path} — run scripts/seed_dossier_template.py first")
    return json.loads(p.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

class DossierItem:
    """One row in the dossier — represents (PO, item) tuple."""

    def __init__(
        self, *,
        po_number: str,
        po_seq: str,
        bqms_code: str,
        item_name: str,
        specification: str,
        unit: str,
        shipping_qty: float,
        dept: str = "MAIN",
        pr_person: str = "",
        receiver: str = "",
        box_weight: float | None = None,
        dim_l: str = "",
        dim_w: str = "",
        dim_h: str = "",
        packing_size: str = "",
        box_qty: float | None = None,
        system_image: bytes | None = None,
        actual_image: bytes | None = None,
    ):
        self.po_number = po_number
        self.po_seq = po_seq
        self.bqms_code = bqms_code
        self.item_name = item_name
        self.specification = specification
        self.unit = unit or "PC"
        self.shipping_qty = shipping_qty
        self.dept = dept
        self.pr_person = pr_person
        self.receiver = receiver
        self.box_weight = box_weight
        self.dim_l = dim_l
        self.dim_w = dim_w
        self.dim_h = dim_h
        # Box Qty (col O) + Packing Size (col N) — MANUAL per-item inputs.
        self.packing_size = packing_size
        # box_qty: blank stays blank (NO default-to-1).
        self.box_qty = box_qty if box_qty not in (None, "") else None
        self.system_image = system_image
        self.actual_image = actual_image


def build_dossier_workbook(
    *,
    items: list[DossierItem],
    customer: str,                  # "SEV" or "SEVT"
    vendor_name: str = "AMA BẮC NINH JSC",
    shipping_no: str = "",
    invoice_no: str = "",
    shipping_date: str = "",        # "DD/MM/YYYY"
    output_path: Path | str,
    template_path: str = TEMPLATE_PATH,
    cells_map_path: str = CELLS_JSON_PATH,
    box_qty_total_override: int | None = None,
    labels: list | None = None,
) -> dict:
    """Build the dossier .xlsx from template + items + images.

    Returns: {success, sheets, images_embedded, warnings: [...]}
    """
    out: dict[str, Any] = {
        "success": False,
        "sheets": [],
        "images_embedded": 0,
        "warnings": [],
        "output_path": str(output_path),
    }

    if not items:
        out["warnings"].append("No items provided")
        return out

    cmap = _load_cells_map(cells_map_path)
    field_map = cmap["field_map"]

    wb = openpyxl.load_workbook(template_path)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Group items by PO
    po_order: list[str] = []
    items_by_po: dict[str, list[DossierItem]] = {}
    for it in items:
        if it.po_number not in items_by_po:
            items_by_po[it.po_number] = []
            po_order.append(it.po_number)
        items_by_po[it.po_number].append(it)

    # -----------------------------------------------------------------------
    # SHEET: packing list
    # -----------------------------------------------------------------------
    if SHEET_PACKING in wb.sheetnames:
        ws = wb[SHEET_PACKING]
        pl = field_map["packing_list"]
        # Update vendor/customer header
        ws[pl["vendor_name_cell"]] = f"VENDOR NAME: {vendor_name}"
        ws[pl["customer_cell"]] = f"CUSTOMER: {customer}"
        first_row = pl["first_data_row"]
        cols = pl["cols"]
        # Write items — incl. packing_size (N), box_qty (O), qty (P)
        for i, it in enumerate(items):
            r = first_row + i
            ws[f"{cols['stt']}{r}"] = i + 1
            ws[f"{cols['dept']}{r}"] = it.dept
            ws[f"{cols['pr_person']}{r}"] = it.pr_person
            ws[f"{cols['po_no']}{r}"] = it.po_number
            ws[f"{cols['bqms_code']}{r}"] = it.bqms_code
            ws[f"{cols['item_name']}{r}"] = it.item_name
            ws[f"{cols['specification']}{r}"] = it.specification
            ws[f"{cols['unit']}{r}"] = it.unit
            if it.box_weight is not None:
                ws[f"{cols['box_weight']}{r}"] = it.box_weight
            ws[f"{cols['dim_l']}{r}"] = it.dim_l
            ws[f"{cols['dim_w']}{r}"] = it.dim_w
            ws[f"{cols['dim_h']}{r}"] = it.dim_h
            # Manual per-item: Packing Size (N) + Box Qty (O); Qty shipped (P)
            ws[f"{cols['packing_size']}{r}"] = it.packing_size
            if it.box_qty is not None:
                ws[f"{cols['box_qty']}{r}"] = it.box_qty
            ws[f"{cols['qty']}{r}"] = it.shipping_qty
        # TOTAL row — "TOTAL" label at total_label_col (E), SUM of total_sum_cols
        total_row = first_row + len(items)
        ws[f"{pl['total_label_col']}{total_row}"] = "TOTAL"
        for sum_col in pl["total_sum_cols"]:
            # Map sum column letter back to which item attribute it sums.
            if sum_col == cols["box_weight"]:
                tot = sum((it.box_weight or 0) for it in items)
            elif sum_col == cols["box_qty"]:
                # PRINT-ONLY override (Thang LOCKED): user-edited TOTAL Box Qty
                # replaces the computed sum WITHOUT rescaling per-row box_qty.
                # null/None → fall back to the computed sum.
                if box_qty_total_override is not None:
                    tot = box_qty_total_override
                else:
                    tot = sum((it.box_qty or 0) for it in items)
            elif sum_col == cols["qty"]:
                tot = sum((it.shipping_qty or 0) for it in items)
            else:
                continue
            ws[f"{sum_col}{total_row}"] = tot
        # Clear any leftover sample data after total_row
        for r in range(total_row + 1, ws.max_row + 1):
            for col_letter in cols.values():
                cell = ws[f"{col_letter}{r}"]
                if cell.value is not None:
                    cell.value = None

    # -----------------------------------------------------------------------
    # SHEET: List Detail
    # -----------------------------------------------------------------------
    if SHEET_LIST_DETAIL in wb.sheetnames:
        ws = wb[SHEET_LIST_DETAIL]
        ld = field_map["list_detail"]
        first_row = ld["first_data_row"]
        cols = ld["cols"]
        img_col = column_index_from_string(cols["image"])
        # Aggregate by bqms_code (sum qty across PO); keep first system image.
        agg: dict[str, dict] = {}
        for it in items:
            if it.bqms_code not in agg:
                # Issue 7 (Thang LOCKED): "{item_name}/ {specification}" — no
                # leading space, one space after the slash. If either part is
                # empty, strip the dangling '/ ' (or trailing '/').
                _name = (it.item_name or "").strip()
                _spec = (it.specification or "").strip()
                name_specs = f"{_name}/ {_spec}".strip().strip("/ ").strip()
                agg[it.bqms_code] = {
                    "name_specs": name_specs,
                    "unit": it.unit,
                    "qty": 0,
                    "system_image": it.system_image,
                }
            agg[it.bqms_code]["qty"] += it.shipping_qty
            if not agg[it.bqms_code]["system_image"] and it.system_image:
                agg[it.bqms_code]["system_image"] = it.system_image
        codes = list(agg.keys())
        for i, code in enumerate(codes):
            r = first_row + i
            d = agg[code]
            ws[f"{cols['no']}{r}"] = i + 1
            ws[f"{cols['item_code']}{r}"] = code
            ws[f"{cols['name_specs']}{r}"] = d["name_specs"]
            ws[f"{cols['unit']}{r}"] = d["unit"]
            ws[f"{cols['qty']}{r}"] = d["qty"]
            # Embed system image at column G of this data row.
            if d["system_image"]:
                if _attach_image_centered(
                    ws, d["system_image"],
                    from_col=img_col, from_row=r, to_col=img_col, to_row=r,
                ):
                    out["images_embedded"] += 1
                else:
                    out["warnings"].append(f"List Detail: image embed failed for {code}")
        # Total row — "Total" label at total_label_col (B), SUM at total_sum_col (I)
        total_row = first_row + len(codes)
        ws[f"{ld['total_label_col']}{total_row}"] = "Total"
        ws[f"{ld['total_sum_col']}{total_row}"] = sum(d["qty"] for d in agg.values())
        # Clear leftover sample rows
        for r in range(total_row + 1, ws.max_row + 1):
            for col_letter in cols.values():
                cell = ws[f"{col_letter}{r}"]
                if cell.value is not None:
                    cell.value = None

    # -----------------------------------------------------------------------
    # SHEET: label  (block per PO, block_height=6, first_block_row=1)
    # -----------------------------------------------------------------------
    if SHEET_LABEL in wb.sheetnames:
        ws = wb[SHEET_LABEL]
        lb = field_map["label"]
        block_h = lb["block_height"]
        first_block = lb["first_block_row"]
        lcells = lb["cells"]  # e.g. {"vendor":"C+0","pr_person":"C+1",...}

        def _label_addr(spec: str, r0: int) -> str:
            """'C+4' + block start r0 → 'C{r0+4}'. Does NOT touch static labels."""
            col, _, off = spec.partition("+")
            return f"{col}{r0 + int(off or 0)}"

        def _write_label_block(r0: int, *, pr_person, po_no, bqms_code, qty):
            ws[_label_addr(lcells["vendor"], r0)] = vendor_name
            ws[_label_addr(lcells["pr_person"], r0)] = pr_person
            ws[_label_addr(lcells["po_no"], r0)] = po_no
            ws[_label_addr(lcells["bqms_code"], r0)] = bqms_code
            ws[_label_addr(lcells["qty"], r0)] = qty

        if labels:
            # ---- EDITABLE LABEL ARRAY (user touched the Label tab) -----------
            # One block PER label entry (duplicates allowed). qty is the EDITED
            # value, NOT a computed PO sum.
            n_labels = len(labels)
            # Template capacity = number of pre-formatted blocks. Measured at
            # build time from the loaded sheet so it tracks the real template:
            #   capacity = (last_block_start_row - first_block_row)//block_h + 1
            capacity = max(0, (ws.max_row - first_block) // block_h + 1)

            if n_labels > capacity:
                # CLONE strategy: replicate the FIRST block's formatting (cell
                # styles + static label text + row heights) downward so the
                # extra duplicated labels stay fully formatted. Merged cells in
                # the label block (none in the sample template, but handled
                # defensively) are re-created at the new offset.
                import copy as _copy
                base_r0 = first_block
                base_merges = [
                    rng for rng in list(ws.merged_cells.ranges)
                    if base_r0 <= rng.min_row < base_r0 + block_h
                ]
                for blk_idx in range(capacity, n_labels):
                    r0 = first_block + blk_idx * block_h
                    for off in range(block_h):
                        src_r = base_r0 + off
                        dst_r = r0 + off
                        # Row height
                        sh = ws.row_dimensions[src_r].height
                        if sh is not None:
                            ws.row_dimensions[dst_r].height = sh
                        # Copy each used cell's value (static labels) + style
                        for col_idx in range(1, ws.max_column + 1):
                            src = ws.cell(row=src_r, column=col_idx)
                            dst = ws.cell(row=dst_r, column=col_idx)
                            dst.value = src.value
                            if src.has_style:
                                dst.font = _copy.copy(src.font)
                                dst.fill = _copy.copy(src.fill)
                                dst.border = _copy.copy(src.border)
                                dst.alignment = _copy.copy(src.alignment)
                                dst.number_format = src.number_format
                                dst.protection = _copy.copy(src.protection)
                    # Re-create merged ranges at the new block offset
                    for rng in base_merges:
                        d = (r0 - base_r0)
                        ws.merge_cells(
                            start_row=rng.min_row + d, start_column=rng.min_col,
                            end_row=rng.max_row + d, end_column=rng.max_col,
                        )

            # Fill one block per label (now guaranteed enough blocks exist).
            for blk_idx, entry in enumerate(labels):
                r0 = first_block + blk_idx * block_h
                _write_label_block(
                    r0,
                    pr_person=entry.get("pr_person", ""),
                    po_no=entry.get("po_number", ""),
                    bqms_code=entry.get("bqms_code", ""),
                    qty=entry.get("qty"),
                )
            # Clear leftover template blocks beyond the labels written.
            for blk_idx in range(n_labels, capacity):
                r0 = first_block + blk_idx * block_h
                if r0 > ws.max_row:
                    break
                for spec in lcells.values():
                    cell = ws[_label_addr(spec, r0)]
                    if cell.value is not None:
                        cell.value = None
        else:
            # ---- LEGACY per-PO derivation (Label tab untouched) --------------
            # For each PO, write 1 block (data offsets read from field_map).
            for blk_idx, po in enumerate(po_order):
                r0 = first_block + blk_idx * block_h
                po_items = items_by_po[po]
                first_it = po_items[0]
                _write_label_block(
                    r0,
                    pr_person=first_it.pr_person,
                    po_no=po,
                    bqms_code=first_it.bqms_code,
                    qty=sum(it.shipping_qty for it in po_items),
                )
            # Clear only the *data* cells of leftover template blocks (preserve
            # the static "BQMS Code:" / "Qty:" labels in those blocks).
            for blk_idx in range(len(po_order), len(po_order) + 8):
                r0 = first_block + blk_idx * block_h
                if r0 > ws.max_row:
                    break
                for spec in lcells.values():
                    cell = ws[_label_addr(spec, r0)]
                    if cell.value is not None:
                        cell.value = None

    # -----------------------------------------------------------------------
    # SHEETS: Cam kết hình ảnh (N) — clone PER ITEM (per bqms_code)
    # -----------------------------------------------------------------------
    # P6 (Thang LOCKED 2026-06-25): one Cam kết sheet PER DossierItem, not per
    # PO. A multi-item PO must produce one commitment sheet per item — each with
    # its own bqms_code / item_name / specification / this-item qty + its own
    # system & actual images. dept/pr_person/receiver are carried per-item.
    # Find the template Cam kết sheet (use "(1)" as base)
    cam_ket_template_name = field_map.get("cam_ket_template_sheet", "Cam kết hình ảnh (1)")
    if cam_ket_template_name not in wb.sheetnames:
        # Fallback: search by regex
        for sn in wb.sheetnames:
            if SHEET_CAM_KET_RE.match(sn):
                cam_ket_template_name = sn
                break

    if cam_ket_template_name in wb.sheetnames:
        template_ws = wb[cam_ket_template_name]
        ck = field_map["cam_ket"]
        n_items = len(items)

        # Strategy: rewrite existing "(1)..(K)" sheets to N items. Delete extras, clone if needed.
        existing_camket = [sn for sn in wb.sheetnames if SHEET_CAM_KET_RE.match(sn)]

        # First, ensure we have exactly n_items camket sheets
        # Clone from template if not enough
        while len(existing_camket) < n_items:
            new_idx = len(existing_camket) + 1
            cloned = wb.copy_worksheet(template_ws)
            cloned.title = f"Cam kết hình ảnh ({new_idx})"
            existing_camket.append(cloned.title)
        # Delete extras
        while len(existing_camket) > n_items:
            extra = existing_camket.pop()
            del wb[extra]

        # Now fill each — one sheet per item (original items order)
        for sheet_idx, it in enumerate(items):
            sheet_name = existing_camket[sheet_idx]
            ws = wb[sheet_name]

            ws[ck["customer"]] = customer
            ws[ck["vendor_name"]] = vendor_name
            ws[ck["department"]] = it.dept
            ws[ck["pr_pic"]] = it.pr_person
            ws[ck["receiver"]] = it.receiver
            # shipping_no → C13 — Part-2 → Part-1 hand-off (do NOT change cell)
            ws[ck["shipping_no"]] = shipping_no
            ws[ck["po_no"]] = it.po_number
            ws[ck["bqms_code"]] = it.bqms_code
            ws[ck["item_name"]] = it.item_name
            ws[ck["specification"]] = it.specification
            ws[ck["quantity"]] = it.shipping_qty  # THIS item's qty, not PO total
            ws[ck["shipping_date"]] = shipping_date

            # Embed images: "Hệ thống" at C12 anchor, "Thực tế" at D12 anchor.
            if it.system_image:
                sc, sr = _anchor_cell(ck["img_system_anchor"])
                if _attach_image_centered(ws, it.system_image,
                                          from_col=sc, from_row=sr, to_col=sc, to_row=sr):
                    out["images_embedded"] += 1
                else:
                    out["warnings"].append(
                        f"Item {it.bqms_code} (PO {it.po_number}): system image embed failed")
            else:
                out["warnings"].append(
                    f"Item {it.bqms_code} (PO {it.po_number}): no system image")

            if it.actual_image:
                ac, ar = _anchor_cell(ck["img_actual_anchor"])
                if _attach_image_centered(ws, it.actual_image,
                                          from_col=ac, from_row=ar, to_col=ac, to_row=ar):
                    out["images_embedded"] += 1
                else:
                    out["warnings"].append(
                        f"Item {it.bqms_code} (PO {it.po_number}): actual image embed failed")
            else:
                out["warnings"].append(
                    f"Item {it.bqms_code} (PO {it.po_number}): no actual image (user didn't upload)")

        # Thang 2026-06-24: template Cam kết GỐC (không hậu tố "(N)") chỉ là nguồn
        # clone, không có data → xoá để khỏi dư 1 sheet trống. Rồi đổi tên clone
        # đầu "(1)" → "Cam kết hình ảnh" cho khớp đúng tên sheet trong form mẫu
        # (mẫu: "Cam kết hình ảnh" · "(2)" · "(3)").
        if (cam_ket_template_name in wb.sheetnames
                and cam_ket_template_name not in existing_camket):
            del wb[cam_ket_template_name]
        if (existing_camket and existing_camket[0] == "Cam kết hình ảnh (1)"
                and "Cam kết hình ảnh" not in wb.sheetnames):
            wb["Cam kết hình ảnh (1)"].title = "Cam kết hình ảnh"

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    wb.save(output_path)
    out["sheets"] = list(wb.sheetnames)
    out["success"] = True
    logger.info(
        "Dossier workbook built: %s (sheets=%s images=%d warnings=%d)",
        output_path, out["sheets"], out["images_embedded"], len(out["warnings"]),
    )
    return out
