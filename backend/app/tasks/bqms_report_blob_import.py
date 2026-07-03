"""BQMS daily report blob auto-import — chỉ cột `report` của bqms_rfq.

Thang 2026-06-01: bqms_rfq auto-import đã bị disable từ 2026-05-18 để bảo vệ
ERP-side edits (V1 prices, scenarios). Nhưng cột `report` (Excel col S — "Báo cáo
DD/MM/YYYY ..." blob) là dữ liệu user input thủ công trên Excel, không bao giờ
được edit từ ERP → an toàn để auto-sync.

Task này chỉ UPDATE cột `report` (+ `updated_at`), không động bất kỳ cột nào
khác. Chart "Xu hướng số mã yêu cầu" ở /reports/daily phụ thuộc cột này để hiển
30 ngày báo cáo gần nhất.

Chạy mỗi 5 phút. force=True bỏ qua mtime check.
"""

from __future__ import annotations

import logging
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import psycopg2
import psycopg2.extras

from app.core.procrastinate_app import app, SYNC_DSN

logger = logging.getLogger(__name__)

EXCEL_PATH = Path(
    "/data/onedrive-staging/Puplic/BQMS/Thong ke hoi hang BQMS.xlsx"
)
SHEET_NAME = "TONG HOP BQMS"
COL_RFQ = 2          # column C: RFQ No.
COL_BQMS = 3         # column D: BQMS code
COL_REPORT = 18      # column S: Báo cáo


@app.periodic(cron="*/5 * * * *")
@app.task(name="bqms_report_blob_import", queue="etl")
def bqms_report_blob_import(timestamp: int = 0, force: bool = False) -> dict[str, Any]:
    """Update bqms_rfq.report from Excel col 18. Touch nothing else."""
    started = datetime.now(timezone.utc)
    t0 = time.monotonic()
    logger.info("bqms_report_blob_import: starting (force=%s)", force)

    out: dict[str, Any] = {"updated": 0, "scanned": 0, "errors": [], "mtime_skipped": False}

    if not EXCEL_PATH.exists():
        out["errors"].append(f"Excel missing: {EXCEL_PATH}")
        return out

    # Check mtime vs last successful run (unless force=True)
    file_mtime = datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime, tz=timezone.utc)
    if not force:
        try:
            with psycopg2.connect(SYNC_DSN) as conn, conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT MAX(completed_at) FROM etl_sync_log
                     WHERE sync_type = 'report_blob_import' AND status = 'success'
                    """
                )
                row = cur.fetchone()
                last_run = row[0] if row and row[0] else None
            if last_run and last_run >= file_mtime:
                logger.info("bqms_report_blob_import: SKIPPED (file unchanged since %s)", last_run)
                out["mtime_skipped"] = True
                return out
        except Exception as exc:
            logger.warning("mtime check failed: %s", exc)

    # Read Excel
    try:
        import openpyxl
    except ImportError:
        out["errors"].append("openpyxl missing")
        return out

    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            out["errors"].append(f"sheet missing: {SHEET_NAME}")
            wb.close()
            return out
        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        out["errors"].append(f"read excel failed: {exc}")
        return out

    # Build update batch: (rfq_no, bqms_code) -> report blob (only when blob looks like "Báo cáo ...")
    updates: list[tuple[str, str, str]] = []
    for r in rows[1:]:  # skip header
        if not r or len(r) <= max(COL_RFQ, COL_BQMS, COL_REPORT):
            continue
        rfq_no = r[COL_RFQ]
        bqms = r[COL_BQMS]
        blob = r[COL_REPORT]
        if not blob or not rfq_no:
            continue
        blob_str = str(blob).strip()
        if not blob_str.lower().startswith("báo cáo") and not blob_str.lower().startswith("bao cao"):
            continue
        updates.append((str(rfq_no).strip(), str(bqms or "").strip(), blob_str))

    out["scanned"] = len(updates)

    if not updates:
        logger.info("bqms_report_blob_import: no report blobs found")
        _log_sync(file_mtime, success=True, updated=0)
        return out

    # UPSERT into bqms_rfq.report ONLY (preserve all other columns)
    conn = psycopg2.connect(SYNC_DSN)
    try:
        with conn.cursor() as cur:
            for rfq_no, bqms, blob in updates:
                try:
                    cur.execute(
                        """
                        UPDATE bqms_rfq
                           SET report = %s,
                               updated_at = NOW()
                         WHERE rfq_number = %s
                           AND (bqms_code = %s OR (%s = '' AND bqms_code IS NULL))
                           AND (report IS NULL OR report <> %s)
                        """,
                        (blob, rfq_no, bqms, bqms, blob),
                    )
                    out["updated"] += cur.rowcount
                except Exception as exc:
                    out["errors"].append(f"{rfq_no}/{bqms}: {str(exc)[:200]}")
        conn.commit()
    finally:
        conn.close()

    duration = round(time.monotonic() - t0, 2)
    logger.info(
        "bqms_report_blob_import: done updated=%d scanned=%d errors=%d in %ss",
        out["updated"], out["scanned"], len(out["errors"]), duration,
    )
    _log_sync(file_mtime, success=True, updated=out["updated"])
    return out


def _log_sync(file_mtime: datetime, success: bool, updated: int) -> None:
    """Record run in etl_sync_log so mtime check can dedup next run."""
    try:
        with psycopg2.connect(SYNC_DSN) as conn, conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO etl_sync_log
                    (sync_type, status, started_at, completed_at, rows_inserted, source_file)
                VALUES
                    ('report_blob_import', %s, NOW(), NOW(), %s, %s)
                """,
                ("success" if success else "error", updated, str(EXCEL_PATH)),
            )
            conn.commit()
    except Exception as exc:
        logger.warning("_log_sync failed: %s", exc)
