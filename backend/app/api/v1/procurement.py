"""Procurement Management — ERP admin endpoints for managing vendor bidding.

Đợt 1 rebuild (Thang 2026-06-18):
  - Mời NCC qua TÀI KHOẢN ĐĂNG NHẬP (vendor_accounts), KHÔNG còn magic-link token.
    `POST /batches/{id}/invite` tạo/upsert procurement_rfq_invitations + gửi email
    chứa LINK ĐĂNG NHẬP cổng NCC (ncc.songchau.vn/login?next=/batches/{id}).
  - `GET /batches/{id}/matrix` — ma trận so sánh: item (hàng) × NCC được mời (cột),
    mỗi ô = báo giá MỚI NHẤT của NCC đó cho item đó; highlight giá thấp nhất.
  - `GET /procurement/stats` — KPI strip (IMV-style) cho dashboard admin.
  - Invitations + vendor_quotes giờ round-aware (round_number).
  - Status columns vẫn là TEXT + CHECK (KHÔNG convert enum — để Đợt 2).
  - Magic-link cũ (procurement_bid_tokens) được GIỮ cho dữ liệu lịch sử nhưng
    KHÔNG còn là đường mời chính. Public bid endpoints xem ở public_bid.py.
"""

from __future__ import annotations

import json as _json
import logging
import os
import re
import secrets
import time as _time
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import asyncpg
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, RedirectResponse

from app.core.database import get_db
from app.core.rbac import require_role
from app.core.security import TokenData
from app.core.config import settings
from app.utils.email_sender import send_email
from app.services import procurement_docs
from app.services.procurement_notifications import dispatch_procurement_event
from app.services.sourcing_pricing_engine import fetch_fx_to_vnd

logger = logging.getLogger(__name__)
router = APIRouter()

# Roles — write actions (mutations) vs read-only (adds 'staff' + 'viewer' via rbac).
_WRITE_ROLES = ("admin", "manager", "procurement")
_READ_ROLES = ("admin", "manager", "procurement", "staff")


# ---------------------------------------------------------------------------
# Đợt 4 — FX normalize ở khâu QUYẾT chốt thầu (ADDITIVE, KHÔNG đổi số gốc/rank).
#   Quy đổi VND SONG SONG để admin so sánh nhanh USD-bid vs VND-bid. As-of =
#   bid_deadline (lookup exchange_rates effective ≤ deadline; fallback latest).
#   Tỷ giá tính 1 LẦN / loại tiền distinct / batch (cả batch chung 1 as-of) rồi
#   nhân ở Python — KHÔNG LATERAL, KHÔNG migration, KHÔNG cột DB. VND→1 (vô hại);
#   thiếu rate → None + cờ fx_missing (KHÔNG bịa rate, KHÔNG nhân 1 âm thầm).
# ---------------------------------------------------------------------------
async def _fx_map_for_batch(
    conn: asyncpg.Connection,
    currencies: set[str],
    as_of: date | None,
) -> dict[str, Decimal | None]:
    """Resolve VND-rate AS-OF `as_of` cho mỗi loại tiền distinct của batch.

    1 lần/loại tiền (không per-cell). rate=None ⇒ thiếu tỷ giá (KHÔNG bịa).
    VND short-circuit về 1 trong helper. Caller nhân ở Python (additive).
    Currency lạ ngoài enum → DB raise → helper trả None → đúng "thiếu rate".
    """
    out: dict[str, Decimal | None] = {}
    for cur in currencies:
        if not cur:
            continue
        out[cur] = await fetch_fx_to_vnd(conn, cur, as_of)
    return out


def _vnd(amount: float | None, rate: Decimal | None) -> float | None:
    """Nhân amount (per-currency gốc) với rate VND. None nếu thiếu một trong hai."""
    if amount is None or rate is None:
        return None
    return float(Decimal(str(amount)) * rate)


def _gen_bid_token() -> str:
    """32-char urlsafe token (~190 bits entropy)."""
    return secrets.token_urlsafe(32)


def _as_date(v: Any) -> date | None:
    """Coerce a date-ish value to a ``datetime.date`` for asyncpg DATE binding.

    The FE sends ``<input type="date">`` values as ISO ``'YYYY-MM-DD'`` strings.
    asyncpg's DATE codec rejects ``str`` (``DataError: expected a datetime.date``),
    so any client date MUST be parsed before bind. Accepts date/datetime/ISO-string;
    ``None``/empty → ``None``; an unparseable string → 400 (not a 500 crash).
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return date.fromisoformat(v.strip()[:10])
        except ValueError:
            raise HTTPException(400, f"Ngày không hợp lệ: {v!r} (định dạng YYYY-MM-DD)")
    raise HTTPException(400, f"Ngày không hợp lệ: {v!r}")


# ---------------------------------------------------------------------------
# Đợt 2 — Audit trail (single source of truth, DRY)
# ---------------------------------------------------------------------------

async def _audit(
    conn: asyncpg.Connection,
    entity_type: str,
    entity_id: int,
    action: str,
    *,
    actor_id: str | None = None,
    actor_vendor_id: int | None = None,
    detail: dict | None = None,
    from_status: str | None = None,
    to_status: str | None = None,
    ip: str | None = None,
) -> None:
    """Append one row to procurement_audit_log.

    REUSABLE — imported by app/api/vendor/quotes.py too (DRY: this is the
    single source). NEVER opens its own transaction: the caller MUST already
    be inside `async with conn.transaction()` so the audit row is atomic with
    the business write. Append-only (no UPDATE/DELETE).

    Data types: users.id = UUID (actor_id); vendor_accounts.id = BIGINT
    (actor_vendor_id); detail → json.dumps → ::jsonb; ip → ::inet.
    """
    await conn.execute(
        """
        INSERT INTO procurement_audit_log
            (entity_type, entity_id, action, from_status, to_status,
             actor_id, actor_vendor_id, detail, ip)
        VALUES ($1, $2, $3, $4, $5, $6, $7, $8::jsonb, $9::inet)
        """,
        entity_type, int(entity_id), action, from_status, to_status,
        actor_id, actor_vendor_id, _json.dumps(detail or {}), ip,
    )


# ---------------------------------------------------------------------------
# P7 — Internal approval gate (app_config-driven, DEFAULT-OFF)
# ---------------------------------------------------------------------------
#
# Two flags seeded by migration 005 (P1), read via value::text like the
# auto_ap pattern (procurement.py auto-AP block). DEFAULT-OFF so a solo owner
# is unaffected: 'procurement_approval_required'='false',
# 'procurement_approval_allow_self'='true'.

def _cfg_bool(raw: Any, default: bool) -> bool:
    """Coerce an app_config value::text flag to bool (mirrors auto_ap parsing).

    Stored values may be JSON-ish ('true'/'false'/'"true"'/'1'). None → default.
    """
    if raw is None:
        return default
    return str(raw).strip().strip('"').lower() in ("true", "1", "yes")


async def _read_approval_config(conn: asyncpg.Connection) -> dict[str, bool]:
    """Read the two P7 flags. Missing rows fall back to the DEFAULT-OFF posture."""
    rows = await conn.fetch(
        "SELECT key, value::text AS value FROM app_config "
        "WHERE key IN ('procurement_approval_required', 'procurement_approval_allow_self')"
    )
    vals = {r["key"]: r["value"] for r in rows}
    return {
        "approval_required": _cfg_bool(vals.get("procurement_approval_required"), False),
        "allow_self": _cfg_bool(vals.get("procurement_approval_allow_self"), True),
    }


# ---------------------------------------------------------------------------
# Đợt 3 — Award maker-checker gate (app_config-driven, DEFAULT-OFF)
# ---------------------------------------------------------------------------
#
# Three flags seeded by migration 020:
#   procurement_award_approval_enabled        ('false') — master gate, OFF
#   procurement_award_approval_threshold_vnd  ('50000000') — chỉ award >= ngưỡng
#                                                            mới phải duyệt
#   procurement_award_breakglass_enabled      ('false') — cho tự duyệt khẩn cấp
# DEFAULT-OFF → solo owner / batch cũ KHÔNG đổi hành vi (finalize ngay).

def _cfg_num(raw: Any, default: float) -> float:
    """Coerce an app_config value::text → float (mirror _cfg_bool).

    threshold lưu trong app_config dạng JSON number (50000000) → value::text =
    '50000000'; nhưng nếu ai đó lưu dạng JSON string ('"50000000"') thì strip
    quote cho an toàn. KHÔNG reuse _coerce_num (cái đó cho FE input, không strip
    quote JSONB). None / không parse được → default.
    """
    if raw is None:
        return default
    try:
        return float(str(raw).strip().strip('"'))
    except (ValueError, TypeError):
        return default


async def _read_award_approval_config(conn: asyncpg.Connection) -> dict[str, Any]:
    """Đọc 3 cờ Đợt 3 trong 1 query. Thiếu row → DEFAULT-OFF (giữ hành vi cũ)."""
    rows = await conn.fetch(
        "SELECT key, value::text AS value FROM app_config WHERE key IN ("
        "'procurement_award_approval_enabled',"
        "'procurement_award_approval_threshold_vnd',"
        "'procurement_award_breakglass_enabled')"
    )
    vals = {r["key"]: r["value"] for r in rows}
    return {
        "enabled":   _cfg_bool(vals.get("procurement_award_approval_enabled"), False),
        "threshold": _cfg_num(vals.get("procurement_award_approval_threshold_vnd"), 50_000_000.0),
        "breakglass": _cfg_bool(vals.get("procurement_award_breakglass_enabled"), False),
    }


def _parse_deadline(raw: Any) -> datetime | None:
    """Parse an ISO8601 date+time deadline into a tz-aware datetime.

    Tolerates a trailing 'Z' (→ +00:00). A bare date/datetime WITHOUT an offset
    is assumed to be Vietnam local time (+07:00) so the FE may send either a
    full offset string (preferred) or a naive 'YYYY-MM-DDTHH:MM'. Returns None
    for empty input. Raises HTTPException(400) on an unparseable string.
    """
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    if isinstance(raw, datetime):
        dt = raw
    else:
        try:
            dt = datetime.fromisoformat(str(raw).strip().replace("Z", "+00:00"))
        except ValueError:
            raise HTTPException(400, "Hạn báo giá không hợp lệ (ISO8601)")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone(timedelta(hours=7)))
    return dt


def _sealed_active(batch) -> bool:
    """Đợt 2b — True ⇔ batch niêm phong giá VÀ chưa tới hạn. HELPER DUY NHẤT
    gate ở MỌI điểm lộ đơn giá NCC (matrix, decision-sheet, get_batch_admin,
    vendor_full_quote) → đừng hard-code lại so sánh deadline ở đâu khác.

    `batch` là asyncpg.Record/dict có 'sealed_until_deadline' + 'bid_deadline'.
    bid_deadline lưu timestamptz (asyncpg trả tz-aware) → so sánh với now(UTC).
    Hết hạn HOẶC sealed=False → False (hiện giá bình thường). sealed=True nhưng
    CHƯA có hạn → True (giữ kín tới khi đặt hạn). FAIL-CLOSED: thà giấu nhầm
    còn hơn rò. KISS.
    """
    if not batch["sealed_until_deadline"]:
        return False
    dl = batch["bid_deadline"]
    if dl is None:
        return True  # sealed bật nhưng chưa có hạn → giữ kín
    return datetime.now(timezone.utc) < dl


def _coerce_date(raw: Any):
    """Parse an ISO date/datetime → date | None (for DATE columns). Never raises."""
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    try:
        return datetime.fromisoformat(str(raw).strip().replace("Z", "+00:00")).date()
    except (ValueError, TypeError):
        return None


def _coerce_num(raw: Any):
    """Parse a numeric → float | None (for NUMERIC columns). Never raises."""
    if raw in (None, ""):
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def _vendor_portal_base() -> str:
    """Base URL of the supplier LOGIN portal (ncc.songchau.vn)."""
    return (
        getattr(settings, "VENDOR_PORTAL_URL", None)
        or getattr(settings, "PUBLIC_BASE_URL", None)
        # Cổng NCC hiện phục vụ tại erp.songchau.vn/ncc (basePath '/ncc', tạm tới
        # khi có domain riêng). Set env VENDOR_PORTAL_URL khi có domain riêng.
        or "https://erp.songchau.vn/ncc"
    ).rstrip("/")


def _resolve_under_files_base(stored_path: str | None) -> Path:
    """Resolve a server-stored path to a real file pinned UNDER FILES_BASE_PATH.

    Generic sibling of `_resolve_contract_pdf` used by P5 quote-attachment routes.
    Accepts absolute legacy paths or relative ('vendor_uploads/...'); rejects
    anything escaping the root or missing on disk. NEVER accepts a client path.
    """
    if not stored_path:
        raise HTTPException(404, "File không tồn tại")
    base = Path(settings.FILES_BASE_PATH).resolve()
    rel = str(stored_path)
    candidate = (base / rel) if not os.path.isabs(rel) else Path(rel)
    try:
        resolved = candidate.resolve()
        resolved.relative_to(base)
    except (ValueError, OSError):
        logger.warning("Attachment path escapes FILES_BASE_PATH: %r", stored_path)
        raise HTTPException(404, "File không hợp lệ")
    if not resolved.is_file():
        raise HTTPException(404, "File không tồn tại")
    return resolved


def _resolve_contract_pdf(contract_file_path: str | None) -> Path:
    """Resolve a stored contract_file_path to a real file UNDER FILES_BASE_PATH.

    Path-traversal guard: the relative path comes from the DB (server-generated),
    but we still pin it under FILES_BASE_PATH and reject anything that resolves
    outside that root or does not exist. NEVER accepts a client-supplied path.
    """
    if not contract_file_path:
        raise HTTPException(404, "Chưa có file PDF — tạo PDF trước")
    base = Path(settings.FILES_BASE_PATH).resolve()
    rel = str(contract_file_path)
    # Stored value may be relative ('contracts/x.pdf') or absolute legacy path.
    candidate = (base / rel) if not os.path.isabs(rel) else Path(rel)
    try:
        resolved = candidate.resolve()
        resolved.relative_to(base)
    except (ValueError, OSError):
        logger.warning("Contract PDF path escapes FILES_BASE_PATH: %r", contract_file_path)
        raise HTTPException(404, "File PDF không hợp lệ")
    if not resolved.is_file():
        raise HTTPException(404, "File PDF không tồn tại")
    return resolved


# ---------------------------------------------------------------------------
# P5 — Item DRAWING upload / view (sandboxed file store)
# ---------------------------------------------------------------------------
#
# An item's `drawing_url` may carry one of two schemes:
#   * `bqms://<bqms_code>`  — BQMS-imported items; the image is served by the
#     existing /api/v1/bqms/rfq/image endpoint (NEVER a file on disk here).
#   * `file://drawings/<batch_id>/<item_id>/<name>` — an admin-uploaded file
#     stored UNDER FILES_BASE_PATH/drawings. P5 adds this scheme; the GET
#     endpoints stream it back with the right content-type.
# Any other (legacy http/https) value is treated as an external URL and the GET
# redirects to it, so older rows keep rendering. The upload endpoint only ever
# WRITES the file:// scheme.

# Allowed upload extensions → response content-type. dwg has no canonical web
# mime; octet-stream forces a download (the FE shows a "tải xuống" affordance).
_DRAWING_CONTENT_TYPES: dict[str, str] = {
    ".pdf": "application/pdf",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".dwg": "application/octet-stream",
}
_DRAWING_MAX_BYTES = 20 * 1024 * 1024  # 20 MB cap (drawings are heavier than quotes)
_DRAWING_FILE_PREFIX = "file://"


def _drawing_ext(filename: str | None) -> str:
    """Return the lower-cased extension IF it is an allowed drawing type, else ''."""
    if not filename:
        return ""
    name = filename.strip().lower()
    for ext in _DRAWING_CONTENT_TYPES:
        if name.endswith(ext):
            return ext
    return ""


def _resolve_drawing_file(drawing_url: str | None) -> tuple[Path, str]:
    """Resolve a `file://...` drawing_url to (real_path, content_type) UNDER FILES_BASE_PATH.

    Path-traversal guard mirrors `_resolve_contract_pdf`: the stored relative path
    is server-generated, but we still pin it under FILES_BASE_PATH and reject
    anything resolving outside that root or missing on disk. Raises 404 otherwise.
    Caller must have already confirmed the scheme is `file://`.
    """
    rel = str(drawing_url or "")[len(_DRAWING_FILE_PREFIX):].lstrip("/")
    if not rel:
        raise HTTPException(404, "Bản vẽ không hợp lệ")
    base = (Path(settings.FILES_BASE_PATH) / "drawings").resolve()
    candidate = base / rel
    try:
        resolved = candidate.resolve()
        resolved.relative_to(base)
    except (ValueError, OSError):
        logger.warning("Drawing path escapes drawings root: %r", drawing_url)
        raise HTTPException(404, "Bản vẽ không hợp lệ")
    if not resolved.is_file():
        raise HTTPException(404, "File bản vẽ không tồn tại")
    ctype = _DRAWING_CONTENT_TYPES.get(resolved.suffix.lower(), "application/octet-stream")
    return resolved, ctype


def _drawing_response(drawing_url: str | None, *, item_no: int | None = None):
    """Build the right FastAPI response for ANY drawing_url scheme (shared by admin+vendor GET).

    * `file://…`  → stream the sandboxed file (path-guarded).
    * `bqms://X`  → redirect to /api/v1/bqms/rfq/image?bqms_code=X (existing source).
    * `http(s)://…` → redirect to the external URL (legacy rows).
    Raises 404 when there is no drawing or the scheme is unknown.
    """
    url = (drawing_url or "").strip()
    if not url:
        raise HTTPException(404, "Mã hàng này chưa có bản vẽ")
    if url.startswith(_DRAWING_FILE_PREFIX):
        resolved, ctype = _resolve_drawing_file(url)
        # inline so images/PDFs render in the lightbox; dwg downloads via octet-stream.
        suffix = f"_{item_no}" if item_no is not None else ""
        return FileResponse(
            str(resolved),
            media_type=ctype,
            filename=f"ban-ve{suffix}{resolved.suffix.lower()}",
            content_disposition_type="inline",
        )
    if url.startswith("bqms://"):
        code = url[len("bqms://"):]
        from urllib.parse import quote as _q
        return RedirectResponse(url=f"/api/v1/bqms/rfq/image?bqms_code={_q(code)}", status_code=307)
    if url.startswith("http://") or url.startswith("https://"):
        return RedirectResponse(url=url, status_code=307)
    raise HTTPException(404, "Bản vẽ không hợp lệ")


# ---------------------------------------------------------------------------
# Vendor Account Management
# ---------------------------------------------------------------------------

@router.get("/vendors")
async def list_vendor_accounts(
    status: str = Query("all"),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách tài khoản nhà cung cấp."""
    where = "1=1"
    if status == "pending":
        where = "va.is_approved = false"
    elif status == "approved":
        where = "va.is_approved = true"

    rows = await conn.fetch(
        f"""
        SELECT va.id, va.company_name, va.contact_name, va.phone, va.tax_code,
               va.product_categories, va.is_approved, va.approved_at, va.created_at,
               u.email,
               (SELECT COUNT(*) FROM vendor_quotes vq WHERE vq.vendor_id = va.id) AS quote_count
        FROM vendor_accounts va
        JOIN users u ON u.id = va.user_id
        WHERE {where}
        ORDER BY va.created_at DESC
        """
    )
    return {"data": [dict(r) for r in rows], "total": len(rows)}


@router.patch("/vendors/{vendor_id}/approve")
async def approve_vendor(
    vendor_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Duyệt tài khoản nhà cung cấp."""
    va = await conn.fetchrow("SELECT id, user_id FROM vendor_accounts WHERE id = $1", vendor_id)
    if not va:
        raise HTTPException(404, "Nhà cung cấp không tồn tại")

    await conn.execute(
        "UPDATE vendor_accounts SET is_approved = true, approved_by = $1, approved_at = NOW() WHERE id = $2",
        token_data.user_id, vendor_id,
    )
    await conn.execute("UPDATE users SET is_active = true WHERE id = $1", va["user_id"])

    return {"message": "Đã duyệt nhà cung cấp"}


@router.patch("/vendors/{vendor_id}/reject")
async def reject_vendor(
    vendor_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Từ chối tài khoản nhà cung cấp."""
    await conn.execute(
        "UPDATE vendor_accounts SET is_approved = false, notes = 'Từ chối bởi admin' WHERE id = $1",
        vendor_id,
    )
    return {"message": "Đã từ chối nhà cung cấp"}


# ---------------------------------------------------------------------------
# Dashboard KPIs (IMV-style stat strip + tab counts)
# ---------------------------------------------------------------------------

@router.get("/stats")
async def procurement_stats(
    days: int = Query(90, ge=1, le=730),
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """KPI strip cho dashboard admin đấu thầu (IMV-style).

    Powers the StatTile row + entity tab badges. `days` = cửa sổ thời gian cho
    các số liệu "in_window" (mặc định 90 ngày).
    """
    window = f"NOW() - INTERVAL '{int(days)} days'"

    batches = await conn.fetchrow(
        f"""
        SELECT
            COUNT(*)::int AS total,
            COUNT(*) FILTER (WHERE status = 'draft')::int      AS draft,
            COUNT(*) FILTER (WHERE status = 'cho_duyet')::int  AS cho_duyet,
            COUNT(*) FILTER (WHERE status = 'published')::int  AS published,
            COUNT(*) FILTER (WHERE status = 'evaluating')::int AS evaluating,
            COUNT(*) FILTER (WHERE status = 'awarded')::int    AS awarded,
            COUNT(*) FILTER (WHERE status IN ('closed','cancelled'))::int AS closed,
            COUNT(*) FILTER (WHERE published_at >= {window})::int AS published_in_window
        FROM procurement_rfq_batches
        """
    )

    quotes = await conn.fetchrow(
        f"""
        SELECT
            COUNT(*) FILTER (WHERE status = 'submitted')::int AS total_submitted,
            COUNT(*) FILTER (WHERE status = 'submitted' AND submitted_at >= {window})::int AS submitted_in_window,
            COUNT(*) FILTER (WHERE status = 'awarded')::int   AS awarded
        FROM vendor_quotes
        """
    )

    invitations = await conn.fetchrow(
        """
        SELECT
            COUNT(*)::int AS total,
            COUNT(*) FILTER (WHERE viewed_at IS NOT NULL)::int AS viewed,
            COUNT(*) FILTER (WHERE status = 'submitted')::int  AS submitted
        FROM procurement_rfq_invitations
        """
    )
    inv_total = invitations["total"] or 0
    submitted_pct = round((invitations["submitted"] / inv_total) * 100, 1) if inv_total else 0.0

    vendors = await conn.fetchrow(
        """
        SELECT
            COUNT(*) FILTER (WHERE status = 'active' OR is_approved = true)::int AS active,
            COUNT(*) FILTER (WHERE (status = 'pending' OR status IS NULL)
                              AND COALESCE(is_approved, false) = false)::int     AS pending
        FROM vendor_accounts
        """
    )

    # Tab badge counts — mirror IMV /kpi `counts` shape.
    counts = {
        "batches": batches["total"],
        "quotes": quotes["total_submitted"],
        "contracts": await conn.fetchval("SELECT COUNT(*)::int FROM procurement_contracts") or 0,
        "pos": await conn.fetchval("SELECT COUNT(*)::int FROM procurement_pos") or 0,
        "deliveries": await conn.fetchval("SELECT COUNT(*)::int FROM procurement_deliveries") or 0,
        "vendors": await conn.fetchval("SELECT COUNT(*)::int FROM vendor_accounts") or 0,
    }

    return {
        "data": {
            "batches": dict(batches),
            "quotes": dict(quotes),
            "invitations": {
                "total": inv_total,
                "viewed": invitations["viewed"] or 0,
                "submitted_pct": submitted_pct,
            },
            "vendors": dict(vendors),
        },
        "counts": counts,
    }


# ---------------------------------------------------------------------------
# RFQ Batch Management
# ---------------------------------------------------------------------------

@router.get("/batches")
async def list_batches(
    status: str = Query("all"),
    q: str | None = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=50),
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách đợt báo giá + số NCC được mời / số đã báo giá.

    FIX (Đợt 1): status/q được PARAMETER-HOÁ (trước đây nội suy f-string → SQL
    injection). invited_count/submitted_count cho bảng so sánh nhanh.
    """
    where = "1=1"
    params: list = []
    idx = 1
    if status and status != "all":
        where += f" AND b.status = ${idx}"
        params.append(status); idx += 1
    if q:
        where += f" AND (b.title ILIKE ${idx} OR b.batch_code ILIKE ${idx})"
        params.append(f"%{q}%"); idx += 1

    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM procurement_rfq_batches b WHERE {where}", *params
    )

    params_paged = params + [limit, (page - 1) * limit]
    rows = await conn.fetch(
        f"""
        SELECT b.*, u.full_name AS created_by_name,
               (SELECT COUNT(*) FROM procurement_rfq_invitations inv
                 WHERE inv.batch_id = b.id)::int AS invited_count,
               (SELECT COUNT(*) FROM procurement_rfq_invitations inv
                 WHERE inv.batch_id = b.id AND inv.status = 'submitted')::int AS submitted_count
        FROM procurement_rfq_batches b
        LEFT JOIN users u ON u.id = b.created_by
        WHERE {where}
        ORDER BY b.created_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params_paged,
    )
    return {"data": [dict(r) for r in rows], "total": total}


@router.post("/batches")
async def create_batch(
    body: dict[str, Any],
    # V8 (Thang 2026-06-27): widened from admin-only to _WRITE_ROLES (admin /
    # manager / procurement) so a phiên can be opened straight from "Gửi đấu thầu"
    # in Thư viện nguồn cung — not just by an admin. Other write endpoints (thêm
    # mã, mời NCC, công bố…) already use _WRITE_ROLES.
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tạo đợt báo giá mới.

    Optional `bid_deadline` (ISO8601 date+time, 'Z' tolerated) seeds both the
    active-round deadline `bid_deadline` AND `deadline_round1` (round-1 is the
    first active round). The auto-close + reminder sweep keys off `bid_deadline`.
    """
    title = body.get("title")
    if not title:
        raise HTTPException(400, "Tiêu đề đợt báo giá là bắt buộc")

    bid_deadline = _parse_deadline(body.get("bid_deadline"))

    # Generate batch code
    count = await conn.fetchval("SELECT COUNT(*) FROM procurement_rfq_batches") or 0
    batch_code = f"BATCH-2026-{count + 1:04d}"

    batch_id = await conn.fetchval(
        """
        INSERT INTO procurement_rfq_batches
            (batch_code, title, description, award_mode, created_by, notes_internal,
             phu_trach, visibility, bid_deadline, deadline_round1, sealed_until_deadline)
        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $9, $10)
        RETURNING id
        """,
        batch_code, title.strip(),
        body.get("description", "").strip() or None,
        body.get("award_mode", "per_item"),
        token_data.user_id,
        body.get("notes_internal", "").strip() or None,
        (body.get("phu_trach") or "").strip() or None,
        body.get("visibility") or "invited",
        bid_deadline,
        # Đợt 2b [SB]: niêm phong giá tới hạn (per-batch, DEFAULT OFF). Coerce
        # truthy → đề phòng FE gửi "true"/1; cột NOT NULL DEFAULT FALSE.
        bool(body.get("sealed_until_deadline")),
    )

    return {"data": {"id": batch_id, "batch_code": batch_code}, "message": "Đã tạo đợt báo giá"}


# ---------------------------------------------------------------------------
# DRY — single shared item-insert helper (Commercial bidding P2)
# ---------------------------------------------------------------------------
#
# EVERY importer (manual add, BQMS, catalog, IMV, paste, Excel) funnels through
# _insert_rfq_items. It is the SINGLE OWNER of:
#   1. the draft-status guard (batch must be status='draft'),
#   2. MAX(item_no)+1 sequencing,
#   3. per-(batch_id, item_code) dedupe (rows WITHOUT an item_code are never
#      deduped — manual/degraded rows always insert),
#   4. the INSERT (source_kind / source_ref_id + explicit provenance columns),
#   5. the procurement_rfq_batches.item_count UPDATE.
# Callers MUST already be inside `async with conn.transaction()` (single txn).
#
# `specification` is NOT NULL on the base table, so the helper always coalesces
# a non-empty spec from specification_full / product_name / item_code / bqms_code.

# Explicit column allow-list for the items INSERT. Provenance/internal columns
# (source_kind, source_ref_id, target_price, notes_internal, bqms_code,
# source_bqms_rfq_id) live here on the ADMIN write path ONLY — vendor-facing
# serializers (app/api/vendor/*) never SELECT them. Re-audited 2026-06-22.
_ITEM_INSERT_COLS = (
    "specification", "specification_full", "bqms_code", "item_code",
    "product_name", "model", "maker", "quantity", "unit",
    "required_material", "notes", "target_price", "source_bqms_rfq_id",
    "drawing_url", "part_no", "cis_code", "moq", "dimension", "item_deadline",
)


def _norm_spec(row: dict[str, Any]) -> str:
    """Derive a non-empty `specification` (column is NOT NULL)."""
    for key in ("specification", "specification_full", "product_name",
                "item_code", "bqms_code", "model"):
        v = row.get(key)
        if v is not None and str(v).strip():
            return str(v).strip()
    return "(không tên)"


async def _insert_rfq_items(
    conn: asyncpg.Connection,
    batch_id: int,
    rows: list[dict[str, Any]],
    source_kind: str,
) -> dict[str, Any]:
    """Insert RFQ item rows into a DRAFT batch — the single DRY insert path.

    Args:
        conn: connection ALREADY inside `async with conn.transaction()`.
        batch_id: target batch (must exist + be status='draft').
        rows: list of column→value dicts. Recognised keys are the union of
              _ITEM_INSERT_COLS. A row's `item_code` (when truthy) is the dedupe
              key per (batch_id, item_code). `source_ref_id` may be set per row.
              `dedupe_code` (optional) overrides the dedupe key — BQMS dedupes on
              bqms_code via this, not item_code.
        source_kind: provenance tag written to every inserted row.

    Returns: {"imported": int, "skipped": int, "imported_ids": [int], ...}.
    Raises 404 if batch missing, 400 if not draft.
    """
    batch = await conn.fetchrow(
        "SELECT status FROM procurement_rfq_batches WHERE id = $1 FOR UPDATE", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] != "draft":
        raise HTTPException(400, "Chỉ thêm/import item khi batch ở status='draft'")

    max_no = await conn.fetchval(
        "SELECT COALESCE(MAX(item_no), 0) FROM procurement_rfq_items WHERE batch_id = $1",
        batch_id,
    ) or 0

    imported_ids: list[int] = []
    skipped: list[str] = []
    seen_codes: set[str] = set()  # dedupe within this single call too

    for row in rows:
        dedupe_code = row.get("dedupe_code")
        if not dedupe_code:
            ic = row.get("item_code")
            dedupe_code = str(ic).strip() if ic and str(ic).strip() else None

        if dedupe_code:
            if dedupe_code in seen_codes:
                skipped.append(dedupe_code)
                continue
            exists = await conn.fetchval(
                """
                SELECT 1 FROM procurement_rfq_items
                 WHERE batch_id = $1
                   AND (item_code = $2 OR bqms_code = $2)
                 LIMIT 1
                """,
                batch_id, dedupe_code,
            )
            if exists:
                skipped.append(dedupe_code)
                continue
            seen_codes.add(dedupe_code)

        max_no += 1
        values = [row.get(c) for c in _ITEM_INSERT_COLS]
        # Override the NOT NULL specification with a coalesced value.
        values[0] = _norm_spec(row)

        col_list = ", ".join(_ITEM_INSERT_COLS)
        ph = ", ".join(f"${i}" for i in range(1, len(_ITEM_INSERT_COLS) + 1))
        sk_ph = f"${len(_ITEM_INSERT_COLS) + 1}"
        bid_ph = f"${len(_ITEM_INSERT_COLS) + 2}"
        ino_ph = f"${len(_ITEM_INSERT_COLS) + 3}"
        sref_ph = f"${len(_ITEM_INSERT_COLS) + 4}"

        new_id = await conn.fetchval(
            f"""
            INSERT INTO procurement_rfq_items
                ({col_list}, source_kind, batch_id, item_no, source_ref_id)
            VALUES ({ph}, {sk_ph}, {bid_ph}, {ino_ph}, {sref_ph})
            RETURNING id
            """,
            *values, source_kind, batch_id, max_no, row.get("source_ref_id"),
        )
        imported_ids.append(new_id)

    # item_count is owned here — no importer can desync the publish guard.
    await conn.execute(
        "UPDATE procurement_rfq_batches "
        "SET item_count = (SELECT COUNT(*) FROM procurement_rfq_items WHERE batch_id = $1) "
        "WHERE id = $1",
        batch_id,
    )

    return {
        "imported": len(imported_ids),
        "skipped": len(skipped),
        "imported_ids": imported_ids,
        "skipped_codes": skipped,
    }


@router.post("/batches/{batch_id}/items")
async def add_items_to_batch(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Thêm items vào đợt báo giá (chỉ khi batch ở trạng thái nháp)."""
    items = body.get("items", [])
    if not items:
        raise HTTPException(400, "Cần ít nhất 1 item")

    rows: list[dict[str, Any]] = []
    for item in items:
        rows.append({
            "specification": (item.get("specification") or "").strip(),
            "bqms_code": (item.get("bqms_code") or "").strip() or None,
            "item_code": (item.get("item_code") or "").strip() or None,
            "product_name": (item.get("product_name") or "").strip() or None,
            "model": (item.get("model") or "").strip() or None,
            "maker": (item.get("maker") or "").strip() or None,
            "quantity": item.get("quantity", 0),
            "unit": (item.get("unit") or "EA"),
            "required_material": (item.get("required_material") or "").strip() or None,
            "notes": (item.get("notes") or "").strip() or None,
            "target_price": item.get("target_price"),
            "source_bqms_rfq_id": item.get("source_bqms_rfq_id"),
        })

    async with conn.transaction():
        result = await _insert_rfq_items(conn, batch_id, rows, "manual")

    return {
        "message": f"Đã thêm {result['imported']} items",
        "added": result["imported"],
        "imported": result["imported"],
        "skipped": result["skipped"],
    }


@router.patch("/batches/{batch_id}/items/{item_id}")
async def update_rfq_item(
    batch_id: int,
    item_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Sửa INLINE thông số 1 mã hàng (Thang 2026-06-29).

    Cho sửa khi đợt CHƯA chốt/đóng/huỷ. Trường hợp lệ: specification, product_name,
    model, maker, required_material, unit, quantity, target_price. target_price là
    ADMIN-ONLY (không lộ sang cổng NCC). Audit 'item_edit'.
    """
    batch = await conn.fetchrow(
        "SELECT status FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] in ("awarded", "closed", "cancelled"):
        raise HTTPException(400, "Đợt đã chốt/đóng — không sửa được mã hàng")
    if not await conn.fetchval(
        "SELECT 1 FROM procurement_rfq_items WHERE id = $1 AND batch_id = $2", item_id, batch_id
    ):
        raise HTTPException(404, "Mã hàng không thuộc đợt này")

    sets: list[str] = []
    vals: list[Any] = []
    changed: list[str] = []
    for f in ("specification", "product_name", "model", "maker", "required_material", "unit"):
        if f in body:
            v = body[f]
            v = (str(v).strip() or None) if v is not None else None
            if f == "specification" and not v:
                raise HTTPException(400, "Tên/quy cách không được để trống")
            if f == "unit" and not v:
                v = "EA"
            sets.append(f"{f} = ${len(vals) + 3}"); vals.append(v); changed.append(f)
    for f in ("quantity", "target_price"):
        if f in body:
            raw = body[f]
            v = None if raw in (None, "") else float(raw)
            if f == "quantity" and (v is None or v <= 0):
                raise HTTPException(400, "Số lượng phải lớn hơn 0")
            if f == "target_price" and v is not None and v < 0:
                raise HTTPException(400, "Giá mục tiêu không hợp lệ")
            sets.append(f"{f} = ${len(vals) + 3}"); vals.append(v); changed.append(f)

    if not sets:
        raise HTTPException(400, "Không có trường hợp lệ để cập nhật")

    async with conn.transaction():
        await conn.execute(
            f"UPDATE procurement_rfq_items SET {', '.join(sets)} WHERE id = $1 AND batch_id = $2",
            item_id, batch_id, *vals,
        )
        await _audit(
            conn, "rfq_item", item_id, "item_edit",
            actor_id=token_data.user_id,
            detail={"batch_id": batch_id, "fields": changed},
        )
    return {"ok": True, "changed": changed}


@router.get("/items/{item_id}/shared-files")
async def list_item_shared_files(
    item_id: int,
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tập file của mã ĐÃ tick chia sẻ cho NCC (FE 'File mã' tô checkbox)."""
    rows = await conn.fetch(
        "SELECT kind, file_name FROM procurement_rfq_shared_files WHERE item_id = $1", item_id
    )
    return {"shared": [{"kind": r["kind"], "file_name": r["file_name"]} for r in rows]}


@router.post("/items/{item_id}/share-file")
async def toggle_item_shared_file(
    item_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Bật/tắt chia sẻ 1 file (thư mục Raw của RFQ gốc) cho NCC.

    shared=true → INSERT; false → DELETE. Lưu sẵn rfq_number để cổng NCC tải file
    KHÔNG cần JOIN lại + KHÔNG lộ rfq_number. Audit item_share_file/unshare.
    """
    file_name = (body.get("file_name") or "").strip()
    kind = (body.get("kind") or "raw").strip()
    shared = bool(body.get("shared"))
    if not file_name:
        raise HTTPException(400, "Thiếu tên file")
    if kind not in ("raw", "images"):
        raise HTTPException(400, "kind phải là raw hoặc images")

    item = await conn.fetchrow(
        """
        SELECT i.id, i.batch_id, br.rfq_number
          FROM procurement_rfq_items i
          LEFT JOIN bqms_rfq br ON br.id = i.source_bqms_rfq_id
         WHERE i.id = $1
        """,
        item_id,
    )
    if not item:
        raise HTTPException(404, "Mã hàng không tồn tại")
    if not item["rfq_number"]:
        raise HTTPException(400, "Mã không có RFQ nguồn (BQMS) để chia sẻ file")

    async with conn.transaction():
        if shared:
            await conn.execute(
                """
                INSERT INTO procurement_rfq_shared_files (batch_id, item_id, rfq_number, kind, file_name)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (item_id, kind, file_name) DO NOTHING
                """,
                item["batch_id"], item_id, item["rfq_number"], kind, file_name,
            )
        else:
            await conn.execute(
                "DELETE FROM procurement_rfq_shared_files WHERE item_id = $1 AND kind = $2 AND file_name = $3",
                item_id, kind, file_name,
            )
        await _audit(
            conn, "rfq_item", item_id,
            "item_share_file" if shared else "item_unshare_file",
            actor_id=token_data.user_id,
            detail={"batch_id": item["batch_id"], "kind": kind, "file": file_name},
        )
    return {"ok": True, "shared": shared}


@router.patch("/batches/{batch_id}/publish")
async def publish_batch(
    batch_id: int,
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Công bố đợt báo giá cho nhà cung cấp.

    P7 — internal approval gate (DEFAULT-OFF):
      * approval ON  ('procurement_approval_required'=true): batch MUST be
        status='approved' (đã được Duyệt nội bộ) → publish. draft/cho_duyet → 400.
      * approval OFF (default): publish allowed from 'draft' OR 'approved'. When
        publishing straight from 'draft' we first stamp a SYNTHETIC approved row
        (approved_by=NULL, approval_auto=true, approved_at=NOW()) so audit +
        downstream stay uniform regardless of the gate. Unchanged owner UX.
    """
    batch = await conn.fetchrow(
        "SELECT id, status, item_count FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    cfg = await _read_approval_config(conn)
    cur_status = batch["status"]

    if cfg["approval_required"]:
        # Gate ON: only an internally-approved batch may be published.
        if cur_status == "published":
            return {"message": "Đợt đã được công bố."}  # idempotent
        if cur_status != "approved":
            raise HTTPException(400, "Phiên phải được duyệt nội bộ trước khi công bố")
    else:
        # Gate OFF: allow draft → published directly (solo owner), or
        # approved → published (if the gate was toggled off mid-flight).
        if cur_status == "published":
            return {"message": "Đợt đã được công bố."}  # idempotent
        if cur_status not in ("draft", "approved"):
            raise HTTPException(400, "Chỉ có thể công bố đợt ở trạng thái nháp hoặc đã duyệt")

    if (batch["item_count"] or 0) == 0:
        raise HTTPException(400, "Cần thêm ít nhất 1 item trước khi công bố")

    async with conn.transaction():
        # Synthetic approval row when publishing straight from 'draft' (gate OFF).
        # Idempotent: only stamp when not already approved (won't clobber a real
        # approved_by from a prior gate-ON approval).
        if cur_status == "draft":
            await conn.execute(
                "UPDATE procurement_rfq_batches "
                "SET approved_by = NULL, approval_auto = true, approved_at = NOW() "
                "WHERE id = $1 AND approved_at IS NULL",
                batch_id,
            )

        await conn.execute(
            "UPDATE procurement_rfq_batches SET status = 'published', published_at = NOW() WHERE id = $1",
            batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "publish",
            actor_id=token_data.user_id, from_status=cur_status, to_status="published",
            detail={"approval_auto": cur_status == "draft", "approval_required": cfg["approval_required"]},
        )

    return {"message": "Đã công bố đợt báo giá. Nhà cung cấp có thể xem và báo giá."}


# ---------------------------------------------------------------------------
# P7 — Internal approval workflow endpoints
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/submit-for-approval")
async def submit_for_approval(
    batch_id: int,
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P7 — Gửi phiên đi duyệt nội bộ: draft → cho_duyet.

    Chỉ admin (người tạo/quản lý phiên) gửi duyệt. Phiên phải đang ở 'draft'
    và có ít nhất 1 mã. Ghi audit action='submit_for_approval'.
    """
    batch = await conn.fetchrow(
        "SELECT id, status, item_count FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] != "draft":
        raise HTTPException(400, "Chỉ gửi duyệt phiên đang ở trạng thái nháp")
    if (batch["item_count"] or 0) == 0:
        raise HTTPException(400, "Cần thêm ít nhất 1 item trước khi gửi duyệt")

    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_rfq_batches "
            "SET status = 'cho_duyet', submitted_by = $1, submitted_at = NOW() "
            "WHERE id = $2",
            token_data.user_id, batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "submit_for_approval",
            actor_id=token_data.user_id, from_status="draft", to_status="cho_duyet",
        )

    return {"message": "Đã gửi phiên đi duyệt nội bộ."}


@router.post("/batches/{batch_id}/approve")
async def approve_batch(
    batch_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P7 — Duyệt nội bộ: cho_duyet → approved.

    admin/manager duyệt. Nếu 'procurement_approval_allow_self'=false thì người
    gửi duyệt KHÔNG được tự duyệt phiên mình gửi (403). Ghi audit action='approve'.
    """
    batch = await conn.fetchrow(
        "SELECT id, status, submitted_by FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] != "cho_duyet":
        raise HTTPException(400, "Chỉ duyệt phiên đang ở trạng thái chờ duyệt")

    cfg = await _read_approval_config(conn)
    if not cfg["allow_self"] and str(batch["submitted_by"]) == str(token_data.user_id):
        raise HTTPException(403, "Không thể tự duyệt phiên mình gửi")

    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_rfq_batches "
            "SET status = 'approved', approved_by = $1, approved_at = NOW(), approval_auto = false "
            "WHERE id = $2",
            token_data.user_id, batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "approve",
            actor_id=token_data.user_id, from_status="cho_duyet", to_status="approved",
        )

    return {"message": "Đã duyệt phiên nội bộ. Có thể công bố cho NCC."}


@router.post("/batches/{batch_id}/reject-internal")
async def reject_internal(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P7 — Trả lại phiên (từ chối duyệt nội bộ): cho_duyet → draft.

    Body {"reason": str} bắt buộc — lý do trả lại được lưu để admin sửa lại.
    Ghi audit action='reject_internal'.
    """
    reason = (body.get("reason") or "").strip()
    if not reason:
        raise HTTPException(400, "Cần nhập lý do trả lại")

    batch = await conn.fetchrow(
        "SELECT id, status FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] != "cho_duyet":
        raise HTTPException(400, "Chỉ trả lại phiên đang ở trạng thái chờ duyệt")

    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_rfq_batches "
            "SET status = 'draft', approval_rejected_by = $1, approval_rejected_at = NOW(), "
            "    approval_rejection_reason = $2 "
            "WHERE id = $3",
            token_data.user_id, reason, batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "reject_internal",
            actor_id=token_data.user_id, from_status="cho_duyet", to_status="draft",
            detail={"reason": reason},
        )

    return {"message": "Đã trả lại phiên cho người tạo để chỉnh sửa."}


# ---------------------------------------------------------------------------
# P7 — Approval config (admin toggle UI)
# ---------------------------------------------------------------------------

@router.get("/approval-config")
async def get_approval_config(
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P7 + Đợt 3 — Đọc cờ duyệt nội bộ (publish) VÀ cờ maker-checker AWARD
    (read roles, để FE dựng nút publish/submit + banner/nút duyệt-chốt-thầu).
    Writing stays admin-only. DEFAULT-OFF khi chưa cấu hình."""
    cfg = await _read_approval_config(conn)
    award = await _read_award_approval_config(conn)
    # Merge 3 cờ Đợt 3 (prefix award_) vào cùng payload — FE đọc 1 lần.
    cfg["award_approval_enabled"] = award["enabled"]
    cfg["award_approval_threshold_vnd"] = award["threshold"]
    cfg["award_breakglass_enabled"] = award["breakglass"]
    return {"data": cfg}


@router.put("/approval-config")
async def put_approval_config(
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P7 + Đợt 3 — Bật/tắt cờ duyệt nội bộ (publish) VÀ cờ maker-checker AWARD
    (admin). Upsert app_config flags.

    Body (chỉ key nào gửi mới ghi):
      P7:    {"approval_required"?: bool, "allow_self"?: bool}
      Đợt 3: {"award_approval_enabled"?: bool, "award_breakglass_enabled"?: bool,
              "award_approval_threshold_vnd"?: number}
    Bool lưu via to_jsonb(::bool); threshold lưu via to_jsonb(::numeric).
    """
    bool_updates: list[tuple[str, bool]] = []
    if "approval_required" in body:
        bool_updates.append(("procurement_approval_required", bool(body["approval_required"])))
    if "allow_self" in body:
        bool_updates.append(("procurement_approval_allow_self", bool(body["allow_self"])))
    if "award_approval_enabled" in body:
        bool_updates.append(("procurement_award_approval_enabled", bool(body["award_approval_enabled"])))
    if "award_breakglass_enabled" in body:
        bool_updates.append(("procurement_award_breakglass_enabled", bool(body["award_breakglass_enabled"])))

    has_threshold = "award_approval_threshold_vnd" in body
    if not bool_updates and not has_threshold:
        raise HTTPException(400, "Không có cờ nào để cập nhật")

    if has_threshold:
        try:
            threshold_val = float(body["award_approval_threshold_vnd"])
        except (TypeError, ValueError):
            raise HTTPException(400, "Ngưỡng duyệt chốt thầu không hợp lệ (VND)")
        if threshold_val < 0:
            raise HTTPException(400, "Ngưỡng duyệt chốt thầu phải >= 0")

    async with conn.transaction():
        for key, val in bool_updates:
            await conn.execute(
                """
                INSERT INTO app_config (key, value)
                VALUES ($1, to_jsonb($2::bool))
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
                """,
                key, val,
            )
        if has_threshold:
            await conn.execute(
                """
                INSERT INTO app_config (key, value)
                VALUES ('procurement_award_approval_threshold_vnd', to_jsonb($1::numeric))
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
                """,
                threshold_val,
            )

    cfg = await _read_approval_config(conn)
    award = await _read_award_approval_config(conn)
    cfg["award_approval_enabled"] = award["enabled"]
    cfg["award_approval_threshold_vnd"] = award["threshold"]
    cfg["award_breakglass_enabled"] = award["breakglass"]
    return {"data": cfg, "message": "Đã cập nhật cấu hình duyệt"}


@router.patch("/batches/{batch_id}/deadline")
async def set_batch_deadline(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P3 — Đặt/cập nhật hạn báo giá (bid_deadline) của đợt.

    Cho phép ở mọi trạng thái còn mở (draft hoặc published). bid_deadline là hạn
    của VÒNG ĐANG MỞ — sweep tự đóng + nhắc NCC dựa vào cột này. Đồng bộ
    deadline_round{current_round} để khớp cột theo vòng. Body: {"bid_deadline": ISO8601|null}.
    Truyền null để gỡ hạn. Ghi audit action='set_deadline'.
    """
    batch = await conn.fetchrow(
        "SELECT id, status, bid_deadline, current_round FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] not in ("draft", "published"):
        raise HTTPException(400, "Chỉ đặt hạn khi đợt ở trạng thái nháp hoặc đã công bố")

    new_deadline = _parse_deadline(body.get("bid_deadline"))
    cur_round = int(batch["current_round"] or 1)
    round_col = {1: "deadline_round1", 2: "deadline_round2", 3: "deadline_round3"}.get(cur_round, "deadline_round1")

    async with conn.transaction():
        await conn.execute(
            f"UPDATE procurement_rfq_batches SET bid_deadline = $1, {round_col} = $1 WHERE id = $2",
            new_deadline, batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "set_deadline",
            actor_id=token_data.user_id,
            detail={
                "round": cur_round,
                "old": batch["bid_deadline"].isoformat() if batch["bid_deadline"] else None,
                "new": new_deadline.isoformat() if new_deadline else None,
            },
        )
        # Đợt A #5 (Thang 2026-06-28): đổi/gia hạn deadline → re-arm nhắc hạn cho
        # MỌI NCC đang mời (reset reminder_sent_at) để sweep nhắc lại theo hạn MỚI.
        # Vá lỗ "đổi hạn âm thầm" (bất công thông tin) — tái dùng máy nhắc sẵn có,
        # không spam (sweep chỉ bắn khi hạn vào cửa sổ nhắc + idempotent).
        await conn.execute(
            """
            UPDATE procurement_rfq_invitations
               SET reminder_sent_at = NULL
             WHERE batch_id = $1 AND status IN ('invited', 'viewed')
            """,
            batch_id,
        )

    return {
        "data": {"bid_deadline": new_deadline.isoformat() if new_deadline else None},
        "message": "Đã cập nhật hạn báo giá",
    }


@router.patch("/batches/{batch_id}")
async def amend_batch(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 2b [AM] — SỬA metadata phiên sau publish (title/description/notes_internal).

    CHỈ sửa TEXT — KHÔNG đụng item/giá/state-machine (tránh phụ thuộc Đợt 3).
    Cho phép ở draft/published (mượn guard set_batch_deadline). Ghi audit
    action='amend' (old→new) + broadcast notif 'amend' tới TẤT CẢ NCC đã mời
    (reuse loop+dispatch như addendum). Ở draft chưa mời ai → loop rỗng, no-op.

    Body (mọi field optional; bỏ field = không đổi):
      { "title"?: str, "description"?: str|null, "notes_internal"?: str|null }

    Lộ giá: detail['changes'] chỉ chứa TEXT metadata (title/desc/notes), KHÔNG có
    giá/tên NCC; nhánh 'amend' của _build_message KHÔNG đọc changes → notif gửi NCC
    tuyệt đối không kèm nội dung. notes_internal sửa được nhưng KHÔNG vào notif.
    """
    batch = await conn.fetchrow(
        "SELECT id, batch_code, status, title, description, notes_internal "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] not in ("draft", "published"):
        raise HTTPException(400, "Chỉ sửa thông tin khi đợt ở trạng thái nháp hoặc đã công bố")

    # Partial update: chỉ field CÓ trong body mới đụng tới. title không cho rỗng.
    sets: list[str] = []
    vals: list[Any] = []
    changes: dict[str, dict] = {}  # field -> {old, new} cho audit (nội bộ, KHÔNG ra NCC)

    if "title" in body:
        new_title = (body.get("title") or "").strip()
        if not new_title:
            raise HTTPException(400, "Tiêu đề không được để trống")
        if new_title != (batch["title"] or ""):
            sets.append(f"title = ${len(vals) + 1}")
            vals.append(new_title)
            changes["title"] = {"old": batch["title"], "new": new_title}

    for field in ("description", "notes_internal"):
        if field in body:
            new_val = (body.get(field) or "").strip() or None
            if new_val != batch[field]:
                sets.append(f"{field} = ${len(vals) + 1}")
                vals.append(new_val)
                changes[field] = {"old": batch[field], "new": new_val}

    if not sets:
        return {"data": {"id": batch_id}, "message": "Không có thay đổi", "broadcast_to": 0}

    vids = [
        r["vendor_id"]
        for r in await conn.fetch(
            "SELECT DISTINCT vendor_id FROM procurement_rfq_invitations WHERE batch_id = $1",
            batch_id,
        )
    ]

    async with conn.transaction():
        await conn.execute(
            f"UPDATE procurement_rfq_batches SET {', '.join(sets)} WHERE id = ${len(vals) + 1}",
            *vals, batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "amend",
            actor_id=token_data.user_id,
            detail={"batch_id": batch_id, "batch_code": batch["batch_code"], "changes": changes},
        )
        # Broadcast tới NCC đã mời (mẫu addendum/cancel): lần đầu internal=True
        # (team đúng 1 row), các lần sau internal=False (mỗi NCC 1 row). Ở draft
        # chưa mời ai → vids rỗng → loop no-op (không spam). detail KHÔNG chứa changes.
        for i, vid in enumerate(vids):
            await dispatch_procurement_event(
                conn, "batch", batch_id, "amend",
                actor_id=token_data.user_id, awarded_vendor_id=vid,
                internal=(i == 0),
                detail={"batch_id": batch_id, "batch_code": batch["batch_code"]},
            )

    logger.info("[AMEND] batch=%s fields=%s broadcast=%s",
                batch_id, list(changes.keys()), len(vids))
    return {"data": {"id": batch_id}, "message": "Đã cập nhật thông tin phiên", "broadcast_to": len(vids)}


@router.post("/batches/{batch_id}/cancel")
async def cancel_batch(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 1 — HUỶ cả đợt báo giá (buyer thoát hiểm).

    Chỉ huỷ được khi đợt CHƯA đi sâu (status ∈ draft/published). Đã sang
    evaluating/awarded/closed/cancelled → từ chối (có hợp đồng/kết quả rồi).
    Body: {"reason": str}  — lý do BẮT BUỘC (đi vào audit + thông báo NCC).

    Khi huỷ:
      * status → 'cancelled' (re-check trong WHERE để chống race với sweep
        tự-đóng-hạn).
      * ghi audit action='cancel'.
      * NHẮC HẠN tự dừng: cả auto-close lẫn reminder của tasks/procurement_
        deadlines.py đều `WHERE status='published'`, nên batch 'cancelled' không
        bao giờ khớp → không cần guard thêm.
      * THÔNG BÁO mọi NCC đang trong cuộc (invitation ∈ invited/viewed/submitted)
        "không cần báo giá nữa". Internal team chỉ nhận 1 bản (internal=False từ
        NCC thứ 2 trở đi → khử trùng).

    Lộ giá: `reason` do buyer gõ tay vào metadata gửi NCC — buyer tự kiểm KHÔNG
    chứa giá nội bộ/đối thủ. KHÔNG stamp target_price vào bất kỳ detail nào.
    """
    reason = (body.get("reason") or "").strip()
    if not reason:
        raise HTTPException(400, "Lý do huỷ là bắt buộc")

    batch = await conn.fetchrow(
        "SELECT id, status, batch_code, title FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] not in ("draft", "published"):
        raise HTTPException(
            400, "Chỉ huỷ được khi đợt ở trạng thái nháp hoặc đã công bố"
        )

    async with conn.transaction():
        # Re-check status trong WHERE → no-op nếu sweep vừa đẩy sang evaluating.
        updated = await conn.execute(
            "UPDATE procurement_rfq_batches SET status = 'cancelled', updated_at = NOW() "
            "WHERE id = $1 AND status IN ('draft', 'published')",
            batch_id,
        )
        if updated.endswith(" 0"):
            raise HTTPException(409, "Đợt vừa đổi trạng thái, vui lòng tải lại")

        await _audit(
            conn, "batch", batch_id, "cancel",
            actor_id=token_data.user_id,
            from_status=batch["status"], to_status="cancelled",
            detail={"reason": reason, "batch_code": batch["batch_code"]},
        )

        # Thông báo NCC đang trong cuộc. internal=True ở lần đầu (team nhận 1
        # bản), False các lần sau (chỉ ghi vendor row → không nhân bản nội bộ).
        invitees = await conn.fetch(
            "SELECT DISTINCT vendor_id FROM procurement_rfq_invitations "
            "WHERE batch_id = $1 AND status IN ('invited', 'viewed', 'submitted')",
            batch_id,
        )
        for idx, r in enumerate(invitees):
            await dispatch_procurement_event(
                conn, "batch", batch_id, "cancel",
                actor_id=token_data.user_id, awarded_vendor_id=r["vendor_id"],
                internal=(idx == 0),
                detail={
                    "batch_id": batch_id, "batch_code": batch["batch_code"],
                    "reason": reason,
                },
            )
        # Không có NCC nào đang trong cuộc → vẫn báo internal team 1 lần.
        if not invitees:
            await dispatch_procurement_event(
                conn, "batch", batch_id, "cancel",
                actor_id=token_data.user_id,
                detail={
                    "batch_id": batch_id, "batch_code": batch["batch_code"],
                    "reason": reason,
                },
            )

    logger.info(
        "Batch #%s (%s) cancelled by %s — %s",
        batch_id, batch["batch_code"], token_data.user_id, reason,
    )
    return {
        "data": {"id": batch_id, "status": "cancelled"},
        "message": "Đã huỷ đợt báo giá",
    }


@router.patch("/batches/{batch_id}/rank-hint")
async def set_batch_rank_hint(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt11 #15 — Bật/tắt gợi ý vị thế (band-mờ) cho NCC, PER-BATCH.

    NHẠY CẢM: khi BẬT, cổng NCC mới thấy band {dẫn đầu/giữa/cần cải thiện} SAU khi
    nộp — KHÔNG bao giờ thấy giá/tên/thứ hạng số đối thủ. Đây là quyết định sản phẩm,
    mặc định TẮT (migration default FALSE) → endpoint vendor trả 404 cho tới khi
    admin (là buyer) CHỦ ĐỘNG bật. Body: {"enabled": bool}.

    G2 đã chốt = hiện MỌI vòng ⇒ bật → rank_hint_round_from = 1; tắt → 9999 (an toàn).
    Ghi audit action='set_rank_hint'.
    """
    batch = await conn.fetchrow(
        "SELECT id, rank_hint_enabled, rank_hint_round_from FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    enabled = bool(body.get("enabled"))
    # Bật → lộ từ vòng 1 (G2 = mọi vòng); tắt → 9999 = không vòng nào (an toàn cứng).
    round_from = 1 if enabled else 9999

    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_rfq_batches "
            "SET rank_hint_enabled = $1, rank_hint_round_from = $2 WHERE id = $3",
            enabled, round_from, batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "set_rank_hint",
            actor_id=token_data.user_id,
            detail={
                "enabled": enabled,
                "round_from": round_from,
                "old_enabled": bool(batch["rank_hint_enabled"]),
            },
        )

    return {
        "data": {"rank_hint_enabled": enabled, "rank_hint_round_from": round_from},
        "message": (
            "Đã bật gợi ý vị thế cho NCC" if enabled else "Đã tắt gợi ý vị thế cho NCC"
        ),
    }


@router.post("/batches/{batch_id}/remind")
async def remind_batch_vendors(
    batch_id: int,
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P3 — Nhắc thủ công các NCC CHƯA báo giá (gửi lại email mời đăng nhập).

    Nhắm vào invitation của VÒNG ĐANG MỞ có status IN ('invited','viewed')
    (chưa 'submitted', không 'declined'). Đây là kênh THỦ CÔNG nên BỎ QUA guard
    reminder_sent_at, nhưng vẫn SET reminder_sent_at=NOW() để sweep tự động
    không gửi trùng ngay sau đó. Ghi audit action='manual_remind'.
    """
    batch = await conn.fetchrow(
        "SELECT id, status, batch_code, title, item_count, current_round "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] != "published":
        raise HTTPException(400, "Chỉ nhắc NCC khi đợt đang ở trạng thái đã công bố")

    cur_round = int(batch["current_round"] or 1)
    targets = await conn.fetch(
        """
        SELECT inv.id AS invitation_id, inv.vendor_id, va.company_name,
               va.contact_name, u.email
          FROM procurement_rfq_invitations inv
          JOIN vendor_accounts va ON va.id = inv.vendor_id
          LEFT JOIN users u ON u.id = va.user_id
         WHERE inv.batch_id = $1 AND inv.round_number = $2
           AND inv.status IN ('invited', 'viewed')
         ORDER BY va.company_name
        """,
        batch_id, cur_round,
    )

    # Thang 30/06: ĐÃ TẮT gửi email mời/nhắc — admin tự gửi link đăng nhập cho NCC.
    # Giữ endpoint để tương thích FE cũ; KHÔNG gửi email, không stamp email_sent.
    logger.info(
        "remind batch %d round %d: email disabled — no-op (%d targets chưa báo giá)",
        batch_id, cur_round, len(targets),
    )
    return {
        "data": {"sent": [], "failed": [], "round_number": cur_round, "targets": len(targets)},
        "message": "Đã tắt gửi email mời/nhắc — hãy gửi link đăng nhập trực tiếp cho NCC",
    }


@router.get("/batches/{batch_id}")
async def get_batch_admin(
    batch_id: int,
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Chi tiết đợt báo giá (admin view — bao gồm target_price, so sánh giá)."""
    batch = await conn.fetchrow("SELECT * FROM procurement_rfq_batches WHERE id = $1", batch_id)
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    # Đợt 2b [SB] — ĐIỂM DỄ SÓT: comparison[] (vqi.unit_price) + quotes[].total_amount
    # đều lộ giá admin-facing → gate khi niêm phong. SELECT * đã có cả 2 cột.
    sealed = _sealed_active(batch)

    # source_bqms_rfq_number (ADMIN-ONLY) JOIN để FE mở "File mã" từ thư mục Raw
    # của RFQ gốc. source_kind đã có sẵn trong i.*. JOIN theo id → 1 dòng (no twin).
    items = await conn.fetch(
        "SELECT i.*, br.rfq_number AS source_bqms_rfq_number "
        "FROM procurement_rfq_items i "
        "LEFT JOIN bqms_rfq br ON br.id = i.source_bqms_rfq_id "
        "WHERE i.batch_id = $1 ORDER BY i.item_no", batch_id
    )

    # Invited vendor accounts + per-vendor status
    invitations = await conn.fetch(
        """
        SELECT inv.vendor_id, va.company_name, inv.status,
               inv.invited_at, inv.viewed_at, inv.quoted_at, inv.email_status
        FROM procurement_rfq_invitations inv
        JOIN vendor_accounts va ON va.id = inv.vendor_id
        WHERE inv.batch_id = $1
        ORDER BY inv.invited_at DESC
        """,
        batch_id,
    )

    # Get all submitted quotes with vendor info
    quotes = await conn.fetch(
        """
        SELECT vq.id, vq.vendor_id, vq.currency, vq.total_amount, vq.status,
               vq.lead_time_days, vq.moq_notes, vq.notes, vq.round_number, vq.submitted_at,
               va.company_name AS vendor_name
        FROM vendor_quotes vq
        JOIN vendor_accounts va ON va.id = vq.vendor_id
        WHERE vq.batch_id = $1 AND vq.status = 'submitted'
        ORDER BY vq.total_amount ASC NULLS LAST
        """,
        batch_id,
    )

    # Get per-item quotes for comparison table
    # Đợt 2b [SB]: niêm phong → comparison rỗng (KHÔNG chạy vòng query lộ unit_price).
    comparison = []
    for item in ([] if sealed else items):
        item_quotes = await conn.fetch(
            """
            SELECT vqi.unit_price, vqi.quantity, vqi.lead_time_days, vqi.notes,
                   vqi.can_do, vqi.free_charge,
                   vq.vendor_id, vq.currency, va.company_name AS vendor_name
            FROM vendor_quote_items vqi
            JOIN vendor_quotes vq ON vq.id = vqi.quote_id
            JOIN vendor_accounts va ON va.id = vq.vendor_id
            WHERE vqi.item_id = $1 AND vq.status = 'submitted'
            -- Dòng có giá thật (>0, báo được, không FOC) xếp TRƯỚC và tăng dần
            -- (rẻ nhất lên đầu); FOC / "không làm" / giá 0 dồn xuống cuối để KHÔNG
            -- bị hiểu nhầm là "rẻ nhất". can_do/free_charge trả kèm để FE gắn nhãn.
            ORDER BY (vqi.can_do IS NOT FALSE
                      AND vqi.free_charge IS NOT TRUE
                      AND vqi.unit_price IS NOT NULL
                      AND vqi.unit_price > 0) DESC,
                     vqi.unit_price ASC
            """,
            item["id"],
        )
        comparison.append({
            "item": dict(item),
            "quotes": [dict(q) for q in item_quotes],
        })

    # Đợt 2b [SB]: quotes[].total_amount suy ra mặt bằng giá NCC → strip khi niêm
    # phong (GIỮ vendor + status để FE đếm đã-nộp). vendor_name/round/submitted GIỮ.
    quotes_out = [dict(q) for q in quotes]
    if sealed:
        for q in quotes_out:
            q["total_amount"] = None

    return {
        "data": {
            **dict(batch),  # sealed_until_deadline + bid_deadline đã trong * (FE đọc cho toggle/badge)
            "items": [dict(i) for i in items],
            "invitations": [dict(i) for i in invitations],
            "quotes": quotes_out,
            "comparison": comparison,
            "sealed": sealed,
        }
    }


# ---------------------------------------------------------------------------
# P5 — Item drawing: admin upload + authed download
# ---------------------------------------------------------------------------

async def _load_drawing_item(
    conn: asyncpg.Connection, batch_id: int, item_id: int
) -> asyncpg.Record:
    """Fetch the (item_id, batch_id) row + batch status, or 404.

    The item MUST belong to the batch in the path (prevents cross-batch id
    confusion). Returns the item's drawing_url + the batch status so callers can
    enforce the draft|published write-window without a second query.
    """
    row = await conn.fetchrow(
        """
        SELECT it.id, it.item_no, it.drawing_url, b.status AS batch_status
          FROM procurement_rfq_items it
          JOIN procurement_rfq_batches b ON b.id = it.batch_id
         WHERE it.id = $1 AND it.batch_id = $2
        """,
        item_id, batch_id,
    )
    if not row:
        raise HTTPException(404, "Mã hàng không thuộc đợt báo giá này")
    return row


@router.post("/batches/{batch_id}/items/{item_id}/drawing")
async def upload_item_drawing(
    batch_id: int,
    item_id: int,
    file: UploadFile = File(...),
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P5 — Tải lên bản vẽ cho 1 mã hàng (admin/manager/procurement).

    Chỉ cho upload khi đợt ở status='draft' hoặc 'published' (sau khi trao thầu
    -> view-only). Validate đuôi file (pdf/png/jpg/jpeg/dwg) + cap 20MB. Lưu sandboxed
    dưới FILES_BASE_PATH/drawings/{batch_id}/{item_id}/ và set
    procurement_rfq_items.drawing_url = 'file://drawings/{batch_id}/{item_id}/{name}'.
    KHÔNG đụng tới item bqms:// — endpoint này chỉ GHI scheme file://.
    """
    ext = _drawing_ext(file.filename)
    if not ext:
        raise HTTPException(400, "Chỉ chấp nhận PDF, PNG, JPG hoặc DWG")

    content = await file.read()
    if not content:
        raise HTTPException(400, "File rỗng")
    if len(content) > _DRAWING_MAX_BYTES:
        raise HTTPException(400, "File quá lớn (tối đa 20MB)")

    item = await _load_drawing_item(conn, batch_id, item_id)
    if item["batch_status"] not in ("draft", "published"):
        raise HTTPException(400, "Chỉ tải bản vẽ khi đợt ở trạng thái nháp hoặc đã công bố")

    # Sandboxed store: FILES_BASE_PATH/drawings/{batch_id}/{item_id}/drawing{ext}.
    # One canonical filename per item (a re-upload overwrites the previous file).
    rel = f"{batch_id}/{item_id}/drawing{ext}"
    dest = Path(settings.FILES_BASE_PATH) / "drawings" / rel
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(content)

    drawing_url = f"{_DRAWING_FILE_PREFIX}drawings/{rel}"
    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_rfq_items SET drawing_url = $1 WHERE id = $2",
            drawing_url, item_id,
        )
        await _audit(
            conn, "item", item_id, "drawing_upload",
            actor_id=token_data.user_id,
            detail={"batch_id": batch_id, "filename": file.filename,
                    "size": len(content), "ext": ext},
        )

    logger.info("Drawing uploaded for item %d (batch %d, %d bytes)", item_id, batch_id, len(content))
    return {
        "data": {"item_id": item_id, "drawing_url": drawing_url, "has_drawing": True},
        "message": "Đã tải lên bản vẽ",
    }


@router.get("/batches/{batch_id}/items/{item_id}/drawing")
async def download_item_drawing(
    batch_id: int,
    item_id: int,
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P5 — Xem/tải bản vẽ của 1 mã hàng (admin/manager/procurement/staff).

    Stream file:// sandboxed (path-guarded), redirect bqms:// -> /bqms/rfq/image,
    redirect http(s):// (legacy). 404 nếu mã hàng không có bản vẽ. Vendor được mời
    dùng endpoint song song ở /api/vendor/quotes/... (invitation-gated).
    """
    item = await _load_drawing_item(conn, batch_id, item_id)
    return _drawing_response(item["drawing_url"], item_no=item["item_no"])


# ---------------------------------------------------------------------------
# BQMS Import — push BQMS RFQ rows into a bid batch as items
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/import-from-bqms")
async def import_items_from_bqms(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Import items từ BQMS RFQ vào batch (Thang 2026-05-14).

    Body: {"rfq_ids": [int, ...]} HOẶC {"bqms_codes": [str, ...]}
    Mỗi BQMS row → 1 procurement_rfq_item với:
      - bqms_code, specification, quantity (từ expected_qty), unit
      - drawing_url = link tới ảnh BQMS qua /api/v1/bqms/rfq/image?code=<bqms_code>
      - source_bqms_rfq_id link gốc (admin only, không show vendor)
    """
    batch = await conn.fetchrow(
        "SELECT id, status FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] != "draft":
        raise HTTPException(400, "Chỉ import được khi batch ở status='draft'")
    # The draft-status guard + sequencing + dedupe + item_count are owned by
    # _insert_rfq_items; this pre-check stays only to short-circuit early.

    rfq_ids: list[int] = body.get("rfq_ids") or []
    codes: list[str] = body.get("bqms_codes") or []
    if not rfq_ids and not codes:
        raise HTTPException(400, "Cần ít nhất rfq_ids hoặc bqms_codes")

    where_parts = []
    params: list = []
    if rfq_ids:
        where_parts.append(f"id = ANY(${len(params)+1}::bigint[])")
        params.append(rfq_ids)
    if codes:
        where_parts.append(f"bqms_code = ANY(${len(params)+1}::text[])")
        params.append(codes)
    where = " OR ".join(where_parts)

    # Phase 2 (Thang 2026-05-14): pull full Samsung fields + JOIN staging.raw_json
    # cho part_no, cis_code, moq, maker, dimension, deadline, attachments.
    #
    # Twin-defence (Thang 2026-06-28): bqms_rfq chứa cặp (rfq_number, bqms_code)
    # trùng (twin etl vs onedrive_sync). Dù caller hiện gửi id của dòng
    # dedup-winner, endpoint TỰ miễn nhiễm twin bất kể caller: bọc nguồn bằng
    # DISTINCT ON (rfq_number, bqms_code) copy nguyên ORDER BY từ bqms_dedup CTE
    # (bqms.py) để mỗi (rfq_number, bqms_code) chỉ còn 1 dòng user-action.
    rfqs = await conn.fetch(
        f"""
        WITH bqms_dedup AS (
            SELECT DISTINCT ON (rfq_number, bqms_code) *
              FROM bqms_rfq
             ORDER BY rfq_number, bqms_code,
                      (COALESCE(quote_unlocked, false))::int DESC,
                      (quoted_price_bqms_v4 IS NOT NULL)::int DESC,
                      (quoted_price_bqms_v3 IS NOT NULL)::int DESC,
                      (quoted_price_bqms_v2 IS NOT NULL)::int DESC,
                      (quoted_price_bqms_v1 IS NOT NULL)::int DESC,
                      (bqms_push_status IS NOT NULL)::int DESC,
                      updated_at DESC NULLS LAST,
                      id DESC
        )
        SELECT r.id, r.bqms_code, r.specification, r.maker, r.expected_qty,
               r.unit, r.rfq_number, r.notes, r.requester, r.department,
               r.inquiry_date, r.person_in_charge_name,
               s.raw_json
        FROM bqms_dedup r
        LEFT JOIN LATERAL (
            SELECT raw_json FROM bqms_vendor_portal_staging
            WHERE rfq_number = r.rfq_number AND module='bidding'
            ORDER BY id DESC LIMIT 1
        ) s ON true
        WHERE ({where}) AND r.bqms_code IS NOT NULL
        """,
        *params,
    )
    if not rfqs:
        raise HTTPException(400, "Không tìm thấy mã BQMS hợp lệ để import")

    # Phase 2: enrich batch metadata từ first RFQ nếu batch còn trống.
    # Deferred (params only) so the UPDATE runs INSIDE the item-insert
    # transaction below — previously it auto-committed before the insert and
    # would orphan the enrichment if _insert_rfq_items raised.
    first_rfq = rfqs[0] if rfqs else None
    _enrich_params = None
    if first_rfq:
        first_raw = first_rfq["raw_json"] or {}
        if isinstance(first_raw, str):
            try:
                first_raw = _json.loads(first_raw)
            except Exception:
                first_raw = {}
        _enrich_params = (
            (first_raw.get("reqName") or "").strip() or None,
            first_rfq["requester"],
            first_rfq["department"],
            first_rfq["person_in_charge_name"],
            (first_raw.get("criteriaCurrency") or "").strip() or None,
            (first_raw.get("ctrTypeNm") or "").strip() or None,
            (first_raw.get("dday") or "").strip() or None,
            first_rfq["rfq_number"],
            batch_id,
        )

    # Build rows for the shared helper (dedupe on bqms_code via dedupe_code).
    rows: list[dict[str, Any]] = []
    for r in rfqs:
        code = r["bqms_code"]
        spec = (r["specification"] or "").strip() or code

        # Phase 2: extract per-item enriched fields từ raw_json._detail.items
        raw = r["raw_json"] or {}
        if isinstance(raw, str):
            try:
                raw = _json.loads(raw)
            except Exception:
                raw = {}
        detail = (raw.get("_detail") or {}) if isinstance(raw, dict) else {}
        items_arr = detail.get("items") or []
        matched_item: dict = {}
        for it in items_arr:
            if it.get("item_code") == code or it.get("cis_code") == code:
                matched_item = it
                break

        rows.append({
            "dedupe_code": code,  # BQMS dedupes on bqms_code, not item_code
            "specification": spec,
            "bqms_code": code,
            "quantity": float(r["expected_qty"] or 1),
            "unit": r["unit"] or "EA",
            "required_material": (
                matched_item.get("material") or matched_item.get("maker") or r["maker"]
            ),
            "drawing_url": f"bqms://{code}",
            # BẢO MẬT (Thang 2026-06-28): `notes` là cột VENDOR-VISIBLE (SELECT ở
            # public_bid.py + vendor/batches.py) → KHÔNG được ghi số RFQ Samsung
            # nội bộ vào đây. Nguồn gốc đã lưu an toàn ở source_bqms_rfq_id (item)
            # + source_bqms_rfq_number (batch), cả hai admin-only.
            "notes": None,
            "source_bqms_rfq_id": int(r["id"]),
            "source_ref_id": int(r["id"]),
            "maker": (matched_item.get("maker") or r["maker"]),
            "part_no": (matched_item.get("part_no") or "").strip() or None,
            "cis_code": (matched_item.get("cis_code") or "").strip() or None,
            "moq": (matched_item.get("moq") or "").strip() or None,
            "dimension": (matched_item.get("dimension") or "").strip() or None,
            "specification_full": (matched_item.get("specification") or "").strip() or None,
        })

    async with conn.transaction():
        if _enrich_params is not None:
            await conn.execute(
                """
                UPDATE procurement_rfq_batches SET
                    reg_dt            = COALESCE(reg_dt, NOW()),
                    req_name          = COALESCE(req_name, $1),
                    requester         = COALESCE(requester, $2),
                    department        = COALESCE(department, $3),
                    person_in_charge  = COALESCE(person_in_charge, $4),
                    criteria_currency = COALESCE(criteria_currency, $5),
                    ctr_type_name     = COALESCE(ctr_type_name, $6),
                    dday_text         = COALESCE(dday_text, $7),
                    source_bqms_rfq_number = COALESCE(source_bqms_rfq_number, $8)
                WHERE id = $9
                """,
                *_enrich_params,
            )
        result = await _insert_rfq_items(conn, batch_id, rows, "bqms")

    logger.info(
        "Imported %d BQMS items into batch %d (skipped %d duplicates)",
        result["imported"], batch_id, result["skipped"],
    )
    return {
        "data": {
            "imported": result["imported_ids"],
            "skipped_duplicates": result["skipped_codes"],
        },
        "message": f"Đã import {result['imported']} mã từ BQMS",
    }


# ---------------------------------------------------------------------------
# Commercial item sources (P2) — catalog / IMV / paste / Excel importers.
# All four route through the shared _insert_rfq_items helper (DRY): the
# draft-status guard, MAX(item_no) sequencing, per-(batch,item_code) dedupe,
# and item_count UPDATE are owned there, NOT duplicated here.
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/import-from-catalog")
async def import_from_catalog(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Import items từ Thư viện nguồn cung (sourcing_entries) vào batch.

    Body: {"sourcing_entry_ids": [int, ...]}
    Map sourcing_entries -> item_code/product_name/model/maker/specification_full/
    unit/quantity. source_kind='catalog', source_ref_id = sourcing_entries.id.
    Dedupe by (batch_id, item_code) inside the shared helper.
    """
    ids: list[int] = body.get("sourcing_entry_ids") or []
    ids = [int(x) for x in ids if x is not None]
    if not ids:
        raise HTTPException(400, "Cần ít nhất 1 sourcing_entry_id")

    entries = await conn.fetch(
        """
        SELECT id, bqms_code, model, product_name, maker, notes, quantity, hs_code
          FROM sourcing_entries
         WHERE id = ANY($1::bigint[]) AND deleted_at IS NULL
        """,
        ids,
    )
    if not entries:
        raise HTTPException(400, "Không tìm thấy mục nguồn cung hợp lệ để import")

    rows: list[dict[str, Any]] = []
    for e in entries:
        # item_code: no dedicated column on sourcing_entries — use bqms_code, else
        # the model, so dedupe still works for catalog rows that carry a code.
        item_code = (e["bqms_code"] or e["model"] or "").strip() or None
        rows.append({
            "item_code": item_code,
            "bqms_code": (e["bqms_code"] or "").strip() or None,
            "product_name": (e["product_name"] or "").strip() or None,
            "model": (e["model"] or "").strip() or None,
            "maker": (e["maker"] or "").strip() or None,
            "specification_full": (e["product_name"] or e["model"] or "").strip() or None,
            "unit": "EA",
            "quantity": float(e["quantity"]) if e["quantity"] is not None else 1,
            "notes": (e["notes"] or "").strip() or None,
            "source_ref_id": int(e["id"]),
        })

    async with conn.transaction():
        result = await _insert_rfq_items(conn, batch_id, rows, "catalog")

    logger.info(
        "Imported %d catalog items into batch %d (skipped %d)",
        result["imported"], batch_id, result["skipped"],
    )
    return {
        "imported": result["imported"],
        "skipped": result["skipped"],
        "data": {"imported_ids": result["imported_ids"], "skipped_codes": result["skipped_codes"]},
        "message": f"Đã import {result['imported']} mã từ thư viện nguồn cung",
    }


@router.post("/batches/{batch_id}/import-from-imv")
async def import_from_imv(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Import items từ IMV RFQ (imv_rfq) vào batch.

    Body: {"imv_rfq_ids": [int, ...]}
    imv_rfq is keyed (rfq_number, item_code) — many rows per RFQ, so we SELECT
    at ITEM granularity (one imv_rfq.id = one item). Map item_code/product_name/
    model/spec->specification_full/maker/unit/quantity/offered_qty (the item's
    quantity prefers offered_qty when present). due_date+due_time seeds the batch
    bid_deadline when empty; handler_name seeds batch.phu_trach when empty.
    source_kind='imv', source_ref_id = imv_rfq.id.
    """
    ids: list[int] = body.get("imv_rfq_ids") or []
    ids = [int(x) for x in ids if x is not None]
    if not ids:
        raise HTTPException(400, "Cần ít nhất 1 imv_rfq_id")

    recs = await conn.fetch(
        """
        SELECT id, rfq_number, item_code, product_name, model, spec, maker,
               unit, quantity, offered_qty, due_date, due_time, handler_name
          FROM imv_rfq
         WHERE id = ANY($1::bigint[])
        """,
        ids,
    )
    if not recs:
        raise HTTPException(400, "Không tìm thấy mục IMV hợp lệ để import")

    rows: list[dict[str, Any]] = []
    for r in recs:
        # For the item quantity prefer offered_qty (what we chào), else quantity.
        qty = r["offered_qty"] if r["offered_qty"] is not None else r["quantity"]
        rows.append({
            "item_code": (r["item_code"] or "").strip() or None,
            "product_name": (r["product_name"] or "").strip() or None,
            "model": (r["model"] or "").strip() or None,
            "specification_full": (r["spec"] or "").strip() or None,
            "maker": (r["maker"] or "").strip() or None,
            "unit": (r["unit"] or "EA"),
            "quantity": float(qty) if qty is not None else 1,
            # BẢO MẬT: `notes` là cột VENDOR-VISIBLE (SELECT ở vendor/batches.py)
            # → KHÔNG ghi số RFQ IMV (định danh khách hàng) vào đây, NCC dự thầu
            # sẽ thấy. Nguồn gốc lưu admin-only ở source_ref_id (→ imv_rfq.id) +
            # source_kind='imv'. Đồng bộ posture với import-from-bqms (notes=None).
            "notes": None,
            "source_ref_id": int(r["id"]),
        })

    # Seed batch deadline + phu_trach from the first IMV row IF empty (idempotent).
    first = recs[0]
    seed_deadline = None
    if first["due_date"] is not None:
        try:
            dt_time = (first["due_time"] or "").strip() or "00:00"
            # due_time may be 'HH:MM' or 'HH:MM:SS'; tolerate both.
            parts = dt_time.split(":")
            hh = int(parts[0]) if parts and parts[0].isdigit() else 0
            mm = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
            seed_deadline = datetime(
                first["due_date"].year, first["due_date"].month, first["due_date"].day,
                hh, mm, tzinfo=timezone(timedelta(hours=7)),
            )
        except Exception:
            seed_deadline = None

    async with conn.transaction():
        result = await _insert_rfq_items(conn, batch_id, rows, "imv")
        await conn.execute(
            """
            UPDATE procurement_rfq_batches
               SET bid_deadline = COALESCE(bid_deadline, $2),
                   phu_trach    = COALESCE(NULLIF(phu_trach, ''), $3)
             WHERE id = $1
            """,
            batch_id, seed_deadline,
            (first["handler_name"] or "").strip() or None,
        )

    logger.info(
        "Imported %d IMV items into batch %d (skipped %d)",
        result["imported"], batch_id, result["skipped"],
    )
    return {
        "imported": result["imported"],
        "skipped": result["skipped"],
        "data": {"imported_ids": result["imported_ids"], "skipped_codes": result["skipped_codes"]},
        "message": f"Đã import {result['imported']} mã từ IMV",
    }


def _parse_paste_rows(text: str) -> list[dict[str, Any]]:
    """Parse tab/newline-delimited paste into item rows. Never raises.

    Each non-empty line -> columns split on TAB (fallback: 2+ spaces). Column
    order: item_code, product_name, quantity, unit. A row that has NO code in
    column 0 still produces a row (it will be inserted as source_kind='manual'
    by the caller — degraded). Per-row try/except so one bad line can't kill all.
    """
    rows: list[dict[str, Any]] = []
    for raw_line in (text or "").splitlines():
        line = raw_line.rstrip("\r")
        if not line.strip():
            continue
        try:
            if "\t" in line:
                cols = [c.strip() for c in line.split("\t")]
            else:
                import re as _re
                cols = [c.strip() for c in _re.split(r"\s{2,}", line.strip())]
            cols += ["", "", "", ""]  # pad
            item_code = cols[0] or None
            product_name = cols[1] or None
            qty_raw = cols[2]
            unit = cols[3] or "EA"
            try:
                quantity = float(qty_raw.replace(",", "")) if qty_raw else 1
            except (ValueError, AttributeError):
                quantity = 1
            rows.append({
                "item_code": item_code,
                "product_name": product_name,
                "quantity": quantity,
                "unit": unit,
                "_has_code": bool(item_code),
            })
        except Exception:
            # Truly malformed line — skip it, never 500.
            continue
    return rows


@router.post("/batches/{batch_id}/import-paste")
async def import_paste(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Import items bằng cách DÁN bảng (tab/newline) vào batch.

    Body: {"text": "<paste>"}. Columns: item_code, product_name, quantity, unit.
    Rows WITHOUT a code still insert as source_kind='manual' (degraded). Parsing
    is per-row try/except so malformed input NEVER 500s.
    """
    text = body.get("text") or ""
    if not str(text).strip():
        raise HTTPException(400, "Cần dán ít nhất 1 dòng dữ liệu")

    parsed = _parse_paste_rows(str(text))
    if not parsed:
        raise HTTPException(400, "Không phân tích được dòng nào từ dữ liệu dán")

    coded = [r for r in parsed if r.pop("_has_code", False)]
    degraded_rows = [r for r in parsed if not r.get("item_code")]

    imported = 0
    degraded = 0
    skipped = 0
    async with conn.transaction():
        if coded:
            res_c = await _insert_rfq_items(conn, batch_id, coded, "paste")
            imported += res_c["imported"]
            skipped += res_c["skipped"]
        if degraded_rows:
            res_d = await _insert_rfq_items(conn, batch_id, degraded_rows, "manual")
            imported += res_d["imported"]
            degraded = res_d["imported"]

    logger.info(
        "Paste-imported %d items into batch %d (degraded %d, skipped %d)",
        imported, batch_id, degraded, skipped,
    )
    return {
        "imported": imported,
        "degraded": degraded,
        "skipped": skipped,
        "message": f"Đã import {imported} mã ({degraded} dòng thiếu mã -> thủ công"
                   + (f", {skipped} trùng bỏ qua" if skipped else "") + ")",
    }


# Excel header aliases (VI + EN, lower-cased) -> canonical row field.
_EXCEL_HEADER_ALIASES: dict[str, str] = {
    "mã hàng": "item_code", "ma hang": "item_code", "mã": "item_code",
    "item code": "item_code", "item_code": "item_code", "code": "item_code",
    "tên": "product_name", "ten": "product_name", "tên hàng": "product_name",
    "ten hang": "product_name", "product name": "product_name",
    "product_name": "product_name", "name": "product_name", "tên sản phẩm": "product_name",
    "model": "model",
    "spec": "specification_full", "specification": "specification_full",
    "thông số": "specification_full", "thong so": "specification_full",
    "maker": "maker", "hãng": "maker", "hang": "maker", "hãng sản xuất": "maker",
    "brand": "maker",
    "đvt": "unit", "dvt": "unit", "unit": "unit", "đơn vị": "unit", "don vi": "unit",
    "sl": "quantity", "số lượng": "quantity", "so luong": "quantity",
    "quantity": "quantity", "qty": "quantity",
}


@router.post("/batches/{batch_id}/import-excel")
async def import_excel(
    batch_id: int,
    file: UploadFile = File(...),
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Import items từ file Excel (.xlsx) vào batch.

    Multipart UploadFile. Header-mapped with VI+EN aliases (Mã hàng/item_code,
    Tên/product_name, Model, Spec, Maker/Hãng, ĐVT/unit, SL/quantity). Rows with
    NO item_code degrade to source_kind='manual'. Never 500s on a bad cell.
    Returns {imported, skipped, degraded_count}.
    """
    import io
    from openpyxl import load_workbook

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(400, "Chỉ chấp nhận file .xlsx")

    raw = await file.read()
    if not raw:
        raise HTTPException(400, "File rỗng")

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("import-excel: load_workbook failed: %s", exc)
        raise HTTPException(400, "Không đọc được file Excel (định dạng không hợp lệ)")

    ws = wb.active
    coded_rows: list[dict[str, Any]] = []
    degraded_rows: list[dict[str, Any]] = []

    try:
        rows_iter = ws.iter_rows(values_only=True)
        header_row = None
        for hr in rows_iter:
            if hr and any(c is not None and str(c).strip() for c in hr):
                header_row = hr
                break
        if header_row is None:
            wb.close()
            raise HTTPException(400, "File Excel không có dòng tiêu đề")

        # Map each column index -> canonical field via alias table.
        col_map: dict[int, str] = {}
        for idx, cell in enumerate(header_row):
            key = str(cell).strip().lower() if cell is not None else ""
            field = _EXCEL_HEADER_ALIASES.get(key)
            if field:
                col_map[idx] = field

        for data_row in rows_iter:
            try:
                if not data_row or not any(
                    c is not None and str(c).strip() for c in data_row
                ):
                    continue
                rec: dict[str, Any] = {}
                for idx, field in col_map.items():
                    if idx >= len(data_row):
                        continue
                    val = data_row[idx]
                    if val is None:
                        continue
                    sval = str(val).strip()
                    if not sval:
                        continue
                    if field == "quantity":
                        try:
                            rec["quantity"] = float(sval.replace(",", ""))
                        except (ValueError, AttributeError):
                            rec["quantity"] = 1
                    else:
                        rec[field] = sval
                if not rec:
                    continue
                rec.setdefault("unit", "EA")
                rec.setdefault("quantity", 1)
                if rec.get("item_code"):
                    coded_rows.append(rec)
                else:
                    degraded_rows.append(rec)
            except Exception as exc:  # never 500 on a bad cell
                logger.warning("import-excel: skipping malformed row: %s", exc)
                continue
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not coded_rows and not degraded_rows:
        raise HTTPException(400, "Không tìm thấy dòng dữ liệu hợp lệ trong file Excel")

    imported = 0
    degraded_count = 0
    skipped = 0
    async with conn.transaction():
        if coded_rows:
            res_c = await _insert_rfq_items(conn, batch_id, coded_rows, "excel")
            imported += res_c["imported"]
            skipped += res_c["skipped"]
        if degraded_rows:
            res_d = await _insert_rfq_items(conn, batch_id, degraded_rows, "manual")
            imported += res_d["imported"]
            degraded_count = res_d["imported"]

    logger.info(
        "Excel-imported %d items into batch %d (skipped %d, degraded %d)",
        imported, batch_id, skipped, degraded_count,
    )
    return {
        "imported": imported,
        "skipped": skipped,
        "degraded_count": degraded_count,
        "message": f"Đã import {imported} mã từ Excel ({degraded_count} dòng thiếu mã -> thủ công)",
    }


# ---------------------------------------------------------------------------
# Đợt 1 — Invitations via LOGIN vendor accounts (replaces magic-link)
# ---------------------------------------------------------------------------

@router.get("/batches/{batch_id}/invitations")
async def list_invitations(
    batch_id: int,
    round_number: int | None = Query(None),
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách NCC (tài khoản đăng nhập) được mời vào batch + status từng NCC.

    Mỗi dòng kèm `my_quote` = báo giá MỚI NHẤT của NCC đó cho batch (theo round
    nếu lọc), null nếu chưa báo giá. Thay thế danh sách magic-link cũ.
    """
    where = "inv.batch_id = $1"
    params: list = [batch_id]
    if round_number is not None:
        where += " AND inv.round_number = $2"
        params.append(round_number)

    # Đợt 2b [SB] — gate điểm rò bị bỏ sót: tab "NCC được mời" hiện my_quote.
    # total_amount per-NCC → mặt bằng giá. Khi niêm phong + chưa hạn phải ẩn
    # (giữ status/submitted_at để vẫn đếm đã-nộp). Dùng helper _sealed_active DUY NHẤT.
    _seal_b = await conn.fetchrow(
        "SELECT sealed_until_deadline, bid_deadline FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    sealed = bool(_seal_b) and _sealed_active(_seal_b)

    rows = await conn.fetch(
        f"""
        SELECT inv.id AS invitation_id, inv.vendor_id,
               va.company_name, va.contact_name, u.email,
               inv.round_number, inv.status,
               inv.invited_at, inv.viewed_at, inv.quoted_at,
               inv.declined_at, inv.decline_reason,
               inv.email_status, inv.email_sent_at,
               inv.reminder_sent_at, inv.missed_deadline,
               lq.id AS q_id, lq.total_amount AS q_total, lq.currency AS q_currency,
               lq.lead_time_days AS q_lead, lq.status AS q_status, lq.submitted_at AS q_submitted
        FROM procurement_rfq_invitations inv
        JOIN vendor_accounts va ON va.id = inv.vendor_id
        LEFT JOIN users u ON u.id = va.user_id
        LEFT JOIN LATERAL (
            SELECT id, total_amount, currency, lead_time_days, status, submitted_at
            FROM vendor_quotes vq
            WHERE vq.batch_id = inv.batch_id AND vq.vendor_id = inv.vendor_id
            ORDER BY vq.round_number DESC, vq.submitted_at DESC NULLS LAST, vq.id DESC
            LIMIT 1
        ) lq ON true
        WHERE {where}
        ORDER BY inv.invited_at DESC
        """,
        *params,
    )

    data = []
    for r in rows:
        d = dict(r)
        my_quote = None
        if d.pop("q_id") is not None:
            my_quote = {
                "id": r["q_id"],
                # SB: ẩn tổng giá + currency khi niêm phong (giữ status/submitted_at).
                "total_amount": (None if sealed else r["q_total"]),
                "currency": (None if sealed else r["q_currency"]),
                "lead_time_days": (None if sealed else r["q_lead"]),
                "status": r["q_status"],
                "submitted_at": r["q_submitted"],
            }
        for k in ("q_total", "q_currency", "q_lead", "q_status", "q_submitted"):
            d.pop(k, None)
        d["my_quote"] = my_quote
        data.append(d)

    return {"data": data, "sealed": sealed}


@router.post("/batches/{batch_id}/invite")
async def invite_vendors(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt-1 CORE — mời NCC qua TÀI KHOẢN ĐĂNG NHẬP (không phải magic-link token).

    Tạo/upsert procurement_rfq_invitations cho từng vendor_accounts.id và gửi
    email chứa LINK ĐĂNG NHẬP cổng NCC (ncc.songchau.vn/login?next=/batches/{id}).

    Body: {
        "vendor_ids": [int, ...],   # vendor_accounts.id (bắt buộc)
        "round_number": int = 1,
        "send_email": bool = true
    }
    """
    batch = await conn.fetchrow(
        "SELECT id, status, batch_code, title, item_count FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] not in ("draft", "published"):
        raise HTTPException(400, "Chỉ mời được khi batch ở status='draft' hoặc 'published'")
    if (batch["item_count"] or 0) == 0:
        raise HTTPException(400, "Batch chưa có items — push BQMS hoặc thêm thủ công trước")

    vendor_ids: list[int] = body.get("vendor_ids") or []
    if not vendor_ids:
        raise HTTPException(400, "Cần ít nhất 1 vendor_id")
    # Dedupe while preserving order
    vendor_ids = list(dict.fromkeys(int(v) for v in vendor_ids))

    round_number = int(body.get("round_number") or 1)
    # Thang 30/06: BỎ gửi mail mời — admin tự gửi link đăng nhập cổng NCC cho NCC.
    # Vẫn TẠO invitation (quyền truy cập) + thông báo trong cổng; chỉ KHÔNG email.
    do_send_email = False

    base = _vendor_portal_base()
    created: list[dict] = []
    skipped_existing: list[int] = []
    failures: list[dict] = []

    # Auto-publish if still draft (preserve old invite behavior). P7: when the
    # internal-approval gate is ON, a draft must NOT be auto-published by the
    # invite shortcut — it would bypass the gate. Block with the same 400 the
    # publish endpoint raises so the flow stays consistent.
    if batch["status"] == "draft":
        _appr = await _read_approval_config(conn)
        if _appr["approval_required"]:
            raise HTTPException(400, "Phiên phải được duyệt nội bộ trước khi công bố")
        async with conn.transaction():
            # Synthetic approval row (gate OFF) so audit/downstream stay uniform.
            await conn.execute(
                "UPDATE procurement_rfq_batches "
                "SET approved_by = NULL, approval_auto = true, approved_at = NOW() "
                "WHERE id = $1 AND approved_at IS NULL",
                batch_id,
            )
            await conn.execute(
                "UPDATE procurement_rfq_batches SET status='published', published_at=NOW() "
                "WHERE id=$1 AND status='draft'",
                batch_id,
            )
            await _audit(
                conn, "batch", batch_id, "publish",
                actor_id=token_data.user_id, from_status="draft", to_status="published",
                detail={"approval_auto": True, "via": "invite"},
            )

    for vid in vendor_ids:
        va = await conn.fetchrow(
            """
            SELECT va.id, va.company_name, va.contact_name, va.status, va.is_approved, u.email
            FROM vendor_accounts va
            LEFT JOIN users u ON u.id = va.user_id
            WHERE va.id = $1
            """,
            vid,
        )
        if not va:
            failures.append({"vendor_id": vid, "error": "Tài khoản NCC không tồn tại"})
            continue
        is_active = str(va["status"]) == "active" or va["is_approved"] is True
        if not is_active:
            failures.append({"vendor_id": vid, "error": "Tài khoản NCC chưa active"})
            continue

        # Idempotent upsert — ON CONFLICT (batch_id,vendor_id,round_number) DO NOTHING
        async with conn.transaction():
            invitation_id = await conn.fetchval(
                """
                INSERT INTO procurement_rfq_invitations
                    (batch_id, vendor_id, round_number, status, invited_by, invited_at)
                VALUES ($1, $2, $3, 'invited', $4, NOW())
                ON CONFLICT (batch_id, vendor_id, round_number) DO NOTHING
                RETURNING id
                """,
                batch_id, vid, round_number, token_data.user_id,
            )
            if invitation_id is not None:
                await _audit(
                    conn, "invitation", invitation_id, "invite",
                    actor_id=token_data.user_id,
                    detail={"vendor_id": vid, "round": round_number, "email": va["email"]},
                )
                # In-portal notification addressed to THIS vendor (NCC-lite) via
                # awarded_vendor_id → recipient_vendor_id. Best-effort, same txn.
                await dispatch_procurement_event(
                    conn, "invitation", invitation_id, "invite",
                    actor_id=token_data.user_id, awarded_vendor_id=vid,
                    detail={
                        "batch_id": batch_id,
                        "batch_code": batch["batch_code"],
                        "round": round_number,
                    },
                )
        if invitation_id is None:
            skipped_existing.append(vid)
            continue

        email = (va["email"] or "").strip().lower()
        company = va["company_name"]
        rec: dict[str, Any] = {
            "invitation_id": invitation_id,
            "vendor_id": vid,
            "company_name": company,
            "email": email or None,
            "email_sent": False,
        }

        if do_send_email and email and "@" in email:
            login_url = f"{base}/login?next=/batches/{batch_id}"
            subject = f"[Song Châu] Mời báo giá phiên #{batch['batch_code']} — {batch['title']}"
            body_html = _build_login_invitation_email(
                batch_code=batch["batch_code"],
                batch_title=batch["title"],
                invitee_name=va["contact_name"] or company or email,
                login_url=login_url,
                item_count=batch["item_count"] or 0,
            )
            try:
                await send_email([email], subject, body_html)
                await conn.execute(
                    """UPDATE procurement_rfq_invitations
                       SET email_sent=true, email_sent_at=NOW(),
                           email_status='sent', email_subject=$1
                       WHERE id=$2""",
                    subject, invitation_id,
                )
                rec["email_sent"] = True
            except Exception as exc:
                logger.warning("Email mời NCC %s thất bại: %s", vid, exc)
                await conn.execute(
                    """UPDATE procurement_rfq_invitations
                       SET email_status='failed', email_error=$1
                       WHERE id=$2""",
                    str(exc)[:500], invitation_id,
                )
                rec["email_error"] = str(exc)[:200]
        elif do_send_email:
            rec["email_error"] = "NCC không có email hợp lệ"

        created.append(rec)

    logger.info(
        "Invited %d login vendors to batch %d round %d (skipped %d, failures %d)",
        len(created), batch_id, round_number, len(skipped_existing), len(failures),
    )
    return {
        "data": {
            "created": created,
            "skipped_existing": skipped_existing,
            "failures": failures,
        },
        "message": f"Đã mời {len(created)} nhà cung cấp",
    }


def _build_login_invitation_email(
    batch_code: str, batch_title: str, invitee_name: str,
    login_url: str, item_count: int, round_number: int = 1,
) -> str:
    """HTML email body cho lời mời báo giá qua tài khoản ĐĂNG NHẬP (Đợt 1).

    KHÁC magic-link: nút trỏ tới trang ĐĂNG NHẬP cổng NCC, không phải URL token.
    round_number > 1 (Đợt 2 mở vòng V2/V3) hiển thị badge "Vòng N".
    """
    round_badge = (
        f'<p style="margin: 4px 0 0; opacity: 0.9; font-size: 12px;">Vòng {round_number} — mời báo giá lại</p>'
        if round_number > 1 else ""
    )
    return f"""<!DOCTYPE html><html><body style="font-family: -apple-system, Segoe UI, Roboto, sans-serif; background: #f8fafc; padding: 24px;">
<div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.05);">
  <div style="background: linear-gradient(135deg, #4f46e5, #7c3aed); color: white; padding: 24px;">
    <h1 style="margin: 0; font-size: 22px;">Song Châu mời báo giá</h1>
    <p style="margin: 6px 0 0; opacity: 0.9; font-size: 13px;">Phiên đấu thầu #{batch_code}</p>
    {round_badge}
  </div>
  <div style="padding: 24px;">
    <p style="margin: 0 0 16px; color: #1e293b; font-size: 14px;">Kính gửi <strong>{invitee_name}</strong>,</p>
    <p style="margin: 0 0 16px; color: #475569; font-size: 14px; line-height: 1.6;">
      Công ty Song Châu trân trọng mời Quý đơn vị tham gia báo giá cho phiên đấu thầu:
    </p>
    <div style="background: #f8fafc; border-left: 4px solid #4f46e5; padding: 14px 18px; border-radius: 6px; margin: 16px 0;">
      <p style="margin: 0; font-weight: 600; color: #1e293b; font-size: 15px;">{batch_title}</p>
      <p style="margin: 6px 0 0; color: #64748b; font-size: 12px;">{item_count} mã linh kiện · Mã phiên: {batch_code}</p>
    </div>
    <p style="margin: 0 0 18px; color: #475569; font-size: 14px;">
      Quý đơn vị vui lòng <strong>đăng nhập cổng nhà cung cấp Song Châu</strong> để xem chi tiết và gửi báo giá.
    </p>
    <div style="text-align: center; margin: 24px 0;">
      <a href="{login_url}" style="display: inline-block; background: linear-gradient(135deg, #4f46e5, #7c3aed); color: white; padding: 14px 36px; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 15px;">
        Đăng nhập & báo giá
      </a>
    </div>
    <p style="margin: 16px 0 0; color: #94a3b8; font-size: 11px; text-align: center;">
      Nếu nút không hoạt động, mở liên kết: <span style="color: #64748b;">{login_url}</span><br>
      Chưa có tài khoản? Liên hệ bộ phận mua hàng Song Châu để được cấp.
    </p>
  </div>
  <div style="background: #f1f5f9; padding: 14px 24px; color: #64748b; font-size: 11px; text-align: center;">
    Song Châu ERP · Email tự động, vui lòng không trả lời
  </div>
</div>
</body></html>"""


def _build_onboard_email(invitee_name: str, activation_link: str) -> str:
    """HTML email mời NCC MỚI kích hoạt tài khoản (Đợt 1 — admin-invite).

    KHÁC _build_login_invitation_email: nút trỏ tới link KÍCH HOẠT mang token
    (đặt mật khẩu lần đầu), không phải trang đăng nhập.
    """
    return f"""<!DOCTYPE html><html><body style="font-family: -apple-system, Segoe UI, Roboto, sans-serif; background: #f8fafc; padding: 24px;">
<div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.05);">
  <div style="background: linear-gradient(135deg, #4f46e5, #7c3aed); color: white; padding: 24px;">
    <h1 style="margin: 0; font-size: 22px;">Mời tham gia cổng Nhà Cung Cấp</h1>
    <p style="margin: 6px 0 0; opacity: 0.9; font-size: 13px;">Song Châu ERP</p>
  </div>
  <div style="padding: 24px;">
    <p style="margin: 0 0 16px; color: #1e293b; font-size: 14px;">Kính gửi <strong>{invitee_name}</strong>,</p>
    <p style="margin: 0 0 16px; color: #475569; font-size: 14px; line-height: 1.6;">
      Công ty Song Châu trân trọng mời Quý đơn vị tham gia cổng nhà cung cấp.
      Vui lòng nhấn nút bên dưới để <strong>kích hoạt tài khoản và đặt mật khẩu</strong>.
    </p>
    <div style="text-align: center; margin: 24px 0;">
      <a href="{activation_link}" style="display: inline-block; background: linear-gradient(135deg, #4f46e5, #7c3aed); color: white; padding: 14px 36px; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 15px;">
        Kích hoạt tài khoản
      </a>
    </div>
    <p style="margin: 16px 0 0; color: #94a3b8; font-size: 11px; text-align: center;">
      Nếu nút không hoạt động, mở liên kết: <span style="color: #64748b;">{activation_link}</span><br>
      Liên kết kích hoạt có hiệu lực trong 7 ngày.
    </p>
  </div>
  <div style="background: #f1f5f9; padding: 14px 24px; color: #64748b; font-size: 11px; text-align: center;">
    Song Châu ERP · Email tự động, vui lòng không trả lời
  </div>
</div>
</body></html>"""


@router.post("/vendors/invite")
async def onboard_vendor(
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 1 — ONBOARD 1 NCC MỚI (admin chủ động mời, KHÁC tự đăng ký).

    Tạo users(role=vendor, is_active=false) + vendor_accounts(status='pending',
    activation_token TTL 7 ngày). NCC mở link ``/activate/{token}`` (endpoint
    activate ĐÃ CÓ) để đặt mật khẩu lần đầu + active.

    Body: {email*, company_name*, contact_name*, phone?, tax_code?,
           product_categories?: list[str]}.

    EMAIL CHƯA LIVE (M365 trống): endpoint LUÔN trả về `activation_link` để admin
    tự gửi tay; vẫn THỬ gửi email best-effort (try/except, không fail nếu lỗi).
    Idempotent-ish: email đã tồn tại → 400.
    """
    for f in ("email", "company_name", "contact_name"):
        if not (body.get(f) or "").strip():
            raise HTTPException(400, f"Trường '{f}' là bắt buộc")
    email = body["email"].strip().lower()

    existing = await conn.fetchval("SELECT id FROM users WHERE email = $1", email)
    if existing:
        raise HTTPException(400, "Email đã tồn tại")

    token = _gen_bid_token()
    expires = datetime.now(timezone.utc) + timedelta(days=7)
    contact_name = body["contact_name"].strip()

    async with conn.transaction():
        # hashed_password NOT NULL → đặt placeholder '!' (KHÔNG phải bcrypt hợp lệ
        # → verify_password luôn fail) → NCC pending không thể đăng nhập tới khi
        # activate đặt mật khẩu thật. Không để NULL (vi phạm NOT NULL).
        user_id = await conn.fetchval(
            "INSERT INTO users (email, hashed_password, full_name, role, is_active) "
            "VALUES ($1, '!', $2, 'vendor'::role_enum, false) RETURNING id",
            email, contact_name,
        )
        await conn.execute(
            """
            INSERT INTO vendor_accounts
                (user_id, company_name, contact_name, phone, tax_code,
                 product_categories, status, invited_by,
                 activation_token, activation_expires)
            VALUES ($1, $2, $3, $4, $5, $6, 'pending', $7, $8, $9)
            """,
            user_id, body["company_name"].strip(), contact_name,
            (body.get("phone") or "").strip() or None,
            (body.get("tax_code") or "").strip() or None,
            body.get("product_categories") or None,
            token_data.user_id, token, expires,
        )

    activation_link = f"{_vendor_portal_base()}/activate/{token}"
    email_sent = False
    try:
        await send_email(
            [email],
            "[Song Châu] Mời tham gia cổng Nhà Cung Cấp",
            _build_onboard_email(contact_name, activation_link),
        )
        email_sent = True
    except Exception as exc:  # noqa: BLE001 — email best-effort, link vẫn trả về
        logger.warning("Onboard email NCC %s thất bại (best-effort): %s", email, exc)

    logger.info("Vendor onboarded by %s: %s (%s)", token_data.user_id, email,
                body["company_name"])
    return {
        "data": {
            "user_id": user_id,
            "activation_link": activation_link,
            "expires_at": expires.isoformat(),
            "email_sent": email_sent,
        },
        "message": "Đã tạo tài khoản NCC. Gửi link kích hoạt cho họ.",
    }


# ---------------------------------------------------------------------------
# Đợt 1 — Quote comparison MATRIX (items × invited vendors)
# ---------------------------------------------------------------------------

# M3 — cache grade scorecard NCC (TTL per-process). Scorecard là truy vấn nặng
# (nhiều CTE/time-window); ma trận là hot-path nên CACHE để KHÔNG tính lại mỗi
# lần mở. Grade là toàn cục (không theo lô) nên 1 cache dùng chung mọi ma trận/drawer.
_GRADE_CACHE: dict[str, Any] = {"at": 0.0, "data": {}}
_GRADE_CACHE_TTL = 600  # 10 phút


async def _get_vendor_grades(conn: asyncpg.Connection) -> dict[int, dict]:
    """Trả {vendor_id: {grade('A'|'B'|'C'|None), on_time_rate, score}} có cache TTL.

    Tính lại bằng _scorecard_factors (1 lần cho MỌI NCC — tránh N+1) khi cache hết
    hạn; lỗi thì degrade về cache cũ / rỗng (không bao giờ làm vỡ ma trận).
    """
    now = _time.monotonic()
    if _GRADE_CACHE["data"] and (now - _GRADE_CACHE["at"] < _GRADE_CACHE_TTL):
        return _GRADE_CACHE["data"]
    try:
        from app.api.v1.procurement_analytics import (
            _grade, _score_from_factors, _scorecard_factors,
        )
        factors_by_vendor = await _scorecard_factors(conn, 12, 3)
        out: dict[int, dict] = {}
        for vid, v in factors_by_vendor.items():
            score = _score_from_factors(v["factors"])[0] if v["sufficient"] else None
            g = _grade(score)
            ot = v["factors"].get("on_time")
            out[vid] = {
                "grade": g if g in ("A", "B", "C") else None,
                "on_time_rate": (float(ot) * 100.0) if ot is not None else None,
                "score": float(score) if score is not None else None,
            }
        _GRADE_CACHE["data"] = out
        _GRADE_CACHE["at"] = now
        return out
    except Exception:
        logger.warning("vendor grade cache compute failed; degrading", exc_info=True)
        return _GRADE_CACHE["data"] or {}


@router.get("/batches/{batch_id}/matrix")
async def batch_matrix(
    batch_id: int,
    round_number: int | None = Query(None),
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt-1 CORE — ma trận so sánh báo giá.

    Hàng = procurement_rfq_item, cột = NCC được mời (từ procurement_rfq_invitations).
    Mỗi ô = báo giá MỚI NHẤT của NCC đó cho item đó (DISTINCT ON vendor_id ORDER BY
    round_number DESC, submitted_at DESC). Highlight đơn giá thấp nhất theo từng
    loại tiền trong cùng hàng. target_price CHỈ có ở route admin này, không bao giờ
    lộ ra path vendor.
    """
    batch = await conn.fetchrow(
        "SELECT id, batch_code, title, status, award_mode, bid_deadline, sealed_until_deadline "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    # Đợt 2b [SB] — niêm phong: TÍNH 1 LẦN, dùng để mask mọi con số bên dưới.
    sealed = _sealed_active(batch)

    # Determine effective round: explicit, else latest invitation round present.
    eff_round = round_number
    if eff_round is None:
        eff_round = await conn.fetchval(
            "SELECT MAX(round_number) FROM procurement_rfq_invitations WHERE batch_id = $1",
            batch_id,
        ) or 1

    # Columns = vendors invited to this batch (for the given round) + their LATEST
    # submitted quote (across rounds, latest-wins).
    vendor_rows = await conn.fetch(
        """
        SELECT inv.vendor_id, va.company_name, inv.status AS inv_status,
               lq.id AS quote_id, lq.currency, lq.lead_time_days,
               lq.total_amount, lq.submitted_at
        FROM procurement_rfq_invitations inv
        JOIN vendor_accounts va ON va.id = inv.vendor_id
        LEFT JOIN LATERAL (
            SELECT id, currency, lead_time_days, total_amount, submitted_at
            FROM vendor_quotes vq
            WHERE vq.batch_id = inv.batch_id AND vq.vendor_id = inv.vendor_id
              AND vq.status = 'submitted'
            ORDER BY vq.round_number DESC, vq.submitted_at DESC NULLS LAST, vq.id DESC
            LIMIT 1
        ) lq ON true
        WHERE inv.batch_id = $1 AND inv.round_number = $2
        ORDER BY va.company_name
        """,
        batch_id, eff_round,
    )

    grades = await _get_vendor_grades(conn)  # M3 — hạng A/B/C (cache TTL)

    # Đợt 4 — FX normalize: build rate map 1 lần/batch, as-of bid_deadline.
    # Niêm phong (sealed) → KHÔNG tính FX (giá đã ẩn — không được lách qua VND).
    as_of = batch["bid_deadline"].date() if batch["bid_deadline"] else None
    fx_map: dict[str, Decimal | None] = {}
    fx_missing_count = 0
    if not sealed:
        currencies = {v["currency"] for v in vendor_rows if v["currency"]}
        fx_map = await _fx_map_for_batch(conn, currencies, as_of)

    vendors = []
    quote_ids: dict[int, int] = {}  # vendor_id -> latest quote_id
    for v in vendor_rows:
        vid = v["vendor_id"]
        if v["quote_id"] is not None:
            quote_ids[vid] = v["quote_id"]
        # Đợt 4 — tổng quy đổi VND per-vendor (additive). total_amount đã bị ép
        # None khi sealed → _v_total None → _v_vnd tự None (sealed gate giữ).
        _v_total = (
            None if sealed
            else (float(v["total_amount"]) if v["total_amount"] is not None else None)
        )
        _v_rate = fx_map.get(v["currency"]) if v["currency"] else None
        _v_vnd = _vnd(_v_total, _v_rate)
        _v_fx_missing = (
            not sealed and _v_total is not None and bool(v["currency"]) and _v_rate is None
        )
        vendors.append({
            "vendor_id": vid,
            "company_name": v["company_name"],
            "inv_status": v["inv_status"],
            "quote_id": v["quote_id"],
            "currency": v["currency"],
            "lead_time_days": v["lead_time_days"],
            # Đợt 2b [SB]: total_amount suy ra mặt bằng giá NCC → ẩn khi niêm phong.
            # inv_status/submitted_at/currency GIỮ (FE đếm đã-nộp, không lộ giá).
            "total_amount": _v_total,
            # Đợt 4 — quy đổi VND tổng (None khi sealed/thiếu rate). KHÔNG đổi gốc.
            "vnd_equiv_total": _v_vnd,
            "fx_missing": _v_fx_missing,
            "submitted_at": v["submitted_at"],
            "grade": grades.get(vid, {}).get("grade"),  # M3
        })

    # source_kind + source_bqms_rfq_number (ADMIN-ONLY): FE phân biệt mã BQMS vs
    # Nguồn cung + mở "File mã" từ thư mục Raw của RFQ gốc. JOIN br.id = source id
    # (1 dòng, KHÔNG nổ twin). rfq_number TUYỆT ĐỐI không lộ ra path vendor.
    items = await conn.fetch(
        """
        SELECT i.id, i.item_no, i.specification, i.bqms_code, i.quantity, i.unit,
               i.required_material, i.target_price, i.source_kind,
               i.awarded_vendor_id, i.awarded_price, i.awarded_currency,
               br.rfq_number AS source_bqms_rfq_number
        FROM procurement_rfq_items i
        LEFT JOIN bqms_rfq br ON br.id = i.source_bqms_rfq_id
        WHERE i.batch_id = $1
        ORDER BY i.item_no
        """,
        batch_id,
    )

    # Pull all line items for the latest quotes in one query, keyed (quote_id,item_id).
    line_map: dict[tuple[int, int], dict] = {}
    if quote_ids:
        line_rows = await conn.fetch(
            """
            SELECT vqi.quote_id, vqi.item_id, vqi.unit_price, vqi.quantity,
                   vqi.lead_time_days, vqi.notes, vqi.can_do, vqi.free_charge,
                   vqi.offered_qty, vqi.moq,
                   (jsonb_array_length(COALESCE(vqi.attachment_paths, '[]'::jsonb)) > 0) AS has_attachment
            FROM vendor_quote_items vqi
            WHERE vqi.quote_id = ANY($1::bigint[])
            """,
            list(quote_ids.values()),
        )
        for lr in line_rows:
            line_map[(lr["quote_id"], lr["item_id"])] = dict(lr)

    # M4 — giá VÒNG TRƯỚC mỗi (NCC, mã) để tính Δ giảm-giá hiện ngay tại ô matrix.
    # Chỉ khi đang ở vòng > 1 (reverse-auction). DISTINCT ON lấy vòng gần nhất < eff_round.
    prior_map: dict[tuple[int, int], float] = {}
    if eff_round and eff_round > 1:
        prior_rows = await conn.fetch(
            """
            SELECT DISTINCT ON (vq.vendor_id, vqi.item_id)
                   vq.vendor_id, vqi.item_id, vqi.unit_price
            FROM vendor_quotes vq
            JOIN vendor_quote_items vqi ON vqi.quote_id = vq.id
            WHERE vq.batch_id = $1 AND vq.status = 'submitted' AND vq.round_number < $2
            ORDER BY vq.vendor_id, vqi.item_id,
                     vq.round_number DESC, vq.submitted_at DESC NULLS LAST, vq.id DESC
            """,
            batch_id, eff_round,
        )
        for pr in prior_rows:
            if pr["unit_price"] is not None:
                prior_map[(pr["vendor_id"], pr["item_id"])] = float(pr["unit_price"])

    # quote_id -> currency (for cell currency, since unit prices share the quote's currency)
    quote_currency = {v["quote_id"]: v["currency"] for v in vendor_rows if v["quote_id"] is not None}

    out_items = []
    for it in items:
        item_id = it["id"]
        item_qty = float(it["quantity"]) if it["quantity"] is not None else None
        cells: dict[str, dict] = {}
        # Collect (vendor_id, unit_price, currency) for lowest computation.
        priced: list[tuple[int, float, str]] = []

        for vid, qid in quote_ids.items():
            line = line_map.get((qid, item_id))
            cur = quote_currency.get(qid)
            if sealed:
                # Đợt 2b [SB] — NIÊM PHONG: chỉ báo "có/không có dòng báo giá",
                # giấu MỌI con số. priced[] không nạp → khối lowest tự no-op.
                cells[str(vid)] = {
                    "unit_price": None, "quantity": None, "currency": None,
                    "lead_time_days": None, "notes": None, "line_total": None,
                    "is_lowest": False, "can_do": None, "is_foc": False,
                    "offered_qty": None, "moq": None,
                    "has_attachment": bool(line is not None and line["has_attachment"]),
                    "delta": None, "delta_pct": None, "prior_unit_price": None,
                    "sealed": True,
                }
                continue
            if line is None:
                cells[str(vid)] = {
                    "unit_price": None, "quantity": None, "currency": None,
                    "lead_time_days": None, "notes": None,
                    "line_total": None, "is_lowest": False, "can_do": None,
                    "is_foc": False,
                    "offered_qty": None, "moq": None, "has_attachment": False,
                    "delta": None, "delta_pct": None, "prior_unit_price": None,
                }
                continue
            unit_price = float(line["unit_price"]) if line["unit_price"] is not None else None
            cell_qty = float(line["quantity"]) if line["quantity"] is not None else None
            eff_qty = cell_qty if cell_qty is not None else item_qty
            line_total = (unit_price * eff_qty) if (unit_price is not None and eff_qty is not None) else None
            # FOC (miễn phí) = NCC cam kết cung cấp dòng này với giá 0. KHÔNG phải
            # một báo giá có thể so sánh — nên KHÔNG được tính là "giá thấp nhất".
            is_foc = bool(line["free_charge"])
            # M4 — Δ so vòng trước (chỉ giá thật, báo được, không FOC, có giá vòng trước).
            prior_p = prior_map.get((vid, item_id))
            delta = (
                (unit_price - prior_p)
                if (unit_price is not None and unit_price > 0 and not is_foc
                    and line["can_do"] is not False and prior_p is not None and prior_p > 0)
                else None
            )
            delta_pct = (delta / prior_p * 100.0) if (delta is not None and prior_p) else None
            # Đợt 4 — quy đổi VND per-cell (additive). FOC (giá 0) KHÔNG quy đổi.
            # rate=None (ngoại tệ thiếu rate) → vnd_equiv None + fx_missing → FE cảnh báo.
            # VND-bid: rate=1 → vnd_equiv == unit_price (vô hại, default behavior).
            _rate = fx_map.get(cur) if cur else None
            _cell_vnd = None if is_foc else _vnd(unit_price, _rate)
            _cell_vnd_total = None if is_foc else _vnd(line_total, _rate)
            _cell_fx_missing = (
                not is_foc and unit_price is not None and cur is not None and _rate is None
            )
            if _cell_fx_missing:
                fx_missing_count += 1
            cells[str(vid)] = {
                "unit_price": unit_price,
                "quantity": cell_qty,
                "currency": cur,
                "lead_time_days": line["lead_time_days"],
                "notes": line["notes"],
                "line_total": line_total,
                # Đợt 4 — VND-equiv SONG SONG (KHÔNG đổi unit_price/line_total gốc).
                "vnd_equiv": _cell_vnd,
                "vnd_line_total": _cell_vnd_total,
                "fx_missing": _cell_fx_missing,
                "is_lowest": False,  # filled below
                "can_do": line["can_do"],
                "is_foc": is_foc,
                # M2 — surface trực tiếp tại ô để admin quyết không cần mở drawer.
                "offered_qty": float(line["offered_qty"]) if line["offered_qty"] is not None else None,
                "moq": line["moq"],
                "has_attachment": bool(line["has_attachment"]),
                # M4 — Δ vòng trước (None ở vòng 1 hoặc không có mốc).
                "delta": delta,
                "delta_pct": delta_pct,
                "prior_unit_price": prior_p,
            }
            # Chỉ những dòng CÓ giá thật (>0), NCC báo được (can_do != False) và
            # KHÔNG phải FOC mới tham gia so sánh "giá thấp nhất". Bịt bug: giá 0
            # (FOC hoặc nhập nhầm) và dòng "không làm được" trước đây bị tô xanh.
            if (
                unit_price is not None
                and unit_price > 0
                and line["can_do"] is not False
                and not is_foc
                and cur
            ):
                priced.append((vid, unit_price, cur))

        # Lowest per-currency within this row.
        by_currency: dict[str, dict | None] = {"USD": None, "RMB": None, "VND": None}
        min_per_cur: dict[str, tuple[int, float]] = {}
        for vid, price, cur in priced:
            cur_min = min_per_cur.get(cur)
            if cur_min is None or price < cur_min[1]:
                min_per_cur[cur] = (vid, price)
        for cur, (vid, price) in min_per_cur.items():
            by_currency[cur] = {"vendor_id": vid, "unit_price": price}
            # Mark every cell in this currency that equals the min as lowest (ties).
            for v2, p2, c2 in priced:
                if c2 == cur and p2 == price:
                    cells[str(v2)]["is_lowest"] = True

        out_items.append({
            "item_id": item_id,
            "item_no": it["item_no"],
            "specification": it["specification"],
            "bqms_code": it["bqms_code"],
            "quantity": item_qty,
            "unit": it["unit"],
            "required_material": it["required_material"],
            # ADMIN-ONLY: phân biệt nguồn + key mở File mã (Raw). KHÔNG ra path vendor.
            "source_kind": it["source_kind"],
            "source_bqms_rfq_number": it["source_bqms_rfq_number"],
            "target_price": float(it["target_price"]) if it["target_price"] is not None else None,
            "awarded_vendor_id": it["awarded_vendor_id"],
            "awarded_price": float(it["awarded_price"]) if it["awarded_price"] is not None else None,
            "awarded_currency": it["awarded_currency"],
            "cells": cells,
            "lowest": {"by_currency": by_currency},
        })

    return {
        "data": {
            "batch": {
                "id": batch["id"],
                "batch_code": batch["batch_code"],
                "title": batch["title"],
                "status": batch["status"],
                "award_mode": batch["award_mode"],
                "round_number": eff_round,
                # Đợt 2b [SB] — FE: render badge "🔒 Niêm phong tới {sealed_until}"
                # + ẩn nút Tờ trình/award khi sealed. sealed_until=null nếu chưa đặt hạn.
                "sealed": sealed,
                "sealed_until": (
                    batch["bid_deadline"].isoformat()
                    if (sealed and batch["bid_deadline"]) else None
                ),
                # Đợt 4 — meta FX cho FE: ngày as-of + số ô thiếu tỷ giá (banner amber).
                # Khi sealed → fx_map rỗng → missing_count=0 (không lộ, không cảnh báo).
                "fx": {
                    "as_of": as_of.isoformat() if as_of else None,
                    "missing_count": fx_missing_count,
                },
            },
            "vendors": vendors,
            "items": out_items,
        }
    }


# ---------------------------------------------------------------------------
# P5 — Full-quote DRAWER (one vendor's complete quote for the batch) + delta
# ---------------------------------------------------------------------------

@router.get("/batches/{batch_id}/quotes/{vendor_id}")
async def vendor_full_quote(
    batch_id: int,
    vendor_id: int,
    round_number: int | None = Query(None),
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P5 — Báo giá ĐẦY ĐỦ của 1 NCC cho 1 đợt (cho DRAWER trượt phải).

    Trả về:
      * header: currency, lead_time_days, total_amount, submitted_at, round_number,
        moq_notes, notes, attachment (path + tên file để tải).
      * lines[]: mỗi mã hàng → item_code/specification + unit_price, offered_qty,
        moq, can_do, lead_time, notes, attachments[], line_total + GIÁ TRỊ VÒNG
        TRƯỚC (prior_unit_price) để FE tính delta.

    Chỉ admin/manager/procurement (target_price/đối thủ KHÔNG bao giờ lộ path vendor).
    Mặc định lấy báo giá MỚI NHẤT đã 'submitted'; round_number ghi đè để soi 1 vòng.
    """
    batch = await conn.fetchrow(
        "SELECT id, batch_code, title, status, award_mode, bid_deadline, sealed_until_deadline "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    # Đợt 2b [SB] — drawer soi sâu nhất 1 NCC → rò nặng nhất. Chặn HẲN khi niêm
    # phong (không build lines[] lộ unit_price). Defense-in-depth: FE đã thay ô
    # bằng khoá nên không bấm Eye; nếu lọt vẫn 409. Sau hạn xem bình thường.
    if _sealed_active(batch):
        raise HTTPException(
            409,
            "Phiên đang niêm phong giá tới hạn — không xem được chi tiết báo giá NCC cho tới khi qua hạn.",
        )

    # vendor_accounts KHÔNG có cột email — email nằm ở users (JOIN qua user_id),
    # giống mọi endpoint khác (vd line ~460, ~1608, ~2494). Trước đây dòng này
    # SELECT email thẳng từ vendor_accounts → UndefinedColumnError → drawer 500.
    vendor = await conn.fetchrow(
        "SELECT va.id, va.company_name, va.contact_name, u.email "
        "FROM vendor_accounts va LEFT JOIN users u ON u.id = va.user_id "
        "WHERE va.id = $1",
        vendor_id,
    )
    if not vendor:
        raise HTTPException(404, "Nhà cung cấp không tồn tại")

    # The target quote = explicit round, else latest submitted (latest-wins).
    if round_number is not None:
        quote = await conn.fetchrow(
            """
            SELECT id, currency, total_amount, lead_time_days, moq_notes, notes,
                   attachment_path, external_url, status, submitted_at, round_number, valid_until
            FROM vendor_quotes
            WHERE batch_id = $1 AND vendor_id = $2 AND round_number = $3
            ORDER BY submitted_at DESC NULLS LAST, id DESC
            LIMIT 1
            """,
            batch_id, vendor_id, round_number,
        )
    else:
        quote = await conn.fetchrow(
            """
            SELECT id, currency, total_amount, lead_time_days, moq_notes, notes,
                   attachment_path, external_url, status, submitted_at, round_number, valid_until
            FROM vendor_quotes
            WHERE batch_id = $1 AND vendor_id = $2 AND status = 'submitted'
            ORDER BY round_number DESC NULLS LAST, submitted_at DESC NULLS LAST, id DESC
            LIMIT 1
            """,
            batch_id, vendor_id,
        )
    if not quote:
        raise HTTPException(404, "NCC này chưa gửi báo giá cho đợt")

    cur_round = quote["round_number"] or 1

    # Prior-round quote (for delta) — the latest submitted quote in an EARLIER round.
    prior = await conn.fetchrow(
        """
        SELECT id, round_number
        FROM vendor_quotes
        WHERE batch_id = $1 AND vendor_id = $2 AND status = 'submitted'
          AND round_number < $3
        ORDER BY round_number DESC NULLS LAST, submitted_at DESC NULLS LAST, id DESC
        LIMIT 1
        """,
        batch_id, vendor_id, cur_round,
    )
    prior_lines: dict[int, float] = {}
    if prior:
        for r in await conn.fetch(
            "SELECT item_id, unit_price FROM vendor_quote_items WHERE quote_id = $1",
            prior["id"],
        ):
            if r["unit_price"] is not None:
                prior_lines[r["item_id"]] = float(r["unit_price"])

    # Items of the batch (drives row order + item_code/spec; quantity fallback).
    items = await conn.fetch(
        """
        SELECT id, item_no, specification, bqms_code, item_code, quantity, unit,
               required_material, target_price
        FROM procurement_rfq_items
        WHERE batch_id = $1
        ORDER BY item_no
        """,
        batch_id,
    )

    # This quote's line rows keyed by item_id.
    line_rows = await conn.fetch(
        """
        SELECT item_id, unit_price, quantity, offered_qty, moq, lead_time_days,
               notes, can_do, currency, attachment_paths, free_charge
        FROM vendor_quote_items
        WHERE quote_id = $1
        """,
        quote["id"],
    )
    line_by_item: dict[int, dict] = {r["item_id"]: dict(r) for r in line_rows}

    quote_currency = quote["currency"]
    lines = []
    for it in items:
        item_id = it["id"]
        ln = line_by_item.get(item_id)
        item_qty = float(it["quantity"]) if it["quantity"] is not None else None
        if ln is None:
            lines.append({
                "item_id": item_id,
                "item_no": it["item_no"],
                "item_code": it["item_code"] or it["bqms_code"],
                "bqms_code": it["bqms_code"],
                "specification": it["specification"],
                "unit": it["unit"],
                "required_material": it["required_material"],
                "quantity": item_qty,
                "unit_price": None, "offered_qty": None, "moq": None,
                "can_do": None, "lead_time_days": None, "notes": None,
                "currency": None, "line_total": None, "free_charge": False,
                "attachments": [], "prior_unit_price": prior_lines.get(item_id),
                "delta": None, "delta_pct": None,
            })
            continue
        unit_price = float(ln["unit_price"]) if ln["unit_price"] is not None else None
        # offered_qty = 'SL báo' (what the vendor can supply); falls back to the
        # quoted 'SL yêu cầu' (quantity), then the RFQ item qty. SAME basis the
        # submit total_amount uses, so drawer line_total == header total.
        offered_qty = float(ln["offered_qty"]) if ln["offered_qty"] is not None else None
        eff_qty = (
            offered_qty
            if offered_qty is not None
            else (float(ln["quantity"]) if ln["quantity"] is not None else item_qty)
        )
        line_total = (unit_price * eff_qty) if (unit_price is not None and eff_qty is not None) else None
        line_cur = ln["currency"] or quote_currency
        # attachment_paths is JSONB (list); coerce to filename descriptors.
        raw_atts = ln.get("attachment_paths")
        if isinstance(raw_atts, str):
            try:
                raw_atts = _json.loads(raw_atts)
            except (ValueError, TypeError):
                raw_atts = []
        attachments = []
        for idx, p in enumerate(raw_atts or []):
            if p:
                attachments.append({"index": idx, "filename": os.path.basename(str(p))})
        prior_up = prior_lines.get(item_id)
        delta = (unit_price - prior_up) if (unit_price is not None and prior_up is not None) else None
        delta_pct = (delta / prior_up * 100.0) if (delta is not None and prior_up) else None
        lines.append({
            "item_id": item_id,
            "item_no": it["item_no"],
            "item_code": it["item_code"] or it["bqms_code"],
            "bqms_code": it["bqms_code"],
            "specification": it["specification"],
            "unit": it["unit"],
            "required_material": it["required_material"],
            "quantity": item_qty,
            "unit_price": unit_price,
            "offered_qty": offered_qty,
            "moq": ln["moq"],
            "can_do": ln["can_do"],
            "free_charge": bool(ln["free_charge"]),
            "lead_time_days": ln["lead_time_days"],
            "notes": ln["notes"],
            "currency": line_cur,
            "line_total": line_total,
            "attachments": attachments,
            "prior_unit_price": prior_up,
            "delta": delta,
            "delta_pct": delta_pct,
        })

    quote_att = quote["attachment_path"]
    _vg = (await _get_vendor_grades(conn)).get(vendor["id"], {})  # M3 — hạng + on-time
    return {
        "data": {
            "batch": {
                "id": batch["id"], "batch_code": batch["batch_code"],
                "title": batch["title"], "status": batch["status"],
                "award_mode": batch["award_mode"],
            },
            "vendor": {
                "vendor_id": vendor["id"], "company_name": vendor["company_name"],
                "contact_name": vendor["contact_name"], "email": vendor["email"],
            },
            "header": {
                "quote_id": quote["id"],
                "currency": quote_currency,
                "lead_time_days": quote["lead_time_days"],
                "total_amount": float(quote["total_amount"]) if quote["total_amount"] is not None else None,
                "submitted_at": quote["submitted_at"],
                "round_number": cur_round,
                "status": quote["status"],
                "moq_notes": quote["moq_notes"],
                "notes": quote["notes"],
                "valid_until": quote["valid_until"],
                "grade": _vg.get("grade"),               # M3
                "on_time_rate": _vg.get("on_time_rate"),  # M3
                "has_attachment": bool(quote_att),
                "attachment_filename": os.path.basename(quote_att) if quote_att else None,
                "external_url": quote["external_url"],
                "prior_round": prior["round_number"] if prior else None,
            },
            "lines": lines,
        }
    }


@router.get("/quotes/{quote_id}/attachment")
async def download_quote_attachment(
    quote_id: int,
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P5 — Tải file đính kèm cấp BÁO GIÁ (Excel/PDF NCC upload) của 1 quote.

    Path-guard pin dưới FILES_BASE_PATH (mirror `_resolve_contract_pdf`). Chỉ
    admin/manager/procurement. attachment_path do server ghi (sandboxed
    vendor_uploads/{vendor_id}).
    """
    row = await conn.fetchrow(
        "SELECT attachment_path FROM vendor_quotes WHERE id = $1", quote_id
    )
    if not row or not row["attachment_path"]:
        raise HTTPException(404, "Báo giá này không có file đính kèm")
    resolved = _resolve_under_files_base(row["attachment_path"])
    return FileResponse(
        str(resolved),
        media_type="application/octet-stream",
        filename=resolved.name,
    )


@router.get("/quotes/{quote_id}/items/{item_id}/attachment")
async def download_quote_item_attachment(
    quote_id: int,
    item_id: int,
    index: int = Query(0, ge=0),
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P5 — Tải 1 file đính kèm cấp DÒNG (attachment_paths[index]) của 1 mã hàng."""
    row = await conn.fetchrow(
        "SELECT attachment_paths FROM vendor_quote_items WHERE quote_id = $1 AND item_id = $2",
        quote_id, item_id,
    )
    if not row:
        raise HTTPException(404, "Không tìm thấy dòng báo giá")
    raw = row["attachment_paths"]
    if isinstance(raw, str):
        try:
            raw = _json.loads(raw)
        except (ValueError, TypeError):
            raw = []
    paths = raw or []
    if index >= len(paths) or not paths[index]:
        raise HTTPException(404, "File đính kèm không tồn tại")
    resolved = _resolve_under_files_base(str(paths[index]))
    return FileResponse(
        str(resolved),
        media_type="application/octet-stream",
        filename=resolved.name,
    )


@router.get("/batches/{batch_id}/quotes/{vendor_id}/attachments.zip")
async def download_vendor_quote_attachments_zip(
    batch_id: int,
    vendor_id: int,
    round_number: int | None = Query(None),
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tải TẤT CẢ file đính kèm của 1 NCC cho 1 đợt thành 1 file .zip.

    Gom: file cấp-phiếu (vendor_quotes.attachment_path) + mọi file cấp-dòng
    (vendor_quote_items.attachment_paths), gom theo thư mục con đặt theo mã hàng.
    Mọi đường dẫn pin dưới FILES_BASE_PATH (mirror download_quote_attachment); file
    thiếu/đã xoá bị bỏ qua (không fail cả zip). Chỉ admin/manager/procurement.
    """
    import io as _io
    import zipfile as _zip
    from fastapi.responses import StreamingResponse

    if round_number is not None:
        quote = await conn.fetchrow(
            "SELECT id, attachment_path FROM vendor_quotes "
            "WHERE batch_id=$1 AND vendor_id=$2 AND round_number=$3 "
            "ORDER BY submitted_at DESC NULLS LAST, id DESC LIMIT 1",
            batch_id, vendor_id, round_number,
        )
    else:
        quote = await conn.fetchrow(
            "SELECT id, attachment_path FROM vendor_quotes "
            "WHERE batch_id=$1 AND vendor_id=$2 AND status='submitted' "
            "ORDER BY round_number DESC NULLS LAST, submitted_at DESC NULLS LAST, id DESC LIMIT 1",
            batch_id, vendor_id,
        )
    if not quote:
        raise HTTPException(404, "NCC này chưa gửi báo giá cho đợt")

    vendor = await conn.fetchrow(
        "SELECT company_name FROM vendor_accounts WHERE id=$1", vendor_id
    )

    # (label, path): file cấp-phiếu + mọi file cấp-dòng.
    files: list[tuple[str, str]] = []
    if quote["attachment_path"]:
        files.append(("bao-gia", str(quote["attachment_path"])))
    line_rows = await conn.fetch(
        "SELECT vqi.attachment_paths, i.bqms_code, i.item_no "
        "FROM vendor_quote_items vqi JOIN procurement_rfq_items i ON i.id = vqi.item_id "
        "WHERE vqi.quote_id = $1",
        quote["id"],
    )
    for r in line_rows:
        raw_atts = r["attachment_paths"]
        if isinstance(raw_atts, str):
            try:
                raw_atts = _json.loads(raw_atts)
            except (ValueError, TypeError):
                raw_atts = []
        label = str(r["bqms_code"] or f"item{r['item_no']}")
        for p in (raw_atts or []):
            if p:
                files.append((label, str(p)))

    if not files:
        raise HTTPException(404, "NCC này không có file đính kèm nào")

    buf = _io.BytesIO()
    used: set[str] = set()
    with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as zf:
        for label, p in files:
            try:
                resolved = _resolve_under_files_base(p)
                if not resolved.exists():
                    continue
                safe_label = re.sub(r"[^A-Za-z0-9._-]", "_", label) or "file"
                arc = f"{safe_label}/{resolved.name}"
                n = 1
                while arc in used:
                    arc = f"{safe_label}/{n}_{resolved.name}"
                    n += 1
                used.add(arc)
                zf.write(str(resolved), arcname=arc)
            except (ValueError, OSError, HTTPException):
                # _resolve_under_files_base raises HTTPException(404) for a
                # missing/escaping path — skip that one file, never fail the
                # whole archive ("file thiếu/đã xoá bị bỏ qua").
                continue
    if not used:
        raise HTTPException(404, "Không tìm thấy file đính kèm trên đĩa")
    buf.seek(0)
    safe_vendor = re.sub(r"[^A-Za-z0-9._-]", "_", (vendor["company_name"] if vendor else "NCC") or "NCC")
    fname = f"bao-gia-{safe_vendor}-batch{batch_id}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# P5 — Decision sheet (TỜ TRÌNH CHỐT THẦU) — per-currency comparison .xlsx
# ---------------------------------------------------------------------------

_EXPORTS_DIR = Path(getattr(settings, "FILES_BASE_PATH", "/data/files")) / "exports"
try:
    _EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
except (OSError, PermissionError):
    _EXPORTS_DIR = Path(os.getcwd()) / "tmp" / "procurement_exports"
    _EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _cleanup_file(path: Path) -> None:
    try:
        if path.exists():
            path.unlink()
    except Exception:  # noqa: BLE001
        logger.warning("Failed to clean export %s", path, exc_info=True)


@router.get("/batches/{batch_id}/decision-sheet")
async def batch_decision_sheet(
    batch_id: int,
    round_number: int | None = Query(None),
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
    background: BackgroundTasks = None,  # type: ignore[assignment]
):
    """P5 — TỜ TRÌNH CHỐT THẦU: bảng so sánh per-currency (.xlsx) + khối award-summary.

    Mỗi loại tiền là 1 BẢNG riêng (KHÔNG cộng chéo tiền tệ): rows=mã hàng,
    cols=NCC, ô = đơn giá; ô THẤP NHẤT trong hàng (cùng currency) tô xanh + đậm.
    Sau bảng so sánh là khối AWARD-SUMMARY: mỗi mã đã chốt → NCC trúng + giá +
    award_reason + người chốt + thời điểm. Stream .xlsx (admin only).

    Tái dùng logic ma trận (latest-submitted-wins per vendor) + procurement_awards
    (active = superseded_by IS NULL). Mirror styling từ analytics_exports._write_xlsx.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    batch = await conn.fetchrow(
        "SELECT id, batch_code, title, status, award_mode, current_round, bid_deadline, sealed_until_deadline "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    # Đợt 2b [SB] — tờ trình chốt thầu là file ĐẦY ĐỦ đơn giá NCC. Khi niêm phong
    # → CHẶN HẲN export (KISS: đừng sinh .xlsx rỗng giá). Sau hạn export bình thường.
    if _sealed_active(batch):
        raise HTTPException(
            409,
            "Phiên đang niêm phong giá tới hạn — không thể xuất tờ trình chốt thầu cho tới khi qua hạn báo giá.",
        )

    eff_round = round_number
    if eff_round is None:
        eff_round = await conn.fetchval(
            "SELECT MAX(round_number) FROM procurement_rfq_invitations WHERE batch_id = $1",
            batch_id,
        ) or 1

    # Vendors (cols) + their latest submitted quote — same shape as the matrix.
    vendor_rows = await conn.fetch(
        """
        SELECT inv.vendor_id, va.company_name,
               lq.id AS quote_id, lq.currency, lq.total_amount
        FROM procurement_rfq_invitations inv
        JOIN vendor_accounts va ON va.id = inv.vendor_id
        LEFT JOIN LATERAL (
            SELECT id, currency, total_amount
            FROM vendor_quotes vq
            WHERE vq.batch_id = inv.batch_id AND vq.vendor_id = inv.vendor_id
              AND vq.status = 'submitted'
            ORDER BY vq.round_number DESC, vq.submitted_at DESC NULLS LAST, vq.id DESC
            LIMIT 1
        ) lq ON true
        WHERE inv.batch_id = $1 AND inv.round_number = $2
        ORDER BY va.company_name
        """,
        batch_id, eff_round,
    )
    vendors = [dict(v) for v in vendor_rows]
    quote_ids = {v["vendor_id"]: v["quote_id"] for v in vendors if v["quote_id"] is not None}
    quote_currency = {v["quote_id"]: v["currency"] for v in vendors if v["quote_id"] is not None}

    items = await conn.fetch(
        """
        SELECT id, item_no, specification, bqms_code, item_code, quantity, unit, target_price
        FROM procurement_rfq_items WHERE batch_id = $1 ORDER BY item_no
        """,
        batch_id,
    )

    line_map: dict[tuple[int, int], dict] = {}
    if quote_ids:
        for lr in await conn.fetch(
            """SELECT quote_id, item_id, unit_price, quantity, currency,
                      can_do, free_charge
               FROM vendor_quote_items WHERE quote_id = ANY($1::bigint[])""",
            list(quote_ids.values()),
        ):
            line_map[(lr["quote_id"], lr["item_id"])] = dict(lr)

    # Active awards (per-item + per-batch) + actor name, for the summary block.
    awards = await conn.fetch(
        """
        SELECT a.item_id, a.vendor_id, a.awarded_price, a.currency, a.quantity,
               a.award_reason, a.awarded_at, va.company_name,
               COALESCE(u.full_name, u.email) AS awarded_by_name
        FROM procurement_awards a
        JOIN vendor_accounts va ON va.id = a.vendor_id
        LEFT JOIN users u ON u.id = a.awarded_by
        WHERE a.batch_id = $1 AND a.superseded_by IS NULL
        ORDER BY a.item_id NULLS FIRST
        """,
        batch_id,
    )
    award_by_item: dict[int | None, dict] = {a["item_id"]: dict(a) for a in awards}

    # Đợt 4 — FX normalize: build rate map 1 lần/batch, as-of bid_deadline.
    # Sealed đã chặn 409 ở trên → tới đây luôn an toàn để quy đổi (giá đã lộ hợp lệ).
    # Gom currency từ cả vendor-quote lẫn award (award có thể chốt ở ngoại tệ).
    as_of = batch["bid_deadline"].date() if batch["bid_deadline"] else None
    fx_currencies = {v["currency"] for v in vendors if v["currency"]}
    fx_currencies |= {a["currency"] for a in awards if a["currency"]}
    fx_map = await _fx_map_for_batch(conn, fx_currencies, as_of)

    # ── Build workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "So sánh báo giá"

    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=14, color="1E293B")
    sub_font = Font(size=10, color="64748B")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4F46E5")  # brand-600
    sect_font = Font(bold=True, size=11, color="1E293B")
    low_fill = PatternFill("solid", fgColor="DCFCE7")  # emerald-100
    low_font = Font(bold=True, color="047857")          # emerald-700
    money_font = Font(name="Consolas")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    r = 1
    ws.cell(r, 1, f"TỜ TRÌNH CHỐT THẦU — {batch['batch_code']}").font = title_font
    r += 1
    ws.cell(r, 1, batch["title"] or "").font = sub_font
    r += 1
    ws.cell(
        r, 1,
        f"Vòng {eff_round} · Hình thức: "
        f"{'Theo mã hàng' if batch['award_mode'] == 'per_item' else 'Theo cả gói'} · "
        f"Xuất lúc {datetime.now(timezone(timedelta(hours=7))).strftime('%d/%m/%Y %H:%M')}",
    ).font = sub_font
    r += 1
    # Đợt 4 — ghi rõ tỷ giá quy đổi VND theo ngày nào (runtime, KHÔNG lưu DB).
    ws.cell(
        r, 1,
        f"Tỷ giá quy đổi VND theo ngày "
        f"{as_of.strftime('%d/%m/%Y') if as_of else 'mới nhất (chưa đặt hạn báo giá)'}",
    ).font = sub_font
    r += 2

    # Group vendors by the currency they quoted in → one comparison block each.
    cur_to_vendors: dict[str, list[dict]] = {}
    for v in vendors:
        cur = v["currency"]
        if v["quote_id"] is None or not cur:
            continue
        cur_to_vendors.setdefault(cur, []).append(v)

    if not cur_to_vendors:
        ws.cell(r, 1, "Chưa có NCC nào gửi báo giá hợp lệ cho vòng này.").font = sub_font
        r += 2

    for cur in sorted(cur_to_vendors.keys()):
        cur_vendors = cur_to_vendors[cur]
        ws.cell(r, 1, f"BẢNG SO SÁNH — {cur}").font = sect_font
        r += 1
        header_row = r
        ws.cell(r, 1, "Mã hàng").font = hdr_font
        ws.cell(r, 1).fill = hdr_fill
        ws.cell(r, 2, "Mô tả").font = hdr_font
        ws.cell(r, 2).fill = hdr_fill
        ws.cell(r, 3, "SL").font = hdr_font
        ws.cell(r, 3).fill = hdr_fill
        for ci, v in enumerate(cur_vendors, start=4):
            c = ws.cell(r, ci, v["company_name"])
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
        for c in range(1, 4 + len(cur_vendors)):
            ws.cell(header_row, c).border = border
        r += 1

        for it in items:
            ws.cell(r, 1, it["item_code"] or it["bqms_code"] or f"#{it['item_no']}").alignment = left
            ws.cell(r, 2, it["specification"] or "").alignment = left
            qty = float(it["quantity"]) if it["quantity"] is not None else None
            ws.cell(r, 3, qty if qty is not None else "").alignment = right

            # Collect this row's prices for the vendors in THIS currency block.
            row_prices: list[tuple[int, float]] = []  # (col_index, price)
            for ci, v in enumerate(cur_vendors, start=4):
                qid = quote_ids.get(v["vendor_id"])
                line = line_map.get((qid, it["id"])) if qid else None
                price = float(line["unit_price"]) if line and line["unit_price"] is not None else None
                is_foc = bool(line["free_charge"]) if line else False
                cant_do = bool(line and line["can_do"] is False)
                # Hiển thị rõ: FOC = miễn phí, "Không làm" = NCC không cung cấp được.
                if is_foc:
                    display = "FOC"
                elif cant_do:
                    display = "Không làm"
                elif price is not None:
                    display = price
                else:
                    display = "—"
                cell = ws.cell(r, ci, display)
                cell.alignment = right
                cell.font = money_font
                # Chỉ giá thật (>0, báo được, không FOC) mới vào diện "thấp nhất".
                if price is not None and price > 0 and not is_foc and not cant_do:
                    row_prices.append((ci, price))
            # Highlight lowest (ties included) in this currency row.
            if row_prices:
                lo = min(p for _, p in row_prices)
                for ci, p in row_prices:
                    if p == lo:
                        ws.cell(r, ci).fill = low_fill
                        ws.cell(r, ci).font = low_font
            for c in range(1, 4 + len(cur_vendors)):
                ws.cell(r, c).border = border
            r += 1

        # Per-vendor total row for this currency block.
        ws.cell(r, 1, "Tổng").font = Font(bold=True)
        for ci, v in enumerate(cur_vendors, start=4):
            tot = v["total_amount"]
            c = ws.cell(r, ci, float(tot) if tot is not None else "")
            c.font = Font(bold=True, name="Consolas")
            c.alignment = right
        for c in range(1, 4 + len(cur_vendors)):
            ws.cell(r, c).border = border
        r += 1

        # Đợt 4 — dòng quy đổi "≈ Tổng VND" dưới Tổng (ADDITIVE, không đổi Tổng gốc).
        # VND-block: rate=1 → trùng Tổng (vô hại). Thiếu rate → in "thiếu tỷ giá"
        # (minh bạch, KHÔNG bịa số). cur = currency của block hiện tại.
        _rate = fx_map.get(cur)
        _vnd_font = Font(italic=True, size=9, color="64748B")
        _vnd_money_font = Font(italic=True, size=9, color="64748B", name="Consolas")
        ws.cell(r, 1, "≈ Tổng VND").font = _vnd_font
        for ci, v in enumerate(cur_vendors, start=4):
            tot = v["total_amount"]
            tot_f = float(tot) if tot is not None else None
            vnd = _vnd(tot_f, _rate)
            if vnd is not None:
                disp = vnd
            elif tot_f is not None:
                disp = "thiếu tỷ giá"
            else:
                disp = ""
            c = ws.cell(r, ci, disp)
            c.font = _vnd_money_font
            c.alignment = right
        for c in range(1, 4 + len(cur_vendors)):
            ws.cell(r, c).border = border
        r += 2

    # ── Award-summary block ──
    ws.cell(r, 1, "KẾT QUẢ CHỐT THẦU").font = sect_font
    r += 1
    # Đợt 4 — thêm cột 9 "Quy đổi VND" (ADDITIVE; Đơn giá gốc cột 3 giữ nguyên).
    summ_hdr = ["Mã hàng", "NCC trúng", "Đơn giá", "Tiền tệ", "SL", "Lý do chọn", "Người chốt", "Thời điểm", "Quy đổi VND"]
    for ci, h in enumerate(summ_hdr, start=1):
        c = ws.cell(r, ci, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = center
    r += 1

    item_code_by_id = {it["id"]: (it["item_code"] or it["bqms_code"] or f"#{it['item_no']}") for it in items}
    if not awards:
        ws.cell(r, 1, "Chưa chốt thầu.").font = sub_font
        r += 1
    else:
        for a in awards:
            if a["item_id"] is None:
                label = "Cả gói (toàn bộ mã)"
            else:
                label = item_code_by_id.get(a["item_id"], f"#{a['item_id']}")
            # Đợt 4 — quy đổi VND giá trúng (cột 9). Thiếu rate → "thiếu tỷ giá".
            _aw_price = float(a["awarded_price"]) if a["awarded_price"] is not None else None
            _aw_vnd = _vnd(_aw_price, fx_map.get(a["currency"]))
            if _aw_vnd is not None:
                _aw_vnd_disp = _aw_vnd
            elif _aw_price is not None:
                _aw_vnd_disp = "thiếu tỷ giá"
            else:
                _aw_vnd_disp = ""
            vals = [
                label,
                a["company_name"],
                float(a["awarded_price"]) if a["awarded_price"] is not None else "",
                a["currency"] or "",
                float(a["quantity"]) if a["quantity"] is not None else "",
                a["award_reason"] or "",
                a["awarded_by_name"] or "",
                a["awarded_at"].astimezone(timezone(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")
                if a["awarded_at"] else "",
                _aw_vnd_disp,
            ]
            for ci, val in enumerate(vals, start=1):
                c = ws.cell(r, ci, val)
                c.border = border
                if ci in (3, 5):
                    c.font = money_font
                    c.alignment = right
                elif ci == 9:
                    c.font = money_font
                    c.alignment = right
                elif ci == 6:
                    c.alignment = left
                else:
                    c.alignment = left
            r += 1

    # Column widths.
    widths = {1: 22, 2: 40, 3: 8}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    # Đợt 4 — award-summary giờ 9 cột (thêm "Quy đổi VND") → đảm bảo cột 9 có width.
    max_col = max(9, 3 + max((len(vs) for vs in cur_to_vendors.values()), default=0))
    for col in range(4, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    safe_code = re.sub(r"[^A-Za-z0-9_.-]+", "_", batch["batch_code"] or f"batch{batch_id}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = _EXPORTS_DIR / f"to-trinh-chot-thau_{safe_code}_{ts}.xlsx"
    wb.save(out_path)
    if background is not None:
        background.add_task(_cleanup_file, out_path)

    download_name = f"To-trinh-chot-thau_{safe_code}.xlsx"
    return FileResponse(
        str(out_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name,
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


# ---------------------------------------------------------------------------
# Đợt 3 — Award finalize (DRY) + total helper for maker-checker gate
# ---------------------------------------------------------------------------

async def _award_total_vnd(
    conn: asyncpg.Connection, batch_id: int
) -> tuple[float, bool]:
    """Tổng giá trị award ACTIVE của batch quy phần VND + cờ có dòng ngoại tệ.

    GIẢ ĐỊNH ĐA TIỀN TỆ (KISS, fail-safe): award KHÔNG lưu tỷ giá (write-back chỉ
    VND 1:1). Nên CHỈ cộng các dòng currency='VND'; nếu batch có BẤT KỲ dòng
    ngoại tệ (USD/...) → trả has_foreign=True ⇒ caller LUÔN cần duyệt (fail-safe:
    thà bắt duyệt nhầm còn hơn lọt công nợ ngoại tệ chưa quy đổi).

    Dòng giá trị:
      * per_batch (item_id IS NULL): awarded_price ĐÃ là total_amount → cộng thẳng.
      * per_item  (item_id NOT NULL): awarded_price × COALESCE(quantity,1).
    Đọc procurement_awards active (superseded_by IS NULL).
    """
    rows = await conn.fetch(
        "SELECT awarded_price, currency, quantity, item_id "
        "FROM procurement_awards WHERE batch_id = $1 AND superseded_by IS NULL",
        batch_id,
    )
    total_vnd = 0.0
    has_foreign = False
    for r in rows:
        cur = (r["currency"] or "VND").upper()
        price = float(r["awarded_price"] or 0)
        line = price if r["item_id"] is None else price * float(r["quantity"] or 1)
        if cur == "VND":
            total_vnd += line
        else:
            has_foreign = True
    return total_vnd, has_foreign


async def _finalize_award(
    conn: asyncpg.Connection,
    *,
    batch_id: int,
    actor: str,
    old_status: str,
    award_mode: str,
    award_round: int,
    award_reason: str,
    winning_quote_ids: set[int],
    awarded_vendor_ids: set[int],
) -> None:
    """Đợt 3 — Bước CUỐI của một award: flip quote-statuses + set batch='awarded'
    + audit + notify. KHÔNG gồm contract/PO/write-back (vẫn là bước admin thủ công
    SAU đó — giữ KISS/YAGNI; finalize chỉ "chốt kết quả thầu").

    DRY — gọi từ HAI đường:
      (1) award_batch khi finalize-NGAY (approval OFF HOẶC total < threshold);
      (2) approve_award khi checker DUYỆT (approval ON + total >= threshold).
    Logic Y HỆT block C cũ — chỉ tách hàm, KHÔNG đổi hành vi.

    PHẢI chạy trong transaction của caller (giống _audit — KHÔNG tự mở txn).
    """
    # (C1) Flip quote statuses: winners → awarded, remaining submitted → rejected.
    if winning_quote_ids:
        win_list = list(winning_quote_ids)
        await conn.execute(
            "UPDATE vendor_quotes SET status = 'awarded' WHERE id = ANY($1::bigint[])",
            win_list,
        )
        rejected = await conn.fetch(
            "UPDATE vendor_quotes SET status = 'rejected' "
            "WHERE batch_id = $1 AND status = 'submitted' AND id != ALL($2::bigint[]) "
            "RETURNING id",
            batch_id, win_list,
        )
        for q in win_list:
            await _audit(conn, "quote", q, "status_change", actor_id=actor, to_status="awarded")
        for r in rejected:
            await _audit(conn, "quote", r["id"], "status_change", actor_id=actor, to_status="rejected")

    # (C2) Finalize batch. award_status được caller set (none khi finalize-ngay,
    # approved khi đến từ đường duyệt) — ở đây CHỈ set status/awarded_at/closed_at.
    await conn.execute(
        "UPDATE procurement_rfq_batches "
        "SET status = 'awarded', awarded_at = NOW(), closed_at = COALESCE(closed_at, NOW()) "
        "WHERE id = $1",
        batch_id,
    )

    # (C3) Audit + notify (y hệt block cũ). old_status='awarded' → re_award.
    action = "re_award" if old_status == "awarded" else "award"
    await _audit(
        conn, "batch", batch_id, action,
        actor_id=actor, from_status=old_status, to_status="awarded",
        detail={"award_mode": award_mode, "reason": award_reason,
                "winner_count": len(winning_quote_ids), "round": award_round},
    )
    sole_vendor = next(iter(awarded_vendor_ids)) if len(awarded_vendor_ids) == 1 else None
    await dispatch_procurement_event(
        conn, "award", batch_id, action,
        actor_id=actor, awarded_vendor_id=sole_vendor,
        detail={"batch_id": batch_id, "award_mode": award_mode,
                "winner_count": len(winning_quote_ids), "round": award_round},
    )


@router.post("/batches/{batch_id}/award")
async def award_batch(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt-2 — Chọn nhà cung cấp trúng thầu (per_item HOẶC per_batch).

    Ghi vào procurement_awards (1 row active mỗi (batch,item) hoặc per-batch),
    đồng bộ procurement_rfq_items.awarded_* (giữ create-contract chạy), flip
    vendor_quotes winner→awarded / losers→rejected, set batch status='awarded',
    awarded_at. RE-AWARD: supersede award active cũ (set superseded_by) rồi
    insert row mới. Winner lấy từ VIEW v_latest_vendor_quote (vòng mới nhất).

    award_mode đọc từ batch row (không phải body). award_reason BẮT BUỘC.

    Body per_item:  {"awards":[{"item_id","vendor_id","price"?,"currency"?,"quantity"?}],
                     "award_reason","criteria"?}
    Body per_batch: {"vendor_id","award_reason","criteria"?}
    """
    batch = await conn.fetchrow(
        "SELECT id, status, award_mode, current_round, award_status "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    # 'awarded' allowed → enables RE-AWARD; 'evaluating' allowed (admin opened matrix).
    if batch["status"] not in ("published", "evaluating", "awarded"):
        raise HTTPException(400, "Đợt báo giá chưa công bố hoặc đã đóng/huỷ")

    old_status = batch["status"]

    # [MC] Đã có đề xuất chốt thầu đang treo → chặn award lại (tránh chồng đề xuất /
    # ghi đè awards đang chờ duyệt). Phải duyệt hoặc từ chối đề xuất hiện tại trước.
    if batch["award_status"] == "proposed":
        raise HTTPException(
            409,
            "Đang có đề xuất chốt thầu chờ duyệt. Hãy duyệt hoặc từ chối trước khi award lại.",
        )

    # [RG] RE-AWARD GUARD — chống công nợ TRÙNG. Khi RE-AWARD (batch đã từng
    # awarded) mà ĐÃ có hợp đồng HOẶC PO chưa cancelled → cấm award lại tới khi
    # huỷ chứng từ tài chính hiện hành. Award lần đầu (chưa awarded) KHÔNG chặn.
    # Chỉ ĐỌC → fail-fast 409 NGOÀI transaction.
    if old_status == "awarded":
        active_doc = await conn.fetchrow(
            "SELECT 'contract' AS kind, contract_no AS ref FROM procurement_contracts "
            "WHERE batch_id = $1 AND status <> 'cancelled' "
            "UNION ALL "
            "SELECT 'po' AS kind, po_no AS ref FROM procurement_pos "
            "WHERE batch_id = $1 AND status <> 'cancelled' "
            "LIMIT 1",
            batch_id,
        )
        if active_doc:
            label = "Hợp đồng" if active_doc["kind"] == "contract" else "PO"
            raise HTTPException(
                409,
                f"Huỷ {label} hiện tại ({active_doc['ref']}) trước khi award lại "
                "(tránh sinh công nợ trùng).",
            )

    award_reason = (body.get("award_reason") or "").strip()
    if not award_reason:
        raise HTTPException(400, "award_reason là bắt buộc")
    criteria = body.get("criteria")

    award_round = batch["current_round"] or 1
    award_mode = batch["award_mode"]
    actor = token_data.user_id

    award_ids: list[int] = []
    superseded_count = 0
    winning_quote_ids: set[int] = set()
    awarded_vendor_ids: set[int] = set()

    async with conn.transaction():
        if criteria is not None:
            await conn.execute(
                "UPDATE procurement_rfq_batches SET criteria = $1 WHERE id = $2",
                criteria, batch_id,
            )

        if award_mode == "per_batch":
            vendor_id = body.get("vendor_id")
            if not vendor_id:
                raise HTTPException(400, "vendor_id là bắt buộc")
            vendor_id = int(vendor_id)

            row = await conn.fetchrow(
                "SELECT * FROM v_latest_vendor_quote WHERE batch_id = $1 AND vendor_id = $2",
                batch_id, vendor_id,
            )
            if not row:
                raise HTTPException(400, "NCC chưa có báo giá hợp lệ")
            winning_quote_ids.add(row["quote_id"])
            awarded_vendor_ids.add(vendor_id)

            # Collect prior active awards for this batch (per-batch row + any per_item rows).
            prev_awards = await conn.fetch(
                "SELECT id, item_id FROM procurement_awards WHERE batch_id = $1 AND superseded_by IS NULL",
                batch_id,
            )

            # Clear the prior ACTIVE per-batch row (item_id IS NULL) out of the partial
            # unique index (uq_pa_batch_perbatch_active) BEFORE inserting the new one —
            # the index is non-deferrable so two active per-batch rows would collide
            # mid-transaction. Self-reference is FK-safe; repointed to new id below.
            for pa in prev_awards:
                if pa["item_id"] is None:
                    await conn.execute(
                        "UPDATE procurement_awards SET superseded_by = id WHERE id = $1",
                        pa["id"],
                    )

            new_award_id = await conn.fetchval(
                """
                INSERT INTO procurement_awards
                    (batch_id, item_id, vendor_id, quote_id, quote_item_id,
                     awarded_price, currency, quantity, award_reason, awarded_by)
                VALUES ($1, NULL, $2, $3, NULL, $4, $5, NULL, $6, $7)
                RETURNING id
                """,
                batch_id, vendor_id, row["quote_id"], row["total_amount"],
                row["currency"] or "VND", award_reason, actor,
            )
            award_ids.append(new_award_id)

            # Mirror awarded_* onto every item from the winner's latest quote.
            # CHỦ ĐÍCH: per_batch nghĩa là NCC này thắng CẢ phiên, nên mọi dòng của
            # họ đều được award — kể cả dòng FOC (awarded_price=0 là GIÁ THẬT vì
            # miễn phí). Không lọc FOC ở đây: bỏ dòng FOC khỏi award sẽ làm thiếu
            # mặt hàng trong hợp đồng/PO.
            vqis = await conn.fetch(
                "SELECT id, item_id, unit_price FROM vendor_quote_items WHERE quote_id = $1",
                row["quote_id"],
            )
            for vqi in vqis:
                await conn.execute(
                    """UPDATE procurement_rfq_items
                       SET awarded_vendor_id = $1, awarded_price = $2, awarded_currency = $3,
                           awarded_round = $4, awarded_quote_item_id = $5
                       WHERE id = $6 AND batch_id = $7""",
                    vendor_id, vqi["unit_price"], row["currency"] or "VND",
                    award_round, vqi["id"], vqi["item_id"], batch_id,
                )

            # Repoint ALL prior active awards (per-batch + per_item) at the new award.
            for pa in prev_awards:
                await conn.execute(
                    "UPDATE procurement_awards SET superseded_by = $1 WHERE id = $2",
                    new_award_id, pa["id"],
                )
                superseded_count += 1

            await _audit(
                conn, "award", new_award_id,
                "re_award" if prev_awards else "award",
                actor_id=actor, to_status="awarded",
                detail={
                    "mode": "per_batch", "vendor_id": vendor_id,
                    "quote_id": row["quote_id"], "currency": row["currency"],
                    "total_amount": float(row["total_amount"]) if row["total_amount"] is not None else None,
                    "reason": award_reason,
                    "superseded_award_ids": [pa["id"] for pa in prev_awards] or None,
                },
            )

        else:  # per_item
            awards = body.get("awards") or []
            if not awards:
                raise HTTPException(400, "Cần ít nhất 1 dòng award")

            for award in awards:
                item_id = int(award["item_id"])
                vendor_id = int(award["vendor_id"])

                row = await conn.fetchrow(
                    "SELECT * FROM v_latest_vendor_quote WHERE batch_id = $1 AND vendor_id = $2",
                    batch_id, vendor_id,
                )
                if not row:
                    raise HTTPException(400, "NCC chưa có báo giá hợp lệ")
                winning_quote_ids.add(row["quote_id"])
                awarded_vendor_ids.add(vendor_id)

                qi = await conn.fetchrow(
                    "SELECT id, unit_price, quantity FROM vendor_quote_items "
                    "WHERE quote_id = $1 AND item_id = $2",
                    row["quote_id"], item_id,
                )
                # body OVERRIDES quote-item, else quote-item, else rfq_item.quantity.
                price = award.get("price")
                if price is None:
                    price = qi["unit_price"] if qi else None
                quantity = award.get("quantity")
                if quantity is None:
                    quantity = (qi["quantity"] if qi and qi["quantity"] is not None else None)
                if quantity is None:
                    quantity = await conn.fetchval(
                        "SELECT quantity FROM procurement_rfq_items WHERE id = $1", item_id
                    )
                currency = award.get("currency") or row["currency"] or "VND"
                quote_item_id = qi["id"] if qi else None

                prev = await conn.fetchrow(
                    "SELECT id, vendor_id FROM procurement_awards "
                    "WHERE batch_id = $1 AND item_id = $2 AND superseded_by IS NULL",
                    batch_id, item_id,
                )

                # Clear the prior row out of the ACTIVE partial unique index BEFORE
                # inserting the new active row. The partial UNIQUE index
                # (uq_pa_batch_item_active) is checked per-statement (non-deferrable),
                # so two active rows for the same (batch,item) would violate it even
                # mid-transaction. We point prev.superseded_by → prev.id (self) first
                # (FK-safe: the row exists), then repoint to the new id after insert.
                if prev:
                    await conn.execute(
                        "UPDATE procurement_awards SET superseded_by = id WHERE id = $1",
                        prev["id"],
                    )

                new_award_id = await conn.fetchval(
                    """
                    INSERT INTO procurement_awards
                        (batch_id, item_id, vendor_id, quote_id, quote_item_id,
                         awarded_price, currency, quantity, award_reason, awarded_by)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                    RETURNING id
                    """,
                    batch_id, item_id, vendor_id, row["quote_id"], quote_item_id,
                    price, currency, quantity, award_reason, actor,
                )
                award_ids.append(new_award_id)

                if prev:
                    await conn.execute(
                        "UPDATE procurement_awards SET superseded_by = $1 WHERE id = $2",
                        new_award_id, prev["id"],
                    )
                    superseded_count += 1

                await conn.execute(
                    """UPDATE procurement_rfq_items
                       SET awarded_vendor_id = $1, awarded_price = $2, awarded_currency = $3,
                           awarded_round = $4, awarded_quote_item_id = $5
                       WHERE id = $6 AND batch_id = $7""",
                    vendor_id, price, currency, award_round, quote_item_id,
                    item_id, batch_id,
                )

                await _audit(
                    conn, "award", new_award_id,
                    "re_award" if prev else "award",
                    actor_id=actor,
                    from_status=str(prev["vendor_id"]) if prev else None,
                    to_status="awarded",
                    detail={
                        "mode": "per_item", "item_id": item_id, "vendor_id": vendor_id,
                        "price": float(price) if price is not None else None,
                        "currency": currency, "quote_id": row["quote_id"],
                        "superseded_award_id": prev["id"] if prev else None,
                        "reason": award_reason,
                    },
                )

        # ── [MC] Đường rẽ maker-checker: đọc config + tính tổng award (sau block B) ──
        # Block B ở trên ĐÃ ghi procurement_awards + mirror rfq_items.awarded_* y như
        # cũ (cần để tính tổng + để approve dựng lại winners). Ở đây chỉ quyết định:
        # finalize NGAY hay TREO chờ duyệt.
        cfg = await _read_award_approval_config(conn)
        total_vnd, has_foreign = await _award_total_vnd(conn, batch_id)
        # Cần duyệt khi: gate BẬT VÀ (có dòng ngoại tệ — fail-safe — HOẶC tổng VND
        # >= ngưỡng). Gate TẮT (mặc định) → luôn finalize-ngay (hành vi cũ y hệt).
        need_approval = cfg["enabled"] and (has_foreign or total_vnd >= cfg["threshold"])

        if not need_approval:
            # ĐƯỜNG 1 — finalize NGAY (hành vi cũ). award_status giữ 'none'.
            await _finalize_award(
                conn, batch_id=batch_id, actor=actor, old_status=old_status,
                award_mode=award_mode, award_round=award_round,
                award_reason=award_reason,
                winning_quote_ids=winning_quote_ids,
                awarded_vendor_ids=awarded_vendor_ids,
            )
        else:
            # ĐƯỜNG 2 — TREO chờ NGƯỜI THỨ HAI duyệt. KHÔNG finalize: KHÔNG flip
            # vendor_quotes, KHÔNG set status='awarded'. batch.status về 'evaluating'
            # (nếu đang published) để khớp ngữ nghĩa "đang xét", award_status='proposed'.
            # 8 trạng thái status KHÔNG bị refactor — chỉ thêm cột award_status.
            await conn.execute(
                "UPDATE procurement_rfq_batches "
                "SET award_status = 'proposed', award_proposed_by = $1, "
                "    award_proposed_at = NOW(), "
                "    status = CASE WHEN status = 'published' THEN 'evaluating' ELSE status END "
                "WHERE id = $2",
                actor, batch_id,
            )
            await _audit(
                conn, "batch", batch_id, "award_proposed",
                actor_id=actor, from_status=old_status, to_status="proposed",
                detail={
                    "award_mode": award_mode, "reason": award_reason,
                    "winner_count": len(winning_quote_ids), "round": award_round,
                    "total_vnd": total_vnd, "has_foreign_currency": has_foreign,
                    "threshold_vnd": cfg["threshold"],
                    "prev_status": old_status,  # approve dựng lại action re_award/award
                },
            )
            # Notify nội bộ "có award chờ duyệt" — KHÔNG báo NCC (awarded_vendor_id=None).
            await dispatch_procurement_event(
                conn, "award", batch_id, "award_proposed",
                actor_id=actor, awarded_vendor_id=None,
                detail={"batch_id": batch_id, "award_mode": award_mode,
                        "winner_count": len(winning_quote_ids), "round": award_round},
            )

    return {
        "message": ("Đã gửi đề xuất chốt thầu — chờ người thứ hai duyệt"
                    if need_approval else "Đã chọn nhà cung cấp trúng thầu"),
        "data": {
            "awarded_count": len(award_ids),
            "award_ids": award_ids,
            "superseded_count": superseded_count,
            "award_status": "proposed" if need_approval else "none",
        },
    }


# ---------------------------------------------------------------------------
# Đợt 3 [MC] — Maker-checker AWARD: duyệt / từ chối đề xuất chốt thầu
# ---------------------------------------------------------------------------

@router.post("/batches/{batch_id}/approve-award")
async def approve_award(
    batch_id: int,
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 3 [MC] — DUYỆT đề xuất chốt thầu (award_status proposed → approved),
    rồi FINALIZE (dùng chung _finalize_award với đường finalize-ngay — DRY).

    SoD: checker PHẢI KHÁC proposer. Cùng người → 403, TRỪ KHI breakglass bật:
    cho duyệt nhưng _audit 'award_breakglass' (cờ cảnh báo compliance) + notify
    internal team. awards đã được ghi sẵn lúc PROPOSE (block B của award_batch) →
    đây chỉ flip quote-statuses + set batch='awarded' + notify NCC trúng thầu.
    """
    batch = await conn.fetchrow(
        "SELECT id, status, award_mode, current_round, award_status, award_proposed_by "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["award_status"] != "proposed":
        raise HTTPException(400, "Không có đề xuất chốt thầu nào đang chờ duyệt")

    actor = token_data.user_id
    cfg = await _read_award_approval_config(conn)
    is_self = str(batch["award_proposed_by"]) == str(actor)
    breakglass_used = False
    if is_self:
        if not cfg["breakglass"]:
            raise HTTPException(
                403, "Không thể tự duyệt đề xuất chốt thầu mình tạo (cần người thứ hai)"
            )
        breakglass_used = True  # cho phép nhưng đánh cờ cảnh báo + notify team

    # Dựng lại winners/vendors từ awards ACTIVE đã ghi lúc propose.
    award_rows = await conn.fetch(
        "SELECT DISTINCT quote_id, vendor_id FROM procurement_awards "
        "WHERE batch_id = $1 AND superseded_by IS NULL",
        batch_id,
    )
    winning_quote_ids = {r["quote_id"] for r in award_rows if r["quote_id"] is not None}
    awarded_vendor_ids = {r["vendor_id"] for r in award_rows if r["vendor_id"] is not None}

    # Lấy prev_status + reason từ audit 'award_proposed' để dựng action đúng
    # (re_award nếu batch từng awarded trước khi propose).
    proposed = await conn.fetchrow(
        "SELECT detail::text AS detail FROM procurement_audit_log "
        "WHERE entity_type = 'batch' AND entity_id = $1 AND action = 'award_proposed' "
        "ORDER BY id DESC LIMIT 1",
        batch_id,
    )
    pdetail: dict[str, Any] = {}
    if proposed and proposed["detail"]:
        try:
            pdetail = _json.loads(proposed["detail"])
        except (ValueError, TypeError):
            pdetail = {}
    old_status = pdetail.get("prev_status", "evaluating")
    award_reason = pdetail.get("reason", "")
    award_round = batch["current_round"] or 1

    async with conn.transaction():
        if breakglass_used:
            # Cờ cảnh báo compliance + notify team. Ghi TRƯỚC finalize.
            await _audit(
                conn, "batch", batch_id, "award_breakglass",
                actor_id=actor, from_status="proposed", to_status="proposed",
                detail={"proposer": str(batch["award_proposed_by"]), "approver": str(actor),
                        "note": "Tự duyệt qua break-glass — cần hậu kiểm"},
            )
            # actor_id=None → KHÔNG loại actor khỏi fan-out → chính người tự duyệt
            # cũng có 1 dòng nhắc "đã dùng break-glass" (lưu vết).
            await dispatch_procurement_event(
                conn, "award", batch_id, "award_breakglass",
                actor_id=None, awarded_vendor_id=None,
                detail={"batch_id": batch_id},
            )

        # Set award_status=approved + ghi approver, RỒI finalize (dùng chung — DRY).
        await conn.execute(
            "UPDATE procurement_rfq_batches "
            "SET award_status = 'approved', award_approved_by = $1, award_approved_at = NOW() "
            "WHERE id = $2",
            actor, batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "award_approved",
            actor_id=actor, from_status="proposed", to_status="approved",
            detail={"proposer": str(batch["award_proposed_by"]), "approver": str(actor),
                    "breakglass": breakglass_used},
        )
        await _finalize_award(
            conn, batch_id=batch_id, actor=actor, old_status=old_status,
            award_mode=batch["award_mode"], award_round=award_round,
            award_reason=award_reason,
            winning_quote_ids=winning_quote_ids,
            awarded_vendor_ids=awarded_vendor_ids,
        )

    return {"message": "Đã duyệt & chốt thầu", "data": {"breakglass": breakglass_used}}


@router.post("/batches/{batch_id}/reject-award")
async def reject_award(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 3 [MC] — TỪ CHỐI đề xuất chốt thầu: award_status proposed → none,
    batch về 'evaluating'. ROLLBACK các award đã ghi lúc propose (supersede +
    clear rfq_items.awarded_*) để KHÔNG còn award active 'ma'. Body {reason}.

    SoD: KHÔNG chặn self-reject — proposer tự rút đề xuất của mình là hợp lệ
    (không phát sinh công nợ). KISS.
    """
    reason = (body.get("reason") or "").strip()
    if not reason:
        raise HTTPException(400, "Cần nhập lý do từ chối")

    batch = await conn.fetchrow(
        "SELECT id, award_status, award_proposed_by FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["award_status"] != "proposed":
        raise HTTPException(400, "Không có đề xuất chốt thầu nào đang chờ duyệt")

    actor = token_data.user_id
    async with conn.transaction():
        # Rollback awards 'ma': supersede mọi award active của batch (self-point,
        # FK-safe vì row tồn tại). KHÔNG flip vendor_quotes (chưa từng flip lúc propose).
        await conn.execute(
            "UPDATE procurement_awards SET superseded_by = id "
            "WHERE batch_id = $1 AND superseded_by IS NULL",
            batch_id,
        )
        # Clear mirror awarded_* trên rfq_items (đồng bộ với supersede).
        await conn.execute(
            "UPDATE procurement_rfq_items "
            "SET awarded_vendor_id = NULL, awarded_price = NULL, awarded_currency = NULL, "
            "    awarded_round = NULL, awarded_quote_item_id = NULL "
            "WHERE batch_id = $1",
            batch_id,
        )
        # Reset cờ + về evaluating.
        await conn.execute(
            "UPDATE procurement_rfq_batches "
            "SET award_status = 'none', status = 'evaluating', "
            "    award_proposed_by = NULL, award_proposed_at = NULL "
            "WHERE id = $1",
            batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "award_rejected",
            actor_id=actor, from_status="proposed", to_status="none",
            detail={"proposer": str(batch["award_proposed_by"]), "approver": str(actor),
                    "reason": reason},
        )
    return {"message": "Đã từ chối đề xuất chốt thầu. Đợt quay lại trạng thái xét."}


# ---------------------------------------------------------------------------
# P6 — Award → catalog write-back (the "sourcing flywheel")
# ---------------------------------------------------------------------------
#
# When a batch is AWARDED, the owner confirms "Lưu giá trúng vào Thư viện nguồn
# cung". Each winning (active, not-yet-written-back) award becomes a
# sourcing_supplier_prices row so the catalog self-enriches for the next RFQ.
#
# IDEMPOTENCY: primary guard is procurement_awards.written_back_to_sourcing —
# sourcing_supplier_prices has NO unique index suitable for ON CONFLICT, so we
# NEVER rely on ON CONFLICT. A secondary exact-duplicate guard (same
# entry+supplier+currency+cost) avoids re-adding a row on a partial re-run, but
# the flag is the source of truth.
#
# COST SEMANTICS: the awarded vendor price IS Song Chau's cost for that item →
# award price maps to sourcing_supplier_prices.cost_amount, currency = award
# currency. We DO NOT invent FX here: procurement_awards carries no exchange
# rate, so we only set cost_vnd_equiv when currency='VND' (1:1) and leave it
# NULL otherwise for the existing sourcing pricing engine to fill later.

# Mirror of sourcing.model_norm GENERATED column (UPPER + strip non-alnum).
_MODEL_NORM_RE = re.compile(r"[^A-Z0-9]")


def _model_norm(s: str | None) -> str:
    """UPPER + strip non-alphanumeric — matches sourcing_entries.model_norm."""
    if not s:
        return ""
    return _MODEL_NORM_RE.sub("", s.upper())


@router.post("/batches/{batch_id}/write-back-sourcing")
async def write_back_sourcing(
    batch_id: int,
    body: dict[str, Any] | None = None,
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """P6 — Lưu giá trúng vào Thư viện nguồn cung (award → sourcing write-back).

    For each ACTIVE award (superseded_by IS NULL) of the batch that has NOT yet
    been written back, resolve its target sourcing_entry and INSERT a
    sourcing_supplier_prices row (the awarded price = Song Chau's COST).

    Resolution:
      1. award.rfq_item.source_kind='catalog' → source_ref_id = sourcing_entries.id
         (DIRECT link).
      2. Else match an existing sourcing_entries row by normalized item code
         (model_norm against item_code/bqms_code/model), then by raw model.
      3. No target found → SKIP (do NOT auto-create a sourcing_entry in P6),
         report in skipped[] with the item_code.

    Body (optional): {"award_ids": [int]} → write back a subset; default = ALL
    active awards of the batch.

    Idempotent: awards already written_back are reported in already_done and
    never re-processed. ONE transaction.

    Returns: {written:[{award_id,item_code,supplier_name,price,sourcing_entry_id}],
              skipped:[{item_code,reason}], already_done:int}.
    """
    body = body or {}
    batch = await conn.fetchrow(
        "SELECT id, status, award_status FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    # Đợt 3 [MC] — CỔNG TÀI CHÍNH: khi award đang chờ duyệt (proposed) thì awards
    # active đã tồn tại nhưng CHƯA được người-thứ-hai duyệt → cấm đẩy giá nguồn
    # (phát sinh công nợ/cost) cho tới khi approve-award. Fail-fast trước transaction.
    if batch["award_status"] == "proposed":
        raise HTTPException(
            409,
            "Đề xuất chốt thầu chưa được duyệt — không thể đẩy giá vào Thư viện nguồn "
            "cung khi đang chờ duyệt (cổng tài chính).",
        )

    subset = body.get("award_ids")
    subset_ids: list[int] | None = None
    if subset:
        subset_ids = [int(x) for x in subset if x is not None]

    actor = token_data.user_id

    # Active awards of the batch, joined to their rfq item for provenance/matching.
    # Per-batch awards (item_id IS NULL) carry no single item → cannot be matched
    # to one sourcing_entry, so they are reported as skipped (per_batch_award).
    rows = await conn.fetch(
        """
        SELECT a.id AS award_id, a.item_id, a.vendor_id, a.awarded_price,
               a.currency, a.written_back_to_sourcing,
               i.source_kind, i.source_ref_id,
               i.item_code, i.bqms_code, i.model, i.product_name,
               va.company_name AS supplier_name
          FROM procurement_awards a
          LEFT JOIN procurement_rfq_items i ON i.id = a.item_id
          LEFT JOIN vendor_accounts va ON va.id = a.vendor_id
         WHERE a.batch_id = $1 AND a.superseded_by IS NULL
         ORDER BY a.id
        """,
        batch_id,
    )

    written: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    already_done = 0

    async with conn.transaction():
        for r in rows:
            award_id = r["award_id"]
            if subset_ids is not None and award_id not in subset_ids:
                continue

            if r["written_back_to_sourcing"]:
                already_done += 1
                continue

            item_code = (r["item_code"] or r["bqms_code"] or r["model"] or "").strip() or None

            # Per-batch award has no single item — cannot map to one entry.
            if r["item_id"] is None:
                skipped.append({"item_code": item_code, "reason": "per_batch_award"})
                continue

            price = r["awarded_price"]
            if price is None:
                skipped.append({"item_code": item_code, "reason": "no_price"})
                continue

            supplier_name = (r["supplier_name"] or "").strip()
            if not supplier_name:
                skipped.append({"item_code": item_code, "reason": "no_supplier"})
                continue

            currency = r["currency"] or "VND"

            # ── 1. Resolve target sourcing_entry ──
            entry_id: int | None = None
            if r["source_kind"] == "catalog" and r["source_ref_id"] is not None:
                entry_id = await conn.fetchval(
                    "SELECT id FROM sourcing_entries WHERE id = $1 AND deleted_at IS NULL",
                    int(r["source_ref_id"]),
                )

            if entry_id is None:
                # Match by normalized code (model_norm) against any code-ish field,
                # then by raw model. Latest non-deleted entry wins.
                norm = _model_norm(r["item_code"] or r["bqms_code"] or r["model"])
                if norm:
                    entry_id = await conn.fetchval(
                        """
                        SELECT id FROM sourcing_entries
                         WHERE deleted_at IS NULL AND model_norm = $1
                         ORDER BY inquiry_date DESC NULLS LAST, id DESC
                         LIMIT 1
                        """,
                        norm,
                    )
                if entry_id is None and (r["model"] or "").strip():
                    entry_id = await conn.fetchval(
                        """
                        SELECT id FROM sourcing_entries
                         WHERE deleted_at IS NULL AND model = $1
                         ORDER BY inquiry_date DESC NULLS LAST, id DESC
                         LIMIT 1
                        """,
                        r["model"].strip(),
                    )

            if entry_id is None:
                # Unmatched commercial item not yet in the catalog → SKIP (no
                # auto-create in P6); report so the owner can add it manually.
                skipped.append({"item_code": item_code, "reason": "no_catalog_entry"})
                continue

            entry_id = int(entry_id)
            # VND-only 1:1 equiv; other currencies left NULL for the pricing engine.
            cost_vnd_equiv = float(price) if currency == "VND" else None
            fx_used = 1.0 if currency == "VND" else None

            # ── 2. Secondary exact-duplicate guard (flag is the primary guard) ──
            dup_id = await conn.fetchval(
                """
                SELECT id FROM sourcing_supplier_prices
                 WHERE sourcing_entry_id = $1 AND supplier_name = $2
                   AND currency = $3 AND cost_amount = ROUND($4::numeric, 4)
                 ORDER BY id DESC LIMIT 1
                """,
                entry_id, supplier_name, currency, price,
            )
            if dup_id is not None:
                ssp_id = int(dup_id)
            else:
                ssp_id = await conn.fetchval(
                    """
                    INSERT INTO sourcing_supplier_prices
                        (sourcing_entry_id, supplier_name, currency, cost_amount,
                         cost_vnd_equiv, exchange_rate_used, notes, is_primary)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, false)
                    RETURNING id
                    """,
                    entry_id, supplier_name, currency, price,
                    cost_vnd_equiv, fx_used,
                    f"Tự động từ đấu thầu NCC (batch #{batch_id}, award #{award_id})",
                )
                ssp_id = int(ssp_id)

            # ── 3. Stamp the award as written back ──
            await conn.execute(
                """
                UPDATE procurement_awards
                   SET written_back_to_sourcing = true,
                       written_back_at = NOW(),
                       written_back_by = $1,
                       sourcing_supplier_price_id = $2
                 WHERE id = $3
                """,
                actor, ssp_id, award_id,
            )

            matched_by = (
                "source_ref" if r["source_kind"] == "catalog"
                and r["source_ref_id"] is not None else "code_or_model"
            )

            await _audit(
                conn, "award", award_id, "catalog_write_back",
                actor_id=actor,
                detail={
                    "batch_id": batch_id, "item_code": item_code,
                    "sourcing_entry_id": entry_id,
                    "sourcing_supplier_price_id": ssp_id,
                    "supplier_name": supplier_name,
                    "currency": currency,
                    "price": float(price) if price is not None else None,
                    "matched_by": matched_by,
                    "duplicate_reused": dup_id is not None,
                },
            )

            written.append({
                "award_id": award_id,
                "item_code": item_code,
                "supplier_name": supplier_name,
                "price": float(price) if price is not None else None,
                "sourcing_entry_id": entry_id,
                "matched_by": matched_by,
            })

    return {
        "message": (
            f"Đã lưu {len(written)} giá trúng vào thư viện nguồn cung"
            + (f", bỏ qua {len(skipped)}" if skipped else "")
            + (f", {already_done} đã lưu trước đó" if already_done else "")
        ),
        "data": {
            "written": written,
            "skipped": skipped,
            "already_done": already_done,
        },
    }


@router.patch("/batches/{batch_id}/evaluating")
async def set_batch_evaluating(
    batch_id: int,
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt-2 (KISS) — chuyển published → evaluating khi admin mở ma trận so sánh.

    Award vẫn chạy trực tiếp từ 'published'; endpoint này chỉ để đánh dấu UX.
    """
    batch = await conn.fetchrow(
        "SELECT id, status FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")
    if batch["status"] != "published":
        raise HTTPException(400, "Chỉ chuyển sang đánh giá từ trạng thái đã công bố")

    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_rfq_batches SET status = 'evaluating', evaluating_at = NOW() WHERE id = $1",
            batch_id,
        )
        await _audit(
            conn, "batch", batch_id, "status_change",
            actor_id=token_data.user_id, from_status="published", to_status="evaluating",
        )

    return {"message": "Đã chuyển sang trạng thái đánh giá"}


@router.post("/batches/{batch_id}/open-round")
async def open_round(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt-2 — Mở vòng đấu giá kế tiếp (V2/V3, reverse-auction).

    Guards: round_number 2..3 và <= max_rounds; round_number == current_round+1
    (không nhảy/lùi); batch ở 'published'/'evaluating'/'awarded' (mở lại sau
    award tạm cho reverse-auction; nếu đang 'awarded' → set lại 'published').
    Mỗi vendor phải active VÀ đã có invitation vòng (round-1) cho batch này
    (carry-forward). Insert invitation vòng N (ON CONFLICT DO NOTHING), gửi
    email login, ghi audit.

    Body: {"round_number":int, "vendor_ids":[int], "deadline"?:ISO8601,
           "message"?:str, "send_email"?:bool=true}
    """
    batch = await conn.fetchrow(
        "SELECT id, status, batch_code, title, item_count, current_round, max_rounds "
        "FROM procurement_rfq_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    round_number = int(body.get("round_number") or 0)
    max_rounds = batch["max_rounds"] or 1
    current_round = batch["current_round"] or 1

    if round_number < 2 or round_number > 3 or round_number > max_rounds:
        raise HTTPException(400, f"Vượt số vòng tối đa (max_rounds={max_rounds})")
    if round_number != current_round + 1:
        raise HTTPException(400, f"Chỉ mở được vòng kế tiếp (vòng {current_round + 1})")
    if batch["status"] not in ("published", "evaluating", "awarded"):
        raise HTTPException(400, "Chỉ mở vòng mới khi đợt đã công bố")

    vendor_ids = body.get("vendor_ids") or []
    if not vendor_ids:
        raise HTTPException(400, "Cần ít nhất 1 vendor_id")
    vendor_ids = list(dict.fromkeys(int(v) for v in vendor_ids))

    deadline_dt = _parse_deadline(body.get("deadline"))
    # Thang 30/06: BỎ gửi mail mời lại — admin tự gửi link đăng nhập cho NCC.
    do_send_email = False
    prev_round = round_number - 1

    base = _vendor_portal_base()
    created: list[dict] = []
    skipped_existing: list[int] = []
    failures: list[dict] = []

    async with conn.transaction():
        # Advance round + status + per-round deadline columns. ROUND-AWARE:
        # bid_deadline tracks the LATEST active round so the auto-close +
        # reminder sweep keys off the round that is now open. We update both
        # the legacy deadline_vN twin (kept for back-compat) and the canonical
        # deadline_roundN; bid_deadline = the new round's deadline (the value
        # just provided, else the existing roundN deadline).
        v_col = {2: "deadline_v2", 3: "deadline_v3"}[round_number]
        round_col = {2: "deadline_round2", 3: "deadline_round3"}[round_number]
        await conn.execute(
            f"""UPDATE procurement_rfq_batches
                   SET current_round = $1,
                       status = 'published',
                       {v_col}     = COALESCE($2, {v_col}),
                       {round_col} = COALESCE($2, {round_col}),
                       bid_deadline = COALESCE($2, {round_col}, bid_deadline)
                 WHERE id = $3""",
            round_number, deadline_dt, batch_id,
        )

        # A NEW round reopens the reminder window: clear reminder_sent_at for
        # the carried-forward invitations of THIS batch so the sweep can remind
        # again for the new round (idempotency guard resets per round).
        await conn.execute(
            "UPDATE procurement_rfq_invitations "
            "SET reminder_sent_at = NULL WHERE batch_id = $1",
            batch_id,
        )

        for vid in vendor_ids:
            va = await conn.fetchrow(
                """
                SELECT va.id, va.company_name, va.contact_name, va.status, va.is_approved, u.email
                FROM vendor_accounts va
                LEFT JOIN users u ON u.id = va.user_id
                WHERE va.id = $1
                """,
                vid,
            )
            if not va:
                failures.append({"vendor_id": vid, "error": "Tài khoản NCC không tồn tại"})
                continue
            is_active = str(va["status"]) == "active" or va["is_approved"] is True
            if not is_active:
                failures.append({"vendor_id": vid, "error": "Tài khoản NCC chưa active"})
                continue
            # Carry-forward only: must have a previous-round invitation for THIS batch.
            had_prev = await conn.fetchval(
                "SELECT 1 FROM procurement_rfq_invitations "
                "WHERE batch_id = $1 AND vendor_id = $2 AND round_number = $3",
                batch_id, vid, prev_round,
            )
            if not had_prev:
                failures.append({"vendor_id": vid, "error": f"NCC chưa tham gia vòng {prev_round}"})
                continue

            invitation_id = await conn.fetchval(
                """
                INSERT INTO procurement_rfq_invitations
                    (batch_id, vendor_id, round_number, status, invited_by, invited_at)
                VALUES ($1, $2, $3, 'invited', $4, NOW())
                ON CONFLICT (batch_id, vendor_id, round_number) DO NOTHING
                RETURNING id
                """,
                batch_id, vid, round_number, token_data.user_id,
            )
            if invitation_id is None:
                skipped_existing.append(vid)
                continue

            await _audit(
                conn, "invitation", invitation_id, "open_round",
                actor_id=token_data.user_id,
                detail={"vendor_id": vid, "round": round_number},
            )

            email = (va["email"] or "").strip().lower()
            rec: dict[str, Any] = {
                "invitation_id": invitation_id,
                "vendor_id": vid,
                "email_sent": False,
            }
            if do_send_email and email and "@" in email:
                login_url = f"{base}/login?next=/batches/{batch_id}"
                subject = (
                    f"[Song Châu] Vòng {round_number} — mời báo giá lại phiên "
                    f"#{batch['batch_code']} — {batch['title']}"
                )
                body_html = _build_login_invitation_email(
                    batch_code=batch["batch_code"],
                    batch_title=batch["title"],
                    invitee_name=va["contact_name"] or va["company_name"] or email,
                    login_url=login_url,
                    item_count=batch["item_count"] or 0,
                    round_number=round_number,
                )
                try:
                    await send_email([email], subject, body_html)
                    await conn.execute(
                        """UPDATE procurement_rfq_invitations
                           SET email_sent=true, email_sent_at=NOW(),
                               email_status='sent', email_subject=$1
                           WHERE id=$2""",
                        subject, invitation_id,
                    )
                    rec["email_sent"] = True
                except Exception as exc:
                    logger.warning("Email mở vòng NCC %s thất bại: %s", vid, exc)
                    await conn.execute(
                        """UPDATE procurement_rfq_invitations
                           SET email_status='failed', email_error=$1
                           WHERE id=$2""",
                        str(exc)[:500], invitation_id,
                    )
                    rec["email_error"] = str(exc)[:200]
            elif do_send_email:
                rec["email_error"] = "NCC không có email hợp lệ"

            created.append(rec)

        await _audit(
            conn, "batch", batch_id, "open_round",
            actor_id=token_data.user_id, from_status=batch["status"], to_status="published",
            detail={
                "round": round_number,
                "vendor_count": len(created),
                "deadline": deadline_dt.isoformat() if deadline_dt else None,
            },
        )

    logger.info(
        "Opened round %d for batch %d: created=%d skipped=%d failures=%d",
        round_number, batch_id, len(created), len(skipped_existing), len(failures),
    )
    return {
        "data": {
            "round_number": round_number,
            "created": created,
            "skipped_existing": skipped_existing,
            "failures": failures,
        },
        "message": f"Đã mở vòng {round_number} cho {len(created)} nhà cung cấp",
    }


@router.get("/batches/{batch_id}/audit")
async def batch_audit_timeline(
    batch_id: int,
    entity_type: str | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt-2 — Dòng thời gian audit của 1 đợt + các thực thể con.

    Gồm: batch row + invitation/quote/award rows thuộc batch. JOIN tên người
    thao tác (users.full_name) và tên NCC (vendor_accounts.company_name).
    ?entity_type= lọc về 1 loại (batch|invitation|quote|award).
    """
    batch = await conn.fetchval(
        "SELECT id FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    rows = await conn.fetch(
        """
        SELECT al.id, al.entity_type, al.entity_id, al.action,
               al.from_status, al.to_status,
               al.actor_id, u.full_name AS actor_name,
               al.actor_vendor_id, va.company_name AS actor_vendor_name,
               al.detail, al.created_at
        FROM procurement_audit_log al
        LEFT JOIN users u ON u.id = al.actor_id
        LEFT JOIN vendor_accounts va ON va.id = al.actor_vendor_id
        WHERE (
                (al.entity_type = 'batch'      AND al.entity_id = $1)
             OR (al.entity_type = 'invitation' AND al.entity_id IN
                    (SELECT id FROM procurement_rfq_invitations WHERE batch_id = $1))
             OR (al.entity_type = 'quote'      AND al.entity_id IN
                    (SELECT id FROM vendor_quotes WHERE batch_id = $1))
             OR (al.entity_type = 'award'      AND al.entity_id IN
                    (SELECT id FROM procurement_awards WHERE batch_id = $1))
            )
          AND ($2::text IS NULL OR al.entity_type = $2)
        ORDER BY al.created_at DESC
        LIMIT $3
        """,
        batch_id, entity_type, limit,
    )

    out = []
    for r in rows:
        d = dict(r)
        if isinstance(d.get("detail"), str):
            try:
                d["detail"] = _json.loads(d["detail"])
            except (ValueError, TypeError):
                pass
        out.append(d)
    return {"data": out}


# ═══════════════════════════════════════════════════════════════════
# Phase 2 — Contracts (Thang 2026-05-14)
# Sau khi award batch → admin tạo contract draft → gửi NCC sign → active.
# ═══════════════════════════════════════════════════════════════════

@router.get("/contracts")
async def list_contracts(
    status: str | None = Query(None),
    vendor_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách hợp đồng NCC."""
    where = "1=1"
    params: list = []
    idx = 1
    if status and status != "all":
        where += f" AND c.status = ${idx}"
        params.append(status); idx += 1
    if vendor_id:
        where += f" AND c.vendor_id = ${idx}"
        params.append(vendor_id); idx += 1

    total = await conn.fetchval(f"SELECT COUNT(*) FROM procurement_contracts c WHERE {where}", *params)
    params_paged = params + [limit, (page - 1) * limit]
    rows = await conn.fetch(
        f"""
        SELECT c.*, b.batch_code, b.title AS batch_title,
               (SELECT COUNT(*) FROM procurement_contract_items WHERE contract_id = c.id) AS item_count,
               (SELECT COUNT(*) FROM procurement_pos WHERE contract_id = c.id) AS po_count
        FROM procurement_contracts c
        LEFT JOIN procurement_rfq_batches b ON b.id = c.batch_id
        WHERE {where}
        ORDER BY c.created_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params_paged,
    )
    return {"data": [dict(r) for r in rows], "total": total}


@router.get("/contracts/{contract_id}")
async def get_contract(
    contract_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Chi tiết hợp đồng + items + linked POs."""
    row = await conn.fetchrow(
        """
        SELECT c.*, b.batch_code, b.title AS batch_title
        FROM procurement_contracts c
        LEFT JOIN procurement_rfq_batches b ON b.id = c.batch_id
        WHERE c.id = $1
        """,
        contract_id,
    )
    if not row:
        raise HTTPException(404)
    items = await conn.fetch(
        "SELECT * FROM procurement_contract_items WHERE contract_id = $1 ORDER BY item_no",
        contract_id,
    )
    pos = await conn.fetch(
        "SELECT id, po_no, po_date, total_amount, status FROM procurement_pos WHERE contract_id = $1 ORDER BY po_date DESC",
        contract_id,
    )
    return {
        "data": {
            **dict(row),
            "items": [dict(i) for i in items],
            "pos": [dict(p) for p in pos],
        }
    }


@router.post("/batches/{batch_id}/create-contract")
async def create_contract_from_batch(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Auto-create contract draft từ ACTIVE awards của batch cho 1 vendor.

    Body: {"vendor_id": int, "payment_terms": str?, "delivery_terms": str?,
           "warranty_terms": str?, "effective_date": date?, "expiry_date": date?}

    HARDENED (Đợt 3):
      - Đọc winning items từ procurement_awards (active = superseded_by IS NULL)
        JOIN procurement_rfq_items để lấy spec/unit/bqms_code — đây là single
        source of truth của award, không dựa vào awarded_* snapshot có thể lệch.
      - INSERT contract + items + audit GÓI TRONG MỘT TRANSACTION.
      - total_amount tính từ awarded_price × quantity của chính các award đó.
      - CHẶN tạo trùng: nếu đã có contract chưa cancelled cho (batch,vendor) → 400.
      - GIỮ fallback magic-link (vendor_id NULL) nhưng LOG khi xảy ra.
    """
    vendor_id = body.get("vendor_id")
    if not vendor_id:
        raise HTTPException(400, "vendor_id bắt buộc")
    vendor_id = int(vendor_id)

    batch = await conn.fetchrow("SELECT id, batch_code, title, award_status FROM procurement_rfq_batches WHERE id = $1", batch_id)
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    # Đợt 3 [MC] — CỔNG TÀI CHÍNH: khi award đang chờ duyệt (proposed) thì awards
    # active đã tồn tại nhưng CHƯA qua người-thứ-hai → cấm tạo hợp đồng (→ PO →
    # công nợ) cho tới khi approve-award. Fail-fast trước mọi INSERT.
    if batch["award_status"] == "proposed":
        raise HTTPException(
            409,
            "Đề xuất chốt thầu chưa được duyệt — không thể tạo hợp đồng/PO khi đang "
            "chờ duyệt (cổng tài chính).",
        )

    # Reject a duplicate non-cancelled contract for the same (batch, vendor).
    dup = await conn.fetchrow(
        """SELECT id, contract_no, status FROM procurement_contracts
           WHERE batch_id = $1 AND vendor_id = $2 AND status <> 'cancelled'
           ORDER BY id DESC LIMIT 1""",
        batch_id, vendor_id,
    )
    if dup:
        raise HTTPException(
            409,
            f"Đã có hợp đồng {dup['contract_no']} (status={dup['status']}) cho NCC này. "
            "Huỷ hợp đồng cũ trước khi tạo mới.",
        )

    # ACTIVE awards for this vendor on this batch, joined to rfq_items for the
    # spec/unit/bqms_code. per_item awards carry item_id; a per_batch award
    # (item_id NULL) covers every item the winner quoted → fall back to the
    # rfq_items.awarded_* mirror in that case.
    award_items = await conn.fetch(
        """
        SELECT a.item_id, a.awarded_price, a.currency, a.quantity AS award_qty,
               i.id AS rfq_item_id, i.item_no, i.bqms_code, i.specification,
               i.unit, i.quantity AS rfq_qty
        FROM procurement_awards a
        JOIN procurement_rfq_items i ON i.id = a.item_id
        WHERE a.batch_id = $1 AND a.vendor_id = $2
          AND a.item_id IS NOT NULL AND a.superseded_by IS NULL
        ORDER BY i.item_no
        """,
        batch_id, vendor_id,
    )

    rows: list[dict] = []
    award_currency = "VND"
    if award_items:
        for a in award_items:
            qty = a["award_qty"] if a["award_qty"] is not None else a["rfq_qty"]
            rows.append({
                "rfq_item_id": int(a["rfq_item_id"]),
                "item_no": a["item_no"],
                "bqms_code": a["bqms_code"],
                "specification": a["specification"],
                "quantity": float(qty or 0),
                "unit": a["unit"] or "EA",
                "unit_price": float(a["awarded_price"] or 0),
            })
        award_currency = award_items[0]["currency"] or "VND"
    else:
        # No per_item awards → maybe a per_batch award. Fall back to the
        # awarded_* mirror on rfq_items (kept in sync by award_batch).
        per_batch = await conn.fetchrow(
            """SELECT id, currency FROM procurement_awards
               WHERE batch_id = $1 AND vendor_id = $2
                 AND item_id IS NULL AND superseded_by IS NULL""",
            batch_id, vendor_id,
        )
        mirror = await conn.fetch(
            """SELECT id, item_no, bqms_code, specification, unit, quantity,
                      awarded_price, awarded_currency
               FROM procurement_rfq_items
               WHERE batch_id = $1 AND awarded_vendor_id = $2
               ORDER BY item_no""",
            batch_id, vendor_id,
        )
        if not mirror:
            raise HTTPException(400, "Chưa có item nào được award cho NCC này")
        for it in mirror:
            rows.append({
                "rfq_item_id": int(it["id"]),
                "item_no": it["item_no"],
                "bqms_code": it["bqms_code"],
                "specification": it["specification"],
                "quantity": float(it["quantity"] or 0),
                "unit": it["unit"] or "EA",
                "unit_price": float(it["awarded_price"] or 0),
            })
        award_currency = (mirror[0]["awarded_currency"]
                          or (per_batch["currency"] if per_batch else None) or "VND")

    # Get vendor info
    vendor = await conn.fetchrow(
        """
        SELECT va.company_name, va.contact_name, va.phone, va.address, va.tax_code, u.email
        FROM vendor_accounts va LEFT JOIN users u ON u.id = va.user_id
        WHERE va.id = $1
        """,
        vendor_id,
    )
    if not vendor:
        # Magic-link vendor — pull from latest quote (legacy path; should be rare
        # now that magic-link is removed). Log it for visibility.
        logger.warning(
            "create-contract: vendor_id=%s has no vendor_accounts row (legacy magic-link?) — "
            "pulling identity from latest awarded quote (batch %d)", vendor_id, batch_id,
        )
        q = await conn.fetchrow(
            "SELECT submitter_name AS company_name, submitter_email AS email, submitter_phone AS phone "
            "FROM vendor_quotes WHERE batch_id = $1 AND vendor_id IS NULL AND status='awarded' ORDER BY id DESC LIMIT 1",
            batch_id,
        )
        vendor = q

    # total from the contract items we are about to insert (no snapshot drift).
    total = sum(r["unit_price"] * r["quantity"] for r in rows)
    currency = award_currency

    async with conn.transaction():
        seq = await conn.fetchval("SELECT nextval('procurement_contract_seq')")
        contract_no = f"SC-CT-{datetime.now().year}-{seq:04d}"

        contract_id = await conn.fetchval(
            """
            INSERT INTO procurement_contracts (
                contract_no, batch_id, vendor_id, vendor_name, vendor_email, vendor_phone,
                vendor_tax_code, vendor_address, total_amount, currency,
                payment_terms, delivery_terms, warranty_terms,
                effective_date, expiry_date, contract_date,
                status, created_by, notes
            ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,'draft',$17,$18)
            RETURNING id
            """,
            contract_no, batch_id, vendor_id,
            (vendor and vendor.get("company_name")) or "Unknown vendor",
            vendor and vendor.get("email"),
            vendor and vendor.get("phone"),
            vendor and vendor.get("tax_code"),
            vendor and vendor.get("address"),
            total, currency,
            body.get("payment_terms") or "Thanh toán 100% trong 30 ngày sau giao hàng",
            body.get("delivery_terms") or "Giao tại kho Song Châu",
            body.get("warranty_terms") or "Bảo hành theo tiêu chuẩn nhà sản xuất",
            _as_date(body.get("effective_date")) or datetime.now().date(),
            _as_date(body.get("expiry_date")),
            datetime.now().date(),
            token_data.user_id,
            f"Auto-created từ phiên {batch['batch_code']}",
        )

        for r in rows:
            await conn.execute(
                """
                INSERT INTO procurement_contract_items (
                    contract_id, rfq_item_id, item_no, bqms_code,
                    specification, quantity, unit, unit_price, lead_time_days, notes
                ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                """,
                contract_id, r["rfq_item_id"], r["item_no"], r["bqms_code"],
                r["specification"], r["quantity"], r["unit"],
                r["unit_price"], None, None,
            )

        await _audit(
            conn, "contract", contract_id, "create",
            actor_id=token_data.user_id, to_status="draft",
            detail={
                "batch_id": batch_id, "vendor_id": vendor_id,
                "contract_no": contract_no, "item_count": len(rows),
                "total_amount": total, "currency": currency,
            },
        )
        await dispatch_procurement_event(
            conn, "contract", contract_id, "create",
            actor_id=token_data.user_id, awarded_vendor_id=vendor_id,
            detail={
                "batch_id": batch_id, "contract_id": contract_id,
                "contract_no": contract_no,
                "total_amount": total, "currency": currency,
            },
        )

    logger.info("Contract %s created by %s for vendor %s (batch %d, total %s %s)",
                contract_no, token_data.user_id, vendor_id, batch_id, total, currency)
    return {"data": {"id": contract_id, "contract_no": contract_no}, "message": f"Đã tạo {contract_no}"}


@router.put("/contracts/{contract_id}")
async def update_contract(
    contract_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Cập nhật terms/dates của contract (chỉ khi status='draft')."""
    row = await conn.fetchrow("SELECT status FROM procurement_contracts WHERE id = $1", contract_id)
    if not row:
        raise HTTPException(404)
    if row["status"] not in ("draft", "sent"):
        raise HTTPException(400, f"Không sửa được contract ở status='{row['status']}'")

    fields = ["payment_terms", "delivery_terms", "warranty_terms",
              "effective_date", "expiry_date", "contract_date", "notes"]
    date_fields = {"effective_date", "expiry_date", "contract_date"}
    updates = []
    params: list = []
    idx = 1
    for f in fields:
        if f in body and body[f] is not None:
            val = _as_date(body[f]) if f in date_fields else body[f]
            updates.append(f"{f} = ${idx}")
            params.append(val); idx += 1
    if not updates:
        raise HTTPException(400, "Không có trường nào")
    params.append(contract_id)
    await conn.execute(
        f"UPDATE procurement_contracts SET {', '.join(updates)}, updated_at=NOW() WHERE id = ${idx}",
        *params,
    )
    return {"message": "Đã cập nhật"}


@router.post("/contracts/{contract_id}/generate-pdf")
async def generate_contract_pdf_endpoint(
    contract_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 3 — Render PDF hợp đồng (Gotenberg, Song Châu letterhead).

    Allowed status IN (draft, sent, signed, active). Ghi contract_file_path +
    pdf_generated_at. KHÔNG đổi status. Idempotent (re-render đè cùng tên file).
    """
    c = await conn.fetchrow(
        "SELECT id, contract_no, status FROM procurement_contracts WHERE id = $1", contract_id
    )
    if not c:
        raise HTTPException(404, "Hợp đồng không tồn tại")
    if c["status"] not in ("draft", "sent", "signed", "active"):
        raise HTTPException(400, f"Status='{c['status']}' — không tạo PDF được")

    try:
        rel_path = await procurement_docs.generate_contract_pdf(conn, contract_id)
    except Exception as exc:
        logger.exception("generate_contract_pdf failed for #%s: %s", contract_id, exc)
        raise HTTPException(502, f"Tạo PDF lỗi: {exc}")

    async with conn.transaction():
        row = await conn.fetchrow(
            """UPDATE procurement_contracts
               SET contract_file_path = $1, pdf_generated_at = NOW(), updated_at = NOW()
               WHERE id = $2
               RETURNING contract_file_path, pdf_generated_at""",
            rel_path, contract_id,
        )
        await _audit(
            conn, "contract", contract_id, "generate_pdf",
            actor_id=token_data.user_id,
            detail={"contract_no": c["contract_no"], "path": rel_path},
        )

    return {
        "data": {
            "contract_file_path": row["contract_file_path"],
            "pdf_generated_at": row["pdf_generated_at"],
        },
        "message": "Đã tạo PDF hợp đồng",
    }


@router.get("/contracts/{contract_id}/pdf")
async def download_contract_pdf(
    contract_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 3 — Tải PDF hợp đồng đã render (admin/manager/staff/procurement)."""
    c = await conn.fetchrow(
        "SELECT contract_no, contract_file_path FROM procurement_contracts WHERE id = $1",
        contract_id,
    )
    if not c:
        raise HTTPException(404, "Hợp đồng không tồn tại")
    resolved = _resolve_contract_pdf(c["contract_file_path"])
    return FileResponse(
        str(resolved),
        media_type="application/pdf",
        filename=f"{c['contract_no']}.pdf",
    )


@router.post("/contracts/{contract_id}/send-to-vendor")
async def send_contract_to_vendor(
    contract_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Gửi hợp đồng cho NCC: chuyển draft→sent + thông báo in-app cổng NCC.

    (2026-06-30) NCC truy cập hợp đồng qua ĐĂNG NHẬP cổng (login-scoped) —
    KHÔNG phụ thuộc email. M365 hiện TRỐNG nên gửi email là BEST-EFFORT và TUYỆT
    ĐỐI không được chặn flip draft→sent: trước đây email lỗi raise 500 khiến HĐ
    kẹt 'draft' và NCC không bao giờ nhìn thấy (draft bị ẩn với NCC). Tự sinh PDF
    nếu admin chưa bấm 'Sinh PDF'. Ghi audit + dispatch in-app cho NCC.
    REUSE sent_to_vendor_at (không thêm cột sent_at).
    """
    c = await conn.fetchrow("SELECT * FROM procurement_contracts WHERE id = $1", contract_id)
    if not c:
        raise HTTPException(404, "Hợp đồng không tồn tại")
    if c["status"] != "draft":
        raise HTTPException(400, f"Status='{c['status']}' — chỉ gửi được khi 'draft'")

    # Auto-sinh PDF nếu chưa có (để 'nhấn Gửi NCC' luôn chạy một phát, không bắt
    # admin bấm 'Sinh PDF' trước). HĐ phải có PDF để NCC xem/ký.
    file_path = c["contract_file_path"]
    if not file_path:
        try:
            file_path = await procurement_docs.generate_contract_pdf(conn, contract_id)
        except Exception as exc:
            logger.exception("send-to-vendor: auto-generate PDF lỗi cho HĐ #%s", contract_id)
            raise HTTPException(502, f"Không tạo được PDF hợp đồng: {exc}")
        await conn.execute(
            "UPDATE procurement_contracts SET contract_file_path=$1, pdf_generated_at=NOW(), "
            "updated_at=NOW() WHERE id=$2",
            file_path, contract_id,
        )

    # PRIMARY (atomic): flip draft→sent + audit + dispatch in-app cho NCC. NCC
    # thấy HĐ ngay khi đăng nhập cổng — không cần email.
    async with conn.transaction():
        flipped = await conn.fetchval(
            "UPDATE procurement_contracts SET status='sent', sent_to_vendor_at=NOW(), "
            "updated_at=NOW() WHERE id=$1 AND status='draft' RETURNING id",
            contract_id,
        )
        if not flipped:
            raise HTTPException(409, "Hợp đồng đã đổi trạng thái, vui lòng tải lại")
        await _audit(
            conn, "contract", contract_id, "send_to_vendor",
            actor_id=token_data.user_id, from_status="draft", to_status="sent",
            detail={"contract_no": c["contract_no"], "vendor_email": c["vendor_email"]},
        )
        await dispatch_procurement_event(
            conn, "contract", contract_id, "send_to_vendor",
            actor_id=token_data.user_id, awarded_vendor_id=c["vendor_id"],
            detail={
                "batch_id": c["batch_id"], "contract_no": c["contract_no"],
                "total_amount": float(c["total_amount"]) if c["total_amount"] is not None else None,
                "currency": c["currency"],
            },
        )

    # BEST-EFFORT email (M365 có thể trống) — KHÔNG chặn flip ở trên.
    email_sent = False
    if c["vendor_email"]:
        try:
            link = f"{_vendor_portal_base()}/contracts/{contract_id}"
            subject = f"[Song Châu] Hợp đồng {c['contract_no']} — vui lòng ký xác nhận"
            await send_email([c["vendor_email"]], subject, _contract_email_html(dict(c), link))
            email_sent = True
        except Exception as exc:
            logger.warning("send-to-vendor: email best-effort lỗi cho HĐ #%s: %s", contract_id, exc)

    msg = (
        "Đã gửi hợp đồng cho NCC (email + cổng đăng nhập)."
        if email_sent
        else "Đã chuyển hợp đồng cho NCC — NCC đăng nhập cổng để xem & ký (email chưa cấu hình)."
    )
    return {"message": msg, "data": {"status": "sent", "email_sent": email_sent}}


@router.post("/contracts/{contract_id}/mark-signed")
async def mark_contract_signed(
    contract_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin", "manager", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đường ký GIẤY/thủ công: sent → signed (KHÔNG nhảy thẳng 'active').

    REPURPOSED (Đợt 3): trước đây nhảy thẳng sang 'active' bỏ qua 'signed' —
    nay chỉ set 'signed' (kèm signed_at, signed_by_vendor, signed_by_user) và
    cùng đi qua bước /activate như đường e-sign của NCC.

    Body: {"signed_by": str}
    """
    row = await conn.fetchrow(
        "SELECT status, contract_no, batch_id, vendor_id FROM procurement_contracts WHERE id = $1",
        contract_id,
    )
    if not row:
        raise HTTPException(404, "Hợp đồng không tồn tại")
    if row["status"] != "sent":
        raise HTTPException(400, f"Status='{row['status']}' — chỉ mark-signed được khi 'sent'")

    signed_by = (body.get("signed_by") or "").strip() or "Ký thủ công (admin xác nhận)"
    async with conn.transaction():
        await conn.execute(
            """UPDATE procurement_contracts
               SET status='signed', signed_at=NOW(), signed_by_vendor=$1,
                   signed_by_user=$2, updated_at=NOW()
               WHERE id = $3""",
            signed_by, token_data.user_id, contract_id,
        )
        await _audit(
            conn, "contract", contract_id, "mark_signed",
            actor_id=token_data.user_id, from_status="sent", to_status="signed",
            detail={"contract_no": row["contract_no"], "signed_by": signed_by, "channel": "manual"},
        )
        await dispatch_procurement_event(
            conn, "contract", contract_id, "mark_signed",
            actor_id=token_data.user_id, awarded_vendor_id=row["vendor_id"],
            detail={
                "batch_id": row["batch_id"], "contract_no": row["contract_no"],
                "signed_by": signed_by,
            },
        )
    return {"message": "Đã đánh dấu đã ký (signed)"}


@router.post("/contracts/{contract_id}/activate")
async def activate_contract(
    contract_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 3 — signed → active (sau khi NCC đã e-sign hoặc admin mark-signed).

    GUARD status='signed'. Ghi signed_by_user = admin kích hoạt. Sau bước này
    create-po mới nhận status='active'.
    """
    row = await conn.fetchrow("SELECT status, contract_no FROM procurement_contracts WHERE id = $1", contract_id)
    if not row:
        raise HTTPException(404, "Hợp đồng không tồn tại")
    if row["status"] != "signed":
        raise HTTPException(400, f"Status='{row['status']}' — chỉ kích hoạt được khi 'signed'")

    async with conn.transaction():
        await conn.execute(
            """UPDATE procurement_contracts
               SET status='active', signed_by_user=COALESCE(signed_by_user, $1), updated_at=NOW()
               WHERE id = $2""",
            token_data.user_id, contract_id,
        )
        await _audit(
            conn, "contract", contract_id, "activate",
            actor_id=token_data.user_id, from_status="signed", to_status="active",
            detail={"contract_no": row["contract_no"]},
        )
    return {"message": "Đã kích hoạt hợp đồng (active)"}


def _contract_email_html(c: dict, link: str) -> str:
    return f"""<!DOCTYPE html><html><body style="font-family:-apple-system,Segoe UI,Roboto,sans-serif;background:#f8fafc;padding:24px">
<div style="max-width:600px;margin:0 auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,0.05)">
<div style="background:linear-gradient(135deg,#059669,#0d9488);color:white;padding:24px">
<h1 style="margin:0;font-size:22px">Hợp đồng từ Song Châu</h1>
<p style="margin:6px 0 0;opacity:0.9;font-size:13px">{c['contract_no']}</p></div>
<div style="padding:24px"><p style="margin:0 0 16px;color:#475569">Kính gửi <strong>{c['vendor_name']}</strong>,</p>
<p style="margin:0 0 16px;color:#475569;line-height:1.6">Song Châu xin gửi hợp đồng để Quý đơn vị xem xét và ký xác nhận:</p>
<table style="width:100%;border-collapse:collapse;margin:16px 0;font-size:13px">
<tr><td style="padding:8px;border-bottom:1px solid #e2e8f0;color:#64748b">Mã hợp đồng</td><td style="padding:8px;border-bottom:1px solid #e2e8f0;font-weight:600">{c['contract_no']}</td></tr>
<tr><td style="padding:8px;border-bottom:1px solid #e2e8f0;color:#64748b">Tổng giá trị</td><td style="padding:8px;border-bottom:1px solid #e2e8f0;font-weight:600">{float(c['total_amount']):,.0f} {c['currency']}</td></tr>
<tr><td style="padding:8px;border-bottom:1px solid #e2e8f0;color:#64748b">Thanh toán</td><td style="padding:8px;border-bottom:1px solid #e2e8f0">{c['payment_terms']}</td></tr>
<tr><td style="padding:8px;border-bottom:1px solid #e2e8f0;color:#64748b">Giao hàng</td><td style="padding:8px;border-bottom:1px solid #e2e8f0">{c['delivery_terms']}</td></tr>
</table>
<div style="text-align:center;margin:24px 0"><a href="{link}" style="display:inline-block;background:linear-gradient(135deg,#059669,#0d9488);color:white;padding:14px 36px;border-radius:8px;text-decoration:none;font-weight:600">Xem & ký hợp đồng</a></div></div>
<div style="background:#f1f5f9;padding:14px 24px;color:#64748b;font-size:11px;text-align:center">Song Châu ERP · Email tự động</div></div></body></html>"""


# ═══════════════════════════════════════════════════════════════════
# Phase 2 — POs (MRO Purchase Orders)
# ═══════════════════════════════════════════════════════════════════

@router.get("/pos")
async def list_pos(
    status: str | None = Query(None),
    contract_id: int | None = Query(None),
    vendor_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách MRO Purchase Orders."""
    where = "1=1"; params: list = []; idx = 1
    if status and status != "all":
        where += f" AND p.status = ${idx}"; params.append(status); idx += 1
    if contract_id:
        where += f" AND p.contract_id = ${idx}"; params.append(contract_id); idx += 1
    if vendor_id:
        where += f" AND p.vendor_id = ${idx}"; params.append(vendor_id); idx += 1

    total = await conn.fetchval(f"SELECT COUNT(*) FROM procurement_pos p WHERE {where}", *params)
    params_paged = params + [limit, (page - 1) * limit]
    rows = await conn.fetch(
        f"""
        SELECT p.*, c.contract_no,
               (SELECT COUNT(*) FROM procurement_po_items WHERE po_id = p.id) AS item_count,
               (SELECT COUNT(*) FROM procurement_deliveries WHERE po_id = p.id) AS delivery_count
        FROM procurement_pos p
        LEFT JOIN procurement_contracts c ON c.id = p.contract_id
        WHERE {where}
        ORDER BY p.po_date DESC, p.created_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params_paged,
    )
    return {"data": [dict(r) for r in rows], "total": total}


@router.get("/pos/{po_id}")
async def get_po(
    po_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Chi tiết PO + items + deliveries."""
    row = await conn.fetchrow(
        """SELECT p.*, c.contract_no FROM procurement_pos p
           LEFT JOIN procurement_contracts c ON c.id = p.contract_id WHERE p.id = $1""",
        po_id,
    )
    if not row:
        raise HTTPException(404)
    items = await conn.fetch(
        "SELECT * FROM procurement_po_items WHERE po_id = $1 ORDER BY item_no", po_id,
    )
    deliveries = await conn.fetch(
        "SELECT * FROM procurement_deliveries WHERE po_id = $1 ORDER BY created_at DESC", po_id,
    )
    return {"data": {**dict(row), "items": [dict(i) for i in items], "deliveries": [dict(d) for d in deliveries]}}


@router.get("/pos/{po_id}/pdf")
async def download_po_pdf(
    po_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 8 #9 — Tải PDF ĐƠN ĐẶT HÀNG (Gotenberg, Song Châu letterhead). On-demand render."""
    p = await conn.fetchrow("SELECT po_no FROM procurement_pos WHERE id = $1", po_id)
    if not p:
        raise HTTPException(404, "Đơn đặt hàng không tồn tại")
    try:
        rel = await procurement_docs.generate_po_pdf(conn, po_id)
    except ValueError:
        raise HTTPException(404, "Đơn đặt hàng không tồn tại")
    resolved = _resolve_under_files_base(rel)
    return FileResponse(
        str(resolved),
        media_type="application/pdf",
        filename=f"DonDatHang_{p['po_no']}.pdf",
    )


@router.post("/contracts/{contract_id}/create-po")
async def create_po_from_contract(
    contract_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tạo PO từ contract — copy items hoặc subset.

    Body: {
        "requested_delivery_date": date?,
        "delivery_address": str?,
        "notes": str?,
        "item_ids": [int]? (nếu None → lấy hết items của contract),
        "qty_overrides": {item_id: qty}?
    }
    """
    c = await conn.fetchrow(
        "SELECT * FROM procurement_contracts WHERE id = $1 AND status = 'active'",
        contract_id,
    )
    if not c:
        raise HTTPException(404, "Contract không tồn tại hoặc chưa active")

    # Đợt 3 [MC] — CỔNG TÀI CHÍNH (defense-in-depth): contract đang ở batch có
    # award 'proposed' chưa duyệt → cấm tạo PO (→ công nợ). Thực tế contract
    # không thể tạo khi proposed, nhưng chặn trực tiếp ở đây cho chắc.
    if c["batch_id"] is not None:
        b_award = await conn.fetchval(
            "SELECT award_status FROM procurement_rfq_batches WHERE id = $1",
            c["batch_id"],
        )
        if b_award == "proposed":
            raise HTTPException(
                409,
                "Đề xuất chốt thầu chưa được duyệt — không thể tạo PO khi đang chờ "
                "duyệt (cổng tài chính).",
            )

    selected_ids = body.get("item_ids")
    qty_overrides = body.get("qty_overrides") or {}

    where_items = "contract_id = $1"
    params: list = [contract_id]
    if selected_ids:
        where_items += " AND id = ANY($2::bigint[])"
        params.append(selected_ids)

    items = await conn.fetch(
        f"SELECT * FROM procurement_contract_items WHERE {where_items} ORDER BY item_no", *params,
    )
    if not items:
        raise HTTPException(400, "Không có item nào")

    total = 0.0
    for it in items:
        qt = float(qty_overrides.get(str(it["id"]), it["quantity"]))
        total += qt * float(it["unit_price"] or 0)

    async with conn.transaction():
        seq = await conn.fetchval("SELECT nextval('procurement_po_seq')")
        po_no = f"SC-PO-{datetime.now().year}-{seq:04d}"

        po_id = await conn.fetchval(
            """
            INSERT INTO procurement_pos (
                po_no, contract_id, batch_id, vendor_id, vendor_name,
                po_date, requested_delivery_date, total_amount, currency,
                delivery_address, notes, status, created_by
            ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,'open',$12)
            RETURNING id
            """,
            po_no, contract_id, c["batch_id"], c["vendor_id"], c["vendor_name"],
            datetime.now().date(), _as_date(body.get("requested_delivery_date")),
            total, c["currency"],
            body.get("delivery_address") or "Kho Song Châu — 123 Đường ABC, Q.7, TP.HCM",
            body.get("notes"), token_data.user_id,
        )

        for it in items:
            qt = float(qty_overrides.get(str(it["id"]), it["quantity"]))
            await conn.execute(
                """
                INSERT INTO procurement_po_items (
                    po_id, contract_item_id, item_no, bqms_code,
                    specification, ordered_qty, unit, unit_price, notes
                ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                """,
                po_id, int(it["id"]), it["item_no"], it["bqms_code"],
                it["specification"], qt, it["unit"], float(it["unit_price"] or 0), it["notes"],
            )

        await _audit(
            conn, "po", po_id, "create",
            actor_id=token_data.user_id, to_status="open",
            detail={
                "po_no": po_no, "contract_id": contract_id,
                "vendor_id": c["vendor_id"], "item_count": len(items),
                "total_amount": total, "currency": c["currency"],
            },
        )
        await dispatch_procurement_event(
            conn, "po", po_id, "create",
            actor_id=token_data.user_id, awarded_vendor_id=c["vendor_id"],
            detail={
                "batch_id": c["batch_id"], "po_id": po_id, "po_no": po_no,
                "total_amount": total, "currency": c["currency"],
            },
        )

    logger.info("PO %s created from contract %d (total %s %s)", po_no, contract_id, total, c["currency"])
    return {"data": {"id": po_id, "po_no": po_no}, "message": f"Đã tạo {po_no}"}


@router.put("/pos/{po_id}/status")
async def update_po_status(
    po_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Update PO status + payment_status."""
    new_status = body.get("status")
    payment = body.get("payment_status")
    if new_status and new_status not in ("draft", "open", "partially_delivered", "delivered", "closed", "cancelled"):
        raise HTTPException(400)
    if payment and payment not in ("pending", "partial", "paid"):
        raise HTTPException(400)
    po = await conn.fetchrow("SELECT status, payment_status FROM procurement_pos WHERE id = $1", po_id)
    if not po:
        raise HTTPException(404)
    updates: list = []; params: list = []; idx = 1
    if new_status:
        updates.append(f"status = ${idx}"); params.append(new_status); idx += 1
        if new_status == "closed":
            updates.append("closed_at = NOW()")
    if payment:
        updates.append(f"payment_status = ${idx}"); params.append(payment); idx += 1
    if not updates:
        raise HTTPException(400, "Không có gì update")
    updates.append("updated_at = NOW()")
    params.append(po_id)
    async with conn.transaction():
        await conn.execute(f"UPDATE procurement_pos SET {', '.join(updates)} WHERE id = ${idx}", *params)
        if new_status and new_status != po["status"]:
            await _audit(
                conn, "po", po_id, "status_change",
                actor_id=token_data.user_id,
                from_status=po["status"], to_status=new_status,
                detail={"payment_status": payment} if payment else None,
            )
    return {"message": "Đã update"}


@router.post("/pos/{po_id}/cancel")
async def cancel_po(
    po_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Huỷ toàn bộ PO (admin từ chối giao hàng = huỷ PO).

    Body: {"reason": str (bắt buộc)}.
    Chỉ huỷ được khi PO đang ở draft/open/partially_delivered;
    PO đã delivered/closed/cancelled → 400.
    Status → 'cancelled', lý do append vào notes, closed_at = NOW().
    """
    reason = (body.get("reason") or "").strip()
    if not reason:
        raise HTTPException(400, "Cần lý do huỷ (reason)")

    po = await conn.fetchrow(
        "SELECT status, batch_id, vendor_id, po_no FROM procurement_pos WHERE id = $1", po_id
    )
    if not po:
        raise HTTPException(404)
    if po["status"] not in ("draft", "open", "partially_delivered"):
        raise HTTPException(400, f"PO ở status='{po['status']}' — không thể huỷ")

    async with conn.transaction():
        await conn.execute(
            """UPDATE procurement_pos
               SET status = 'cancelled',
                   notes = COALESCE(notes, '') || E'\n[Huỷ] ' || $1,
                   closed_at = NOW(),
                   updated_at = NOW()
               WHERE id = $2""",
            reason, po_id,
        )
        await _audit(
            conn, "po", po_id, "cancel",
            actor_id=token_data.user_id,
            from_status=po["status"], to_status="cancelled",
            detail={"reason": reason},
        )
        await dispatch_procurement_event(
            conn, "po", po_id, "cancel",
            actor_id=token_data.user_id, awarded_vendor_id=po["vendor_id"],
            detail={"batch_id": po["batch_id"], "po_no": po["po_no"], "reason": reason},
        )
    return {"message": "Đã huỷ PO"}


# ═══════════════════════════════════════════════════════════════════
# Phase 2 — Deliveries (Giao hàng)
# ═══════════════════════════════════════════════════════════════════

@router.get("/deliveries")
async def list_deliveries(
    status: str | None = Query(None),
    po_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách giao hàng."""
    where = "1=1"; params: list = []; idx = 1
    if status and status != "all":
        where += f" AND d.status = ${idx}"; params.append(status); idx += 1
    if po_id:
        where += f" AND d.po_id = ${idx}"; params.append(po_id); idx += 1

    total = await conn.fetchval(f"SELECT COUNT(*) FROM procurement_deliveries d WHERE {where}", *params)
    params_paged = params + [limit, (page - 1) * limit]
    rows = await conn.fetch(
        f"""
        SELECT d.*, p.po_no, p.vendor_name,
               (SELECT COUNT(*) FROM procurement_delivery_items WHERE delivery_id = d.id) AS item_count,
               (SELECT SUM(delivered_qty) FROM procurement_delivery_items WHERE delivery_id = d.id) AS total_qty
        FROM procurement_deliveries d
        LEFT JOIN procurement_pos p ON p.id = d.po_id
        WHERE {where}
        ORDER BY d.created_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params_paged,
    )
    return {"data": [dict(r) for r in rows], "total": total}


@router.get("/deliveries/{delivery_id}")
async def get_delivery(
    delivery_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Chi tiết giao hàng + items đã giao."""
    row = await conn.fetchrow(
        """SELECT d.*, p.po_no, p.vendor_name FROM procurement_deliveries d
           LEFT JOIN procurement_pos p ON p.id = d.po_id WHERE d.id = $1""",
        delivery_id,
    )
    if not row:
        raise HTTPException(404)
    items = await conn.fetch(
        """SELECT di.*, pi.bqms_code, pi.specification, pi.ordered_qty, pi.unit
           FROM procurement_delivery_items di
           JOIN procurement_po_items pi ON pi.id = di.po_item_id
           WHERE di.delivery_id = $1
           ORDER BY pi.item_no""",
        delivery_id,
    )
    return {"data": {**dict(row), "items": [dict(i) for i in items]}}


@router.post("/deliveries/{delivery_id}/generate-note")
async def generate_delivery_note_endpoint(
    delivery_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 8 #2 — Render PHIẾU GIAO NHẬN PDF (Gotenberg, Song Châu letterhead).

    Ghi delivery_note_path. KHÔNG đổi status. Idempotent (re-render đè cùng tên).
    """
    d = await conn.fetchrow(
        "SELECT id, delivery_no FROM procurement_deliveries WHERE id = $1", delivery_id
    )
    if not d:
        raise HTTPException(404, "Lô giao không tồn tại")
    try:
        rel_path = await procurement_docs.generate_delivery_note_pdf(conn, delivery_id)
    except Exception as exc:
        logger.exception("generate_delivery_note_pdf failed for #%s: %s", delivery_id, exc)
        raise HTTPException(502, f"Tạo PDF lỗi: {exc}")

    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_deliveries SET delivery_note_path = $1, updated_at = NOW() WHERE id = $2",
            rel_path, delivery_id,
        )
        await _audit(
            conn, "delivery", delivery_id, "generate_note",
            actor_id=token_data.user_id,
            detail={"delivery_no": d["delivery_no"], "path": rel_path},
        )
    return {"data": {"delivery_note_path": rel_path}, "message": "Đã tạo Phiếu Giao Nhận"}


@router.get("/deliveries/{delivery_id}/note")
async def download_delivery_note(
    delivery_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "procurement", "warehouse")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 8 #2 — Tải PHIẾU GIAO NHẬN PDF đã render. Auto-render nếu chưa có."""
    d = await conn.fetchrow(
        "SELECT delivery_no, delivery_note_path FROM procurement_deliveries WHERE id = $1",
        delivery_id,
    )
    if not d:
        raise HTTPException(404, "Lô giao không tồn tại")
    rel = d["delivery_note_path"]
    if not rel:
        # Lazy-generate on first download so the admin never hits "chưa có PDF".
        # In a txn: file-write + path-store commit together (no orphan regenerate).
        try:
            async with conn.transaction():
                rel = await procurement_docs.generate_delivery_note_pdf(conn, delivery_id)
                await conn.execute(
                    "UPDATE procurement_deliveries SET delivery_note_path = $1, updated_at = NOW() WHERE id = $2",
                    rel, delivery_id,
                )
        except ValueError:
            raise HTTPException(404, "Lô giao không tồn tại")
    resolved = _resolve_under_files_base(rel)
    return FileResponse(
        str(resolved),
        media_type="application/pdf",
        filename=f"PhieuGiaoNhan_{d['delivery_no']}.pdf",
    )


@router.get("/deliveries/{delivery_id}/documents/{idx}")
async def download_delivery_doc_admin(
    delivery_id: int,
    idx: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "procurement", "warehouse")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 8 #6 — Admin tải chứng từ CO/CQ NCC đã upload cho lô giao."""
    d = await conn.fetchrow(
        "SELECT documents FROM procurement_deliveries WHERE id = $1", delivery_id
    )
    if not d:
        raise HTTPException(404, "Lô giao không tồn tại")
    docs = d["documents"]
    if isinstance(docs, str):
        docs = _json.loads(docs or "[]")
    if not docs or idx < 0 or idx >= len(docs):
        raise HTTPException(404, "Chứng từ không tồn tại")
    entry = docs[idx]
    resolved = _resolve_under_files_base(entry.get("path"))
    # attachment: chặn browser render inline file NCC upload (stored-XSS guard).
    return FileResponse(
        str(resolved),
        filename=entry.get("name") or f"document_{idx}",
        content_disposition_type="attachment",
    )


@router.patch("/deliveries/{delivery_id}/items/{item_id}/quality")
async def set_delivery_item_quality(
    delivery_id: int,
    item_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES, "warehouse")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 8 #5 — Buyer chấm kết quả CHẤT LƯỢNG từng mã trong lô giao.

    Endpoint riêng (không nhồi vào PUT status) để cập nhật quality_status/notes của
    1 dòng. Trực tiếp unblock yếu tố 'chất lượng' (0.15) của vendor scorecard —
    trước đây mọi dòng mặc định 'ok' nên điểm luôn hoàn hảo.
    """
    qs = body.get("quality_status")
    if qs not in ("ok", "minor_defect", "rejected"):
        raise HTTPException(400, "quality_status phải là ok / minor_defect / rejected")
    # Chỉ chấm chất lượng khi hàng ĐÃ ĐẾN/ĐÃ NHẬN — tránh chấm 'rejected' trên lô
    # còn đang vận chuyển (sẽ làm hỏng yếu tố chất lượng scorecard trước khi kiểm).
    dlv = await conn.fetchrow(
        "SELECT status FROM procurement_deliveries WHERE id = $1", delivery_id
    )
    if not dlv:
        raise HTTPException(404, "Lô giao không tồn tại")
    if dlv["status"] not in ("arrived", "received"):
        raise HTTPException(400, "Chỉ chấm chất lượng khi lô đã đến hoặc đã nhận hàng")
    row = await conn.fetchrow(
        "SELECT id FROM procurement_delivery_items WHERE id = $1 AND delivery_id = $2",
        item_id, delivery_id,
    )
    if not row:
        raise HTTPException(404, "Dòng giao hàng không tồn tại")
    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_delivery_items SET quality_status = $1, "
            "notes = COALESCE($2, notes) WHERE id = $3",
            qs, body.get("notes"), item_id,
        )
        await _audit(
            conn, "delivery", delivery_id, "quality_change",
            actor_id=token_data.user_id,
            detail={"item_id": item_id, "quality_status": qs, "notes": body.get("notes")},
        )
    return {"message": "Đã cập nhật chất lượng"}


@router.patch("/deliveries/{delivery_id}/items/{item_id}/confirm-qty")
async def set_delivery_item_confirmed_qty(
    delivery_id: int,
    item_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES, "warehouse")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đợt 10 #4 — BUYER xác nhận số THỰC NHẬN từng mã trong lô giao.

    confirmed_qty là số NỘI BỘ (NCC KHÔNG thấy/sửa). delivered_qty vẫn là số NCC
    khai — endpoint này KHÔNG đụng vào nó. confirmed_by ép = token_data.user_id
    (KHÔNG nhận từ body → không spoof). Validate: số ≥ 0, ≤ delivered_qty
    (DECISION D2=A — không owe NCC nhiều hơn số họ giao). PO %-progress admin sau
    đó tính từ COALESCE(confirmed_qty, delivered_qty) (DECISION D1=B) — lô cũ chưa
    xác nhận fallback delivered_qty nên zero-breaking.

    Mirror set_delivery_item_quality (cùng guard status arrived/received).
    """
    # (1) Parse + validate confirmed_qty: bắt buộc, số, ≥ 0.
    raw = body.get("confirmed_qty")
    if raw is None:
        raise HTTPException(400, "Thiếu confirmed_qty")
    try:
        cq = float(raw)
    except (TypeError, ValueError):
        raise HTTPException(400, "confirmed_qty phải là số")
    if cq < 0:
        raise HTTPException(400, "confirmed_qty phải ≥ 0")
    # (2) Guard status: chỉ xác nhận khi lô ĐÃ ĐẾN / ĐÃ NHẬN (mirror quality).
    dlv = await conn.fetchrow(
        "SELECT status, po_id FROM procurement_deliveries WHERE id = $1", delivery_id
    )
    if not dlv:
        raise HTTPException(404, "Lô giao không tồn tại")
    if dlv["status"] not in ("arrived", "received"):
        raise HTTPException(400, "Chỉ xác nhận số nhận khi lô đã đến hoặc đã nhận hàng")
    # (3) Fetch dòng giao (scope WHERE id AND delivery_id → chống cross-delivery write).
    row = await conn.fetchrow(
        "SELECT id, delivered_qty FROM procurement_delivery_items "
        "WHERE id = $1 AND delivery_id = $2",
        item_id, delivery_id,
    )
    if not row:
        raise HTTPException(404, "Dòng giao hàng không tồn tại")
    # (4) VALIDATE (D2=A): confirmed_qty ≤ delivered_qty — không xác nhận nhận
    #     nhiều hơn số NCC khai giao.
    delivered = float(row["delivered_qty"] or 0)
    if cq > delivered:
        raise HTTPException(
            400, f"Số nhận ({cq}) không được vượt số giao ({delivered})"
        )
    async with conn.transaction():
        await conn.execute(
            "UPDATE procurement_delivery_items "
            "SET confirmed_qty = $1, confirmed_by = $2, confirmed_at = NOW() "
            "WHERE id = $3",
            cq, token_data.user_id, item_id,
        )
        await _audit(
            conn, "delivery", delivery_id, "confirm_qty",
            actor_id=token_data.user_id,
            detail={
                "item_id": item_id,
                "confirmed_qty": cq,
                "delivered_qty": delivered,
            },
        )
        # (5) [DECISION D1=B] Recompute PO status từ COALESCE(confirmed_qty,
        #     delivered_qty) gộp qua các lô NON-rejected của PO. Xác nhận số nhận
        #     thấp hơn có thể HẠ 'delivered' → 'partially_delivered'. COALESCE đảm
        #     bảo lô chưa xác nhận vẫn dùng delivered_qty (zero-breaking lô cũ).
        po_id = dlv["po_id"]
        if po_id is not None:
            po_row = await conn.fetchrow(
                "SELECT status FROM procurement_pos WHERE id = $1", po_id
            )
            totals = await conn.fetchrow(
                """
                SELECT
                    (SELECT SUM(ordered_qty) FROM procurement_po_items
                      WHERE po_id = $1) AS ordered,
                    (SELECT COALESCE(SUM(COALESCE(di.confirmed_qty, di.delivered_qty)), 0)
                       FROM procurement_delivery_items di
                       JOIN procurement_deliveries d ON d.id = di.delivery_id
                      WHERE d.po_id = $1 AND d.status NOT IN ('rejected', 'returned')) AS delivered
                """,
                po_id,
            )
            if po_row and totals and totals["ordered"]:
                ratio = float(totals["delivered"] or 0) / float(totals["ordered"])
                new_status = (
                    "delivered" if ratio >= 1
                    else ("partially_delivered" if ratio > 0 else "open")
                )
                # Chỉ chuyển trong nhóm giao-hàng (open/partial/delivered) — KHÔNG
                # đè closed/cancelled (status machine ngoài luồng giao).
                if (
                    new_status != po_row["status"]
                    and po_row["status"] in ("open", "partially_delivered", "delivered")
                ):
                    await conn.execute(
                        "UPDATE procurement_pos SET status = $1, updated_at = NOW(), "
                        "actual_delivery_date = CASE WHEN $1 = 'delivered' "
                        "THEN COALESCE(actual_delivery_date, NOW()::date) "
                        "ELSE actual_delivery_date END WHERE id = $2",
                        new_status, po_id,
                    )
                    await _audit(
                        conn, "po", po_id, "status_change",
                        actor_id=token_data.user_id,
                        from_status=po_row["status"], to_status=new_status,
                        detail={"trigger": "confirm_qty", "delivery_id": delivery_id},
                    )
    return {"message": "Đã xác nhận số nhận"}


@router.post("/pos/{po_id}/deliveries")
async def create_delivery(
    po_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Admin tạo giao hàng (thường NCC sẽ tự tạo qua public endpoint).

    Body: {
        "delivery_method": str?,
        "tracking_no": str?,
        "items": [{po_item_id: int, delivered_qty: float, quality_status: str?, notes: str?}],
        "notes": str?
    }
    """
    po = await conn.fetchrow(
        "SELECT id, vendor_id, status, batch_id FROM procurement_pos WHERE id = $1", po_id
    )
    if not po:
        raise HTTPException(404)
    if po["status"] not in ("open", "partially_delivered"):
        raise HTTPException(400, f"PO ở status='{po['status']}' — không tạo delivery được")

    items_in = body.get("items") or []
    if not items_in:
        raise HTTPException(400, "Cần ít nhất 1 item giao")

    # Đợt 8 #3 — trường đóng gói/invoice (GĐ1): để kho đối chiếu + lên Phiếu Giao Nhận.
    inv_date = _coerce_date(body.get("invoice_date"))

    async with conn.transaction():
        seq = await conn.fetchval("SELECT nextval('procurement_delivery_seq')")
        delivery_no = f"SC-DEL-{datetime.now().year}-{seq:04d}"

        delivery_id = await conn.fetchval(
            """
            INSERT INTO procurement_deliveries (
                delivery_no, po_id, vendor_id, delivery_method, tracking_no,
                status, notes, created_by,
                vendor_invoice_no, invoice_date, packing_qty, packing_unit, gross_weight
            ) VALUES ($1,$2,$3,$4,$5,'pending',$6,$7,$8,$9,$10,$11,$12)
            RETURNING id
            """,
            delivery_no, po_id, po["vendor_id"],
            body.get("delivery_method") or "vendor_delivery",
            body.get("tracking_no"),
            body.get("notes"),
            token_data.user_id,
            (body.get("vendor_invoice_no") or None),
            inv_date,
            _coerce_num(body.get("packing_qty")),
            (body.get("packing_unit") or None),
            _coerce_num(body.get("gross_weight")),
        )

        for it in items_in:
            await conn.execute(
                """
                INSERT INTO procurement_delivery_items (
                    delivery_id, po_item_id, delivered_qty, quality_status, notes
                ) VALUES ($1,$2,$3,$4,$5)
                """,
                delivery_id, int(it["po_item_id"]),
                float(it["delivered_qty"]),
                it.get("quality_status") or "ok",
                it.get("notes"),
            )
            # Increment delivered_qty on PO item
            await conn.execute(
                "UPDATE procurement_po_items SET delivered_qty = delivered_qty + $1 WHERE id = $2",
                float(it["delivered_qty"]), int(it["po_item_id"]),
            )

        await _audit(
            conn, "delivery", delivery_id, "create",
            actor_id=token_data.user_id, to_status="pending",
            detail={
                "delivery_no": delivery_no, "po_id": po_id,
                "vendor_id": po["vendor_id"], "item_count": len(items_in),
            },
        )
        await dispatch_procurement_event(
            conn, "delivery", delivery_id, "create",
            actor_id=token_data.user_id, awarded_vendor_id=po["vendor_id"],
            detail={"batch_id": po["batch_id"], "delivery_no": delivery_no, "po_id": po_id},
        )

        # Auto-update PO status based on received totals.
        # [DECISION D1=B] numerator = COALESCE(confirmed_qty, delivered_qty) gộp
        # qua các lô NON-rejected của PO (số THỰC NHẬN nếu buyer đã xác nhận, else
        # số NCC khai). Tại thời điểm tạo lô confirmed_qty luôn NULL ⇒ COALESCE
        # rơi về delivered_qty ⇒ zero-breaking so với trước.
        totals = await conn.fetchrow(
            """
            SELECT
                (SELECT SUM(ordered_qty) FROM procurement_po_items
                  WHERE po_id = $1) AS ordered,
                (SELECT COALESCE(SUM(COALESCE(di.confirmed_qty, di.delivered_qty)), 0)
                   FROM procurement_delivery_items di
                   JOIN procurement_deliveries d ON d.id = di.delivery_id
                  WHERE d.po_id = $1 AND d.status NOT IN ('rejected', 'returned')) AS delivered
            """,
            po_id,
        )
        if totals and totals["delivered"] is not None:
            ratio = float(totals["delivered"]) / float(totals["ordered"]) if totals["ordered"] else 0
            new_status = "delivered" if ratio >= 1 else ("partially_delivered" if ratio > 0 else "open")
            if new_status != po["status"]:
                # On reaching 'delivered', stamp actual_delivery_date (first time only).
                await conn.execute(
                    "UPDATE procurement_pos SET status = $1, updated_at = NOW(), "
                    "actual_delivery_date = CASE WHEN $1 = 'delivered' "
                    "THEN COALESCE(actual_delivery_date, NOW()::date) ELSE actual_delivery_date END "
                    "WHERE id = $2",
                    new_status, po_id,
                )
                await _audit(
                    conn, "po", po_id, "status_change",
                    actor_id=token_data.user_id,
                    from_status=po["status"], to_status=new_status,
                    detail={"trigger": "delivery_recorded", "delivery_no": delivery_no},
                )

    return {"data": {"id": delivery_id, "delivery_no": delivery_no}, "message": f"Đã tạo {delivery_no}"}


@router.put("/deliveries/{delivery_id}/status")
async def update_delivery_status(
    delivery_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES, "warehouse")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Update delivery status (pending → shipping → arrived → received / rejected).

    Roles aligned with the rest of the PO lifecycle (_WRITE_ROLES = admin/manager/
    procurement) plus warehouse, so the same staff who create/cancel can progress it.
    """
    new_status = body.get("status")
    if new_status not in ("pending", "shipping", "arrived", "received", "rejected", "returned"):
        raise HTTPException(400)
    dlv = await conn.fetchrow(
        """SELECT d.status, d.vendor_id, d.delivery_no, p.batch_id
           FROM procurement_deliveries d
           LEFT JOIN procurement_pos p ON p.id = d.po_id
           WHERE d.id = $1""",
        delivery_id,
    )
    if not dlv:
        raise HTTPException(404)
    extra = []
    params: list = [new_status]
    idx = 2
    if new_status == "received":
        extra.append(f"received_at = NOW(), received_by = ${idx}")
        params.append(token_data.user_id); idx += 1
    if new_status == "rejected" and body.get("rejection_reason"):
        extra.append(f"rejection_reason = ${idx}")
        params.append(body["rejection_reason"]); idx += 1
    params.append(delivery_id)
    extra_sql = (", " + ", ".join(extra)) if extra else ""
    async with conn.transaction():
        await conn.execute(
            f"UPDATE procurement_deliveries SET status = $1{extra_sql}, updated_at = NOW() WHERE id = ${idx}",
            *params,
        )
        if new_status != dlv["status"]:
            await _audit(
                conn, "delivery", delivery_id, "status_change",
                actor_id=token_data.user_id,
                from_status=dlv["status"], to_status=new_status,
                detail={"rejection_reason": body.get("rejection_reason")} if new_status == "rejected" else None,
            )
            # Only the terminal received/rejected transitions notify; skip the
            # noisy intermediate shipping/arrived/pending status changes.
            if new_status in ("received", "rejected"):
                await dispatch_procurement_event(
                    conn, "delivery", delivery_id, new_status,
                    actor_id=token_data.user_id, awarded_vendor_id=dlv["vendor_id"],
                    detail={
                        "batch_id": dlv["batch_id"], "delivery_no": dlv["delivery_no"],
                        "rejection_reason": body.get("rejection_reason"),
                    },
                )

        # ================================================================
        # === PROCUREMENT AUTO-AP (công nợ phải trả) — BEHAVIOR-CHANGE  ===
        # === requires owner sign-off before enabling.                  ===
        # ================================================================
        # When PROCUREMENT_AUTO_AP_ENABLED is TRUE, a delivery transitioning
        # to status='received' auto-creates ONE accounts_payable row for THAT
        # delivery (amount = value of that delivery only), idempotent via a
        # UNIQUE index on accounts_payable.delivery_id.
        #
        # GATED default-OFF (app.core.config.settings.PROCUREMENT_AUTO_AP_ENABLED,
        # env var PROCUREMENT_AUTO_AP_ENABLED). With the flag FALSE this entire
        # block is a NO-OP — no AP is created — so deploying the code does NOT
        # change any financial behavior. Every line below only runs once Thang
        # flips the flag on.  ⚠️ OWNER REVIEW REQUIRED.
        # Only on a REAL transition into received (dlv['status'] is the prior
        # status) — a re-PUT of received on an already-received delivery is a
        # no-op for AP (also guarded by the uq_ap_procurement_delivery idempotency).
        if new_status == "received" and dlv["status"] != "received":
            # Resolve the flag: app_config 'procurement_auto_ap_enabled' is an
            # INSTANT runtime kill-switch (no redeploy) that overrides the
            # env/settings default. Read inside a SAVEPOINT so a config-read
            # error can NEVER poison/abort the delivery-receipt txn.
            _ap_on = settings.PROCUREMENT_AUTO_AP_ENABLED
            try:
                async with conn.transaction():  # savepoint isolates the read
                    _ov = await conn.fetchval(
                        "SELECT value::text FROM app_config WHERE key = 'procurement_auto_ap_enabled'"
                    )
                if _ov is not None:
                    _ap_on = _ov.strip().strip('"').lower() in ("true", "1", "yes")
            except Exception:  # noqa: BLE001 — never block receipt on a config read
                pass

            if _ap_on:  # ⚠️ BEHAVIOR-CHANGE GUARD (env default OR app_config)
                from app.services import chain_service

                # BEST-EFFORT: the auto-AP hook must NEVER abort or roll back
                # the delivery-receipt response. The AP write runs in a NESTED
                # `async with conn.transaction()` (savepoint) wrapped in
                # try/except. If anything fails we log a WARNING and let the
                # OUTER transaction (status UPDATE + audit + notification)
                # commit normally.
                #
                # The savepoint is load-bearing: a bare try/except that only
                # logs is NOT enough. The outer `async with conn.transaction()`
                # opened above means any in-statement error marks the WHOLE
                # transaction aborted in asyncpg — every subsequent statement
                # would then raise InFailedSQLTransactionError and still break
                # the receipt. The nested transaction rolls back ONLY the
                # auto-AP write and leaves the outer transaction clean.
                try:
                    async with conn.transaction():  # savepoint — auto-AP only
                        ap_id = await chain_service.ensure_ap_for_procurement_delivery(
                            conn, delivery_id, created_by=token_data.user_id,
                        )
                    logger.info(
                        "Procurement auto-AP for delivery %s -> ap_id=%s",
                        delivery_id, ap_id,
                    )
                except Exception as exc:  # noqa: BLE001
                    # W3-06 — NEVER swallow: the auto-AP SAVEPOINT above has
                    # already rolled back (so the delivery-receipt still commits),
                    # but the công-nợ-phải-trả row is now MISSING. Log at ERROR
                    # *and* bell-notify every admin so a human creates the AP by
                    # hand instead of the gap dying silently in the log. The notify
                    # runs in its OWN savepoint (inside chain_service) so it can't
                    # poison the still-open outer receipt transaction.
                    logger.error(
                        "Procurement auto-AP hook failed for delivery %s: %s",
                        delivery_id, exc, exc_info=True,
                    )
                    await chain_service.notify_admins_hook_failure(
                        conn,
                        hook="auto-AP",
                        ref_type="procurement_delivery",
                        ref_id=delivery_id,
                        error=f"delivery {delivery_id}: {exc}",
                        link="/finance/reconcile",
                    )
        # ============ END PROCUREMENT AUTO-AP BEHAVIOR-CHANGE BLOCK ====
    return {"message": "Đã update"}


# ---------------------------------------------------------------------------
# Đợt 2a #12 — Q&A (hỏi đáp / làm rõ RFQ) + Addendum (phụ lục broadcast)
# ---------------------------------------------------------------------------
#
# Admin cockpit cho thread Q&A. Bảng procurement_rfq_messages (_018):
#   * question → NCC hỏi (vendor side: app/api/vendor/messages.py).
#   * answer   → admin trả lời RIÊNG 1 NCC (gửi tới đúng NCC đó).
#   * addendum → admin BROADCAST phụ lục tới TẤT CẢ NCC đã mời (ẩn danh người hỏi).
#
# Notif TÁI DÙNG procurement_quote (KHÔNG ALTER TYPE) — deep-link
# /vendor-bidding/{batch_id} (admin) + rfq/[id] (cổng NCC qua metadata.batch_id).
# Mỗi POST: _audit (append-only) + dispatch (best-effort). KHÔNG lộ giá/tên đối thủ.


def _serialize_admin_message(row: asyncpg.Record) -> dict[str, Any]:
    """Row procurement_rfq_messages → dict cho admin FE.

    `author` = 'vendor' | 'admin'. KHÔNG trả author_admin_id (UUID) ra FE — admin
    chỉ cần biết bên nào viết, không cần định danh người dùng.
    """
    atts = row["attachments"]
    if isinstance(atts, str):
        try:
            atts = _json.loads(atts)
        except (TypeError, ValueError):
            atts = []
    return {
        "id": row["id"],
        "kind": row["kind"],
        "author": "vendor" if row["is_vendor_author"] else "admin",
        "body": row["body"],
        "attachments": atts or [],
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
    }


@router.get("/batches/{batch_id}/message-threads")
async def list_message_threads(
    batch_id: int,
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Cột trái panel Q&A: danh sách NCC đã mời + đếm câu hỏi CHƯA ĐỌC của từng NCC.

    `unread_count_admin` = số câu hỏi (kind='question') admin chưa đọc trong thread
    NCC đó → chấm "cần trả lời". `last_at` = mốc tin nhắn mới nhất (sort).
    """
    rows = await conn.fetch(
        """
        SELECT inv.vendor_id, va.company_name,
               COUNT(m.id) FILTER (
                   WHERE m.kind = 'question' AND m.read_by_admin_at IS NULL
               ) AS unread_count_admin,
               MAX(m.created_at) AS last_at
          FROM procurement_rfq_invitations inv
          JOIN vendor_accounts va ON va.id = inv.vendor_id
          LEFT JOIN procurement_rfq_messages m
                 ON m.batch_id = inv.batch_id AND m.vendor_id = inv.vendor_id
         WHERE inv.batch_id = $1
         GROUP BY inv.vendor_id, va.company_name
         ORDER BY unread_count_admin DESC, last_at DESC NULLS LAST
        """,
        batch_id,
    )
    return {
        "threads": [
            {
                "vendor_id": r["vendor_id"],
                "company_name": r["company_name"],
                "unread_count_admin": int(r["unread_count_admin"] or 0),
                "last_at": r["last_at"].isoformat() if r["last_at"] else None,
            }
            for r in rows
        ]
    }


@router.get("/batches/{batch_id}/messages")
async def admin_list_messages(
    batch_id: int,
    vendor_id: int = Query(..., description="Xem thread của 1 NCC (bắt buộc)"),
    token_data: TokenData = Depends(require_role(*_READ_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Thread Q&A của 1 NCC (question + answer) + addendum (để admin thấy ngữ cảnh).

    `vendor_id` bắt buộc (KISS — tránh dump toàn bộ). Side-effect: đánh dấu các câu
    hỏi của NCC đó là admin đã đọc (clear chấm unread).
    """
    rows = await conn.fetch(
        """
        SELECT id, kind, body, attachments, created_at,
               (author_vendor_id IS NOT NULL) AS is_vendor_author
          FROM procurement_rfq_messages
         WHERE batch_id = $1
           AND ( (vendor_id = $2 AND kind IN ('question','answer'))
                 OR kind = 'addendum' )
         ORDER BY created_at ASC
        """,
        batch_id, vendor_id,
    )
    await conn.execute(
        """
        UPDATE procurement_rfq_messages SET read_by_admin_at = NOW()
         WHERE batch_id = $1 AND vendor_id = $2
           AND kind = 'question' AND read_by_admin_at IS NULL
        """,
        batch_id, vendor_id,
    )
    return {"messages": [_serialize_admin_message(r) for r in rows]}


@router.post("/batches/{batch_id}/messages")
async def admin_answer_message(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Admin TRẢ LỜI RIÊNG 1 NCC (kind='answer'). Notif chỉ tới NCC đích.

    Body: { "vendor_id": int, "body": str, "attachments"?: [str] }
    Verify NCC thuộc batch (đã mời) — 400 nếu chưa mời. Notif `awarded_vendor_id`
    + `internal=False` → đúng 1 row tới NCC, KHÔNG báo lại team.
    """
    vendor_id = body.get("vendor_id")
    if vendor_id is None:
        raise HTTPException(400, "Thiếu vendor_id")
    vendor_id = int(vendor_id)

    text = str(body.get("body") or "").strip()
    if not text:
        raise HTTPException(400, "Nội dung trả lời không được để trống")
    if len(text) > 4000:
        raise HTTPException(400, "Nội dung quá dài (tối đa 4000 ký tự)")

    batch = await conn.fetchrow(
        "SELECT batch_code FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    invited = await conn.fetchrow(
        "SELECT 1 FROM procurement_rfq_invitations "
        "WHERE batch_id = $1 AND vendor_id = $2 LIMIT 1",
        batch_id, vendor_id,
    )
    if not invited:
        raise HTTPException(400, "NCC chưa được mời đợt này")

    # Admin attachment text-only M1: chấp nhận list path do admin cung cấp (admin
    # tin cậy); không có hệ admin-upload nên thường rỗng. KHÔNG sandbox (YAGNI M1).
    atts = body.get("attachments") or []
    if not isinstance(atts, list):
        atts = []

    async with conn.transaction():
        msg_id = await conn.fetchval(
            """
            INSERT INTO procurement_rfq_messages
                (batch_id, vendor_id, kind, author_admin_id, body, attachments)
            VALUES ($1, $2, 'answer', $3, $4, $5::jsonb)
            RETURNING id
            """,
            batch_id, vendor_id, token_data.user_id, text, _json.dumps(atts),
        )
        await _audit(
            conn, "rfq_message", msg_id, "answer",
            actor_id=token_data.user_id,
            detail={"batch_id": batch_id, "vendor_id": vendor_id},
        )
        await dispatch_procurement_event(
            conn, "rfq_message", msg_id, "answer",
            actor_id=token_data.user_id, awarded_vendor_id=vendor_id,
            internal=False,  # CHỈ NCC đích nhận chuông, KHÔNG báo lại team
            detail={"batch_id": batch_id, "batch_code": batch["batch_code"]},
        )

    logger.info("[RFQ_QA] dispatched answer batch=%s vendor=%s msg=%s",
                batch_id, vendor_id, msg_id)
    return {"id": msg_id, "ok": True}


@router.post("/batches/{batch_id}/addendum")
async def admin_post_addendum(
    batch_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role(*_WRITE_ROLES)),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Admin ĐĂNG PHỤ LỤC (kind='addendum') BROADCAST tới TẤT CẢ NCC đã mời.

    Body: { "body": str, "attachments"?: [str] }
    Shape A: 1 row vendor_id NULL (DRY) + loop notif từng NCC mời (mẫu invite_vendors).
    Lần dispatch đầu internal=True (team biết đã đăng); các lần sau internal=False
    (chỉ thêm row vendor, không nhân bản team). Ẩn danh người hỏi → công bằng.
    """
    text = str(body.get("body") or "").strip()
    if not text:
        raise HTTPException(400, "Nội dung phụ lục không được để trống")
    if len(text) > 4000:
        raise HTTPException(400, "Nội dung quá dài (tối đa 4000 ký tự)")

    batch = await conn.fetchrow(
        "SELECT batch_code FROM procurement_rfq_batches WHERE id = $1", batch_id
    )
    if not batch:
        raise HTTPException(404, "Đợt báo giá không tồn tại")

    atts = body.get("attachments") or []
    if not isinstance(atts, list):
        atts = []

    vids = [
        r["vendor_id"]
        for r in await conn.fetch(
            "SELECT DISTINCT vendor_id FROM procurement_rfq_invitations WHERE batch_id = $1",
            batch_id,
        )
    ]

    async with conn.transaction():
        msg_id = await conn.fetchval(
            """
            INSERT INTO procurement_rfq_messages
                (batch_id, vendor_id, kind, author_admin_id, body, attachments)
            VALUES ($1, NULL, 'addendum', $2, $3, $4::jsonb)
            RETURNING id
            """,
            batch_id, token_data.user_id, text, _json.dumps(atts),
        )
        await _audit(
            conn, "rfq_message", msg_id, "addendum",
            actor_id=token_data.user_id,
            detail={"batch_id": batch_id, "vendor_count": len(vids)},
        )
        for i, vid in enumerate(vids):
            await dispatch_procurement_event(
                conn, "rfq_message", msg_id, "addendum",
                actor_id=token_data.user_id, awarded_vendor_id=vid,
                internal=(i == 0),  # team biết đúng 1 lần; mỗi NCC 1 row
                detail={"batch_id": batch_id, "batch_code": batch["batch_code"]},
            )

    logger.info("[RFQ_QA] addendum broadcast=%s batch=%s msg=%s",
                len(vids), batch_id, msg_id)
    return {"id": msg_id, "broadcast_to": len(vids)}
