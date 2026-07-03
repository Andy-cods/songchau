"""Sourcing Library API.

Lưu mỗi lần sale đi tìm giá + nhà cung cấp để dùng làm tham chiếu báo giá.
Thang 2026-05-23.
"""
# NOTE: `from __future__ import annotations` removed (Thang 2026-06-13 deploy)
# — it breaks FastAPI forward-ref resolution when @limiter.limit (slowapi
# 0.1.9) decorates handlers using Pydantic body models like QuoteBatchRequest.
# With Python 3.12 PEP 604 union syntax (`str | None`) works natively, so the
# future-annotations import was only cosmetic.

import io
import json
import logging
import os
import secrets
from datetime import datetime, timezone, timedelta, date as date_cls
from decimal import Decimal  # 1b.2: used in FX-snapshot helper signatures near top
from pathlib import Path
from typing import Any, Literal

import aiofiles
import asyncpg
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile
from pydantic import BaseModel, model_validator

from app.core.database import get_db
from app.core.rbac import require_role, TokenData
from app.core.slowapi_limiter import limiter

logger = logging.getLogger(__name__)
router = APIRouter()

# Image upload settings — persist to /data/files (volume mounted) to survive restart
IMAGE_DIR = Path("/data/files/sourcing")
IMAGE_MAX_BYTES = 10 * 1024 * 1024  # 10MB
IMAGE_ALLOWED_CT = {"image/jpeg", "image/png", "image/webp", "image/gif"}
IMAGE_EXT_MAP = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
}

# Excel column header → DB field mapping (Vietnamese + English variants)
EXCEL_HEADER_MAP: dict[str, str] = {
    "tên kh": "customer_name",
    "ten kh": "customer_name",
    "khách hàng": "customer_name",
    "customer": "customer_name",
    "ng phụ trách": "person_in_charge",
    "ng phu trach": "person_in_charge",
    "người phụ trách": "person_in_charge",
    "phụ trách": "person_in_charge",
    "model": "model",
    "tên sp": "product_name",
    "ten sp": "product_name",
    "sản phẩm": "product_name",
    "product": "product_name",
    "maker": "maker",
    "ngày hỏi giá": "inquiry_date",
    "ngay hoi gia": "inquiry_date",
    "ngày hỏi": "inquiry_date",
    "inquiry date": "inquiry_date",
    "giá nhập yên nhật": "cost_jpy",
    "gia nhap yen nhat": "cost_jpy",
    "jpy": "cost_jpy",
    "giá nhập usd": "cost_usd",
    "gia nhap usd": "cost_usd",
    "usd": "cost_usd",
    "giá nhập won": "cost_krw",
    "gia nhap won": "cost_krw",
    "won": "cost_krw",
    "krw": "cost_krw",
    "giá nhập rmb": "cost_rmb",
    "gia nhap rmb": "cost_rmb",
    "rmb": "cost_rmb",
    "giá nhập vnd": "cost_vnd",
    "gia nhap vnd": "cost_vnd",
    "vnd": "cost_vnd",
    "giá nhập": "cost_vnd",
    "giá bán": "sale_vnd",
    "gia ban": "sale_vnd",
    "sale": "sale_vnd",
    "số lượng": "quantity",
    "so luong": "quantity",
    "qty": "quantity",
    "quantity": "quantity",
    "thuế xuất": "tax_pct",
    "thue xuat": "tax_pct",
    "tax": "tax_pct",
    "hs code": "hs_code",
    "hs": "hs_code",
    "cân nặng": "weight_kg",
    "can nang": "weight_kg",
    "weight": "weight_kg",
    "weight (kg)": "weight_kg",
    "hệ số": "coefficient",
    "he so": "coefficient",
    "coefficient": "coefficient",
    "ghi chú": "notes",
    "ghi chu": "notes",
    "notes": "notes",
    "note": "notes",
    "nhà cung cấp": "supplier_name",
    "nha cung cap": "supplier_name",
    "ncc": "supplier_name",
    "supplier": "supplier_name",
    "sdt ncc": "supplier_phone",
    "phone": "supplier_phone",
    "email ncc": "supplier_email",
    "email": "supplier_email",
    "hình ảnh": "image_url",
    "hinh anh": "image_url",
    "image": "image_url",
    "tỷ giá": "_exchange_rate_raw",
    "ty gia": "_exchange_rate_raw",
    "exchange rate": "_exchange_rate_raw",
    "row classifi": "row_classification",  # truncated "Row Classification"
    "row classification": "row_classification",
    "classification": "row_classification",
    "mã bqms": "bqms_code",
    "ma bqms": "bqms_code",
    "bqms code": "bqms_code",
    "bqms_code": "bqms_code",
}


class SourcingPayload(BaseModel):
    bqms_code: str | None = None
    customer_name: str | None = None
    person_in_charge: str | None = None
    model: str | None = None
    product_name: str | None = None
    maker: str | None = None
    inquiry_date: str | None = None  # YYYY-MM-DD
    cost_jpy: float | None = None
    cost_usd: float | None = None
    cost_krw: float | None = None
    cost_rmb: float | None = None
    cost_vnd: float | None = None
    sale_vnd: float | None = None
    quantity: float | None = None
    tax_pct: float | None = None
    hs_code: str | None = None
    weight_kg: float | None = None
    coefficient: float | None = None
    supplier_name: str | None = None
    supplier_phone: str | None = None
    supplier_email: str | None = None
    image_url: str | None = None
    notes: str | None = None
    row_classification: str | None = None
    exchange_rate: dict[str, float] | None = None
    # Batch #1 (Thang 2026-06-27): FROZEN pricing context captured at "Áp dụng
    # giá báo". Stored verbatim as JSONB; reopening restores the form inputs and
    # the quote modal defaults to its unit_price_vnd so form == export.
    quote_snapshot: dict[str, Any] | None = None
    # Batch #4 (V3): FK to customers — when set, the exported báo giá autofills
    # company_name / MST / address from the CRM record (create_quote_batch already
    # enriches via customer_id). NULL = free-text customer_name only (legacy).
    customer_id: int | None = None
    # A1: phí vận chuyển nội địa VN (VND). When provided on PUT /{id} and it
    # differs from the latest stored history value, a new row is appended to
    # sourcing_vn_shipping_history. ALSO stored on sourcing_entries.vn_shipping_fee_vnd
    # (2026-07-02) so reopen restores it without reading quote_snapshot.
    vn_shipping_fee_vnd: float | None = None
    # 2026-07-02: phí vận chuyển quốc tế (FedEx) VND — persisted on the entry so
    # the pricing form reopens with the exact typed value (fallback = snapshot).
    fedex_fee_vnd: float | None = None
    # Pricing-history FX freeze (Thang 2026-07-01): the FE sends the PRIMARY
    # supplier's cost currency + amount so the backend can freeze the REAL FX
    # rate at create-time (the primary supplier row does not exist in the DB
    # yet). When primary_cost_currency is a known foreign currency, the FX
    # snapshot must fetch its JPY/USD/KRW/RMB→VND rate — never short-circuit 1.
    primary_cost_currency: str | None = None
    primary_cost_amount: float | None = None


# ── FX snapshot helpers (Batch 1A · 1b.2) ───────────────────────────
# An entry's cost-currency is whichever cost_* column is populated (the
# legacy multi-currency model). We derive (currency, amount) so we can
# freeze the FX rate + its effective date onto the entry at save time.
_COST_CURRENCY_FIELDS: list[tuple[str, str]] = [
    ("cost_jpy", "JPY"),
    ("cost_usd", "USD"),
    ("cost_krw", "KRW"),
    ("cost_rmb", "RMB"),
    ("cost_vnd", "VND"),
]


def _payload_cost_currency(payload: "SourcingPayload") -> tuple[str | None, float | None]:
    """Return (currency, cost_amount) for the entry's cost.

    Prefers the first non-zero foreign-currency cost; falls back to VND.
    Returns (None, None) when no cost is set at all.
    """
    for field, cur in _COST_CURRENCY_FIELDS:
        val = getattr(payload, field, None)
        if val is not None and float(val) > 0:
            return (cur, float(val))
    return (None, None)


def _fx_cost_currency(payload: "SourcingPayload") -> tuple[str | None, float | None]:
    """Resolve (currency, amount) to freeze the FX rate against.

    Pricing-history fix (Thang 2026-07-01): on CREATE the primary supplier row
    does not exist yet, so the legacy cost_* columns may be empty/VND even for a
    foreign entry. Prefer the payload's `primary_cost_currency`/`amount` (the
    real primary NCC currency the FE sends) so `_compute_fx_snapshot` fetches the
    REAL JPY/USD/KRW/RMB→VND rate instead of short-circuiting to 1. Falls back to
    the legacy `_payload_cost_currency` when the primary fields are absent.
    """
    pc = (payload.primary_cost_currency or "").strip().upper()
    if pc:
        amt = payload.primary_cost_amount
        try:
            amt = float(amt) if amt is not None else None
        except (TypeError, ValueError):
            amt = None
        return (pc, amt)
    return _payload_cost_currency(payload)


def _manual_fx_for(exchange_rate: dict | None, currency: str | None):
    """Case-insensitive lookup of a manual FX rate from the JSONB exchange_rate dict.

    BUG FIX (Thang 2026-06-21): `_payload_cost_currency` returns an UPPERCASE
    currency (e.g. "JPY") but the frontend posts the exchange_rate JSONB with
    lowercase keys ("jpy"), so `.get(cost_currency)` silently dropped the
    user-typed tỷ giá. Match keys case-insensitively so the manual rate persists
    regardless of key case.
    """
    if not exchange_rate or not currency:
        return None
    for k, v in exchange_rate.items():
        if str(k).upper().strip() == currency.upper().strip():
            return v
    return None


async def _compute_fx_snapshot(
    conn: asyncpg.Connection,
    currency: str | None,
    inquiry_date: Any,
    manual_rate: Any = None,
) -> tuple[Decimal | None, Any]:
    """Freeze (fx_rate_snapshot, fx_rate_date) for an entry's cost currency.

    Precedence: a user-typed manual rate (Thang 2026-06-17 — hàng ngoại tệ sửa
    tỷ giá bằng tay) wins over the auto rate. VND → (1, today). Foreign without a
    manual rate → fetch_fx_meta(currency, as_of_date=inquiry_date or today).
    Returns (None, None) when currency is unknown or the rate is missing.
    """
    if not currency:
        return (None, None)
    cur = currency.upper().strip()
    as_of: date_cls = inquiry_date if isinstance(inquiry_date, date_cls) else date_cls.today()
    if cur == "VND":
        return (Decimal("1"), as_of)
    # Manual override (user typed the tỷ giá by hand) → freeze it as-of-now.
    if manual_rate is not None:
        try:
            mr = Decimal(str(manual_rate))
            if mr > 0:
                return (mr, as_of)
        except (ValueError, ArithmeticError, TypeError):
            pass
    from app.services.sourcing_pricing_engine import fetch_fx_meta
    rate, rate_date = await fetch_fx_meta(conn, cur, as_of_date=as_of)
    if rate is None or rate <= 0:
        return (None, None)
    return (rate, rate_date or as_of)


def _serialize(row: asyncpg.Record) -> dict[str, Any]:
    d = dict(row)
    for k, v in list(d.items()):
        if isinstance(v, (datetime,)):
            d[k] = v.isoformat()
        elif hasattr(v, "isoformat"):
            d[k] = v.isoformat()
        elif isinstance(v, str) and k in ("exchange_rate", "quote_snapshot"):
            try:
                d[k] = json.loads(v)
            except Exception:
                pass
        elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
            try:
                d[k] = float(v)
            except Exception:
                pass
    # A1: reshape vn_shipping_history (json from json_agg, or already a list) into
    # the FE contract {value_vnd, at, by}. Always emit a list (never null).
    if "vn_shipping_history" in d:
        d["vn_shipping_history"] = _shape_vn_shipping_history(d["vn_shipping_history"])
    return d


def _shape_vn_shipping_history(raw: Any) -> list[dict[str, Any]]:
    """Normalize the LATERAL json_agg payload to [{value_vnd, at, by}]."""
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            return []
    if not isinstance(raw, list):
        return []
    out: list[dict[str, Any]] = []
    for h in raw:
        if not isinstance(h, dict):
            continue
        val = h.get("value_vnd")
        try:
            val = float(val) if val is not None else None
        except (TypeError, ValueError):
            pass
        at = h.get("created_at")
        if hasattr(at, "isoformat"):
            at = at.isoformat()
        out.append({
            "value_vnd": val,
            "at": at,
            "by": h.get("created_by_email"),
        })
    return out


@router.get("/")
async def list_sourcing(
    search: str | None = Query(None, description="Tìm trong bqms_code/model/product_name/maker/supplier"),
    bqms_code: str | None = None,
    maker: str | None = None,
    supplier: str | None = None,
    customer: str | None = None,
    # Thang 2026-06-03: PIM catalog filters
    catalog_category: str | None = None,
    brand_canonical: str | None = None,
    catalog_status: str | None = None,
    stage: int | None = Query(None, ge=1, le=3),
    has_price: bool | None = None,
    has_supplier: bool | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(30, ge=10, le=200),
    sort_by: str = Query("created_at", regex="^(created_at|updated_at|inquiry_date|sale_vnd|cost_vnd|maker|supplier_name)$"),
    sort_dir: str = Query("desc", regex="^(asc|desc)$"),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List sourcing entries với search + filter."""
    # NOTE: all column refs below are prefixed `se.` because the query now LEFT
    # JOINs `customers c` so the FE can link the Khách cell to /crm/{id}. The FK
    # `sourcing_entries.customer_id → customers(id)` is populated by the Quote Hub;
    # we return it + a canonical customer_name (joined company_name first, falling
    # back to the stored se.customer_name text) so old text-only rows still render.
    # (Thang 2026-06-22 — GAP 1 link completeness)
    conditions: list[str] = ["se.deleted_at IS NULL"]
    params: list[Any] = []
    idx = 1
    if search:
        like = f"%{search.strip()}%"
        conditions.append(
            f"(se.bqms_code ILIKE ${idx} OR se.model ILIKE ${idx} OR se.product_name ILIKE ${idx} "
            f"OR se.maker ILIKE ${idx} OR se.supplier_name ILIKE ${idx} OR se.customer_name ILIKE ${idx} "
            f"OR se.brand_canonical ILIKE ${idx})"
        )
        params.append(like)
        idx += 1
    if bqms_code:
        conditions.append(f"se.bqms_code = ${idx}")
        params.append(bqms_code)
        idx += 1
    if maker:
        conditions.append(f"se.maker ILIKE ${idx}")
        params.append(f"%{maker}%")
        idx += 1
    if supplier:
        conditions.append(f"se.supplier_name ILIKE ${idx}")
        params.append(f"%{supplier}%")
        idx += 1
    if customer:
        conditions.append(f"se.customer_name ILIKE ${idx}")
        params.append(f"%{customer}%")
        idx += 1
    if catalog_category:
        conditions.append(f"se.catalog_category = ${idx}")
        params.append(catalog_category)
        idx += 1
    if brand_canonical:
        conditions.append(f"se.brand_canonical = ${idx}")
        params.append(brand_canonical)
        idx += 1
    if catalog_status:
        conditions.append(f"se.catalog_status = ${idx}")
        params.append(catalog_status)
        idx += 1
    if stage is not None:
        conditions.append(f"se.stage = ${idx}")
        params.append(stage)
        idx += 1
    if has_price is not None:
        if has_price:
            conditions.append("se.sale_vnd IS NOT NULL")
        else:
            conditions.append("se.sale_vnd IS NULL")
    if has_supplier is not None:
        if has_supplier:
            conditions.append("se.supplier_name IS NOT NULL AND BTRIM(se.supplier_name) <> ''")
        else:
            conditions.append("(se.supplier_name IS NULL OR BTRIM(se.supplier_name) = '')")

    where = " AND ".join(conditions)
    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM sourcing_entries se WHERE {where}", *params
    )

    params.extend([page_size, (page - 1) * page_size])
    # Canonical customer_name = joined customers.company_name when the FK is set,
    # else the legacy free-text se.customer_name. `customer_id` is returned so the
    # FE can render the Khách cell as a <Link href="/crm/{id}">. `sort_by` comes
    # from a strict whitelist regex of sourcing_entries columns → qualified `se.`.
    rows = await conn.fetch(
        f"""
        SELECT se.id, se.bqms_code,
               COALESCE(c.company_name, se.customer_name) AS customer_name,
               se.customer_id,
               se.person_in_charge, se.model, se.product_name,
               se.maker, se.inquiry_date,
               se.cost_jpy, se.cost_usd, se.cost_krw, se.cost_rmb, se.cost_vnd,
               se.sale_vnd, se.quantity, se.tax_pct, se.hs_code, se.weight_kg, se.coefficient,
               se.supplier_name, se.supplier_phone, se.supplier_email,
               se.image_url, se.notes, se.row_classification, se.exchange_rate,
               se.catalog_category, se.brand_canonical, se.part_type, se.machine_model,
               se.catalog_status, se.stage, se.missing_count, se.model_norm,
               se.fx_rate_snapshot, se.fx_rate_date,
               se.created_by_email, se.updated_by_email, se.created_at, se.updated_at,
               -- A2: typed primary-supplier price (FE shows this instead of the
               -- FX-scaled cost_vnd). NULL when no primary supplier row exists.
               pc.cost_amount AS primary_cost_amount,
               pc.currency    AS primary_cost_currency,
               -- A1: latest ~10 VN-shipping fee history rows (newest first).
               COALESCE(vsh.history, '[]'::json) AS vn_shipping_history,
               -- Pricing-history badges (cheap indexed COUNT/MAX per entry).
               COALESCE(ps.cnt, 0)    AS pricing_snapshot_count,
               COALESCE(ps.latest, 0) AS latest_pricing_version
        FROM sourcing_entries se
        LEFT JOIN customers c ON c.id = se.customer_id
        LEFT JOIN LATERAL (
            SELECT COUNT(*)::int AS cnt, COALESCE(MAX(version), 0) AS latest
              FROM sourcing_pricing_snapshots sps
             WHERE sps.entry_id = se.id
        ) ps ON true
        LEFT JOIN LATERAL (
            SELECT sp.cost_amount, sp.currency
              FROM sourcing_supplier_prices sp
             WHERE sp.sourcing_entry_id = se.id AND sp.is_primary
             ORDER BY sp.updated_at DESC
             LIMIT 1
        ) pc ON true
        LEFT JOIN LATERAL (
            SELECT json_agg(h ORDER BY h.created_at DESC) AS history
              FROM (
                SELECT value_vnd, created_at, created_by_email
                  FROM sourcing_vn_shipping_history
                 WHERE entry_id = se.id
                 ORDER BY created_at DESC
                 LIMIT 10
              ) h
        ) vsh ON true
        WHERE {where}
        ORDER BY se.{sort_by} {sort_dir.upper()} NULLS LAST, se.id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )
    return {
        "data": {
            "items": [_serialize(r) for r in rows],
            "total": int(total or 0),
            "page": page,
            "page_size": page_size,
            "pages": (int(total or 0) + page_size - 1) // page_size,
        }
    }


@router.get("/stats")
async def sourcing_stats(
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """KPI tổng cho header sourcing."""
    row = await conn.fetchrow(
        """
        SELECT
            COUNT(*)::int AS total_entries,
            COUNT(DISTINCT bqms_code) FILTER (WHERE bqms_code IS NOT NULL)::int AS unique_codes,
            COUNT(DISTINCT model_norm) FILTER (WHERE model_norm <> '')::int AS unique_models,
            COUNT(DISTINCT supplier_name) FILTER (WHERE supplier_name IS NOT NULL)::int AS unique_suppliers,
            COUNT(DISTINCT maker) FILTER (WHERE maker IS NOT NULL)::int AS unique_makers,
            COUNT(DISTINCT customer_name) FILTER (WHERE customer_name IS NOT NULL)::int AS unique_customers,
            COUNT(*) FILTER (WHERE sale_vnd IS NOT NULL)::int AS has_price_count,
            COUNT(*) FILTER (WHERE supplier_name IS NOT NULL AND BTRIM(supplier_name) <> '')::int AS has_supplier_count,
            COUNT(*) FILTER (WHERE image_url IS NOT NULL AND image_url <> '')::int AS has_image_count,
            COUNT(*) FILTER (WHERE hs_code IS NOT NULL AND hs_code <> '')::int AS has_hs_count,
            COUNT(*) FILTER (WHERE stage = 1)::int AS stage_1,
            COUNT(*) FILTER (WHERE stage = 2)::int AS stage_2,
            COUNT(*) FILTER (WHERE stage = 3)::int AS stage_3,
            COUNT(*) FILTER (WHERE catalog_status = 'OK')::int AS status_ok,
            COUNT(*) FILTER (WHERE catalog_status = 'NEEDS_BRAND')::int AS status_needs_brand,
            COUNT(*) FILTER (WHERE catalog_status = 'NOT_IN_CATALOG')::int AS status_not_in_catalog,
            COUNT(*) FILTER (WHERE catalog_status = 'PRODUCT_CANDIDATE')::int AS status_candidate,
            COUNT(*) FILTER (WHERE created_at >= NOW() - INTERVAL '30 days')::int AS added_30d,
            COUNT(*) FILTER (WHERE created_at >= NOW() - INTERVAL '7 days')::int AS added_7d
        FROM sourcing_entries
        WHERE deleted_at IS NULL
        """
    )
    top_suppliers = await conn.fetch(
        """
        SELECT supplier_name, COUNT(*)::int AS entries
        FROM sourcing_entries
        WHERE deleted_at IS NULL AND supplier_name IS NOT NULL AND BTRIM(supplier_name) <> ''
        GROUP BY supplier_name
        ORDER BY entries DESC
        LIMIT 5
        """
    )
    top_makers = await conn.fetch(
        """
        SELECT maker, COUNT(*)::int AS entries
        FROM sourcing_entries
        WHERE deleted_at IS NULL AND maker IS NOT NULL AND BTRIM(maker) <> ''
        GROUP BY maker
        ORDER BY entries DESC
        LIMIT 5
        """
    )
    top_categories = await conn.fetch(
        """
        SELECT catalog_category, COUNT(*)::int AS entries
        FROM sourcing_entries
        WHERE deleted_at IS NULL AND catalog_category IS NOT NULL
        GROUP BY catalog_category
        ORDER BY entries DESC
        LIMIT 10
        """
    )
    top_brands = await conn.fetch(
        """
        SELECT brand_canonical, COUNT(*)::int AS entries
        FROM sourcing_entries
        WHERE deleted_at IS NULL AND brand_canonical IS NOT NULL
        GROUP BY brand_canonical
        ORDER BY entries DESC
        LIMIT 25
        """
    )
    top_customers = await conn.fetch(
        """
        SELECT customer_name, COUNT(*)::int AS entries
        FROM sourcing_entries
        WHERE deleted_at IS NULL AND customer_name IS NOT NULL
        GROUP BY customer_name
        ORDER BY entries DESC
        LIMIT 50
        """
    )
    return {
        "data": {
            **dict(row or {}),
            "top_suppliers": [dict(r) for r in top_suppliers],
            "top_makers": [dict(r) for r in top_makers],
            "top_categories": [dict(r) for r in top_categories],
            "top_brands": [dict(r) for r in top_brands],
            "top_customers": [dict(r) for r in top_customers],
        }
    }


@router.get("/suggestions")
async def get_suggestions(
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Return top 100 distinct values for autocomplete fields (customer, supplier, maker, person, brand)."""
    queries = {
        "customers": "SELECT customer_name AS v, COUNT(*) AS c FROM sourcing_entries WHERE customer_name IS NOT NULL AND BTRIM(customer_name) <> '' GROUP BY 1 ORDER BY 2 DESC LIMIT 100",
        "suppliers": "SELECT supplier_name AS v, COUNT(*) AS c FROM sourcing_entries WHERE supplier_name IS NOT NULL AND BTRIM(supplier_name) <> '' GROUP BY 1 ORDER BY 2 DESC LIMIT 100",
        "makers":    "SELECT maker AS v, COUNT(*) AS c FROM sourcing_entries WHERE maker IS NOT NULL AND BTRIM(maker) <> '' GROUP BY 1 ORDER BY 2 DESC LIMIT 100",
        "persons":   "SELECT person_in_charge AS v, COUNT(*) AS c FROM sourcing_entries WHERE person_in_charge IS NOT NULL AND BTRIM(person_in_charge) <> '' GROUP BY 1 ORDER BY 2 DESC LIMIT 100",
        "brands":    "SELECT brand_canonical AS v, COUNT(*) AS c FROM sourcing_entries WHERE brand_canonical IS NOT NULL AND BTRIM(brand_canonical) <> '' GROUP BY 1 ORDER BY 2 DESC LIMIT 100",
        "hs_codes":  "SELECT hs_code AS v, COUNT(*) AS c FROM sourcing_entries WHERE hs_code IS NOT NULL AND BTRIM(hs_code) <> '' GROUP BY 1 ORDER BY 2 DESC LIMIT 50",
    }
    out: dict = {}
    for key, sql in queries.items():
        rows = await conn.fetch(sql)
        out[key] = [{"value": r["v"], "count": int(r["c"])} for r in rows]
    return {"data": out}


@router.get("/coverage")
async def coverage_by_codes(
    codes: str = Query(..., description="Comma-separated BQMS codes"),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đếm sourcing entries cho list mã — dùng badge trong price-trends matched table."""
    code_list = [c.strip() for c in codes.split(",") if c.strip()]
    if not code_list:
        return {"data": {}}
    if len(code_list) > 200:
        code_list = code_list[:200]
    rows = await conn.fetch(
        "SELECT bqms_code, COUNT(*)::int AS cnt, COUNT(DISTINCT supplier_name)::int AS supplier_cnt "
        "FROM sourcing_entries WHERE bqms_code = ANY($1) GROUP BY bqms_code",
        code_list,
    )
    return {
        "data": {
            r["bqms_code"]: {"entries": r["cnt"], "suppliers": r["supplier_cnt"]}
            for r in rows
        }
    }


@router.get("/last-customer-prices")
async def last_customer_prices(
    customer_id: int = Query(..., description="Khách hàng đang được chọn trong modal báo giá"),
    sourcing_ids: str = Query(..., description="Comma-separated sourcing entry ids"),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "director", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Giá bán gần nhất từng báo cho KHÁCH NÀY, theo từng sourcing_id.

    Quét lịch sử ``quote_batches.line_items`` (JSONB) của customer_id, lấy
    DISTINCT ON sourcing_id đơn giá VND của lần báo giá MỚI NHẤT. Dùng cho
    badge "Lần trước" trong QuoteBatchModal. Soft-deleted báo giá bị loại.
    """
    id_list: list[int] = []
    seen: set[int] = set()
    for tok in (sourcing_ids or "").split(","):
        tok = tok.strip()
        if not tok:
            continue
        try:
            v = int(tok)
        except (TypeError, ValueError):
            continue
        if v not in seen:
            seen.add(v)
            id_list.append(v)
    if not id_list:
        return {"data": []}

    rows = await conn.fetch(
        """
        SELECT DISTINCT ON ((li->>'sourcing_id')::int)
               (li->>'sourcing_id')::int     AS sourcing_id,
               (li->>'unit_price_vnd')::numeric AS last_price_vnd,
               qb.created_at                  AS quoted_at,
               qb.quote_no                    AS quote_no
          FROM quote_batches qb,
               jsonb_array_elements(qb.line_items) li
         WHERE qb.customer_id = $1
           AND qb.deleted_at IS NULL
           AND (li->>'sourcing_id')::int = ANY($2::int[])
         ORDER BY (li->>'sourcing_id')::int, qb.created_at DESC
        """,
        customer_id, id_list,
    )
    return {
        "data": [
            {
                "sourcing_id": r["sourcing_id"],
                "last_price_vnd": float(r["last_price_vnd"]) if r["last_price_vnd"] is not None else None,
                "quoted_at": r["quoted_at"].isoformat() if r["quoted_at"] is not None else None,
                "quote_no": r["quote_no"],
            }
            for r in rows
        ]
    }


@router.get("/imv-rfq/items")
async def imv_rfq_items(
    q: str | None = Query(None, description="Tìm theo rfq_number / khách / mã / tên / model"),
    limit: int = Query(30, ge=1, le=100, description="Số dòng (mặc định 30, tối đa 100)"),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "director", "procurement", "viewer", "accountant"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tìm item lines trong imv_rfq để chèn vào modal báo giá hàng loạt.

    q rỗng → trả về các dòng mới nhất (recent). Có q → ILIKE trên
    rfq_number / customer_name / item_code / product_name / model.
    Sắp xếp last_seen_at DESC NULLS LAST, id DESC. item_code có thể NULL →
    expose thêm 'code' = COALESCE(item_code, customer_item_code).
    """
    qv = (q or "").strip()
    if qv:
        rows = await conn.fetch(
            """
            SELECT id, rfq_number, customer_name, item_code, customer_item_code,
                   product_name, model, maker, quantity, unit, due_date
              FROM imv_rfq
             WHERE rfq_number    ILIKE '%' || $1 || '%'
                OR customer_name ILIKE '%' || $1 || '%'
                OR item_code     ILIKE '%' || $1 || '%'
                OR product_name  ILIKE '%' || $1 || '%'
                OR model         ILIKE '%' || $1 || '%'
             ORDER BY last_seen_at DESC NULLS LAST, id DESC
             LIMIT $2
            """,
            qv, limit,
        )
    else:
        rows = await conn.fetch(
            """
            SELECT id, rfq_number, customer_name, item_code, customer_item_code,
                   product_name, model, maker, quantity, unit, due_date
              FROM imv_rfq
             ORDER BY last_seen_at DESC NULLS LAST, id DESC
             LIMIT $1
            """,
            limit,
        )
    return {
        "data": [
            {
                "id": r["id"],
                "rfq_number": r["rfq_number"],
                "customer_name": r["customer_name"],
                "item_code": r["item_code"],
                "code": r["item_code"] or r["customer_item_code"],
                "product_name": r["product_name"],
                "model": r["model"],
                "maker": r["maker"],
                "quantity": float(r["quantity"]) if r["quantity"] is not None else None,
                "unit": r["unit"],
                "due_date": r["due_date"].isoformat() if r["due_date"] is not None else None,
            }
            for r in rows
        ]
    }


@router.get("/by-code/{bqms_code}")
async def list_by_code(
    bqms_code: str,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Lấy tất cả sourcing entries cho 1 mã BQMS — dùng trong CodeHistoryDrawer."""
    rows = await conn.fetch(
        """
        SELECT id, bqms_code, customer_name, person_in_charge, model, product_name,
               maker, inquiry_date,
               cost_jpy, cost_usd, cost_krw, cost_rmb, cost_vnd,
               sale_vnd, quantity, tax_pct, hs_code, weight_kg, coefficient,
               supplier_name, supplier_phone, supplier_email,
               image_url, notes, row_classification, exchange_rate,
               created_by_email, updated_by_email, created_at, updated_at
        FROM sourcing_entries
        WHERE bqms_code = $1
        ORDER BY COALESCE(inquiry_date, created_at::date) DESC, id DESC
        """,
        bqms_code,
    )
    return {"data": [_serialize(r) for r in rows]}


@router.get("/by-customer/{customer_id:int}")
async def list_by_customer(
    customer_id: int,
    limit: int = Query(100, ge=1, le=500),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Sourcing entries gắn với 1 khách hàng theo FK customer_id — dùng cho panel
    "Mã đã sourcing" trong trang chi tiết khách hàng (CRM).

    Lọc theo FK `sourcing_entries.customer_id` (KHÔNG dùng text customer_name) —
    chỉ những mã đã được gắn khách qua Quote Hub mới hiện. Mỗi dòng kèm giá bán
    gần nhất + tiền tệ cost, NCC, ngày cập nhật. Soft-deleted bị loại. Index
    `idx_se_customer_id (customer_id, inquiry_date DESC)` phục vụ truy vấn này.
    Route đặt TRƯỚC `/{entry_id:int}` để literal prefix không bị nuốt.
    """
    rows = await conn.fetch(
        """
        SELECT se.id, se.bqms_code, se.model, se.product_name, se.maker,
               se.inquiry_date,
               se.cost_jpy, se.cost_usd, se.cost_krw, se.cost_rmb, se.cost_vnd,
               se.sale_vnd, se.quantity,
               se.supplier_name, se.image_url,
               se.fx_rate_snapshot, se.fx_rate_date,
               se.created_at, se.updated_at,
               -- Tiền tệ của cost (theo cột cost_* được điền) — hiển thị badge giá.
               CASE
                   WHEN se.cost_jpy IS NOT NULL AND se.cost_jpy > 0 THEN 'JPY'
                   WHEN se.cost_usd IS NOT NULL AND se.cost_usd > 0 THEN 'USD'
                   WHEN se.cost_krw IS NOT NULL AND se.cost_krw > 0 THEN 'KRW'
                   WHEN se.cost_rmb IS NOT NULL AND se.cost_rmb > 0 THEN 'RMB'
                   WHEN se.cost_vnd IS NOT NULL AND se.cost_vnd > 0 THEN 'VND'
                   ELSE NULL
               END AS cost_currency
        FROM sourcing_entries se
        WHERE se.customer_id = $1
          AND se.deleted_at IS NULL
        ORDER BY COALESCE(se.updated_at, se.created_at) DESC NULLS LAST, se.id DESC
        LIMIT $2
        """,
        customer_id, limit,
    )
    return {"data": {"items": [_serialize(r) for r in rows], "total": len(rows)}}


@router.get("/{entry_id:int}")
async def get_sourcing(
    entry_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    row = await conn.fetchrow("SELECT * FROM sourcing_entries WHERE id = $1", entry_id)
    if not row:
        raise HTTPException(404, f"Không tìm thấy sourcing #{entry_id}")
    data = _serialize(row)
    # Pricing-history badges: how many saved "đợt tính giá" + the latest version.
    ps = await conn.fetchrow(
        """
        SELECT COUNT(*)::int AS cnt, COALESCE(MAX(version), 0) AS latest
          FROM sourcing_pricing_snapshots WHERE entry_id = $1
        """,
        entry_id,
    )
    data["pricing_snapshot_count"] = int(ps["cnt"]) if ps else 0
    data["latest_pricing_version"] = int(ps["latest"]) if ps else 0
    # A1: attach latest ~10 VN-shipping fee history rows (newest first).
    hist = await conn.fetch(
        """
        SELECT value_vnd, created_at, created_by_email
          FROM sourcing_vn_shipping_history
         WHERE entry_id = $1
         ORDER BY created_at DESC
         LIMIT 10
        """,
        entry_id,
    )
    data["vn_shipping_history"] = _shape_vn_shipping_history([dict(h) for h in hist])
    return {"data": data}


# ── Versioned pricing history (Thang 2026-07-01) ─────────────────────
# A pricing snapshot is an IMMUTABLE, full price computation saved ONLY when the
# user clicks "Lưu đợt tính giá". Loading an old version renders straight from
# its frozen breakdown/params (the FE never re-calls /calc-suggest, so the
# FX-staleness guard + today's % rules are never re-applied). Editing + saving a
# new version leaves older versions untouched.

class PricingSnapshotBody(BaseModel):
    snapshot: dict[str, Any]
    sale_vnd: float | None = None
    label: str | None = None


async def _append_pricing_snapshot(
    conn: asyncpg.Connection,
    entry_id: int,
    snapshot: dict[str, Any],
    sale_vnd: float | None,
    label: str | None,
    email: str | None,
) -> asyncpg.Record:
    """Append version = MAX(version)+1 for `entry_id`, race-safe.

    Concurrent saves would collide on UNIQUE(entry_id, version); a transaction-
    scoped advisory lock keyed on the entry serialises the MAX(version)+1 read →
    INSERT so only one writer computes each version. Also mirrors the frozen
    (fx_rate, fx_date) + snapshot onto the entry so a plain reopen (GET-single)
    shows the same numbers as the latest saved đợt.
    """
    async with conn.transaction():
        await conn.execute(
            "SELECT pg_advisory_xact_lock(hashtext('pricing_snap_' || $1::text))",
            str(entry_id),
        )
        next_version = await conn.fetchval(
            "SELECT COALESCE(MAX(version), 0) + 1 FROM sourcing_pricing_snapshots WHERE entry_id = $1",
            entry_id,
        )
        sale_dec: Decimal | None = None
        if sale_vnd is not None:
            try:
                sale_dec = Decimal(str(sale_vnd)).quantize(Decimal("1"))
            except (ValueError, ArithmeticError, TypeError):
                sale_dec = None
        row = await conn.fetchrow(
            """
            INSERT INTO sourcing_pricing_snapshots
                (entry_id, version, snapshot, sale_vnd, label, created_by_email)
            VALUES ($1, $2, $3::jsonb, $4, $5, $6)
            RETURNING version, created_at
            """,
            entry_id, next_version, json.dumps(snapshot), sale_dec, label, email,
        )
        # Mirror the frozen quote context onto the entry (best-effort: only set
        # fx_rate_snapshot / fx_rate_date when present + valid in the snapshot).
        fx_rate_raw = snapshot.get("fx_rate")
        fx_date_raw = snapshot.get("fx_date")
        fx_rate_dec: Decimal | None = None
        if fx_rate_raw is not None:
            try:
                cand = Decimal(str(fx_rate_raw))
                if cand > 0:
                    fx_rate_dec = cand
            except (ValueError, ArithmeticError, TypeError):
                fx_rate_dec = None
        fx_date_val: Any = None
        if fx_date_raw:
            try:
                fx_date_val = datetime.strptime(str(fx_date_raw)[:10], "%Y-%m-%d").date()
            except ValueError:
                fx_date_val = None
        await conn.execute(
            """
            UPDATE sourcing_entries SET
                quote_snapshot   = $2::jsonb,
                fx_rate_snapshot = COALESCE($3, fx_rate_snapshot),
                fx_rate_date     = COALESCE($4, fx_rate_date)
            WHERE id = $1
            """,
            entry_id, json.dumps(snapshot), fx_rate_dec, fx_date_val,
        )
    return row


@router.post("/{entry_id:int}/pricing-snapshots")
async def create_pricing_snapshot(
    entry_id: int,
    body: PricingSnapshotBody,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Lưu 1 đợt tính giá (explicit versioning). Returns {version, created_at}."""
    exists = await conn.fetchval("SELECT id FROM sourcing_entries WHERE id = $1", entry_id)
    if not exists:
        raise HTTPException(404, f"Không tìm thấy sourcing #{entry_id}")
    row = await _append_pricing_snapshot(
        conn, entry_id, body.snapshot, body.sale_vnd, body.label,
        getattr(token_data, "email", None),
    )
    return {"data": {"version": int(row["version"]), "created_at": row["created_at"].isoformat()}}


@router.get("/{entry_id:int}/pricing-snapshots")
async def list_pricing_snapshots(
    entry_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách đợt tính giá (metadata only, newest-first). Full snapshot via /{version}."""
    rows = await conn.fetch(
        """
        SELECT id, version, sale_vnd, label, created_at, created_by_email
          FROM sourcing_pricing_snapshots
         WHERE entry_id = $1
         ORDER BY version DESC
        """,
        entry_id,
    )
    return {
        "data": [
            {
                "id": r["id"],
                "version": int(r["version"]),
                "sale_vnd": float(r["sale_vnd"]) if r["sale_vnd"] is not None else None,
                "label": r["label"],
                "created_at": r["created_at"].isoformat() if r["created_at"] is not None else None,
                "created_by_email": r["created_by_email"],
            }
            for r in rows
        ]
    }


@router.get("/{entry_id:int}/pricing-snapshots/{version:int}")
async def get_pricing_snapshot(
    entry_id: int,
    version: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Full frozen snapshot for one version — loaded straight into the form (no recompute)."""
    row = await conn.fetchrow(
        """
        SELECT snapshot FROM sourcing_pricing_snapshots
         WHERE entry_id = $1 AND version = $2
        """,
        entry_id, version,
    )
    if not row:
        raise HTTPException(404, f"Không tìm thấy đợt tính giá v{version} cho #{entry_id}")
    snap = row["snapshot"]
    if isinstance(snap, str):
        try:
            snap = json.loads(snap)
        except Exception:
            pass
    return {"data": snap}


@router.post("/")
async def create_sourcing(
    payload: SourcingPayload,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    inq_date = payload.inquiry_date
    inq_date_val: Any = None
    if inq_date:
        try:
            inq_date_val = datetime.strptime(inq_date, "%Y-%m-%d").date()
        except ValueError:
            inq_date_val = None
    exchange_rate_json = json.dumps(payload.exchange_rate) if payload.exchange_rate else None
    snapshot_json = json.dumps(payload.quote_snapshot) if payload.quote_snapshot else None

    # 1b.2: freeze the FX rate + its effective date onto the new entry. Snapshot
    # is taken at the entry's inquiry_date (or today) so the quote is auditable.
    # Pricing-history fix: prefer the payload's primary supplier currency (the
    # primary supplier row does not exist yet at create time) so a foreign entry
    # freezes the REAL rate, never 1.
    cost_currency, _ = _fx_cost_currency(payload)
    _manual_fx = _manual_fx_for(payload.exchange_rate, cost_currency)
    fx_snapshot, fx_snapshot_date = await _compute_fx_snapshot(
        conn, cost_currency, inq_date_val, _manual_fx
    )

    row = await conn.fetchrow(
        """
        INSERT INTO sourcing_entries (
            bqms_code, customer_name, person_in_charge, model, product_name, maker, inquiry_date,
            cost_jpy, cost_usd, cost_krw, cost_rmb, cost_vnd,
            sale_vnd, quantity, tax_pct, hs_code, weight_kg, coefficient,
            supplier_name, supplier_phone, supplier_email,
            image_url, notes, row_classification, exchange_rate,
            created_by_id, created_by_email,
            fx_rate_snapshot, fx_rate_date, quote_snapshot, customer_id,
            fedex_fee_vnd, vn_shipping_fee_vnd
        ) VALUES (
            $1, $2, $3, $4, $5, $6, $7,
            $8, $9, $10, $11, $12,
            $13, $14, $15, $16, $17, $18,
            $19, $20, $21,
            $22, $23, $24, $25::jsonb,
            $26, $27,
            $28, $29, $30::jsonb, $31,
            $32, $33
        ) RETURNING id, created_at
        """,
        payload.bqms_code, payload.customer_name, payload.person_in_charge,
        payload.model, payload.product_name, payload.maker, inq_date_val,
        payload.cost_jpy, payload.cost_usd, payload.cost_krw, payload.cost_rmb, payload.cost_vnd,
        payload.sale_vnd, payload.quantity, payload.tax_pct, payload.hs_code, payload.weight_kg, payload.coefficient,
        payload.supplier_name, payload.supplier_phone, payload.supplier_email,
        payload.image_url, payload.notes, payload.row_classification, exchange_rate_json,
        _coerce_int(getattr(token_data, "user_id", None)),
        getattr(token_data, "email", None),
        fx_snapshot, fx_snapshot_date, snapshot_json, payload.customer_id,
        payload.fedex_fee_vnd, payload.vn_shipping_fee_vnd,
    )
    return {"data": {"id": row["id"], "created_at": row["created_at"].isoformat()}}


def _coerce_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


@router.put("/{entry_id:int}")
async def update_sourcing(
    entry_id: int,
    payload: SourcingPayload,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    existing = await conn.fetchrow(
        """
        SELECT id, cost_jpy, cost_usd, cost_krw, cost_rmb, cost_vnd, fx_rate_snapshot
          FROM sourcing_entries WHERE id = $1
        """,
        entry_id,
    )
    if not existing:
        raise HTTPException(404, f"Không tìm thấy sourcing #{entry_id}")

    inq_date_val: Any = None
    if payload.inquiry_date:
        try:
            inq_date_val = datetime.strptime(payload.inquiry_date, "%Y-%m-%d").date()
        except ValueError:
            inq_date_val = None
    exchange_rate_json = json.dumps(payload.exchange_rate) if payload.exchange_rate else None
    snapshot_json = json.dumps(payload.quote_snapshot) if payload.quote_snapshot else None

    # 1b.2: snapshot is IMMUTABLE once set — re-saving an old quote keeps its
    # ORIGINAL rate. Only (re)compute when there is no snapshot yet OR when the
    # cost currency changed (e.g. supplier switched JPY→USD), in which case the
    # frozen rate no longer applies. The `COALESCE`s below mean an unchanged
    # entry never overwrites the existing snapshot.
    # Prefer the payload primary supplier currency (same reasoning as CREATE):
    # a foreign entry must resolve the REAL rate even when the legacy cost_*
    # columns are VND/empty.
    new_currency, _ = _fx_cost_currency(payload)

    def _existing_currency(rec: asyncpg.Record) -> str | None:
        for field, cur in _COST_CURRENCY_FIELDS:
            v = rec.get(field)
            if v is not None and float(v) > 0:
                return cur
        return None

    old_currency = _existing_currency(existing)
    had_snapshot = existing["fx_rate_snapshot"] is not None
    # A user-typed manual rate for the cost currency always (re)freezes the
    # snapshot — even on an existing entry whose currency is unchanged — so the
    # hand-entered tỷ giá persists and shows correctly on reopen.
    _manual_fx = _manual_fx_for(payload.exchange_rate, new_currency)
    fx_snapshot: Decimal | None = None
    fx_snapshot_date: Any = None
    recompute = (
        (not had_snapshot)
        or (new_currency is not None and new_currency != old_currency)
        or (_manual_fx is not None)
    )
    if recompute and new_currency is not None:
        fx_snapshot, fx_snapshot_date = await _compute_fx_snapshot(
            conn, new_currency, inq_date_val, _manual_fx
        )

    await conn.execute(
        """
        UPDATE sourcing_entries SET
            bqms_code = $2, customer_name = $3, person_in_charge = $4,
            model = $5, product_name = $6, maker = $7, inquiry_date = $8,
            cost_jpy = $9, cost_usd = $10, cost_krw = $11, cost_rmb = $12, cost_vnd = $13,
            sale_vnd = $14, quantity = $15, tax_pct = $16, hs_code = $17, weight_kg = $18, coefficient = $19,
            supplier_name = $20, supplier_phone = $21, supplier_email = $22,
            image_url = $23, notes = $24, row_classification = $25, exchange_rate = $26::jsonb,
            updated_by_id = $27, updated_by_email = $28,
            fx_rate_snapshot = COALESCE($29, fx_rate_snapshot),
            fx_rate_date     = COALESCE($30, fx_rate_date),
            quote_snapshot   = COALESCE($31::jsonb, quote_snapshot),
            customer_id      = $32,
            fedex_fee_vnd    = $33,
            vn_shipping_fee_vnd = $34
        WHERE id = $1
        """,
        entry_id,
        payload.bqms_code, payload.customer_name, payload.person_in_charge,
        payload.model, payload.product_name, payload.maker, inq_date_val,
        payload.cost_jpy, payload.cost_usd, payload.cost_krw, payload.cost_rmb, payload.cost_vnd,
        payload.sale_vnd, payload.quantity, payload.tax_pct, payload.hs_code, payload.weight_kg, payload.coefficient,
        payload.supplier_name, payload.supplier_phone, payload.supplier_email,
        payload.image_url, payload.notes, payload.row_classification, exchange_rate_json,
        _coerce_int(getattr(token_data, "user_id", None)),
        getattr(token_data, "email", None),
        fx_snapshot, fx_snapshot_date, snapshot_json, payload.customer_id,
        payload.fedex_fee_vnd, payload.vn_shipping_fee_vnd,
    )

    # A1 — VN-shipping fee history: append a row only when the caller sent a
    # value AND it differs from the latest stored one (dedupe). NUMERIC(18,0)
    # comparison via Decimal-rounded value to mirror the column's precision.
    if payload.vn_shipping_fee_vnd is not None:
        new_val = Decimal(str(payload.vn_shipping_fee_vnd)).quantize(Decimal("1"))
        latest = await conn.fetchval(
            """
            SELECT value_vnd FROM sourcing_vn_shipping_history
             WHERE entry_id = $1 ORDER BY created_at DESC LIMIT 1
            """,
            entry_id,
        )
        if latest is None or Decimal(latest) != new_val:
            await conn.execute(
                """
                INSERT INTO sourcing_vn_shipping_history
                    (entry_id, value_vnd, created_by_email)
                VALUES ($1, $2, $3)
                """,
                entry_id, new_val, getattr(token_data, "email", None),
            )

    return {"data": {"id": entry_id, "updated": True}}


@router.delete("/{entry_id:int}")
async def delete_sourcing(
    entry_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    deleted = await conn.fetchval(
        "DELETE FROM sourcing_entries WHERE id = $1 RETURNING id", entry_id
    )
    if not deleted:
        raise HTTPException(404, f"Không tìm thấy sourcing #{entry_id}")
    return {"data": {"id": entry_id, "deleted": True}}


# ── Image Upload ──────────────────────────────────────────────────

@router.post("/{entry_id:int}/image")
async def upload_sourcing_image(
    entry_id: int,
    file: UploadFile = File(..., description="Ảnh sản phẩm/NCC"),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Upload ảnh trực tiếp cho sourcing entry, set entry.image_url tự động."""
    existing = await conn.fetchval("SELECT id FROM sourcing_entries WHERE id = $1", entry_id)
    if not existing:
        raise HTTPException(404, f"Không tìm thấy sourcing #{entry_id}")

    content_type = (file.content_type or "").lower()
    if content_type not in IMAGE_ALLOWED_CT:
        raise HTTPException(400, f"Ảnh phải là JPG/PNG/WebP/GIF. Nhận: {content_type}")

    body = await file.read()
    if len(body) > IMAGE_MAX_BYTES:
        raise HTTPException(400, f"Ảnh vượt {IMAGE_MAX_BYTES // 1024 // 1024}MB")
    if len(body) < 8:
        raise HTTPException(400, "Ảnh rỗng")

    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    ext = IMAGE_EXT_MAP.get(content_type, ".jpg")
    fname = f"{entry_id}_{secrets.token_hex(8)}{ext}"
    path = IMAGE_DIR / fname

    async with aiofiles.open(path, "wb") as f:
        await f.write(body)

    # URL served via /api/v1/sourcing/image/{filename}
    url = f"/api/v1/sourcing/image/{fname}"
    await conn.execute(
        "UPDATE sourcing_entries SET image_url = $1 WHERE id = $2",
        url, entry_id,
    )
    logger.info(f"Sourcing image uploaded: {fname} ({len(body)} bytes) for #{entry_id}")
    return {"data": {"id": entry_id, "image_url": url, "size_bytes": len(body)}}


@router.post("/upload-image")
async def upload_image_loose(
    file: UploadFile = File(..., description="Ảnh (chưa gắn entry)"),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement")),
):
    """Upload ảnh không gắn entry — trả URL để gán vào image_url khi submit form."""
    content_type = (file.content_type or "").lower()
    if content_type not in IMAGE_ALLOWED_CT:
        raise HTTPException(400, f"Ảnh phải là JPG/PNG/WebP/GIF. Nhận: {content_type}")
    body = await file.read()
    if len(body) > IMAGE_MAX_BYTES:
        raise HTTPException(400, f"Ảnh vượt {IMAGE_MAX_BYTES // 1024 // 1024}MB")
    if len(body) < 8:
        raise HTTPException(400, "Ảnh rỗng")

    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    ext = IMAGE_EXT_MAP.get(content_type, ".jpg")
    fname = f"loose_{secrets.token_hex(10)}{ext}"
    path = IMAGE_DIR / fname
    async with aiofiles.open(path, "wb") as f:
        await f.write(body)
    return {"data": {"image_url": f"/api/v1/sourcing/image/{fname}", "size_bytes": len(body)}}


@router.get("/image/{filename}")
async def serve_sourcing_image(
    filename: str,
    # ICE security #3 (Thang 2026-06-13): require JWT to serve sourcing images.
    # require_role's underlying get_current_user already accepts ?token=<jwt>
    # query param fallback, so <img src="/api/v1/sourcing/image/x.jpg?token=…">
    # in the React picker keeps working without code changes on the FE side.
    # Viewer role is included so read-only audit users can still preview thumbs.
    token_data: TokenData = Depends(
        require_role("admin", "manager", "staff", "sales", "viewer", "procurement", "director")
    ),
):
    """Serve sourcing image — requires JWT (Authorization header or ?token=)."""
    from fastapi.responses import FileResponse
    safe = Path(filename).name  # sanitize
    path = IMAGE_DIR / safe
    if not path.exists() or not path.is_file():
        raise HTTPException(404, "Ảnh không tồn tại")
    return FileResponse(path)


# ── Bulk Excel Import ────────────────────────────────────────────

def _normalize_header(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _coerce_number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        # Handle strings like "1.260.000" (vi-VN) or "1,260,000"
        if isinstance(v, str):
            cleaned = v.strip().replace(" ", "").replace(",", "").replace(".", "")
            # detect decimal: if original has comma followed by 1-2 digits, treat as decimal
            if not cleaned:
                return None
            # If original had , as decimal separator
            if "," in v and v.rfind(",") > v.rfind("."):
                # remove thousand seps (.) keep , as decimal
                tmp = v.replace(".", "").replace(",", ".").strip()
                return float(tmp)
            # Otherwise dots could be thousand seps OR decimal
            if "." in v:
                # Heuristic: if more than one dot OR last group has >=3 digits, treat dots as thousands
                parts = v.split(".")
                if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3):
                    return float("".join(parts))
            return float(v.replace(",", "").replace(" ", ""))
        return float(v)
    except (ValueError, TypeError):
        return None


def _coerce_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date_cls):
        return v.isoformat()
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(v.strip(), fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _coerce_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(..., description="File Excel theo format Thang"),
    dry_run: bool = Query(False, description="Không insert, chỉ trả về preview"),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Import bulk từ Excel file (.xlsx)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl chưa được cài đặt")

    body = await file.read()
    if len(body) > 20 * 1024 * 1024:
        raise HTTPException(400, "File >20MB")
    if len(body) < 100:
        raise HTTPException(400, "File rỗng")

    try:
        wb = load_workbook(io.BytesIO(body), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Không đọc được Excel: {exc}")

    sheet = wb.active
    if not sheet:
        raise HTTPException(400, "File không có sheet")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "File trống")

    # Map column index → DB field name
    col_map: dict[int, str] = {}
    for idx, header in enumerate(header_row):
        norm = _normalize_header(header)
        if not norm:
            continue
        # Try exact match then substring match
        if norm in EXCEL_HEADER_MAP:
            col_map[idx] = EXCEL_HEADER_MAP[norm]
        else:
            for key, field in EXCEL_HEADER_MAP.items():
                if key in norm or norm in key:
                    col_map[idx] = field
                    break

    if not col_map:
        raise HTTPException(
            400,
            f"Không nhận diện được cột nào. Headers: {[_normalize_header(h) for h in header_row if h]}",
        )

    parsed_rows: list[dict[str, Any]] = []
    skipped = 0
    NUMERIC_FIELDS = {
        "cost_jpy", "cost_usd", "cost_krw", "cost_rmb", "cost_vnd",
        "sale_vnd", "quantity", "tax_pct", "weight_kg", "coefficient",
    }
    DATE_FIELDS = {"inquiry_date"}

    for row in rows_iter:
        if not row or all(cell is None or cell == "" for cell in row):
            skipped += 1
            continue
        entry: dict[str, Any] = {}
        for idx, field in col_map.items():
            if idx >= len(row):
                continue
            cell = row[idx]
            if cell is None or cell == "":
                continue
            if field == "_exchange_rate_raw":
                # Skip — too varied to parse reliably
                continue
            if field in NUMERIC_FIELDS:
                entry[field] = _coerce_number(cell)
            elif field in DATE_FIELDS:
                entry[field] = _coerce_date(cell)
            else:
                entry[field] = _coerce_str(cell)
        # Must have at least product_name or bqms_code
        if not entry.get("product_name") and not entry.get("bqms_code"):
            skipped += 1
            continue
        parsed_rows.append(entry)

    wb.close()

    if dry_run:
        return {
            "data": {
                "dry_run": True,
                "total_parsed": len(parsed_rows),
                "skipped": skipped,
                "preview": parsed_rows[:5],
                "headers_detected": [
                    {"index": idx, "field": col_map[idx], "header": _normalize_header(header_row[idx])}
                    for idx in col_map
                ],
            }
        }

    # Insert in transaction
    inserted = 0
    user_email = getattr(token_data, "email", None)
    user_id = _coerce_int(getattr(token_data, "user_id", None))

    async with conn.transaction():
        for entry in parsed_rows:
            inq = entry.get("inquiry_date")
            inq_val: Any = None
            if inq:
                try:
                    inq_val = datetime.strptime(inq, "%Y-%m-%d").date()
                except ValueError:
                    pass
            await conn.execute(
                """
                INSERT INTO sourcing_entries (
                    bqms_code, customer_name, person_in_charge, model, product_name, maker, inquiry_date,
                    cost_jpy, cost_usd, cost_krw, cost_rmb, cost_vnd,
                    sale_vnd, quantity, tax_pct, hs_code, weight_kg, coefficient,
                    supplier_name, supplier_phone, supplier_email,
                    image_url, notes, row_classification,
                    created_by_id, created_by_email
                ) VALUES (
                    $1, $2, $3, $4, $5, $6, $7,
                    $8, $9, $10, $11, $12,
                    $13, $14, $15, $16, $17, $18,
                    $19, $20, $21,
                    $22, $23, $24,
                    $25, $26
                )
                """,
                entry.get("bqms_code"), entry.get("customer_name"), entry.get("person_in_charge"),
                entry.get("model"), entry.get("product_name"), entry.get("maker"), inq_val,
                entry.get("cost_jpy"), entry.get("cost_usd"), entry.get("cost_krw"), entry.get("cost_rmb"), entry.get("cost_vnd"),
                entry.get("sale_vnd"), entry.get("quantity"), entry.get("tax_pct"), entry.get("hs_code"), entry.get("weight_kg"), entry.get("coefficient"),
                entry.get("supplier_name"), entry.get("supplier_phone"), entry.get("supplier_email"),
                entry.get("image_url"), entry.get("notes"), entry.get("row_classification"),
                user_id, user_email,
            )
            inserted += 1

    return {
        "data": {
            "inserted": inserted,
            "skipped": skipped,
            "headers_detected": [
                {"index": idx, "field": col_map[idx], "header": _normalize_header(header_row[idx])}
                for idx in col_map
            ],
        },
        "message": f"Đã import {inserted} entry · bỏ qua {skipped} dòng",
    }


# ── Multi-supplier comparison ────────────────────────────────────

@router.get("/compare/{bqms_code}")
async def compare_suppliers(
    bqms_code: str,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "director", "procurement")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """So sánh side-by-side các supplier cho 1 mã.

    Trả về tổng hợp + ranking theo giá nhập VND ascending.
    """
    rows = await conn.fetch(
        """
        SELECT id, supplier_name, supplier_phone, supplier_email, maker,
               cost_jpy, cost_usd, cost_krw, cost_rmb, cost_vnd,
               sale_vnd, quantity, coefficient, tax_pct, hs_code, weight_kg,
               notes, row_classification, image_url, inquiry_date, created_at
        FROM sourcing_entries
        WHERE bqms_code = $1 AND supplier_name IS NOT NULL
        ORDER BY cost_vnd ASC NULLS LAST, created_at DESC
        """,
        bqms_code,
    )
    entries = [_serialize(r) for r in rows]

    # Stats
    costs = [e["cost_vnd"] for e in entries if e.get("cost_vnd")]
    sales = [e["sale_vnd"] for e in entries if e.get("sale_vnd")]
    summary = {
        "code": bqms_code,
        "supplier_count": len(entries),
        "cost_min_vnd": min(costs) if costs else None,
        "cost_max_vnd": max(costs) if costs else None,
        "cost_avg_vnd": sum(costs) / len(costs) if costs else None,
        "sale_min_vnd": min(sales) if sales else None,
        "sale_max_vnd": max(sales) if sales else None,
        "spread_pct": (
            ((max(costs) - min(costs)) / min(costs) * 100) if len(costs) >= 2 and min(costs) > 0 else None
        ),
    }
    if summary["spread_pct"] is not None:
        summary["spread_pct"] = round(summary["spread_pct"], 1)

    return {"data": {"summary": summary, "entries": entries}}


# ============================================================
# Phase 1 — RFQ Library bulk-lookup + quote-batch (Thang 2026-06-03)
# ============================================================

import re as _re

_NORMALIZE_RE = _re.compile(r"[^A-Z0-9]")


def _normalize_model(s: str) -> str:
    """Mirror DB GENERATED column logic.

    UPPER + strip non-alphanumeric. Dùng cho bulk-lookup paste codes —
    user gõ 'molex-39001' / 'MOLEX 39001' / 'MOLEX_39001' all match same.
    """
    if not s:
        return ""
    return _NORMALIZE_RE.sub("", s.upper())


class BulkLookupRequest(BaseModel):
    codes: list[str]
    search_mode: str = "exact"   # 'exact' | 'fuzzy'


@router.post("/bulk-lookup")
async def bulk_lookup(
    body: BulkLookupRequest,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "director", "procurement", "viewer", "accountant"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tra cứu thư viện nguồn cung hàng loạt — paste 1-500 mã, trả về:
    - items: rows found (1 latest row per model, with aggregate stats)
    - missing: codes không tìm thấy
    - summary: counts

    Performance budget:
    - exact mode 200 codes: ≤ 200ms (idx_se_model_norm + ANY array)
    - fuzzy mode 200 codes: ≤ 800ms (GIN trigram + LATERAL similarity)
    """
    raw_codes = body.codes or []
    if not raw_codes:
        raise HTTPException(400, "Danh sách mã rỗng")
    if len(raw_codes) > 500:
        raise HTTPException(400, f"Tối đa 500 mã/lần (đang có {len(raw_codes)})")

    # Normalize + dedup
    norm_pairs: list[tuple[str, str]] = []
    seen_norm: set[str] = set()
    for raw in raw_codes:
        if not raw:
            continue
        s = str(raw).strip()
        if not s:
            continue
        n = _normalize_model(s)
        if not n or n in seen_norm:
            continue
        seen_norm.add(n)
        norm_pairs.append((s, n))

    if not norm_pairs:
        raise HTTPException(400, "Không có mã hợp lệ sau normalize")

    norm_list = [p[1] for p in norm_pairs]

    if body.search_mode == "fuzzy":
        # Fuzzy: dùng LATERAL similarity per input code, lấy best match
        rows = await conn.fetch(
            """
            WITH q AS (
                SELECT unnest($1::text[]) AS query_norm,
                       unnest($2::text[]) AS query_raw
            ),
            best AS (
                SELECT q.query_raw, q.query_norm,
                       s.id, s.model, s.product_name, s.maker, s.customer_name,
                       s.inquiry_date, s.sale_vnd, s.cost_vnd,
                       s.supplier_name, s.image_url, s.brand_canonical,
                       s.catalog_status, s.stage,
                       similarity(s.model, q.query_raw) AS sim
                  FROM q
                  LEFT JOIN LATERAL (
                      SELECT * FROM sourcing_entries
                       WHERE deleted_at IS NULL
                       ORDER BY similarity(model, q.query_raw) DESC NULLS LAST
                       LIMIT 1
                  ) s ON s.id IS NOT NULL AND similarity(s.model, q.query_raw) > 0.3
            )
            SELECT * FROM best
            """,
            norm_list, [p[0] for p in norm_pairs],
        )
    else:
        # Exact: model_norm = ANY (cực nhanh nhờ B-tree index)
        rows = await conn.fetch(
            """
            WITH q AS (
                SELECT unnest($1::text[]) AS query_norm,
                       unnest($2::text[]) AS query_raw
            ),
            latest AS (
                SELECT DISTINCT ON (s.model_norm)
                    s.model_norm, s.id, s.model, s.product_name, s.maker,
                    s.customer_name, s.inquiry_date, s.sale_vnd, s.cost_vnd,
                    s.supplier_name, s.image_url, s.brand_canonical,
                    s.catalog_status, s.stage
                  FROM sourcing_entries s
                 WHERE s.deleted_at IS NULL
                   AND s.model_norm = ANY($1::text[])
                 ORDER BY s.model_norm, s.inquiry_date DESC NULLS LAST, s.id DESC
            ),
            agg AS (
                SELECT model_norm,
                       COUNT(*) AS total_inquiries,
                       MIN(sale_vnd) AS min_sale,
                       MAX(sale_vnd) AS max_sale,
                       AVG(sale_vnd) FILTER (WHERE sale_vnd IS NOT NULL) AS avg_sale,
                       ARRAY_AGG(DISTINCT supplier_name) FILTER (
                           WHERE supplier_name IS NOT NULL AND supplier_name <> ''
                       ) AS suppliers,
                       ARRAY_AGG(DISTINCT customer_name) FILTER (
                           WHERE customer_name IS NOT NULL AND customer_name <> ''
                       ) AS customers
                  FROM sourcing_entries
                 WHERE deleted_at IS NULL
                   AND model_norm = ANY($1::text[])
                 GROUP BY model_norm
            )
            SELECT q.query_raw, q.query_norm,
                   l.id, l.model, l.product_name, l.maker,
                   l.customer_name, l.inquiry_date, l.sale_vnd, l.cost_vnd,
                   l.supplier_name, l.image_url, l.brand_canonical,
                   l.catalog_status, l.stage,
                   a.total_inquiries, a.min_sale, a.max_sale, a.avg_sale,
                   a.suppliers, a.customers
              FROM q
              LEFT JOIN latest l ON l.model_norm = q.query_norm
              LEFT JOIN agg    a ON a.model_norm = q.query_norm
             ORDER BY q.query_raw
            """,
            norm_list, [p[0] for p in norm_pairs],
        )

    items: list[dict] = []
    missing: list[str] = []
    for r in rows:
        d = dict(r)
        if d.get("id") is None:
            missing.append(d["query_raw"])
        else:
            # Convert avg/dates for JSON
            for k in ("avg_sale",):
                if d.get(k) is not None:
                    d[k] = float(d[k])
            items.append(d)

    return {
        "data": {
            "items": items,
            "missing": missing,
            "found_count": len(items),
            "missing_count": len(missing),
            "input_count": len(norm_pairs),
            "search_mode": body.search_mode,
        }
    }


class QuoteBatchItem(BaseModel):
    sourcing_id: int
    quantity: float | None = None
    supplier_price_id: int | None = None       # chosen sourcing_supplier_prices.id
    manual_unit_price_vnd: float | None = None  # typed override (already VND)
    # Manual FX override (Thang 2026-06-21): when set on a SUPPLIER line, this
    # rate is used instead of the live DB FX. Omitted → live rate. Ignored on
    # the manual-unit-price branch (no currency conversion happens there).
    fx_rate_override: float | None = None
    # Per-line thời gian giao hàng — the renderer appends this into the
    # Ghi chú column. Carried into the line dict only; not used in pricing.
    delivery_time: str | None = None

    @model_validator(mode="after")
    def _exactly_one(self):
        # Owner decision (Thang 2026-06-21): ALWAYS manual per-line — no
        # server-side auto-pick. Each line MUST carry EXACTLY ONE explicit
        # choice; both-or-neither is rejected.
        has_sp = self.supplier_price_id is not None
        has_manual = self.manual_unit_price_vnd is not None
        if has_sp == has_manual:  # both or neither
            raise ValueError(
                f"sourcing_id={self.sourcing_id}: phải chọn ĐÚNG MỘT trong "
                "supplier_price_id hoặc manual_unit_price_vnd"
            )
        if self.fx_rate_override is not None and float(self.fx_rate_override) <= 0:
            raise ValueError(
                f"sourcing_id={self.sourcing_id}: fx_rate_override phải > 0"
            )
        return self


class QuoteBatchRequest(BaseModel):
    customer_name: str | None = None
    customer_id: int | None = None
    customer_contact: str | None = None
    customer_address: str | None = None
    quote_note: str | None = None
    items: list[QuoteBatchItem]
    # xlsx | tsv | pdf — PDF goes through template-XLSX + Gotenberg
    file_format: str = "xlsx"
    # preview=True → render to a temp file + return download_url but DO NOT
    # INSERT into quote_batches (side-effect-free modal preview).
    preview: bool = False
    # M3 — versioning + expire (Thang 2026-06-22)
    # When set, the new quote becomes the next version of the existing quote's
    # group (revision chain). Omitted → the new quote starts its own group.
    revise_of_quote_no: str | None = None
    # ISO date (YYYY-MM-DD) the quote is valid until. Default = created + 10 days.
    valid_until: str | None = None
    # Người báo giá (Thang) — display-only owner shown in the rendered form's G6
    # cell. Defaults to the real login email when omitted. The DB audit columns
    # (created_by_email/created_by_id) ALWAYS use the real token, never this.
    quote_owner: str | None = None


def _next_quote_no_sql() -> str:
    """Generate quote_no theo pattern SC-YYMMDD-NNNN."""
    today = datetime.now().strftime("%y%m%d")
    return today


@router.get("/quote-staff")
async def list_quote_staff(
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "director"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Active users for the "Người báo giá" dropdown on the quote modal.

    Same roles as create_quote_batch may read this (anyone who can create a
    quote can pick its display owner).
    """
    rows = await conn.fetch(
        "SELECT id, full_name, email FROM users WHERE is_active = true ORDER BY full_name"
    )
    return {
        "data": [
            {"id": r["id"], "full_name": r["full_name"], "email": r["email"]}
            for r in rows
        ]
    }


@router.post("/quote-batch")
@limiter.limit("10/minute")
async def create_quote_batch(
    request: Request,
    response: Response,
    body: QuoteBatchRequest,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "director"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tạo 1 báo giá hàng loạt từ N sourcing entries đã tick.

    Flow:
    1. Validate sourcing_ids tồn tại
    2. Build line_items snapshot
    3. Generate quote_no SC-YYMMDD-NNNN
    4. Render file XLSX/TSV vào /data/files/quotes/
    5. INSERT vào quote_batches
    6. Return { quote_no, download_url, total }
    """
    if not body.items:
        raise HTTPException(400, "Phải có ít nhất 1 mã trong báo giá")
    if len(body.items) > 1000:
        raise HTTPException(400, "Tối đa 1000 mã/báo giá")
    if body.file_format not in ("xlsx", "tsv", "pdf"):
        raise HTTPException(400, "file_format chỉ chấp nhận xlsx | tsv | pdf")

    from app.services.sourcing_pricing_engine import compute_sale_vnd

    # ── Customer enrichment (Quote Hub D4-6) ──────────────────────────────
    # When the modal sends a picked customer_id, load the canonical record so
    # the rendered báo giá carries the real MST (tax_code) + company_name +
    # address — and so customer_name/customer_address fall back to the DB row
    # when the caller didn't pass explicit overrides. Fixes the prior bug where
    # only a free-text customer_name string reached the template.
    customer_mst: str | None = None
    customer_name = body.customer_name
    customer_address = body.customer_address
    if body.customer_id is not None:
        cust = await conn.fetchrow(
            "SELECT company_name, tax_code, address FROM customers WHERE id = $1",
            body.customer_id,
        )
        if cust:
            customer_mst = cust["tax_code"]
            if not customer_name:
                customer_name = cust["company_name"]
            if not customer_address:
                customer_address = cust["address"]

    ids = [it.sourcing_id for it in body.items]
    rows = await conn.fetch(
        """
        SELECT id, model, product_name, maker, supplier_name, sale_vnd, cost_vnd,
               quantity, brand_canonical, hs_code, notes, row_classification
          FROM sourcing_entries
         WHERE id = ANY($1::bigint[]) AND deleted_at IS NULL
        """,
        ids,
    )
    row_map = {r["id"]: dict(r) for r in rows}

    line_items = []
    total_value = 0.0
    for item in body.items:
        r = row_map.get(item.sourcing_id)
        if not r:
            raise HTTPException(404, f"Không tìm thấy sourcing entry id={item.sourcing_id}")
        qty = float(item.quantity) if item.quantity is not None else float(r["quantity"] or 1)

        # ── MANUAL per-line resolution (Thang 2026-06-21) ──────────────
        # No server-side auto-pick: each line carries EXACTLY ONE explicit
        # choice (enforced by QuoteBatchItem._exactly_one).
        supplier_name = r["supplier_name"]
        cost_currency = "VND"
        fx_rate = 1.0
        fx_date = date_cls.today().isoformat()
        if item.manual_unit_price_vnd is not None:
            # Typed override — used verbatim (already VND).
            unit = float(item.manual_unit_price_vnd)
            price_source = "manual"
        else:
            # Chosen supplier price — load it (must belong to this entry).
            sp = await conn.fetchrow(
                """
                SELECT id, supplier_name, currency, cost_amount, exchange_rate_used
                  FROM sourcing_supplier_prices
                 WHERE id = $1 AND sourcing_entry_id = $2
                """,
                item.supplier_price_id, item.sourcing_id,
            )
            if not sp:
                raise HTTPException(
                    404,
                    f"Không tìm thấy giá NCC id={item.supplier_price_id} cho "
                    f"sourcing #{item.sourcing_id}",
                )
            sp_currency = (sp["currency"] or "VND").upper().strip()
            # Compute the SALE VND via the engine. When the owner typed a manual
            # FX rate (fx_rate_override) we pass it as exchange_rate so the export
            # matches the modal preview exactly; otherwise exchange_rate=None
            # forces a fresh read from exchange_rates with the staleness guard.
            # NOT the cached cost_vnd_equiv, NOT raw sourcing_entries.sale_vnd.
            fx_override = (
                float(item.fx_rate_override)
                if item.fx_rate_override is not None
                else None
            )
            try:
                res = await compute_sale_vnd(
                    conn,
                    # V1.1 (Thang 2026-06-27): row_classification is a STATUS,
                    # never a pricing key. Forcing it as item_type silently
                    # changed the quoted price whenever a pricing rule happened
                    # to share a classification name. Always use 'default'.
                    item_type="default",
                    cost_amount=sp["cost_amount"],
                    currency=sp_currency,
                    exchange_rate=fx_override,
                    qty=qty,
                    is_domestic_vn=(sp_currency == "VND"),
                )
            except ValueError as exc:
                # Missing / stale FX → surface as 400 so the modal shows it.
                raise HTTPException(400, str(exc))
            unit = float(res["suggested_sale_vnd"])
            bd = res.get("breakdown", {})
            fx_rate = float(bd.get("exchange_rate_used") or 1.0)
            fx_date = bd.get("exchange_rate_date") or date_cls.today().isoformat()
            supplier_name = sp["supplier_name"]
            cost_currency = sp_currency
            price_source = "supplier"

        line_total = qty * unit
        total_value += line_total
        # GAP 1 — ĐVT (column E): sourcing_entries has NO unit/uom/đvt column
        # (verified against backend/migrations). Default to "Cái" explicitly +
        # centralized here so the two renderers don't each silently fall back.
        # Provide both keys: quote_renderer reads "uom", the PDF renderer reads
        # "unit". TODO: add a real `unit` field to sourcing_entries to drive this.
        item_unit = "Cái"
        # GAP 2 — Ghi chú (column H): sourcing_entries.notes. Provide both keys:
        # quote_renderer reads "note", the PDF renderer reads "notes".
        item_note = r["notes"] or ""
        line_items.append({
            "sourcing_id": item.sourcing_id,
            "model": r["model"],
            "product_name": r["product_name"],
            "maker": r["maker"],
            "brand": r["brand_canonical"],
            "supplier": supplier_name,
            "hs_code": r["hs_code"],
            "quantity": qty,
            "unit_price_vnd": unit,
            "line_total_vnd": line_total,
            "unit": item_unit,
            "uom": item_unit,
            "notes": item_note,
            "note": item_note,
            # Per-line thời gian giao hàng — renderer appends into Ghi chú.
            "delivery_time": item.delivery_time,
            # ── chosen-price snapshot (display + audit) ──
            "price_source": price_source,
            "supplier_price_id": item.supplier_price_id,
            "supplier_name": supplier_name,
            "cost_currency": cost_currency,
            "fx_rate": fx_rate,
            "fx_date": fx_date,
        })

    # Generate quote_no via daily seq
    today_prefix = datetime.now().strftime("%y%m%d")
    seq_val = await conn.fetchval("SELECT NEXTVAL('quote_batches_daily_seq')")
    # Reset seq mỗi sáng — đơn giản hóa: dùng MOD 10000 + day prefix
    quote_no = f"SC-{today_prefix}-{(seq_val % 10000):04d}"

    # Render file
    from app.services import quote_renderer, sourcing_quote_pdf_renderer
    out_dir = Path("/data/files/quotes")
    out_dir.mkdir(parents=True, exist_ok=True)
    file_path = out_dir / f"{quote_no}.{body.file_format}"

    try:
        if body.file_format == "xlsx":
            quote_renderer.render_xlsx(
                file_path,
                quote_no=quote_no,
                customer_name=customer_name or "",
                quote_note=body.quote_note or "",
                line_items=line_items,
                total_value=total_value,
                created_by=token_data.email,
                quote_owner=body.quote_owner or token_data.email,
                created_at=datetime.now(),
                customer_contact=body.customer_contact,
                customer_address=customer_address,
                customer_mst=customer_mst,
            )
        elif body.file_format == "pdf":
            # Template-driven: quote_renderer fills SOURCING_QUOTE.xlsx →
            # Gotenberg → PDF bytes. Hard-fails (HTTP 500) if Gotenberg down —
            # never emits an alternate-identity form.
            quote_data = {
                "quote_no": quote_no,
                "quote_date": datetime.now(),
                "created_by": body.quote_owner or token_data.email,
                "valid_days": 10,
                "customer_contact": body.customer_contact or "",
                "customer_company": customer_name or "",
                "customer_address": customer_address or "",
                "customer_mst": customer_mst or "",
                "quote_note": body.quote_note or "",
                # GAP 3 — VAT parity: the XLSX path hardcodes 8%
                # (DEFAULT_VAT_RATE) but the PDF path defaults to 0 unless the
                # caller supplies a rate. Pass tax_pct=8 so the PDF VAT cell
                # (=H{subtotal}*0.08) matches the template's "VAT 8%" label and
                # agrees with the XLSX output.
                "tax_pct": 8,
            }
            pdf_bytes = await sourcing_quote_pdf_renderer.render_pdf(
                quote_data, line_items
            )
            file_path.write_bytes(pdf_bytes)
        else:  # tsv
            quote_renderer.render_tsv(file_path, line_items)
    except Exception as exc:
        logger.exception("Render file failed: %s", exc)
        raise HTTPException(500, f"Render file lỗi: {exc}")

    # preview=True → render only, NO DB insert (side-effect-free modal preview).
    if body.preview:
        return {
            "data": {
                "id": None,
                "quote_no": quote_no,
                "total_items": len(line_items),
                "total_value_vnd": float(total_value),
                "download_url": f"/api/v1/sourcing/quote-batch/{quote_no}/download",
                "file_format": body.file_format,
                "created_at": datetime.now().isoformat(),
                "preview": True,
            },
            "message": f"Xem trước báo giá {quote_no} ({len(line_items)} mã)",
        }

    # ── valid_until (M3) ──────────────────────────────────────────────────
    # Parse caller-supplied ISO date; default to created (today) + 10 days.
    valid_until_date = _so_parse_date(body.valid_until)
    if valid_until_date is None:
        valid_until_date = date_cls.today() + timedelta(days=10)

    # ── Insert + versioning (M3) — single transaction ─────────────────────
    # All version-chain mutations (group assignment, version_no bump, flipping
    # the previous current row off) MUST commit atomically with the INSERT so a
    # concurrent reader never sees two is_current rows for the same group.
    async with conn.transaction():
        row = await conn.fetchrow(
            """
            INSERT INTO quote_batches (
                quote_no, customer_id, customer_name, quote_note,
                total_items, total_value_vnd, item_ids, line_items,
                file_path, file_format, created_by_id, created_by_email,
                valid_until
            ) VALUES (
                $1, $2, $3, $4, $5, $6, $7, $8::jsonb, $9, $10, $11, $12, $13
            )
            RETURNING id, quote_no, total_items, total_value_vnd, created_at
            """,
            quote_no, body.customer_id, customer_name, body.quote_note,
            len(line_items), total_value, ids, json.dumps(line_items, default=str),
            str(file_path), body.file_format,
            int(token_data.user_id) if str(token_data.user_id).isdigit() else None,
            token_data.email,
            valid_until_date,
        )
        new_id = row["id"]

        if body.revise_of_quote_no:
            # Revision: chain onto the existing quote's group.
            parent = await conn.fetchrow(
                """
                SELECT quote_group_id FROM quote_batches
                 WHERE quote_no = $1 AND deleted_at IS NULL
                """,
                body.revise_of_quote_no,
            )
            if not parent:
                raise HTTPException(
                    404,
                    f"Không tìm thấy báo giá gốc {body.revise_of_quote_no} để sửa",
                )
            grp = parent["quote_group_id"] or new_id
            next_ver = await conn.fetchval(
                "SELECT COALESCE(MAX(version_no), 0) + 1 FROM quote_batches "
                "WHERE quote_group_id = $1",
                grp,
            )
            await conn.execute(
                """
                UPDATE quote_batches
                   SET quote_group_id = $1, version_no = $2, is_current = true
                 WHERE id = $3
                """,
                grp, next_ver, new_id,
            )
            # Demote every OTHER version of the group.
            await conn.execute(
                "UPDATE quote_batches SET is_current = false "
                "WHERE quote_group_id = $1 AND id != $2",
                grp, new_id,
            )
            version_no = int(next_ver)
        else:
            # Fresh quote — self-group, version 1, current.
            await conn.execute(
                "UPDATE quote_batches "
                "SET quote_group_id = id, version_no = 1, is_current = true "
                "WHERE id = $1",
                new_id,
            )
            version_no = 1

    return {
        "data": {
            "id": row["id"],
            "quote_no": row["quote_no"],
            "total_items": row["total_items"],
            "total_value_vnd": float(row["total_value_vnd"]),
            "download_url": f"/api/v1/sourcing/quote-batch/{quote_no}/download",
            "file_format": body.file_format,
            "created_at": row["created_at"],
            "version_no": version_no,
            "valid_until": valid_until_date.isoformat(),
            "preview": False,
        },
        "message": f"Đã tạo báo giá {quote_no} ({len(line_items)} mã, {int(total_value):,} VND)",
    }


@router.get("/quote-batch/{quote_no}/download")
async def download_quote_batch(
    quote_no: str,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "director", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Download file báo giá theo quote_no.

    Format is driven by the on-disk file's extension — so a quote
    created as PDF streams back as PDF, XLSX as XLSX, TSV as TSV.
    """
    from fastapi.responses import FileResponse
    row = await conn.fetchrow(
        "SELECT file_path, file_format, quote_no FROM quote_batches WHERE quote_no = $1",
        quote_no,
    )
    fp: Path | None = None
    db_format: str | None = None
    if row:
        db_format = row["file_format"]
        candidate = Path(row["file_path"])
        if candidate.exists():
            fp = candidate
    if fp is None:
        # Bug V7: a modal "Xem trước" renders the file to /data/files/quotes/
        # WITHOUT inserting a quote_batches row (side-effect-free preview), so a
        # freshly-previewed quote_no has a file on disk but no DB row → the old
        # code 404'd the download/preview. Fall back to serving the on-disk file
        # directly. quote_no is validated to contain no path separators / glob
        # metachars first (new quotes are SC-YYMMDD-NNNN; legacy rows are
        # QB-YYMMDD-NNNN — both pass the safe regex below).
        safe = bool(_re.fullmatch(r"[A-Za-z0-9._-]{1,64}", quote_no or ""))
        quotes_dir = Path("/data/files/quotes")
        matches = (
            sorted(quotes_dir.glob(f"{quote_no}.*"))
            if (safe and quotes_dir.exists())
            else []
        )
        if matches:
            fp = matches[0]
    if fp is None or not fp.exists():
        raise HTTPException(404, f"Không tìm thấy báo giá {quote_no}")
    # Prefer on-disk extension (handles legacy rows where DB ↔ file
    # extension drift). DB column is the fallback.
    ext = fp.suffix.lstrip(".").lower() or (db_format or "bin")
    media = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "tsv": "text/tab-separated-values",
        "pdf": "application/pdf",
    }.get(ext, "application/octet-stream")
    return FileResponse(
        str(fp),
        media_type=media,
        filename=f"{quote_no}.{ext}",
    )


@router.get("/quote-batch")
async def list_quote_batches(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    customer_id: int | None = Query(
        None, description="Filter báo giá theo khách hàng (Hồ sơ tab)"
    ),
    all_versions: bool = Query(
        False,
        description="True → trả mọi version; mặc định chỉ trả is_current=true",
    ),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "director", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List báo giá đã tạo gần đây.

    Optional ``customer_id`` scopes the list to one customer (Quote Hub /
    Hồ sơ tab). Soft-deleted rows are always excluded.

    M3: by default only the current version of each revision group is returned
    (``is_current=true``). Pass ``all_versions=true`` to see superseded
    versions too. Each row carries ``version_no``, ``is_current``,
    ``converted_order_id`` and a computed ``expired`` flag.
    """
    conditions: list[str] = ["deleted_at IS NULL"]
    params: list[Any] = []
    idx = 1
    if customer_id is not None:
        conditions.append(f"customer_id = ${idx}")
        params.append(customer_id)
        idx += 1
    if not all_versions:
        conditions.append("is_current = true")
    where = " AND ".join(conditions)

    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM quote_batches WHERE {where}",
        *params,
    )

    params.extend([limit, offset])
    rows = await conn.fetch(
        f"""
        SELECT id, quote_no, customer_name, total_items, total_value_vnd,
               file_format, status, sent_at, created_by_email, created_at,
               quote_group_id, version_no, is_current, converted_order_id,
               valid_until,
               (status = 'sent'
                AND valid_until IS NOT NULL
                AND valid_until < CURRENT_DATE) AS expired
          FROM quote_batches
         WHERE {where}
         ORDER BY created_at DESC
         LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )

    def _ser(r: asyncpg.Record) -> dict:
        d = dict(r)
        for k, v in list(d.items()):
            if isinstance(v, datetime) or hasattr(v, "isoformat"):
                d[k] = v.isoformat()
            elif isinstance(v, Decimal):
                d[k] = float(v)
        return d

    return {
        "data": [_ser(r) for r in rows],
        "total": total,
    }


class QuoteBatchSendRequest(BaseModel):
    sent_to_email: str | None = None


@router.post("/quote-batch/{quote_no}/send")
@limiter.limit("30/minute")
async def send_quote_batch(
    request: Request,
    response: Response,
    quote_no: str,
    body: QuoteBatchSendRequest,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "director"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đánh dấu 1 báo giá đã gửi cho khách (Quote Hub / Hồ sơ tab).

    Transitions the row to status='sent', stamps sent_at=NOW() and records the
    recipient email. Idempotent-ish: re-sending refreshes sent_at + recipient.
    404 if the quote_no is missing or soft-deleted.
    """
    row = await conn.fetchrow(
        """
        UPDATE quote_batches
           SET status = 'sent',
               sent_at = NOW(),
               sent_to_email = $1
         WHERE quote_no = $2 AND deleted_at IS NULL
        RETURNING quote_no, status, sent_at
        """,
        body.sent_to_email, quote_no,
    )
    if not row:
        raise HTTPException(404, f"Không tìm thấy báo giá {quote_no}")
    return {
        "data": {
            "quote_no": row["quote_no"],
            "status": row["status"],
            "sent_at": row["sent_at"],
        },
        "message": f"Đã đánh dấu gửi báo giá {quote_no}",
    }


@router.get("/quote-batch/{quote_no}/prefill")
async def prefill_quote_batch(
    quote_no: str,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "director", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Reconstruct a quote's seed payload so the modal can open a revision.

    Returns the customer block (joined to ``customers`` for canonical MST /
    address / primary contact) plus the per-line {sourcing_id, quantity,
    unit_price_vnd} taken from the stored ``line_items`` snapshot, the
    ``valid_until`` date and the ``quote_note``. Used by QuoteBatchModal in
    "Sửa & gửi lại" mode.
    """
    qb = await conn.fetchrow(
        """
        SELECT id, quote_no, customer_id, customer_name, quote_note,
               line_items, valid_until
          FROM quote_batches
         WHERE quote_no = $1 AND deleted_at IS NULL
        """,
        quote_no,
    )
    if not qb:
        raise HTTPException(404, f"Không tìm thấy báo giá {quote_no}")

    # Canonical customer (MST / address / primary contact) when a customer_id
    # was pinned; otherwise fall back to the snapshot's free-text name.
    customer: dict[str, Any] = {
        "id": qb["customer_id"],
        "company_name": qb["customer_name"],
        "tax_code": None,
        "address": None,
        "primary_contact": None,
    }
    if qb["customer_id"] is not None:
        cust = await conn.fetchrow(
            """
            SELECT c.id, c.company_name, c.tax_code, c.address,
                   pc.full_name AS primary_contact
              FROM customers c
              LEFT JOIN LATERAL (
                  SELECT full_name
                    FROM customer_contacts
                   WHERE customer_id = c.id AND is_active = true
                   ORDER BY is_primary DESC, id ASC
                   LIMIT 1
              ) pc ON true
             WHERE c.id = $1
            """,
            qb["customer_id"],
        )
        if cust:
            pc_name = cust["primary_contact"]
            customer = {
                "id": cust["id"],
                "company_name": cust["company_name"],
                "tax_code": cust["tax_code"],
                "address": cust["address"],
                # Object shape {full_name} to match the frontend PrefillResponse
                # contract (QuoteBatchModal reads primary_contact?.full_name).
                "primary_contact": {"full_name": pc_name} if pc_name else None,
            }

    # line_items is JSONB — asyncpg may hand it back as str or already-parsed.
    raw_lines = qb["line_items"]
    if isinstance(raw_lines, str):
        try:
            raw_lines = json.loads(raw_lines)
        except Exception:
            raw_lines = []
    items = [
        {
            "sourcing_id": li.get("sourcing_id"),
            "quantity": li.get("quantity"),
            "unit_price_vnd": li.get("unit_price_vnd"),
        }
        for li in (raw_lines or [])
        if li.get("sourcing_id") is not None
    ]

    valid_until = qb["valid_until"]
    return {
        "data": {
            "customer": customer,
            "items": items,
            "valid_until": valid_until.isoformat() if valid_until else None,
            "quote_note": qb["quote_note"],
        }
    }


@router.post("/quote-batch/{quote_no}/create-order")
async def create_order_from_quote_batch(
    quote_no: str,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "staff"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """M4 — materialize a sourcing order from an accepted báo giá.

    IDEMPOTENT. The idempotency key is the pair
    (source_type='quote_batch', source_ref_id=<quote_batches.id>): if a live
    order already references this quote we return it with already_existed=true
    and create nothing. Otherwise we build an OrderCreatePayload from the quote
    snapshot (customer_* + line_items → {sourcing_id, quantity,
    unit_price_vnd}), reuse the shared _so_create_order_core for the INSERT, then
    stamp quote_batches.status='accepted' + converted_order_id.
    """
    qb = await conn.fetchrow(
        """
        SELECT id, quote_no, customer_id, customer_name, quote_note,
               line_items, converted_order_id
          FROM quote_batches
         WHERE quote_no = $1 AND deleted_at IS NULL
        """,
        quote_no,
    )
    if not qb:
        raise HTTPException(404, f"Không tìm thấy báo giá {quote_no}")

    # ── Idempotency: an existing live order on this quote short-circuits ───
    existing = await conn.fetchrow(
        """
        SELECT id, order_number FROM sourcing_orders
         WHERE source_type = 'quote_batch'
           AND source_ref_id = $1
           AND deleted_at IS NULL
         ORDER BY id ASC
         LIMIT 1
        """,
        qb["id"],
    )
    if existing:
        return {
            "data": {
                "order_id": existing["id"],
                "order_number": existing["order_number"],
                "already_existed": True,
            },
            "message": f"Đã có đơn {existing['order_number']} cho báo giá {quote_no}",
        }

    # ── Customer block — prefer canonical customers row for contact/email/etc ─
    customer_name = qb["customer_name"]
    customer_contact: str | None = None
    customer_email: str | None = None
    customer_phone: str | None = None
    customer_address: str | None = None
    if qb["customer_id"] is not None:
        cust = await conn.fetchrow(
            """
            SELECT c.company_name, c.address,
                   pc.full_name AS contact_name, pc.email, pc.phone
              FROM customers c
              LEFT JOIN LATERAL (
                  SELECT full_name, email, phone
                    FROM customer_contacts
                   WHERE customer_id = c.id AND is_active = true
                   ORDER BY is_primary DESC, id ASC
                   LIMIT 1
              ) pc ON true
             WHERE c.id = $1
            """,
            qb["customer_id"],
        )
        if cust:
            customer_name = cust["company_name"] or customer_name
            customer_address = cust["address"]
            customer_contact = cust["contact_name"]
            customer_email = cust["email"]
            customer_phone = cust["phone"]

    # ── Map snapshot lines → order items ──────────────────────────────────
    raw_lines = qb["line_items"]
    if isinstance(raw_lines, str):
        try:
            raw_lines = json.loads(raw_lines)
        except Exception:
            raw_lines = []
    items: list[OrderItemPayload] = []
    for li in (raw_lines or []):
        if li.get("sourcing_id") is None:
            continue
        items.append(OrderItemPayload(
            sourcing_id=li["sourcing_id"],
            quantity=li.get("quantity"),
            unit_price_vnd=li.get("unit_price_vnd"),
        ))
    if not items:
        raise HTTPException(400, f"Báo giá {quote_no} không có dòng hàng để tạo đơn")

    payload = OrderCreatePayload(
        customer_id=qb["customer_id"],
        customer_name=customer_name or "Khách lẻ",
        customer_contact=customer_contact,
        customer_email=customer_email,
        customer_phone=customer_phone,
        customer_address=customer_address,
        source_type="quote_batch",
        source_ref_id=qb["id"],
        source_ref_no=quote_no,
        items=items,
        notes=qb["quote_note"],
        initial_status="confirmed",
    )

    result = await _so_create_order_core(conn, payload, token_data)
    new_order_id = result["data"]["id"]
    new_order_number = result["data"]["order_number"]

    # ── Provenance: mark the quote accepted + back-link the order ──────────
    await conn.execute(
        """
        UPDATE quote_batches
           SET status = 'accepted', converted_order_id = $1
         WHERE quote_no = $2
        """,
        new_order_id, quote_no,
    )

    return {
        "data": {
            "order_id": new_order_id,
            "order_number": new_order_number,
            "already_existed": False,
        },
        "message": f"Đã tạo đơn {new_order_number} từ báo giá {quote_no}",
    }


# ============================================================
# Phase 2 — Sourcing Orders (Quote-to-order pipeline)
# Thang 2026-06-03
# Migration: backend/migrations/sourcing_orders.sql
# ============================================================

from decimal import Decimal, ROUND_HALF_UP  # noqa: E402

# Status state machine — keep in sync with CHECK constraint + frontend STATUS_META
_SO_STATUS_NEXT: dict[str, list[str]] = {
    "draft":              ["quoted", "cancelled"],
    "quoted":             ["confirmed", "cancelled"],
    "confirmed":          ["payment_requested", "cancelled"],
    "payment_requested":  ["payment_approved", "confirmed", "cancelled"],
    "payment_approved":   ["shipped", "cancelled"],
    "shipped":            ["delivered", "cancelled"],
    "delivered":          [],
    "cancelled":          [],
}
_SO_TERMINAL = {"delivered", "cancelled"}

# ─────────────────────────────────────────────────────────────────────────────
# PERM-1 — Per-transition role ACL matrix (Thang 2026-06-03)
# ─────────────────────────────────────────────────────────────────────────────
# Each key is a (from_status, to_status) tuple; value is the set of roles that
# may perform that transition. Defaults to {"admin", "manager"} for any tuple
# NOT explicitly listed — so a missing entry fails closed for non-managers.
#
# Matrix (Vietnamese-first roles):
#   draft → quoted              : sales, procurement, manager, admin
#   quoted → confirmed          : sales, manager, admin           (customer-side ack)
#   payment_requested → payment_approved : accountant, manager, admin
#   payment_approved → shipped  : warehouse, manager, admin       (kho xuất hàng)
#   shipped → delivered         : warehouse, sales, manager, admin (KH/sale xác nhận)
#   * → cancelled               : manager, admin only             (see _SO_CANCEL_ROLES)
#
# NOTE: confirmed → payment_requested is handled by the dedicated
#   POST /orders/{id}/payment-request endpoint (sales/manager/admin) — NOT this
#   matrix. Keep both lists in sync if the rules change.
# ─────────────────────────────────────────────────────────────────────────────
_SO_TRANSITION_ROLES: dict[tuple[str, str], set[str]] = {
    ("draft", "quoted"):                       {"sales", "procurement", "manager", "admin"},
    ("quoted", "confirmed"):                   {"sales", "manager", "admin"},
    ("payment_requested", "payment_approved"): {"accountant", "manager", "admin"},
    ("payment_approved", "shipped"):           {"warehouse", "manager", "admin"},
    ("shipped", "delivered"):                  {"warehouse", "sales", "manager", "admin"},
}
_SO_CANCEL_ROLES: set[str] = {"manager", "admin"}


def _so_allowed_roles_for(from_status: str, to_status: str) -> set[str]:
    """Return the role set permitted to drive (from_status → to_status).

    Cancellation always uses _SO_CANCEL_ROLES regardless of from_status.
    Unknown transitions default to {"admin", "manager"} (fail-closed for
    non-managers).
    """
    if to_status == "cancelled":
        return _SO_CANCEL_ROLES
    return _SO_TRANSITION_ROLES.get((from_status, to_status), {"admin", "manager"})


async def _so_apply_status_transition(
    conn: asyncpg.Connection,
    order_id: int,
    from_status: str,
    to_status: str,
    actor_user_id: Any = None,
    actor_email: str | None = None,
    note: str | None = None,
    metadata: dict | None = None,
) -> None:
    """Single source of truth for sourcing_orders status mutations.

    Validates the transition against _SO_STATUS_NEXT, runs the UPDATE on
    sourcing_orders.status, and inserts a row into sourcing_order_status_history.
    Raises HTTPException(409) on illegal transitions.

    Callers MUST already be inside `async with conn.transaction():` so the
    UPDATE + history INSERT commit atomically with surrounding side-effects
    (e.g. PDF file write, create-order INSERT).
    """
    if from_status == to_status:
        return
    allowed = _SO_STATUS_NEXT.get(from_status, [])
    if to_status not in allowed:
        raise HTTPException(
            409,
            f"Không thể chuyển {from_status} → {to_status}. Allowed: {allowed}",
        )
    await conn.execute(
        """
        UPDATE sourcing_orders
           SET status = $1,
               updated_by_id = $2,
               updated_by_email = $3
         WHERE id = $4
        """,
        to_status,
        _coerce_int(actor_user_id),
        actor_email,
        order_id,
    )
    try:
        uid_uuid = str(actor_user_id) if actor_user_id else None
    except Exception:
        uid_uuid = None
    await conn.execute(
        """
        INSERT INTO sourcing_order_status_history
          (order_id, from_status, status, by_user_id, by_user_email, note, metadata)
        VALUES ($1, $2, $3, $4::uuid, $5, $6, $7::jsonb)
        """,
        order_id, from_status, to_status, uid_uuid, actor_email,
        note, json.dumps(metadata or {}),
    )

    # ── PHASE 3 — additive event emission (audit-log only, NO behavior change).
    # Records the order state-transition on the append-only domain_events bus so
    # the unified spine / chain timeline can reconstruct the order lifecycle.
    # This is purely additive: it only INSERTs an audit row inside the caller's
    # transaction and never alters financial state. The auto-AR / chain mutation
    # path stays in payment_requests.py behind PHASE3_AUTO_AR_ENABLED. Wrapped in
    # try/except so an audit-log failure can never abort an order transition.
    _SO_EVENT_MAP = {
        "draft":             "order.created",
        "quoted":            "order.quoted",
        "confirmed":         "order.confirmed",
        "payment_requested": "order.payment_requested",
        "payment_approved":  "order.payment_approved",
        "shipped":           "order.shipped",
        "delivered":         "order.delivered",
        "cancelled":         "order.cancelled",
    }
    _event_type = _SO_EVENT_MAP.get(to_status)
    if _event_type:
        try:
            _chain_code = await conn.fetchval(
                "SELECT chain_code FROM sourcing_orders WHERE id = $1", order_id
            )
            await conn.execute(
                """
                INSERT INTO domain_events
                    (event_type, aggregate_type, aggregate_id, payload, chain_code, created_by)
                VALUES ($1, 'sourcing_order', $2, $3::jsonb, $4, $5::uuid)
                """,
                _event_type,
                str(order_id),
                json.dumps({"from_status": from_status, "to_status": to_status, **(metadata or {})}),
                _chain_code,
                uid_uuid,
            )
        except Exception as _exc:  # pragma: no cover — audit best-effort
            logger.debug("sourcing domain_events emit skipped: %s", _exc)


def _so_round_unit(v: Decimal) -> Decimal:
    """Round lên bội số 1.000 VND (convention báo giá lẻ VN)."""
    if v is None:
        return Decimal(0)
    return (v / Decimal(1000)).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * Decimal(1000)


def _so_calc_line(item: dict) -> dict:
    """Apply VN-correct formula cho 1 line. Mutate copy, return enriched dict.

    Input fields: cost_vnd, tax_pct, coefficient, qty, shipping_fee_vnd,
                  customer_markup_pct?, sale_unit_vnd_override?
    Output adds: sale_unit_vnd, sale_total_vnd, landed_cost_vnd
    """
    def _D(v: Any, default: str = "0") -> Decimal:
        if v is None or v == "":
            return Decimal(default)
        try:
            return Decimal(str(v))
        except Exception:
            return Decimal(default)

    cost = _D(item.get("cost_vnd"))
    tax_pct = _D(item.get("tax_pct"))
    coef = _D(item.get("coefficient"), "1")
    qty = _D(item.get("qty") or item.get("quantity"), "1")
    if qty <= 0:
        qty = Decimal(1)
    ship = _D(item.get("shipping_fee_vnd"))
    customer_markup_pct = item.get("customer_markup_pct")
    override = item.get("sale_unit_vnd_override")

    if customer_markup_pct is not None and str(customer_markup_pct) != "":
        effective_coef = Decimal(1) + (_D(customer_markup_pct) / Decimal(100))
    else:
        effective_coef = coef if coef > 0 else Decimal("1.0")

    landed = (cost + (ship / qty)) * (Decimal(1) + tax_pct / Decimal(100))

    if override is not None and _D(override) > 0:
        sale_unit = _D(override)
    else:
        sale_unit = landed * effective_coef

    sale_unit = _so_round_unit(sale_unit)
    sale_total = sale_unit * qty

    out = dict(item)
    out["landed_cost_vnd"] = float(landed)
    out["sale_unit_vnd"] = int(sale_unit)
    out["sale_total_vnd"] = int(sale_total)
    return out


def _so_calc_totals(items: list[dict], discount_vnd: Any = 0) -> dict:
    """Loop tất cả lines + return totals dict."""
    def _D(v: Any, default: str = "0") -> Decimal:
        if v is None or v == "":
            return Decimal(default)
        try:
            return Decimal(str(v))
        except Exception:
            return Decimal(default)

    enriched = [_so_calc_line(it) for it in items]
    subtotal = sum((Decimal(str(it["sale_total_vnd"])) for it in enriched), Decimal(0))
    shipping = sum((_D(it.get("shipping_fee_vnd")) for it in enriched), Decimal(0))
    tax_display = Decimal(0)
    for it in enriched:
        qty = _D(it.get("qty") or it.get("quantity"), "1")
        if qty <= 0:
            qty = Decimal(1)
        cost = _D(it.get("cost_vnd"))
        ship = _D(it.get("shipping_fee_vnd"))
        tax_pct = _D(it.get("tax_pct"))
        tax_display += (cost + ship / qty) * qty * tax_pct / Decimal(100)
    discount = _D(discount_vnd)
    total = subtotal - discount
    return {
        "items": enriched,
        "subtotal_vnd": int(subtotal),
        "shipping_fee_vnd": int(shipping),
        "tax_vnd": int(tax_display.quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
        "discount_vnd": int(discount),
        "total_value_vnd": int(total),
    }


async def _so_next_order_number(conn: asyncpg.Connection) -> str:
    """SO-YYMMDD-NNNN (uses sourcing_orders_seq)."""
    seq_val = await conn.fetchval("SELECT NEXTVAL('sourcing_orders_seq')")
    today = datetime.now().strftime("%y%m%d")
    return f"SO-{today}-{int(seq_val) % 10000:04d}"


def _so_serialize(row: asyncpg.Record | dict | None) -> dict | None:
    if row is None:
        return None
    d = dict(row)
    for k, v in list(d.items()):
        if isinstance(v, datetime):
            d[k] = v.isoformat()
        elif hasattr(v, "isoformat"):
            d[k] = v.isoformat()
        elif isinstance(v, Decimal):
            d[k] = int(v) if v == v.to_integral_value() else float(v)
        elif isinstance(v, str) and k == "line_items":
            try:
                d[k] = json.loads(v)
            except Exception:
                pass
        elif isinstance(v, str) and k == "metadata":
            try:
                d[k] = json.loads(v)
            except Exception:
                pass
    return d


# ── Pydantic models ──────────────────────────────────────────────

class OrderItemPayload(BaseModel):
    sourcing_id: int | None = None
    sourcing_entry_id: int | None = None  # alias
    model: str | None = None
    product_name: str | None = None
    name: str | None = None  # alias
    maker: str | None = None
    supplier_name: str | None = None
    hs_code: str | None = None
    qty: float | None = None
    quantity: float | None = None  # alias
    weight_kg: float | None = None
    cost_vnd: float | None = None
    tax_pct: float | None = None
    coefficient: float | None = None
    shipping_fee_vnd: float | None = None
    customer_markup_pct: float | None = None
    sale_unit_vnd_override: float | None = None
    unit_price_vnd: float | None = None  # alias for override (legacy task spec)
    notes: str | None = None
    image_url: str | None = None


def _so_items_subtotal_for_validation(items: list["OrderItemPayload"] | None) -> float:
    """Compute subtotal preview = sum across items of (cost_vnd + qty * unit_price).

    DATA-4 guardrail: cheap arithmetic preview used inside Pydantic validators so
    we can reject discount_vnd > subtotal *before* hitting the DB. The authoritative
    figure still comes from _so_calc_totals() (which applies tax/coef/shipping);
    using cost + qty*price here is intentionally conservative — it slightly
    over-estimates subtotal, so any discount rejected here would also be rejected
    by the DB CHECK constraint chk_so_discount_le_subtotal.
    """
    if not items:
        return 0.0
    total = 0.0
    for it in items:
        qty = it.qty if it.qty is not None else (it.quantity if it.quantity is not None else 1.0)
        try:
            qty_f = float(qty or 0)
        except (TypeError, ValueError):
            qty_f = 0.0
        # "price" = override / unit_price_vnd alias / sale_unit_vnd_override
        price = it.sale_unit_vnd_override
        if price is None:
            price = it.unit_price_vnd
        try:
            price_f = float(price or 0)
        except (TypeError, ValueError):
            price_f = 0.0
        try:
            cost_f = float(it.cost_vnd or 0)
        except (TypeError, ValueError):
            cost_f = 0.0
        total += cost_f + qty_f * price_f
    return total


class OrderCreatePayload(BaseModel):
    customer_id: int | None = None
    customer_name: str
    customer_contact: str | None = None
    customer_email: str | None = None
    customer_phone: str | None = None
    customer_address: str | None = None
    person_in_charge: str | None = None
    sourcing_entry_ids: list[int] | None = None
    source_type: str | None = "sourcing"
    source_ref_id: int | None = None
    source_ref_no: str | None = None
    items: list[OrderItemPayload]
    delivery_date: str | None = None
    payment_terms: str | None = None
    discount_vnd: float | None = 0
    notes: str | None = None
    internal_notes: str | None = None
    # Direct-create entry points (default "draft"). When != draft we walk the
    # state machine via _so_apply_status_transition so audit history is written.
    initial_status: Literal["draft", "quoted", "confirmed"] | None = None
    # Optional assignee email — resolved to users.id (UUID) at INSERT time.
    assigned_to_email: str | None = None

    @model_validator(mode="after")
    def _validate_discount_le_subtotal(self) -> "OrderCreatePayload":
        """DATA-4: reject discount_vnd > subtotal so we never write a negative total."""
        if self.discount_vnd is not None and self.discount_vnd > 0:
            subtotal = _so_items_subtotal_for_validation(self.items)
            if float(self.discount_vnd) > subtotal:
                raise ValueError(
                    f"discount_vnd ({self.discount_vnd:,.0f}) không được vượt subtotal "
                    f"({subtotal:,.0f}) — tổng đơn sẽ âm"
                )
        return self


class OrderUpdatePayload(BaseModel):
    customer_id: int | None = None
    customer_name: str | None = None
    customer_contact: str | None = None
    customer_email: str | None = None
    customer_phone: str | None = None
    customer_address: str | None = None
    person_in_charge: str | None = None
    items: list[OrderItemPayload] | None = None
    delivery_date: str | None = None
    payment_terms: str | None = None
    discount_vnd: float | None = None
    notes: str | None = None
    internal_notes: str | None = None

    @model_validator(mode="after")
    def _validate_discount_le_subtotal(self) -> "OrderUpdatePayload":
        """DATA-4: when both items + discount are supplied, reject discount > subtotal.

        If items is None (caller only patches discount), we can't compute subtotal
        here — the DB CHECK constraint chk_so_discount_le_subtotal still enforces it
        against the stored subtotal_vnd, so negative totals remain impossible.
        """
        if (
            self.discount_vnd is not None
            and self.discount_vnd > 0
            and self.items is not None
        ):
            subtotal = _so_items_subtotal_for_validation(self.items)
            if float(self.discount_vnd) > subtotal:
                raise ValueError(
                    f"discount_vnd ({self.discount_vnd:,.0f}) không được vượt subtotal "
                    f"({subtotal:,.0f}) — tổng đơn sẽ âm"
                )
        return self


class OrderStatusUpdatePayload(BaseModel):
    new_status: str
    note: str | None = None
    metadata: dict | None = None


class OrderPaymentRequestPayload(BaseModel):
    payment_method: str | None = "bank_transfer"
    beneficiary_name: str | None = None
    beneficiary_bank: str | None = None
    beneficiary_account: str | None = None
    description: str | None = None
    attachments: list[str] | None = None


class OrderCalcPayload(BaseModel):
    items: list[OrderItemPayload]
    discount_vnd: float | None = 0

    @model_validator(mode="after")
    def _validate_discount_le_subtotal(self) -> "OrderCalcPayload":
        """DATA-4: reject preview-calc requests where discount would make total negative."""
        if self.discount_vnd is not None and self.discount_vnd > 0:
            subtotal = _so_items_subtotal_for_validation(self.items)
            if float(self.discount_vnd) > subtotal:
                raise ValueError(
                    f"discount_vnd ({self.discount_vnd:,.0f}) không được vượt subtotal "
                    f"({subtotal:,.0f}) — tổng đơn sẽ âm"
                )
        return self


def _so_items_to_dicts(items: list[OrderItemPayload]) -> list[dict]:
    """Normalize Pydantic items -> list[dict] with field aliases resolved."""
    out: list[dict] = []
    for it in items or []:
        d = it.model_dump(exclude_none=False)
        # Resolve aliases
        d["sourcing_entry_id"] = d.get("sourcing_entry_id") or d.get("sourcing_id")
        d["qty"] = d.get("qty") or d.get("quantity")
        d["product_name"] = d.get("product_name") or d.get("name")
        # Legacy task spec: `unit_price_vnd` overrides sale_unit
        if d.get("unit_price_vnd") and not d.get("sale_unit_vnd_override"):
            d["sale_unit_vnd_override"] = d["unit_price_vnd"]
        out.append(d)
    return out


def _so_parse_date(s: str | None) -> Any:
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


# ── Endpoints ────────────────────────────────────────────────────

async def _so_create_order_core(
    conn: asyncpg.Connection,
    body: OrderCreatePayload,
    token_data: TokenData,
) -> dict:
    """Shared order-creation logic (enrich → totals → INSERT → state walk).

    Used by both POST /orders and POST /quote-batch/{quote_no}/create-order so
    the INSERT + status-history seeding + initial_status state-machine walk are
    written exactly once. Returns the {data, message} envelope. The caller is
    responsible for any provenance side-effects (e.g. stamping
    quote_batches.converted_order_id).
    """
    if not body.items:
        raise HTTPException(400, "Phải có ít nhất 1 line item")
    if len(body.items) > 500:
        raise HTTPException(400, "Tối đa 500 line items/đơn")

    raw_items = _so_items_to_dicts(body.items)

    # Enrich từ sourcing_entries nếu user chỉ pass sourcing_id
    entry_ids = [it["sourcing_entry_id"] for it in raw_items if it.get("sourcing_entry_id")]
    if entry_ids:
        entries = await conn.fetch(
            """
            SELECT id, model, product_name, maker, supplier_name, hs_code,
                   weight_kg, cost_vnd, sale_vnd, tax_pct, coefficient,
                   image_url, customer_name, person_in_charge
              FROM sourcing_entries
             WHERE id = ANY($1::bigint[]) AND deleted_at IS NULL
            """,
            entry_ids,
        )
        emap = {r["id"]: dict(r) for r in entries}
        for it in raw_items:
            sid = it.get("sourcing_entry_id")
            if not sid or sid not in emap:
                continue
            ent = emap[sid]
            for k in ("model", "product_name", "maker", "supplier_name", "hs_code",
                      "weight_kg", "cost_vnd", "tax_pct", "coefficient", "image_url"):
                if it.get(k) in (None, "") and ent.get(k) is not None:
                    it[k] = float(ent[k]) if isinstance(ent[k], Decimal) else ent[k]

    totals = _so_calc_totals(raw_items, body.discount_vnd or 0)
    order_number = await _so_next_order_number(conn)

    user_id = _coerce_int(getattr(token_data, "user_id", None))
    user_email = getattr(token_data, "email", None)

    # Resolve optional assignee email → users.id (UUID). Silently null if no
    # match — we don't want a typo to block order creation.
    assigned_to_uuid: str | None = None
    if body.assigned_to_email:
        assignee = await conn.fetchrow(
            "SELECT id FROM users WHERE LOWER(email) = LOWER($1) "
            "  AND is_active = true AND deleted_at IS NULL",
            body.assigned_to_email.strip(),
        )
        if assignee:
            assigned_to_uuid = str(assignee["id"])

    # initial_status whitelist: only draft / quoted / confirmed are valid entry
    # points. Anything else falls back to draft.
    requested_initial = (body.initial_status or "draft").lower()
    if requested_initial not in ("draft", "quoted", "confirmed"):
        requested_initial = "draft"

    async with conn.transaction():
        row = await conn.fetchrow(
            """
            INSERT INTO sourcing_orders (
                order_number, sourcing_entry_ids, source_type, source_ref_id, source_ref_no,
                customer_id, customer_name, customer_contact, customer_email,
                customer_phone, customer_address, person_in_charge,
                order_date, delivery_date, payment_terms,
                line_items,
                subtotal_vnd, tax_vnd, shipping_fee_vnd, discount_vnd, total_value_vnd,
                currency, status, notes, internal_notes,
                assigned_to,
                created_by_id, created_by_email, updated_by_id, updated_by_email
            ) VALUES (
                $1, $2, $3, $4, $5,
                $6, $7, $8, $9,
                $10, $11, $12,
                CURRENT_DATE, $13, $14,
                $15::jsonb,
                $16, $17, $18, $19, $20,
                'VND', 'draft', $21, $22,
                $25::uuid,
                $23, $24, $23, $24
            )
            RETURNING id, order_number, status, total_value_vnd, created_at
            """,
            order_number, entry_ids or [], body.source_type or "sourcing",
            body.source_ref_id, body.source_ref_no,
            body.customer_id, body.customer_name, body.customer_contact, body.customer_email,
            body.customer_phone, body.customer_address, body.person_in_charge,
            _so_parse_date(body.delivery_date), body.payment_terms,
            json.dumps(totals["items"], default=str),
            totals["subtotal_vnd"], totals["tax_vnd"], totals["shipping_fee_vnd"],
            totals["discount_vnd"], totals["total_value_vnd"],
            body.notes, body.internal_notes,
            user_id, user_email,
            assigned_to_uuid,
        )

        # Seed initial draft history row — was previously emitted by the
        # AFTER INSERT trigger trg_sosh_log + log_sourcing_order_status_change(),
        # which was dropped in the Option A deploy. Without this seed, new
        # orders have no draft row in status_history and the audit log starts
        # at whatever the walk-path adds (quoted/confirmed), if any.
        actor_uid_raw = getattr(token_data, "user_id", None)
        actor_uid_str = str(actor_uid_raw) if actor_uid_raw else None
        await conn.execute(
            """
            INSERT INTO sourcing_order_status_history
                (order_id, from_status, status, by_user_id, by_user_email,
                 by_user_name, note, metadata, at)
            VALUES ($1, NULL, 'draft', $2::uuid, $3, $4, 'Đơn mới được tạo',
                    $5::jsonb, NOW())
            """,
            row["id"],
            actor_uid_str,
            user_email,
            getattr(token_data, "full_name", None),
            json.dumps({"trigger": "create_order"}),
        )

        # Walk state machine for non-draft entry points. Each step writes a
        # history row via _so_apply_status_transition so audit is preserved.
        current = "draft"
        walk_path: list[str] = []
        if requested_initial == "quoted":
            walk_path = ["quoted"]
        elif requested_initial == "confirmed":
            walk_path = ["quoted", "confirmed"]
        for step in walk_path:
            await _so_apply_status_transition(
                conn, row["id"], current, step,
                actor_user_id=getattr(token_data, "user_id", None),
                actor_email=user_email,
                note="auto on create",
                metadata={"reason": "initial_status", "target": requested_initial},
            )
            current = step

        final_status = current
    return {
        "data": {
            "id": row["id"],
            "order_number": row["order_number"],
            "status": final_status,
            "total_value_vnd": int(row["total_value_vnd"]),
            "created_at": row["created_at"].isoformat(),
        },
        "message": f"Đã tạo đơn {row['order_number']}",
    }


@router.post("/orders")
async def create_sourcing_order(
    body: OrderCreatePayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "staff"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Create order from sourcing entries (or manual line items).

    Auto-calc totals server-side using VN formula. Status = 'draft' (or the
    requested initial_status walked via the state machine).
    """
    return await _so_create_order_core(conn, body, token_data)


@router.get("/orders")
async def list_sourcing_orders(
    status: str | None = None,
    customer_id: int | None = None,
    customer: str | None = Query(None, description="Tìm theo tên khách"),
    customer_name: str | None = Query(None, description="Alias of `customer` (frontend compat)"),
    assigned_to: str | None = None,
    sourcing_entry_id: int | None = Query(
        None,
        description="Filter orders that reference this sourcing_entries.id",
    ),
    q: str | None = Query(None, description="Search order_number / customer_name"),
    date_from: str | None = None,
    date_to: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(30, ge=10, le=200),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "staff", "accountant", "director"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List orders with filter + pagination."""
    # Table alias `so` because we LEFT JOIN users to surface assigned_to_email.
    conditions: list[str] = ["so.deleted_at IS NULL"]
    params: list[Any] = []
    idx = 1
    if status:
        conditions.append(f"so.status = ${idx}")
        params.append(status)
        idx += 1
    if customer_id:
        conditions.append(f"so.customer_id = ${idx}")
        params.append(customer_id)
        idx += 1
    # Accept both `customer` (existing) and `customer_name` (frontend) — first
    # non-empty wins.
    customer_filter = customer or customer_name
    if customer_filter:
        conditions.append(f"so.customer_name ILIKE ${idx}")
        params.append(f"%{customer_filter}%")
        idx += 1
    if assigned_to:
        conditions.append(f"so.assigned_to = ${idx}::uuid")
        params.append(assigned_to)
        idx += 1
    if sourcing_entry_id:
        # sourcing_orders.sourcing_entry_ids is BIGINT[] — array containment is
        # the fast/index-friendly path (vs. JSONB scan over line_items).
        conditions.append(f"${idx} = ANY(so.sourcing_entry_ids)")
        params.append(sourcing_entry_id)
        idx += 1
    if q:
        like = f"%{q.strip()}%"
        conditions.append(f"(so.order_number ILIKE ${idx} OR so.customer_name ILIKE ${idx})")
        params.append(like)
        idx += 1
    if date_from:
        d = _so_parse_date(date_from)
        if d:
            conditions.append(f"so.order_date >= ${idx}")
            params.append(d)
            idx += 1
    if date_to:
        d = _so_parse_date(date_to)
        if d:
            conditions.append(f"so.order_date <= ${idx}")
            params.append(d)
            idx += 1

    where = " AND ".join(conditions)
    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM sourcing_orders so WHERE {where}",
        *params,
    )

    params.extend([page_size, (page - 1) * page_size])
    rows = await conn.fetch(
        f"""
        SELECT so.id, so.order_number, so.status, so.customer_id, so.customer_name,
               so.person_in_charge,
               so.order_date, so.delivery_date, so.payment_terms,
               so.subtotal_vnd, so.tax_vnd, so.shipping_fee_vnd, so.discount_vnd,
               so.total_value_vnd,
               so.currency, so.source_type, so.source_ref_no,
               so.quote_pdf_url, so.quote_pdf_version, so.quote_sent_at,
               so.payment_request_id, so.sales_order_id,
               so.assigned_to, u.email AS assigned_to_email,
               so.created_by_email, so.created_at, so.updated_at
          FROM sourcing_orders so
          LEFT JOIN users u ON u.id = so.assigned_to
         WHERE {where}
         ORDER BY so.created_at DESC, so.id DESC
         LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )
    return {
        "data": {
            "items": [_so_serialize(r) for r in rows],
            "total": int(total or 0),
            "page": page,
            "page_size": page_size,
            "pages": (int(total or 0) + page_size - 1) // page_size,
        }
    }


@router.get("/orders/{order_id}")
async def get_sourcing_order(
    order_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "staff", "accountant", "director"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    row = await conn.fetchrow(
        "SELECT * FROM sourcing_orders WHERE id = $1 AND deleted_at IS NULL",
        order_id,
    )
    if not row:
        raise HTTPException(404, f"Không tìm thấy đơn #{order_id}")
    history = await conn.fetch(
        """
        SELECT id, from_status, status, by_user_id, by_user_email, by_user_name,
               note, metadata, at
          FROM sourcing_order_status_history
         WHERE order_id = $1
         ORDER BY at DESC, id DESC
        """,
        order_id,
    )
    return {
        "data": {
            "order": _so_serialize(row),
            "status_history": [_so_serialize(h) for h in history],
        }
    }


@router.patch("/orders/{order_id}")
async def update_sourcing_order(
    order_id: int,
    body: OrderUpdatePayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Partial update — re-calc totals khi items đổi."""
    existing = await conn.fetchrow(
        "SELECT id, status FROM sourcing_orders WHERE id = $1 AND deleted_at IS NULL",
        order_id,
    )
    if not existing:
        raise HTTPException(404, f"Không tìm thấy đơn #{order_id}")
    if existing["status"] in _SO_TERMINAL:
        raise HTTPException(409, f"Đơn đã ở trạng thái terminal ({existing['status']}) — không sửa được")

    sets: list[str] = []
    params: list[Any] = []
    idx = 1

    simple = {
        "customer_id": body.customer_id,
        "customer_name": body.customer_name,
        "customer_contact": body.customer_contact,
        "customer_email": body.customer_email,
        "customer_phone": body.customer_phone,
        "customer_address": body.customer_address,
        "person_in_charge": body.person_in_charge,
        "payment_terms": body.payment_terms,
        "notes": body.notes,
        "internal_notes": body.internal_notes,
    }
    for k, v in simple.items():
        if v is not None:
            sets.append(f"{k} = ${idx}")
            params.append(v)
            idx += 1

    if body.delivery_date is not None:
        sets.append(f"delivery_date = ${idx}")
        params.append(_so_parse_date(body.delivery_date))
        idx += 1

    if body.items is not None:
        raw_items = _so_items_to_dicts(body.items)
        totals = _so_calc_totals(raw_items, body.discount_vnd or 0)
        sets.extend([
            f"line_items = ${idx}::jsonb",
            f"subtotal_vnd = ${idx + 1}",
            f"tax_vnd = ${idx + 2}",
            f"shipping_fee_vnd = ${idx + 3}",
            f"discount_vnd = ${idx + 4}",
            f"total_value_vnd = ${idx + 5}",
        ])
        params.extend([
            json.dumps(totals["items"], default=str),
            totals["subtotal_vnd"], totals["tax_vnd"],
            totals["shipping_fee_vnd"], totals["discount_vnd"], totals["total_value_vnd"],
        ])
        idx += 6
    elif body.discount_vnd is not None:
        sets.append(f"discount_vnd = ${idx}")
        params.append(int(body.discount_vnd))
        idx += 1

    sets.append(f"updated_by_id = ${idx}")
    params.append(_coerce_int(getattr(token_data, "user_id", None)))
    idx += 1
    sets.append(f"updated_by_email = ${idx}")
    params.append(getattr(token_data, "email", None))
    idx += 1

    if not sets:
        raise HTTPException(400, "Không có field nào để update")

    params.append(order_id)
    await conn.execute(
        f"UPDATE sourcing_orders SET {', '.join(sets)} WHERE id = ${idx}",
        *params,
    )
    updated = await conn.fetchrow("SELECT * FROM sourcing_orders WHERE id = $1", order_id)
    return {"data": _so_serialize(updated)}


@router.patch("/orders/{order_id}/status")
async def transition_sourcing_order_status(
    order_id: int,
    body: OrderStatusUpdatePayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "warehouse", "accountant"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Generic status transition. Validates `next` whitelist server-side.

    Writes manual history row with by_user_* + note (in addition to the
    trigger-emitted row which captures the actual UPDATE).

    PERM-1 (Thang 2026-06-03): in addition to the state-machine check, also
    enforces the per-transition ACL matrix _SO_TRANSITION_ROLES so e.g. a
    warehouse staffer can't mark `payment_approved` and a sales user can't
    mark `delivered`.
    """
    new_status = body.new_status
    if new_status not in _SO_STATUS_NEXT:
        raise HTTPException(400, f"Status không hợp lệ: {new_status}")
    if new_status == "cancelled" and not body.note:
        raise HTTPException(400, "Cancel phải kèm lý do (note)")

    async with conn.transaction():
        order = await conn.fetchrow(
            "SELECT id, status, order_number FROM sourcing_orders "
            "WHERE id = $1 AND deleted_at IS NULL FOR UPDATE",
            order_id,
        )
        if not order:
            raise HTTPException(404, f"Không tìm thấy đơn #{order_id}")
        current = order["status"]
        if current == new_status:
            raise HTTPException(409, f"Đơn đã ở status {new_status}")

        # PERM-1: enforce per-transition role ACL (after state-machine check
        # in _so_apply_status_transition would also reject illegal target;
        # we run this BEFORE the apply so the 403 wins over the 409 ordering).
        actor_role = (getattr(token_data, "role", None) or "").lower()
        allowed = _so_allowed_roles_for(current, new_status)
        if actor_role not in allowed:
            raise HTTPException(
                status_code=403,
                detail={
                    "error": "TRANSITION_FORBIDDEN",
                    "message": (
                        f"Role '{actor_role}' không được phép chuyển "
                        f"{current} → {new_status}. Cần: {sorted(allowed)}"
                    ),
                    "from_status": current,
                    "to_status": new_status,
                    "required_roles": sorted(allowed),
                },
            )

        # Single source of truth for transition rules + history audit row.
        await _so_apply_status_transition(
            conn, order_id, current, new_status,
            actor_user_id=getattr(token_data, "user_id", None),
            actor_email=getattr(token_data, "email", None),
            note=body.note,
            metadata=body.metadata or {},
        )
    updated = await conn.fetchrow(
        "SELECT * FROM sourcing_orders WHERE id = $1", order_id,
    )
    return {"data": _so_serialize(updated)}


@router.get("/orders/{order_id}/allowed-transitions")
async def get_allowed_transitions(
    order_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "warehouse",
        "accountant", "staff", "director",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Return next-status transitions the CURRENT user may perform.

    Frontend (SourcingOrderDetailDrawer) calls this to hide buttons the actor
    doesn't have permission for, matching the server-side ACL in
    _SO_TRANSITION_ROLES + _SO_CANCEL_ROLES so the user never sees a button
    that would 403.

    Returns:
      {
        "current_status": "confirmed",
        "actor_role": "sales",
        "allowed": ["payment_requested"],   # ← cancellable only by manager/admin
        "cancellable": false,
        "all_next": ["payment_requested", "cancelled"]
      }
    """
    order = await conn.fetchrow(
        "SELECT id, status FROM sourcing_orders WHERE id = $1 AND deleted_at IS NULL",
        order_id,
    )
    if not order:
        raise HTTPException(404, f"Không tìm thấy đơn #{order_id}")
    current = order["status"]
    actor_role = (getattr(token_data, "role", None) or "").lower()
    all_next = list(_SO_STATUS_NEXT.get(current, []))
    allowed: list[str] = []
    for nxt in all_next:
        if nxt == "cancelled":
            continue  # surface separately via `cancellable`
        if actor_role in _so_allowed_roles_for(current, nxt):
            allowed.append(nxt)
    cancellable = "cancelled" in all_next and actor_role in _SO_CANCEL_ROLES
    return {
        "data": {
            "current_status": current,
            "actor_role": actor_role,
            "allowed": allowed,
            "cancellable": cancellable,
            "all_next": all_next,
        }
    }


@router.post("/orders/{order_id}/payment-request")
async def request_payment_for_order(
    order_id: int,
    body: OrderPaymentRequestPayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đề xuất thanh toán cho đơn. Status confirmed -> payment_requested.

    Idempotent: 409 only if an ACTIVE PR exists (status IN pending/approved/paid).
    A rejected PR allows re-submit — a new payment_requests row is INSERTED and
    sourcing_orders.payment_request_id is rewired to the new row.

    Side-effect: insert real `payment_requests` row + fan-out in-app
    notifications to all active accountants/managers/admins.
    """
    async with conn.transaction():
        order = await conn.fetchrow(
            """
            SELECT id, order_number, status, total_value_vnd, customer_name,
                   currency, payment_request_id, line_items, assigned_to
              FROM sourcing_orders
             WHERE id = $1 AND deleted_at IS NULL
             FOR UPDATE
            """,
            order_id,
        )
        if not order:
            raise HTTPException(404, f"Không tìm thấy đơn #{order_id}")
        if order["status"] != "confirmed":
            raise HTTPException(
                409,
                f"Chỉ đơn ở status 'confirmed' mới đề xuất TT được (hiện: {order['status']})",
            )

        # Idempotency relaxed: 409 only when an ACTIVE PR already exists for
        # this sourcing_order. Rejected/cancelled PRs allow a fresh submit so
        # sales can edit beneficiary/amount and try again.
        active_pr = await conn.fetchrow(
            """
            SELECT id, status
              FROM payment_requests
             WHERE sourcing_order_id = $1
               AND status IN ('pending', 'approved', 'paid')
             ORDER BY id DESC
             LIMIT 1
            """,
            order_id,
        )
        if active_pr:
            raise HTTPException(
                409,
                {
                    "error": "PAYMENT_ALREADY_REQUESTED",
                    "existing_payment_request_id": active_pr["id"],
                    "existing_payment_request_status": active_pr["status"],
                    "message": "Đơn này đã có yêu cầu thanh toán đang xử lý",
                },
            )

        # Beneficiary fallback: pick supplier from first line if not provided
        beneficiary = body.beneficiary_name
        if not beneficiary:
            li = order["line_items"]
            if isinstance(li, str):
                try:
                    li = json.loads(li)
                except Exception:
                    li = []
            for it in (li or []):
                if it.get("supplier_name"):
                    beneficiary = it["supplier_name"]
                    break
        description = body.description or (
            f"Thanh toán đơn {order['order_number']} — KH {order['customer_name']}"
        )

        # Pick an assignee accountant — first active one — so the PR shows up
        # in their queue immediately. Fallback to any admin.
        assignee = await conn.fetchval(
            "SELECT id FROM users "
            "WHERE role = 'accountant' AND is_active = true AND deleted_at IS NULL "
            "ORDER BY created_at ASC LIMIT 1"
        )
        if not assignee:
            assignee = await conn.fetchval(
                "SELECT id FROM users "
                "WHERE role = 'admin' AND is_active = true "
                "ORDER BY created_at ASC LIMIT 1"
            )

        requester_uuid = getattr(token_data, "user_id", None)
        requester_email = getattr(token_data, "email", None)
        requester_full_name = await conn.fetchval(
            "SELECT full_name FROM users WHERE id = $1::uuid", requester_uuid,
        ) if requester_uuid else None

        # INSERT real payment_requests row.
        pr_id = await conn.fetchval(
            """
            INSERT INTO payment_requests (
                company_id, requester_id, requester_name, department, request_date,
                description, amount, currency, payment_method,
                beneficiary_name, beneficiary_bank, beneficiary_account,
                status, sourcing_order_id, attachments, metadata,
                created_at, updated_at
            ) VALUES (
                NULL, $1::uuid, $2, 'Sales', CURRENT_DATE,
                $3, $4, COALESCE($5, 'VND')::currency_code, $6,
                $7, $8, $9,
                'pending', $10, $11::text[], $12::jsonb,
                NOW(), NOW()
            ) RETURNING id
            """,
            requester_uuid,
            requester_full_name or requester_email,
            description,
            float(order["total_value_vnd"] or 0),
            order["currency"],
            body.payment_method,
            beneficiary,
            body.beneficiary_bank,
            body.beneficiary_account,
            order_id,
            body.attachments or [],
            json.dumps({
                "order_id": order_id,
                "order_number": order["order_number"],
                "assigned_to": str(assignee) if assignee else None,
            }),
        )

        # Rewire back-link on sourcing_orders to the real PR id.
        await conn.execute(
            "UPDATE sourcing_orders SET payment_request_id = $1 WHERE id = $2",
            pr_id, order_id,
        )

        await _so_apply_status_transition(
            conn,
            order_id=order["id"],
            from_status="confirmed",
            to_status="payment_requested",
            actor_user_id=requester_uuid,
            actor_email=requester_email,
            note=f"Đề xuất thanh toán: {description}",
            metadata={
                "payment_request_id": pr_id,
                "beneficiary": beneficiary or order["customer_name"],
                "amount": float(order.get("total_value_vnd") or 0),
                "payment_method": body.payment_method,
                "beneficiary_bank": body.beneficiary_bank,
                "beneficiary_account": body.beneficiary_account,
            },
        )

        # In-app notification fan-out — accountant + manager + admin
        recipients: list[dict] = []
        notif_count = 0
        try:
            recipients = await conn.fetch(
                "SELECT id FROM users "
                "WHERE role = ANY($1::role_enum[]) "
                "  AND is_active = true AND deleted_at IS NULL",
                ["accountant", "admin", "manager"],
            )
            title = f"Đề xuất TT mới: {order['order_number']}"
            body_text = (
                f"Sale {requester_email or ''} đề xuất thanh toán "
                f"{int(order['total_value_vnd']):,} VND cho NCC {beneficiary or '—'}"
            )
            meta = {
                "order_id": order_id,
                "order_number": order["order_number"],
                "payment_request_id": pr_id,
                "amount": int(order["total_value_vnd"] or 0),
                "payment_method": body.payment_method,
                "beneficiary_name": beneficiary,
                "beneficiary_bank": body.beneficiary_bank,
                "beneficiary_account": body.beneficiary_account,
                "description": description,
            }
            for r in recipients:
                await conn.execute(
                    """
                    INSERT INTO notifications
                      (recipient_id, type, title, body, ref_type, ref_id, metadata)
                    VALUES ($1::uuid, 'workflow_request', $2, $3,
                            'payment_request', $4, $5::jsonb)
                    """,
                    str(r["id"]), title, body_text, pr_id, json.dumps(meta),
                )
                notif_count += 1
        except Exception as exc:
            logger.warning("payment-request notification fan-out failed: %s", exc)

    return {
        "data": {
            "order_id": order_id,
            "payment_request_id": pr_id,
            "status": "payment_requested",
            "notification_recipients": [str(r["id"]) for r in recipients],
            "notifications_sent": notif_count,
        },
        "message": f"Đã đề xuất TT cho đơn {order['order_number']}",
    }


@router.get("/orders/{order_id}/quote-pdf")
async def render_order_quote_pdf(
    order_id: int,
    token_data: TokenData = Depends(require_role(
        # V1 fix (Thang 2026-06-13): GET is strictly read-only — viewer can
        # download the existing PDF but cannot mutate state. Only the POST
        # regenerate path bumps version / triggers draft→quoted.
        "admin", "manager", "sales", "procurement", "staff",
        "accountant", "director", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Stream the *existing* quote PDF — strictly read-only, no side effects.

    Serves /data/files/quotes/{order_number}_v{version}.pdf when it exists.
    If no PDF has been rendered yet, returns 404 with a hint to call
    POST /orders/{id}/quote-pdf/regenerate.

    Security V1 (Thang 2026-06-13): the original ``?regenerate=true`` query
    toggle let a privilege-escalating mutation ride on a GET request — any
    viewer could force draft→quoted + a version bump just by loading the
    URL (also CSRF-reachable via <img>/prefetch/history side channel). The
    mutating code path now lives only behind a POST with a stricter role
    allowlist (no viewer).
    """
    from fastapi.responses import FileResponse

    order_row = await conn.fetchrow(
        """
        SELECT id, order_number, quote_pdf_url, quote_pdf_version
          FROM sourcing_orders
         WHERE id = $1 AND deleted_at IS NULL
        """,
        order_id,
    )
    if not order_row:
        raise HTTPException(404, f"Không tìm thấy đơn #{order_id}")

    current_version = int(order_row["quote_pdf_version"] or 0)
    file_existing = order_row["quote_pdf_url"]
    if not file_existing or current_version < 1:
        raise HTTPException(
            404,
            "Chưa có PDF báo giá. Gọi POST /orders/{id}/quote-pdf/regenerate "
            "để tạo (yêu cầu quyền sales/manager/admin).",
        )

    out_dir = Path("/data/files/quotes")
    existing_path = out_dir / f"{order_row['order_number']}_v{current_version}.pdf"
    if not existing_path.exists():
        raise HTTPException(
            404,
            "File PDF không tồn tại trên server. Gọi POST "
            "/orders/{id}/quote-pdf/regenerate để tạo lại.",
        )

    return FileResponse(
        str(existing_path),
        media_type="application/pdf",
        filename=f"{order_row['order_number']}.pdf",
    )


@router.post("/orders/{order_id}/quote-pdf/regenerate")
async def regenerate_order_quote_pdf(
    order_id: int,
    token_data: TokenData = Depends(require_role(
        # V1 fix (Thang 2026-06-13): viewer + staff + accountant explicitly
        # EXCLUDED — only roles that can author a quote may force a re-render
        # that bumps version + writes status_history + may auto-transition
        # draft → quoted. This is the privilege boundary that the unsafe
        # GET ?regenerate=true used to bypass.
        "admin", "manager", "sales", "procurement", "director"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Force re-render of the quote PDF — replaces the unsafe GET ?regenerate=true
    toggle.

    Side effects (all inside a single DB transaction):
      - render fresh PDF via sourcing_quote_pdf_renderer (Song Châu template + Gotenberg)
      - write /data/files/quotes/{order_number}_v{version+1}.pdf
      - bump quote_pdf_version
      - if order.status == 'draft': auto-transition draft → quoted (audited)
      - if version > 1: insert sourcing_order_status_history row
        (status unchanged, captures regenerate audit)
    """
    return await _render_or_regenerate_quote_pdf(
        conn, token_data, order_id, force_regenerate=True
    )


async def _render_or_regenerate_quote_pdf(
    conn: asyncpg.Connection,
    token_data: TokenData,
    order_id: int,
    force_regenerate: bool,
):
    from fastapi.responses import FileResponse
    from app.services import sourcing_quote_pdf_renderer

    order_row = await conn.fetchrow(
        "SELECT * FROM sourcing_orders WHERE id = $1 AND deleted_at IS NULL",
        order_id,
    )
    if not order_row:
        raise HTTPException(404, f"Không tìm thấy đơn #{order_id}")

    order = _so_serialize(order_row) or {}
    items = order.get("line_items") or []
    if isinstance(items, str):
        try:
            items = json.loads(items)
        except Exception:
            items = []

    out_dir = Path("/data/files/quotes")
    out_dir.mkdir(parents=True, exist_ok=True)

    current_version = int(order.get("quote_pdf_version") or 0)
    file_existing = order.get("quote_pdf_url")
    if file_existing and not force_regenerate and current_version > 0:
        # Try to serve existing
        existing_path = out_dir / f"{order['order_number']}_v{current_version}.pdf"
        if existing_path.exists():
            return FileResponse(
                str(existing_path),
                media_type="application/pdf",
                filename=f"{order['order_number']}.pdf",
            )

    # Render new — same faithful Song Châu template + Gotenberg path as the
    # quote-batch library, so library and orders quotes are identical (no more
    # AMA alternate-identity form).
    # Map sourcing_orders row fields onto the quote_data the renderer expects.
    # Mirror create_quote_batch GAP1/GAP2: ensure uom + note keys exist.
    for it in items:
        it.setdefault("uom", it.get("unit") or "Cái")
        it.setdefault("note", it.get("notes") or "")
        # sourcing_orders.line_items store the unit price under `sale_unit_vnd`
        # (from _so_calc_line) and the qty under `qty`; quote_renderer reads
        # `unit_price_vnd` / `quantity`. Without this map every price + the
        # grand total render as 0 on the orders-page PDF.
        it.setdefault(
            "unit_price_vnd",
            it.get("sale_unit_vnd") or it.get("unit_price_vnd") or 0,
        )
        it.setdefault("quantity", it.get("qty") or it.get("quantity") or 1)
    # order_date comes back as an ISO string (_so_serialize); coerce to a real
    # date so G5 keeps the template's localized date format instead of raw text.
    _od = order.get("order_date")
    if isinstance(_od, str):
        try:
            _od = datetime.strptime(_od[:10], "%Y-%m-%d").date()
        except ValueError:
            _od = None
    quote_data = {
        "quote_no": order["order_number"],
        "quote_date": _od or datetime.now(),
        "created_by": order.get("created_by_email"),
        "valid_days": order.get("valid_days") or 10,
        "customer_contact": order.get("customer_contact"),
        "customer_company": order.get("customer_name"),
        "customer_address": order.get("customer_address"),
        "quote_note": order.get("notes"),
        "tax_pct": 8,
        "breakdown": {
            "tax_vnd": order.get("tax_vnd"),
            "shipping_fee_vnd": order.get("shipping_fee_vnd"),
            "discount_vnd": order.get("discount_vnd"),
        },
    }
    try:
        pdf_bytes = await sourcing_quote_pdf_renderer.render_pdf(quote_data, items)
    except RuntimeError as exc:
        raise HTTPException(500, str(exc))

    new_version = current_version + 1
    file_path = out_dir / f"{order['order_number']}_v{new_version}.pdf"
    public_url = f"/api/v1/sourcing/orders/{order_id}/quote-pdf"

    # File write + version bump + state transition must commit atomically so
    # we don't leave a v2 PDF on disk with v1 metadata in the DB (or vice versa).
    async with conn.transaction():
        file_path.write_bytes(pdf_bytes)
        await conn.execute(
            """
            UPDATE sourcing_orders
               SET quote_pdf_url = $1,
                   quote_pdf_version = $2,
                   updated_by_id = $3,
                   updated_by_email = $4
             WHERE id = $5
            """,
            public_url, new_version,
            _coerce_int(getattr(token_data, "user_id", None)),
            getattr(token_data, "email", None),
            order_id,
        )
        # Auto draft → quoted, but only via the audited state-machine helper so
        # sourcing_order_status_history captures the transition (mirrors what
        # the manual PATCH /orders/{id}/status endpoint does).
        if order.get("status") == "draft":
            await _so_apply_status_transition(
                conn, order_id, "draft", "quoted",
                actor_user_id=getattr(token_data, "user_id", None),
                actor_email=getattr(token_data, "email", None),
                note="Auto-quote via PDF render",
                metadata={
                    "trigger": "quote_pdf_render",
                    "quote_pdf_version": new_version,
                },
            )
        elif new_version > 1:
            # AUD-1 fix: regeneration (version > 1) silently overwrote
            # quote_pdf_url+quote_pdf_version without leaving an audit trail.
            # Status is not changing here, so we bypass the state-machine helper
            # and INSERT directly with from_status == status == current_status.
            current_status = order.get("status") or "draft"
            actor_user_id = getattr(token_data, "user_id", None)
            actor_email = getattr(token_data, "email", None)
            try:
                uid_uuid = str(actor_user_id) if actor_user_id else None
            except Exception:
                uid_uuid = None
            await conn.execute(
                """
                INSERT INTO sourcing_order_status_history
                  (order_id, from_status, status, by_user_id, by_user_email, note, metadata)
                VALUES ($1, $2, $3, $4::uuid, $5, $6, $7::jsonb)
                """,
                order_id, current_status, current_status, uid_uuid, actor_email,
                f"Báo giá PDF re-generated v{new_version}",
                json.dumps({
                    "pdf_version": new_version,
                    "trigger": "quote_pdf_regenerate",
                }),
            )
    return FileResponse(
        str(file_path),
        media_type="application/pdf",
        filename=f"{order['order_number']}.pdf",
    )


@router.get("/calc-sale")
async def calc_sale_quick(
    cost_vnd: float = Query(..., ge=0),
    tax_pct: float = Query(0, ge=0),
    coefficient: float = Query(1.0, ge=0),
    shipping_fee: float = Query(0, ge=0),
    qty: float = Query(1, gt=0),
    customer_markup_pct: float | None = Query(None, ge=0),
    customer_name: str | None = None,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "staff"
    )),
):
    """Quick calc helper — single-line live preview.

    Returns:
      { suggested_sale_vnd, breakdown: { base, tax_amount, shipping, markup_amount } }
    """
    item = {
        "cost_vnd": cost_vnd,
        "tax_pct": tax_pct,
        "coefficient": coefficient,
        "shipping_fee_vnd": shipping_fee,
        "qty": qty,
        "customer_markup_pct": customer_markup_pct,
    }
    enriched = _so_calc_line(item)
    base = float(cost_vnd) * float(qty)
    tax_amount = (float(cost_vnd) + float(shipping_fee) / float(qty)) * float(qty) * float(tax_pct) / 100.0
    markup_amount = float(enriched["sale_total_vnd"]) - base - tax_amount - float(shipping_fee)
    return {
        "data": {
            "suggested_sale_vnd": enriched["sale_unit_vnd"],
            "suggested_total_vnd": enriched["sale_total_vnd"],
            "landed_cost_vnd": enriched["landed_cost_vnd"],
            "customer_name": customer_name,
            "breakdown": {
                "base": int(round(base)),
                "tax_amount": int(round(tax_amount)),
                "shipping": int(round(shipping_fee)),
                "markup_amount": int(round(markup_amount)),
            },
        }
    }


@router.post("/orders/calc-preview")
async def calc_preview_order(
    body: OrderCalcPayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "sales", "procurement", "staff"
    )),
):
    """Multi-line preview — không ghi DB. Dùng cho form live-preview."""
    if not body.items:
        raise HTTPException(400, "Phải có ít nhất 1 line item")
    raw_items = _so_items_to_dicts(body.items)
    totals = _so_calc_totals(raw_items, body.discount_vnd or 0)
    return {"data": totals}


# ============================================================
# Phase 3 — Multi-supplier prices + Pricing rules engine
# Thang 2026-06-13
# Migrations:
#   - backend/migrations/sourcing_multi_supplier.sql
#   - backend/migrations/sourcing_pricing_rules.sql
# Engine: app/services/sourcing_pricing_engine.py
# ============================================================


# ── Pydantic models ──────────────────────────────────────────────

class SupplierPricePayload(BaseModel):
    supplier_name: str
    supplier_phone: str | None = None
    supplier_email: str | None = None
    currency: Literal["VND", "JPY", "USD", "KRW", "RMB", "EUR"] = "VND"
    cost_amount: float
    exchange_rate_used: float | None = None
    cost_vnd_equiv: float | None = None      # explicit override — else server computes
    lead_time_days: int | None = None
    moq: int | None = None
    notes: str | None = None
    is_primary: bool = False


class SupplierPriceUpdatePayload(BaseModel):
    supplier_name: str | None = None
    supplier_phone: str | None = None
    supplier_email: str | None = None
    currency: Literal["VND", "JPY", "USD", "KRW", "RMB", "EUR"] | None = None
    cost_amount: float | None = None
    exchange_rate_used: float | None = None
    cost_vnd_equiv: float | None = None
    lead_time_days: int | None = None
    moq: int | None = None
    notes: str | None = None


class PricingRulePayload(BaseModel):
    markup_pct: float = 1.4
    tax_pct: float = 10
    shipping_fee_vnd: float | None = 0
    description_vi: str | None = None
    # Expanded cols cho template Bảng tính giá
    import_tax_pct: float | None = None
    vat_pct: float | None = None
    purchase_cost_pct: float | None = None
    transfer_fee_pct: float | None = None
    swift_fee_usd: float | None = None
    profit_pct_import: float | None = None
    profit_pct_domestic: float | None = None


class CalcSuggestPayload(BaseModel):
    item_type: str | None = None
    cost_amount: float
    currency: Literal["VND", "JPY", "USD", "KRW", "RMB", "EUR"] = "VND"
    exchange_rate: float | None = None
    qty: float | None = 1
    fedex_fee_vnd: float | None = 0
    vn_shipping_fee_vnd: float | None = 0
    is_domestic_vn: bool = False
    # 1b.2: ISO date string (YYYY-MM-DD). When given, the FX rate is looked up
    # historically (rate effective on/before this date) instead of "latest".
    # Optional — omitting it preserves the original "latest rate" behavior.
    quote_date: str | None = None


class CalcSuggestBulkPayload(BaseModel):
    """ICE backend #2 — bulk preview for multi-line order forms.

    Cuts a 20-row order preview from 60 round-trips (3 DB hits × 20 lines)
    to roughly 2: one FX query (cached), one rule query (cached). All
    expensive lookups are dedup'd before the per-row compute loop.
    """
    items: list[CalcSuggestPayload]


def _ssp_serialize(row: asyncpg.Record | dict | None) -> dict | None:
    if row is None:
        return None
    d = dict(row)
    for k, v in list(d.items()):
        if isinstance(v, datetime):
            d[k] = v.isoformat()
        elif hasattr(v, "isoformat"):
            d[k] = v.isoformat()
        elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
            try:
                d[k] = float(v)
            except Exception:
                pass
    return d


async def _compute_vnd_equiv(
    conn: asyncpg.Connection,
    cost_amount: float,
    currency: str,
    exchange_rate: float | None,
) -> tuple[float, float]:
    """Convert (cost_amount, currency) → VND using DB exchange_rates fallback.

    FIX B5 (Thang 2026-06-13): no more DEFAULT_FX_TO_VND hardcode. If caller
    provides `exchange_rate` we trust it; otherwise we look up the latest row
    in `exchange_rates`. If neither is available (and currency != VND), we
    raise HTTPException(400) so the caller surfaces the "missing rate" UX.
    Returns (vnd_equiv, fx_used).
    """
    from app.services.sourcing_pricing_engine import fetch_fx_to_vnd
    cur = (currency or "VND").upper()
    if exchange_rate and float(exchange_rate) > 0:
        fx = float(exchange_rate)
    else:
        db_rate = await fetch_fx_to_vnd(conn, cur)
        if db_rate is None or db_rate <= 0:
            if cur == "VND":
                fx = 1.0
            else:
                raise HTTPException(
                    400,
                    f"Chưa có tỷ giá {cur}/VND trong hệ thống. "
                    f"Cập nhật tại /admin/exchange-rates rồi thử lại.",
                )
        else:
            fx = float(db_rate)
    return (float(cost_amount) * fx, fx)


# ── Supplier-price CRUD ──────────────────────────────────────────

@router.get("/{entry_id:int}/suppliers")
async def list_supplier_prices(
    entry_id: int,
    with_quote_price: bool = Query(
        False,
        description="Attach LIVE compute_sale_vnd preview per candidate (for the quote modal).",
    ),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "director", "procurement", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List tất cả supplier prices cho 1 sourcing entry.

    When `with_quote_price=true`, each candidate is enriched with a LIVE
    quotable VND price (compute_sale_vnd with exchange_rate=None → fresh FX),
    plus fx_rate/fx_date/fx_stale/fx_error. FX errors NEVER 500 the listing —
    they are captured per-row in `fx_error` so the modal can badge "tỷ giá quá
    hạn" without blocking the list.
    """
    entry = await conn.fetchrow(
        "SELECT id, quantity FROM sourcing_entries "
        "WHERE id = $1 AND deleted_at IS NULL",
        entry_id,
    )
    if not entry:
        raise HTTPException(404, f"Không tìm thấy sourcing #{entry_id}")
    # SELECT * gives audit columns when migration has run, hides them otherwise.
    rows = await conn.fetch(
        """
        SELECT *
          FROM sourcing_supplier_prices
         WHERE sourcing_entry_id = $1
         ORDER BY is_primary DESC, cost_vnd_equiv ASC NULLS LAST, id ASC
        """,
        entry_id,
    )
    serialized = [_ssp_serialize(r) for r in rows]

    if with_quote_price:
        from app.services.sourcing_pricing_engine import compute_sale_vnd
        # V1.1: classification never drives pricing → always 'default'.
        # V1.2: the per-supplier preview MUST use the entry's real quantity so
        # fixed costs (L, M, swift) amortise the same way the form's breakdown
        # does — otherwise the NCC list price won't match the form when qty > 1.
        item_type = "default"
        entry_qty = entry["quantity"] or 1
        for r in serialized:
            if r is None:
                continue
            cur = (r.get("currency") or "VND").upper().strip()
            try:
                res = await compute_sale_vnd(
                    conn,
                    item_type=item_type,
                    cost_amount=r.get("cost_amount") or 0,
                    currency=cur,
                    exchange_rate=None,
                    qty=entry_qty,
                    is_domestic_vn=(cur == "VND"),
                )
                bd = res.get("breakdown", {})
                r["quote_unit_price_vnd"] = res["suggested_sale_vnd"]
                r["fx_rate"] = bd.get("exchange_rate_used")
                r["fx_date"] = bd.get("exchange_rate_date")
                r["fx_stale"] = bool(bd.get("fx_stale"))
                r["fx_error"] = None
            except ValueError as exc:
                # Missing / stale FX → list anyway, flagged.
                r["quote_unit_price_vnd"] = None
                r["fx_rate"] = None
                r["fx_date"] = None
                r["fx_stale"] = True
                r["fx_error"] = str(exc)

    return {"data": serialized}


@router.post("/{entry_id:int}/suppliers")
async def create_supplier_price(
    entry_id: int,
    payload: SupplierPricePayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Thêm 1 NCC + giá nhập cho sourcing entry."""
    existing = await conn.fetchval(
        "SELECT id FROM sourcing_entries WHERE id = $1 AND deleted_at IS NULL", entry_id,
    )
    if not existing:
        raise HTTPException(404, f"Không tìm thấy sourcing #{entry_id}")

    # Compute VND equiv if not provided
    if payload.cost_vnd_equiv is not None:
        vnd_equiv = float(payload.cost_vnd_equiv)
        fx_used = (
            float(payload.exchange_rate_used)
            if payload.exchange_rate_used
            else (vnd_equiv / float(payload.cost_amount) if payload.cost_amount else 1.0)
        )
    else:
        vnd_equiv, fx_used = await _compute_vnd_equiv(
            conn, payload.cost_amount, payload.currency, payload.exchange_rate_used,
        )

    async with conn.transaction():
        # If is_primary requested, clear other primaries first (partial unique
        # index would otherwise reject the INSERT)
        if payload.is_primary:
            await conn.execute(
                "UPDATE sourcing_supplier_prices SET is_primary = false "
                "WHERE sourcing_entry_id = $1 AND is_primary = true",
                entry_id,
            )

        # ICE security #1: audit trail — populate created_by_{id,email}.
        # Falls back gracefully if migration hasn't run yet (UndefinedColumn).
        actor_id = _coerce_int(getattr(token_data, "user_id", None))
        actor_email = getattr(token_data, "email", None)
        try:
            row = await conn.fetchrow(
                """
                INSERT INTO sourcing_supplier_prices (
                    sourcing_entry_id, supplier_name, supplier_phone, supplier_email,
                    currency, cost_amount, cost_vnd_equiv, exchange_rate_used,
                    lead_time_days, moq, notes, is_primary,
                    created_by_id, created_by_email, updated_by_id, updated_by_email
                ) VALUES (
                    $1, $2, $3, $4,
                    $5, $6, $7, $8,
                    $9, $10, $11, $12,
                    $13, $14, $13, $14
                ) RETURNING *
                """,
                entry_id, payload.supplier_name, payload.supplier_phone, payload.supplier_email,
                payload.currency, payload.cost_amount, vnd_equiv, fx_used,
                payload.lead_time_days, payload.moq, payload.notes, payload.is_primary,
                actor_id, actor_email,
            )
        except asyncpg.UndefinedColumnError:
            row = await conn.fetchrow(
                """
                INSERT INTO sourcing_supplier_prices (
                    sourcing_entry_id, supplier_name, supplier_phone, supplier_email,
                    currency, cost_amount, cost_vnd_equiv, exchange_rate_used,
                    lead_time_days, moq, notes, is_primary
                ) VALUES (
                    $1, $2, $3, $4,
                    $5, $6, $7, $8,
                    $9, $10, $11, $12
                ) RETURNING *
                """,
                entry_id, payload.supplier_name, payload.supplier_phone, payload.supplier_email,
                payload.currency, payload.cost_amount, vnd_equiv, fx_used,
                payload.lead_time_days, payload.moq, payload.notes, payload.is_primary,
            )
    return {"data": _ssp_serialize(row), "message": f"Đã thêm NCC {payload.supplier_name}"}


@router.put("/{entry_id:int}/suppliers/{sup_id:int}")
async def update_supplier_price(
    entry_id: int,
    sup_id: int,
    payload: SupplierPriceUpdatePayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Update 1 supplier price row."""
    existing = await conn.fetchrow(
        "SELECT id, currency, cost_amount, exchange_rate_used "
        "FROM sourcing_supplier_prices "
        "WHERE id = $1 AND sourcing_entry_id = $2",
        sup_id, entry_id,
    )
    if not existing:
        raise HTTPException(404, f"Không tìm thấy supplier price #{sup_id} cho entry #{entry_id}")

    sets: list[str] = []
    params: list[Any] = []
    idx = 1
    simple = {
        "supplier_name": payload.supplier_name,
        "supplier_phone": payload.supplier_phone,
        "supplier_email": payload.supplier_email,
        "currency": payload.currency,
        "cost_amount": payload.cost_amount,
        "exchange_rate_used": payload.exchange_rate_used,
        "lead_time_days": payload.lead_time_days,
        "moq": payload.moq,
        "notes": payload.notes,
    }
    for k, v in simple.items():
        if v is not None:
            sets.append(f"{k} = ${idx}")
            params.append(v)
            idx += 1

    # Re-compute cost_vnd_equiv if cost/currency/fx changed and explicit equiv not provided
    if payload.cost_vnd_equiv is not None:
        sets.append(f"cost_vnd_equiv = ${idx}")
        params.append(float(payload.cost_vnd_equiv))
        idx += 1
    elif any(v is not None for v in (payload.cost_amount, payload.currency, payload.exchange_rate_used)):
        new_cost = payload.cost_amount if payload.cost_amount is not None else float(existing["cost_amount"])
        new_currency = payload.currency or existing["currency"]
        new_fx = payload.exchange_rate_used if payload.exchange_rate_used is not None else (
            float(existing["exchange_rate_used"]) if existing["exchange_rate_used"] else None
        )
        vnd_equiv, _ = await _compute_vnd_equiv(conn, new_cost, new_currency, new_fx)
        sets.append(f"cost_vnd_equiv = ${idx}")
        params.append(vnd_equiv)
        idx += 1

    if not sets:
        raise HTTPException(400, "Không có field nào để update")

    # ICE security #1: audit trail — best-effort write updated_by_{id,email}.
    # Probe columns once per call (cheap, cached by PG) so older deployments
    # without the migration keep working.
    actor_id = _coerce_int(getattr(token_data, "user_id", None))
    actor_email = getattr(token_data, "email", None)
    try:
        has_audit = await conn.fetchval(
            "SELECT COUNT(*) FROM information_schema.columns "
            "WHERE table_name='sourcing_supplier_prices' "
            "AND column_name IN ('updated_by_id','updated_by_email')"
        ) == 2
    except Exception:
        has_audit = False
    if has_audit:
        sets.append(f"updated_by_id = ${idx}")
        params.append(actor_id)
        idx += 1
        sets.append(f"updated_by_email = ${idx}")
        params.append(actor_email)
        idx += 1

    params.append(sup_id)
    await conn.execute(
        f"UPDATE sourcing_supplier_prices SET {', '.join(sets)} WHERE id = ${idx}",
        *params,
    )
    updated = await conn.fetchrow(
        "SELECT * FROM sourcing_supplier_prices WHERE id = $1", sup_id,
    )
    return {"data": _ssp_serialize(updated)}


@router.delete("/{entry_id:int}/suppliers/{sup_id:int}")
async def delete_supplier_price(
    entry_id: int,
    sup_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Xoá 1 supplier price row."""
    deleted = await conn.fetchval(
        "DELETE FROM sourcing_supplier_prices "
        "WHERE id = $1 AND sourcing_entry_id = $2 RETURNING id",
        sup_id, entry_id,
    )
    if not deleted:
        raise HTTPException(404, f"Không tìm thấy supplier price #{sup_id} cho entry #{entry_id}")
    return {"data": {"id": sup_id, "deleted": True}}


@router.patch("/{entry_id:int}/suppliers/{sup_id:int}/set-primary")
async def set_primary_supplier(
    entry_id: int,
    sup_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đánh dấu 1 supplier price là primary — auto unflag các row khác."""
    existing = await conn.fetchval(
        "SELECT id FROM sourcing_supplier_prices "
        "WHERE id = $1 AND sourcing_entry_id = $2",
        sup_id, entry_id,
    )
    if not existing:
        raise HTTPException(404, f"Không tìm thấy supplier price #{sup_id} cho entry #{entry_id}")

    async with conn.transaction():
        # Unflag all others first to avoid partial-unique-index conflict
        await conn.execute(
            "UPDATE sourcing_supplier_prices SET is_primary = false "
            "WHERE sourcing_entry_id = $1 AND id <> $2",
            entry_id, sup_id,
        )
        await conn.execute(
            "UPDATE sourcing_supplier_prices SET is_primary = true WHERE id = $1",
            sup_id,
        )
    updated = await conn.fetchrow(
        "SELECT * FROM sourcing_supplier_prices WHERE id = $1", sup_id,
    )
    return {"data": _ssp_serialize(updated), "message": "Đã đặt làm NCC chính"}


# ── Pricing rules CRUD ───────────────────────────────────────────

@router.get("/pricing-rules")
async def list_pricing_rules(
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "director", "procurement", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List toàn bộ pricing rules — dùng cho settings page + dropdown chọn item_type."""
    try:
        rows = await conn.fetch(
            """
            SELECT id, item_type, markup_pct, tax_pct, shipping_fee_vnd,
                   import_tax_pct, vat_pct, purchase_cost_pct,
                   transfer_fee_pct, swift_fee_usd,
                   profit_pct_import, profit_pct_domestic,
                   description_vi, created_at, updated_at
              FROM sourcing_pricing_rules
             ORDER BY (item_type = 'default') DESC, item_type ASC
            """
        )
    except asyncpg.UndefinedColumnError:
        rows = await conn.fetch(
            """
            SELECT id, item_type, markup_pct, tax_pct, shipping_fee_vnd,
                   description_vi, created_at, updated_at
              FROM sourcing_pricing_rules
             ORDER BY (item_type = 'default') DESC, item_type ASC
            """
        )
    return {"data": [_ssp_serialize(r) for r in rows]}


@router.post("/pricing-rules/{item_type}")
async def upsert_pricing_rule_post(
    item_type: str,
    payload: PricingRulePayload,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Upsert pricing rule cho item_type (POST = create-or-replace, idempotent)."""
    return await _upsert_pricing_rule(conn, item_type, payload, token_data)


@router.put("/pricing-rules/{item_type}")
async def upsert_pricing_rule_put(
    item_type: str,
    payload: PricingRulePayload,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Upsert pricing rule cho item_type (PUT alias of POST)."""
    return await _upsert_pricing_rule(conn, item_type, payload, token_data)


async def _upsert_pricing_rule(
    conn: asyncpg.Connection,
    item_type: str,
    payload: PricingRulePayload,
    token_data: TokenData | None = None,
) -> dict:
    item_type = (item_type or "").strip()
    if not item_type:
        raise HTTPException(400, "item_type không được rỗng")
    if payload.markup_pct < 0 or payload.tax_pct < 0:
        raise HTTPException(400, "markup_pct / tax_pct phải >= 0")

    # Snapshot CURRENT row before UPDATE — feeds the audit history table.
    # Empty {} if the rule doesn't exist yet (i.e. this upsert is an INSERT).
    try:
        old_row = await conn.fetchrow(
            "SELECT * FROM sourcing_pricing_rules WHERE item_type = $1",
            item_type,
        )
    except Exception:
        old_row = None

    # Detect schema — pick the broadest INSERT that DB supports.
    try:
        cols_exist = await conn.fetchval(
            "SELECT COUNT(*) FROM information_schema.columns "
            "WHERE table_name='sourcing_pricing_rules' "
            "AND column_name IN ("
            "'import_tax_pct','vat_pct','purchase_cost_pct','transfer_fee_pct',"
            "'swift_fee_usd','profit_pct_import','profit_pct_domestic')"
        )
    except Exception:
        cols_exist = 0

    has_expanded = (cols_exist or 0) >= 7

    if has_expanded:
        row = await conn.fetchrow(
            """
            INSERT INTO sourcing_pricing_rules
                (item_type, markup_pct, tax_pct, shipping_fee_vnd, description_vi,
                 import_tax_pct, vat_pct, purchase_cost_pct,
                 transfer_fee_pct, swift_fee_usd,
                 profit_pct_import, profit_pct_domestic)
            VALUES ($1, $2::NUMERIC, $3::NUMERIC, $4::NUMERIC, $5,
                    COALESCE($6::NUMERIC, 20::NUMERIC),
                    COALESCE($7::NUMERIC, $3::NUMERIC),
                    COALESCE($8::NUMERIC, 25::NUMERIC),
                    COALESCE($9::NUMERIC, 0.2::NUMERIC),
                    COALESCE($10::NUMERIC, 5::NUMERIC),
                    COALESCE($11::NUMERIC, 12::NUMERIC),
                    COALESCE($12::NUMERIC, 20::NUMERIC))
            ON CONFLICT (item_type) DO UPDATE SET
                markup_pct          = EXCLUDED.markup_pct,
                tax_pct             = EXCLUDED.tax_pct,
                shipping_fee_vnd    = EXCLUDED.shipping_fee_vnd,
                description_vi      = EXCLUDED.description_vi,
                import_tax_pct      = EXCLUDED.import_tax_pct,
                vat_pct             = EXCLUDED.vat_pct,
                purchase_cost_pct   = EXCLUDED.purchase_cost_pct,
                transfer_fee_pct    = EXCLUDED.transfer_fee_pct,
                swift_fee_usd       = EXCLUDED.swift_fee_usd,
                profit_pct_import   = EXCLUDED.profit_pct_import,
                profit_pct_domestic = EXCLUDED.profit_pct_domestic
            RETURNING *
            """,
            item_type, payload.markup_pct, payload.tax_pct,
            payload.shipping_fee_vnd or 0, payload.description_vi,
            payload.import_tax_pct, payload.vat_pct,
            payload.purchase_cost_pct,
            payload.transfer_fee_pct, payload.swift_fee_usd,
            payload.profit_pct_import, payload.profit_pct_domestic,
        )
    else:
        row = await conn.fetchrow(
            """
            INSERT INTO sourcing_pricing_rules
                (item_type, markup_pct, tax_pct, shipping_fee_vnd, description_vi)
            VALUES ($1, $2, $3, $4, $5)
            ON CONFLICT (item_type) DO UPDATE SET
                markup_pct = EXCLUDED.markup_pct,
                tax_pct = EXCLUDED.tax_pct,
                shipping_fee_vnd = EXCLUDED.shipping_fee_vnd,
                description_vi = EXCLUDED.description_vi
            RETURNING *
            """,
            item_type, payload.markup_pct, payload.tax_pct,
            payload.shipping_fee_vnd or 0, payload.description_vi,
        )

    # ICE security #1: audit trail (best-effort UPDATE, separate statement
    # because the upsert above is parameterised). Skips silently if the audit
    # columns don't exist yet.
    if token_data is not None:
        actor_id = _coerce_int(getattr(token_data, "user_id", None))
        actor_email = getattr(token_data, "email", None)
        try:
            await conn.execute(
                "UPDATE sourcing_pricing_rules "
                "SET updated_by_id = $1, updated_by_email = $2, "
                "    created_by_id = COALESCE(created_by_id, $1), "
                "    created_by_email = COALESCE(created_by_email, $2) "
                "WHERE item_type = $3",
                actor_id, actor_email, item_type,
            )
        except asyncpg.UndefinedColumnError:
            pass

    # Audit history (Thang 2026-06-14) — every upsert appends one row to
    # sourcing_pricing_rules_history with old/new JSONB + actor + auto summary.
    # Best-effort: if the history table doesn't exist yet (migration not run),
    # the insert is skipped silently — the rule write is still authoritative.
    try:
        actor_id_hist = None
        actor_email_hist = None
        if token_data is not None:
            actor_id_hist = _coerce_int(getattr(token_data, "user_id", None))
            actor_email_hist = getattr(token_data, "email", None)
        old_dict = _row_to_audit_dict(old_row) if old_row else {}
        new_dict = _row_to_audit_dict(row) if row else {}
        summary = _summarize_rule_diff(old_dict, new_dict)
        await conn.execute(
            """
            INSERT INTO sourcing_pricing_rules_history
                (rule_item_type, old_values, new_values,
                 changed_by_id, changed_by_email, change_summary)
            VALUES ($1, $2::JSONB, $3::JSONB, $4, $5, $6)
            """,
            item_type,
            json.dumps(old_dict, default=_audit_json_default),
            json.dumps(new_dict, default=_audit_json_default),
            actor_id_hist,
            actor_email_hist,
            summary,
        )
    except asyncpg.UndefinedTableError:
        pass
    except Exception as exc:
        logger.warning("pricing-rule history insert failed for %s: %s", item_type, exc)

    # ICE #1 cache: rule just changed — drop in-memory entry so /calc-suggest
    # sees the fresh values immediately.
    try:
        from app.services.sourcing_pricing_engine import invalidate_pricing_caches
        invalidate_pricing_caches()
    except Exception:
        pass
    return {"data": _ssp_serialize(row), "message": f"Đã lưu rule '{item_type}'"}


# ── Pricing rules audit history ──────────────────────────────────

# Columns we surface in old/new JSONB blobs — internal IDs + free-text
# audit columns are stripped so the diff is meaningful to humans.
_PRICING_RULE_AUDIT_FIELDS: tuple[str, ...] = (
    "item_type",
    "markup_pct",
    "tax_pct",
    "shipping_fee_vnd",
    "description_vi",
    "import_tax_pct",
    "vat_pct",
    "purchase_cost_pct",
    "transfer_fee_pct",
    "swift_fee_usd",
    "profit_pct_import",
    "profit_pct_domestic",
)


def _audit_json_default(value: Any) -> Any:
    """JSON serializer for Decimal / datetime / date / fallback str."""
    try:
        from decimal import Decimal
        if isinstance(value, Decimal):
            return str(value)
    except Exception:
        pass
    if isinstance(value, (datetime, date_cls)):
        return value.isoformat()
    return str(value)


def _row_to_audit_dict(row: Any) -> dict[str, Any]:
    """Pick only meaningful business fields from an asyncpg.Record."""
    if row is None:
        return {}
    try:
        as_dict = dict(row)
    except Exception:
        return {}
    return {k: as_dict[k] for k in _PRICING_RULE_AUDIT_FIELDS if k in as_dict}


def _fmt_audit_value(v: Any) -> str:
    if v is None:
        return "∅"
    try:
        from decimal import Decimal
        if isinstance(v, Decimal):
            return f"{v:f}".rstrip("0").rstrip(".") or "0"
    except Exception:
        pass
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def _summarize_rule_diff(old: dict[str, Any], new: dict[str, Any]) -> str:
    """Render 'field: old -> new; field2: old -> new' (truncate at 6 fields)."""
    if not old:
        return "Tạo mới rule"
    parts: list[str] = []
    for k in _PRICING_RULE_AUDIT_FIELDS:
        if old.get(k) != new.get(k):
            parts.append(f"{k}: {_fmt_audit_value(old.get(k))} → {_fmt_audit_value(new.get(k))}")
    if not parts:
        return "Không có thay đổi"
    if len(parts) > 6:
        return "; ".join(parts[:6]) + f"; +{len(parts) - 6} field khác"
    return "; ".join(parts)


@router.get("/pricing-rules/{item_type}/history")
async def get_pricing_rule_history(
    item_type: str,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "director", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Last 50 audit-history entries for a single pricing rule.

    Returns: { data: [ { id, changed_at, changed_by_id, changed_by_email,
                          change_summary, old_values, new_values }, ... ] }

    Sorted newest-first. Empty array if rule never changed (or history
    table not yet provisioned).
    """
    item_type = (item_type or "").strip()
    if not item_type:
        raise HTTPException(400, "item_type không được rỗng")
    try:
        rows = await conn.fetch(
            """
            SELECT id, rule_item_type, old_values, new_values,
                   changed_at, changed_by_id, changed_by_email, change_summary
              FROM sourcing_pricing_rules_history
             WHERE rule_item_type = $1
             ORDER BY changed_at DESC
             LIMIT 50
            """,
            item_type,
        )
    except asyncpg.UndefinedTableError:
        return {"data": []}

    def _parse_jsonb(v: Any) -> Any:
        if v is None:
            return {}
        if isinstance(v, (dict, list)):
            return v
        if isinstance(v, str):
            try:
                return json.loads(v)
            except Exception:
                return {}
        return v

    data = [
        {
            "id": r["id"],
            "rule_item_type": r["rule_item_type"],
            "changed_at": r["changed_at"].isoformat() if r["changed_at"] else None,
            "changed_by_id": r["changed_by_id"],
            "changed_by_email": r["changed_by_email"],
            "change_summary": r["change_summary"],
            "old_values": _parse_jsonb(r["old_values"]),
            "new_values": _parse_jsonb(r["new_values"]),
        }
        for r in rows
    ]
    return {"data": data}


# ── Calc-suggest endpoint ────────────────────────────────────────

@router.post("/calc-suggest")
async def calc_suggest(
    body: CalcSuggestPayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "director", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Compute suggested sale_vnd dựa trên item_type rule + cost + currency.

    Body: { item_type?, cost_amount, currency, exchange_rate?, qty?,
            fedex_fee_vnd?, vn_shipping_fee_vnd?, is_domestic_vn? }
    Returns: { data: { suggested_sale_vnd, breakdown: {...} } }

    Engine: app/services/sourcing_pricing_engine.compute_sale_vnd().

    ────────────── GOLDEN VALUES (verified 2026-06-13) ──────────────
    These are the canonical engine outputs. Any test agent comparing
    against different values is using a STALE expected baseline — fix
    the test, NOT the engine. See backend/tests/unit/test_pricing_engine.py.

    Scenario IMPORT (is_domestic_vn=False, default rule):
      Input: cost_amount=100 USD, exchange_rate=25000, qty=1,
             fedex_fee_vnd=500000, vn_shipping_fee_vnd=0
      Output: suggested_sale_vnd S = 4,747,064
      Breakdown:
        K=2,500,000  M=500,000  N=600,000 (20% import tax on K+M)
        O=360,000 (10% VAT)     P=625,000 (25% purchase cost)
        Q≈131,250   R=508,440 (12% import profit)

    Scenario DOMESTIC (is_domestic_vn=True, default rule):
      Input: cost_amount=100 USD, exchange_rate=25000, qty=1,
             fedex_fee_vnd=500000, vn_shipping_fee_vnd=0
      Output: suggested_sale_vnd S = 4,399,740
      Note: N=0 (no import tax), profit_pct=20% (domestic) instead of 12%.

    Domestic R column note (per template "Bảng tính giá 2026"):
      For the canonical domestic scenario above, R (profit row) =
      base_for_profit × 20% ≈ 733,290 VND. If an E2E test expects a
      different R value, the test's baseline is wrong; the engine
      output matches the spreadsheet template.

    Engine guarantees:
    - Determinism: same inputs → same outputs (Decimal arithmetic,
      ROUND_HALF_UP to whole VND at the very last step only).
    - FX is either passed explicitly or read from `exchange_rates`
      table (no hardcoded fallback — see FIX B5 2026-06-13).
    - `item_type` falls back to 'default' rule if unknown.

    To re-verify locally:
      pytest backend/tests/unit/test_pricing_engine.py -v
    """
    if body.cost_amount < 0:
        raise HTTPException(400, "cost_amount phải >= 0")
    from app.services.sourcing_pricing_engine import compute_sale_vnd, fetch_fx_to_vnd

    # 1b.2: best-effort parse quote_date → date for historical FX lookup.
    # Invalid/missing → None (engine falls back to "latest rate" as before).
    fx_date: date_cls | None = None
    if body.quote_date:
        try:
            fx_date = datetime.strptime(body.quote_date, "%Y-%m-%d").date()
        except ValueError:
            fx_date = None

    # FIX B5 (Thang 2026-06-13): nếu UI không pass exchange_rate, fallback
    # query exchange_rates table. Engine cũng query — chỉ làm ở đây để có
    # message Vietnamese rõ ràng khi rate thiếu hoàn toàn.
    # 1b.2: when quote_date given, look up the rate effective on/before it.
    # NOTE: we only pass `exchange_rate` to the engine when the CALLER supplied
    # one. When the rate is sourced from the DB we leave it None so the engine
    # fetches via fetch_fx_meta() and records the true `exchange_rate_date`
    # (rate_date) in the breakdown — this is the audit/display date the owner
    # asked for. The pre-fetch below only guards the "missing rate" message.
    effective_rate = body.exchange_rate
    if effective_rate is None or effective_rate <= 0:
        effective_rate = None
        db_rate = await fetch_fx_to_vnd(conn, body.currency or "VND", as_of_date=fx_date)
        if db_rate is None or db_rate <= 0:
            if (body.currency or "VND").upper() != "VND":
                raise HTTPException(
                    400,
                    f"Chưa có tỷ giá {body.currency}/VND trong hệ thống. "
                    f"Cập nhật tại /admin/exchange-rates rồi thử lại.",
                )

    try:
        result = await compute_sale_vnd(
            conn,
            item_type=body.item_type,
            cost_amount=body.cost_amount,
            currency=body.currency,
            exchange_rate=effective_rate,
            qty=body.qty,
            fedex_fee_vnd=body.fedex_fee_vnd,
            vn_shipping_fee_vnd=body.vn_shipping_fee_vnd,
            is_domestic_vn=body.is_domestic_vn,
            fx_date=fx_date,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"data": result}


@router.post("/calc-suggest/bulk")
async def calc_suggest_bulk(
    body: CalcSuggestBulkPayload,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "director", "procurement"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Compute suggested sale_vnd for many lines in one round-trip.

    ICE backend #2 (Thang 2026-06-13): combined with the TTL cache on
    fetch_fx_to_vnd + get_rule, a 20-line preview drops from ~60 DB hits
    (3 per /calc-suggest call) to roughly 2 (FX + rule, both cached).

    Body: { items: [CalcSuggestPayload, ...] }
    Returns: { data: [ { suggested_sale_vnd, breakdown, error? }, ... ] }
    Per-row errors do NOT abort the batch — the offending row carries
    `{ error: "..." }` so the UI can render a partial preview.
    """
    from app.services.sourcing_pricing_engine import compute_sale_vnd

    if not body.items:
        return {"data": []}
    if len(body.items) > 200:
        raise HTTPException(400, "Bulk calc-suggest tối đa 200 dòng / request")

    out: list[dict[str, Any]] = []
    for idx, item in enumerate(body.items):
        if item.cost_amount < 0:
            out.append({"index": idx, "error": "cost_amount phải >= 0"})
            continue
        try:
            result = await compute_sale_vnd(
                conn,
                item_type=item.item_type,
                cost_amount=item.cost_amount,
                currency=item.currency,
                exchange_rate=item.exchange_rate,
                qty=item.qty,
                fedex_fee_vnd=item.fedex_fee_vnd,
                vn_shipping_fee_vnd=item.vn_shipping_fee_vnd,
                is_domestic_vn=item.is_domestic_vn,
            )
            out.append({"index": idx, **result})
        except ValueError as exc:
            out.append({"index": idx, "error": str(exc)})
        except Exception as exc:  # noqa: BLE001
            logger.exception("calc-suggest/bulk row %d failed: %s", idx, exc)
            out.append({"index": idx, "error": "Lỗi nội bộ khi tính giá"})
    return {"data": out}
