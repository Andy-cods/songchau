"""
Analytics Exports API — generic CSV/XLSX/PNG export for any analytics panel.

Endpoint (mount under /api/v1/analytics):
  POST /exports — Serialize any panel (xnk / trends / forecast) to file.

Design notes (Thang 2026-06-04):
  - Avoid HTTP loopback: re-call handler functions from xnk_analytics.py /
    analytics_trends.py / forecast.py directly (import + invoke).
  - For PNG export the frontend renders the canvas to a base64 dataURL
    (html2canvas) and sends it; backend only decodes + streams as file.
  - CSV uses UTF-8 BOM so Excel auto-detects encoding for Vietnamese chars.
  - XLSX uses openpyxl (already in deps), with bold header row + autosize.
  - Files written to /data/files/exports/ then streamed via FileResponse and
    background-cleaned to avoid disk bloat.
"""

from __future__ import annotations

import base64
import csv
import io
import logging
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Awaitable

import asyncpg
import openpyxl
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.api.v1 import analytics_trends, forecast, xnk_analytics
from app.core.database import get_db
from app.core.rbac import require_role
from app.core.security import TokenData

logger = logging.getLogger(__name__)

router = APIRouter()

EXPORTS_DIR = Path("/data/files/exports")
try:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
except (OSError, PermissionError):
    # Fall back to a tmp dir on dev environments (Windows/macOS) where
    # /data/files/exports/ can't be created.
    EXPORTS_DIR = Path.cwd() / "tmp" / "analytics_exports"
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Request schema
# ---------------------------------------------------------------------------


class ExportRequest(BaseModel):
    scope: str = Field(..., description='"xnk" | "trends" | "forecast"')
    panel: str = Field(..., description='Panel key, e.g. "kpi", "monthly_trend", "volatility"')
    format: str = Field(..., description='"csv" | "xlsx" | "png"')
    filters: dict[str, Any] = Field(default_factory=dict)
    chart_dataurl: str | None = Field(
        default=None,
        description="Base64 data URL of rendered chart (only used when format='png').",
    )


# ---------------------------------------------------------------------------
# Panel registry — (scope, panel) → (handler_fn, default_label, row_extractor)
# ---------------------------------------------------------------------------
#
# row_extractor takes the handler's returned dict and produces a list[dict]
# (each dict is one row) so CSV/XLSX writers can be agnostic of the panel
# shape. Returning empty list is legal (writers emit just a header + 1 empty
# row noting "no data").


def _extract_kpi_xnk(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    # KPI is a single-row "metric → value" projection
    return [{"metric": k, "value": v} for k, v in d.items()]


def _extract_hs_distribution(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    return [
        {
            "lower": b.get("lower"),
            "upper": b.get("upper"),
            "count": b.get("count"),
        }
        for b in (d.get("bins") or [])
    ]


def _extract_monthly_trend(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list((payload.get("data") or {}).get("months") or [])


def _extract_top_sellers(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list(payload.get("data") or [])


def _extract_trends_kpi(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    return [{"metric": k, "value": v} for k, v in d.items()]


def _extract_multi_series(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    out: list[dict[str, Any]] = []
    for s in d.get("series") or []:
        code = s.get("bqms_code")
        for p in s.get("points") or []:
            out.append({
                "bqms_code": code,
                "ym": p.get("ym"),
                "median_v1_vnd": p.get("median_v1_vnd"),
                "n": p.get("n"),
            })
    return out


def _extract_by_customer(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    out: list[dict[str, Any]] = []
    for c in d.get("customers") or []:
        name = c.get("buyer_name")
        for p in c.get("points") or []:
            out.append({
                "buyer_name": name,
                "ym": p.get("ym"),
                "avg_unit_price": p.get("avg_unit_price"),
                "qty": p.get("qty"),
            })
    return out


def _extract_by_supplier(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    out: list[dict[str, Any]] = []
    for s in d.get("suppliers") or []:
        name = s.get("supplier_name")
        for p in s.get("points") or []:
            out.append({
                "supplier_name": name,
                "ym": p.get("ym"),
                "avg_cost_vnd": p.get("avg_cost_vnd"),
                "qty": p.get("qty"),
            })
    return out


def _extract_volatility(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list(payload.get("data") or [])


def _extract_forecast_products(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list((payload.get("data") or {}).get("rows") or [])


def _extract_forecast_kpi(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    flat = {k: v for k, v in d.items() if not isinstance(v, dict)}
    rows = [{"metric": k, "value": v} for k, v in flat.items()]
    top = d.get("top_sku")
    if isinstance(top, dict):
        for k, v in top.items():
            rows.append({"metric": f"top_sku.{k}", "value": v})
    return rows


def _extract_top_predicted(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list((payload.get("data") or {}).get("rows") or [])


def _extract_confidence_distribution(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list((payload.get("data") or {}).get("buckets") or [])


def _extract_funnel(payload: dict[str, Any]) -> list[dict[str, Any]]:
    d = payload.get("data") or {}
    return [{"metric": k, "value": v} for k, v in d.items()]


def _extract_reorder(payload: dict[str, Any]) -> list[dict[str, Any]]:
    return list((payload.get("data") or {}).get("rows") or [])


# Registry: (scope, panel) -> (handler, label, extractor)
PANEL_REGISTRY: dict[tuple[str, str], tuple[Callable[..., Awaitable[Any]], str, Callable[[dict], list[dict]]]] = {
    # XNK Analytics
    ("xnk", "kpi"): (xnk_analytics.xnk_analytics_kpi, "XNK KPI", _extract_kpi_xnk),
    ("xnk", "hs_distribution"): (xnk_analytics.xnk_hs_distribution, "XNK HS Distribution", _extract_hs_distribution),
    ("xnk", "monthly_trend"): (xnk_analytics.xnk_monthly_trend, "XNK Monthly Trend", _extract_monthly_trend),
    ("xnk", "top_sellers"): (xnk_analytics.xnk_top_sellers, "XNK Top Sellers", _extract_top_sellers),
    # Price Trends
    ("trends", "kpi"): (analytics_trends.price_trends_kpi, "Trends KPI", _extract_trends_kpi),
    ("trends", "multi_series"): (analytics_trends.price_trends_multi_series, "Trends Multi-Series", _extract_multi_series),
    ("trends", "by_customer"): (analytics_trends.price_trends_by_customer, "Trends by Customer", _extract_by_customer),
    ("trends", "by_supplier"): (analytics_trends.price_trends_by_supplier, "Trends by Supplier", _extract_by_supplier),
    ("trends", "volatility"): (analytics_trends.price_trends_volatility, "Trends Volatility", _extract_volatility),
    # Forecast
    ("forecast", "products"): (forecast.list_forecast_products, "Forecast Products", _extract_forecast_products),
    ("forecast", "kpi"): (forecast.forecast_kpi, "Forecast KPI", _extract_forecast_kpi),
    ("forecast", "top_predicted"): (forecast.forecast_top_predicted, "Forecast Top Predicted", _extract_top_predicted),
    ("forecast", "confidence_distribution"): (
        forecast.forecast_confidence_distribution, "Forecast Confidence Distribution", _extract_confidence_distribution,
    ),
    ("forecast", "funnel"): (forecast.forecast_funnel, "Forecast Funnel", _extract_funnel),
    ("forecast", "reorder_suggestions"): (forecast.forecast_reorder_suggestions, "Forecast Reorder", _extract_reorder),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


_SAFE_NAME = re.compile(r"[^A-Za-z0-9_.-]+")


def _safe_filename(stem: str, ext: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    clean = _SAFE_NAME.sub("_", stem)[:60].strip("_") or "export"
    return f"{clean}_{ts}_{uuid.uuid4().hex[:6]}.{ext}"


def _cleanup(path: Path) -> None:
    try:
        if path.exists():
            path.unlink()
    except Exception:
        logger.warning("Failed to cleanup export file %s", path, exc_info=True)


def _serialize_cell(value: Any) -> Any:
    """Coerce odd types (datetime, None) for csv/xlsx cells."""
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return str(value)
    return value


def _write_csv(rows: list[dict[str, Any]], path: Path) -> None:
    # UTF-8 BOM so Excel auto-detects encoding for Vietnamese text.
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        if not rows:
            f.write("(no data)\n")
            return
        # Preserve key order across rows: union, first-seen wins.
        fieldnames: list[str] = []
        seen: set[str] = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    fieldnames.append(k)
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow({k: _serialize_cell(r.get(k)) for k in fieldnames})


def _write_xlsx(rows: list[dict[str, Any]], path: Path, sheet_label: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_label or "Sheet1")[:31]  # Excel sheet name limit
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    if not rows:
        ws.cell(row=1, column=1, value="(no data)")
        wb.save(path)
        return

    fieldnames: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)

    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, name in enumerate(fieldnames, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_serialize_cell(row.get(name)))

    # Auto-fit column widths (rough)
    for c_idx, name in enumerate(fieldnames, start=1):
        max_len = len(str(name))
        for row in rows:
            v = row.get(name)
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 60)

    wb.save(path)


def _write_png_from_dataurl(dataurl: str, path: Path) -> None:
    # Accept "data:image/png;base64,XXXX" or raw base64
    if "," in dataurl:
        _, b64 = dataurl.split(",", 1)
    else:
        b64 = dataurl
    try:
        raw = base64.b64decode(b64, validate=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid chart_dataurl: {exc}") from exc
    path.write_bytes(raw)


async def _call_handler(
    handler: Callable[..., Awaitable[Any]],
    filters: dict[str, Any],
    token_data: TokenData,
    conn: asyncpg.Connection,
) -> Any:
    """Invoke a panel handler, passing only kwargs it understands.

    Endpoint handlers use FastAPI Query() defaults — those become plain default
    values when called as regular Python functions, so anything missing in
    `filters` falls back to that default. We deliberately drop unknown keys to
    avoid TypeErrors.
    """
    import inspect

    sig = inspect.signature(handler)
    accepted = set(sig.parameters.keys())
    kwargs: dict[str, Any] = {}
    for k, v in (filters or {}).items():
        if k in accepted:
            kwargs[k] = v
    if "token_data" in accepted:
        kwargs["token_data"] = token_data
    if "conn" in accepted:
        kwargs["conn"] = conn
    try:
        return await handler(**kwargs)
    except HTTPException:
        raise
    except TypeError as exc:
        raise HTTPException(
            status_code=400, detail=f"Invalid filters for panel: {exc}"
        ) from exc


# ---------------------------------------------------------------------------
# POST /exports
# ---------------------------------------------------------------------------


@router.post("/exports")
async def create_export(
    body: ExportRequest,
    background: BackgroundTasks,
    token_data: TokenData = Depends(
        # allow_viewer=False (defense-in-depth 02/07): /exports gọi trực tiếp handler
        # panel trends/* (giá nội bộ) nên phải chặn viewer TẠI CHỖ, không chỉ dựa vào
        # việc đây là POST (viewer đã bị chặn non-GET). Phòng khi thêm biến thể GET.
        require_role("admin", "manager", "staff", "sales", "director", "procurement", allow_viewer=False)
    ),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Serialize a panel to CSV/XLSX/PNG and stream it back as an attachment."""
    fmt = body.format.lower().strip()
    if fmt not in {"csv", "xlsx", "png"}:
        raise HTTPException(status_code=400, detail="format must be csv|xlsx|png")

    key = (body.scope.lower().strip(), body.panel.lower().strip())
    entry = PANEL_REGISTRY.get(key)
    if entry is None:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown (scope, panel) = {key}. Known panels: "
            + ", ".join(f"{s}/{p}" for s, p in sorted(PANEL_REGISTRY.keys())),
        )
    handler, label, extractor = entry

    stem = f"{key[0]}_{key[1]}"

    # PNG: skip data fetch, decode dataURL straight from request body.
    if fmt == "png":
        if not body.chart_dataurl:
            raise HTTPException(
                status_code=400, detail="chart_dataurl required for png export"
            )
        out_path = EXPORTS_DIR / _safe_filename(stem, "png")
        _write_png_from_dataurl(body.chart_dataurl, out_path)
        background.add_task(_cleanup, out_path)
        return FileResponse(
            out_path,
            media_type="image/png",
            filename=out_path.name,
            headers={"Content-Disposition": f'attachment; filename="{out_path.name}"'},
        )

    # CSV / XLSX: re-invoke handler in-process (no HTTP loopback).
    payload = await _call_handler(handler, body.filters or {}, token_data, conn)
    if not isinstance(payload, dict):
        # Handlers return dicts; coerce to be safe.
        payload = {"data": payload}
    rows = extractor(payload) or []

    if fmt == "csv":
        out_path = EXPORTS_DIR / _safe_filename(stem, "csv")
        _write_csv(rows, out_path)
        background.add_task(_cleanup, out_path)
        return FileResponse(
            out_path,
            media_type="text/csv; charset=utf-8",
            filename=out_path.name,
            headers={"Content-Disposition": f'attachment; filename="{out_path.name}"'},
        )

    # xlsx
    out_path = EXPORTS_DIR / _safe_filename(stem, "xlsx")
    _write_xlsx(rows, out_path, label)
    background.add_task(_cleanup, out_path)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_path.name,
        headers={"Content-Disposition": f'attachment; filename="{out_path.name}"'},
    )
