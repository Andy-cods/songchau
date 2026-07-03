"""Samsung BQMS API — KPI, records, RFQ parsing, quotation generation, sync, deliveries."""

# NOTE: `from __future__ import annotations` removed (Thang 2026-06-14) —
# breaks @limiter.limit (slowapi 0.1.9) Pydantic forward-ref resolution for
# QuoteBatchRequest + PushToSecRequest bodies. Python 3.12 has native PEP 604
# union syntax so the future-annotations import was only cosmetic.

import io
import json as _json
import logging
import math
import time

from pydantic import BaseModel, ConfigDict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, Query, Request, Response, UploadFile
from fastapi.exceptions import HTTPException
from fastapi.responses import StreamingResponse
import asyncpg

from app.core.database import get_db
from app.core.rbac import require_role
from app.core.security import TokenData
from app.core.slowapi_limiter import limiter
from app.services.bqms_service import BQMSService

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# BQMS user-edit guard (Thang 2026-05-15): when `bqms_user_edit_disabled` flag
# is true in app_config, ALL endpoints that mutate BQMS data return 403.
# Use as a single-line guard at top of each edit endpoint:
#     await _assert_bqms_edit_enabled(conn)
# Read-only endpoints are NOT affected.
# ─────────────────────────────────────────────────────────────────────────────

_BQMS_EDIT_DISABLED_MSG = (
    "BQMS user editing is currently disabled. Data is sourced from Samsung "
    "scrape only. Toggle app_config.bqms_user_edit_disabled=false to re-enable."
)


async def _assert_bqms_edit_enabled(conn: asyncpg.Connection) -> None:
    """Raise 403 nếu BQMS user-edit bị tắt qua app_config flag."""
    try:
        val = await conn.fetchval(
            "SELECT value FROM app_config WHERE key='bqms_user_edit_disabled'"
        )
    except Exception:
        return  # If app_config table or row missing, allow (fail-open)
    # jsonb value can be Python bool/str/dict — normalize to bool
    disabled = False
    if isinstance(val, bool):
        disabled = val
    elif isinstance(val, str):
        disabled = val.strip().strip('"').lower() in ("true", "1", "yes")
    if disabled:
        raise HTTPException(status_code=403, detail=_BQMS_EDIT_DISABLED_MSG)
router = APIRouter()

# Shared service instance
_bqms_service = BQMSService()


# ---------------------------------------------------------------------------
# Sync — Samsung API
# ---------------------------------------------------------------------------

# In-memory job tracking (lightweight; use Redis for production multi-worker)
_sync_jobs: dict[int, dict] = {}


@router.post("/sync")
async def trigger_sync(
    background_tasks: BackgroundTasks,
    date_from: date = Query(..., description="Ngày bắt đầu (YYYY-MM-DD)"),
    date_to: date = Query(..., description="Ngày kết thúc (YYYY-MM-DD)"),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """
    Trigger a sync of Samsung BQMS PO list.

    The sync runs in the background. Use GET /bqms/sync/status/{job_id} to check progress.
    """
    if date_from > date_to:
        raise HTTPException(
            status_code=400,
            detail="date_from phải nhỏ hơn hoặc bằng date_to",
        )

    # Quick validation: range should not exceed 365 days
    if (date_to - date_from).days > 365:
        raise HTTPException(
            status_code=400,
            detail="Khoảng thời gian tối đa là 365 ngày",
        )

    # Create a sync log entry first (synchronously to return the ID)
    sync_id = await conn.fetchval(
        """
        INSERT INTO etl_sync_log (sync_type, source_file, status)
        VALUES ('bqms_po', $1, 'queued')
        RETURNING id
        """,
        f"Samsung API {date_from} - {date_to}",
    )

    # Check for already-running sync
    running = await conn.fetchval(
        "SELECT id FROM etl_sync_log WHERE sync_type = 'bqms_po' AND status = 'running' LIMIT 1"
    )
    if running:
        raise HTTPException(400, f"Đồng bộ đang chạy (job #{running}). Vui lòng đợi hoàn thành.")

    logger.info(
        "BQMS sync triggered: job_id=%d, range=%s→%s, by=%s",
        sync_id, date_from, date_to, token_data.user_id,
    )

    # Run Playwright sync in a separate thread (Playwright needs its own event loop)
    import threading

    def _thread_sync(sid: int):
        """Run Playwright sync in a dedicated thread with its own asyncio loop."""
        import asyncio as _aio

        async def _do_sync():
            import asyncpg as apg
            from app.etl.bqms_playwright import playwright_fetch_pos
            from app.tasks.bqms_sync import _upsert_pos, _update_sync_log
            from app.core.config import settings as cfg

            db_url = str(cfg.DATABASE_URL).replace("+asyncpg", "").replace("postgresql+asyncpg", "postgresql")
            bconn = await apg.connect(db_url)
            try:
                await bconn.execute(
                    "UPDATE etl_sync_log SET status = 'running', started_at = NOW() WHERE id = $1", sid
                )
            finally:
                await bconn.close()

            result = {"new_pos": 0, "updated_pos": 0, "status": "running"}
            try:
                po_list = await playwright_fetch_pos()
                new_list, upd = _upsert_pos(po_list)
                result["new_pos"] = len(new_list)
                result["updated_pos"] = upd
                result["status"] = "success"

                # Bridge: tất cả PO → bqms_deliveries (trang Giao Hàng)
                from app.tasks.bqms_sync import _bridge_po_to_deliveries
                new_del, upd_del = _bridge_po_to_deliveries(po_list)
                result["deliveries_created"] = new_del
                result["deliveries_updated"] = upd_del
            except Exception as exc:
                result["status"] = "error"
                result["error_message"] = str(exc)[:500]

            _update_sync_log(sid, result)

        _aio.run(_do_sync())

    t = threading.Thread(target=_thread_sync, args=(sync_id,), daemon=True)
    t.start()

    return {
        "job_id": sync_id,
        "status": "queued",
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "message": f"Đồng bộ đã được khởi tạo (job #{sync_id}). Dùng GET /bqms/sync/status/{sync_id} để kiểm tra.",
    }


@router.get("/sync/status/{job_id}")
async def sync_status(
    job_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Check status of a BQMS sync job."""
    result = await _bqms_service.get_sync_status(conn, job_id)
    if not result:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy job #{job_id}")

    # Convert datetime fields for JSON serialization
    for key in ("started_at", "completed_at"):
        if result.get(key) and isinstance(result[key], datetime):
            result[key] = result[key].isoformat()

    return {"data": result}


@router.get("/sync/latest")
async def sync_latest(
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Lấy kết quả đồng bộ gần nhất."""
    row = await conn.fetchrow(
        """
        SELECT id, sync_type, status, started_at, completed_at,
               rows_inserted, rows_updated, rows_skipped, error_message,
               EXTRACT(EPOCH FROM (COALESCE(completed_at, NOW()) - started_at))::int AS duration_seconds
        FROM etl_sync_log
        WHERE sync_type = 'bqms_po'
        ORDER BY started_at DESC
        LIMIT 1
        """
    )
    if not row:
        return {"data": None}

    result = dict(row)
    for key in ("started_at", "completed_at"):
        if result.get(key) and isinstance(result[key], datetime):
            result[key] = result[key].isoformat()

    return {"data": result}


@router.get("/sync/circuit")
async def sync_circuit_status(
    token_data: TokenData = Depends(require_role("admin", "manager")),
):
    """Trạng thái circuit breaker Samsung BQMS — bảo vệ tài khoản khỏi bị khóa."""
    from app.etl.bqms_playwright import _load_circuit, _BACKOFF_SECONDS
    import time as _time
    circuit = _load_circuit()
    failures = circuit.get("failures", 0)
    last_fail = circuit.get("last_failure_at", 0)

    if failures == 0:
        state = "closed"
        wait_remaining = 0
    else:
        idx = min(failures - 1, len(_BACKOFF_SECONDS) - 1)
        wait_total = _BACKOFF_SECONDS[idx]
        elapsed = _time.time() - last_fail
        if elapsed < wait_total:
            state = "open"
            wait_remaining = int(wait_total - elapsed)
        else:
            state = "half-open"
            wait_remaining = 0

    return {
        "state": state,
        "failures": failures,
        "last_error": circuit.get("last_error", ""),
        "wait_remaining_seconds": wait_remaining,
        "wait_remaining_minutes": wait_remaining // 60,
    }


@router.post("/sync/circuit/reset")
async def reset_circuit_breaker(
    token_data: TokenData = Depends(require_role("admin")),
):
    """Admin reset circuit breaker — cho phép thử login lại ngay."""
    from app.etl.bqms_playwright import _record_success
    _record_success()
    return {"message": "Circuit breaker đã reset. Có thể thử đồng bộ lại."}


@router.get("/sync/steps")
async def sync_steps(
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
):
    """Trạng thái từng bước của đồng bộ Samsung — poll real-time."""
    from app.etl.bqms_playwright import get_sync_steps, STEP_DEFINITIONS
    data = get_sync_steps()
    data["definitions"] = STEP_DEFINITIONS
    return data


@router.get("/sync/history")
async def sync_history(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Lịch sử đồng bộ Samsung BQMS."""
    total = await conn.fetchval(
        "SELECT COUNT(*) FROM etl_sync_log WHERE sync_type = 'bqms_po'"
    )
    rows = await conn.fetch(
        """
        SELECT id, sync_type, status, started_at, completed_at,
               rows_inserted, rows_updated, rows_skipped, error_message,
               EXTRACT(EPOCH FROM (COALESCE(completed_at, NOW()) - started_at))::int AS duration_seconds
        FROM etl_sync_log
        WHERE sync_type = 'bqms_po'
        ORDER BY started_at DESC
        LIMIT $1 OFFSET $2
        """,
        limit, offset,
    )

    data = []
    for r in rows:
        d = dict(r)
        for key in ("started_at", "completed_at"):
            if d.get(key) and isinstance(d[key], datetime):
                d[key] = d[key].isoformat()
        data.append(d)

    return {"data": data, "total": total}


# ---------------------------------------------------------------------------
# RFQ — PDF Parsing
# ---------------------------------------------------------------------------

@router.post("/rfq/parse")
async def parse_rfq_pdf(
    file: UploadFile = File(..., description="Samsung RFQ PDF file"),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """
    Upload and parse a Samsung RFQ PDF file.

    Returns structured items extracted from the PDF, including product matching
    against the existing products table.
    """
    # Validate file type
    filename = file.filename or "untitled.pdf"
    if not filename.lower().endswith(".pdf"):
        raise HTTPException(
            status_code=400,
            detail="Chỉ chấp nhận file PDF. Vui lòng upload file .pdf",
        )

    # Read content
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File trống")

    if len(content) > 50 * 1024 * 1024:  # 50 MB
        raise HTTPException(status_code=400, detail="File quá lớn (tối đa 50 MB)")

    logger.info(
        "RFQ PDF upload: file=%s, size=%d, user=%s",
        filename, len(content), token_data.user_id,
    )

    try:
        result = await _bqms_service.parse_rfq_pdf(conn, content, filename)
    except Exception as e:
        logger.error("RFQ PDF parse error: %s", e)
        raise HTTPException(
            status_code=422,
            detail=f"Không thể phân tích file PDF: {e}",
        )

    if not result.get("success"):
        raise HTTPException(
            status_code=422,
            detail=result.get("error", "Lỗi không xác định khi phân tích PDF"),
        )

    return {
        "data": result,
        "message": f"Đã phân tích {result.get('items_count', 0)} items từ RFQ PDF",
    }


# ---------------------------------------------------------------------------
# RFQ — Quotation Generation
# ---------------------------------------------------------------------------

@router.post("/rfq/generate")
async def generate_quotation(
    submission_id: int = Query(..., description="ID của bqms_rfq_submissions"),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """
    Generate Excel quotation files (CAM_KET + QUOTATION) from an existing submission.

    Requires line items to already be saved in bqms_quotation_items.
    """
    try:
        result = await _bqms_service.generate_quotation(conn, submission_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("Quotation generation error: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"Lỗi tạo file báo giá: {e}",
        )

    return {
        "data": result,
        "message": f"Đã tạo file báo giá cho RFQ {result.get('rfq_number', 'N/A')}",
    }


# ---------------------------------------------------------------------------
# RFQ — Submit for Approval
# ---------------------------------------------------------------------------

@router.post("/rfq/submit")
async def submit_quotation(
    submission_id: int = Query(..., description="ID của bqms_rfq_submissions"),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """
    Submit a quotation for manager/admin approval.

    Creates a workflow instance and changes submission status from 'draft' to 'pending'.
    """
    try:
        result = await _bqms_service.submit_quotation(
            conn, submission_id, token_data.user_id
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("Quotation submit error: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"Lỗi khi submit báo giá: {e}",
        )

    return {
        "data": result,
        "message": result.get("message", "Đã gửi báo giá để phê duyệt"),
    }


# ---------------------------------------------------------------------------
# KPI Summary
# ---------------------------------------------------------------------------

@router.get("/kpi")
async def kpi_summary(
    period: str | None = Query(None, description="YYYY-MM"),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """KPI summary computed from real bqms_rfq data."""
    total = await conn.fetchval("SELECT COUNT(*) FROM bqms_rfq")
    won = await conn.fetchval("SELECT COUNT(*) FROM bqms_rfq WHERE result = 'won'")
    lost = await conn.fetchval("SELECT COUNT(*) FROM bqms_rfq WHERE result = 'lost'")
    pending = await conn.fetchval("SELECT COUNT(*) FROM bqms_rfq WHERE result = 'pending' OR result IS NULL")
    makers = await conn.fetchval("SELECT COUNT(DISTINCT maker) FROM bqms_rfq WHERE maker IS NOT NULL")
    deliveries = await conn.fetchval("SELECT COUNT(*) FROM bqms_deliveries")
    samsung_po = await conn.fetchval("SELECT COUNT(*) FROM bqms_samsung_po")
    decided = won + lost
    win_rate = round(won * 100.0 / decided, 1) if decided > 0 else 0
    return {
        "data": {
            "total_rfqs": total,
            "won_count": won,
            "lost_count": lost,
            "pending_count": pending,
            "win_rate_pct": win_rate,
            "maker_count": makers,
            "total_deliveries": deliveries,
            "total_samsung_po": samsung_po,
            "total_items": total,
            "processed": won + lost,
        }
    }


# ---------------------------------------------------------------------------
# Records
# ---------------------------------------------------------------------------

@router.get("/records")
async def list_records(
    status: str | None = Query(None),
    category: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    conditions = ["1=1"]
    params: list = []
    idx = 1

    if status:
        conditions.append(f"br.status = ${idx}")
        params.append(status)
        idx += 1
    if category:
        conditions.append(f"br.category = ${idx}")
        params.append(category)
        idx += 1
    if date_from:
        conditions.append(f"br.synced_at >= ${idx}")
        params.append(date_from)
        idx += 1
    if date_to:
        conditions.append(f"br.synced_at <= ${idx}")
        params.append(date_to)
        idx += 1

    where = " AND ".join(conditions)

    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM bqms_records br WHERE {where}", *params
    )

    params.extend([limit, offset])
    rows = await conn.fetch(
        f"""
        SELECT br.*
        FROM bqms_records br
        WHERE {where}
        ORDER BY br.synced_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )
    return {"data": [dict(r) for r in rows], "total": total}


# ---------------------------------------------------------------------------
# RFQ List
# ---------------------------------------------------------------------------

@router.get("/rfq")
async def list_rfq(
    status: str | None = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    conditions = ["1=1"]
    params: list = []
    idx = 1

    if status:
        conditions.append(f"r.status = ${idx}")
        params.append(status)
        idx += 1

    where = " AND ".join(conditions)

    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM bqms_rfq r WHERE {where}", *params
    )

    params.extend([limit, offset])
    rows = await conn.fetch(
        f"""
        SELECT r.*
        FROM bqms_rfq r
        WHERE {where}
        ORDER BY r.created_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )
    return {"data": [dict(r) for r in rows], "total": total}


# ---------------------------------------------------------------------------
# Pareto Analysis
# ---------------------------------------------------------------------------

@router.get("/analytics/pareto")
async def pareto_analysis(
    period: str | None = Query(None, description="YYYY-MM"),
    top_n: int = Query(20, ge=5, le=100),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Pareto analysis -- top defect categories by frequency."""
    conditions = ["1=1"]
    params: list = []
    idx = 1

    if period:
        conditions.append(f"TO_CHAR(br.synced_at, 'YYYY-MM') = ${idx}")
        params.append(period)
        idx += 1

    where = " AND ".join(conditions)
    params.append(top_n)

    # Group RFQs by maker — Pareto on supplier mix.
    # bqms_records does not have category/defect_qty cols; use bqms_rfq.maker
    # which is the meaningful dimension for our supplier-side workflow.
    where_rfq = where.replace('br.synced_at', 'r.created_at').replace('br.', 'r.')
    rows = await conn.fetch(
        f"""
        SELECT COALESCE(NULLIF(TRIM(r.maker), ''), 'Khong xac dinh') AS category,
               COUNT(*) AS count,
               SUM(COALESCE(r.expected_qty, 0))::numeric AS total_qty,
               ROUND(
                   100.0 * COUNT(*) / NULLIF(SUM(COUNT(*)) OVER (), 0), 2
               ) AS percentage
        FROM bqms_rfq r
        WHERE {where_rfq.replace('1=1', 'TRUE')}
        GROUP BY 1
        ORDER BY count DESC
        LIMIT ${idx}
        """,
        *params,
    )

    # Build cumulative percentage
    data = []
    cumulative = 0.0
    for r in rows:
        pct = float(r["percentage"] or 0)
        cumulative += pct
        data.append({
            **dict(r),
            "cumulative_pct": round(cumulative, 2),
        })

    return {"data": data}


# ---------------------------------------------------------------------------
# RFQ Table — Unified BQMS Page (main endpoint)
# ---------------------------------------------------------------------------

@router.get("/rfq-table")
async def rfq_table(
    year: int | None = Query(None, description="Năm lọc (VD: 2026)"),
    month: int | None = Query(None, description="Tháng lọc (1-12)"),
    search: str | None = Query(None, description="Tìm theo RFQ No, BQMS Code, tên hàng, maker"),
    result_filter: str | None = Query(None, description="pending | won | lost | all"),
    source_filter: str | None = Query(None, description="excel_import | etl | onedrive_sync | manual | all"),
    loai_hang: str | None = Query(None, description="TM | GC | all (Drawing detection)"),
    round_filter: str | None = Query(
        None,
        description=(
            "v1_has | v2_has | v3_has | v4_has | v1_missing | all — "
            "lọc theo trạng thái cột quoted_price_bqms_v1..v4. "
            "v{N}_has = đã có giá V{N}; v1_missing = CHƯA có V1 (mới scrape, "
            "chưa báo giá lần nào)."
        ),
    ),
    page: int = Query(1, ge=1),
    # Thang 2026-06-15 (Batch 2b): default 100→12 — khớp page-size mặc định FE,
    # first paint nhẹ hơn. FE gửi page_size rõ ràng nên default chỉ là fallback.
    page_size: int = Query(12, ge=10, le=500),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """
    Main data endpoint for the unified BQMS table.
    Queries bqms_rfq (6,473 rows). Returns paginated, filterable, sortable data
    with KPI summary and month-group metadata.
    """
    # ─── Param validation (Thang 2026-06-13: polish 2) ─────────────
    # Reject unknown round_filter values với 400 thay vì silent fall-through.
    # Tránh case typo "v1_have" → trả về data không filter mà user không biết.
    _ROUND_FILTER_ALLOWED = {"all", "v1_has", "v2_has", "v3_has", "v4_has", "v1_missing"}
    if round_filter and round_filter.lower() not in _ROUND_FILTER_ALLOWED:
        raise HTTPException(400, "round_filter không hợp lệ")

    # Batch 2C: V-round / D-N tracking columns are added by a later migration
    # (bqms_vround_tracking.sql). Probe once so the SELECT degrades gracefully
    # (NULL literals) when the migration hasn't been applied yet — never 500.
    _has_vround = bool(await conn.fetchval(
        "SELECT EXISTS (SELECT 1 FROM information_schema.columns "
        "WHERE table_name='bqms_rfq' AND column_name='qt_state')"
    ))
    if _has_vround:
        _vround_select = (
            "            r.deadline_dt,\n"
            "            r.qt_state::text AS qt_state,\n"
            "            r.current_round,"
        )
    else:
        _vround_select = (
            "            NULL::timestamptz AS deadline_dt,\n"
            "            NULL::text AS qt_state,\n"
            "            NULL::smallint AS current_round,"
        )

    # ─── Scope conditions (year + month + search) ──────────────────
    # Thang 2026-05-29: tách scope (year/month/search) khỏi secondary filters
    # (result_filter/source_filter/loai_hang) để KPI Trúng/Trượt/Pending hiển
    # đúng TỔNG tháng kể cả khi user đang lọc theo "Chưa báo giá" etc.
    scope_conditions: list[str] = ["1=1"]
    scope_params: list[Any] = []
    sidx = 1

    if year is not None:
        scope_conditions.append(
            f"EXTRACT(YEAR FROM COALESCE(r.inquiry_date, r.created_at::date)) = ${sidx}"
        )
        scope_params.append(year)
        sidx += 1

    if month is not None:
        scope_conditions.append(
            f"EXTRACT(MONTH FROM COALESCE(r.inquiry_date, r.created_at::date)) = ${sidx}"
        )
        scope_params.append(month)
        sidx += 1

    if search:
        like = f"%{search}%"
        # Search bao gồm description từ staging.raw_json._detail.items
        # (e.g. "CNC BRUSH", "BLADE") — bqms_rfq không có cột description
        # nên cần subquery. Per Thang 2026-05-15.
        # CRITICAL FIX (Thang 2026-05-22): prefix all columns with `r.` —
        # this query JOINs with round2_audit CTE which ALSO has `rfq_number`,
        # making bare `rfq_number` ambiguous → AmbiguousColumnError → 500.
        scope_conditions.append(
            f"(r.rfq_number ILIKE ${sidx} OR r.bqms_code ILIKE ${sidx} "
            f"OR r.specification ILIKE ${sidx} OR r.maker ILIKE ${sidx} "
            f"OR EXISTS (SELECT 1 FROM bqms_vendor_portal_staging s "
            f"  WHERE s.module='bidding' AND s.rfq_number = r.rfq_number "
            f"  AND s.raw_json::text ILIKE ${sidx}))"
        )
        scope_params.append(like)
        sidx += 1

    where_kpi = " AND ".join(scope_conditions)

    # ─── Full conditions (scope + secondary filters) ───────────────
    # Inherit scope, then append result_filter / source_filter / loai_hang.
    conditions: list[str] = list(scope_conditions)
    params: list[Any] = list(scope_params)
    idx = sidx

    # Thang 2026-05-16: tách "pending" thành 2 filter:
    #   - tracking (Đang theo dõi): result=pending AND quote_unlocked=true (đã báo giá)
    #   - unquoted (Chưa báo giá): result=pending AND quote_unlocked=false (chưa quyết định)
    # Backward compat: 'pending' vẫn match cả 2 (= union).
    if result_filter and result_filter.lower() != "all":
        rf = result_filter.lower()
        if rf == "tracking":
            conditions.append("r.result::text = 'pending' AND COALESCE(r.quote_unlocked, false) = true")
        elif rf == "unquoted":
            conditions.append("r.result::text = 'pending' AND COALESCE(r.quote_unlocked, false) = false")
        else:
            conditions.append(f"r.result::text = ${idx}")
            params.append(rf)
            idx += 1

    if source_filter and source_filter.lower() != "all":
        conditions.append(f"r.data_source = ${idx}")
        params.append(source_filter.lower())
        idx += 1

    if loai_hang and loai_hang.upper() != "ALL":
        # Drawing classification stored in notes as 'classification=GC' or 'classification=TM'
        # for etl rows. For excel rows, fall back to spec keyword heuristic.
        if loai_hang.upper() == "GC":
            conditions.append(
                f"(r.notes ILIKE '%classification=GC%' OR r.specification ILIKE '%gia c_ng%')"
            )
        elif loai_hang.upper() == "TM":
            conditions.append(
                f"(r.notes ILIKE '%classification=TM%' OR "
                f"(r.notes NOT ILIKE '%classification=GC%' "
                f"AND r.specification NOT ILIKE '%gia c_ng%'))"
            )

    # Thang 2026-06-13: round_filter — lọc theo trạng thái cột quoted_price_bqms_v1..v4.
    # Mục đích: user muốn xem riêng "đã có V1", "đã có V2", ..., hoặc "chưa có V1"
    # (mã mới scrape, chưa báo giá lần nào). Filter ADDITIVE, không phá KPI tổng tháng
    # (KPI vẫn dùng `where_kpi` = scope_conditions, không bị round_filter chi phối).
    if round_filter and round_filter.lower() != "all":
        rfv = round_filter.lower().strip()
        _round_col_map = {
            "v1_has": "r.quoted_price_bqms_v1 IS NOT NULL",
            "v2_has": "r.quoted_price_bqms_v2 IS NOT NULL",
            "v3_has": "r.quoted_price_bqms_v3 IS NOT NULL",
            "v4_has": "r.quoted_price_bqms_v4 IS NOT NULL",
            "v1_missing": "r.quoted_price_bqms_v1 IS NULL",
        }
        if rfv in _round_col_map:
            conditions.append(_round_col_map[rfv])
        # Unknown round_filter values → silently ignored (treated as "all") để
        # không break clients đang gửi value lạ trong cache.

    where = " AND ".join(conditions)

    # Thang 2026-05-22: all queries use `bqms_rfq r` alias since the WHERE
    # clause (built above) qualifies columns with `r.` to avoid ambiguity
    # with the round2_audit JOIN in the paginated SELECT.

    # Thang 2026-06-04 (BUG B fix — push button missing on duplicate twins):
    # `bqms_rfq` has 116 (rfq_number, bqms_code) pairs that are duplicated
    # across data_source='etl' and data_source='onedrive_sync'. The dedup
    # unique index `uq_bqms_rfq_dedup` keys on (rfq_number, bqms_code,
    # source_hash); different source_hash → INSERT-new instead of UPSERT.
    # The etl twin carries user actions (quote_unlocked=true, V1 price,
    # bqms_push_status='saved_temp'), but the onedrive_sync twin has a
    # newer inquiry_date and wins ORDER BY inquiry_date DESC — hiding the
    # push button (`item.quote_unlocked === true` returns false on the
    # shadowing twin).
    #
    # Option A fix: dedupe at query time via `bqms_dedup` CTE. DISTINCT ON
    # (rfq_number, bqms_code) ordered by (quote_unlocked DESC, has-push-state
    # DESC, updated_at DESC) → the row carrying user actions always wins,
    # regardless of which sync ran last. Single-file edit, low risk.
    # Permanent root-cause fix (collapse dedup key) tracked for next sprint.
    # Thang 2026-06-13 (FIX V1-hidden + V2-push-fail):
    # Reorder ORDER BY so V-presence beats push-state. Previous ordering
    # `(bqms_push_status IS NOT NULL)::int DESC` came BEFORE V-presence —
    # if the duplicate twin carried `bqms_push_status='queued'` but NULL
    # for `quoted_price_bqms_v1`, push-state-twin won → frontend rendered
    # the row with V1=NULL (looked "ẩn"). Now any twin holding a real V1
    # price wins regardless of push state.
    #
    # Tiebreaker hierarchy (top wins):
    #   1. quote_unlocked=true  (user actioned)
    #   2. V4 price set         (rounds collapse newest→oldest)
    #   3. V3 price set
    #   4. V2 price set
    #   5. V1 price set
    #   6. bqms_push_status NOT NULL  (push attempted at all)
    #   7. updated_at DESC
    #   8. id DESC
    dedup_cte = """
        bqms_dedup AS (
            SELECT DISTINCT ON (rfq_number, bqms_code) *
              FROM bqms_rfq
             ORDER BY rfq_number, bqms_code,
                      (COALESCE(quote_unlocked, false))::int DESC,
                      (quoted_price_bqms_v4 IS NOT NULL)::int DESC,
                      (quoted_price_bqms_v3 IS NOT NULL)::int DESC,
                      (quoted_price_bqms_v2 IS NOT NULL)::int DESC,
                      (quoted_price_bqms_v1 IS NOT NULL)::int DESC,
                      (bqms_push_status IS NOT NULL)::int DESC,
                      updated_at DESC NULLS LAST,
                      id DESC
        )
    """

    # Total count — respects ALL filters (table is paginated by `where`)
    total: int = await conn.fetchval(
        f"WITH {dedup_cte} SELECT COUNT(*) FROM bqms_dedup r WHERE {where}",
        *params,
    )

    # KPI — chỉ scope (year/month/search), KHÔNG bị result_filter / source / loai_hang
    # chi phối. Đảm bảo "Trúng thầu" / "Trượt thầu" luôn hiển tổng tháng đúng.
    kpi_rows = await conn.fetchrow(
        f"""
        WITH {dedup_cte}
        SELECT
            COUNT(*) FILTER (WHERE r.result::text = 'won')  AS won,
            COUNT(*) FILTER (WHERE r.result::text = 'lost') AS lost,
            COUNT(*) FILTER (
                WHERE r.result::text = 'pending' OR r.result IS NULL
            ) AS pending
        FROM bqms_dedup r
        WHERE {where_kpi}
        """,
        *scope_params,
    )
    won_count  = int(kpi_rows["won"]  or 0)
    lost_count = int(kpi_rows["lost"] or 0)
    pending_count = int(kpi_rows["pending"] or 0)
    decided = won_count + lost_count
    win_rate = round(won_count * 100.0 / decided, 1) if decided > 0 else 0.0

    # Month summary (group headers for the UI)
    month_rows = await conn.fetch(
        f"""
        WITH {dedup_cte}
        SELECT
            EXTRACT(YEAR  FROM COALESCE(r.inquiry_date, r.created_at::date))::int AS yr,
            EXTRACT(MONTH FROM COALESCE(r.inquiry_date, r.created_at::date))::int AS mo,
            COUNT(*)                                                              AS cnt,
            COUNT(*) FILTER (WHERE r.result::text = 'won')                       AS won,
            COUNT(*) FILTER (WHERE r.result::text = 'lost')                      AS lost
        FROM bqms_dedup r
        WHERE {where}
        GROUP BY yr, mo
        ORDER BY yr DESC, mo DESC
        LIMIT 24
        """,
        *params,
    )
    months_data = [
        {
            "year": r["yr"],
            "month": r["mo"],
            "count": int(r["cnt"] or 0),
            "won": int(r["won"] or 0),
            "lost": int(r["lost"] or 0),
        }
        for r in month_rows
    ]

    # Paginated rows
    offset = (page - 1) * page_size
    params_paged = params + [page_size, offset]
    # Phase 2 per Thang 2026-05-12: include requester/department/assigned_to.
    # Phase E (Thang 2026-05-13): include classification_override for user-edit support.
    # Round-2 priority sort (Thang 2026-05-18):
    # RFQ với version >= 2 + audit_log round-2 trong 7 ngày → đẩy lên đầu
    # Thêm 2 fields:
    #   round2_recent_at: timestamp audit log gần nhất (NULL nếu không có)
    #   is_round2_24h:    true nếu round-2 trong 24h gần nhất (UI highlight cam)
    rows = await conn.fetch(
        f"""
        WITH {dedup_cte},
        round2_audit AS (
            SELECT record_id AS rfq_number,
                   MAX(created_at) AS round2_at
              FROM audit_log
             WHERE action IN ('bqms_periodic.round2_invitation',
                              'bqms_periodic.round2_v1_missing_warning')
               AND created_at > NOW() - INTERVAL '7 days'
             GROUP BY record_id
        )
        SELECT
            r.id, r.rfq_number, r.bqms_code, r.specification, r.maker,
            r.expected_qty, r.unit,
            r.purchase_price_rmb, r.purchase_price_vnd,
            r.quoted_price_ama,
            r.quoted_price_bqms_v1, r.quoted_price_bqms_v2,
            r.quoted_price_bqms_v3, r.quoted_price_bqms_v4,
            -- Thang 2026-06-13: per-round quote dates for UI chips + audit.
            r.quoted_dt_v1, r.quoted_dt_v2, r.quoted_dt_v3, r.quoted_dt_v4,
            r.supplier_name, r.result::text AS result, r.notes, r.report,
            r.person_in_charge_name, r.inquiry_date,
            COALESCE(r.inquiry_date, r.created_at::date) AS effective_date,
            r.created_at, r.version, r.data_source,
            r.requester, r.department,
            r.assigned_to::text AS assigned_to,
            r.classification_override,
            -- Thang 2026-06-04: explicit boolean cast so frontend gate
            -- `item.quote_unlocked === true` evaluates correctly. Also
            -- surface push state so UI can render re-push button on
            -- saved_temp rows (round 2-4) even when quote_unlocked got
            -- relocked by a later round.
            COALESCE(r.quote_unlocked, false)::boolean AS quote_unlocked,
            r.bqms_push_status,
            r.bqms_pushed_round,
            -- Thang 2026-06-15 (Batch 2f): ngày đẩy báo giá lên SEC — thay cột STT "#".
            r.bqms_pushed_at,
            -- Batch 2C: V-round / D-N tracking columns (NULL literals when the
            -- bqms_vround_tracking.sql migration is not yet applied).
{_vround_select}
            -- Round-2 priority flags
            a.round2_at AS round2_recent_at,
            (a.round2_at IS NOT NULL
             AND a.round2_at > NOW() - INTERVAL '24 hours') AS is_round2_24h,
            (a.round2_at IS NOT NULL AND r.version >= 2) AS is_round2_priority
        FROM bqms_dedup r
        LEFT JOIN round2_audit a ON a.rfq_number = r.rfq_number
        WHERE {where}
        ORDER BY
            -- Priority 1: Round-2 RFQ trong 7 ngày + version>=2 lên đầu, mới nhất trước
            CASE WHEN a.round2_at IS NOT NULL AND r.version >= 2 THEN 0 ELSE 1 END,
            a.round2_at DESC NULLS LAST,
            -- Priority 2: Fallback theo inquiry_date như cũ
            r.inquiry_date DESC NULLS LAST,
            r.updated_at DESC NULLS LAST,
            r.id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params_paged,
    )

    # Per Thang 2026-05-11: enrich approved rows with bidding metadata
    # (MOQ/CIS/Part No/maker/deadline/classification/...) by post-query
    # JOIN with bqms_vendor_portal_staging.raw_json. This avoids fragile
    # SQL alias-rewriting in the main WHERE clause.
    rfq_numbers = {r["rfq_number"] for r in rows if r["rfq_number"]}
    staging_meta: dict[str, dict] = {}
    if rfq_numbers:
        # FIX duplicate (Thang 2026-05-14): include staging.id so bqms_rfq rows
        # can carry staging_id → button "Báo giá" trên approved row vẫn gọi
        # /vendor-staging/{staging_id}/quote được.
        meta_rows = await conn.fetch(
            """
            SELECT DISTINCT ON (rfq_number)
                rfq_number, id AS staging_id, raw_json
            FROM bqms_vendor_portal_staging
            WHERE module='bidding' AND rfq_number = ANY($1::text[])
            ORDER BY rfq_number, id DESC
            """,
            list(rfq_numbers),
        )
        for mr in meta_rows:
            raw = mr["raw_json"] or {}
            if isinstance(raw, str):
                try:
                    raw = _json.loads(raw)
                except Exception:
                    raw = {}
            detail = raw.get("_detail") or {}
            items_arr = detail.get("items") or []
            staging_meta[mr["rfq_number"]] = {
                "staging_id": int(mr["staging_id"]),
                "raw": raw,
                "detail": detail,
                "items_arr": items_arr,
                "deadline_dt": (raw.get("deadlineDt") or "").strip() or None,
                "reg_dt": (raw.get("regDt") or "").strip() or None,
                "bd_status": (raw.get("progressStatusName") or "").strip() or None,
                "psincharge_name": (raw.get("psinchargeName") or "").strip() or None,
                "ctr_type_nm": (raw.get("ctrTypeNm") or "").strip() or None,
                "currency": (raw.get("criteriaCurrency") or "").strip() or None,
                "dday_html": (raw.get("dday") or "").strip() or None,
                "req_name": (raw.get("reqName") or "").strip() or None,
                "classification": detail.get("classification"),
                "detail_version": detail.get("version"),
                "items_count": len(items_arr),
                "attachments_count": len(detail.get("attachments") or []),
            }

    # Per Thang 2026-05-11: Merge Bidding pending rows into BQMS table.
    # Pending bidding rows live in bqms_vendor_portal_staging (status=pending_review)
    # and aren't yet in bqms_rfq. We surface them here with `is_pending=true`
    # so the user can see + Báo giá from a single unified table.
    # Only include on first page + when filter shows un-quoted rows.
    # Thang 2026-05-16: pending_bidding staging = mã mới chưa báo giá → thuộc
    # nhóm "Chưa báo giá" (unquoted) hoặc "Tất cả" (all). Filter 'tracking'
    # KHÔNG stitch vì đã báo giá = đã có trong bqms_rfq, không còn staging.
    pending_bidding: list[dict] = []
    if page == 1 and (not result_filter or result_filter.lower() in ("all", "pending", "unquoted")):
        # FIX duplicate (Thang 2026-05-14): Sau Phase H, auto-drill UPSERT vào
        # bqms_rfq nhưng staging vẫn status='pending_review' → rfq-table merge
        # cả 2 source → 1 RFQ hiện 2 dòng (staging với nút "Báo giá" + bqms_rfq
        # với VP badge). Filter: chỉ surface staging nếu rfq_number chưa có
        # trong bqms_rfq. bqms_rfq row sẽ tự cõng nút "Báo giá" qua staging_id_map.
        # Thang 2026-06-04: Scope theo `search` để tránh staging stub leak khi
        # user tìm 1 RFQ cụ thể. Trước đây query trả về 200 pending mới nhất
        # → nếu RFQ search bị bqms_rfq filter (year/month) loại bỏ, staging stub
        # vẫn xuất hiện với is_pending=true → drawer ẩn nút "Đẩy lên SEC".
        pending_sql = """
            SELECT
                s.id AS staging_id,
                s.rfq_number,
                s.created_at,
                s.raw_json
            FROM bqms_vendor_portal_staging s
            WHERE s.module = 'bidding' AND s.status = 'pending_review'
              AND NOT EXISTS (
                  SELECT 1 FROM bqms_rfq r WHERE r.rfq_number = s.rfq_number
              )
        """
        pending_args: list[Any] = []
        if search:
            pending_sql += (
                " AND (s.rfq_number ILIKE $1"
                "      OR s.raw_json::text ILIKE $1)"
            )
            pending_args.append(f"%{search}%")
        pending_sql += " ORDER BY s.id DESC LIMIT 200"
        pending_rows = await conn.fetch(pending_sql, *pending_args)
        for r in pending_rows:
            raw = r["raw_json"] or {}
            if isinstance(raw, str):
                try:
                    raw = _json.loads(raw)
                except Exception:
                    raw = {}
            detail = raw.get("_detail") or {}
            items = detail.get("items") or []

            # Per Thang 2026-05-11: expand to 1 row per item so user sees
            # each bqms_code with its own Description + Specification.
            # Fallback: if no items drilled yet, show a single placeholder row.
            iter_items = items if items else [{}]
            # Phase 2 (Thang 2026-05-13): parse Requester/Department từ
            # psinchargeName format "Name/Department/Company"
            _pic_full = (raw.get("psinchargeName") or "").strip()
            _pic_parts = [p.strip() for p in _pic_full.split("/") if p.strip()] if _pic_full else []
            _row_requester = _pic_parts[0] if _pic_parts else None
            _row_department = _pic_parts[1] if len(_pic_parts) >= 2 else None
            for idx_it, it in enumerate(iter_items):
                bqms_code = (it.get("item_code") or "").strip() or None
                description = (it.get("description") or "").strip() or None
                specification = (it.get("specification") or "").strip() or None
                pending_bidding.append({
                    # Negative ID composite: -(staging_id*100 + idx) so multiple
                    # items per staging row stay unique without colliding w/ bqms_rfq IDs.
                    "id": -(int(r["staging_id"]) * 100 + idx_it),
                    "staging_id": int(r["staging_id"]),
                    "is_pending": True,
                    "rfq_number": r["rfq_number"],
                    "bqms_code": bqms_code,
                    "description": description,
                    "specification": specification or (raw.get("reqName") or "").strip() or None,
                    "maker": (it.get("maker") or "").strip() or None,
                    "expected_qty": it.get("qty"),
                    "unit": (it.get("unit") or "").strip() or None,
                    "purchase_price_rmb": None,
                    "purchase_price_vnd": None,
                    "quoted_price_ama": None,
                    "quoted_price_bqms_v1": None,
                    "quoted_price_bqms_v2": None,
                    "quoted_price_bqms_v3": None,
                    "quoted_price_bqms_v4": None,
                    "supplier_name": None,
                    "result": "pending",
                    "notes": None,
                    "report": None,
                    "person_in_charge_name": (raw.get("psinchargeName") or "").split("/")[0].strip() or None,
                    "inquiry_date": None,
                    "effective_date": r["created_at"].date() if r["created_at"] else None,
                    "created_at": r["created_at"],
                    "version": None,
                    "data_source": "bidding_pending",
                    "req_name": (raw.get("reqName") or "").strip() or None,
                    "reg_dt": (raw.get("regDt") or "").strip() or None,
                    "deadline_dt": (raw.get("deadlineDt") or "").strip() or None,
                    "submit_dt": (raw.get("submitDt") or "").strip() or None,
                    "bd_status": (raw.get("progressStatusName") or raw.get("submitGb") or "").strip() or None,
                    "psincharge_name": (raw.get("psinchargeName") or "").strip() or None,
                    "currency": (raw.get("criteriaCurrency") or "").strip() or None,
                    "item_cnt_text": (raw.get("itemCnt") or "").strip() or None,
                    "dday_html": (raw.get("dday") or "").strip() or None,
                    "ctr_type_nm": (raw.get("ctrTypeNm") or "").strip() or None,
                    "classification": detail.get("classification"),
                    "detail_version": detail.get("version"),
                    "items_count": len(items),
                    "attachments_count": len(detail.get("attachments") or []),
                    "detail_error": (detail.get("error") or "").strip() or None,
                    # Per-item enriched fields (renamed from "first_*" since now
                    # per-row, not per-RFQ).
                    "first_maker": (it.get("maker") or "").strip() or None,
                    "first_part_no": (it.get("part_no") or "").strip() or None,
                    "first_cis_code": (it.get("cis_code") or "").strip() or None,
                    "first_moq": (it.get("moq") or "").strip() or None,
                    # Phase 2 (Thang 2026-05-13): cũng hiển thị cho pending
                    "requester": _row_requester,
                    "department": _row_department,
                    "assigned_to": None,
                    "assigned_to_name": None,
                })

    def _serialize(r) -> dict:
        d = dict(r) if not isinstance(r, dict) else r
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
        return d

    def _enrich_with_staging(d: dict) -> dict:
        """Mix bidding metadata from staging (matched by rfq_number)
        into an approved row. Per-item fields (maker/moq/cis_code/part_no)
        are looked up by bqms_code. Only adds keys that aren't set."""
        if d.get("is_pending"):
            return d
        rfq_no = d.get("rfq_number")
        meta = staging_meta.get(rfq_no)
        if not meta:
            return d
        # FIX duplicate (Thang 2026-05-14): inject staging_id so button "Báo giá"
        # trên bqms_rfq row vẫn gọi được endpoint /vendor-staging/{id}/quote.
        d["staging_id"] = meta.get("staging_id")
        # Per-RFQ fields
        for k in ("deadline_dt", "reg_dt", "bd_status", "psincharge_name",
                  "ctr_type_nm", "currency", "dday_html", "req_name",
                  "classification", "detail_version",
                  "items_count", "attachments_count"):
            if not d.get(k):
                d[k] = meta.get(k)
        # Per-item fields — lookup row's bqms_code in items array
        bqms = (d.get("bqms_code") or "").strip()
        if bqms:
            for it in meta.get("items_arr", []):
                if (it.get("item_code") == bqms or it.get("cis_code") == bqms):
                    if not d.get("first_maker"):
                        d["first_maker"] = (it.get("maker") or "").strip() or None
                    if not d.get("first_moq"):
                        d["first_moq"] = (it.get("moq") or "").strip() or None
                    if not d.get("first_cis_code"):
                        d["first_cis_code"] = (it.get("cis_code") or "").strip() or None
                    if not d.get("first_part_no"):
                        d["first_part_no"] = (it.get("part_no") or "").strip() or None
                    # Per Thang 2026-05-11: surface item-level description
                    # ('JIG-MEASUREMENT', 'Pipe Fitting', ...) so the
                    # frontend cell shows it BESIDES specification.
                    if not d.get("description"):
                        d["description"] = (it.get("description") or "").strip() or None
                    # Override specification with the per-item one if available
                    item_spec = (it.get("specification") or "").strip()
                    if item_spec and (not d.get("specification") or len(d.get("specification") or "") < len(item_spec)):
                        d["specification"] = item_spec
                    break
        return d

    # Phase 2 per Thang 2026-05-12: post-lookup users.full_name for assigned_to
    # so frontend can show "Người PT" without changing the WHERE/JOIN above.
    assigned_user_ids = {r["assigned_to"] for r in rows if r["assigned_to"]}
    assigned_name_map: dict[str, str] = {}
    if assigned_user_ids:
        u_rows = await conn.fetch(
            "SELECT id::text AS id, full_name FROM users WHERE id = ANY($1::uuid[])",
            list(assigned_user_ids),
        )
        assigned_name_map = {u["id"]: u["full_name"] for u in u_rows}

    def _add_assigned_name(d: dict) -> dict:
        uid = d.get("assigned_to")
        d["assigned_to_name"] = assigned_name_map.get(uid) if uid else None
        # Phase E (Thang 2026-05-13): classification_override > auto
        # Frontend expects `classification` field; we preserve override info
        # separately so UI can show "user-edited" badge.
        ovr = d.get("classification_override")
        if ovr:
            d["classification_auto"] = d.get("classification")
            d["classification"] = ovr
            d["classification_is_override"] = True
        else:
            d["classification_is_override"] = False
        return d

    # Phase F (Thang 2026-05-13): User muốn thứ tự match sec-bqms — QT mới
    # nhất (regDt mới nhất) lên đầu, KHÔNG quan trọng là pending hay approved.
    # Sort key priority: reg_dt → inquiry_date → effective_date → created_at.
    items_serialized = (
        [_serialize(p) for p in pending_bidding]
        + [_add_assigned_name(_enrich_with_staging(_serialize(r))) for r in rows]
    )

    # Smart quote scenario classification (Thang 2026-05-18 — TH1/TH2/TH3)
    # Annotate every row with scenario + UI metadata. Used by frontend for
    # badges + wizard default round.
    from app.services.bqms_quote_scenario import (
        classify as _classify_scenario,
        scenario_default_round as _scenario_round,
        scenario_meta as _scenario_meta,
        pushable_round as _pushable_round,
    )

    # Batch 2C: D-N countdown helper. Returns integer days until the submission
    # deadline (ceil), negative when overdue, or None when no parseable deadline.
    def _days_to_deadline(it: dict) -> int | None:
        from app.services.bqms_auto_skip_expired import parse_deadline as _pd
        dl = None
        iso = it.get("deadline_dt")
        if isinstance(iso, str) and iso.strip():
            try:
                dl = datetime.fromisoformat(iso.replace("Z", "+00:00"))
            except ValueError:
                dl = _pd(iso)
        if dl is None:
            # Fall back to the raw Samsung deadline string carried on the row.
            for fld in ("deadline_raw", "deadline_dt"):
                raw = it.get(fld)
                if isinstance(raw, str) and raw.strip():
                    dl = _pd(raw)
                    if dl is not None:
                        break
        if dl is None:
            return None
        from datetime import timezone as _tz
        if dl.tzinfo is None:
            dl = dl.replace(tzinfo=_tz.utc)
        delta = dl - datetime.now(_tz.utc)
        # Ceil so "12 hours left" reads as 1 day, "1h overdue" reads as 0.
        return math.ceil(delta.total_seconds() / 86400.0)

    for _it in items_serialized:
        # Samsung-side round: detail_version (from staging _detail.version)
        # falls back to bqms_rfq.version.
        _samsung_round = None
        for _k in ("detail_version", "version"):
            _v = _it.get(_k)
            try:
                if _v is not None and str(_v).strip() != "":
                    _samsung_round = int(_v)
                    break
            except (ValueError, TypeError):
                pass
        _scenario = _classify_scenario(
            quoted_price_bqms_v1=_it.get("quoted_price_bqms_v1"),
            quoted_price_bqms_v2=_it.get("quoted_price_bqms_v2"),
            version=_it.get("version"),
            data_source=_it.get("data_source"),
            samsung_round=_samsung_round,
        )
        _it["scenario"] = _scenario
        _it["scenario_default_round"] = _scenario_round(_scenario, _samsung_round or 1)
        _it["scenario_meta"] = _scenario_meta(_scenario)
        # pushable_round = round to PUSH (highest filled V). Different from
        # scenario_default_round (which is for NEW quote form). Bug 2026-05-20:
        # using scenario_default_round for the push button jumped V1→V2 right
        # after user generated V1, because TH3.scenario_default_round = max(2, ...).
        _it["pushable_round"] = _pushable_round(
            _it.get("quoted_price_bqms_v1"),
            _it.get("quoted_price_bqms_v2"),
            _it.get("quoted_price_bqms_v3"),
            _it.get("quoted_price_bqms_v4"),
        )
        # Batch 2C: D-N countdown — days from now until deadline_dt. Prefer the
        # persisted bqms_rfq.deadline_dt (ISO string post-_serialize); fall back
        # to parsing the raw Samsung deadline string carried on the row. Negative
        # = overdue. None when no parseable deadline.
        _it["days_to_deadline"] = _days_to_deadline(_it)

    def _parse_reg_dt(s: str | None) -> datetime | None:
        """Parse sec-bqms regDt strings like '(GMT+07:00) 5/15/2026 17:00' or
        '5/15/2026 17:00' or '5/15/2026'. Returns datetime or None."""
        if not s:
            return None
        import re as _re_local
        m = _re_local.search(
            r"(\d{1,2})/(\d{1,2})/(\d{4})(?:\s+(\d{1,2}):(\d{2}))?",
            s,
        )
        if not m:
            return None
        try:
            mo, dd, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
            hh = int(m.group(4) or 0)
            mn = int(m.group(5) or 0)
            return datetime(yr, mo, dd, hh, mn)
        except Exception:
            return None

    def _date_part(it: dict) -> datetime:
        """Primary date: reg_dt → inquiry_date → effective_date → created_at."""
        rd = _parse_reg_dt(it.get("reg_dt"))
        if rd:
            return rd
        for fld in ("inquiry_date", "effective_date"):
            v = it.get(fld)
            if isinstance(v, str) and v:
                try:
                    return datetime.fromisoformat(v.replace("Z", "+00:00").split("+")[0])
                except Exception:
                    pass
        ca = it.get("created_at")
        if isinstance(ca, str) and ca:
            try:
                return datetime.fromisoformat(ca.replace("Z", "+00:00").split("+")[0])
            except Exception:
                pass
        return datetime.min

    def _sort_key(it: dict) -> tuple:
        """Sort by (date DESC, rfq_number DESC). reg_dt thường chỉ có ngày
        không kèm giờ → cần tiebreaker. sec-bqms registers QT theo thứ tự
        tăng dần (QT26062664 > QT26062316 nghĩa là 664 đăng ký SAU 316),
        nên dùng rfq_number DESC làm tiebreaker để khớp UI sec-bqms."""
        return (_date_part(it), it.get("rfq_number") or "")

    items_serialized.sort(key=_sort_key, reverse=True)

    return {
        "data": {
            "items": items_serialized,
            "total": total + len(pending_bidding),
            "pending_bidding_count": len(pending_bidding),
            "page": page,
            "page_size": page_size,
            "total_pages": math.ceil(total / page_size) if total > 0 else 1,
            "kpis": {
                "total_month": total,
                "won": won_count,
                "lost": lost_count,
                "pending": pending_count + len(pending_bidding),
                "win_rate": win_rate,
            },
            "months": months_data,
        }
    }


# ---------------------------------------------------------------------------
# RFQ — Inline Price Edit
# ---------------------------------------------------------------------------

_PRICE_ALLOWED_FIELDS = frozenset({
    "quoted_price_bqms_v1",
    "quoted_price_bqms_v2",
    "quoted_price_bqms_v3",
    "quoted_price_bqms_v4",
    "purchase_price_rmb",
    "purchase_price_vnd",
    "quoted_price_ama",
    "notes",
})


# Phase E (Thang 2026-05-13): User-editable classification override
@router.patch("/rfq/{rfq_id}/classification")
async def update_rfq_classification(
    rfq_id: int,
    body: dict,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales",
        "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """User override TM/GC classification. body = {classification: 'TM'|'GC'|null}.
    null hoặc 'auto' = revert về auto-detect (set classification_override=NULL).
    """
    val = body.get("classification")
    if isinstance(val, str):
        val = val.strip().upper()
        if val in ("AUTO", ""):
            val = None
        elif val not in ("TM", "GC"):
            raise HTTPException(400, "classification phải là 'TM', 'GC', hoặc null/auto")
    elif val is not None:
        raise HTTPException(400, "classification phải là string hoặc null")

    row = await conn.fetchrow(
        "UPDATE bqms_rfq SET classification_override = $1, updated_at = NOW() "
        "WHERE id = $2 RETURNING id, classification_override, notes",
        val, rfq_id,
    )
    if not row:
        raise HTTPException(404, f"RFQ #{rfq_id} không tồn tại")

    # Audit
    try:
        await conn.execute(
            """
            INSERT INTO audit_log (user_id, action, table_name, record_id, new_data, created_at)
            VALUES ($1::uuid, 'bqms.classification_override', 'bqms_rfq', $2, $3::jsonb, NOW())
            """,
            token_data.user_id, str(rfq_id),
            _json.dumps({"classification_override": val}),
        )
    except Exception as _exc:
        logger.warning("audit_log classification failed: %s", _exc)

    return {
        "data": {
            "id": row["id"],
            "classification_override": row["classification_override"],
            "classification_is_override": row["classification_override"] is not None,
        },
        "message": f"Classification đã set = {val or 'auto'}",
    }


@router.patch("/rfq/{rfq_id}/price")
async def update_rfq_price(
    rfq_id: int,
    body: dict,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """
    Inline price/notes edit for a single bqms_rfq row.
    body: {"field": "quoted_price_bqms_v1", "value": 15000}
    """
    field = body.get("field")
    if field not in _PRICE_ALLOWED_FIELDS:
        raise HTTPException(
            status_code=400,
            detail=f"Trường '{field}' không được phép chỉnh sửa. Cho phép: {sorted(_PRICE_ALLOWED_FIELDS)}",
        )

    value = body.get("value")

    # Validate numeric fields
    if field != "notes" and value is not None:
        try:
            value = float(value)
        except (TypeError, ValueError):
            raise HTTPException(
                status_code=400,
                detail="Giá trị phải là số",
            )

    # Thang 2026-05-15: khi user inline-edit V1-V4 → set Người PT = user đó.
    # Áp dụng cho các field giá báo (quoted_price_bqms_v*) và notes nếu nó là
    # hành động báo giá thực sự.
    _PRICE_ROUND_FIELDS = {
        "quoted_price_bqms_v1", "quoted_price_bqms_v2",
        "quoted_price_bqms_v3", "quoted_price_bqms_v4",
    }
    update_assignee = field in _PRICE_ROUND_FIELDS and value is not None
    if update_assignee:
        result = await conn.execute(
            f"UPDATE bqms_rfq SET {field} = $1, assigned_to = $3::uuid, "
            f"updated_at = NOW() WHERE id = $2",
            value, rfq_id, token_data.user_id,
        )
    else:
        result = await conn.execute(
            f"UPDATE bqms_rfq SET {field} = $1, updated_at = NOW() WHERE id = $2",
            value, rfq_id,
        )

    if result == "UPDATE 0":
        raise HTTPException(status_code=404, detail=f"Không tìm thấy RFQ #{rfq_id}")

    # Write to bqms_quote_log for vN price fields (round audit trail)
    # This powers the daily morning report's "báo giá hôm nay" counts.
    _ROUND_FIELDS = {
        "quoted_price_bqms_v1": 1,
        "quoted_price_bqms_v2": 2,
        "quoted_price_bqms_v3": 3,
        "quoted_price_bqms_v4": 4,
    }
    if field in _ROUND_FIELDS and value is not None:
        try:
            # Fetch current item_type so log captures TM/GC snapshot
            item_type = await conn.fetchval(
                "SELECT item_type FROM bqms_rfq WHERE id = $1", rfq_id
            )
            await conn.execute(
                """
                INSERT INTO bqms_quote_log
                  (rfq_id, round, quoted_price, quoted_currency, item_type, quoted_by, notes)
                VALUES ($1, $2, $3, 'USD', $4, $5, 'inline-edit from RFQ table')
                """,
                rfq_id, _ROUND_FIELDS[field], float(value), item_type, token_data.user_id,
            )
        except Exception as exc:
            logger.warning("quote_log insert failed for rfq=%s: %s", rfq_id, exc)

    logger.info(
        "RFQ price updated: id=%d, field=%s, by=%s", rfq_id, field, token_data.user_id
    )
    return {"message": "Đã cập nhật"}


# ---------------------------------------------------------------------------
# RFQ — Skip / Unskip (per Thang 2026-05-11)
# Mark a bqms_rfq row as 'skipped' (= "không báo giá nữa"). Also propagates
# to bqms_vendor_portal_staging if a matching staging row exists, so the
# row stops appearing in pending lists.
# ---------------------------------------------------------------------------

@router.post("/rfq/{rfq_id}/skip")
async def skip_rfq(
    rfq_id: int,
    body: dict | None = None,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Mark RFQ as 'skipped' (set result='skipped' on bqms_rfq + staging.status='skipped').
    Pass {\"unskip\": true} in body to revert: result='pending' + staging.status='pending_review'.
    """
    body = body or {}
    unskip = bool(body.get("unskip"))

    rfq = await conn.fetchrow(
        "SELECT id, rfq_number, result::text AS result FROM bqms_rfq WHERE id = $1",
        rfq_id,
    )
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} không tồn tại")

    new_result = "pending" if unskip else "skipped"
    new_staging_status = "pending_review" if unskip else "skipped"

    await conn.execute(
        "UPDATE bqms_rfq SET result = $1::rfq_result, result_updated_by = $2::uuid, "
        "result_date = CURRENT_DATE, updated_at = NOW() WHERE id = $3",
        new_result, token_data.user_id, rfq_id,
    )

    # Propagate to staging row(s) for the same rfq_number, if any
    affected = 0
    if rfq["rfq_number"]:
        result = await conn.execute(
            "UPDATE bqms_vendor_portal_staging SET status = $1, "
            "reviewed_by = $2::uuid, reviewed_at = NOW() "
            "WHERE rfq_number = $3 AND module = 'bidding'",
            new_staging_status, token_data.user_id, rfq["rfq_number"],
        )
        try:
            affected = int(result.split()[-1]) if result else 0
        except Exception:
            affected = 0

    logger.info(
        "RFQ %sskip: rfq_id=%d rfq_number=%s by=%s staging_affected=%d",
        "un-" if unskip else "", rfq_id, rfq["rfq_number"], token_data.user_id, affected,
    )
    return {
        "message": "Đã bỏ skip" if unskip else "Đã skip RFQ",
        "data": {
            "rfq_id": rfq_id,
            "rfq_number": rfq["rfq_number"],
            "result": new_result,
            "staging_affected": affected,
        },
    }


# ---------------------------------------------------------------------------
# RFQ — Đánh dấu kết quả Thắng/Thua/Đang chờ (Tính năng A)
# ---------------------------------------------------------------------------

@router.patch("/rfq/{rfq_id}/result")
async def update_rfq_result(
    rfq_id: int,
    body: dict,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales",
        "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Đánh dấu kết quả RFQ nội bộ. body = {result: 'won'|'lost'|'pending'}."""
    val = body.get("result")
    if not isinstance(val, str):
        raise HTTPException(400, "result phải là string")
    val = val.strip().lower()
    if val not in ("won", "lost", "pending"):
        raise HTTPException(400, "result phải là 'won', 'lost', hoặc 'pending'")

    row = await conn.fetchrow(
        "UPDATE bqms_rfq SET result = $1::rfq_result, "
        "result_updated_by = $2::uuid, result_date = CURRENT_DATE, "
        "updated_at = NOW() WHERE id = $3 "
        "RETURNING id, rfq_number, result::text AS result",
        val, token_data.user_id, rfq_id,
    )
    if not row:
        raise HTTPException(404, f"RFQ #{rfq_id} không tồn tại")

    # Audit (best-effort, mirror /classification)
    try:
        await conn.execute(
            "INSERT INTO audit_log (user_id, action, table_name, record_id, new_data, created_at) "
            "VALUES ($1::uuid, 'bqms.result_mark', 'bqms_rfq', $2, $3::jsonb, NOW())",
            token_data.user_id, str(rfq_id), _json.dumps({"result": val}),
        )
    except Exception as _exc:
        logger.warning("audit_log result failed: %s", _exc)

    return {
        "message": "Đã cập nhật kết quả",
        "ok": True,
        "result": row["result"],
        "data": {
            "rfq_id": row["id"],
            "rfq_number": row["rfq_number"],
            "result": row["result"],
        },
    }


# ---------------------------------------------------------------------------
# RFQ — Generate Báo giá round 2/3/4 files (per Thang 2026-05-11)
# ---------------------------------------------------------------------------

class _GenerateRoundBody(BaseModel):
    """Optional body for /generate-round. When provided, items override the
    DB-derived items (per-bqms spec/maker/qty/price edits from TM wizard).
    Image overrides are handled separately via /quote-image-override (file
    upload), so we don't need image bytes here — autofill_service reads
    /data/quote-overrides/{rfq}/{code}__product_photo.{ext} automatically.
    """
    model_config = ConfigDict(extra="ignore")
    items: list[dict] | None = None


@router.post("/rfq/{rfq_id}/generate-round")
async def generate_quote_round(
    rfq_id: int,
    round_n: int = Query(..., ge=1, le=4, description="Round number 1-4"),
    flow_type: str = Query("tm", description="tm | gc"),
    new_price: float | None = Query(None, description="Optional: V_n price to set"),
    body: _GenerateRoundBody | None = None,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Generate quotation files for round N (1-4) using current bqms_rfq data.

    Per Thang 2026-05-11: lần 1 dùng /quote (drill detail). Lần 2-4 reuse
    bqms_rfq data + price V_n + write into [QT]_AMA BAC NINH_L{n} subfolder.

    If `new_price` provided: also UPDATE bqms_rfq.quoted_price_bqms_v{n} first.
    """
    if round_n not in (1, 2, 3, 4):
        raise HTTPException(400, "round must be 1-4")

    # Bug #1 fix (Thang 2026-05-19): quote_unlocked guard.
    # Pre-check the RFQ is in a quotable state. Prevents concurrent writes
    # khi 2 user click "Báo giá" cùng lúc.
    pre_row = await conn.fetchrow(
        """
        SELECT quote_unlocked, result::text AS result,
               deadline_dt, result_updated_by,
               quoted_price_bqms_v1, quoted_price_bqms_v2,
               quoted_price_bqms_v3, quoted_price_bqms_v4
          FROM bqms_rfq WHERE id = $1
        """,
        rfq_id,
    )
    if not pre_row:
        raise HTTPException(404, f"RFQ #{rfq_id} không tồn tại")
    # Deadline-aware close guard (Thang 2026-06-24): a machine-set result='closed'
    # (result_updated_by IS NULL) must NOT block an RFQ whose deadline is still
    # live — re-opened round-2 RFQs kept a stale 'closed' from round 1. Treat as
    # blocked only for won/lost, a human-finalized close, or a passed deadline.
    _now = datetime.now(timezone.utc)
    _dl = pre_row.get("deadline_dt")
    if _dl is not None and _dl.tzinfo is not None:
        _deadline_open = _dl >= _now
    elif _dl is not None:
        # asyncpg returned a naive datetime — compare on calendar date instead.
        _deadline_open = _dl.date() >= _now.date()
    else:
        _deadline_open = False
    _human_closed = pre_row.get("result_updated_by") is not None
    if pre_row["result"] in ("won", "lost") or (
        pre_row["result"] == "closed" and (_human_closed or not _deadline_open)
    ):
        raise HTTPException(
            status_code=409,
            detail=f"RFQ đã ở trạng thái '{pre_row['result']}' (deadline đã qua hoặc đã chốt) — không thể báo giá thêm",
        )
    if not pre_row["quote_unlocked"]:
        raise HTTPException(
            409,
            "RFQ chưa được unlock báo giá. Click 'Báo giá' trên staging row trước "
            "(POST /vendor-staging/{id}/quote sẽ set quote_unlocked=true).",
        )

    # 1. Optionally update the V_n price — Bug #2 fix: warn nếu overwrite
    if new_price is not None:
        existing_v = pre_row[f"quoted_price_bqms_v{round_n}"]
        if existing_v is not None and abs(float(existing_v) - float(new_price)) > 0.01:
            logger.warning(
                "Overwriting V%d for rfq=%s: %s → %s (user=%s)",
                round_n, rfq_id, existing_v, new_price, token_data.user_id,
            )
            # Audit log overwrite — leave trail for forensics
            try:
                await conn.execute(
                    """
                    INSERT INTO audit_log
                        (action, table_name, record_id, new_data, created_at)
                    VALUES ('bqms.rfq.price_overwrite', 'bqms_rfq', $1, $2::jsonb, NOW())
                    """,
                    str(rfq_id),
                    _json.dumps({
                        "round_n": round_n,
                        "old_price": float(existing_v),
                        "new_price": float(new_price),
                        "user_id": str(token_data.user_id),
                    }),
                )
            except Exception as exc:
                logger.warning("overwrite audit failed: %s", exc)

        # Thang 2026-05-15: cũng set assigned_to để Người PT = user vừa báo giá
        # Thang 2026-06-13: pin quoted_dt_v{n} = CURRENT_DATE so XLSX cell C4 + UI
        # chip "V{n}: dd/mm" match exactly the submission date.
        await conn.execute(
            f"UPDATE bqms_rfq SET quoted_price_bqms_v{round_n} = $1, "
            f"quoted_dt_v{round_n} = CURRENT_DATE, "
            f"assigned_to = $3::uuid, updated_at = NOW() WHERE id = $2",
            float(new_price), rfq_id, token_data.user_id,
        )
        # Phase G (Thang 2026-05-13): log lịch sử báo giá để Detail page hiện
        # đầy đủ V1→V4 timeline. Bug #5 fix: dùng currency thống nhất 'USD'
        # (xnk_price_lookup + samsung_po unit_price đều USD; VND chỉ dùng
        # cho purchase_price_vnd).
        try:
            await conn.execute(
                """
                INSERT INTO bqms_quote_log
                    (rfq_id, round, quoted_price, quoted_currency, item_type, quoted_by, notes)
                VALUES ($1, $2, $3, 'USD', $4, $5, $6)
                """,
                rfq_id, round_n, float(new_price),
                'GC' if (flow_type or 'tm').lower() == 'gc' else 'TM',
                token_data.user_id,
                f"L{round_n} from generate-round endpoint (flow={flow_type})",
            )
        except Exception as exc:
            logger.warning("quote_log insert failed for rfq=%s round=%d: %s",
                           rfq_id, round_n, exc)

    # 2. Fetch all items for this RFQ (group by rfq_number)
    rfq = await conn.fetchrow(
        "SELECT rfq_number FROM bqms_rfq WHERE id = $1", rfq_id,
    )
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} not found")
    rfq_number = rfq["rfq_number"]

    rows = await conn.fetch(
        """
        SELECT id, rfq_number, bqms_code, specification, maker,
               expected_qty, unit, supplier_name, person_in_charge_name,
               quoted_price_bqms_v1, quoted_price_bqms_v2,
               quoted_price_bqms_v3, quoted_price_bqms_v4,
               notes
        FROM bqms_rfq
        WHERE rfq_number = $1
        ORDER BY id
        """,
        rfq_number,
    )
    if not rows:
        raise HTTPException(404, f"No items for RFQ {rfq_number}")

    # Build items in autofill_service shape
    price_field = f"quoted_price_bqms_v{round_n}"
    items = []
    for r in rows:
        # Detect TM/GC from notes (etl rows store classification=TM/GC there)
        loai = "GC" if "classification=GC" in (r["notes"] or "") else "TM"
        v_n = r[price_field]
        items.append({
            "don_hang": r["rfq_number"],
            "bqms": r["bqms_code"],
            "spec": r["specification"],
            "short_name": (r["specification"] or "")[:40],
            "maker": r["maker"],
            "so_luong": r["expected_qty"],
            "don_vi": r["unit"],
            "loai_hang": loai,
            "unit_price": v_n,
            "suggested_price": v_n,
        })

    # Thang 2026-05-15: TM wizard preview allows per-item override of
    # spec/maker/qty/price before file generation. When body.items is
    # provided, overlay overrides onto DB-derived items keyed by bqms code.
    # Image overrides are handled via /quote-image-override (separate
    # endpoint) — autofill_service reads /data/quote-overrides automatically.
    if body and body.items:
        override_map: dict[str, dict] = {}
        for it in body.items:
            code = (it.get("bqms") or it.get("bqms_code") or "").strip()
            if code:
                override_map[code] = it
        for it in items:
            ov = override_map.get((it["bqms"] or "").strip())
            if not ov:
                continue
            if "spec" in ov and ov["spec"] is not None:
                it["spec"] = str(ov["spec"])
                it["short_name"] = it["spec"][:40]
            if "maker" in ov and ov["maker"] is not None:
                it["maker"] = str(ov["maker"])
            if "so_luong" in ov and ov["so_luong"] not in (None, ""):
                try:
                    it["so_luong"] = int(float(ov["so_luong"]))
                except (TypeError, ValueError):
                    pass
            if "unit_price" in ov and ov["unit_price"] not in (None, ""):
                try:
                    p = float(ov["unit_price"])
                    it["unit_price"] = p
                    it["suggested_price"] = p
                except (TypeError, ValueError):
                    pass

        # Thang 2026-05-15 (Issue 10 + 12): persist toàn bộ chỉnh sửa từ TM
        # wizard vào bqms_rfq — giá V_n + Người PT + spec/maker/qty. Lần sau
        # user mở lại wizard, /quotations/lookup sẽ trả về dữ liệu đã sửa.
        price_col = f"quoted_price_bqms_v{round_n}"
        for it in items:
            code = (it.get("bqms") or "").strip()
            ov = override_map.get(code)
            if not code or not ov:
                continue
            sets: list[str] = []
            args: list[Any] = []
            try:
                p = float(ov.get("unit_price") or 0)
            except (TypeError, ValueError):
                p = 0.0
            if p > 0:
                args.append(p)
                sets.append(f"{price_col} = ${len(args)}")
                args.append(token_data.user_id)
                sets.append(f"assigned_to = ${len(args)}::uuid")
            if "spec" in ov and ov["spec"] is not None:
                args.append(str(ov["spec"]))
                sets.append(f"specification = ${len(args)}")
            if "maker" in ov and ov["maker"] is not None:
                args.append(str(ov["maker"]))
                sets.append(f"maker = ${len(args)}")
            if "so_luong" in ov and ov["so_luong"] not in (None, ""):
                try:
                    args.append(int(float(ov["so_luong"])))
                    sets.append(f"expected_qty = ${len(args)}")
                except (TypeError, ValueError):
                    pass
            if not sets:
                continue
            args.append(rfq_number)
            args.append(code)
            sql = (
                f"UPDATE bqms_rfq SET {', '.join(sets)}, updated_at = NOW() "
                f"WHERE rfq_number = ${len(args)-1} AND bqms_code = ${len(args)}"
            )
            await conn.execute(sql, *args)

    # 3. Run autofill_job with round_n in output_dir name
    from app.services.tools.autofill_service import run_autofill_job

    cam_ket_tpl = await conn.fetchval(
        "SELECT file_path FROM quotation_templates "
        "WHERE template_type = 'cam_ket' AND is_default = true LIMIT 1"
    )
    commercial_tpl = await conn.fetchval(
        "SELECT file_path FROM quotation_templates "
        "WHERE template_type = 'commercial' AND is_default = true LIMIT 1"
    )

    # Insert quotation row first to get quotation_id
    quotation_id = await conn.fetchval(
        """
        INSERT INTO quotations
            (rfq_no, source_type, items, total_items, created_by, status, flow_type, quote_level)
        VALUES ($1, 'excel', $2::jsonb, $3, $4::uuid, 'processing', $5, $6)
        RETURNING id
        """,
        rfq_number,
        _json.dumps(items, default=str, ensure_ascii=False),
        len(items),
        token_data.user_id,
        flow_type,
        round_n,
    )

    # Patch the autofill_service to write into L{round_n} subfolder.
    # Thang 2026-05-21: passes round_n directly to autofill (replaces the
    # old "always write L1, then shutil.move to L{round_n}" hack which was
    # destroying previously-generated V{round_n} files on each regenerate).
    # Inside autofill, quote_round_subfolder() now archives any pre-existing
    # L{round_n} folder to `.archived_<ts>/` so V1 history is preserved when
    # user re-generates V1 (and so on).

    result = await run_autofill_job(
        conn=conn,
        quotation_id=quotation_id,
        items=items,
        cam_ket_template=cam_ket_tpl,
        commercial_template=commercial_tpl,
        flow_type=flow_type,
        round_n=round_n,
    )

    return {
        "data": {
            "quotation_id": quotation_id,
            "rfq_id": rfq_id,
            "rfq_number": rfq_number,
            "round_n": round_n,
            "flow_type": flow_type,
            "new_price_set": new_price,
            "files": result.get("files", []),
            "errors": result.get("errors", []),
        },
        "message": f"Đã tạo file báo giá lần {round_n} cho {rfq_number}",
    }


# ---------------------------------------------------------------------------
# Rename file / folder inside RFQ folder — Thang 2026-05-15
# ---------------------------------------------------------------------------


@router.get("/rfq/{rfq_id}/subfolders")
async def list_rfq_subfolders(
    rfq_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List subfolders inside RFQ root (excluding raw/ and images/).

    Used by the rename-folder UI to pick which L{n} folder to rename.
    """
    from app.etl.bqms_bidding_scraper import find_existing_rfq_folder

    rfq = await conn.fetchrow("SELECT rfq_number FROM bqms_rfq WHERE id = $1", rfq_id)
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} not found")
    root = find_existing_rfq_folder(rfq["rfq_number"])
    if root is None or not root.exists():
        return {"data": {"root": None, "subfolders": []}}
    EXCLUDE = {"raw", "images"}
    subs = []
    try:
        for child in sorted(root.iterdir(), key=lambda p: p.name.lower()):
            if child.is_dir() and child.name not in EXCLUDE:
                subs.append({
                    "name": child.name,
                    "path": str(child),
                })
    except OSError as exc:
        logger.warning("list_rfq_subfolders: %s", exc)
    return {"data": {"root": str(root), "subfolders": subs}}


class _RenameBody(BaseModel):
    """Body for rename-file / rename-folder endpoints.

    old_path: absolute path on disk (returned by /bidding/folder or
              quotation history). MUST resolve inside the RFQ folder root
              (no path traversal).
    new_name: just the basename (no slashes). Extension is preserved
              automatically if user omits it. For folders, no extension
              expected.
    """
    model_config = ConfigDict(extra="ignore")
    old_path: str
    new_name: str


def _safe_resolve_in_root(target: Path, root: Path) -> Path:
    """Resolve target under root, refusing path-traversal."""
    try:
        resolved = target.resolve()
        resolved.relative_to(root.resolve())
        return resolved
    except (ValueError, OSError) as exc:
        raise HTTPException(400, f"Path outside RFQ folder: {target} (root={root})") from exc


@router.post("/rfq/{rfq_id}/rename-file")
async def rename_rfq_file(
    rfq_id: int,
    body: _RenameBody,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Rename a file inside the RFQ folder (Excel/PDF/attachment).

    Updates `quotations.output_xlsx` / `output_pdf` rows if the renamed
    path matches.
    """
    from app.etl.bqms_bidding_scraper import find_existing_rfq_folder

    # Locate RFQ root folder via DB → find_existing_rfq_folder
    rfq = await conn.fetchrow(
        "SELECT rfq_number FROM bqms_rfq WHERE id = $1", rfq_id,
    )
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} not found")
    rfq_number = rfq["rfq_number"]
    root = find_existing_rfq_folder(rfq_number)
    if root is None or not root.exists():
        raise HTTPException(404, f"RFQ folder not found for {rfq_number}")

    # Validate new_name: no path separators, not empty, length cap
    new_name_raw = (body.new_name or "").strip()
    if not new_name_raw:
        raise HTTPException(400, "Tên mới không được rỗng")
    if "/" in new_name_raw or "\\" in new_name_raw or new_name_raw in (".", ".."):
        raise HTTPException(400, "Tên mới không được chứa / hoặc \\")
    if len(new_name_raw) > 200:
        raise HTTPException(400, "Tên mới quá dài (>200 ký tự)")

    old_p = Path(body.old_path)
    old_resolved = _safe_resolve_in_root(old_p, root)
    if not old_resolved.exists():
        raise HTTPException(404, f"File không tồn tại: {old_p.name}")
    if not old_resolved.is_file():
        raise HTTPException(400, "Đường dẫn này không phải file (có thể là folder — dùng /rename-folder)")

    # Preserve original extension if user omitted it
    new_name = new_name_raw
    old_suffix = old_resolved.suffix.lower()
    if old_suffix and Path(new_name).suffix.lower() != old_suffix:
        new_name = new_name + old_suffix

    new_path = old_resolved.parent / new_name
    if new_path == old_resolved:
        return {"data": {"old": str(old_resolved), "new": str(new_path)}, "message": "Tên không đổi"}
    if new_path.exists():
        raise HTTPException(409, f"File đã tồn tại: {new_name}")

    old_resolved.rename(new_path)
    logger.info("rename-file: %s → %s (user=%s)", old_resolved, new_path, token_data.email)

    # Update DB references in quotations table
    try:
        n1 = await conn.execute(
            "UPDATE quotations SET output_xlsx = $1 WHERE output_xlsx = $2",
            str(new_path), str(old_resolved),
        )
        n2 = await conn.execute(
            "UPDATE quotations SET output_pdf = $1 WHERE output_pdf = $2",
            str(new_path), str(old_resolved),
        )
        logger.info("rename-file: DB updates xlsx=%s pdf=%s", n1, n2)
    except Exception as exc:
        logger.warning("rename-file: DB update failed (rename still applied on disk): %s", exc)

    return {
        "data": {"old_path": str(old_resolved), "new_path": str(new_path), "new_name": new_name},
        "message": f"Đã đổi tên: {old_resolved.name} → {new_name}",
    }


@router.post("/rfq/{rfq_id}/rename-folder")
async def rename_rfq_folder(
    rfq_id: int,
    body: _RenameBody,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Rename a subfolder INSIDE the RFQ folder root (e.g. _AMA BAC NINH_L1).

    Renaming the RFQ root itself is NOT allowed — that breaks
    find_existing_rfq_folder + scraper which match by RFQ number prefix.
    Updates `quotations.output_xlsx` / `output_pdf` rows whose paths start
    with the renamed folder prefix.
    """
    from app.etl.bqms_bidding_scraper import find_existing_rfq_folder

    rfq = await conn.fetchrow("SELECT rfq_number FROM bqms_rfq WHERE id = $1", rfq_id)
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} not found")
    rfq_number = rfq["rfq_number"]
    root = find_existing_rfq_folder(rfq_number)
    if root is None or not root.exists():
        raise HTTPException(404, f"RFQ folder not found for {rfq_number}")

    new_name = (body.new_name or "").strip()
    if not new_name:
        raise HTTPException(400, "Tên mới không được rỗng")
    if "/" in new_name or "\\" in new_name or new_name in (".", ".."):
        raise HTTPException(400, "Tên mới không được chứa / hoặc \\")
    if len(new_name) > 200:
        raise HTTPException(400, "Tên mới quá dài (>200 ký tự)")

    old_p = Path(body.old_path)
    old_resolved = _safe_resolve_in_root(old_p, root)
    if not old_resolved.exists():
        raise HTTPException(404, f"Folder không tồn tại: {old_p.name}")
    if not old_resolved.is_dir():
        raise HTTPException(400, "Đường dẫn này không phải folder (có thể là file — dùng /rename-file)")
    # Forbid renaming the RFQ root itself (scraper / image lookup relies on prefix match)
    if old_resolved.resolve() == root.resolve():
        raise HTTPException(
            400,
            "Không thể đổi tên folder gốc RFQ (sẽ làm hỏng scrape / image lookup). "
            "Chỉ đổi tên các subfolder bên trong.",
        )

    new_path = old_resolved.parent / new_name
    if new_path == old_resolved:
        return {"data": {"old": str(old_resolved), "new": str(new_path)}, "message": "Tên không đổi"}
    if new_path.exists():
        raise HTTPException(409, f"Folder đã tồn tại: {new_name}")

    old_resolved.rename(new_path)
    logger.info("rename-folder: %s → %s (user=%s)", old_resolved, new_path, token_data.email)

    # Update DB references: any output_xlsx / output_pdf path that lived
    # under the renamed folder must be rewritten to the new prefix.
    old_prefix = str(old_resolved) + "/"
    new_prefix = str(new_path) + "/"
    try:
        await conn.execute(
            "UPDATE quotations SET output_xlsx = $2 || substr(output_xlsx, length($1) + 1) "
            "WHERE output_xlsx LIKE $1 || '%'",
            old_prefix, new_prefix,
        )
        await conn.execute(
            "UPDATE quotations SET output_pdf = $2 || substr(output_pdf, length($1) + 1) "
            "WHERE output_pdf LIKE $1 || '%'",
            old_prefix, new_prefix,
        )
    except Exception as exc:
        logger.warning("rename-folder: DB update failed: %s", exc)

    return {
        "data": {"old_path": str(old_resolved), "new_path": str(new_path), "new_name": new_name},
        "message": f"Đã đổi tên folder: {old_resolved.name} → {new_name}",
    }


# ---------------------------------------------------------------------------
# GC Quote Wizard — per Thang 2026-05-11
# ---------------------------------------------------------------------------

from pydantic import BaseModel as _WizardBase, Field as _WizardField  # noqa: E402


@router.get("/rfq/{rfq_id}/wizard-items")
async def rfq_wizard_items(
    rfq_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List items + pre-fill saved wizard state (materials/parts/processes).

    Thang 2026-05-15 (Issue 12): khi user mở lại GC wizard cho 1 RFQ đã báo
    giá trước đó, các trường materials/parts/others/processes/nego/jig_name
    được load từ `quotations.items` (lần báo giá GC gần nhất) để user không
    phải gõ lại từ đầu.
    """
    rfq = await conn.fetchrow(
        "SELECT rfq_number FROM bqms_rfq WHERE id = $1", rfq_id,
    )
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} not found")
    rows = await conn.fetch(
        "SELECT bqms_code, specification, expected_qty "
        "FROM bqms_rfq WHERE rfq_number = $1 ORDER BY id",
        rfq["rfq_number"],
    )

    # Load latest GC quotation items (if any) — pre-fill wizard
    prior_items: dict[str, dict] = {}
    try:
        prior_row = await conn.fetchrow(
            """
            SELECT items FROM quotations
            WHERE rfq_no = $1 AND flow_type = 'gc' AND deleted_at IS NULL
            ORDER BY created_at DESC LIMIT 1
            """,
            rfq["rfq_number"],
        )
        if prior_row and prior_row["items"]:
            raw = prior_row["items"]
            if isinstance(raw, str):
                raw = _json.loads(raw)
            if isinstance(raw, list):
                for it in raw:
                    if isinstance(it, dict):
                        code = (it.get("bqms_code") or it.get("bqms") or "").strip()
                        if code:
                            prior_items[code] = it
    except Exception as exc:
        logger.warning("wizard-items: load prior GC quote failed: %s", exc)

    def _enrich(r) -> dict:
        code = r["bqms_code"] or ""
        out = {
            "bqms_code": code,
            "spec": r["specification"] or "",
            "qty": int(r["expected_qty"] or 1),
        }
        prior = prior_items.get(code)
        if prior:
            # carry over wizard-specific structured fields
            for key in ("jig_name", "materials", "parts", "others", "processes", "nego"):
                if key in prior:
                    out[key] = prior[key]
        return out

    return {"data": [_enrich(r) for r in rows]}


@router.post("/rfq/{rfq_id}/force-rescan")
async def force_rescan_rfq(
    rfq_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Force drill detail + download attachments + extract images cho 1 RFQ.

    Thang 2026-05-15 (Issue 14): user bấm "Quét ngay" trên image section khi
    folder chưa có. Chạy sync (~30-60s) qua samsung_session_lock để không
    đụng cron scrape đang chạy.

    Idempotent: nếu folder/files đã có rồi → return ngay không drill lại.
    """
    rfq = await conn.fetchrow(
        "SELECT rfq_number FROM bqms_rfq WHERE id = $1", rfq_id,
    )
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} not found")
    rfq_number = rfq["rfq_number"]

    # Find staging row (need raw_json for download_files_for_rfq)
    staging = await conn.fetchrow(
        """
        SELECT id, raw_json FROM bqms_vendor_portal_staging
        WHERE module='bidding' AND rfq_number = $1
        ORDER BY id DESC LIMIT 1
        """,
        rfq_number,
    )
    if not staging:
        raise HTTPException(404, f"No staging row for {rfq_number} — chờ cron list-scrape")

    raw_row = staging["raw_json"]
    if isinstance(raw_row, str):
        raw_row = _json.loads(raw_row or "{}")
    if not isinstance(raw_row, dict):
        raw_row = {}

    # Skip if already drilled
    from app.etl.bqms_bidding_scraper import find_existing_rfq_folder, download_files_for_rfq
    folder = find_existing_rfq_folder(rfq_number)
    has_images = bool(folder and (folder / "images").exists() and any(
        (folder / "images").glob("*.png")
    ))
    has_detail = bool((raw_row.get("_detail") or {}).get("items"))
    if folder and has_images and has_detail:
        return {
            "data": {"rfq_number": rfq_number, "folder": str(folder), "already_drilled": True},
            "message": "Folder + ảnh + detail đã có sẵn",
        }

    # Acquire Samsung session lock then drill
    from app.services.samsung_session_lock import samsung_session_lock
    from app.core.config import settings as cfg
    import asyncpg as apg
    db_url = (
        str(cfg.DATABASE_URL)
        .replace("+asyncpg", "")
        .replace("postgresql+asyncpg", "postgresql")
    )
    pool = await apg.create_pool(db_url, min_size=1, max_size=2)
    try:
        async with samsung_session_lock(pool, who=f"force-rescan-{rfq_number}",
                                        timeout_seconds=180):
            result = await download_files_for_rfq(rfq_number, raw_row, db_pool=pool)

        # Merge fresh_detail into staging.raw_json + upsert bqms_rfq
        if result.get("fresh_detail"):
            new_raw = {**raw_row, "_detail": result["fresh_detail"]}
            async with pool.acquire() as c:
                await c.execute(
                    "UPDATE bqms_vendor_portal_staging SET raw_json = $1::jsonb "
                    "WHERE id = $2",
                    _json.dumps(new_raw, default=str, ensure_ascii=False),
                    staging["id"],
                )
            from app.etl.bqms_bidding_scraper import upsert_bqms_rfq_for_one_staging_row
            n = await upsert_bqms_rfq_for_one_staging_row(pool, new_raw)
            result["bqms_rfq_upserts"] = n
    finally:
        await pool.close()

    logger.info(
        "force-rescan RFQ %s by user=%s: files=%d images=%d upserts=%s",
        rfq_number, token_data.email,
        len(result.get("attachments") or []),
        len(result.get("images") or []),
        result.get("bqms_rfq_upserts", "?"),
    )
    return {
        "data": {
            "rfq_number": rfq_number,
            "folder": str(folder) if folder else None,
            "files_downloaded": len(result.get("attachments") or []),
            "images_extracted": len(result.get("images") or []),
            "items_drilled": len((result.get("fresh_detail") or {}).get("items") or []),
        },
        "message": f"Quét xong RFQ {rfq_number}",
    }


class WizardMaterialIn(_WizardBase):
    name: str
    w: float | None = None
    l: float | None = None
    h: float | None = None
    qty: float = 1
    unit_price: float = 0


class WizardPartIn(_WizardBase):
    name: str
    qty: float = 1
    unit_price: float = 0


class WizardOtherIn(_WizardBase):
    description: str = ""
    qty: float = 1
    unit_price: float = 0


class WizardProcessIn(_WizardBase):
    name: str
    time_hr: float = 0
    unit_price: float = 0


class WizardItemIn(_WizardBase):
    bqms_code: str
    jig_name: str
    spec: str = ""
    qty: float = 1
    materials: list[WizardMaterialIn] = _WizardField(default_factory=list)
    parts: list[WizardPartIn] = _WizardField(default_factory=list)
    others: list[WizardOtherIn] = _WizardField(default_factory=list)
    processes: list[WizardProcessIn] = _WizardField(default_factory=list)
    nego: float = 0


class WizardFinalizeIn(_WizardBase):
    rfq_id: int
    round_n: int = _WizardField(..., ge=1, le=4)
    items: list[WizardItemIn] = _WizardField(..., min_length=1)
    # Thang 2026-06-13 (Bug fix T3): yyyy-mm-dd from frontend; None → server
    # falls back to today in VN tz (Asia/Ho_Chi_Minh) inside the GC builder.
    current_date: str | None = None


@router.post("/quote-wizard/finalize-gc")
async def quote_wizard_finalize_gc(
    payload: WizardFinalizeIn,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Build a GC quotation xlsx from wizard data (materials + processes per item).

    Per Thang 2026-05-11: before this, GC quote was auto-generated from
    bqms_rfq data (no material/process detail). Wizard lets user fill in
    real values which get injected into the QUOTATION_GC.xlsx template.
    """
    rfq = await conn.fetchrow(
        "SELECT rfq_number FROM bqms_rfq WHERE id = $1", payload.rfq_id,
    )
    if not rfq:
        raise HTTPException(404, f"RFQ #{payload.rfq_id} not found")
    rfq_number = rfq["rfq_number"]

    # Compute totals per item for storing back into bqms_rfq.quoted_price_bqms_v{n}
    def _calc_total(it: WizardItemIn) -> float:
        mat = sum((m.qty or 0) * (m.unit_price or 0) for m in it.materials)
        parts = sum((p.qty or 0) * (p.unit_price or 0) for p in it.parts)
        others = sum((o.qty or 0) * (o.unit_price or 0) for o in it.others)
        proc = sum((p.time_hr or 0) * (p.unit_price or 0) for p in it.processes)
        sub = mat + parts + others + proc
        mgmt = sub * 0.05
        profit = mgmt
        return sub + mgmt + profit - (it.nego or 0)

    item_totals = {it.bqms_code: _calc_total(it) for it in payload.items}

    # Locate GC template
    gc_template = await conn.fetchval(
        "SELECT file_path FROM quotation_templates "
        "WHERE template_type = 'gc' AND is_default = true LIMIT 1"
    )
    if not gc_template:
        raise HTTPException(500, "GC template not configured in DB")

    # Find or create the per-RFQ folder + L{round_n} subfolder
    from app.etl.bqms_bidding_scraper import (
        find_existing_rfq_folder, quote_round_subfolder, ensure_rfq_folder_on_scrape,
    )
    parent = find_existing_rfq_folder(rfq_number) or ensure_rfq_folder_on_scrape(
        rfq_number, {},
    )
    if parent is None:
        raise HTTPException(500, f"Could not create folder for {rfq_number}")
    out_dir = quote_round_subfolder(parent, rfq_number, payload.round_n)

    # Build the GC xlsx using template + wizard data
    from app.services.tools.gc_template_quotation import (
        fill_gc_quotation_from_wizard,
    )
    import os
    out_xlsx = os.path.join(out_dir, f"QUOTATION_GC_{rfq_number}.xlsx")

    # Discover images for the selected bqms codes:
    #  1) Phase D (Thang 2026-05-12): check /data/quote-overrides/{rfq}/{code}__product_photo.png
    #     first — user-uploaded override has HIGHEST priority
    #  2) Fallback: auto-discovery from RFQ folder images/ subfolder.
    images_map: dict[str, bytes] = {}
    override_dir = Path(f"/data/quote-overrides/{rfq_number}")

    for it in payload.items:
        code = (it.bqms_code or "").strip()
        if not code:
            continue
        # Check override first
        for ext in (".png", ".jpg", ".jpeg"):
            ovr = override_dir / f"{code}__product_photo{ext}"
            if ovr.exists():
                try:
                    images_map[code] = ovr.read_bytes()
                    logger.info("GC wizard image OVERRIDE: %s → %s (%d B)",
                                code, ovr.name, len(images_map[code]))
                except Exception as exc:
                    logger.warning("GC override read failed %s: %s", ovr, exc)
                break

    # Fallback: auto-discover from images/ for codes without override
    images_dir = parent / "images"
    if images_dir.exists():
        codes = sorted(
            (it.bqms_code.strip() for it in payload.items
             if it.bqms_code and it.bqms_code.strip() not in images_map),
            key=lambda s: -len(s),
        )
        used_files: set[str] = set()
        for img_path in sorted(images_dir.glob("*.png")):
            if img_path.name in used_files:
                continue
            for code in codes:
                if code and img_path.stem.startswith(code) and code not in images_map:
                    try:
                        images_map[code] = img_path.read_bytes()
                        used_files.add(img_path.name)
                        logger.info(
                            "GC wizard image: %s → %s (%d B)",
                            code, img_path.name, len(images_map[code]),
                        )
                    except Exception as exc:
                        logger.warning("GC wizard image read failed %s: %s", img_path, exc)
                    break
    logger.info(
        "GC wizard images_map: %d/%d items have images",
        len(images_map), len(payload.items),
    )

    # Thang 2026-06-13 (Bug fix T3): pass current_date from frontend so the
    # date stamped on each GC sheet matches what the user sees in the UI.
    # `None` → builder defaults to today in VN tz inside _coerce_quote_date.
    result = fill_gc_quotation_from_wizard(
        template_path=gc_template,
        wizard_items=[it.model_dump() for it in payload.items],
        images_map=images_map,
        rfq_no=rfq_number,
        output_path=out_xlsx,
        quote_date=payload.current_date,
    )

    # Thang 2026-05-15 (Issue 11): chỉ giữ 1 file Excel COMBINED + nhiều
    # PDF tách per-item. Per-item xlsx chỉ dùng làm input cho Gotenberg
    # conversion → xoá sau khi render PDF xong.
    from app.services.gotenberg_service import convert_xlsx_to_pdf
    per_item_outputs: list[dict[str, str]] = []
    for entry in (result.get("per_item_files") or []):
        per_xlsx = entry["xlsx"]
        per_pdf = per_xlsx.replace(".xlsx", ".pdf")
        try:
            await convert_xlsx_to_pdf(per_xlsx, per_pdf)
            entry["pdf"] = per_pdf
        except Exception as exc:
            logger.warning("Gotenberg per-item conversion failed for %s: %s", per_xlsx, exc)
            entry["pdf"] = None
        # Cleanup: xoá per-item xlsx (user không cần — chỉ giữ combined)
        try:
            Path(per_xlsx).unlink(missing_ok=True)
        except OSError as exc:
            logger.warning("cleanup per-item xlsx %s failed: %s", per_xlsx, exc)
        entry["xlsx"] = None
        per_item_outputs.append(entry)

    # Combined PDF (review): bỏ — user chỉ muốn 1 xlsx combined + PDF tách.
    out_pdf = None

    # INSERT quotation row
    import json as _j
    quotation_id = await conn.fetchval(
        """
        INSERT INTO quotations
            (rfq_no, source_type, items, total_items, created_by, status, flow_type, quote_level,
             output_xlsx, output_pdf)
        VALUES ($1, 'excel', $2::jsonb, $3, $4::uuid, 'completed', 'gc', $5, $6, $7)
        RETURNING id
        """,
        rfq_number,
        _j.dumps([it.model_dump() for it in payload.items], default=str, ensure_ascii=False),
        len(payload.items),
        token_data.user_id,
        payload.round_n,
        out_xlsx,
        out_pdf,
    )

    # Update bqms_rfq.quoted_price_bqms_v{round_n} per item + auto-tracklog
    # current user as the assignee (Phase 2 per Thang 2026-05-12). The "Người PT"
    # column on the BQMS table reads users.full_name via assigned_to.
    for bqms_code, total in item_totals.items():
        # Thang 2026-05-15: assigned_to GHI ĐÈ (không COALESCE) — user nào
        # nhấn báo giá thì cột Người PT cập nhật về user đó.
        # Thang 2026-06-13: also pin quoted_dt_v{n} = CURRENT_DATE for GC flow.
        await conn.execute(
            f"UPDATE bqms_rfq "
            f"SET quoted_price_bqms_v{payload.round_n} = $1, "
            f"    quoted_dt_v{payload.round_n} = CURRENT_DATE, "
            f"    assigned_to = $4::uuid, "
            f"    updated_at = NOW() "
            f"WHERE rfq_number = $2 AND bqms_code = $3",
            float(total), rfq_number, bqms_code, token_data.user_id,
        )

    # Thang 2026-05-15 (Issue 11): 1 file Excel combined + N file PDF per-item.
    # Bỏ gc_quotation_pdf (combined) + gc_quotation_xlsx_item (tách) khỏi list.
    files = [{"type": "gc_quotation_xlsx", "path": out_xlsx}]
    for e in per_item_outputs:
        if e.get("pdf"):
            files.append({
                "type": "gc_quotation_pdf_item",
                "path": e["pdf"],
                "bqms_code": e["bqms_code"],
            })

    # Audit log per item — dashboard "Tổng quan" reads these to show
    # today's quoting activity. Per Thang 2026-05-11 + Phase 4.1 (Thang 2026-05-12):
    # ALSO insert into bqms_quote_log so daily_report.morning_report() picks up
    # GC submissions (currently it only sees inline-edit quote logs).
    try:
        for bqms_code, total in item_totals.items():
            await conn.execute(
                """
                INSERT INTO audit_log
                    (user_id, action, table_name, record_id, new_data, created_at)
                VALUES ($1::uuid, 'bqms.quote.gc', 'bqms_rfq', $2, $3::jsonb, NOW())
                """,
                token_data.user_id, f"{rfq_number}:{bqms_code}",
                _json.dumps({
                    "rfq_number": rfq_number,
                    "bqms_code": bqms_code,
                    "round_n": payload.round_n,
                    "total": float(total),
                    "flow": "gc",
                    "quotation_id": quotation_id,
                }),
            )
            # Phase 4.1: link to BQMS via bqms_quote_log so daily report shows
            # GC quotes in "SL báo giá được" breakdown (not just inline edits).
            try:
                rfq_row = await conn.fetchrow(
                    "SELECT id FROM bqms_rfq WHERE rfq_number = $1 AND bqms_code = $2 LIMIT 1",
                    rfq_number, bqms_code,
                )
                if rfq_row:
                    await conn.execute(
                        """
                        INSERT INTO bqms_quote_log
                          (rfq_id, round, quoted_price, quoted_currency, item_type, quoted_by, notes)
                        VALUES ($1, $2, $3, 'VND', 'GC', $4::uuid, $5)
                        """,
                        rfq_row["id"], payload.round_n, float(total),
                        token_data.user_id,
                        f"GC wizard finalize quotation_id={quotation_id}",
                    )
            except Exception as qexc:
                logger.warning("bqms_quote_log insert failed for %s/%s: %s",
                               rfq_number, bqms_code, qexc)
    except Exception as exc:
        logger.warning("audit_log insert failed: %s", exc)

    # Phase 6 (Thang 2026-05-12 "Full Vision"): award +1 EXP per quoted item
    # to user's primary pet. Best-effort — pet bug must NOT fail the quote.
    try:
        from app.services.pet_service import award_exp as _award_pet_exp
        for bqms_code in item_totals.keys():
            await _award_pet_exp(
                conn, str(token_data.user_id),
                "quote_submitted", delta=1,
                source_ref=f"{rfq_number}:{bqms_code}",
            )
    except Exception as exc:
        logger.warning("pet award_exp failed: %s", exc)

    return {
        "data": {
            "quotation_id": quotation_id,
            "rfq_id": payload.rfq_id,
            "rfq_number": rfq_number,
            "round_n": payload.round_n,
            "files": files,
            "totals": item_totals,
            "errors": result.get("errors", []),
        },
        "message": f"Đã tạo báo giá GC lần {payload.round_n} cho {rfq_number} ({len(payload.items)} mã)",
    }


# ---------------------------------------------------------------------------
# Quote file management: re-render PDF after manual xlsx edit
# Per Thang 2026-05-11: allow user to edit saved xlsx and re-export PDF.
# ---------------------------------------------------------------------------

class _RegenPdfIn(_WizardBase):
    xlsx_path: str = _WizardField(..., description="Absolute path to the xlsx file under /data/onedrive-staging/")


@router.post("/quote-file/regen-pdf")
async def quote_file_regen_pdf(
    body: _RegenPdfIn,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Re-render PDF from a (possibly user-edited) xlsx file.

    User flow: download xlsx → edit locally → upload via /file-browser → click
    "Tạo lại PDF" on the table row → this endpoint regenerates the .pdf
    next to the .xlsx using Gotenberg (same converter as initial generation).
    """
    import os
    from pathlib import Path

    # Security: only allow paths under onedrive-staging
    p = Path(body.xlsx_path).resolve()
    allowed_root = Path("/data/onedrive-staging").resolve()
    try:
        p.relative_to(allowed_root)
    except ValueError:
        raise HTTPException(403, "path outside allowed root")
    if not p.exists():
        raise HTTPException(404, f"xlsx not found: {p}")
    if p.suffix.lower() != ".xlsx":
        raise HTTPException(400, "only .xlsx allowed")

    out_pdf = str(p.with_suffix(".pdf"))
    from app.services.gotenberg_service import convert_xlsx_to_pdf
    try:
        await convert_xlsx_to_pdf(str(p), out_pdf)
    except Exception as exc:
        raise HTTPException(500, f"Gotenberg conversion failed: {exc}")

    try:
        await conn.execute(
            """
            INSERT INTO audit_log
                (user_id, action, table_name, record_id, new_data, created_at)
            VALUES ($1::uuid, 'bqms.quote.regen_pdf', 'quotations', $2, $3::jsonb, NOW())
            """,
            token_data.user_id, str(p),
            _json.dumps({"xlsx": str(p), "pdf": out_pdf}),
        )
    except Exception as exc:
        logger.warning("regen audit log failed: %s", exc)

    return {
        "data": {"xlsx": str(p), "pdf": out_pdf, "regenerated": True},
        "message": "Đã render lại PDF",
    }


@router.get("/scrape-control/status")
async def scrape_control_status(
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Read the bqms_periodic_scrape on/off flag."""
    row = await conn.fetchrow(
        "SELECT value::text AS value_text, updated_at "
        "FROM app_config WHERE key='bqms_periodic_scrape_enabled'"
    )
    enabled = False
    if row:
        # value_text is the JSON literal: 'true' or 'false' (lowercase, no quotes)
        v = (row["value_text"] or "").strip().lower()
        enabled = v == "true"
    return {
        "data": {
            "enabled": enabled,
            "updated_at": row["updated_at"].isoformat() if row and row["updated_at"] else None,
        }
    }


@router.post("/scrape-control/toggle")
async def scrape_control_toggle(
    enabled: bool = Query(..., description="true=enable, false=disable"),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Turn the BQMS periodic scrape on/off at runtime."""
    val = "true" if enabled else "false"
    await conn.execute(
        """
        INSERT INTO app_config (key, value, updated_at, updated_by)
        VALUES ('bqms_periodic_scrape_enabled', $1::jsonb, NOW(), $2::uuid)
        ON CONFLICT (key) DO UPDATE SET
            value = EXCLUDED.value,
            updated_at = EXCLUDED.updated_at,
            updated_by = EXCLUDED.updated_by
        """,
        val, token_data.user_id,
    )
    # Audit log
    try:
        await conn.execute(
            """
            INSERT INTO audit_log
                (user_id, action, table_name, record_id, new_data, created_at)
            VALUES ($1::uuid, 'bqms.scrape_toggle', 'app_config', 'bqms_periodic_scrape_enabled', $2::jsonb, NOW())
            """,
            token_data.user_id,
            _json.dumps({"enabled": enabled}),
        )
    except Exception as exc:
        logger.warning("audit log toggle failed: %s", exc)
    return {"data": {"enabled": enabled}, "message": f"Cron scrape đã {'BẬT' if enabled else 'TẮT'}"}


# ─────────────────────────────────────────────────────────────────────────────
# BQMS Scraper Settings (Thang 2026-06-22)
# ─────────────────────────────────────────────────────────────────────────────
# Admin-only console: toggle the 6 Samsung-scraper cron flags + manage the
# Samsung login credentials at RUNTIME (cross-process: sc-api writes app_config,
# sc-worker/sc-scheduler read it via the credential resolver) WITHOUT a restart.
#
# Flag keys live in app_config as bqms_<key>_enabled. Excel auto-import is a
# LOCAL file importer (not a Samsung scraper) so it is intentionally excluded
# from this toggle set.
# ─────────────────────────────────────────────────────────────────────────────

# (ui_key → app_config key). These are the 6 Samsung scrapers, all paused.
_SCRAPER_FLAG_KEYS: dict[str, str] = {
    "periodic_scrape": "bqms_periodic_scrape_enabled",
    "smart_sync":      "bqms_smart_sync_enabled",
    "smart_rescan":    "bqms_smart_rescan_enabled",
    "code_track":      "bqms_code_track_enabled",
    "state_tick":      "bqms_state_tick_enabled",
    "won_sync":        "bqms_won_sync_periodic_enabled",
}


def _coerce_flag(val: Any) -> bool:
    """Normalize a JSONB app_config value (bool / 'true' / 1) to a Python bool."""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().strip('"').lower() in ("true", "1", "yes")
    return False


async def _read_scraper_flags(conn: asyncpg.Connection) -> dict[str, bool]:
    """Read the 6 scraper flags from app_config. Missing row → False (paused)."""
    rows = await conn.fetch(
        "SELECT key, value #>> '{}' AS v FROM app_config WHERE key = ANY($1::text[])",
        list(_SCRAPER_FLAG_KEYS.values()),
    )
    by_key = {r["key"]: r["v"] for r in rows}
    return {
        ui_key: _coerce_flag(by_key.get(cfg_key))
        for ui_key, cfg_key in _SCRAPER_FLAG_KEYS.items()
    }


@router.get("/scraper-settings")
async def get_scraper_settings(
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Admin console snapshot: the 6 scraper flags + credential metadata.

    NEVER returns the Samsung password value — only `password_set` + `source`.
    """
    from app.services.bqms_credentials import get_bqms_credentials_meta

    flags = await _read_scraper_flags(conn)

    meta = get_bqms_credentials_meta()  # {username, password_set, source}

    # When the password override lives in the DB, surface its updated_at.
    cred_updated_at = None
    if meta["source"] == "db":
        row = await conn.fetchrow(
            "SELECT updated_at FROM app_config WHERE key = 'bqms_password'"
        )
        if row and row["updated_at"]:
            cred_updated_at = row["updated_at"].isoformat()

    return {
        "flags": flags,
        "credentials": {
            "username": meta["username"],
            "password_set": meta["password_set"],
            "source": meta["source"],          # 'db' (override) | 'env' (fallback)
            "updated_at": cred_updated_at,
        },
    }


class ScraperFlagUpdate(BaseModel):
    model_config = ConfigDict(extra="forbid")
    key: str | None = None
    value: bool | None = None
    flags: dict[str, bool] | None = None  # bulk form


@router.put("/scraper-settings/flags")
async def update_scraper_flags(
    body: ScraperFlagUpdate,
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Upsert one or many scraper flags. Returns the full new flag set.

    Single:  { "key": "won_sync", "value": true }
    Bulk:    { "flags": { "won_sync": true, "state_tick": false } }
    """
    # Build the set of (ui_key → bool) to apply.
    updates: dict[str, bool] = {}
    if body.flags is not None:
        updates.update(body.flags)
    if body.key is not None:
        if body.value is None:
            raise HTTPException(status_code=400, detail="`value` is required when `key` is given")
        updates[body.key] = body.value

    if not updates:
        raise HTTPException(status_code=400, detail="Provide `key`+`value` or `flags`")

    invalid = [k for k in updates if k not in _SCRAPER_FLAG_KEYS]
    if invalid:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown flag(s): {', '.join(invalid)}. "
                   f"Valid: {', '.join(_SCRAPER_FLAG_KEYS)}",
        )

    # Safety interlock: ENABLING any scraper requires a successful Samsung test-login
    # within the last 24h. This stops re-enabling a scraper with an outdated password
    # (repeated bad logins lock the Samsung account). Disabling is ALWAYS allowed so
    # the kill-switch works even when credentials are wrong.
    if any(bool(v) for v in updates.values()):
        ok_recent = await conn.fetchval(
            """
            SELECT (value #>> '{}')::timestamptz > NOW() - INTERVAL '24 hours'
            FROM app_config WHERE key = 'bqms_last_login_ok_at'
            """
        )
        if not ok_recent:
            raise HTTPException(
                status_code=409,
                detail="Cần Test đăng nhập Samsung thành công (trong 24 giờ) trước khi "
                       "bật scrape — tránh spam mật khẩu cũ làm khoá tài khoản.",
            )

    for ui_key, enabled in updates.items():
        cfg_key = _SCRAPER_FLAG_KEYS[ui_key]
        await conn.execute(
            """
            INSERT INTO app_config (key, value, updated_at, updated_by)
            VALUES ($1, $2::jsonb, NOW(), $3::uuid)
            ON CONFLICT (key) DO UPDATE SET
                value      = EXCLUDED.value,
                updated_at = EXCLUDED.updated_at,
                updated_by = EXCLUDED.updated_by
            """,
            cfg_key, "true" if enabled else "false", token_data.user_id,
        )
        try:
            await conn.execute(
                """
                INSERT INTO audit_log
                    (user_id, action, table_name, record_id, new_data, created_at)
                VALUES ($1::uuid, 'bqms.scraper_flag', 'app_config', $2, $3::jsonb, NOW())
                """,
                token_data.user_id, cfg_key, _json.dumps({"enabled": enabled}),
            )
        except Exception as exc:
            logger.warning("audit log scraper_flag failed: %s", exc)

    flags = await _read_scraper_flags(conn)
    logger.info("bqms scraper flags updated by %s: %s", token_data.email, updates)
    return {"flags": flags}


class ScraperCredsUpdate(BaseModel):
    model_config = ConfigDict(extra="forbid")
    username: str | None = None
    password: str | None = None


@router.put("/scraper-settings/credentials")
async def update_scraper_credentials(
    body: ScraperCredsUpdate,
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Upsert the Samsung credential override into app_config (runtime change).

    `username` and `password` are independent — send either or both. The
    password is stored to app_config 'bqms_password' but is NEVER returned and
    NEVER logged. Busts the in-process credential cache so this API process
    sees the change immediately; other processes pick it up on cache expiry.
    """
    from app.services.bqms_credentials import (
        bust_bqms_credentials_cache,
        get_bqms_credentials_meta,
    )

    if body.username is None and body.password is None:
        raise HTTPException(status_code=400, detail="Provide `username` and/or `password`")

    async def _upsert(key: str, value: str) -> None:
        await conn.execute(
            """
            INSERT INTO app_config (key, value, updated_at, updated_by)
            VALUES ($1, to_jsonb($2::text), NOW(), $3::uuid)
            ON CONFLICT (key) DO UPDATE SET
                value      = EXCLUDED.value,
                updated_at = EXCLUDED.updated_at,
                updated_by = EXCLUDED.updated_by
            """,
            key, value, token_data.user_id,
        )

    if body.username is not None:
        await _upsert("bqms_username", body.username.strip())
    if body.password is not None:
        await _upsert("bqms_password", body.password)

    # Audit — record WHAT changed, never the secret value.
    try:
        await conn.execute(
            """
            INSERT INTO audit_log
                (user_id, action, table_name, record_id, new_data, created_at)
            VALUES ($1::uuid, 'bqms.scraper_credentials', 'app_config', 'bqms_credentials', $2::jsonb, NOW())
            """,
            token_data.user_id,
            _json.dumps({
                "username_changed": body.username is not None,
                "password_changed": body.password is not None,
            }),
        )
    except Exception as exc:
        logger.warning("audit log scraper_credentials failed: %s", exc)

    bust_bqms_credentials_cache()
    meta = get_bqms_credentials_meta()

    row = await conn.fetchrow(
        "SELECT MAX(updated_at) AS ts FROM app_config WHERE key IN ('bqms_username','bqms_password')"
    )
    updated_at = row["ts"].isoformat() if row and row["ts"] else None

    logger.info(
        "bqms credentials updated by %s (username_changed=%s password_changed=%s)",
        token_data.email, body.username is not None, body.password is not None,
    )
    return {
        "username": meta["username"],
        "password_set": meta["password_set"],
        "updated_at": updated_at,
    }


@router.post("/scraper-settings/test-login")
async def test_scraper_login(
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Run ONE Samsung login with the CURRENT resolved credentials.

    Wrapped in samsung_session_lock (short timeout) so it can't collide with a
    running scraper/push. Busts the credential cache first so a just-saved
    password is used. Does NOT enable any flag. Returns { ok, message }.
    """
    from app.services.bqms_credentials import bust_bqms_credentials_cache
    from app.services.samsung_session_lock import samsung_session_lock
    from app.etl.bqms_playwright import playwright_bqms_login
    from app.core.config import settings as cfg
    import asyncpg as apg
    import asyncio

    # Use the freshly-saved override, not a stale cached value.
    bust_bqms_credentials_cache()

    db_url = (
        str(cfg.DATABASE_URL)
        .replace("+asyncpg", "")
        .replace("postgresql+asyncpg", "postgresql")
    )
    pool = await apg.create_pool(db_url, min_size=1, max_size=2)
    try:
        async with samsung_session_lock(
            pool, who=f"test-login-{token_data.email}", timeout_seconds=45,
        ):
            # Hard outer bound so a hung Samsung/Playwright login can't hang the
            # request indefinitely (the 45s above only bounds lock ACQUISITION).
            cookies = await asyncio.wait_for(playwright_bqms_login(), timeout=90)
        ok = bool(cookies and cookies.get("JSESSIONID"))
        message = (
            "Đăng nhập Samsung BQMS thành công."
            if ok else
            "Đăng nhập trả về nhưng không có session cookie — kiểm tra lại tài khoản."
        )
    except Exception as exc:  # noqa: BLE001
        # Surface the failure reason (Samsung error text) but never the password.
        ok = False
        message = f"Đăng nhập thất bại: {exc}"
        logger.warning("bqms test-login failed for %s: %s", token_data.email, exc)
    finally:
        await pool.close()

    # On success, stamp the time so the flag-enable gate (update_scraper_flags) knows
    # a recent login PASSED — this is the safety interlock that stops anyone from
    # re-enabling a scraper while an outdated Samsung password would get spammed and
    # lock the account.
    if ok:
        try:
            await conn.execute(
                """
                INSERT INTO app_config (key, value, updated_at, updated_by)
                VALUES ('bqms_last_login_ok_at', to_jsonb(NOW()::text), NOW(), $1::uuid)
                ON CONFLICT (key) DO UPDATE SET
                    value      = EXCLUDED.value,
                    updated_at = EXCLUDED.updated_at,
                    updated_by = EXCLUDED.updated_by
                """,
                token_data.user_id,
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("failed to stamp bqms_last_login_ok_at: %s", exc)

    logger.info("bqms test-login by %s → ok=%s", token_data.email, ok)
    return {"ok": ok, "message": message}


# Phase F (Thang 2026-05-13): Data-gap tracking — thống kê RFQ còn thiếu detail
# + endpoint quét bù manual để user theo dõi tiến độ auto-drill.
@router.get("/data-gaps")
async def bqms_data_gaps(
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales",
        "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Thống kê RFQ pending còn thiếu data (items grid / folder / images).

    Trả về:
      - by_state: count theo trạng thái (has_items / empty_items / no_detail / missing_folder)
      - missing_list: top 50 RFQ thiếu items chi tiết (rfq_number, samsung_item_cnt, scraped_at)
      - last_cron: thông tin cron run cuối (timestamp, duration, drilled_count)
    """
    # 1) State breakdown — count by data completeness state
    state_rows = await conn.fetch(
        """
        SELECT
            CASE
                WHEN raw_json->'_detail' IS NULL THEN 'no_detail'
                WHEN jsonb_array_length(COALESCE(raw_json->'_detail'->'items','[]'::jsonb)) > 0 THEN 'has_items'
                ELSE 'empty_items'
            END AS state,
            COUNT(*) AS n
        FROM bqms_vendor_portal_staging
        WHERE module = 'bidding' AND status = 'pending_review'
        GROUP BY state
        """
    )
    by_state = {r["state"]: int(r["n"]) for r in state_rows}
    total_pending = sum(by_state.values())

    # 2) List of RFQs missing items (Samsung says >0 items but our _detail.items empty)
    missing_rows = await conn.fetch(
        """
        SELECT
            id,
            rfq_number,
            raw_json->>'itemCnt' AS samsung_item_cnt,
            raw_json->>'regDt'   AS reg_dt,
            raw_json->>'reqName' AS req_name,
            scraped_at,
            jsonb_array_length(COALESCE(raw_json->'_detail'->'items','[]'::jsonb)) AS our_items
        FROM bqms_vendor_portal_staging
        WHERE module = 'bidding' AND status = 'pending_review'
          AND jsonb_array_length(COALESCE(raw_json->'_detail'->'items','[]'::jsonb)) = 0
        ORDER BY id DESC
        LIMIT 50
        """
    )
    missing_list = [
        {
            "id": int(r["id"]),
            "rfq_number": r["rfq_number"],
            "samsung_item_cnt": (r["samsung_item_cnt"] or "").strip() or None,
            "reg_dt": (r["reg_dt"] or "").strip() or None,
            "req_name": (r["req_name"] or "").strip()[:80] or None,
            "scraped_at": r["scraped_at"].isoformat() if r["scraped_at"] else None,
            "our_items": int(r["our_items"] or 0),
        }
        for r in missing_rows
    ]

    # 3) Last periodic-scrape run from app_config (cron may write summary here)
    last_cron_row = await conn.fetchrow(
        "SELECT value::text AS v, updated_at FROM app_config "
        "WHERE key = 'bqms_periodic_scrape_last_run'"
    )
    last_cron = None
    if last_cron_row:
        try:
            last_cron = {
                "summary": _json.loads(last_cron_row["v"]),
                "updated_at": last_cron_row["updated_at"].isoformat() if last_cron_row["updated_at"] else None,
            }
        except Exception:
            last_cron = {"summary": last_cron_row["v"], "updated_at": None}

    # 4) Smart-rescan state (Phase F Thang 2026-05-13)
    rescan_state_row = await conn.fetchrow(
        "SELECT value::text AS v, updated_at FROM app_config "
        "WHERE key = 'bqms_smart_rescan_state'"
    )
    rescan_state = None
    if rescan_state_row:
        try:
            rescan_state = _json.loads(rescan_state_row["v"])
            rescan_state["updated_at"] = (
                rescan_state_row["updated_at"].isoformat() if rescan_state_row["updated_at"] else None
            )
        except Exception:
            rescan_state = None

    # Smart-rescan enabled flag (default true)
    rescan_enabled_row = await conn.fetchrow(
        "SELECT value::text AS v FROM app_config WHERE key = 'bqms_smart_rescan_enabled'"
    )
    rescan_enabled = True
    if rescan_enabled_row:
        try:
            v = rescan_enabled_row["v"].strip().lower().strip('"')
            rescan_enabled = v in ("true", "1", "yes")
        except Exception:
            pass

    # 5) Total scope across all sources (for context)
    total_rfq = await conn.fetchval("SELECT COUNT(*) FROM bqms_rfq")

    # Phase G (Thang 2026-05-13): Smart Code-Track state + gap breakdown
    ct_state_row = await conn.fetchrow(
        "SELECT value::text AS v, updated_at FROM app_config "
        "WHERE key = 'bqms_code_track_state'"
    )
    ct_state = None
    if ct_state_row:
        try:
            ct_state = _json.loads(ct_state_row["v"])
            ct_state["updated_at"] = (
                ct_state_row["updated_at"].isoformat() if ct_state_row["updated_at"] else None
            )
        except Exception:
            ct_state = None
    ct_enabled_row = await conn.fetchrow(
        "SELECT value::text AS v FROM app_config WHERE key = 'bqms_code_track_enabled'"
    )
    ct_enabled = True  # default ON
    if ct_enabled_row:
        try:
            v = ct_enabled_row["v"].strip().lower().strip('"')
            ct_enabled = v in ("true", "1", "yes")
        except Exception:
            pass
    ct_breakdown_rows = await conn.fetch(
        "SELECT gap_type, COUNT(*) AS n FROM bqms_row_gaps "
        "WHERE healed_at IS NULL GROUP BY gap_type"
    )
    ct_breakdown = {r["gap_type"]: int(r["n"]) for r in ct_breakdown_rows}
    ct_healed_today = await conn.fetchval(
        "SELECT COUNT(*) FROM bqms_row_gaps WHERE healed_at >= CURRENT_DATE"
    )
    ct_cooldown_count = await conn.fetchval(
        "SELECT COUNT(DISTINCT rfq_number) FROM bqms_row_gaps "
        "WHERE last_attempt_at > NOW() - INTERVAL '10 minutes' AND healed_at IS NULL"
    )

    return {
        "data": {
            "total_pending": total_pending,
            "total_rfq_db": int(total_rfq or 0),
            "by_state": by_state,
            "missing_list": missing_list,
            "last_cron": last_cron,
            "smart_rescan": {
                "enabled": rescan_enabled,
                "state": rescan_state,  # {status, gaps_before, processed, finished_at, ...}
            },
            "code_track": {
                "enabled": ct_enabled,
                "last_run": ct_state,
                "gap_breakdown": ct_breakdown,
                "healed_today": int(ct_healed_today or 0),
                "pending_cooldown": int(ct_cooldown_count or 0),
            },
        }
    }


@router.post("/data-gaps/toggle-code-track")
async def toggle_code_track(
    enabled: bool = Query(...),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Bật/tắt Smart Code-Track engine (chạy mỗi 3 phút, self-healing 10 loại gap)."""
    val = "true" if enabled else "false"
    await conn.execute(
        """
        INSERT INTO app_config (key, value, updated_at, updated_by)
        VALUES ('bqms_code_track_enabled', $1::jsonb, NOW(), $2::uuid)
        ON CONFLICT (key) DO UPDATE SET
            value = EXCLUDED.value,
            updated_at = EXCLUDED.updated_at,
            updated_by = EXCLUDED.updated_by
        """,
        val, token_data.user_id,
    )
    return {"data": {"enabled": enabled},
            "message": f"Smart Code-Track đã {'BẬT' if enabled else 'TẮT'}"}


@router.get("/data-gaps/healing-log")
async def healing_log(
    limit: int = Query(50, ge=1, le=200),
    kind: str | None = Query(None, description="Filter by gap_type"),
    only_healed: bool = Query(False, description="Only show healed_at IS NOT NULL"),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Recent heal attempts/outcomes for the UI 'Lịch sử heal' tab."""
    rows = await conn.fetch(
        """
        SELECT id, rfq_number, rfq_id, staging_id, gap_type, evidence,
               detected_at, last_attempt_at, drill_attempts, healed_at, last_error
        FROM bqms_row_gaps
        WHERE ($1::text IS NULL OR gap_type = $1)
          AND ($2 = false OR healed_at IS NOT NULL)
        ORDER BY COALESCE(healed_at, last_attempt_at, detected_at) DESC
        LIMIT $3
        """,
        kind, only_healed, limit,
    )
    out = []
    for r in rows:
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
        out.append(d)
    return {"data": out, "total": len(out)}


@router.post("/data-gaps/toggle-smart-rescan")
async def toggle_smart_rescan(
    enabled: bool = Query(...),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Bật/tắt smart auto-rescan task (chạy mỗi 5 phút, drill khi có gap)."""
    val = "true" if enabled else "false"
    await conn.execute(
        """
        INSERT INTO app_config (key, value, updated_at, updated_by)
        VALUES ('bqms_smart_rescan_enabled', $1::jsonb, NOW(), $2::uuid)
        ON CONFLICT (key) DO UPDATE SET
            value = EXCLUDED.value,
            updated_at = EXCLUDED.updated_at,
            updated_by = EXCLUDED.updated_by
        """,
        val, token_data.user_id,
    )
    return {"data": {"enabled": enabled}, "message": f"Smart auto-rescan đã {'BẬT' if enabled else 'TẮT'}"}


@router.post("/data-gaps/rescan")
async def bqms_data_gaps_rescan(
    max_rfqs: int = Query(50, ge=1, le=200, description="Max RFQs to drill this run"),
    budget_seconds: int = Query(1800, ge=60, le=3600, description="Hard time cap"),
    token_data: TokenData = Depends(require_role("admin", "manager")),
):
    """Quét bù: drill detail cho các RFQ pending còn thiếu items.

    Chạy synchronous (block tới khi xong hoặc hết budget). Trả về số RFQ
    đã drill + số file tải. Dùng cùng logic với periodic cron — chỉ khác
    là user trigger ngay thay vì chờ 30 phút tiếp theo.
    """
    import asyncpg
    from app.tasks.bqms_periodic_scrape import _auto_drill_new_rfqs
    from app.core.config import settings as _settings

    db_url = (
        str(_settings.DATABASE_URL)
        .replace("+asyncpg", "")
        .replace("postgresql+asyncpg", "postgresql")
    )
    pool = await asyncpg.create_pool(db_url, min_size=1, max_size=2)
    try:
        processed, total_files = await _auto_drill_new_rfqs(
            pool, max_per_cycle=max_rfqs, budget_seconds=budget_seconds,
        )
        return {
            "data": {
                "rfqs_processed": int(processed),
                "files_downloaded": int(total_files),
            },
            "message": f"Quét bù xong — {processed} RFQ đã drill, {total_files} file tải",
        }
    finally:
        await pool.close()


@router.get("/stats/win-lost")
async def stats_win_lost(
    period: str = Query("week", description="day | week | month"),
    weeks: int = Query(8, ge=1, le=52, description="how many periods back"),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Win/Lost statistics by period — reads directly from Selection Result
    staging (module='selection_result'). Per Thang 2026-05-11:
      - selectionResultName='Selected'   → won
      - selectionResultName='Unselected' → lost
    Grouped by week/month of deadlineDt (closing date)."""
    if period not in ("day", "week", "month"):
        raise HTTPException(400, "period must be day|week|month")

    trunc = {"day": "day", "week": "week", "month": "month"}[period]

    # Pull all selection results + parse date in Python (format varies)
    # Use both selectionResult (RES01=won, RES02=lost) AND selectionResultName
    # to be robust against text variations ("Selected" vs "Selection").
    rows = await conn.fetch(
        """
        SELECT
            id,
            rfq_number,
            raw_json->>'selectionResult'     AS result_code,
            raw_json->>'selectionResultName' AS result_name,
            raw_json->>'deadlineDt' AS deadline_dt,
            raw_json->>'regDt'      AS reg_dt,
            scraped_at::date         AS scraped_date
        FROM bqms_vendor_portal_staging
        WHERE module = 'selection_result'
        """,
    )

    from datetime import datetime as _dt, timedelta as _td
    import re as _re

    def _parse_date(s: str | None) -> _dt | None:
        if not s: return None
        # Strip "(GMT+xx:xx) " prefix
        s2 = _re.sub(r"\(GMT[+\-]\d{2}:\d{2}\)\s*", "", s).strip()
        for fmt in ("%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return _dt.strptime(s2.split(" ")[0] if fmt == "%m/%d/%Y" else s2[:len(_dt.now().strftime(fmt))], fmt)
            except (ValueError, TypeError):
                try:
                    return _dt.strptime(s2, fmt)
                except (ValueError, TypeError):
                    continue
        return None

    def _trunc_date(d: _dt) -> _dt:
        if period == "day":
            return d.replace(hour=0, minute=0, second=0, microsecond=0)
        if period == "week":
            return (d - _td(days=d.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        # month
        return d.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    cutoff = _dt.now() - _td(days=weeks * (7 if period == "week" else (1 if period == "day" else 30)))

    by_period: dict[str, dict[str, int]] = {}
    seen_rfq: dict[str, str] = {}  # dedup: keep first result per rfq_number
    for r in rows:
        # Use selectionResult code first (more reliable): RES01=won, RES02=lost
        code = (r["result_code"] or "").strip().upper()
        name = (r["result_name"] or "").strip().lower()
        if code == "RES01" or name.startswith("select"):
            key = "won"
        elif code == "RES02" or name.startswith("unselect"):
            key = "lost"
        else:
            continue
        rfq = r["rfq_number"] or ""
        if rfq and rfq in seen_rfq:
            continue
        seen_rfq[rfq] = key
        # Parse date: prefer deadlineDt → regDt → scraped_date fallback
        dt = (_parse_date(r["deadline_dt"]) or _parse_date(r["reg_dt"])
              or (_dt.combine(r["scraped_date"], _dt.min.time()) if r["scraped_date"] else None))
        if not dt or dt < cutoff:
            continue
        ps = _trunc_date(dt).date().isoformat()
        by_period.setdefault(ps, {"won": 0, "lost": 0})
        by_period[ps][key] += 1

    totals = {"won": 0, "lost": 0}
    for v in by_period.values():
        totals["won"] += v["won"]
        totals["lost"] += v["lost"]
    total_decided = totals["won"] + totals["lost"]
    win_rate = (totals["won"] / total_decided * 100) if total_decided > 0 else 0.0

    return {
        "data": {
            "period": period,
            "weeks": weeks,
            "by_period": [
                {"start": ps, "won": v["won"], "lost": v["lost"], "total": v["won"] + v["lost"]}
                for ps, v in sorted(by_period.items(), reverse=True)
            ],
            "totals": {**totals, "decided": total_decided, "win_rate_pct": round(win_rate, 1)},
            "source": "bqms_vendor_portal_staging (selection_result module)",
            "raw_count": len(rows),
            "deduped_count": len(seen_rfq),
        }
    }


@router.post("/scrape-trigger/selection-result")
async def trigger_selection_result_scrape(
    limit: int = Query(0, ge=0, le=2000, description="0 = all available rows"),
    token_data: TokenData = Depends(require_role("admin", "manager")),
):
    """Manually fire a Selection Result scrape NOW (synchronous, ~1-2 min).
    Use this to refresh win/lost data without waiting for cron.
    Per Thang 2026-05-11: cron is OFF by default, this gives manual control."""
    import asyncpg
    from app.etl.bqms_l1_l3_scraper import scrape_selection_result
    from app.core.config import settings

    db_url = (
        str(settings.DATABASE_URL)
        .replace("+asyncpg", "")
        .replace("postgresql+asyncpg", "postgresql")
    )
    pool = await asyncpg.create_pool(db_url, min_size=1, max_size=2)
    try:
        result = await scrape_selection_result(
            limit=limit, save_raw_json=False, db_pool=pool, auto_mark_result=True,
        )
        return {
            "data": {
                "list_count": result.get("list_count", 0),
                "total_available": result.get("total_available", 0),
                "staging_inserts": result.get("staging_inserts", 0),
                "won_marked": result.get("won_marked", 0),
                "lost_marked": result.get("lost_marked", 0),
                "duration_seconds": result.get("duration_seconds"),
            },
            "message": "Selection Result scrape complete",
        }
    finally:
        await pool.close()


@router.post("/admin/reextract-images")
async def admin_reextract_images(
    rfq_number: str | None = Query(None, description="If set, only this RFQ; else ALL recent folders"),
    limit: int = Query(50, ge=1, le=500, description="Max folders to process"),
    token_data: TokenData = Depends(require_role("admin", "manager")),
):
    """Re-run smart image extraction for existing folders (clean up duplicates
    from the pre-2026-05-11 naive extraction). Skips folders where images/
    is empty or raw/ is empty.

    Per Thang 2026-05-11: smart dedup needs to be applied retroactively to
    folders whose images got assigned identical content across multiple
    item codes.
    """
    from app.etl.bqms_bidding_scraper import (
        RFQ_ROOT, _extract_images_for_rfq_folder, find_existing_rfq_folder,
    )

    folders: list = []
    if rfq_number:
        f = find_existing_rfq_folder(rfq_number)
        if f:
            folders.append(f)
    else:
        # Walk recent month folders
        from datetime import datetime as _dt
        now = _dt.now()
        for y in (now.year, now.year - 1):
            year_root = RFQ_ROOT / f"RFQ {y}"
            if not year_root.exists():
                continue
            for m in range(12, 0, -1):
                month_root = year_root / f"THANG {m}"
                if not month_root.exists():
                    continue
                for d in month_root.iterdir():
                    if d.is_dir() and (d / "raw").exists():
                        folders.append(d)
                        if len(folders) >= limit:
                            break
                if len(folders) >= limit:
                    break
            if len(folders) >= limit:
                break

    summary = []
    for folder in folders:
        raw = folder / "raw"
        images = folder / "images"
        if not raw.exists():
            continue
        # Clean old images first
        if images.exists():
            for p in images.glob("*.png"):
                try: p.unlink()
                except Exception: pass
            for p in images.glob("*.jpg"):
                try: p.unlink()
                except Exception: pass
            for p in images.glob("*.jpeg"):
                try: p.unlink()
                except Exception: pass
        try:
            n = _extract_images_for_rfq_folder(raw, images)
            summary.append({"folder": folder.name, "images": n})
        except Exception as exc:
            summary.append({"folder": folder.name, "error": str(exc)[:200]})

    return {
        "data": {"processed": len(summary), "summary": summary[:50]},
        "message": f"Đã re-extract {len(summary)} folder(s)",
    }


@router.post("/admin/reset-data")
async def admin_reset_data(
    confirm: str = Query(..., description="Phải nhập đúng 'RESET' để xác nhận"),
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """WIPE all BQMS scraped data — for go-live with fresh state.
    Deletes from: bqms_rfq, bqms_vendor_portal_staging, bqms_won_quotations,
    bqms_deliveries, quotations (bqms-related), bqms_quote_log, bqms_quote_batches.
    DOES NOT delete: quotation_templates, users, audit_log.

    Per Thang 2026-05-11: wanted a clean slate before production go-live.
    """
    if confirm != "RESET":
        raise HTTPException(400, "confirm phải là 'RESET' (in hoa)")

    counts: dict[str, int] = {}
    for tbl, where in [
        ("bqms_quote_batch_items", "1=1"),
        ("bqms_quote_batches", "1=1"),
        ("bqms_quote_log", "1=1"),
        ("rfq_quotations", "1=1"),
        ("bqms_quotation_items", "1=1"),
        ("quotations", "rfq_no IS NOT NULL"),
        ("bqms_won_quotations", "1=1"),
        ("bqms_deliveries", "1=1"),
        ("bqms_rfq", "1=1"),
        ("bqms_vendor_portal_staging", "1=1"),
    ]:
        try:
            row = await conn.fetchrow(f"SELECT COUNT(*) AS c FROM {tbl} WHERE {where}")
            counts[tbl] = int(row["c"])
            await conn.execute(f"DELETE FROM {tbl} WHERE {where}")
        except Exception as exc:
            logger.warning("reset %s failed: %s", tbl, exc)
            counts[tbl] = -1

    # Audit log
    try:
        await conn.execute(
            """
            INSERT INTO audit_log
                (user_id, action, table_name, record_id, new_data, created_at)
            VALUES ($1::uuid, 'bqms.admin_reset_data', 'multi', 'reset', $2::jsonb, NOW())
            """,
            token_data.user_id, _json.dumps(counts),
        )
    except Exception:
        pass

    return {"data": {"deleted": counts}, "message": "Đã reset toàn bộ data BQMS"}


@router.get("/activity/recent")
async def quote_activity_recent(
    days: int = Query(7, ge=1, le=30),
    limit: int = Query(50, ge=1, le=200),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Recent BQMS quoting + scrape activity for dashboard 'Tổng quan' widget.
    Per Thang 2026-05-11: surfaces today's quote events + Closed/Round2 detections
    from periodic scrape."""
    rows = await conn.fetch(
        """
        SELECT
            id, action, record_id, new_data, user_email, created_at
        FROM audit_log
        WHERE action LIKE 'bqms.%' OR action LIKE 'bqms_periodic.%'
        AND created_at > NOW() - ($1 || ' days')::interval
        ORDER BY created_at DESC
        LIMIT $2
        """,
        str(days), limit,
    )
    items = []
    for r in rows:
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
        items.append(d)
    return {"data": {"items": items, "days": days}}


# ---------------------------------------------------------------------------
# RFQ — Quotation History for a specific RFQ row
# ---------------------------------------------------------------------------

@router.get("/rfq/{rfq_id}/history")
async def rfq_history(
    rfq_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Get quotation history for a specific bqms_rfq row (matched by rfq_number)."""
    rfq = await conn.fetchrow(
        "SELECT id, rfq_number FROM bqms_rfq WHERE id = $1", rfq_id
    )
    if not rfq:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy RFQ #{rfq_id}")

    rfq_number = rfq["rfq_number"]

    # Phase G (Thang 2026-05-13): expose bqms_quote_log entries trong response
    # để user thấy đầy đủ lịch sử giá (mọi lần báo giá đều log 1 row, kể cả
    # khi sửa lại giá V_n nhiều lần).
    quote_log_rows = []
    try:
        log_rows = await conn.fetch(
            """
            SELECT ql.id, ql.rfq_id, ql.round, ql.quoted_price, ql.quoted_currency,
                   ql.item_type, ql.notes, ql.created_at,
                   u.full_name AS quoted_by_name, u.email AS quoted_by_email
            FROM bqms_quote_log ql
            LEFT JOIN users u ON u.id = ql.quoted_by
            JOIN bqms_rfq r ON r.id = ql.rfq_id
            WHERE r.rfq_number = $1
            ORDER BY ql.created_at DESC
            LIMIT 100
            """,
            rfq_number,
        )
        for lr in log_rows:
            d = dict(lr)
            for k, v in d.items():
                if isinstance(v, (date, datetime)):
                    d[k] = v.isoformat()
            if d.get("quoted_price") is not None:
                d["quoted_price"] = float(d["quoted_price"])
            quote_log_rows.append(d)
    except Exception as exc:
        logger.warning("quote_log fetch failed for rfq=%s: %s", rfq_number, exc)

    # Query quotations table — include file paths for download/preview
    try:
        rows = await conn.fetch(
            """
            SELECT id, rfq_no, status, total_items, filled_items,
                   output_xlsx, output_pdf, created_at
            FROM quotations
            WHERE rfq_no = $1
            ORDER BY created_at DESC
            LIMIT 50
            """,
            rfq_number,
        )
    except Exception:
        rows = []

    import os

    def _ser(r: asyncpg.Record) -> dict:
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()

        # Build file list from output directory
        files = []
        pdf_path = d.get("output_pdf")
        xlsx_path = d.get("output_xlsx")
        qid = d["id"]

        if pdf_path and os.path.exists(os.path.dirname(pdf_path)):
            out_dir = os.path.dirname(pdf_path)
            for fname in sorted(os.listdir(out_dir)):
                fpath = os.path.join(out_dir, fname)
                fsize = os.path.getsize(fpath) if os.path.isfile(fpath) else 0
                is_pdf = fname.endswith(".pdf")
                is_cam_ket = "cam_ket" in fname.lower()
                ftype = ("cam_ket_" if is_cam_ket else "quotation_") + ("pdf" if is_pdf else "xlsx")
                files.append({
                    "type": ftype,
                    "filename": fname,
                    "size": fsize,
                    "path": fpath,
                    "download_url": f"/api/v1/quotations/download/{qid}/{ftype}",
                    "preview_url": f"/api/v1/quotations/preview/{qid}/{ftype}" if is_pdf else None,
                })

        d["files"] = files
        return d

    return {
        "data": [_ser(r) for r in rows],
        "rfq_number": rfq_number,
        "total": len(rows),
        # Phase G (Thang 2026-05-13): full price-change history từ bqms_quote_log
        "quote_log": quote_log_rows,
    }


# ---------------------------------------------------------------------------
# Deliveries
# ---------------------------------------------------------------------------

# Status normalization: Vietnamese → English
_STATUS_NORM = {
    "chua_giao": "pending",
    "dang_giao": "in_transit",
    "da_giao": "delivered",
    "hoan_tat": "completed",
    "giao_mot_phan": "partial",
}

# Reverse map for DB queries: group → all DB enum values
_STATUS_GROUPS = {
    "pending": ("chua_giao", "pending"),
    "in_transit": ("dang_giao", "in_transit", "picked_up", "customs_clearance"),
    "delivered": ("da_giao", "delivered", "completed", "hoan_tat"),
}


def _build_delivery_filters(
    status: str | None, month: int | None, year: int | None,
    date_from: date | None = None, date_to: date | None = None,
) -> tuple[str, list]:
    """Build WHERE clause for delivery queries. Returns (where_sql, params)."""
    conditions = ["1=1"]
    params: list = []
    idx = 1

    if status:
        group = _STATUS_GROUPS.get(status)
        if group:
            placeholders = ", ".join(f"${idx + i}" for i in range(len(group)))
            conditions.append(f"d.delivery_status::text IN ({placeholders})")
            params.extend(group)
            idx += len(group)
        else:
            conditions.append(f"d.delivery_status::text = ${idx}")
            params.append(status)
            idx += 1
    if date_from:
        # Thang 2026-06-01: dùng OR — show rows nếu po_date HOẶC delivery_date
        # nằm trong khoảng. Trước đây COALESCE(po_date, delivery_date) sẽ ẩn
        # các đơn po_date cũ (vd Feb) nhưng lịch giao tháng 6.
        conditions.append(f"(d.po_date >= ${idx} OR d.delivery_date >= ${idx})")
        params.append(date_from)
        idx += 1
    if date_to:
        conditions.append(f"(d.po_date <= ${idx} OR d.delivery_date <= ${idx})")
        params.append(date_to)
        idx += 1
    if month:
        # Match nếu MONTH(po_date) = M HOẶC MONTH(delivery_date) = M.
        conditions.append(
            f"(EXTRACT(MONTH FROM d.po_date) = ${idx} "
            f"OR EXTRACT(MONTH FROM d.delivery_date) = ${idx})"
        )
        params.append(month)
        idx += 1
    if year:
        conditions.append(
            f"(EXTRACT(YEAR FROM d.po_date) = ${idx} "
            f"OR EXTRACT(YEAR FROM d.delivery_date) = ${idx})"
        )
        params.append(year)
        idx += 1

    return " AND ".join(conditions), params


@router.get("/deliveries/kpi")
async def delivery_kpi(
    status: str | None = Query(None),
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=2020, le=2099),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """KPI thống kê giao hàng — tính trên toàn bộ dữ liệu đã lọc (không phân trang).

    Thang 2026-06-02: dedup theo (po_number, bqms_code) cho các field "per-PO-line"
    (total_order_value, total_orders, delivered/in_transit/pending count).
    Lý do: 1 PO line có thể được giao thành nhiều partial → tạo nhiều row trong
    bqms_deliveries với CÙNG `amount`. SUM(amount) naive bị nhân 2-3 lần.
    Trạng thái của 1 line = trạng thái của row partial MỚI NHẤT (theo delivery_date).

    Per-shipment fields (total_delivered_vnd) vẫn SUM trên row vì mỗi partial có
    value khác nhau (đại diện cho giá trị shipment đó).
    """
    where, params = _build_delivery_filters(status, month, year)

    # Query 1: per-pair stats (dedup by line)
    pair_row = await conn.fetchrow(
        f"""
        WITH ranked AS (
            SELECT
                d.po_number, d.bqms_code,
                d.delivery_status::text AS status,
                d.amount,
                ROW_NUMBER() OVER (
                    PARTITION BY d.po_number, d.bqms_code
                    ORDER BY d.delivery_date DESC NULLS LAST, d.id DESC
                ) AS rn
              FROM bqms_deliveries d
             WHERE {where}
        ),
        pair_state AS (
            SELECT
                po_number, bqms_code,
                MAX(status) FILTER (WHERE rn = 1) AS latest_status,
                MAX(amount) AS amount
              FROM ranked
             GROUP BY po_number, bqms_code
        )
        SELECT
            COUNT(*)::int AS total_orders,
            COUNT(*) FILTER (
                WHERE latest_status IN ('da_giao','delivered','completed','hoan_tat')
            )::int AS delivered_count,
            COUNT(*) FILTER (
                WHERE latest_status IN ('dang_giao','in_transit','picked_up','customs_clearance')
            )::int AS in_transit_count,
            COUNT(*) FILTER (
                WHERE latest_status IN ('chua_giao','pending')
            )::int AS pending_count,
            COALESCE(SUM(amount), 0)::bigint AS total_order_value
        FROM pair_state
        """,
        *params,
    )

    # Query 2: per-row (partial shipment) sum — naive SUM since values differ per partial
    delivered_vnd = await conn.fetchval(
        f"""
        SELECT COALESCE(SUM(d.total_delivered_value_vnd) FILTER (
            WHERE d.delivery_status::text IN ('da_giao','delivered','completed','hoan_tat')
        ), 0)::bigint
        FROM bqms_deliveries d
        WHERE {where}
        """,
        *params,
    )

    return {**dict(pair_row), "total_delivered_vnd": int(delivered_vnd or 0)}


@router.get("/deliveries/revenue-stats")
async def delivery_revenue_stats(
    status: str | None = Query(None),
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=2020, le=2099),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    q: str | None = Query(None),
    driver_id: int | None = Query(None),
    po_number: str | None = Query(None),
    bqms_code: str | None = Query(None),
    group_by: str = Query("day"),
    breakdown_limit: int = Query(20, ge=1, le=100),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Doanh thu PO: summary + timeseries + breakdown.

    `total_amount_vnd` = SUM(amount) toàn bộ PO trong filter (giá trị đặt hàng).
    `delivered_amount_vnd` = SUM(total_delivered_value_vnd) khi đã giao (doanh thu thực).
    Summary numbers MUST khớp với /deliveries/kpi khi cùng filter.
    """
    if group_by not in {"day", "month", "driver", "po", "bqms", "recipient", "origin", "status"}:
        raise HTTPException(400, "group_by không hợp lệ")

    conds = ["1=1"]
    params: list = []
    idx = 1

    if status:
        grp = _STATUS_GROUPS.get(status)
        if grp:
            placeholders = ", ".join(f"${idx + i}" for i in range(len(grp)))
            conds.append(f"d.delivery_status::text IN ({placeholders})")
            params.extend(grp)
            idx += len(grp)
        else:
            conds.append(f"d.delivery_status::text = ${idx}")
            params.append(status)
            idx += 1
    if date_from:
        conds.append(f"COALESCE(d.po_date, d.delivery_date) >= ${idx}")
        params.append(date_from); idx += 1
    if date_to:
        conds.append(f"COALESCE(d.po_date, d.delivery_date) <= ${idx}")
        params.append(date_to); idx += 1
    if month:
        conds.append(f"EXTRACT(MONTH FROM COALESCE(d.po_date, d.delivery_date)) = ${idx}")
        params.append(month); idx += 1
    if year:
        conds.append(f"EXTRACT(YEAR FROM COALESCE(d.po_date, d.delivery_date)) = ${idx}")
        params.append(year); idx += 1
    if driver_id:
        conds.append(f"d.driver_id = ${idx}")
        params.append(driver_id); idx += 1
    if po_number:
        conds.append(f"d.po_number ILIKE ${idx}")
        params.append(f"%{po_number}%"); idx += 1
    if bqms_code:
        conds.append(f"d.bqms_code ILIKE ${idx}")
        params.append(f"%{bqms_code}%"); idx += 1
    if q:
        conds.append(
            f"(d.po_number ILIKE ${idx} OR d.bqms_code ILIKE ${idx} "
            f"OR d.recipient_name ILIKE ${idx} OR d.specification ILIKE ${idx})"
        )
        params.append(f"%{q}%"); idx += 1
    where_sql = " AND ".join(conds)

    # Thang 2026-06-02: dedup theo (po, bqms_code) cho các field per-PO-line.
    # Xem note trong /deliveries/kpi để biết lý do.
    pair_summary = await conn.fetchrow(
        f"""
        WITH ranked AS (
            SELECT
                d.po_number, d.bqms_code,
                d.delivery_status::text AS status,
                d.amount, d.quantity,
                ROW_NUMBER() OVER (
                    PARTITION BY d.po_number, d.bqms_code
                    ORDER BY d.delivery_date DESC NULLS LAST, d.id DESC
                ) AS rn
              FROM bqms_deliveries d
             WHERE {where_sql}
        ),
        pair_state AS (
            SELECT
                po_number, bqms_code,
                MAX(status) FILTER (WHERE rn = 1) AS latest_status,
                MAX(amount) AS amount,
                MAX(quantity) AS quantity
              FROM ranked
             GROUP BY po_number, bqms_code
        )
        SELECT
            COUNT(*)::int AS total_orders,
            COALESCE(SUM(amount), 0)::bigint AS total_amount_vnd,
            COUNT(*) FILTER (
                WHERE latest_status IN ('da_giao','delivered','completed','hoan_tat')
            )::int AS delivered_count,
            COUNT(*) FILTER (
                WHERE latest_status IN ('dang_giao','in_transit','picked_up','customs_clearance')
            )::int AS in_transit_count,
            COUNT(*) FILTER (
                WHERE latest_status IN ('chua_giao','pending')
            )::int AS pending_count,
            COALESCE(SUM(amount) FILTER (
                WHERE latest_status IN ('chua_giao','pending')
            ), 0)::bigint AS pending_amount_vnd,
            COALESCE(SUM(amount) FILTER (
                WHERE latest_status IN ('dang_giao','in_transit','picked_up','customs_clearance')
            ), 0)::bigint AS in_transit_amount_vnd,
            COALESCE(SUM(quantity), 0)::bigint AS total_qty
        FROM pair_state
        """,
        *params,
    )

    # Per-row (partial shipment) sums — values differ per partial nên SUM naive OK
    row_summary = await conn.fetchrow(
        f"""
        SELECT
            COALESCE(SUM(d.total_delivered_value_vnd) FILTER (
                WHERE d.delivery_status::text IN ('da_giao','delivered','completed','hoan_tat')
            ), 0)::bigint AS delivered_amount_vnd,
            COALESCE(SUM(d.actual_delivered_qty), 0)::bigint AS delivered_qty
        FROM bqms_deliveries d
        WHERE {where_sql}
        """,
        *params,
    )
    summary = {**dict(pair_summary), **dict(row_summary)}
    summary["delivery_rate"] = (
        round(summary["delivered_count"] / summary["total_orders"] * 100, 2)
        if summary["total_orders"] > 0 else 0.0
    )
    summary["avg_order_value"] = (
        int(summary["total_amount_vnd"] // summary["total_orders"])
        if summary["total_orders"] > 0 else 0
    )

    if group_by == "month":
        ts_expr = "TO_CHAR(COALESCE(d.po_date, d.delivery_date), 'YYYY-MM')"
        ts_label = "TO_CHAR(COALESCE(d.po_date, d.delivery_date), 'MM/YYYY')"
    else:
        ts_expr = "TO_CHAR(COALESCE(d.po_date, d.delivery_date), 'YYYY-MM-DD')"
        ts_label = "TO_CHAR(COALESCE(d.po_date, d.delivery_date), 'DD/MM')"

    ts_rows = await conn.fetch(
        f"""
        SELECT
            {ts_expr} AS bucket,
            {ts_label} AS label,
            COUNT(*)::int AS count,
            COALESCE(SUM(d.amount), 0)::bigint AS total_amount,
            COALESCE(SUM(d.total_delivered_value_vnd) FILTER (
                WHERE d.delivery_status::text IN ('da_giao','delivered','completed','hoan_tat')
            ), 0)::bigint AS delivered_amount,
            COUNT(*) FILTER (
                WHERE d.delivery_status::text IN ('da_giao','delivered','completed','hoan_tat')
            )::int AS delivered_count
        FROM bqms_deliveries d
        WHERE {where_sql}
            AND COALESCE(d.po_date, d.delivery_date) IS NOT NULL
        GROUP BY bucket, label
        ORDER BY bucket
        LIMIT 400
        """,
        *params,
    )
    timeseries = [dict(r) for r in ts_rows]

    bk_select = None
    bk_join = ""
    bk_group = "key"
    if group_by == "driver":
        bk_select = "COALESCE(drv.full_name, '— Chưa gán —') AS key, d.driver_id AS group_id"
        bk_join = "LEFT JOIN bqms_contacts drv ON drv.id = d.driver_id"
        bk_group = "key, d.driver_id"
    elif group_by == "po":
        bk_select = "COALESCE(d.po_number, '—') AS key, NULL::int AS group_id"
    elif group_by == "bqms":
        bk_select = "COALESCE(d.bqms_code, '—') AS key, NULL::int AS group_id"
    elif group_by == "recipient":
        bk_select = "COALESCE(NULLIF(d.recipient_name,''), '— Trống —') AS key, NULL::int AS group_id"
    elif group_by == "origin":
        bk_select = "COALESCE(NULLIF(d.country_origin,''), '— Trống —') AS key, NULL::int AS group_id"
    elif group_by == "status":
        bk_select = "d.delivery_status::text AS key, NULL::int AS group_id"

    breakdown: list = []
    if bk_select:
        bk_rows = await conn.fetch(
            f"""
            SELECT
                {bk_select},
                COUNT(*)::int AS count,
                COALESCE(SUM(d.amount), 0)::bigint AS total_amount,
                COALESCE(SUM(d.total_delivered_value_vnd) FILTER (
                    WHERE d.delivery_status::text IN ('da_giao','delivered','completed','hoan_tat')
                ), 0)::bigint AS delivered_amount,
                COUNT(*) FILTER (
                    WHERE d.delivery_status::text IN ('da_giao','delivered','completed','hoan_tat')
                )::int AS delivered_count,
                COUNT(*) FILTER (
                    WHERE d.delivery_status::text IN ('chua_giao','pending')
                )::int AS pending_count
            FROM bqms_deliveries d
            {bk_join}
            WHERE {where_sql}
            GROUP BY {bk_group}
            ORDER BY total_amount DESC NULLS LAST
            LIMIT ${idx}
            """,
            *params, breakdown_limit,
        )
        breakdown = [dict(r) for r in bk_rows]

    return {
        "summary": summary,
        "timeseries": timeseries,
        "breakdown": breakdown,
        "group_by": group_by,
        "filters_applied": {
            "status": status, "month": month, "year": year,
            "date_from": str(date_from) if date_from else None,
            "date_to": str(date_to) if date_to else None,
            "q": q, "driver_id": driver_id,
            "po_number": po_number, "bqms_code": bqms_code,
        },
    }


@router.get("/deliveries/export")
async def export_deliveries_excel(
    status: str | None = Query(None),
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=2020, le=2099),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Xuất danh sách giao hàng ra Excel giống format THỐNG KÊ PO."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    where, params = _build_delivery_filters(status, month, year)

    rows = await conn.fetch(
        f"""
        SELECT d.*
        FROM bqms_deliveries d
        WHERE {where}
        ORDER BY d.po_date DESC NULLS LAST, d.id DESC
        LIMIT 10000
        """,
        *params,
    )
    if len(rows) == 10000:
        raise HTTPException(400, "Quá nhiều bản ghi (>10,000). Vui lòng lọc theo tháng/năm/trạng thái.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "THỐNG KÊ PO"

    # Column mapping: DB column → Vietnamese header
    COL_MAP = [
        ("po_date", "Ngày PO"),
        ("po_number", "Số PO"),
        ("shipping_no", "Shipping No"),
        ("quotation_no", "Số QT"),
        ("bqms_code", "BQMS code"),
        ("item_name", "Item Name"),
        ("specification", "Spec"),
        ("quantity", "SL"),
        ("unit", "Đơn vị"),
        ("unit_price", "Đơn giá"),
        ("amount", "Thành tiền"),
        ("sev_type", "SEV/T"),
        ("buyer_email", "MAIL PUR"),
        ("recipient_name", "TÊN NGƯỜI NHẬN"),
        ("receiving_warehouse", "KHO NHẬN"),
        ("buyer_phone", "SĐT PUR"),
        ("delivery_status", "TÌNH TRẠNG"),
        ("delivery_date", "NGÀY GIAO HÀNG"),
        ("actual_delivered_qty", "SL GIAO THỰC TẾ"),
        ("delivery_info", "THÔNG TIN GIAO HÀNG"),
        ("delivery_method", "CÁCH THỨC GIAO HÀNG"),
        ("country_origin", "XUẤT XỨ"),
        ("total_delivered_value_vnd", "TỔNG GIÁ TRỊ ĐÃ GIAO\n(VND)"),
    ]

    # Status display mapping
    STATUS_DISPLAY = {
        "chua_giao": "chưa giao", "pending": "chưa giao",
        "dang_giao": "đang giao", "in_transit": "đang giao",
        "picked_up": "đã lấy hàng", "customs_clearance": "thông quan",
        "da_giao": "đã giao", "delivered": "đã giao",
        "completed": "hoàn tất", "hoan_tat": "hoàn tất",
    }

    # Header styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write header
    for col_idx, (_, header) in enumerate(COL_MAP, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, record in enumerate(rows, 2):
        rec = dict(record)
        for col_idx, (db_col, _) in enumerate(COL_MAP, 1):
            value = rec.get(db_col)
            if db_col == "delivery_status" and value:
                value = STATUS_DISPLAY.get(str(value), str(value))
            if db_col in ("po_date", "delivery_date") and value:
                if hasattr(value, "strftime"):
                    value = value.strftime("%d/%m/%Y")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if db_col in ("quantity", "unit_price", "amount", "actual_delivered_qty", "total_delivered_value_vnd"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    # Auto-fit column widths
    COL_WIDTHS = [12, 14, 14, 14, 18, 22, 30, 8, 8, 12, 14, 8, 14, 18, 25, 14, 12, 14, 10, 25, 16, 10, 18]
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Giao_hang_{month or 'all'}_{year or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/deliveries")
async def delivery_tracking(
    status: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=2020, le=2099),
    search: str | None = Query(None, description="Tìm theo PO, BQMS code, shipping no, specification"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    where, params = _build_delivery_filters(status, month, year, date_from, date_to)
    idx = len(params) + 1

    # Thang 2026-06-01: thêm search param để click PO từ global search bar
    # navigate trực tiếp tới row PO đó (kết hợp ?po=...&year=all).
    if search:
        s = search.strip()
        if s:
            like = f"%{s}%"
            where = (
                f"({where}) AND ("
                f"d.po_number ILIKE ${idx} OR d.bqms_code ILIKE ${idx} "
                f"OR d.shipping_no ILIKE ${idx} OR d.specification ILIKE ${idx} "
                f"OR d.item_name ILIKE ${idx} OR d.quotation_no ILIKE ${idx})"
            )
            params.append(like)
            idx += 1

    # Issue 3 (Thang 2026-06-25): dedup at query level — show ONE row per
    # (po_number, bqms_code) = LATEST shipment, while ALL shipment rows stay in
    # the DB (history preserved). COUNT must reflect deduped groups for paging.
    total = await conn.fetchval(
        f"""
        SELECT COUNT(*) FROM (
            SELECT 1 FROM bqms_deliveries d
            WHERE {where}
            GROUP BY d.po_number, d.bqms_code
        ) x
        """,
        *params,
    )

    actual_offset = (page - 1) * limit if page > 1 else offset

    params.extend([limit, actual_offset])
    rows = await conn.fetch(
        f"""
        WITH base AS (
            SELECT d.*,
                   drv.full_name AS driver_name,
                   drv.phone AS driver_phone,
                   drv.license_plate AS driver_license_plate,
                   drv.vehicle_type AS driver_vehicle_type
            FROM bqms_deliveries d
            LEFT JOIN bqms_contacts drv
                ON drv.id = d.driver_id AND drv.is_driver = true
            WHERE {where}
        ),
        ranked AS (
            SELECT *,
                   ROW_NUMBER() OVER (
                       PARTITION BY po_number, bqms_code
                       ORDER BY COALESCE(actual_delivered_at, delivery_date::timestamptz) DESC NULLS LAST,
                                updated_at DESC NULLS LAST,
                                id DESC
                   ) AS rn,
                   COUNT(*) OVER (PARTITION BY po_number, bqms_code) AS shipment_count
            FROM base
        )
        SELECT * FROM ranked
        WHERE rn = 1
        ORDER BY po_date DESC NULLS LAST, id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )

    # Normalize status for frontend
    data = []
    for r in rows:
        d = dict(r)
        raw_status = str(d.get("delivery_status", ""))
        d["delivery_status_normalized"] = _STATUS_NORM.get(raw_status, raw_status)
        data.append(d)

    return {"data": data, "total": total}


@router.get("/deliveries/shipments")
async def delivery_shipments(
    po_number: str = Query(..., description="PO number (exact)"),
    bqms_code: str = Query(..., description="BQMS code (exact)"),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """ENHANCEMENT P4 (Thang LOCKED 2026-06-25): full shipment history for one
    (po_number, bqms_code) pair. The /deliveries list dedups to the LATEST
    shipment per pair; this returns ALL shipment rows for the detail slide-over,
    ordered latest-first (same ordering as the dedup ROW_NUMBER)."""
    rows = await conn.fetch(
        """
        SELECT id, shipping_no, delivery_date, actual_delivered_at,
               actual_delivered_qty, quantity, delivery_status,
               total_delivered_value_vnd, data_source
        FROM bqms_deliveries
        WHERE po_number = $1 AND bqms_code = $2
        ORDER BY COALESCE(actual_delivered_at, delivery_date::timestamptz) DESC NULLS LAST,
                 updated_at DESC NULLS LAST,
                 id DESC
        """,
        po_number,
        bqms_code,
    )

    data = []
    for r in rows:
        d = dict(r)
        raw_status = str(d.get("delivery_status", ""))
        d["delivery_status_normalized"] = _STATUS_NORM.get(raw_status, raw_status)
        data.append(d)

    return {"data": data}


@router.post("/deliveries")
async def create_delivery(
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tạo mới delivery record."""
    required = ["po_number", "bqms_code", "specification"]
    for f in required:
        if not body.get(f):
            raise HTTPException(400, f"Trường '{f}' là bắt buộc")

    row = await conn.fetchrow(
        """
        INSERT INTO bqms_deliveries (
            po_number, bqms_code, item_name, specification, quantity, unit,
            unit_price, delivery_status, delivery_date,
            shipping_no, sev_type, buyer_email, recipient_name,
            delivery_method, notes, data_source
        ) VALUES (
            $1, $2, $3, $4, $5, $6,
            $7, $8::delivery_status, $9,
            $10, $11, $12, $13,
            $14, $15, 'manual'
        )
        RETURNING *
        """,
        body.get("po_number"),
        body.get("bqms_code"),
        body.get("item_name"),
        body.get("specification"),
        body.get("quantity"),
        body.get("unit", "EA"),
        body.get("unit_price"),
        body.get("delivery_status", "chua_giao"),
        body.get("delivery_date"),
        body.get("shipping_no"),
        body.get("sev_type"),
        body.get("buyer_email"),
        body.get("recipient_name"),
        body.get("delivery_method"),
        body.get("notes"),
    )
    return {"data": dict(row), "message": "Đã tạo đơn giao hàng"}


@router.put("/deliveries/{delivery_id}")
async def update_delivery(
    delivery_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Cập nhật delivery record với optimistic lock."""
    from app.core.concurrency import emit_record_changed

    existing = await conn.fetchrow(
        "SELECT id, version FROM bqms_deliveries WHERE id = $1", delivery_id
    )
    if not existing:
        raise HTTPException(404, "Đơn giao hàng không tồn tại")

    # Optimistic lock check (if client sends version)
    expected_version = body.pop("version", None)
    if expected_version is not None and existing["version"] != expected_version:
        raise HTTPException(
            409,
            f"Người khác vừa cập nhật đơn này (version DB: {existing['version']}, "
            f"client gửi: {expected_version}). Vui lòng tải lại."
        )

    ALLOWED_FIELDS = {
        "po_number", "bqms_code", "item_name", "specification", "quantity", "unit",
        "unit_price", "delivery_status", "delivery_date", "shipping_no",
        "sev_type", "buyer_email", "buyer_phone", "recipient_name",
        "receiving_warehouse", "delivery_method",
        "notes", "actual_delivered_at", "actual_delivered_qty", "delivery_info",
        "country_origin",  # user-editable per request 2026-05-08
        "driver_id",  # Phase G (Thang 2026-05-13): assign delivery person
    }

    sets = []
    params: list[Any] = []
    idx = 1
    for key, value in body.items():
        if key in ALLOWED_FIELDS:
            if key == "delivery_status":
                sets.append(f"{key} = ${idx}::delivery_status")
            else:
                sets.append(f"{key} = ${idx}")
            params.append(value)
            idx += 1

    if not sets:
        raise HTTPException(400, "Không có trường nào để cập nhật")

    sets.append("updated_at = NOW()")
    params.append(delivery_id)

    row = await conn.fetchrow(
        f"UPDATE bqms_deliveries SET {', '.join(sets)} WHERE id = ${idx} RETURNING *",
        *params,
    )

    # Emit real-time event
    await emit_record_changed("bqms_delivery", delivery_id, "updated", token_data.user_id)

    return {"data": dict(row), "message": "Đã cập nhật đơn giao hàng"}


@router.patch("/deliveries/{delivery_id}/status")
async def update_delivery_status(
    delivery_id: int,
    body: dict[str, Any],
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Cập nhật trạng thái giao hàng."""
    new_status = body.get("status")
    if not new_status:
        raise HTTPException(400, "Trường 'status' là bắt buộc")

    VALID_STATUSES = {
        "chua_giao", "dang_giao", "da_giao", "hoan_tat",
        "pending", "picked_up", "in_transit", "customs_clearance", "delivered", "completed",
    }
    if new_status not in VALID_STATUSES:
        raise HTTPException(400, f"Trạng thái không hợp lệ: {new_status}")

    existing = await conn.fetchrow(
        "SELECT id, delivery_status FROM bqms_deliveries WHERE id = $1", delivery_id
    )
    if not existing:
        raise HTTPException(404, "Đơn giao hàng không tồn tại")

    update_fields = "delivery_status = $2::delivery_status, updated_at = NOW()"
    params: list[Any] = [delivery_id, new_status]

    # Auto-set actual_delivered_at when marking as delivered/completed
    if new_status in ("da_giao", "delivered", "completed", "hoan_tat"):
        update_fields += ", actual_delivered_at = COALESCE(actual_delivered_at, NOW())"

    row = await conn.fetchrow(
        f"UPDATE bqms_deliveries SET {update_fields} WHERE id = $1 RETURNING *",
        *params,
    )

    # Real-time sync
    from app.core.concurrency import emit_record_changed
    await emit_record_changed("bqms_delivery", delivery_id, "status_changed", token_data.user_id,
                              metadata={"new_status": new_status})

    # Event-driven notification for manager + warehouse
    try:
        from app.services.event_notifications import dispatch_delivery_status_change
        await dispatch_delivery_status_change(
            conn, delivery_id, new_status, actor_user_id=str(token_data.user_id),
        )
    except Exception as exc:
        logger.warning("delivery-status notification failed: %s", exc)

    return {"data": dict(row), "message": f"Đã cập nhật trạng thái: {new_status}"}


# ---------------------------------------------------------------------------
# Tạo hồ sơ giao hàng (Create Delivery Dossier) — Thang 2026-05-16
# ---------------------------------------------------------------------------


class _DossierItem(BaseModel):
    """Per-item payload for dossier creation."""
    model_config = ConfigDict(extra="ignore")
    po_number: str
    po_seq: str = ""
    bqms_code: str
    item_name: str = ""
    specification: str = ""
    shipping_qty: float
    dept: str = "MAIN"
    pr_person: str = ""
    receiver: str = ""           # Người nhận hàng (Cam kết hình ảnh sheet)
    unit: str = "PC"
    dim_l: str = ""
    dim_w: str = ""
    dim_h: str = ""
    box_weight: float | None = None  # Box Weight per item (blank stays blank, no auto-sync)
    packing_size: str = ""       # Packing Size MM (col N) — MANUAL per-item input
    box_qty: float | None = None     # Box Qty (col O) — MANUAL per-item input, blank allowed


class _DossierLabel(BaseModel):
    """One editable Label "tem" block (FE Label tab). qty is the EDITED value."""
    model_config = ConfigDict(extra="ignore")
    po_number: str = ""
    pr_person: str = ""
    bqms_code: str = ""
    qty: float | None = None


class _CreateDossierBody(BaseModel):
    model_config = ConfigDict(extra="ignore")
    sev_type: str  # 'SEV' or 'SEVT'
    items: list[_DossierItem]
    vendor_invoice_no: str
    invoice_date: str       # YYYY-MM-DD
    etd: str                # YYYY-MM-DD
    packing_qty: float = 1
    packing_unit: str = "Box"
    volume: float = 0
    volume_unit: str = "M3"
    gross_weight: float = 0
    weight_unit: str = "KG"
    remark: str = ""
    shipping_manager: str = "AMA Bac Ninh JSC"
    # PRINT-ONLY Box-Qty TOTAL override for Packing List (null = computed sum).
    box_qty_total_override: float | None = None
    # Editable Label tab payload. None/absent → builder uses legacy per-PO path.
    labels: list[_DossierLabel] | None = None


@router.post("/deliveries/create-dossier")
async def create_delivery_dossier(
    body: _CreateDossierBody,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Enqueue a delivery dossier job. Returns job_id for poll.

    Validates:
      - sev_type ∈ {SEV, SEVT}
      - items not empty + all share company (rejected if mixed)
      - all (po_number, bqms_code) tuples exist in bqms_deliveries
    """
    if body.sev_type not in ("SEV", "SEVT"):
        raise HTTPException(400, "sev_type must be SEV or SEVT")
    if not body.items:
        raise HTTPException(400, "items không được rỗng")

    # Concurrency guards (B3 — Thang 2026-05-18):
    # 1. Per-user cap: max 3 active (queued|running) jobs at a time
    active_count = await conn.fetchval(
        "SELECT COUNT(*) FROM bqms_dossier_jobs "
        "WHERE user_id = $1::uuid AND status IN ('queued','running')",
        token_data.user_id,
    )
    if active_count and int(active_count) >= 3:
        raise HTTPException(
            429,
            "Bạn đang có 3 hồ sơ đang xử lý — chờ xong rồi tạo tiếp",
        )
    # 2. Global queue cap: max 10 jobs in queue → tránh overload
    queued_total = await conn.fetchval(
        "SELECT COUNT(*) FROM bqms_dossier_jobs WHERE status IN ('queued','running')"
    )
    if queued_total and int(queued_total) >= 10:
        raise HTTPException(
            503,
            "Hàng đợi đầy — vui lòng thử lại sau vài phút",
        )

    # Verify items + collect delivery row IDs
    distinct_pos: list[str] = []
    delivery_ids: list[int] = []
    found_company_set: set[str] = set()
    for it in body.items:
        if it.po_number not in distinct_pos:
            distinct_pos.append(it.po_number)
        row = await conn.fetchrow(
            "SELECT id, sev_type FROM bqms_deliveries WHERE po_number = $1 AND bqms_code = $2 "
            "ORDER BY id DESC LIMIT 1",
            it.po_number, it.bqms_code,
        )
        if not row:
            raise HTTPException(400, f"Không tìm thấy delivery row cho PO {it.po_number} / BQMS {it.bqms_code}")
        delivery_ids.append(int(row["id"]))
        if row["sev_type"]:
            found_company_set.add(row["sev_type"])
    if found_company_set and len(found_company_set) > 1:
        raise HTTPException(400,
            f"Mix SEV/SEVT không cho phép — items có {sorted(found_company_set)}. Tạo riêng từng hồ sơ.")

    # Insert job row
    job_row = await conn.fetchrow(
        """
        INSERT INTO bqms_dossier_jobs
            (user_id, sev_type, po_numbers, delivery_row_ids, form_data, status, progress_step)
        VALUES ($1::uuid, $2, $3, $4, $5::jsonb, 'queued', 'Đang chờ Samsung session')
        RETURNING id
        """,
        token_data.user_id, body.sev_type, distinct_pos, delivery_ids,
        _json.dumps(body.model_dump(), default=str, ensure_ascii=False),
    )
    job_id = int(job_row["id"])

    # Enqueue Procrastinate task — App not auto-opened in FastAPI context.
    try:
        from app.tasks.bqms_dossier import bqms_create_delivery_dossier as task_fn
        from app.core.procrastinate_app import app as proc_app
        async with proc_app.open_async():
            prc_job_id = await task_fn.defer_async(job_id=job_id)
        await conn.execute(
            "UPDATE bqms_dossier_jobs SET procrastinate_job_id = $1 WHERE id = $2",
            prc_job_id, job_id,
        )
    except Exception as exc:
        logger.exception("defer task failed for dossier job=%d: %s", job_id, exc)
        await conn.execute(
            "UPDATE bqms_dossier_jobs SET status='failed', error=$1 WHERE id=$2",
            f"defer failed: {exc}", job_id,
        )
        raise HTTPException(500, f"Không enqueue được task: {exc}")

    return {
        "data": {
            "job_id": job_id,
            "sev_type": body.sev_type,
            "po_numbers": distinct_pos,
            "items_count": len(body.items),
        },
        "message": f"Đã tạo job #{job_id}. Theo dõi qua GET /deliveries/dossier-job/{job_id}",
    }


@router.get("/deliveries/dossier-job/{job_id}")
async def get_dossier_job(
    job_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Poll job status. Frontend gọi mỗi 4s khi modal mở.

    Phản hồi extra fields cho concurrency UX:
      - queue_position: số job 'queued' xếp trước nếu chưa chạy (0 = đang chạy / đã xong)
      - eta_seconds: estimate dựa trên queue_position × 600s (cap 1h)
    """
    row = await conn.fetchrow(
        "SELECT * FROM bqms_dossier_jobs WHERE id = $1", job_id,
    )
    if not row:
        raise HTTPException(404, f"Job #{job_id} không tồn tại")
    out = dict(row)
    # Parse JSON fields for frontend
    for k in ("form_data", "files", "confirm_preview"):
        v = out.get(k)
        if isinstance(v, str):
            try:
                out[k] = _json.loads(v)
            except Exception:
                pass

    # Confirm checkpoint UX: image URL + countdown remaining (5-min auto-cancel)
    if out.get("status") == "awaiting_confirm":
        out["confirm_image_url"] = (
            f"/api/v1/bqms/deliveries/dossier-job/{job_id}/confirm-image"
        )
        from datetime import datetime, timezone
        ac = out.get("awaiting_confirm_at")
        if isinstance(ac, datetime):
            elapsed = (datetime.now(timezone.utc) - ac).total_seconds()
            out["confirm_remaining_seconds"] = max(0, int(300 - elapsed))

    # Queue position (only when status='queued')
    queue_position = 0
    eta_seconds = 0
    if out.get("status") == "queued":
        q_row = await conn.fetchrow(
            """
            SELECT
                (SELECT COUNT(*) FROM bqms_dossier_jobs
                  WHERE status = 'queued' AND id < $1) AS ahead,
                (SELECT COUNT(*) FROM bqms_dossier_jobs
                  WHERE status = 'running') AS running
            """,
            job_id,
        )
        if q_row:
            queue_position = int(q_row["ahead"]) + int(q_row["running"])
            # Rough ETA: each job ~10 min (600s). Cap at 1 hour.
            eta_seconds = min(3600, queue_position * 600)
    out["queue_position"] = queue_position
    out["eta_seconds"] = eta_seconds

    # Detect stuck jobs (status=running but last_heartbeat_at > 5 min ago)
    if out.get("status") == "running" and out.get("last_heartbeat_at"):
        from datetime import datetime, timezone, timedelta
        hb = out["last_heartbeat_at"]
        if isinstance(hb, datetime):
            age = (datetime.now(timezone.utc) - hb).total_seconds()
            out["heartbeat_age_seconds"] = int(age)
            if age > 300:
                out["stuck_warning"] = (
                    f"Job có thể bị treo — không có heartbeat trong {int(age // 60)} phút"
                )
    return {"data": out}


@router.post("/deliveries/dossier-job/{job_id}/upload-image")
async def upload_dossier_image(
    job_id: int,
    bqms_code: str | None = Form(None),  # optional — derived from item_key if absent
    slot: str = Form("actual"),  # 'actual' | 'system' (user replaces system image)
    item_key: str = Form(None),  # composite `${po_number}|${po_seq}|${bqms_code}`
    file: UploadFile = File(...),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Upload "Hệ thống" hoặc "Thực tế" image cho 1 item trong job.

    Cam kết sheets are PER ITEM (po_number, po_seq, bqms_code), so the filename
    is keyed by the sanitized composite `item_key`. The same bqms_code can
    appear on >1 sheet (different PO/seq) without colliding.

    Saved at `/data/bqms-push-evidence/dossier/{job_id}/{sanitized_key}_{slot}.png`
    where sanitized_key = re.sub(r"[^A-Za-z0-9_-]", "_", item_key). If item_key
    is absent (older client) it falls back to bqms_code (backward compatible).
    Task sẽ pick up khi build Excel.
    """
    import re as _re
    # Thang 2026-06-26 FIX: item_key (po|seq|code) is the source of truth. The FE
    # sends item_key but historically forgot bqms_code → this endpoint 422'd and
    # silently dropped EVERY pasted "Thực tế" image. Derive bqms_code from item_key
    # when absent so the upload (and the embedded Excel image) succeed.
    if not bqms_code and item_key:
        bqms_code = (item_key.rsplit("|", 1)[-1] or "").strip()
    if not bqms_code:
        raise HTTPException(400, "Thiếu bqms_code (hoặc item_key) để định danh ảnh")
    if not _re.match(r"^[A-Z0-9\-_]+$", bqms_code):
        raise HTTPException(400, "bqms_code chỉ chấp nhận [A-Z0-9-_]")
    if slot not in ("actual", "system"):
        raise HTTPException(400, "slot phải là 'actual' hoặc 'system'")
    # Sanitize the composite item_key with the SAME rule the build task uses in
    # _evidence_key(), so the saved filename matches the task's reconstructed
    # filename byte-for-byte. Fall back to bqms_code when item_key is omitted.
    file_key = _re.sub(r"[^A-Za-z0-9_\-]", "_", item_key) if item_key else bqms_code
    ext = Path(file.filename or "").suffix.lower()
    if ext not in (".png", ".jpg", ".jpeg"):
        raise HTTPException(400, "Chỉ chấp nhận .png/.jpg/.jpeg")

    job = await conn.fetchrow(
        "SELECT id, status FROM bqms_dossier_jobs WHERE id = $1", job_id,
    )
    if not job:
        raise HTTPException(404, f"Job #{job_id} không tồn tại")
    if job["status"] in ("done", "failed"):
        raise HTTPException(409, f"Job đã ở status={job['status']}, không upload thêm được")

    target_dir = Path(f"/data/bqms-push-evidence/dossier/{job_id}")
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{file_key}_{slot}.png"
    MAX = 5 * 1024 * 1024
    total = 0
    with open(target, "wb") as f:
        while True:
            chunk = await file.read(64 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > MAX:
                f.close()
                target.unlink(missing_ok=True)
                raise HTTPException(413, "Ảnh quá lớn (>5MB)")
            f.write(chunk)

    return {
        "data": {"job_id": job_id, "bqms_code": bqms_code, "slot": slot,
                 "item_key": item_key, "file_key": file_key,
                 "path": str(target), "size_bytes": total},
        "message": f"Đã upload ảnh {slot} cho {bqms_code}",
    }


# ---------------------------------------------------------------------------
# Confirm checkpoint — user kiểm tra 100% trước khi scraper bấm Create Delivery
# (KHÔNG HOÀN TÁC). Job ở status 'awaiting_confirm' chờ tín hiệu confirm/cancel.
# ---------------------------------------------------------------------------


async def _set_confirm_signal(conn, job_id: int, signal: str) -> dict:
    """Set confirm_signal nếu job đang ở awaiting_confirm. Task poll sẽ nhặt."""
    row = await conn.fetchrow(
        "SELECT id, status FROM bqms_dossier_jobs WHERE id = $1", job_id,
    )
    if not row:
        raise HTTPException(404, f"Job #{job_id} không tồn tại")
    if row["status"] != "awaiting_confirm":
        raise HTTPException(
            409,
            f"Job đang ở status={row['status']}, không phải đang chờ xác nhận",
        )
    await conn.execute(
        "UPDATE bqms_dossier_jobs SET confirm_signal = $1 WHERE id = $2",
        signal, job_id,
    )
    return {"job_id": job_id, "signal": signal}


@router.post("/deliveries/dossier-job/{job_id}/confirm")
async def confirm_dossier_job(
    job_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """User xác nhận → scraper sẽ bấm Create Delivery (tạo Delivery thật)."""
    data = await _set_confirm_signal(conn, job_id, "confirm")
    return {"data": data, "message": "Đã xác nhận — đang tạo Delivery trên Samsung"}


@router.post("/deliveries/dossier-job/{job_id}/cancel")
async def cancel_dossier_job(
    job_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """User huỷ tại checkpoint → đóng popup, KHÔNG tạo Delivery."""
    data = await _set_confirm_signal(conn, job_id, "cancel")
    return {"data": data, "message": "Đã huỷ — không tạo Delivery"}


@router.get("/deliveries/dossier-job/{job_id}/confirm-image")
async def get_dossier_confirm_image(
    job_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
):
    """Stream screenshot popup Create Delivery đã điền (cho bước kiểm tra)."""
    from fastapi.responses import FileResponse
    shot = Path(f"/data/bqms-push-evidence/dossier/{job_id}/confirm_preview.png")
    if not shot.exists():
        raise HTTPException(404, "Chưa có screenshot kiểm tra")
    return FileResponse(str(shot), media_type="image/png")


# ---------------------------------------------------------------------------
# Re-edit / regenerate (EXCEL-ONLY) — mở lại 1 hồ sơ ĐÃ HOÀN TẤT, sửa form +
# ảnh, dựng lại file Excel ghi đè trong output_folder đã lưu.
#
# SAFETY (lý do tồn tại của feature): regenerate KHÔNG chạy scraper Samsung,
# KHÔNG acquire samsung_session_lock, KHÔNG mở popup Create Delivery (không hoàn
# tác — đã làm rồi), KHÔNG re-parse Shipping No, KHÔNG chạy UPDATE cộng dồn
# actual_delivered_qty trên bqms_deliveries (sẽ double-count). CHỈ Excel.
# Reuse shipping_no/output_folder đã lưu.
# ---------------------------------------------------------------------------


@router.get("/deliveries/dossier-jobs")
async def list_dossier_jobs(
    limit: int = Query(50, ge=1, le=200),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Liệt kê hồ sơ gần đây cho picker "mở lại để sửa"."""
    rows = await conn.fetch(
        """
        SELECT id, sev_type, po_numbers, status, output_folder,
               created_at, updated_at,
               (form_data->>'vendor_invoice_no') AS invoice_no
          FROM bqms_dossier_jobs
         ORDER BY created_at DESC
         LIMIT $1
        """,
        limit,
    )
    return {"data": [dict(r) for r in rows]}


@router.get("/deliveries/dossier-job/{job_id}/image")
async def get_dossier_job_image(
    job_id: int,
    item_key: str = Query(...),
    slot: str = Query("actual"),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
):
    """Stream lại ảnh evidence ĐÃ upload để wizard hydrate khi sửa.

    Đọc từ `/data/bqms-push-evidence/dossier/{job_id}/{sanitize(item_key)}_{slot}.png`
    với sanitize IDENTICAL upload endpoint + task (_evidence_key).
    """
    import re as _re
    if slot not in ("actual", "system"):
        raise HTTPException(400, "slot phải là 'actual' hoặc 'system'")
    file_key = _re.sub(r"[^A-Za-z0-9_\-]", "_", item_key)
    p = Path(f"/data/bqms-push-evidence/dossier/{job_id}/{file_key}_{slot}.png")
    if not p.exists():
        raise HTTPException(404, "Không có ảnh đã upload cho item này")
    from fastapi.responses import FileResponse
    return FileResponse(
        str(p), media_type="image/png",
        headers={"Cache-Control": "private, max-age=60"},
    )


@router.get("/deliveries/dossier-job/{job_id}/file")
async def download_dossier_job_file(
    job_id: int,
    kind: str = Query("excel"),       # excel | delivery_note | po
    po: str | None = Query(None),     # bắt buộc khi kind='po'
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tải file đã tạo của 1 hồ sơ: Excel / Delivery Note PDF / PO PDF.

    Đường dẫn lấy từ `bqms_dossier_jobs.files` (JSONB). Chống path-traversal:
    file BẮT BUỘC nằm trong thư mục gốc BBGH (.../Giao hàng).
    """
    import os
    import json as _json2
    row = await conn.fetchrow(
        "SELECT files FROM bqms_dossier_jobs WHERE id = $1", job_id,
    )
    if not row:
        raise HTTPException(404, f"Job #{job_id} không tồn tại")
    files = row["files"]
    if isinstance(files, str):
        files = _json2.loads(files) if files else {}
    files = files or {}

    if kind == "excel":
        path = files.get("excel")
    elif kind == "delivery_note":
        path = files.get("delivery_note")
    elif kind == "po":
        if not po:
            raise HTTPException(400, "thiếu tham số 'po'")
        path = next(
            (p.get("path") for p in (files.get("po_pdfs") or [])
             if str(p.get("po")) == str(po)),
            None,
        )
    else:
        raise HTTPException(400, "kind phải là excel | delivery_note | po")

    if not path:
        raise HTTPException(404, "Hồ sơ chưa có file này")

    base = os.path.realpath("/data/onedrive-staging/Puplic/BQMS/Giao hàng")
    real = os.path.realpath(str(path))
    if real != base and not real.startswith(base + os.sep):
        raise HTTPException(403, "Đường dẫn không hợp lệ")
    if not os.path.isfile(real):
        raise HTTPException(404, "File không còn trên máy chủ")

    fname = os.path.basename(real)
    low = real.lower()
    media = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if low.endswith(".xlsx")
        else "application/pdf" if low.endswith(".pdf")
        else "application/octet-stream"
    )
    from fastapi.responses import FileResponse
    # Starlette encodes non-ASCII filenames (RFC 6266 filename*) automatically.
    return FileResponse(real, media_type=media, filename=fname)


def _zip_folder_to_bytes(folder_path: str) -> bytes:
    """Zip toàn bộ file trong folder (đệ quy) → bytes. Chạy trong threadpool."""
    import io
    import os
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rootd, _dirs, fnames in os.walk(folder_path):
            for fn in fnames:
                fp = os.path.join(rootd, fn)
                zf.write(fp, os.path.relpath(fp, folder_path))
    return buf.getvalue()


@router.get("/deliveries/dossier-job/{job_id}/folder.zip")
async def download_dossier_folder_zip(
    job_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tải TOÀN BỘ thư mục hồ sơ (1 đợt giao) dưới dạng .zip — Excel + PDF + ảnh."""
    import asyncio as _asyncio
    import os
    from urllib.parse import quote
    row = await conn.fetchrow(
        "SELECT output_folder FROM bqms_dossier_jobs WHERE id = $1", job_id,
    )
    if not row or not row["output_folder"]:
        raise HTTPException(404, "Hồ sơ chưa có thư mục")
    base = os.path.realpath("/data/onedrive-staging/Puplic/BQMS/Giao hàng")
    real = os.path.realpath(str(row["output_folder"]))
    if real != base and not real.startswith(base + os.sep):
        raise HTTPException(403, "Đường dẫn không hợp lệ")
    if not os.path.isdir(real):
        raise HTTPException(404, "Thư mục không còn trên máy chủ")

    data = await _asyncio.to_thread(_zip_folder_to_bytes, real)
    zip_name = os.path.basename(real) + ".zip"
    from fastapi.responses import Response
    return Response(
        content=data,
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{quote(zip_name)}",
            "Content-Length": str(len(data)),
        },
    )


@router.post("/deliveries/dossier-job/{job_id}/update-regenerate")
async def update_regenerate_dossier_job(
    job_id: int,
    body: _CreateDossierBody,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Sửa form_data của 1 hồ sơ ĐÃ HOÀN TẤT rồi dựng lại CHỈ file Excel.

    SAFETY: chỉ enqueue task Excel-only `bqms_regenerate_dossier_excel`. Task này
    KHÔNG chạy scraper Samsung / popup Create Delivery / qty-accumulation. Reuse
    shipping_no + output_folder đã lưu trên job row.
    """
    job = await conn.fetchrow(
        "SELECT id, status FROM bqms_dossier_jobs WHERE id = $1", job_id,
    )
    if not job:
        raise HTTPException(404, f"Job #{job_id} không tồn tại")
    if job["status"] != "done":
        raise HTTPException(409, "hồ sơ đang xử lý hoặc chưa hoàn tất")

    await conn.execute(
        """
        UPDATE bqms_dossier_jobs
           SET form_data    = $1::jsonb,
               status       = 'regenerating',
               progress_pct = 0,
               progress_step = 'Đang cập nhật Excel',
               updated_at   = NOW()
         WHERE id = $2
        """,
        _json.dumps(body.model_dump(), default=str, ensure_ascii=False),
        job_id,
    )

    # Enqueue Excel-only regenerate task (KHÔNG phải task scraper) — match
    # create-dossier defer style.
    try:
        from app.tasks.bqms_dossier import bqms_regenerate_dossier_excel as task_fn
        from app.core.procrastinate_app import app as proc_app
        async with proc_app.open_async():
            prc_job_id = await task_fn.defer_async(job_id=job_id)
        await conn.execute(
            "UPDATE bqms_dossier_jobs SET procrastinate_job_id = $1 WHERE id = $2",
            prc_job_id, job_id,
        )
    except Exception as exc:
        logger.exception("defer regenerate failed for dossier job=%d: %s", job_id, exc)
        await conn.execute(
            "UPDATE bqms_dossier_jobs SET status='failed', error=$1 WHERE id=$2",
            f"defer regenerate failed: {exc}", job_id,
        )
        raise HTTPException(500, f"Không enqueue được task: {exc}")

    return {"data": {"job_id": job_id}, "message": "Đang cập nhật hồ sơ"}


# ---------------------------------------------------------------------------
# Dossier prefill + system image streaming endpoints (M40 wizard support)
# ---------------------------------------------------------------------------


@router.post("/deliveries/dossier-prefill")
async def dossier_prefill(
    body: dict,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Prefill data cho dossier wizard.

    Body: `{delivery_ids: [int]}` — list bqms_deliveries.id user đã chọn từ trang Giao hàng.

    Trả về structured data cho wizard:
    - sev_type detected (reject if mixed)
    - distinct_po_numbers
    - items (per delivery row, đã enriched từ bqms_rfq + system image URL)
    - default form values (vendor_invoice_no theo {DDMMYYYY}-{N}, dates today)
    """
    delivery_ids = body.get("delivery_ids") or []
    if not isinstance(delivery_ids, list) or not delivery_ids:
        raise HTTPException(400, "delivery_ids không được rỗng")
    try:
        delivery_ids = [int(x) for x in delivery_ids]
    except Exception:
        raise HTTPException(400, "delivery_ids phải là list số nguyên")

    # Thang 2026-06-01: dùng LATERAL pick-latest để KHÔNG nhân đôi rows.
    # bqms_rfq có thể có nhiều entries cùng bqms_code (RFQ qua nhiều lần) →
    # LEFT JOIN thẳng sẽ làm 1 delivery row × N RFQ rows = N items hiển sai.
    rows = await conn.fetch(
        """
        SELECT d.id, d.po_number, d.bqms_code, d.sev_type,
               d.item_name, d.specification, d.unit,
               COALESCE(d.quantity, 0) AS quantity,
               COALESCE(d.actual_delivered_qty, 0) AS actual_qty,
               d.recipient_name, d.receiving_warehouse,
               r.rfq_number, r.person_in_charge_name, r.department
          FROM bqms_deliveries d
          LEFT JOIN LATERAL (
              SELECT rfq_number, person_in_charge_name, department
                FROM bqms_rfq rr
               WHERE rr.bqms_code = d.bqms_code
               ORDER BY COALESCE(rr.inquiry_date, rr.created_at::date) DESC NULLS LAST,
                        rr.id DESC
               LIMIT 1
          ) r ON true
         WHERE d.id = ANY($1::bigint[])
         ORDER BY d.po_number, d.bqms_code
        """,
        delivery_ids,
    )
    if not rows:
        raise HTTPException(404, "Không tìm thấy delivery rows")

    # History lookup: per bqms_code, find LATEST values from past successful dossier jobs.
    # This auto-prefills fields like dim_l/w/h, box_weight, dept, pr_person, receiver
    # so user doesn't have to re-type them for the same BQMS code.
    bqms_codes = [r["bqms_code"] for r in rows if r["bqms_code"]]
    history: dict[str, dict] = {}
    if bqms_codes:
        hist_rows = await conn.fetch(
            """
            WITH unfolded AS (
              SELECT
                item->>'bqms_code' AS bqms_code,
                item AS data,
                j.created_at,
                ROW_NUMBER() OVER (
                  PARTITION BY item->>'bqms_code'
                  ORDER BY j.created_at DESC
                ) AS rn
              FROM bqms_dossier_jobs j
              CROSS JOIN LATERAL jsonb_array_elements(j.form_data->'items') AS item
              WHERE j.status = 'done'
                AND item->>'bqms_code' = ANY($1::text[])
            )
            SELECT bqms_code, data
              FROM unfolded
             WHERE rn = 1
            """,
            bqms_codes,
        )
        for h in hist_rows:
            try:
                history[h["bqms_code"]] = _json.loads(h["data"]) if isinstance(h["data"], str) else h["data"]
            except Exception:
                continue

    sev_types = {r["sev_type"] for r in rows if r["sev_type"]}
    if len(sev_types) > 1:
        raise HTTPException(
            400,
            f"Mix SEV/SEVT không cho phép — items có {sorted(sev_types)}. "
            "Tạo riêng từng hồ sơ.",
        )
    sev_type = next(iter(sev_types)) if sev_types else "SEV"

    # Build vendor_invoice_no theo pattern {DDMMYYYY}-{N} với counter từ DB
    from datetime import datetime as _dt
    today_prefix = _dt.now().strftime("%d%m%Y")
    counter_row = await conn.fetchrow(
        """
        SELECT COUNT(*) + 1 AS next_n
          FROM bqms_dossier_jobs
         WHERE form_data->>'vendor_invoice_no' LIKE $1
        """,
        f"{today_prefix}-%",
    )
    next_n = int(counter_row["next_n"]) if counter_row else 1

    # Build items với system image URL (relative — frontend prefix với API base)
    from app.services.dossier_image_resolver import find_system_image
    items_out = []
    distinct_pos: list[str] = []
    for r in rows:
        if r["po_number"] not in distinct_pos:
            distinct_pos.append(r["po_number"])
        # Smart image lookup: tries override → exact RFQ → broad scan (last 2 years)
        sys_img_path = find_system_image(r["bqms_code"], r["rfq_number"])
        spec = r["specification"] or ""
        # History lookup (if any) — auto-prefill dims/box_weight/dept/etc.
        h = history.get(r["bqms_code"]) or {}

        def hist_str(key: str, fallback: str) -> str:
            v = h.get(key)
            return str(v) if v not in (None, "") else fallback

        # Thang 2026-06-01: trỏ system_image_url thẳng tới /rfq/image — nguồn ảnh
        # chung với trang BQMS list (5 layer priority: P0 picker-pinned →
        # P1 per-RFQ override → P2 code-override folder → P2.5 image index →
        # P3 FS scan). Trước đây dùng endpoint riêng /dossier-system-image chỉ
        # 4 layer + mất context rfq_number → user pin ảnh ở BQMS không thấy ở wizard.
        from urllib.parse import quote as _q
        if sys_img_path:
            params = f"bqms_code={_q(r['bqms_code'])}"
            if r["rfq_number"]:
                params += f"&rfq_number={_q(r['rfq_number'])}"
            img_url = f"/api/v1/bqms/rfq/image?{params}"
        else:
            img_url = None

        items_out.append({
            "delivery_id": int(r["id"]),
            "po_number": r["po_number"],
            "po_seq": "",  # bqms_deliveries không lưu po_seq — scraper sẽ tự match qua grid
            "bqms_code": r["bqms_code"],
            "rfq_number": r["rfq_number"],  # expose cho frontend dùng cho picker + image URL
            # Item Name ← new split 'item_name' column (deployed today). Fallback
            # to job-history, then first spec segment / bqms_code.
            "item_name": (r["item_name"] or "").strip() or hist_str(
                "item_name", spec.split(",")[0].strip()[:40] if spec else r["bqms_code"]),
            "specification": hist_str("specification", spec),
            "unit": h.get("unit") or r["unit"] or "PC",
            "ordered_qty": float(r["quantity"] or 0),
            "remaining_qty": max(0, float(r["quantity"] or 0) - float(r["actual_qty"] or 0)),
            # Issue 5/6 (Thang LOCKED): per-PO Dept ← delivery 'receiving_warehouse'
            # (Kho Nhận), fallback RFQ 'department' only when empty. NEW SOURCE WINS
            # over legacy job-history prefill — do NOT read history for these.
            "dept": (r["receiving_warehouse"] or r["department"] or "").strip(),
            # PR Person ← delivery 'recipient_name' (Người nhận). New source wins.
            "pr_person": (r["recipient_name"] or "").strip(),
            # Receiver ← BLANK (Cam kết C10 left empty, no prefill). Thang LOCKED.
            "receiver": "",
            "dim_l": hist_str("dim_l", ""),
            "dim_w": hist_str("dim_w", ""),
            "dim_h": hist_str("dim_h", ""),
            "box_weight": float(h.get("box_weight") or 0),
            "has_system_image": sys_img_path is not None,
            "system_image_url": img_url,
            "has_history": bool(h),
        })

    # Multi-delivery history per PO (Thang 2026-05-21): query past dossiers
    # that touched any of the selected POs so the UI can show "Đây là lần N
    # của PO này" and the user knows about previous shipments.
    delivery_history: list[dict] = []
    # Parallel record carrying the FULL parsed form_data dict per attempt so we
    # can prefill the HEADER from the last attempt without an extra DB round-trip.
    # (delivery_history itself only keeps a trimmed `items` projection.)
    attempts_full: list[dict] = []
    if distinct_pos:
        hist_rows = await conn.fetch(
            """
            SELECT po_number, dossier_id, attempt_no, shipping_no,
                   invoice_no, status, is_partial, output_folder,
                   form_data, created_at
              FROM v_po_delivery_history
             WHERE po_number = ANY($1::text[])
               AND sev_type = $2
             ORDER BY po_number, attempt_no
            """,
            distinct_pos, sev_type,
        )
        for r in hist_rows:
            # Parse form_data once (may arrive as JSON string or dict).
            fd = r["form_data"] or {}
            if isinstance(fd, str):
                try:
                    fd = _json.loads(fd)
                except Exception:
                    fd = {}
            if not isinstance(fd, dict):
                fd = {}
            # Extract qty shipped per item from form_data.items[]
            items_info = []
            try:
                for it in (fd.get("items") or []):
                    if it.get("po_number") == r["po_number"]:
                        items_info.append({
                            "bqms_code": it.get("bqms_code"),
                            "shipping_qty": it.get("shipping_qty"),
                        })
            except Exception:
                pass
            delivery_history.append({
                "po_number": r["po_number"],
                "dossier_id": r["dossier_id"],
                "attempt_no": r["attempt_no"],
                "shipping_no": r["shipping_no"],
                "invoice_no": r["invoice_no"],
                "status": r["status"],
                "is_partial": r["is_partial"],
                "output_folder": r["output_folder"],
                "items": items_info,
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            })
            attempts_full.append({
                "po_number": r["po_number"],
                "attempt_no": r["attempt_no"],
                "status": r["status"],
                "created_at": r["created_at"],
                "form_data": fd,
            })

    # Compute next attempt_no per PO so UI shows "lần 2" / "lần 3" preview.
    next_attempt_by_po: dict[str, int] = {}
    for po in distinct_pos:
        prev = [h for h in delivery_history if h["po_number"] == po
                and h["status"] in ("done", "queued", "running",
                                    "invoice_ready", "po_downloaded", "excel_built")]
        next_attempt_by_po[po] = (
            max((h["attempt_no"] for h in prev), default=0) + 1
        )

    # Header prefill from the LAST attempt of the dominant PO (Thang 2026-06-22).
    # Goal: a repeat delivery of the same PO should not require re-typing the
    # invoice/packing header — reuse the previous attempt's values.
    #
    # Dominant PO = the selected PO with the MOST RECENT prior attempt. We rank
    # all candidate attempts by (created_at, attempt_no) desc and take the top
    # one's PO; that PO's latest attempt becomes the header source. This also
    # naturally satisfies "first selected PO that HAS history" — only POs with
    # history produce candidates.
    #
    # "Completed dossier" filter: prefer status='done' (fully finished). If no
    # 'done' attempt exists yet, fall back to the broader in-pipeline set so a
    # quick repeat delivery (while the prior job is still running) still prefills.
    # form_data here is already parsed to a dict (or {} on garbage) above.
    header_from_last_attempt = None

    def _pick_dominant_header(status_whitelist: tuple[str, ...]):
        cands = [a for a in attempts_full if a["status"] in status_whitelist]
        if not cands:
            return None
        # Sort newest-first; created_at may be None → push to the end.
        from datetime import datetime as __dt, timezone as __tz
        _min = __dt.min.replace(tzinfo=__tz.utc)

        def _key(a):
            ca = a["created_at"]
            if ca is not None and ca.tzinfo is None:
                ca = ca.replace(tzinfo=__tz.utc)
            return (ca or _min, a["attempt_no"] or 0)

        cands.sort(key=_key, reverse=True)
        return cands[0]

    chosen = _pick_dominant_header(("done",)) or _pick_dominant_header(
        ("invoice_ready", "po_downloaded", "excel_built", "running", "queued")
    )
    if chosen:
        fd = chosen["form_data"] if isinstance(chosen["form_data"], dict) else {}

        def _h(key):
            v = fd.get(key)
            return v if v not in (None, "") else None

        # box_l/w/h are NOT stored in the header — _DossierItem stores them
        # per-item as dim_l/dim_w/dim_h. Derive from the first item that has a
        # non-empty dimension; omit (null) if none. Documented choice.
        box_l = box_w = box_h = None
        try:
            for it in (fd.get("items") or []):
                if not isinstance(it, dict):
                    continue
                dl, dw, dh = it.get("dim_l"), it.get("dim_w"), it.get("dim_h")
                if any(x not in (None, "") for x in (dl, dw, dh)):
                    box_l = dl if dl not in (None, "") else None
                    box_w = dw if dw not in (None, "") else None
                    box_h = dh if dh not in (None, "") else None
                    break
        except Exception:
            pass

        ca = chosen["created_at"]
        header_from_last_attempt = {
            "attempt_no": chosen["attempt_no"],
            "attempt_date": ca.isoformat() if ca else None,
            "po_number": chosen["po_number"],
            "vendor_invoice_no": _h("vendor_invoice_no"),
            "invoice_date": _h("invoice_date"),
            "etd": _h("etd"),
            "packing_qty": _h("packing_qty"),
            "packing_unit": _h("packing_unit"),
            "volume": _h("volume"),
            "volume_unit": _h("volume_unit"),
            "gross_weight": _h("gross_weight"),
            "weight_unit": _h("weight_unit"),
            "box_l": box_l,
            "box_w": box_w,
            "box_h": box_h,
            "shipping_manager": _h("shipping_manager"),
            "remark": _h("remark"),
        }

    return {
        "data": {
            "sev_type": sev_type,
            "distinct_po_numbers": distinct_pos,
            "items": items_out,
            "delivery_history": delivery_history,
            "next_attempt_by_po": next_attempt_by_po,
            "header_from_last_attempt": header_from_last_attempt,
            "defaults": {
                "vendor_invoice_no": f"{today_prefix}-{next_n:02d}",
                "invoice_date": _dt.now().strftime("%Y-%m-%d"),
                "etd": _dt.now().strftime("%Y-%m-%d"),
                "packing_qty": 1,
                "packing_unit": "Box",
                "volume": 0.001,
                "volume_unit": "M3",
                "gross_weight": 1.0,
                "weight_unit": "KG",
                "shipping_manager": "AMA Bac Ninh JSC",
                "remark": "",
            },
        },
    }


@router.get("/deliveries/dossier-system-image/{bqms_code}")
async def dossier_system_image(
    bqms_code: str,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "warehouse", "sales", "procurement", "accountant",
    )),
):
    """Stream system image cho 1 BQMS code (từ RFQ folder).

    Returns 404 nếu không tìm thấy. Cache-Control 1h.
    """
    import re as _re
    if not _re.match(r"^[A-Z0-9\-_]+$", bqms_code):
        raise HTTPException(400, "bqms_code chỉ chấp nhận [A-Z0-9-_]")

    from app.services.dossier_image_resolver import find_system_image
    img_path = find_system_image(bqms_code, None)
    if not img_path or not img_path.exists():
        raise HTTPException(404, f"Không tìm thấy ảnh hệ thống cho {bqms_code}")

    from fastapi.responses import FileResponse
    media_type = "image/png" if img_path.suffix.lower() == ".png" else "image/jpeg"
    return FileResponse(
        str(img_path), media_type=media_type,
        headers={"Cache-Control": "private, max-age=3600"},
    )


# ---------------------------------------------------------------------------
# Contacts (DANH BẠ)
# ---------------------------------------------------------------------------

@router.get("/contacts")
async def list_contacts(
    q: str | None = Query(None),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh bạ liên hệ Samsung — 1,028 contacts từ DANH BẠ sheet."""
    if q:
        rows = await conn.fetch(
            """
            SELECT * FROM bqms_contacts
            WHERE full_name ILIKE $1 OR email_username ILIKE $1 OR phone ILIKE $1
            ORDER BY full_name
            """,
            f"%{q}%",
        )
    else:
        rows = await conn.fetch("SELECT * FROM bqms_contacts ORDER BY full_name")
    return {"data": [dict(r) for r in rows], "total": len(rows)}


# Driver CRUD endpoints (was here, extracted to bqms_drivers.py in PR-1
# 2026-05-13). bqms_drivers_router is mounted with the same /bqms prefix
# so all /drivers/* routes remain unchanged from the client's perspective.


# ---------------------------------------------------------------------------
# Direct PO API — /bqms/po/* (sitemap-derived endpoints, item 1+2)
#
# Plan: BQMS_SITEMAP/BQMS_SITEMAP.md §2.2 / §2.3
# ---------------------------------------------------------------------------

from pydantic import BaseModel as _BaseModel, Field as _Field


class _POInfoItem(_BaseModel):
    poNo: str
    poSeq: str
    poStatus: str = "PO2"
    secureKey: str


class _ConfirmRequest(_BaseModel):
    pos: list[_POInfoItem] = _Field(..., min_length=1, max_length=200)


@router.get("/po/all")
async def fetch_all_pos_endpoint(
    date_from: str = Query(..., description="YYYYMMDD or YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYYMMDD or YYYY-MM-DD"),
    status: str = Query("", description="'Y'=confirmed, 'N'=not, '' = both"),
    company_code: str = Query("", description="C5H0 / C5H2 / ..."),
    page_size: int = Query(99999, ge=10, le=99999),
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
):
    """Fetch ALL POs from Samsung BQMS for a date range (item 1).

    Read-only — calls Samsung selectPOAcceptList.do directly. Logs in via
    Playwright session, then issues a single httpx POST.
    """
    from app.etl.bqms_po_api import fetch_all_pos
    statuses = [s for s in (status or "").split(",") if s.strip()] or None
    pos = await fetch_all_pos(
        date_from=date_from,
        date_to=date_to,
        status_codes=statuses,
        company_code=company_code,
        page_size=page_size,
    )
    return {
        "data": {
            "items": pos,
            "total": len(pos),
            "confirmed": sum(1 for p in pos if p.get("SP_CONFIRM_FLAG") == "Y"),
            "not_confirmed": sum(1 for p in pos if p.get("SP_CONFIRM_FLAG") == "N"),
        }
    }


@router.post("/po/confirm")
async def confirm_pos_endpoint(
    body: _ConfirmRequest,
    token_data: TokenData = Depends(require_role("manager", "admin")),
):
    """Confirm a batch of POs on Samsung BQMS (item 2 — PRODUCTION WRITE).

    Manager / admin only. Each PO must include the latest secureKey from a
    fresh /po/all fetch — Samsung rejects stale tokens.
    """
    from app.etl.bqms_po_api import confirm_pos
    payload = [p.model_dump() for p in body.pos]
    try:
        result = await confirm_pos(payload)
    except Exception as exc:
        logger.exception("BQMS confirm_pos failed")
        raise HTTPException(status_code=502, detail=f"BQMS confirm error: {exc}")
    return {"data": result, "message": f"Đã gửi confirm cho {len(payload)} PO"}


@router.post("/po/cancel-confirm")
async def cancel_confirm_pos_endpoint(
    body: _ConfirmRequest,
    token_data: TokenData = Depends(require_role("admin")),
):
    """Cancel previously-confirmed POs (admin recovery — undoes /po/confirm)."""
    from app.etl.bqms_po_api import cancel_confirm_pos
    payload = [p.model_dump() for p in body.pos]
    try:
        result = await cancel_confirm_pos(payload)
    except Exception as exc:
        logger.exception("BQMS cancel_confirm_pos failed")
        raise HTTPException(status_code=502, detail=f"BQMS cancel-confirm error: {exc}")
    return {"data": result, "message": f"Đã hủy confirm cho {len(payload)} PO"}



# ---------------------------------------------------------------------------
# Won quotations (sheet TRUNG BG of Thong ke hoi hang BQMS.xlsx)
# ---------------------------------------------------------------------------

from pydantic import BaseModel


class _WonQuotationPatch(BaseModel):
    hs_code: str | None = None
    goods_description: str | None = None


@router.get("/won-quotations")
async def list_won_quotations(
    search: str | None = Query(None, description="Tìm theo HS code, BQMS code, RFQ No, NCC, mô tả"),
    has_hs: str | None = Query(None, description="filled | missing | all"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=10, le=500),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List won quotations from sheet TRUNG BG.

    Filters:
    - search: matched against rfq_number, bqms_code, hs_code, supplier_name,
      specification, description, goods_description (case-insensitive substring).
    - has_hs: 'filled' = has HS code; 'missing' = NULL or empty.
    """
    conds = ["1=1"]
    params: list[Any] = []
    idx = 1

    if search:
        like = f"%{search.strip()}%"
        conds.append(
            f"(w.rfq_number ILIKE ${idx} OR w.bqms_code ILIKE ${idx} OR w.hs_code ILIKE ${idx} "
            f"OR w.supplier_name ILIKE ${idx} OR w.specification ILIKE ${idx} "
            f"OR w.description ILIKE ${idx} OR w.goods_description ILIKE ${idx})"
        )
        params.append(like)
        idx += 1

    if has_hs == "filled":
        conds.append("w.hs_code IS NOT NULL AND w.hs_code <> ''")
    elif has_hs == "missing":
        conds.append("(w.hs_code IS NULL OR w.hs_code = '')")

    where = " AND ".join(conds)
    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM bqms_won_quotations w WHERE {where}", *params
    )

    offset = (page - 1) * page_size
    params_paged = params + [page_size, offset]
    # Phase 4.2 (Thang 2026-05-12): LEFT JOIN bqms_contracts để show contract_no
    # badge trên won-quotations table. Match by rfq_number (request_no).
    rows = await conn.fetch(
        f"""
        SELECT w.id, w.rfq_number, w.bqms_code, w.person_in_charge_name,
               w.description, w.specification, w.quantity, w.unit,
               w.po_price, w.po_deadline, w.supplier_name,
               w.hs_code, w.goods_description, w.customs_char_count, w.notes,
               w.synced_at, w.created_at,
               c.id          AS contract_id,
               c.contract_no AS contract_no,
               c.status      AS contract_status
        FROM bqms_won_quotations w
        LEFT JOIN LATERAL (
            SELECT id, contract_no, status
            FROM bqms_contracts
            WHERE request_no = w.rfq_number
            ORDER BY contract_start DESC NULLS LAST, id DESC
            LIMIT 1
        ) c ON true
        WHERE {where}
        ORDER BY w.synced_at DESC NULLS LAST, w.id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params_paged,
    )

    def _ser(r: asyncpg.Record) -> dict:
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                try:
                    d[k] = float(v)
                except Exception:
                    pass
        return d

    return {
        "data": {
            "items": [_ser(r) for r in rows],
            "total": int(total or 0),
            "page": page,
            "page_size": page_size,
        }
    }


# ---------------------------------------------------------------------------
# Phase 3.2 (Thang 2026-05-12): bulk HS code lookup — user paste list of
# bqms_codes → return existing HS code + description for each. Speeds up
# the Trúng BG / HS code research from one-by-one to copy-paste-tra hàng loạt.
# ---------------------------------------------------------------------------

class _BulkHsLookupIn(_BaseModel):
    codes: list[str]


@router.post("/hs-code/bulk-lookup")
async def hs_code_bulk_lookup(
    body: _BulkHsLookupIn,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales",
        "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Bulk lookup HS code + goods_description for a list of BQMS codes.

    Strategy: query bqms_won_quotations grouped by bqms_code, return the most
    recent row's HS data for each code. If no row exists, returns null fields
    so frontend can highlight missing items for manual entry.
    """
    codes_clean = [c.strip() for c in (body.codes or []) if c and c.strip()]
    if not codes_clean:
        return {"data": {"items": [], "missing": []}}
    if len(codes_clean) > 200:
        raise HTTPException(400, "Tối đa 200 mã/lần")

    rows = await conn.fetch(
        """
        SELECT DISTINCT ON (bqms_code)
            bqms_code, hs_code, goods_description, description,
            specification, supplier_name, po_price,
            synced_at, customs_char_count
        FROM bqms_won_quotations
        WHERE bqms_code = ANY($1::text[])
        ORDER BY bqms_code, synced_at DESC NULLS LAST, id DESC
        """,
        codes_clean,
    )

    found_map = {}
    for r in rows:
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                try:
                    d[k] = float(v)
                except Exception:
                    pass
        found_map[d["bqms_code"]] = d

    items = []
    missing = []
    for code in codes_clean:
        if code in found_map:
            items.append(found_map[code])
        else:
            missing.append(code)
            items.append({
                "bqms_code": code,
                "hs_code": None,
                "goods_description": None,
                "description": None,
                "specification": None,
                "supplier_name": None,
                "po_price": None,
                "synced_at": None,
                "customs_char_count": None,
            })

    return {
        "data": {
            "items": items,
            "missing": missing,
            "found_count": len(codes_clean) - len(missing),
            "missing_count": len(missing),
            "total": len(codes_clean),
        }
    }


@router.get("/staging/contracts")
async def list_staging_contracts(
    status: str | None = Query(None, description="pending_review | merged | all"),
    search: str | None = Query(None, description="Filter by rfq_number/contract_no/specification"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=10, le=500),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List Contract Mgmt staging rows (raw scrape from sec-bqms.com).
    Per Thang 2026-05-11: surface contract module data as a tab inside
    /bqms/won-quotations so user can review pending contracts before merge.
    """
    conds = ["module = 'contract'"]
    params: list[Any] = []
    idx = 1
    if status and status != "all":
        conds.append(f"status = ${idx}"); params.append(status); idx += 1
    if search:
        conds.append(
            f"(rfq_number ILIKE ${idx} OR contract_no ILIKE ${idx} OR specification ILIKE ${idx})"
        )
        params.append(f"%{search.strip()}%"); idx += 1
    where = " AND ".join(conds)

    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM bqms_vendor_portal_staging WHERE {where}", *params,
    )
    offset = (page - 1) * page_size
    rows = await conn.fetch(
        f"""
        SELECT id, scraped_at, status, rfq_number, contract_no, contract_period,
               item_code, description, specification, quantity, unit, raw_json
        FROM bqms_vendor_portal_staging
        WHERE {where}
        ORDER BY scraped_at DESC, id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params, page_size, offset,
    )

    def _ser(r: asyncpg.Record) -> dict:
        d = dict(r)
        raw = d.pop("raw_json", None) or {}
        if isinstance(raw, str):
            try: raw = _json.loads(raw)
            except Exception: raw = {}
        # Lift fields from raw_json.contract for richer display
        c = (raw.get("contract") or {})
        d["contract_amount"] = c.get("amount")
        d["contract_status"] = c.get("status")
        d["contract_subject"] = c.get("subject")
        d["contract_kind"] = c.get("contract_kind")
        d["created_by"] = c.get("created_by")
        d["items_count"] = len((c.get("items") or [raw.get("item")]) or [])
        for k, v in list(d.items()):
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                try: d[k] = float(v)
                except Exception: pass
        return d

    return {
        "data": {
            "items": [_ser(r) for r in rows],
            "total": int(total or 0),
            "page": page,
            "page_size": page_size,
        }
    }


@router.get("/staging/mro")
async def list_staging_mro(
    status: str | None = Query(None, description="pending_review | merged | all"),
    search: str | None = Query(None, description="Filter by rfq_number/PO/item code"),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=10, le=500),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List MRO P/O Receipt staging rows (module='po' from MRO scraper).
    Per Thang 2026-05-11: surface MRO data as a tab inside /bqms/deliveries.
    """
    conds = ["module = 'po'"]
    params: list[Any] = []
    idx = 1
    if status and status != "all":
        conds.append(f"status = ${idx}"); params.append(status); idx += 1
    if search:
        conds.append(
            f"(rfq_number ILIKE ${idx} OR item_code ILIKE ${idx} OR specification ILIKE ${idx})"
        )
        params.append(f"%{search.strip()}%"); idx += 1
    where = " AND ".join(conds)

    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM bqms_vendor_portal_staging WHERE {where}", *params,
    )
    offset = (page - 1) * page_size
    rows = await conn.fetch(
        f"""
        SELECT id, scraped_at, status, rfq_number,
               item_code, description, specification, quantity, unit, raw_json
        FROM bqms_vendor_portal_staging
        WHERE {where}
        ORDER BY scraped_at DESC, id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params, page_size, offset,
    )

    def _ser(r: asyncpg.Record) -> dict:
        d = dict(r)
        raw = d.pop("raw_json", None) or {}
        if isinstance(raw, str):
            try: raw = _json.loads(raw)
            except Exception: raw = {}
        # Lift MRO-specific fields from raw_json
        d["po_no"] = raw.get("PO_NO")
        d["po_status"] = raw.get("PO_STATUS_NAME")
        d["vendor"] = raw.get("SP_NAME")
        d["plant"] = raw.get("PLANT_NAME")
        d["buying_price"] = raw.get("BUYING_PRICE")
        d["buying_amount"] = raw.get("BUYING_AMOUNT")
        d["buying_currency"] = raw.get("BUYING_CURRENCY")
        d["receiver"] = raw.get("RECEIVER_NAME")
        d["delivery_address"] = raw.get("DELIVERY_ADDRESS")
        d["req_delivery_date"] = raw.get("REQ_DELIVERY_DATE")
        d["po_qty"] = raw.get("PO_QTY")
        d["item_cate"] = raw.get("ITEM_CATE")
        d["manufacturer"] = raw.get("MANUFACTURER")
        d["cis_code"] = raw.get("CIS_CODE")
        for k, v in list(d.items()):
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                try: d[k] = float(v)
                except Exception: pass
        return d

    return {
        "data": {
            "items": [_ser(r) for r in rows],
            "total": int(total or 0),
            "page": page,
            "page_size": page_size,
        }
    }


@router.patch("/won-quotations/{won_id}")
async def update_won_quotation(
    won_id: int,
    body: _WonQuotationPatch,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Edit hs_code and/or goods_description for a won quotation row.

    Only these two fields are user-editable; everything else syncs from
    Excel TRUNG BG sheet via the auto-importer.
    """
    sets = []
    params: list[Any] = []
    idx = 1

    if body.hs_code is not None:
        sets.append(f"hs_code = ${idx}")
        params.append(body.hs_code.strip() or None)
        idx += 1
    if body.goods_description is not None:
        sets.append(f"goods_description = ${idx}")
        params.append(body.goods_description.strip() or None)
        idx += 1

    if not sets:
        raise HTTPException(400, "Chưa có trường nào để cập nhật")

    params.append(won_id)
    sql = f"UPDATE bqms_won_quotations SET {', '.join(sets)} WHERE id = ${idx} RETURNING id, hs_code, goods_description"
    row = await conn.fetchrow(sql, *params)
    if not row:
        raise HTTPException(404, f"Won quotation id={won_id} không tồn tại")

    return {"data": dict(row), "message": "Cập nhật thành công"}


# ---------------------------------------------------------------------------
# Thang 2026-06-13: Manual won-data refresh per RFQ.
# Dispatches Procrastinate task `bqms_sync_won_for_rfq` which acquires
# the Samsung session lock, drills contract list, and UPSERTs into
# bqms_won_quotations. UI binds this to a "Cập nhật Trúng BG" button on
# the WonQuotation drawer / table row.
# ---------------------------------------------------------------------------

@router.post("/won-quotations/refresh/{rfq_number}")
async def refresh_won_quotation(
    rfq_number: str,
    token_data: TokenData = Depends(
        require_role("admin", "manager", "staff", "sales", "procurement")
    ),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Manually refresh won-quotation data for one RFQ.

    Dispatches a background Procrastinate task that:
      1. Acquires Samsung session lock.
      2. Drills the Samsung Vendor Portal contract list.
      3. Upserts rows from staging into bqms_won_quotations.

    Returns the deferred job_id immediately; client polls
    /push-queue/status or audit_log for completion.
    """
    rfq_number = (rfq_number or "").strip()
    if not rfq_number:
        raise HTTPException(400, "rfq_number is required")

    # Sanity check: RFQ must exist in bqms_rfq (otherwise the refresh would
    # never produce a matching won row).
    exists = await conn.fetchval(
        "SELECT 1 FROM bqms_rfq WHERE rfq_number = $1 LIMIT 1", rfq_number,
    )
    if not exists:
        raise HTTPException(
            404, f"RFQ {rfq_number} không tồn tại trong bqms_rfq",
        )

    # Defer the Procrastinate task. Import inside the handler so the FastAPI
    # process doesn't pay the procrastinate import cost when the endpoint
    # is never called.
    try:
        from app.tasks.bqms_won_sync import bqms_sync_won_for_rfq
        job_id = await bqms_sync_won_for_rfq.defer_async(
            rfq_number=rfq_number,
            user_id=str(token_data.user_id),
            timestamp=int(datetime.now().timestamp()),
        )
    except Exception as exc:
        logger.exception("defer bqms_sync_won_for_rfq failed: %s", exc)
        raise HTTPException(500, f"Không dispatch được task: {exc}")

    # Audit — surface the trigger in BQMS history drawer.
    try:
        await conn.execute(
            """
            INSERT INTO audit_log
                (user_id, action, table_name, record_id, new_data, created_at)
            VALUES ($1::uuid, 'bqms.won_sync.dispatched', 'bqms_won_quotations',
                    $2, $3::jsonb, NOW())
            """,
            token_data.user_id, rfq_number,
            _json.dumps({"rfq_number": rfq_number, "job_id": job_id}),
        )
    except Exception as exc:
        logger.warning("audit_log won_sync_dispatched failed: %s", exc)

    logger.info(
        "won_sync dispatched: rfq=%s job_id=%s by=%s",
        rfq_number, job_id, token_data.user_id,
    )
    return {
        "data": {
            "rfq_number": rfq_number,
            "job_id": job_id,
            "status": "queued",
        },
        "message": (
            f"Đã dispatch task cập nhật Trúng BG cho RFQ {rfq_number}. "
            "Theo dõi tiến độ qua audit_log hoặc poll /won-quotations sau ~1 phút."
        ),
    }


# ---------------------------------------------------------------------------
# Thang 2026-06-13 (Bug fix T4): Global won-data refresh — NO rfq_number path
# param. UI's RefreshWonButton calls POST /api/v1/bqms/won/refresh with no
# body. Dispatches the periodic won-sync drain task (idempotent contract
# staging -> bqms_won_quotations UPSERT, no Samsung scrape) so user gets an
# instant "all pending picked up" sync without per-RFQ targeting.
# Kept alongside /won-quotations/refresh/{rfq_number} (per-RFQ) so both
# entry points work — the per-RFQ one is still used by the WonQuotation
# drawer's single-row refresh.
# ---------------------------------------------------------------------------

@router.post("/won/refresh")
async def refresh_won_quotations_global(
    token_data: TokenData = Depends(
        require_role("admin", "manager", "staff", "sales")
    ),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Refresh ALL pending won/contract data in one go.

    Dispatches the `bqms_sync_won_for_all_pending` task which drains all
    pending rows from contract staging into bqms_won_quotations. No Samsung
    scrape is triggered — this is the "pick up everything already scraped"
    sync, idempotent and UNGATED (vs the periodic task which is OFF by
    default behind app_config).

    Returns the deferred job_id immediately; client toasts a success
    message and invalidates the BQMS RFQ table cache.
    """
    try:
        from app.tasks.bqms_won_sync import bqms_sync_won_for_all_pending
        job_id = await bqms_sync_won_for_all_pending.defer_async(
            user_id=str(token_data.user_id),
            timestamp=int(datetime.now().timestamp()),
        )
    except Exception as exc:
        logger.exception("defer bqms_won_sync_periodic failed: %s", exc)
        raise HTTPException(500, f"Không dispatch được task: {exc}")

    # Audit — surface the global trigger in the BQMS history drawer.
    try:
        await conn.execute(
            """
            INSERT INTO audit_log
                (user_id, action, table_name, record_id, new_data, created_at)
            VALUES ($1::uuid, 'bqms.won_sync.dispatched_global',
                    'bqms_won_quotations', NULL, $2::jsonb, NOW())
            """,
            token_data.user_id,
            _json.dumps({"scope": "all_pending", "job_id": job_id}),
        )
    except Exception as exc:
        logger.warning("audit_log won_sync_dispatched_global failed: %s", exc)

    logger.info(
        "won_sync GLOBAL dispatched: job_id=%s by=%s",
        job_id, token_data.user_id,
    )
    return {
        "data": {
            "task_id": job_id,
            "message": "Đang đồng bộ dữ liệu trúng...",
        },
        "message": "Đang đồng bộ dữ liệu trúng...",
    }


# ---------------------------------------------------------------------------
# Origin summary (multi-select rows -> 2-col table BQMS code | Xuất xứ)
# ---------------------------------------------------------------------------

class _OriginSummaryRequest(BaseModel):
    ids: list[int]


@router.post("/deliveries/origin-summary")
async def deliveries_origin_summary(
    body: _OriginSummaryRequest,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """For a list of delivery row IDs, return [{bqms_code, country_origin}].
    Used by the "Thống kê xuất xứ" multi-select feature in the deliveries
    table — output mirrors the form Thang sketched (BQMS code | Xuất xứ).
    Sorted by bqms_code for stable ordering.
    """
    if not body.ids:
        return {"data": {"items": [], "total": 0}}
    if len(body.ids) > 1000:
        raise HTTPException(400, "Tối đa 1000 dòng/lần thống kê")

    rows = await conn.fetch(
        """
        SELECT DISTINCT bqms_code, country_origin
        FROM bqms_deliveries
        WHERE id = ANY($1::bigint[])
          AND bqms_code IS NOT NULL
          AND bqms_code <> ''
        ORDER BY bqms_code
        """,
        body.ids,
    )

    items = [
        {"bqms_code": r["bqms_code"], "country_origin": r["country_origin"] or ""}
        for r in rows
    ]
    return {"data": {"items": items, "total": len(items)}}


# ─── Bulk delivery lookup (paste BQMS codes) — Thang 2026-06-01 ──────
# Tương tự /bqms/hs-code/bulk-lookup ở trang Trúng BG. User paste danh sách
# BQMS code → trả về toàn bộ delivery rows cho các code đó (bao gồm: PO, qty,
# shipping_no, delivery_date, status, sev_type). Mã không có delivery → vẫn
# trả 1 placeholder để user thấy mã đó chưa giao.


class _BulkDeliveryLookupRequest(BaseModel):
    codes: list[str]


@router.post("/deliveries/bulk-lookup")
async def deliveries_bulk_lookup(
    body: _BulkDeliveryLookupRequest,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant", "viewer"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tra cứu giao hàng hàng loạt theo danh sách BQMS code.

    Body: {codes: ["Z0000002-...", "...", ...]} (tối đa 200 mã).

    Trả về:
        - items: list delivery rows (1 mã có thể có nhiều dòng = nhiều PO/lần giao).
        - missing_codes: mã không tìm thấy delivery nào.
        - found_codes: mã có ít nhất 1 delivery.
    """
    # Normalize input: trim, uppercase, dedup, drop empty
    raw_codes = body.codes or []
    codes = sorted({c.strip().upper() for c in raw_codes if c and c.strip()})
    if not codes:
        raise HTTPException(400, "Danh sách mã rỗng")
    if len(codes) > 200:
        raise HTTPException(400, f"Tối đa 200 mã/lần (đang có {len(codes)})")

    rows = await conn.fetch(
        """
        SELECT
            d.id, d.po_number, d.po_date, d.bqms_code, d.specification, d.unit,
            d.quantity, d.unit_price, d.amount,
            d.actual_delivered_qty, d.shipping_no, d.delivery_date,
            d.delivery_status, d.delivery_status_normalized,
            d.sev_type, d.country_origin, d.recipient_name,
            d.receiving_warehouse, d.delivery_method,
            d.quotation_no, d.total_delivered_value_vnd,
            d.created_at, d.updated_at
        FROM bqms_deliveries d
        WHERE UPPER(d.bqms_code) = ANY($1::text[])
        ORDER BY d.bqms_code, d.po_number, d.po_date DESC NULLS LAST, d.id DESC
        """,
        codes,
    )

    items = [dict(r) for r in rows]
    found = {r["bqms_code"].upper() for r in rows if r["bqms_code"]}
    missing = [c for c in codes if c not in found]

    # Aggregate per code: count + sum qty + last delivery
    summary: dict[str, dict] = {}
    for r in rows:
        key = (r["bqms_code"] or "").upper()
        if key not in summary:
            summary[key] = {
                "bqms_code": r["bqms_code"],
                "count": 0,
                "total_quantity": 0,
                "total_delivered_qty": 0,
                "last_delivery_date": None,
                "last_shipping_no": None,
                "latest_status": None,
                "po_numbers": set(),
            }
        s = summary[key]
        s["count"] += 1
        try:
            s["total_quantity"] += float(r["quantity"] or 0)
        except Exception:
            pass
        try:
            s["total_delivered_qty"] += float(r["actual_delivered_qty"] or 0)
        except Exception:
            pass
        if r["po_number"]:
            s["po_numbers"].add(r["po_number"])
        if r["delivery_date"] and (s["last_delivery_date"] is None or r["delivery_date"] > s["last_delivery_date"]):
            s["last_delivery_date"] = r["delivery_date"]
            s["last_shipping_no"] = r["shipping_no"]
            s["latest_status"] = r["delivery_status"]

    summary_list = []
    for code in codes:
        s = summary.get(code)
        if s:
            summary_list.append({
                "bqms_code": s["bqms_code"],
                "count": s["count"],
                "total_quantity": s["total_quantity"],
                "total_delivered_qty": s["total_delivered_qty"],
                "remaining_qty": max(0, s["total_quantity"] - s["total_delivered_qty"]),
                "last_delivery_date": s["last_delivery_date"],
                "last_shipping_no": s["last_shipping_no"],
                "latest_status": s["latest_status"],
                "po_numbers": sorted(s["po_numbers"]),
                "found": True,
            })
        else:
            summary_list.append({
                "bqms_code": code,
                "count": 0,
                "total_quantity": 0,
                "total_delivered_qty": 0,
                "remaining_qty": 0,
                "last_delivery_date": None,
                "last_shipping_no": None,
                "latest_status": None,
                "po_numbers": [],
                "found": False,
            })

    return {
        "data": {
            "items": items,
            "summary": summary_list,
            "found_codes": sorted(found),
            "missing_codes": missing,
            "found_count": len(found),
            "missing_count": len(missing),
            "total_rows": len(items),
        }
    }


# ---------------------------------------------------------------------------
# Vendor Portal scraper — Contract Mgmt
# ---------------------------------------------------------------------------

@router.post("/scrape-contracts")
async def scrape_contracts_trigger(
    limit: int = Query(10, ge=0, le=50, description="Max contracts to drill (0 = all on first page)"),
    drill_items: bool = Query(True, description="Click into each contract to fetch Item Information"),
    dry_run: bool = Query(True, description="When true, save raw JSON only — no INSERT into staging"),
    token_data: TokenData = Depends(require_role("admin")),
):
    """Trigger a one-shot scrape of Vendor Portal -> Contract Mgmt.

    Returns the run summary (run_id, counts, json_path). When dry_run=False,
    rows are also INSERTed into `bqms_vendor_portal_staging` with status
    'pending_review' for human approval before merging into bqms_won_quotations.

    Synchronous — blocks for the duration of the scrape (~12s/contract).
    Single login per call enforced by the scraper module itself.
    """
    import threading
    import asyncio as _aio

    result_holder: dict[str, Any] = {}

    def _thread_scrape() -> None:
        async def _do() -> None:
            from app.etl.bqms_contract_scraper import scrape_contracts
            from app.core.config import settings as cfg
            import asyncpg as apg

            db_pool = None
            if not dry_run:
                db_url = (
                    str(cfg.DATABASE_URL)
                    .replace("+asyncpg", "")
                    .replace("postgresql+asyncpg", "postgresql")
                )
                db_pool = await apg.create_pool(db_url, min_size=1, max_size=2)
            try:
                payload = await scrape_contracts(
                    limit=limit,
                    drill_items=drill_items,
                    save_raw_json=True,
                    db_pool=db_pool,
                )
                result_holder["payload"] = payload
            except Exception as exc:
                logger.exception("scrape_contracts failed")
                result_holder["error"] = str(exc)[:500]
            finally:
                if db_pool is not None:
                    await db_pool.close()

        _aio.run(_do())

    t = threading.Thread(target=_thread_scrape, daemon=True)
    t.start()
    # Cap thread wait at 5 minutes (50 contracts * ~12s headroom) — protects
    # the request worker if BQMS hangs.
    t.join(timeout=300)

    if t.is_alive():
        raise HTTPException(504, "Scrape timed out after 5 minutes")

    if "error" in result_holder:
        raise HTTPException(500, f"Scrape failed: {result_holder['error']}")

    payload = result_holder.get("payload") or {}
    # Strip the heavy `items` array from response — caller reads json_path or
    # GET /bqms/vendor-staging for the actual rows.
    summary = {k: v for k, v in payload.items() if k != "items"}
    summary["item_total"] = sum(len(c.get("items") or []) for c in payload.get("items", []))
    summary["mode"] = "dry_run" if dry_run else "staged"
    logger.info(
        "scrape_contracts done by user=%s: run_id=%s drilled=%d mode=%s",
        token_data.user_id, summary.get("run_id"), summary.get("drilled_count"), summary["mode"],
    )
    return {"data": summary}


@router.post("/scrape-mro-po")
async def scrape_mro_po_trigger(
    limit: int = Query(0, ge=0, le=200, description="Max rows (0 = all on first page)"),
    dry_run: bool = Query(True, description="When true, save raw JSON only — no INSERT into staging"),
    token_data: TokenData = Depends(require_role("admin")),
):
    """Trigger a one-shot scrape of Vendor Portal -> Execution > MRO > P/O Receipt.

    The MRO P/O list is inline-rich (item_code, spec, qty, unit_price all in
    list columns) so this scraper is single-pass — no per-row drill needed.
    Faster and cheaper than the Contract Mgmt scraper.

    Rows land in bqms_vendor_portal_staging with module='po'.
    """
    import threading
    import asyncio as _aio

    result_holder: dict[str, Any] = {}

    def _thread_scrape() -> None:
        async def _do() -> None:
            from app.etl.bqms_mro_scraper import scrape_mro_po
            from app.core.config import settings as cfg
            import asyncpg as apg

            db_pool = None
            if not dry_run:
                db_url = (
                    str(cfg.DATABASE_URL)
                    .replace("+asyncpg", "")
                    .replace("postgresql+asyncpg", "postgresql")
                )
                db_pool = await apg.create_pool(db_url, min_size=1, max_size=2)
            try:
                payload = await scrape_mro_po(
                    limit=limit,
                    save_raw_json=True,
                    db_pool=db_pool,
                )
                result_holder["payload"] = payload
            except Exception as exc:
                logger.exception("scrape_mro_po failed")
                result_holder["error"] = str(exc)[:500]
            finally:
                if db_pool is not None:
                    await db_pool.close()

        _aio.run(_do())

    t = threading.Thread(target=_thread_scrape, daemon=True)
    t.start()
    t.join(timeout=120)
    if t.is_alive():
        raise HTTPException(504, "MRO scrape timed out after 2 minutes")
    if "error" in result_holder:
        raise HTTPException(500, f"MRO scrape failed: {result_holder['error']}")

    payload = result_holder.get("payload") or {}
    summary = {k: v for k, v in payload.items() if k != "items"}
    summary["mode"] = "dry_run" if dry_run else "staged"
    logger.info(
        "scrape_mro_po done by user=%s: run_id=%s rows=%d mode=%s",
        token_data.user_id, summary.get("run_id"), summary.get("list_count"), summary["mode"],
    )
    return {"data": summary}


@router.post("/scrape-bidding")
async def scrape_bidding_trigger(
    limit: int = Query(0, ge=0, le=200, description="Max rows (0 = all on chosen page)"),
    dry_run: bool = Query(True, description="When true, save raw JSON only — no INSERT into staging"),
    drill_details: bool = Query(False, description="Click each subject to fetch Quotation Amount items"),
    page_size: int = Query(100, ge=10, le=100, description="Rows per page (10/30/50/100)"),
    page_num: int = Query(1, ge=1, le=100, description="Page number (1-based, ~98 pages of 10)"),
    smart_skip: bool = Query(True, description="Skip drill+staging for QTs already in bqms_rfq with images"),
    token_data: TokenData = Depends(require_role("admin")),
):
    """Trigger a one-shot scrape of Bidding · Quotation Submit.

    Bidding is REQUEST-level (1 RFQ = 1 row) — invitations to quote, not
    won items. Lands in staging with module='bidding' for visibility only;
    the merge-approved endpoint does NOT push these into bqms_won_quotations.

    Has an IBSheet locale popup — handled internally by the scraper.
    """
    import threading
    import asyncio as _aio

    result_holder: dict[str, Any] = {}

    def _thread_scrape() -> None:
        async def _do() -> None:
            from app.etl.bqms_bidding_scraper import scrape_bidding
            from app.core.config import settings as cfg
            import asyncpg as apg

            db_pool = None
            if not dry_run:
                db_url = (
                    str(cfg.DATABASE_URL)
                    .replace("+asyncpg", "")
                    .replace("postgresql+asyncpg", "postgresql")
                )
                db_pool = await apg.create_pool(db_url, min_size=1, max_size=2)
            try:
                payload = await scrape_bidding(
                    limit=limit,
                    save_raw_json=True,
                    db_pool=db_pool,
                    drill_details=drill_details,
                    page_size=page_size,
                    page_num=page_num,
                    smart_skip=smart_skip,
                )
                result_holder["payload"] = payload
            except Exception as exc:
                logger.exception("scrape_bidding failed")
                result_holder["error"] = str(exc)[:500]
            finally:
                if db_pool is not None:
                    await db_pool.close()

        _aio.run(_do())

    t = threading.Thread(target=_thread_scrape, daemon=True)
    t.start()
    # Drill mode: 10 RFQ × ~10s/drill ≈ 100s + login overhead → cap 5 min
    timeout_s = 300 if drill_details else 120
    t.join(timeout=timeout_s)
    if t.is_alive():
        raise HTTPException(504, f"Bidding scrape timed out after {timeout_s}s")
    if "error" in result_holder:
        raise HTTPException(500, f"Bidding scrape failed: {result_holder['error']}")

    payload = result_holder.get("payload") or {}
    summary = {k: v for k, v in payload.items() if k != "items"}
    summary["mode"] = "dry_run" if dry_run else "staged"
    logger.info(
        "scrape_bidding done by user=%s: run_id=%s rows=%d/%d mode=%s",
        token_data.user_id, summary.get("run_id"),
        summary.get("list_count"), summary.get("total_available"), summary["mode"],
    )
    return {"data": summary}


# PR-2 (Thang 2026-05-13): /bidding/folder + /rfq/image + /bidding/folder/file
# extracted to bqms_images.py — mounted with same /bqms prefix.


@router.post("/scrape-announcement")
async def scrape_announcement_trigger(
    limit: int = Query(0, ge=0, le=200),
    dry_run: bool = Query(True),
    page_size: int = Query(100, ge=10, le=100),
    page_num: int = Query(1, ge=1, le=100),
    token_data: TokenData = Depends(require_role("admin")),
):
    """Phase L1 — Bidding · Quotation Announcement (menu 5).
    REQUEST-level invitations, không auto-merge."""
    import threading
    import asyncio as _aio
    result_holder: dict[str, Any] = {}

    def _thread() -> None:
        async def _do() -> None:
            from app.etl.bqms_l1_l3_scraper import scrape_announcement
            from app.core.config import settings as cfg
            import asyncpg as apg
            db_pool = None
            if not dry_run:
                db_url = str(cfg.DATABASE_URL).replace("+asyncpg", "").replace("postgresql+asyncpg", "postgresql")
                db_pool = await apg.create_pool(db_url, min_size=1, max_size=2)
            try:
                payload = await scrape_announcement(
                    limit=limit, save_raw_json=True, db_pool=db_pool,
                    page_size=page_size, page_num=page_num,
                )
                result_holder["payload"] = payload
            except Exception as exc:
                logger.exception("scrape_announcement failed")
                result_holder["error"] = str(exc)[:500]
            finally:
                if db_pool: await db_pool.close()
        _aio.run(_do())

    t = threading.Thread(target=_thread, daemon=True); t.start(); t.join(timeout=180)
    if t.is_alive(): raise HTTPException(504, "Announcement scrape timed out")
    if "error" in result_holder: raise HTTPException(500, result_holder["error"])
    p = result_holder.get("payload") or {}
    return {"data": {k: v for k, v in p.items() if k != "items"}}


@router.post("/scrape-selection-result")
async def scrape_selection_trigger(
    limit: int = Query(0, ge=0, le=200),
    dry_run: bool = Query(True),
    auto_mark_result: bool = Query(True, description="Auto-update bqms_rfq.result based on Selected/Unselected"),
    page_size: int = Query(100, ge=10, le=100),
    page_num: int = Query(1, ge=1, le=100),
    token_data: TokenData = Depends(require_role("admin")),
):
    """Phase L3 — Selection Result (menu 18).
    Side effect: auto-mark bqms_rfq.result = 'won' / 'lost' from Selected/Unselected.
    """
    import threading
    import asyncio as _aio
    result_holder: dict[str, Any] = {}

    def _thread() -> None:
        async def _do() -> None:
            from app.etl.bqms_l1_l3_scraper import scrape_selection_result
            from app.core.config import settings as cfg
            import asyncpg as apg
            db_pool = None
            if not dry_run:
                db_url = str(cfg.DATABASE_URL).replace("+asyncpg", "").replace("postgresql+asyncpg", "postgresql")
                db_pool = await apg.create_pool(db_url, min_size=1, max_size=2)
            try:
                payload = await scrape_selection_result(
                    limit=limit, save_raw_json=True, db_pool=db_pool,
                    auto_mark_result=auto_mark_result,
                    page_size=page_size, page_num=page_num,
                )
                result_holder["payload"] = payload
            except Exception as exc:
                logger.exception("scrape_selection_result failed")
                result_holder["error"] = str(exc)[:500]
            finally:
                if db_pool: await db_pool.close()
        _aio.run(_do())

    t = threading.Thread(target=_thread, daemon=True); t.start(); t.join(timeout=180)
    if t.is_alive(): raise HTTPException(504, "Selection scrape timed out")
    if "error" in result_holder: raise HTTPException(500, result_holder["error"])
    p = result_holder.get("payload") or {}
    return {"data": {k: v for k, v in p.items() if k != "items"}}


async def _ensure_ondemand_drill(
    conn: asyncpg.Connection,
    staging_id: int,
    rfq_number: str,
    user_id,
) -> dict[str, Any]:
    """Trigger (or re-use) a single-RFQ on-demand drill for `staging_id`.

    Thang 2026-06-23 — fix nút Báo giá "không hoạt động" khi RFQ chưa drill.

    REUSES the existing quote-batch infra (bqms_quote_batches +
    bqms_quote_batch_items + `quote_one_rfq_task`). That task already:
      • acquires `samsung_session_lock` (serialize với push/scrape — KHÔNG mở
        session Samsung song song)
      • login → drill detail + download + extract ảnh
      • merge fresh_detail vào staging.raw_json + upsert bqms_rfq
    nên đây là con đường drill-1-RFQ AN TOÀN nhất, không phải viết scraper mới.

    Dedupe: nếu đã có batch (1-RFQ) cho staging này đang pending/running thì
    tái sử dụng, KHÔNG enqueue trùng (tránh nhiều user bấm → nhiều login).

    Returns a dict for the endpoint `summary`:
      {status:'drilling', batch_id, batch_item_id, reused:bool, message}
    or {status:'drill_enqueue_failed', message} if enqueue raised.
    """
    # 1. Dedupe / poll — xem batch drill 1-RFQ gần nhất (15 phút) cho staging này.
    #    FE poll bằng cách re-call CHÍNH endpoint /quote (mọi role gọi được),
    #    nên ta phân loại trạng thái job ở đây để trả về cho FE.
    existing = await conn.fetchrow(
        """
        SELECT bi.batch_id, bi.id AS item_id, bi.status, bi.error_message
          FROM bqms_quote_batch_items bi
          JOIN bqms_quote_batches b ON b.id = bi.batch_id
         WHERE bi.staging_id = $1
           AND b.total_count = 1
           AND b.created_at > NOW() - INTERVAL '15 minutes'
         ORDER BY bi.id DESC
         LIMIT 1
        """,
        staging_id,
    )
    if existing and existing["status"] in ("pending", "running"):
        # Job đang chạy → đừng enqueue trùng, bảo FE tiếp tục poll.
        return {
            "status": "drilling",
            "batch_id": existing["batch_id"],
            "batch_item_id": existing["item_id"],
            "reused": True,
            "message": (
                f"Đang tải chi tiết RFQ {rfq_number} từ Samsung… "
                f"nút sẽ tự mở khoá khi xong."
            ),
        }
    if existing and existing["status"] == "error":
        # Job vừa fail (login lỗi / Samsung timeout). KHÔNG enqueue lại tự động
        # để tránh vòng lặp vô hạn — báo rõ cho user, cron vẫn sẽ tự thử ngầm.
        return {
            "status": "drill_failed",
            "batch_id": existing["batch_id"],
            "message": (
                f"Tải chi tiết RFQ {rfq_number} thất bại: "
                f"{(existing['error_message'] or 'lỗi không xác định')[:200]}. "
                f"Cron sẽ tự thử lại ngầm — chờ 3-5 phút rồi bấm lại."
            ),
        }
    if existing and existing["status"] == "done":
        # Job ĐÃ drill xong nhưng _detail.items vẫn rỗng → RFQ này thực sự không
        # có item để báo giá (đã Closed / hết hạn / Samsung trả grid rỗng).
        # KHÔNG enqueue lại (đã thử 1 lần) — báo rõ cho user.
        return {
            "status": "drill_empty",
            "batch_id": existing["batch_id"],
            "message": (
                f"Đã tải chi tiết RFQ {rfq_number} nhưng Samsung không trả về "
                f"dòng linh kiện nào (có thể RFQ đã đóng/hết hạn). "
                f"Kiểm tra lại trên cổng Samsung."
            ),
        }
    # Chưa có batch nào trong 15 phút → enqueue một lượt drill mới bên dưới.

    # 2. Tạo 1-RFQ batch + item (cùng schema create_quote_batch dùng).
    batch_id = await conn.fetchval(
        "INSERT INTO bqms_quote_batches (created_by, total_count, pending_count) "
        "VALUES ($1, 1, 1) RETURNING id",
        user_id,
    )
    item_id = await conn.fetchval(
        "INSERT INTO bqms_quote_batch_items "
        "(batch_id, staging_id, rfq_number, status) "
        "VALUES ($1, $2, $3, 'pending') RETURNING id",
        batch_id, staging_id, rfq_number,
    )

    # 3. Enqueue Procrastinate job (giống create_quote_batch).
    try:
        from app.tasks.bqms_quote_batch import quote_one_rfq_task
        from app.core.procrastinate_app import app as proc_app
        async with proc_app.open_async():
            job_id = await quote_one_rfq_task.defer_async(
                batch_item_id=item_id,
                staging_id=staging_id,
                user_id=str(user_id),
            )
        await conn.execute(
            "UPDATE bqms_quote_batch_items SET procrastinate_job_id=$1 WHERE id=$2",
            job_id, item_id,
        )
    except Exception as exc:
        logger.exception(
            "on-demand drill enqueue failed for staging #%d (RFQ %s)",
            staging_id, rfq_number,
        )
        await conn.execute(
            "UPDATE bqms_quote_batches SET status='error' WHERE id=$1", batch_id,
        )
        return {
            "status": "drill_enqueue_failed",
            "message": (
                f"Không enqueue được job tải chi tiết RFQ {rfq_number}: {exc}. "
                f"Hệ thống có cron tự drill ngầm — chờ 3-5 phút rồi bấm lại."
            ),
        }

    return {
        "status": "drilling",
        "batch_id": batch_id,
        "batch_item_id": item_id,
        "reused": False,
        "message": (
            f"Đang tải chi tiết RFQ {rfq_number} từ Samsung… "
            f"nút sẽ tự mở khoá khi xong (~30-90 giây)."
        ),
    }


@router.post("/vendor-staging/{staging_id}/quote")
async def quote_bidding_staging(
    staging_id: int,
    # Phase H (Thang 2026-05-13): download_files param giữ lại để backward
    # compat URL nhưng KHÔNG còn tác dụng — scrape (30min/5min/3min cron)
    # đã làm hết drill+download+extract. Click "Báo giá" chỉ MỞ KHÓA V1-V4.
    download_files: bool = Query(False, description="DEPRECATED — scrape làm hết, button chỉ unlock V1-V4"),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """LIGHTWEIGHT unlock — Phase H (Thang 2026-05-13).

    New design — Báo giá button KHÔNG còn trigger scrape:
      1. Idempotent UPSERT bqms_rfq từ staging._detail.items (nếu chưa có)
      2. SET bqms_rfq.quote_unlocked=true → frontend bật L1-L4 buttons
      3. SET bqms_rfq.assigned_to = current_user.id
      4. Mark staging.status='approved' + reviewed_at

    KHÔNG còn drill detail / download files / extract images here. Tất cả
    việc nặng đó nằm trong 3 cron task:
      - bqms_periodic_scrape (30 min) — bulk list + drill new RFQs
      - bqms_smart_rescan (5 min) — drill RFQs missing _detail.items
      - bqms_smart_code_track (3 min) — self-heal 10 kinds of gaps

    ON-DEMAND DRILL (Thang 2026-06-23 — fix nút Báo giá "không hoạt động"):
    Nếu raw_json._detail.items vẫn RỖNG (cron chưa kịp drill RFQ này), thay vì
    trả warning thụ động bắt user ngồi chờ + bấm lại, endpoint sẽ ENQUEUE NGAY
    một Procrastinate job `quote_one_rfq_task` (qua hạ tầng quote-batch có sẵn:
    1 batch / 1 RFQ) để drill riêng RFQ này. Task đó:
      - acquire `samsung_session_lock` (serialize với push/scrape — KHÔNG mở
        session Samsung song song, an toàn) → login → drill detail + download
        + extract ảnh → merge fresh_detail vào staging.raw_json → upsert bqms_rfq.
    Endpoint trả `status='drilling'` + `batch_id`. Frontend poll
    GET /vendor-staging/quote-batch/{batch_id}; khi item 'done' thì tự gọi lại
    POST .../quote — lúc này _detail.items đã có → đi nhánh unlock bình thường.
    Dedupe: nếu đã có batch drill đang chạy cho staging này thì tái sử dụng,
    không enqueue trùng. Endpoint vẫn hoàn tất <1 giây — KHÔNG có Playwright
    session chạy trong request (job nặng nằm trên sc-worker).
    """
    row = await conn.fetchrow(
        "SELECT id, module, rfq_number, status, raw_json "
        "FROM bqms_vendor_portal_staging WHERE id = $1",
        staging_id,
    )
    if not row:
        raise HTTPException(404, f"Staging row #{staging_id} not found")
    if row["module"] != "bidding":
        raise HTTPException(400, f"module='{row['module']}' — quote chỉ áp dụng cho 'bidding'")
    if row["status"] not in ("pending_review", "approved"):
        raise HTTPException(409, f"Row đang ở status='{row['status']}', không thể báo lại")

    rfq_number = (row["rfq_number"] or "").strip()
    if not rfq_number:
        raise HTTPException(400, f"Row #{staging_id} thiếu rfq_number")

    raw = row["raw_json"]
    if not isinstance(raw, dict):
        raw = _json.loads(raw or "{}")

    items_in_detail = (raw.get("_detail") or {}).get("items") or []
    summary: dict[str, Any] = {
        "items_in_detail": len(items_in_detail),
        "rfq_number": rfq_number,  # Thang 2026-05-23: trả rfq_number để FE có thể filter+highlight
    }

    if not items_in_detail:
        # SMART UNLOCK (Thang 2026-06-23) — _detail.items rỗng KHÔNG có nghĩa là
        # chưa drill: scrape sau (etl/onedrive_sync) thường GHI ĐÈ staging.raw_json
        # về list-level (xoá _detail), nhưng các row bqms_rfq đã upsert trước đó
        # (drill cũ / cron / lần Báo giá trước) VẪN CÒN. Nếu đã có row → chỉ cần
        # UNLOCK ngay (tức thì), KHÔNG drill lại 30-90s. Đây chính là điểm làm nút
        # Báo giá "xoay mãi": có sẵn row mà vẫn đi drill.
        existing_rows = await conn.fetchval(
            "SELECT COUNT(*) FROM bqms_rfq WHERE rfq_number = $1", rfq_number,
        )
        if existing_rows and int(existing_rows) > 0:
            unlocked = await conn.fetchval(
                """
                UPDATE bqms_rfq
                SET quote_unlocked = true, assigned_to = $1::uuid, updated_at = NOW()
                WHERE rfq_number = $2
                RETURNING (SELECT COUNT(*) FROM bqms_rfq WHERE rfq_number = $2)
                """,
                token_data.user_id, rfq_number,
            )
            await conn.execute(
                "UPDATE bqms_vendor_portal_staging "
                "SET status='approved', reviewed_by=$1, reviewed_at=NOW() "
                "WHERE id=$2 AND status IN ('pending_review','approved')",
                token_data.user_id, staging_id,
            )
            summary["bqms_rfq_upserts"] = 0
            summary["quote_unlocked"] = int(unlocked or 0)
            summary["staging_status"] = "approved"
            logger.info(
                "quote staging #%d (RFQ %s): _detail rỗng nhưng đã có %s bqms_rfq "
                "row → UNLOCK trực tiếp (no drill) unlocked=%d",
                staging_id, rfq_number, existing_rows, summary["quote_unlocked"],
            )
            return {"data": summary}

        # Chưa có row bqms_rfq nào → ON-DEMAND DRILL (Thang 2026-06-23). RFQ thật
        # sự chưa được drill lần nào. Enqueue NGAY 1 job drill riêng RFQ này qua
        # hạ tầng quote-batch có sẵn; FE poll rồi tự unlock khi xong. KHÔNG mở
        # Playwright trong request (job nặng trên sc-worker).
        drill = await _ensure_ondemand_drill(conn, staging_id, rfq_number, token_data.user_id)
        summary.update(drill)
        summary["bqms_rfq_upserts"] = 0
        summary["quote_unlocked"] = 0
        summary["staging_status"] = "pending_review"
        logger.info(
            "quote staging #%d (RFQ %s): empty _detail.items — on-demand drill %s (batch=%s)",
            staging_id, rfq_number, drill.get("status"), drill.get("batch_id"),
        )
        return {"data": summary}

    # 1. Idempotent UPSERT bqms_rfq (nếu cron đã làm thì no-op)
    # Thang 2026-06-22 (fix nút Báo giá chập chờn): KHÔNG còn tạo asyncpg pool
    # MỚI mỗi request (apg.create_pool). Đó là nguyên nhân fail ngẫu nhiên khi
    # nhiều user bấm cùng lúc / scraper đang chạy → tổng connection vượt
    # max_connections của Postgres → TooManyConnectionsError/timeout. Giờ dùng
    # GLOBAL pool (đã init sẵn, size 20) cho upsert + connection request-scoped
    # đã inject (Depends(get_db)) cho 2 UPDATE → không mở connection mới.
    from app.etl.bqms_bidding_scraper import upsert_bqms_rfq_for_one_staging_row
    from app.core.database import db_pool as _global_db_pool

    n = await upsert_bqms_rfq_for_one_staging_row(_global_db_pool.pool(), raw)
    summary["bqms_rfq_upserts"] = n

    # 2 + 3. UNLOCK V1-V4 buttons + assign current user (idempotent) — chạy
    # trên request-scoped conn, KHÔNG mở connection mới.
    # Thang 2026-05-15: assigned_to GHI ĐÈ (không COALESCE) — user nào nhấn
    # "Báo giá" thì cột Người PT cập nhật về user đó.
    unlocked = await conn.fetchval(
        """
        UPDATE bqms_rfq
        SET quote_unlocked = true,
            assigned_to = $1::uuid,
            updated_at = NOW()
        WHERE rfq_number = $2
        RETURNING (SELECT COUNT(*) FROM bqms_rfq WHERE rfq_number = $2)
        """,
        token_data.user_id, rfq_number,
    )
    summary["quote_unlocked"] = int(unlocked or 0)

    # 4. Mark staging approved
    await conn.execute(
        """
        UPDATE bqms_vendor_portal_staging
        SET status='approved', reviewed_by=$1, reviewed_at=NOW()
        WHERE id=$2 AND status IN ('pending_review','approved')
        """,
        token_data.user_id, staging_id,
    )
    summary["staging_status"] = "approved"

    logger.info(
        "quote staging #%d (RFQ %s) UNLOCKED by user=%s: upserts=%d unlocked=%d",
        staging_id, rfq_number, token_data.user_id,
        summary["bqms_rfq_upserts"], summary["quote_unlocked"],
    )
    return {"data": summary}


@router.post("/vendor-staging/{staging_id}/skip")
async def skip_bidding_staging(
    staging_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Mark a bidding staging row as "Skip — review sau" — đặt status='skipped'.
    Khác với reject: skip có thể đổi ý sau."""
    row = await conn.fetchrow(
        "UPDATE bqms_vendor_portal_staging "
        "SET status='skipped', reviewed_by=$1, reviewed_at=NOW() "
        "WHERE id = $2 AND status='pending_review' "
        "RETURNING id, status",
        token_data.user_id, staging_id,
    )
    if not row:
        raise HTTPException(409, f"Staging #{staging_id} không tồn tại hoặc đã được duyệt trước đó")
    return {"data": dict(row)}


# ---------------------------------------------------------------------------
# Option B — background queue: batch /quote via Procrastinate worker
# ---------------------------------------------------------------------------

from pydantic import BaseModel, Field as PydField


class QuoteBatchRequest(BaseModel):
    staging_ids: list[int] = PydField(..., min_length=1, max_length=200)


@router.post("/vendor-staging/quote-batch")
@limiter.limit("10/minute")
async def create_quote_batch(
    request: Request,
    response: Response,
    payload: QuoteBatchRequest,
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Tạo 1 batch /quote chạy nền — enqueue N task vào Procrastinate (queue='bqms_quote').
    Mỗi task = 1 RFQ với own login, chạy tuần tự trên sc-worker (~30-90s/RFQ).

    Frontend poll GET /vendor-staging/quote-batch/{id} để xem progress.
    Trả về batch_id ngay (không block)."""
    ids = list({sid for sid in payload.staging_ids if sid > 0})
    if not ids:
        raise HTTPException(400, "staging_ids rỗng")

    # Validate: chỉ accept module='bidding' rows ở status='pending_review' or 'approved'
    rows = await conn.fetch(
        "SELECT id, rfq_number, module, status FROM bqms_vendor_portal_staging "
        "WHERE id = ANY($1::int[])",
        ids,
    )
    found_ids = {r["id"] for r in rows}
    missing = [i for i in ids if i not in found_ids]
    if missing:
        raise HTTPException(404, f"Không tìm thấy staging rows: {missing}")
    bad = [r["id"] for r in rows if r["module"] != "bidding"
           or r["status"] not in ("pending_review", "approved")]
    if bad:
        raise HTTPException(400, f"Rows không phải bidding/pending_review: {bad}")

    # Tạo batch row
    batch_id = await conn.fetchval(
        "INSERT INTO bqms_quote_batches (created_by, total_count, pending_count) "
        "VALUES ($1, $2, $2) RETURNING id",
        token_data.user_id, len(rows),
    )

    # Tạo per-row item rows
    item_id_by_staging: dict[int, int] = {}
    for r in rows:
        item_id = await conn.fetchval(
            "INSERT INTO bqms_quote_batch_items "
            "(batch_id, staging_id, rfq_number, status) "
            "VALUES ($1, $2, $3, 'pending') RETURNING id",
            batch_id, r["id"], r["rfq_number"],
        )
        item_id_by_staging[r["id"]] = item_id

    # Enqueue Procrastinate tasks. The app is not opened by FastAPI lifespan,
    # so we open it temporarily for the duration of the enqueue burst.
    try:
        from app.tasks.bqms_quote_batch import quote_one_rfq_task
        from app.core.procrastinate_app import app as proc_app
        async with proc_app.open_async():
            for r in rows:
                item_id = item_id_by_staging[r["id"]]
                job_id = await quote_one_rfq_task.defer_async(
                    batch_item_id=item_id,
                    staging_id=r["id"],
                    user_id=str(token_data.user_id),
                )
                await conn.execute(
                    "UPDATE bqms_quote_batch_items SET procrastinate_job_id=$1 WHERE id=$2",
                    job_id, item_id,
                )
    except Exception as exc:
        logger.exception("quote-batch: failed to enqueue tasks for batch #%d", batch_id)
        await conn.execute(
            "UPDATE bqms_quote_batches SET status='error' WHERE id=$1",
            batch_id,
        )
        raise HTTPException(500, f"Enqueue lỗi: {exc}")

    logger.info(
        "quote-batch #%d created by user=%s — %d RFQs enqueued to bqms_quote",
        batch_id, token_data.user_id, len(rows),
    )
    return {"data": {"batch_id": batch_id, "total_count": len(rows)}}


@router.get("/vendor-staging/quote-batch/{batch_id}")
async def get_quote_batch_status(
    batch_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Polled bởi frontend — trả về batch summary + per-row progress."""
    batch = await conn.fetchrow(
        "SELECT id, created_at, completed_at, created_by, total_count, "
        "       pending_count, running_count, done_count, error_count, status "
        "FROM bqms_quote_batches WHERE id = $1",
        batch_id,
    )
    if not batch:
        raise HTTPException(404, f"Batch #{batch_id} không tồn tại")

    items = await conn.fetch(
        "SELECT id, staging_id, rfq_number, status, items_count, files_count, "
        "       images_count, upserts_count, classification, error_message, "
        "       started_at, completed_at "
        "FROM bqms_quote_batch_items WHERE batch_id = $1 ORDER BY id",
        batch_id,
    )
    return {
        "data": {
            "batch": dict(batch),
            "items": [dict(it) for it in items],
        }
    }


@router.get("/vendor-staging/quote-batch")
async def list_recent_quote_batches(
    limit: int = Query(20, ge=1, le=100),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List N batches gần nhất — dùng cho lịch sử."""
    rows = await conn.fetch(
        "SELECT id, created_at, completed_at, created_by, total_count, "
        "       pending_count, running_count, done_count, error_count, status "
        "FROM bqms_quote_batches ORDER BY id DESC LIMIT $1",
        limit,
    )
    return {"data": {"batches": [dict(r) for r in rows]}}


@router.post("/bidding/{staging_id}/download-files")
async def download_files_for_bidding_staging(
    staging_id: int,
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Download all attachments for a single bidding RFQ to local VPS folder
    `Puplic/BQMS/RFQ/RFQ <year>/THANG <month>/<rfq_number>/raw/` and extract
    images from RFQ_*.xlsx into `<folder>/images/`. Per user 2026-05-08:
    local-VPS only, no OneDrive upload.

    The staging row's raw_json supplies the secureKey + form fields needed
    to navigate directly to the RFQ detail in BQMS.
    """
    row = await conn.fetchrow(
        "SELECT id, module, rfq_number, raw_json FROM bqms_vendor_portal_staging WHERE id = $1",
        staging_id,
    )
    if not row:
        raise HTTPException(404, f"Staging row #{staging_id} not found")
    if row["module"] != "bidding":
        raise HTTPException(400, f"Row #{staging_id} module='{row['module']}' — only 'bidding' supported")

    rfq_number = (row["rfq_number"] or "").strip()
    if not rfq_number:
        raise HTTPException(400, f"Row #{staging_id} has no rfq_number")

    raw = row["raw_json"]
    if not isinstance(raw, dict):
        import json as _j
        raw = _j.loads(raw or "{}")

    import threading
    import asyncio as _aio

    result_holder: dict[str, Any] = {}

    def _thread_dl() -> None:
        async def _do() -> None:
            from app.etl.bqms_bidding_scraper import download_files_for_rfq
            from app.core.config import settings as cfg
            import asyncpg as apg

            db_url = (
                str(cfg.DATABASE_URL)
                .replace("+asyncpg", "")
                .replace("postgresql+asyncpg", "postgresql")
            )
            pool = await apg.create_pool(db_url, min_size=1, max_size=2)
            try:
                summary = await download_files_for_rfq(rfq_number, raw, db_pool=pool)
                result_holder["summary"] = summary
            except Exception as exc:
                logger.exception("download_files_for_rfq failed")
                result_holder["error"] = str(exc)[:500]
            finally:
                await pool.close()

        _aio.run(_do())

    t = threading.Thread(target=_thread_dl, daemon=True)
    t.start()
    # Login + nav + many downloads — give it 5 minutes max
    t.join(timeout=300)
    if t.is_alive():
        raise HTTPException(504, "Download timed out after 5 minutes")
    if "error" in result_holder:
        raise HTTPException(500, f"Download failed: {result_holder['error']}")

    summary = result_holder.get("summary") or {}
    logger.info(
        "download-files by user=%s: rfq=%s files=%d images=%d folder_existed=%s",
        token_data.user_id, rfq_number,
        summary.get("downloaded_count"), summary.get("images_extracted"),
        summary.get("folder_pre_existed"),
    )
    return {"data": summary}


@router.get("/vendor-staging")
async def list_vendor_staging(
    status: str | None = Query(None, description="Filter by status: pending_review|approved|rejected|merged"),
    module: str = Query("contract", description="contract | po | bidding"),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List staging rows from bqms_vendor_portal_staging for human review."""
    where = ["module = $1"]
    params: list[Any] = [module]
    if status:
        params.append(status)
        where.append(f"status = ${len(params)}")

    params.extend([limit, offset])
    # Pull useful Bidding-specific fields out of raw_json so the table can show
    # Excel BC BQMS THANG-equivalent columns without a per-row /detail call.
    sql = f"""
        SELECT id, scraped_at, scrape_run_id, module, rfq_number, contract_no,
               contract_period, item_code, description, specification,
               quantity, unit, status, review_notes, reviewed_at, merged_at,
               -- Top-level scraped fields
               raw_json->>'reqName'            AS req_name,
               raw_json->>'regDt'              AS reg_dt,
               raw_json->>'deadlineDt'         AS deadline_dt,
               raw_json->>'submitDt'           AS submit_dt,
               raw_json->>'progressStatusName' AS bd_status,
               raw_json->>'psinchargeName'     AS psincharge_name,
               raw_json->>'criteriaCurrency'   AS currency,
               raw_json->>'itemCnt'            AS item_cnt_text,
               raw_json->>'dday'               AS dday_html,
               raw_json->>'ctrTypeNm'          AS ctr_type_nm,
               -- Drilled detail summary
               raw_json #>> '{{_detail,classification}}'   AS classification,
               raw_json #>> '{{_detail,version}}'          AS detail_version,
               jsonb_array_length(COALESCE(raw_json #> '{{_detail,items}}', '[]'::jsonb))         AS items_count,
               jsonb_array_length(COALESCE(raw_json #> '{{_detail,attachments}}', '[]'::jsonb))   AS attachments_count,
               raw_json #>> '{{_detail,error}}'            AS detail_error,
               raw_json #>> '{{_detail,items,0,maker}}'    AS first_maker,
               raw_json #>> '{{_detail,items,0,part_no}}'  AS first_part_no,
               raw_json #>> '{{_detail,items,0,cis_code}}' AS first_cis_code,
               raw_json #>> '{{_detail,items,0,moq}}'      AS first_moq
        FROM bqms_vendor_portal_staging
        WHERE {' AND '.join(where)}
        ORDER BY scraped_at DESC, id DESC
        LIMIT ${len(params) - 1} OFFSET ${len(params)}
    """
    rows = await conn.fetch(sql, *params)

    count_sql = f"SELECT COUNT(*) FROM bqms_vendor_portal_staging WHERE {' AND '.join(where)}"
    total = await conn.fetchval(count_sql, *params[: len(params) - 2])

    return {
        "data": {
            "items": [dict(r) for r in rows],
            "total": int(total or 0),
            "limit": limit,
            "offset": offset,
        }
    }


@router.get("/vendor-staging/{staging_id}")
async def get_vendor_staging_row(
    staging_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Full raw JSON for a single staging row — used by the review UI."""
    row = await conn.fetchrow(
        "SELECT * FROM bqms_vendor_portal_staging WHERE id = $1",
        staging_id,
    )
    if not row:
        raise HTTPException(404, f"Staging row #{staging_id} không tồn tại")
    return {"data": dict(row)}


class _StagingDecision(BaseModel):
    decision: str  # 'approve' | 'reject'
    notes: str | None = None


@router.post("/vendor-staging/{staging_id}/decide")
async def decide_vendor_staging_row(
    staging_id: int,
    body: _StagingDecision,
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Approve or reject a staging row. Approved rows are merged into
    bqms_won_quotations by a follow-up job (B7 — not wired yet)."""
    if body.decision not in ("approve", "reject"):
        raise HTTPException(400, "decision phải là 'approve' hoặc 'reject'")
    new_status = "approved" if body.decision == "approve" else "rejected"

    row = await conn.fetchrow(
        """
        UPDATE bqms_vendor_portal_staging
        SET status = $1,
            review_notes = $2,
            reviewed_by = $3,
            reviewed_at = NOW()
        WHERE id = $4
          AND status = 'pending_review'
        RETURNING id, status, reviewed_at
        """,
        new_status, body.notes, token_data.user_id, staging_id,
    )
    if not row:
        raise HTTPException(
            409,
            f"Staging row #{staging_id} không tồn tại hoặc đã được duyệt trước đó",
        )
    return {"data": dict(row)}


# ─── Helpers for B7 merge ─────────────────────────────────────────

import json as _json
import re as _re
from datetime import date as _date


def _parse_money(raw: str | None) -> float | None:
    """Parse '5,000,000 VND' or '139,000' -> 5000000.0 / 139000.0."""
    if not raw:
        return None
    s = _re.sub(r"[^\d.\-]", "", str(raw).replace(",", ""))
    try:
        return float(s) if s else None
    except ValueError:
        return None


def _parse_period_end(raw: str | None) -> _date | None:
    """Parse 'Contract Period' like '5/7/2026 ~ 8/7/2026' -> date(2026,8,7).

    BQMS format is D/M/YYYY (per Phan Van Thao sample). End date is the
    second segment after '~'. Falls back to None on any parse error.
    """
    if not raw:
        return None
    parts = [p.strip() for p in str(raw).split("~")]
    target = parts[-1] if len(parts) > 1 else parts[0]
    m = _re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", target)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return _date(y, mo, d)
    except ValueError:
        return None


def _parse_yyyymmdd(raw: str | None) -> _date | None:
    """Parse '20260807' -> date(2026,8,7) — used by MRO REQ_DELIVERY_DATE."""
    if not raw:
        return None
    s = "".join(ch for ch in str(raw) if ch.isdigit())
    if len(s) != 8:
        return None
    try:
        return _date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    except ValueError:
        return None


def _po_row_to_delivery_fields(staging_row, raw: dict) -> dict:
    """Map a module='po' staging row + its raw_json into bqms_deliveries
    column values (NOT bqms_won_quotations — corrected 2026-05-09).

    MRO P/O Receipt → Giao hàng function. raw_json is the per-row MRO grid
    record (PO_NO, ITEM_CODE, BUYING_PRICE, REQ_DELIVERY_DATE, …).
    """
    # PO_CONFIRM_DT is epoch milliseconds (string) — convert to date
    po_dt = None
    raw_dt = raw.get("PO_CONFIRM_DT")
    if raw_dt:
        try:
            po_dt = _date.fromtimestamp(int(raw_dt) / 1000)
        except (TypeError, ValueError):
            po_dt = None
    return {
        "po_number": (raw.get("PO_NO") or "").strip() or None,
        "po_date": po_dt,
        "quotation_no": (staging_row["rfq_number"] or "").strip() or None,
        "bqms_code": (staging_row["item_code"] or "").strip() or None,
        "specification": staging_row["specification"] or raw.get("SPECIFICATION") or None,
        "quantity": staging_row["quantity"],
        "unit": staging_row["unit"] or "EA",
        "unit_price": _parse_money(raw.get("BUYING_PRICE")),
        "amount": _parse_money(raw.get("BUYING_AMOUNT")),
        "recipient_name": raw.get("RECEIVER_NAME") or None,
        "receiving_warehouse": raw.get("DELIVERY_ADDRESS") or None,
        "delivery_date": _parse_yyyymmdd(raw.get("REQ_DELIVERY_DATE")),
    }


def _contract_row_to_won_fields(staging_row, raw: dict) -> dict:
    """Map a module='contract' staging row + its raw_json (which contains
    {contract:{basic_info,...}, item:{...}}) into bqms_won_quotations columns."""
    contract = (raw.get("contract") or {}) if isinstance(raw, dict) else {}
    item = (raw.get("item") or {}) if isinstance(raw, dict) else {}
    basic = contract.get("basic_info") or {}
    return {
        "rfq_number": (staging_row["rfq_number"] or "").strip() or None,
        "bqms_code": (staging_row["item_code"] or "").strip() or None,
        "description": staging_row["description"],
        "specification": staging_row["specification"],
        "quantity": staging_row["quantity"],
        "unit": staging_row["unit"] or "EA",
        "po_price": _parse_money(item.get("unit_price")),
        "po_deadline": _parse_period_end(basic.get("Contract Period")),
        "supplier_name": basic.get("Vendor name") or None,
    }


@router.get("/coverage/excel-vs-portal")
async def excel_vs_portal_coverage(
    year: int | None = Query(None, description="Filter by inquiry year (default: all)"),
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Phase J — Coverage report for "đã có thể bỏ Excel chưa".

    Compares bqms_rfq rows by data_source:
      - excel_import / onedrive_sync = legacy Excel chain
      - etl = Vendor Portal direct scrape
      - manual = nhập tay

    Returns counts + overlap analysis (RFQ that exist in BOTH chains, only Excel,
    only Portal). Used by /admin/vendor-staging "Coverage" panel.
    """
    where = "1=1"
    params: list[Any] = []
    if year:
        where = "EXTRACT(YEAR FROM COALESCE(inquiry_date, created_at::date)) = $1"
        params = [year]

    rows = await conn.fetch(
        f"""
        SELECT data_source, COUNT(*) AS n,
               COUNT(DISTINCT rfq_number) AS distinct_rfq
        FROM bqms_rfq
        WHERE {where}
        GROUP BY data_source
        ORDER BY data_source
        """,
        *params,
    )
    by_source = {r["data_source"] or "unknown": {"items": int(r["n"]), "rfq": int(r["distinct_rfq"])} for r in rows}

    # Overlap: RFQ numbers seen in BOTH legacy excel and etl
    overlap_row = await conn.fetchrow(
        f"""
        WITH legacy AS (
          SELECT DISTINCT rfq_number FROM bqms_rfq
          WHERE {where}
            AND data_source IN ('excel_import','onedrive_sync')
            AND rfq_number IS NOT NULL
        ),
        portal AS (
          SELECT DISTINCT rfq_number FROM bqms_rfq
          WHERE {where}
            AND data_source = 'etl'
            AND rfq_number IS NOT NULL
        )
        SELECT
          (SELECT COUNT(*) FROM legacy) AS legacy_only_or_overlap,
          (SELECT COUNT(*) FROM portal) AS portal_only_or_overlap,
          (SELECT COUNT(*) FROM legacy l JOIN portal p USING(rfq_number)) AS overlap,
          (SELECT COUNT(*) FROM legacy l WHERE NOT EXISTS (SELECT 1 FROM portal p WHERE p.rfq_number = l.rfq_number)) AS only_legacy,
          (SELECT COUNT(*) FROM portal p WHERE NOT EXISTS (SELECT 1 FROM legacy l WHERE l.rfq_number = p.rfq_number)) AS only_portal
        """,
        *params,
    )

    legacy_total = int(overlap_row["legacy_only_or_overlap"] or 0)
    portal_total = int(overlap_row["portal_only_or_overlap"] or 0)
    overlap = int(overlap_row["overlap"] or 0)
    only_legacy = int(overlap_row["only_legacy"] or 0)
    only_portal = int(overlap_row["only_portal"] or 0)

    # Coverage % = (overlap + only_portal) / (legacy + only_portal) — how much
    # of the legacy chain we've already covered + new items the portal caught
    union = legacy_total + only_portal
    coverage_pct = round(100.0 * (overlap + only_portal) / union, 1) if union > 0 else 0.0

    return {
        "data": {
            "year": year,
            "by_source": by_source,
            "rfq_overlap": {
                "legacy_total_distinct": legacy_total,
                "portal_total_distinct": portal_total,
                "overlap": overlap,
                "only_legacy": only_legacy,
                "only_portal": only_portal,
            },
            "coverage_pct": coverage_pct,
            "ready_to_deprecate_excel": coverage_pct >= 95.0 and only_legacy < 10,
        }
    }


@router.post("/vendor-staging/merge-approved")
async def merge_approved_vendor_staging(
    module: str = Query("contract", description="contract | po"),
    token_data: TokenData = Depends(require_role("admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Merge all staging rows with status='approved' into the function's
    target table — module-specific dispatch:

      module='contract' (Contract Mgmt)  → bqms_won_quotations  (Trúng BG)
      module='po'       (MRO P/O Receipt)→ bqms_deliveries      (Giao hàng)

    For each approved-but-not-yet-merged row:
      - INSERT into target table
      - UPDATE staging.status='merged', staging.merged_at=NOW()
      - Skip if natural key already exists (re-scrape protection)
    """
    if module not in ("contract", "po"):
        raise HTTPException(400, "module phải là 'contract' hoặc 'po'")

    rows = await conn.fetch(
        """
        SELECT id, rfq_number, item_code, description, specification,
               quantity, unit, raw_json
        FROM bqms_vendor_portal_staging
        WHERE module = $1
          AND status = 'approved'
          AND merged_at IS NULL
        ORDER BY id
        FOR UPDATE
        """,
        module,
    )
    target_table = "bqms_won_quotations" if module == "contract" else "bqms_deliveries"

    if not rows:
        return {
            "data": {
                "module": module,
                "target_table": target_table,
                "merged": 0,
                "skipped_duplicate": 0,
                "errors": [],
            }
        }
    merged = 0
    skipped = 0
    errors: list[dict] = []

    async with conn.transaction():
        for r in rows:
            try:
                raw = r["raw_json"] if isinstance(r["raw_json"], dict) else _json.loads(r["raw_json"] or "{}")
                if not isinstance(raw, dict):
                    raw = {}

                if module == "contract":
                    fields = _contract_row_to_won_fields(r, raw)
                    rfq_number = fields["rfq_number"]
                    bqms_code = fields["bqms_code"]
                    # Dedupe: same (rfq_number, bqms_code) already in won_quotations
                    if bqms_code and rfq_number:
                        exists = await conn.fetchval(
                            "SELECT 1 FROM bqms_won_quotations "
                            "WHERE rfq_number = $1 AND bqms_code = $2 LIMIT 1",
                            rfq_number, bqms_code,
                        )
                        if exists:
                            skipped += 1
                            await conn.execute(
                                "UPDATE bqms_vendor_portal_staging "
                                "SET status='merged', merged_at=NOW(), "
                                "review_notes = COALESCE(review_notes, '') || $1 "
                                "WHERE id = $2",
                                f" [skip: already in {target_table}]", r["id"],
                            )
                            continue
                    await conn.execute(
                        """
                        INSERT INTO bqms_won_quotations
                            (rfq_number, bqms_code, description, specification,
                             quantity, unit, po_price, po_deadline, supplier_name)
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                        """,
                        fields["rfq_number"], fields["bqms_code"],
                        fields["description"], fields["specification"],
                        fields["quantity"], fields["unit"],
                        fields["po_price"], fields["po_deadline"],
                        fields["supplier_name"],
                    )
                else:  # module == "po" → bqms_deliveries
                    fields = _po_row_to_delivery_fields(r, raw)
                    po_number = fields["po_number"]
                    bqms_code = fields["bqms_code"]
                    # Dedupe: same (po_number, bqms_code) already in deliveries
                    if po_number and bqms_code:
                        exists = await conn.fetchval(
                            "SELECT 1 FROM bqms_deliveries "
                            "WHERE po_number = $1 AND bqms_code = $2 LIMIT 1",
                            po_number, bqms_code,
                        )
                        if exists:
                            skipped += 1
                            await conn.execute(
                                "UPDATE bqms_vendor_portal_staging "
                                "SET status='merged', merged_at=NOW(), "
                                "review_notes = COALESCE(review_notes, '') || $1 "
                                "WHERE id = $2",
                                f" [skip: already in {target_table}]", r["id"],
                            )
                            continue
                    await conn.execute(
                        """
                        INSERT INTO bqms_deliveries
                            (po_number, po_date, quotation_no, bqms_code,
                             specification, quantity, unit, unit_price, amount,
                             recipient_name, receiving_warehouse, delivery_date,
                             delivery_status, data_source)
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9,
                                $10, $11, $12, 'chua_giao', 'etl')
                        """,
                        fields["po_number"], fields["po_date"],
                        fields["quotation_no"], fields["bqms_code"],
                        fields["specification"], fields["quantity"],
                        fields["unit"], fields["unit_price"], fields["amount"],
                        fields["recipient_name"], fields["receiving_warehouse"],
                        fields["delivery_date"],
                    )

                await conn.execute(
                    "UPDATE bqms_vendor_portal_staging "
                    "SET status='merged', merged_at=NOW() WHERE id = $1",
                    r["id"],
                )
                merged += 1
            except Exception as exc:
                logger.exception("merge failed for staging row %d (module=%s)", r["id"], module)
                errors.append({"id": r["id"], "error": str(exc)[:300]})

    logger.info(
        "merge-approved by user=%s: module=%s target=%s merged=%d skipped=%d errors=%d",
        token_data.user_id, module, target_table, merged, skipped, len(errors),
    )
    return {
        "data": {
            "module": module,
            "target_table": target_table,
            "merged": merged,
            "skipped_duplicate": skipped,
            "errors": errors,
        }
    }


# ---------------------------------------------------------------------------
# Phase 4.2 (Thang 2026-05-12 audit follow-up):
# Contract endpoints — list / detail / trigger merger.
# ---------------------------------------------------------------------------

@router.post("/contracts/merge")
async def trigger_contracts_merge(
    token_data: TokenData = Depends(require_role("admin", "manager")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Merge tất cả contract staging vào bqms_contracts. Idempotent."""
    from app.services.bqms_contract_merger import merge_contracts
    stats = await merge_contracts(conn)
    return {"data": stats, "message": "Đã merge contracts"}


@router.get("/contracts")
async def list_contracts(
    search: str | None = Query(None),
    status: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=10, le=200),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales",
        "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List hợp đồng đã ký với Samsung."""
    conds = ["1=1"]
    params: list[Any] = []
    idx = 1
    if search:
        like = f"%{search.strip()}%"
        conds.append(
            f"(contract_no ILIKE ${idx} OR request_no ILIKE ${idx} "
            f"OR vendor_name ILIKE ${idx} OR subject ILIKE ${idx})"
        )
        params.append(like)
        idx += 1
    if status:
        conds.append(f"status = ${idx}")
        params.append(status)
        idx += 1
    where = " AND ".join(conds)
    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM bqms_contracts WHERE {where}", *params,
    )
    params.extend([page_size, (page - 1) * page_size])
    rows = await conn.fetch(
        f"""
        SELECT c.id, c.contract_no, c.request_no, c.contract_kind,
               c.contract_type, c.subject, c.status, c.amount, c.currency,
               c.contract_period, c.contract_start, c.contract_end,
               c.vendor_name, c.created_by_samsung, c.reconciliation,
               c.won_quotation_id, c.rfq_id, c.synced_at, c.created_at,
               (SELECT COUNT(*) FROM bqms_contract_items ci WHERE ci.contract_id = c.id) AS item_count
        FROM bqms_contracts c
        WHERE {where}
        ORDER BY c.contract_start DESC NULLS LAST, c.id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )
    def _ser(r):
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                try: d[k] = float(v)
                except Exception: pass
        return d
    return {
        "data": {
            "items": [_ser(r) for r in rows],
            "total": int(total or 0),
            "page": page,
            "page_size": page_size,
        }
    }


@router.get("/contracts/{contract_id}")
async def contract_detail(
    contract_id: int,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales",
        "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    row = await conn.fetchrow(
        "SELECT * FROM bqms_contracts WHERE id = $1", contract_id,
    )
    if not row:
        raise HTTPException(404, "Contract không tồn tại")
    items = await conn.fetch(
        "SELECT * FROM bqms_contract_items WHERE contract_id = $1 ORDER BY item_no",
        contract_id,
    )
    def _ser(r):
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)): d[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                try: d[k] = float(v)
                except Exception: pass
        return d
    return {"data": {"contract": _ser(row), "items": [_ser(it) for it in items]}}


# PR-2 (Thang 2026-05-13): /quote-image-override POST/DELETE/check
# extracted to bqms_images.py — mounted with same /bqms prefix.


# ---------------------------------------------------------------------------
# Phase 4.3 (Thang 2026-05-12): MRO PO list (Samsung POs from scraper).
# Source: bqms_samsung_po. Show with link to bqms_deliveries.
# ---------------------------------------------------------------------------

@router.get("/mro/po")
async def list_mro_po(
    search: str | None = Query(None),
    process_status: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=10, le=200),
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales",
        "procurement", "warehouse", "accountant",
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """List Samsung PO (MRO P/O Receipt)."""
    conds = ["1=1"]
    params: list[Any] = []
    idx = 1
    if search:
        like = f"%{search.strip()}%"
        conds.append(
            f"(p.po_number ILIKE ${idx} OR p.bqms_code ILIKE ${idx} "
            f"OR p.specification ILIKE ${idx} OR p.maker ILIKE ${idx})"
        )
        params.append(like); idx += 1
    if process_status:
        conds.append(f"p.process_status::text = ${idx}")
        params.append(process_status); idx += 1
    where = " AND ".join(conds)
    total = await conn.fetchval(
        f"SELECT COUNT(*) FROM bqms_samsung_po p WHERE {where}", *params,
    )
    params.extend([page_size, (page - 1) * page_size])
    rows = await conn.fetch(
        f"""
        SELECT p.id, p.po_number, p.po_date, p.bqms_code, p.specification,
               p.maker, p.order_qty, p.unit_price, p.amount, p.currency::text AS currency,
               p.preferred_delivery_date, p.process_status::text AS process_status,
               p.vendor_code, p.buyer_name, p.company, p.plant,
               p.shipping_qty, p.gr_qty, p.invoice_qty,
               (SELECT COUNT(*) FROM bqms_deliveries d WHERE d.samsung_po_id = p.id) AS delivery_count
        FROM bqms_samsung_po p
        WHERE {where}
        ORDER BY p.po_date DESC NULLS LAST, p.id DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *params,
    )
    def _ser(r):
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)): d[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                try: d[k] = float(v)
                except Exception: pass
        return d
    return {
        "data": {
            "items": [_ser(r) for r in rows],
            "total": int(total or 0),
            "page": page,
            "page_size": page_size,
        }
    }


# ═══════════════════════════════════════════════════════════════════
# BQMS Auto-Submit (Thang 2026-05-14) — đẩy báo giá lên sec-bqms qua Playwright
# ═══════════════════════════════════════════════════════════════════

@router.get("/rfq/{rfq_id}/push-preview")
async def get_push_preview(
    rfq_id: int,
    round_n: int = Query(1, ge=1, le=4),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Build payload đề xuất cho Push to SEC modal."""
    from app.services.bqms_image_resizer import resize_for_samsung, resolve_image_for_bqms_code
    from datetime import date, timedelta

    rfq = await conn.fetchrow(
        """SELECT id, rfq_number, classification_override, notes, maker
           FROM bqms_rfq WHERE id = $1""", rfq_id,
    )
    if not rfq:
        raise HTTPException(404, f"RFQ #{rfq_id} không tồn tại")

    rfq_number = rfq["rfq_number"]
    classification = (rfq["classification_override"] or "").upper().strip()
    if not classification:
        notes = (rfq["notes"] or "").lower()
        if "classification=gc" in notes:
            classification = "GC"
        elif "classification=tm" in notes:
            classification = "TM"
        else:
            classification = "TM"

    # Thang 2026-06-13 (FIX V2 push fail):
    # Dedup at query time SAME WAY rfq-table does. Without this, duplicate
    # (rfq_number, bqms_code) twins (etl + onedrive_sync) would both come
    # back here; if the twin with NULL V2 ranks first, push-preview reads
    # price_v=NULL → throws "chưa có giá V2" warning + 0-price push.
    # DISTINCT ON ordered by V-presence DESC ensures the row carrying real
    # V{round_n} price wins.
    items_db = await conn.fetch(
        """SELECT DISTINCT ON (rfq_number, bqms_code)
                  id, rfq_number, bqms_code, specification,
                  expected_qty, unit, maker,
                  quoted_price_bqms_v1, quoted_price_bqms_v2,
                  quoted_price_bqms_v3, quoted_price_bqms_v4
           FROM bqms_rfq WHERE rfq_number = $1
           ORDER BY rfq_number, bqms_code,
                    (COALESCE(quote_unlocked, false))::int DESC,
                    (quoted_price_bqms_v4 IS NOT NULL)::int DESC,
                    (quoted_price_bqms_v3 IS NOT NULL)::int DESC,
                    (quoted_price_bqms_v2 IS NOT NULL)::int DESC,
                    (quoted_price_bqms_v1 IS NOT NULL)::int DESC,
                    updated_at DESC NULLS LAST,
                    id DESC""",
        rfq_number,
    )
    if not items_db:
        raise HTTPException(400, "Không có item nào")

    # Fetch REAL description per bqms_code from staging.raw_json._detail.items.
    # bqms_rfq table không có cột description; description thật ("CNC BRUSH",
    # "BLADE") chỉ tồn tại trong staging từ lúc scrape items. Map by bqms_code.
    # Per Thang 2026-05-15: Submission Opinion phải dùng description, không
    # phải specification.
    description_map: dict[str, str] = {}
    try:
        staging_row = await conn.fetchrow(
            """SELECT raw_json FROM bqms_vendor_portal_staging
               WHERE module='bidding' AND rfq_number=$1
               ORDER BY id DESC LIMIT 1""", rfq_number,
        )
        if staging_row and staging_row["raw_json"]:
            raw = staging_row["raw_json"]
            if isinstance(raw, str):
                raw = _json.loads(raw)
            detail = (raw.get("_detail") or {})
            for det_it in (detail.get("items") or []):
                c = (det_it.get("item_code") or "").strip()
                d = (det_it.get("description") or "").strip()
                if c and d:
                    description_map[c] = d
    except Exception as exc:
        logger.warning("push-preview description_map fetch failed for %s: %s",
                       rfq_number, str(exc)[:120])

    # Look up user-pinned primary images in one query (Thang 2026-05-20):
    # picker modal stores choices in bqms_code_primary_image — push-preview must
    # respect that so chosen ảnh thực sự được dùng khi đẩy lên Samsung.
    primary_map: dict[str, str] = {}
    codes_for_primary = [it["bqms_code"] for it in items_db if it.get("bqms_code")]
    if codes_for_primary and round_n == 1:
        try:
            primary_rows = await conn.fetch(
                "SELECT bqms_code, image_path FROM bqms_code_primary_image "
                "WHERE bqms_code = ANY($1::text[])",
                codes_for_primary,
            )
            primary_map = {r["bqms_code"]: r["image_path"] for r in primary_rows}
        except Exception as exc:
            logger.warning("push-preview primary-image lookup failed: %s", exc)

    items_out: list[dict] = []
    warnings: list[str] = []
    for it in items_db:
        code = it["bqms_code"]
        if not code:
            continue
        # Round 1 needs image; round 2+ uses Samsung's stored image (don't re-upload)
        img_src = None
        image_path = None
        image_source = "missing"
        if round_n == 1:
            # PRIORITY 0 — user pinned via picker modal (DB).
            pinned = primary_map.get(code)
            if pinned:
                p = Path(pinned)
                if p.exists():
                    img_src = p
                else:
                    logger.warning("push-preview: pinned primary %s for %s gone",
                                   pinned, code)
            # Fallback to filesystem auto-pick if no pinned or pinned file vanished.
            if img_src is None:
                img_src = resolve_image_for_bqms_code(code, rfq_number)

            if img_src:
                try:
                    resized = resize_for_samsung(img_src)
                    image_path = str(resized)
                    # Mark "override" when user explicitly pinned OR file is under
                    # /quote-overrides OR /bqms-image-uploads OR .user-image-uploads.
                    src_str = str(img_src)
                    is_user = (
                        bool(pinned)
                        or "quote-overrides" in src_str
                        or ".user-image-uploads" in src_str
                        or "bqms-image-uploads" in src_str
                    )
                    image_source = "override" if is_user else "auto"
                except Exception as exc:
                    warnings.append(f"{code}: resize ảnh lỗi - {exc}")
            else:
                warnings.append(f"{code}: không tìm thấy ảnh")
        else:
            # Round 2+: no image upload required, Samsung keeps the V1 image.
            image_source = "skip_round_gt_1"

        price_v = it[f"quoted_price_bqms_v{round_n}"]
        if price_v is None and round_n > 1:
            for r in range(round_n - 1, 0, -1):
                price_v = it[f"quoted_price_bqms_v{r}"]
                if price_v is not None:
                    break
        if price_v is None:
            warnings.append(f"{code}: chưa có giá V{round_n}")

        # Real description from staging (e.g., "CNC BRUSH"), fallback to spec
        real_desc = description_map.get(code) or it["specification"] or code

        items_out.append({
            "rfq_item_id": int(it["id"]),
            "bqms_code": code,
            "description": real_desc[:200],
            "specification": it["specification"],
            "quantity": float(it["expected_qty"] or 0),
            "unit": it["unit"] or "Piece",
            "maker": it["maker"] or "",
            "image_path": image_path,
            "image_source": image_source,
            "quotation_price": float(price_v) if price_v is not None else 0,
            "abandonment": "N",
            "lead_time_days": 30,
        })

    opinion = ", ".join(i["description"] for i in items_out if i["abandonment"] == "N")

    today = date.today()
    try:
        valid = today.replace(month=today.month + 3) if today.month <= 9 else date(today.year + 1, ((today.month + 3) - 1) % 12 + 1, today.day)
    except ValueError:
        valid = today + timedelta(days=90)

    base_folder = _find_round_folder_bqms(rfq_number, round_n)
    attachment_paths: list[str] = []
    if base_folder and base_folder.exists():
        skip_cam_ket = any(
            kw in (rfq["maker"] or "").lower()
            for kw in ["samsung", "sec ", "samsung electro"]
        )
        if classification == "TM":
            for f in sorted(base_folder.iterdir()):
                if not f.is_file():
                    continue
                ln = f.name.lower()
                if ln.endswith(".pdf"):
                    is_cam_ket = any(k in ln for k in ["cam_ket", "cam-ket", "camket", "commit"])
                    if is_cam_ket and skip_cam_ket:
                        continue
                    attachment_paths.append(str(f))
        elif classification == "GC":
            for f in sorted(base_folder.iterdir()):
                if not f.is_file():
                    continue
                ln = f.name.lower()
                if ln.endswith((".xlsx", ".xls", ".pdf")):
                    attachment_paths.append(str(f))
    else:
        warnings.append(f"Chưa có folder L{round_n} — anh quote V{round_n} trong ERP trước")
    if not attachment_paths:
        warnings.append(f"Không tìm thấy file đính kèm trong folder L{round_n}")

    return {
        "data": {
            "rfq_id": rfq_id,
            "rfq_number": rfq_number,
            "classification": classification,
            "round": round_n,
            "items": items_out,
            "submission_opinion": opinion,
            "quote_valid_date": valid.isoformat(),
            "attachment_paths": attachment_paths,
            "warnings": warnings,
        }
    }


def _find_round_folder_bqms(rfq_number: str, round_n: int):
    """Locate /data/onedrive-staging/.../{QT}<sep><meta>/QT_AMABACNINH_L{round}/.

    Folder-name compat:
      - Bare:   `{rfq}` (legacy)
      - OLD:    `{rfq}_<first_item>_<qty>_<date>_<time>` (underscore-joined, pre-2026-05-19)
      - NEW:    `{rfq} {qty_total} {date} {time}`        (SPACE-joined, since 2026-05-19)

    Round-subfolder compat:
      - NEW: `{rfq}_AMABACNINH_L{n}`   (no spaces inside AMABACNINH, since 2026-05-19)
      - OLD: `{rfq}_AMA BAC NINH_L{n}` (with spaces, before 2026-05-19)

    Critical bug fix (Thang 2026-05-20): previously rejected folders that don't
    start with `{rfq}_` — but the 2026-05-19 NEW pretty-name uses `{rfq} ` (space).
    So all newly-generated quote folders were invisible to push-preview.
    """
    from pathlib import Path as _Path
    from datetime import datetime as _dt
    root = _Path("/data/onedrive-staging/Puplic/BQMS/RFQ")
    now = _dt.now()
    patterns = [
        f"{rfq_number}_AMABACNINH_L{round_n}",      # NEW
        f"{rfq_number}_AMA BAC NINH_L{round_n}",    # OLD (backward compat)
    ]
    for y in [now.year, now.year - 1]:
        year_root = root / f"RFQ {y}"
        if not year_root.exists():
            continue
        for m in range(12, 0, -1):
            month_root = year_root / f"THANG {m}"
            if not month_root.exists():
                continue
            for d in month_root.iterdir():
                if not d.is_dir():
                    continue
                # Accept: bare match, underscore-prefix (old pretty), space-prefix (new pretty).
                if (d.name != rfq_number
                        and not d.name.startswith(f"{rfq_number}_")
                        and not d.name.startswith(f"{rfq_number} ")):
                    continue
                for pat in patterns:
                    round_sub = d / pat
                    if round_sub.exists():
                        return round_sub
    return None


class PushToSecRequest(BaseModel):
    items: list[dict[str, Any]]
    submission_opinion: str
    quote_valid_date: str
    attachment_paths: list[str]
    round: int = 1


@router.post("/rfq/{rfq_id}/push-to-sec")
@limiter.limit("5/minute")
async def push_to_sec(
    request: Request,
    response: Response,
    rfq_id: int,
    body: PushToSecRequest,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Dispatch Procrastinate task bqms_submit_quote.

    Thang 2026-05-23: dedupe theo `rfq_number` thay vì id đơn lẻ. 1 RFQ có thể
    có N items (mỗi item là 1 row bqms_rfq) — nếu chỉ check id, user click 9 lần
    cho 9 items sẽ tạo 9 jobs duplicate cho cùng QT26066620. Sửa: check trạng
    thái của BẤT KỲ row nào trong RFQ → 409 nếu có job đang chạy.
    """
    rfq = await conn.fetchrow(
        "SELECT id, rfq_number, bqms_push_status FROM bqms_rfq WHERE id = $1",
        rfq_id,
    )
    if not rfq:
        raise HTTPException(404)

    # Dedupe by rfq_number: nếu BẤT KỲ row nào của RFQ này đang queued/running → 409
    busy = await conn.fetchval(
        """
        SELECT bqms_push_status FROM bqms_rfq
        WHERE rfq_number = $1 AND bqms_push_status IN ('queued', 'running')
        LIMIT 1
        """,
        rfq["rfq_number"],
    )
    if busy:
        raise HTTPException(
            409,
            f"QT {rfq['rfq_number']} đang được đẩy (status={busy}). "
            f"Vào popup ở giữa màn hình bấm \"Hủy queue\" nếu muốn đẩy lại.",
        )

    errors: list[str] = []
    if not body.submission_opinion or not body.submission_opinion.strip():
        errors.append("Submission Opinion rỗng")
    if not body.attachment_paths:
        errors.append("Cần ít nhất 1 file đính kèm")
    # Image required ONLY on round 1 — rounds 2-4 reuse Samsung's stored image.
    image_required = (body.round == 1)
    logger.info(
        "push_to_sec.image_gate rfq=%s round=%d image_required=%s",
        rfq["rfq_number"], body.round, image_required,
    )
    for i, it in enumerate(body.items, 1):
        if it.get("abandonment", "N") == "Y":
            continue
        if image_required and not it.get("image_path"):
            errors.append(f"Item #{i} ({it.get('bqms_code')}) thiếu ảnh (V1 yêu cầu — V2+ không cần)")
        if not it.get("quotation_price") or float(it["quotation_price"]) <= 0:
            errors.append(f"Item #{i} ({it.get('bqms_code')}) thiếu giá")
    if errors:
        raise HTTPException(400, detail={"errors": errors})

    payload = {
        "rfq_number": rfq["rfq_number"],
        "round": body.round,
        "items": body.items,
        "submission_opinion": body.submission_opinion,
        "quote_valid_date": body.quote_valid_date,
        "attachment_paths": body.attachment_paths,
    }
    await conn.execute(
        """UPDATE bqms_rfq SET
            bqms_push_status='queued',
            bqms_push_payload=$1::jsonb,
            bqms_push_error=NULL,
            bqms_push_round_active=$2,
            bqms_push_step_index=0,
            bqms_push_started_at=NOW(),
            bqms_push_heartbeat_at=NOW()
           WHERE id=$3""",
        _json.dumps(payload), int(body.round), rfq_id,
    )

    from app.tasks.bqms_auto_submit import bqms_submit_quote_task
    from app.core.procrastinate_app import app as proc_app
    # FIX (Thang 2026-05-15): Procrastinate app chưa được open trong FastAPI lifespan
    # → cần wrap với open_async() như pattern ở /vendor-staging/quote-batch (line 4111).
    async with proc_app.open_async():
        job_id = await bqms_submit_quote_task.defer_async(
            rfq_id=rfq_id, payload=payload, user_id=str(token_data.user_id),
        )
    await conn.execute(
        "UPDATE bqms_rfq SET bqms_push_job_id=$1 WHERE id=$2",
        str(job_id), rfq_id,
    )
    position = await conn.fetchval(
        """SELECT COUNT(*) FROM procrastinate_jobs
           WHERE queue_name='bqms_push' AND status IN ('todo', 'doing')""",
    )
    logger.info("push_to_sec dispatched: rfq_id=%d job_id=%s pos=%s", rfq_id, job_id, position)
    return {
        "data": {
            "job_id": str(job_id),
            "queue_position": int(position or 1),
            "rfq_number": rfq["rfq_number"],
        },
        "message": f"Đã queue. Vị trí #{position}",
    }


class BatchPushToSecRequest(BaseModel):
    rfq_ids: list[int]
    round: int = 1


@router.post("/push-to-sec/batch")
@limiter.limit("3/minute")
async def push_to_sec_batch(
    request: Request,
    response: Response,
    body: BatchPushToSecRequest,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """ĐẨY NHIỀU mã lên SEC THEO THỨ TỰ trong 1 phiên (Thang 2026-06-29).

    Build payload SERVER-SIDE per RFQ qua get_push_preview (ĐÚNG logic modal đơn) +
    validate y hệt push đơn. Mã thiếu giá/ảnh/file → BỎ QUA (skip) kèm lý do, KHÔNG
    làm hỏng cả mẻ. Mã hợp lệ → set queued + enqueue 1 job `bqms_submit_batch` (giữ
    samsung_session_lock 1 lần, đẩy lần lượt). Tối đa 8 mã/mẻ. Dedup theo rfq_number.
    """
    rfq_ids = list(dict.fromkeys(int(x) for x in (body.rfq_ids or [])))
    if not rfq_ids:
        raise HTTPException(400, "Cần ít nhất 1 mã")
    if len(rfq_ids) > 8:
        raise HTTPException(400, "Tối đa 8 mã mỗi lần đẩy (giới hạn phiên Samsung) — chia nhỏ giúp em.")
    round_n = max(1, min(4, int(body.round or 1)))

    enqueued: list[dict] = []
    skipped: list[dict] = []
    seen_rfq_numbers: set[str] = set()
    batch_rfqs: list[dict] = []

    for rid in rfq_ids:
        rno = None
        try:
            rfq = await conn.fetchrow(
                "SELECT id, rfq_number, bqms_push_status FROM bqms_rfq WHERE id=$1", rid)
            if not rfq:
                skipped.append({"rfq_id": rid, "rfq_number": None, "errors": ["Không tồn tại"]})
                continue
            rno = rfq["rfq_number"]
            if rno in seen_rfq_numbers:
                continue  # cùng 1 RFQ chọn nhiều dòng (nhiều item) → chỉ xử lý 1 lần
            seen_rfq_numbers.add(rno)
            busy = await conn.fetchval(
                "SELECT bqms_push_status FROM bqms_rfq WHERE rfq_number=$1 "
                "AND bqms_push_status IN ('queued','running') LIMIT 1", rno)
            if busy:
                skipped.append({"rfq_id": rid, "rfq_number": rno, "errors": [f"Đang được đẩy (status={busy})"]})
                continue
            # Build payload bằng CHÍNH get_push_preview (không tự dựng lại → không lệch).
            preview = await get_push_preview(rfq_id=rid, round_n=round_n, token_data=token_data, conn=conn)
            data = preview["data"]
            items = data.get("items") or []
            opinion = data.get("submission_opinion") or ""
            attachments = data.get("attachment_paths") or []
            valid_date = data.get("quote_valid_date")
            # Validate y hệt push đơn (push_to_sec).
            errs: list[str] = []
            if not opinion.strip():
                errs.append("Thiếu Submission Opinion")
            if not attachments:
                errs.append(f"Không có file đính kèm (folder L{round_n})")
            image_required = (round_n == 1)
            for i, it in enumerate(items, 1):
                if it.get("abandonment", "N") == "Y":
                    continue
                if image_required and not it.get("image_path"):
                    errs.append(f"#{i} {it.get('bqms_code')}: thiếu ảnh (V1)")
                if not it.get("quotation_price") or float(it["quotation_price"]) <= 0:
                    errs.append(f"#{i} {it.get('bqms_code')}: thiếu giá")
            if errs:
                skipped.append({"rfq_id": rid, "rfq_number": rno, "errors": errs[:5]})
                continue
            payload = {
                "rfq_number": rno, "round": round_n, "items": items,
                "submission_opinion": opinion, "quote_valid_date": valid_date,
                "attachment_paths": attachments,
            }
            # Set queued CHỈ trên dòng rid (giống push đơn) — popup dedup theo rfq_number
            # sẽ chọn đúng dòng đang chạy; tránh kẹt các dòng item khác ở 'queued'.
            await conn.execute(
                """UPDATE bqms_rfq SET bqms_push_status='queued', bqms_push_payload=$1::jsonb,
                   bqms_push_error=NULL, bqms_push_round_active=$2, bqms_push_step_index=0,
                   bqms_push_progress_pct=0, bqms_push_progress_step='Chờ trong hàng đợi...',
                   bqms_push_started_at=NOW(), bqms_push_heartbeat_at=NOW() WHERE id=$3""",
                _json.dumps(payload), round_n, rid)
            batch_rfqs.append({"rfq_id": rid, "payload": payload})
            enqueued.append({"rfq_id": rid, "rfq_number": rno})
        except HTTPException as he:
            skipped.append({"rfq_id": rid, "rfq_number": rno, "errors": [str(he.detail)[:200]]})
        except Exception as exc:
            logger.warning("batch push prep failed rfq_id=%d: %s", rid, exc)
            skipped.append({"rfq_id": rid, "rfq_number": rno, "errors": [str(exc)[:200]]})

    if not batch_rfqs:
        # Tất cả bị bỏ qua → trả 200 (KHÔNG 400) để FE render ĐẦY ĐỦ lý do skip
        # (api.ts ép `detail` của lỗi 400 về string → mất danh sách chi tiết).
        return {
            "data": {"job_id": "", "enqueued": [], "skipped": skipped},
            "message": "Không có mã nào đủ điều kiện để đẩy",
        }

    from app.tasks.bqms_auto_submit import bqms_submit_batch_task
    from app.core.procrastinate_app import app as proc_app
    try:
        async with proc_app.open_async():
            job_id = await bqms_submit_batch_task.defer_async(
                rfqs=batch_rfqs, user_id=str(token_data.user_id))
    except Exception:
        # Enqueue lỗi → reset các dòng vừa set 'queued' về NULL để không kẹt popup.
        logger.exception("push_to_sec_batch defer_async failed")
        for e in enqueued:
            await conn.execute(
                "UPDATE bqms_rfq SET bqms_push_status=NULL, bqms_push_job_id=NULL WHERE id=$1",
                e["rfq_id"])
        raise HTTPException(500, "Không xếp được hàng đợi — vui lòng thử lại sau")
    for e in enqueued:
        await conn.execute(
            "UPDATE bqms_rfq SET bqms_push_job_id=$1 WHERE id=$2", str(job_id), e["rfq_id"])

    logger.info("push_to_sec_batch: enqueued=%d skipped=%d job=%s", len(enqueued), len(skipped), job_id)
    return {
        "data": {"job_id": str(job_id), "enqueued": enqueued, "skipped": skipped},
        "message": f"Đã xếp hàng {len(enqueued)} mã đẩy lần lượt"
        + (f" · bỏ qua {len(skipped)} mã" if skipped else ""),
    }


@router.get("/push-queue/status")
async def get_push_queue_status(
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Danh sách tất cả job queued + running + recent done.

    Thang 2026-05-23: DEDUPE by (rfq_number, bqms_pushed_round) — vì 1 RFQ có
    nhiều bqms_rfq rows (1 row per item) đều bị set status='queued' cùng lúc
    → user thấy 9 popup chồng nhau cho cùng 1 push. Giờ chỉ trả 1 row đại diện
    per (RFQ, round). Thêm cột item_count để hiển thị số mã trong RFQ.
    """
    rows = await conn.fetch(
        """
        WITH ranked AS (
            SELECT r.id, r.rfq_number, r.bqms_push_status, r.bqms_push_job_id,
                   r.bqms_push_error, r.bqms_pushed_at, r.bqms_pushed_round,
                   r.bqms_push_screenshot_path,
                   r.bqms_push_progress_pct, r.bqms_push_progress_step,
                   r.bqms_push_started_at,
                   r.bqms_push_round_active, r.bqms_push_step_index,
                   r.bqms_push_total_steps, r.bqms_push_step_key,
                   j.status AS job_status, j.scheduled_at, j.attempts,
                   ROW_NUMBER() OVER (
                       PARTITION BY r.rfq_number, COALESCE(r.bqms_push_round_active, r.bqms_pushed_round, 0)
                       ORDER BY
                           CASE r.bqms_push_status
                               WHEN 'running' THEN 1
                               WHEN 'queued' THEN 2
                               WHEN 'failed' THEN 3
                               ELSE 4
                           END,
                           r.id ASC
                   ) AS rn,
                   COUNT(*) OVER (PARTITION BY r.rfq_number, COALESCE(r.bqms_push_round_active, r.bqms_pushed_round, 0)) AS item_count
            FROM bqms_rfq r
            LEFT JOIN procrastinate_jobs j ON j.id::text = r.bqms_push_job_id
            WHERE r.bqms_push_status IN ('queued', 'running', 'failed', 'saved_temp')
              AND (
                    -- BUG FIX (Thang 2026-06-29): re-push 1 QT đã đẩy >24h trước thì
                    -- bqms_pushed_at là lần THÀNH CÔNG CŨ → điều kiện "trong 24h" loại
                    -- bỏ cả job mới đang chạy/failed → popup không hiện. Sửa: luôn hiện
                    -- job queued/running; với failed/saved_temp dùng started_at (lần ĐẨY
                    -- HIỆN TẠI), pushed_at chỉ là fallback cho rows cũ chưa có started_at.
                    r.bqms_push_status IN ('queued', 'running')
                 OR r.bqms_push_started_at > NOW() - INTERVAL '24 hours'
                 OR r.bqms_pushed_at      > NOW() - INTERVAL '24 hours'
                  )
        )
        SELECT * FROM ranked
        WHERE rn = 1
        ORDER BY
            CASE bqms_push_status
                WHEN 'running' THEN 1
                WHEN 'queued' THEN 2
                WHEN 'failed' THEN 3
                ELSE 4
            END,
            bqms_pushed_at DESC NULLS LAST
        LIMIT 20
        """,
    )
    def _ser_q(r):
        d = dict(r)
        for k, v in d.items():
            if isinstance(v, (date, datetime)):
                d[k] = v.isoformat()
        d.pop("rn", None)  # internal field
        return d
    return {"data": [_ser_q(r) for r in rows]}


@router.post("/push-queue/cancel/{rfq_number}")
async def cancel_push_queue(
    rfq_number: str,
    round_n: int | None = Query(None, ge=1, le=4, description="Round cụ thể; None=tất cả round"),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Hủy queued job + reset status về NULL.

    Chỉ hủy được status='queued' (chưa chạy). Job đang 'running' phải chờ.
    """
    conditions = ["rfq_number = $1", "bqms_push_status = 'queued'"]
    params: list[Any] = [rfq_number]
    if round_n is not None:
        conditions.append(f"bqms_pushed_round = ${len(params) + 1}")
        params.append(round_n)
    where = " AND ".join(conditions)

    affected = await conn.fetch(
        f"""
        UPDATE bqms_rfq
        SET bqms_push_status = NULL,
            bqms_push_job_id = NULL,
            bqms_push_error = NULL,
            bqms_push_payload = NULL,
            bqms_push_progress_pct = NULL,
            bqms_push_progress_step = NULL,
            bqms_push_started_at = NULL
        WHERE {where}
        RETURNING id, bqms_push_job_id
        """,
        *params,
    )
    # Try cancel procrastinate jobs too (best-effort)
    job_ids = [r["bqms_push_job_id"] for r in affected if r["bqms_push_job_id"]]
    if job_ids:
        await conn.execute(
            "UPDATE procrastinate_jobs SET status = 'cancelled' "
            "WHERE id::text = ANY($1::text[]) AND status = 'todo'",
            list(set(job_ids)),
        )
    logger.info("cancel push queue rfq=%s round=%s rows=%d", rfq_number, round_n, len(affected))
    return {
        "data": {
            "rfq_number": rfq_number,
            "round": round_n,
            "cancelled_rows": len(affected),
        },
        "message": f"Đã hủy {len(affected)} dòng trong queue",
    }


@router.get("/rfq/{rfq_id}/push-screenshot")
async def get_push_screenshot(
    rfq_id: int,
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Stream screenshot evidence."""
    from fastapi.responses import FileResponse
    from pathlib import Path as _Path
    row = await conn.fetchrow(
        "SELECT bqms_push_screenshot_path FROM bqms_rfq WHERE id = $1", rfq_id,
    )
    if not row or not row["bqms_push_screenshot_path"]:
        raise HTTPException(404, "Chưa có screenshot")
    p = _Path(row["bqms_push_screenshot_path"])
    if not p.exists():
        raise HTTPException(404, "Screenshot file đã bị xóa")
    return FileResponse(str(p), media_type="image/png")


@router.post("/rfq/{rfq_id}/push-preview/upload-image")
async def upload_override_image(
    rfq_id: int,
    bqms_code: str = Query(...),
    file: UploadFile = File(...),
    token_data: TokenData = Depends(require_role("admin", "manager", "staff", "sales")),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Admin upload ảnh override."""
    from pathlib import Path as _Path
    rfq = await conn.fetchrow("SELECT rfq_number FROM bqms_rfq WHERE id = $1", rfq_id)
    if not rfq:
        raise HTTPException(404)
    ext = _Path(file.filename or "").suffix.lower()
    if ext not in (".png", ".jpg", ".jpeg"):
        raise HTTPException(400, "Chỉ PNG/JPG")
    ovr_dir = _Path(f"/data/quote-overrides/{rfq['rfq_number']}")
    ovr_dir.mkdir(parents=True, exist_ok=True)
    out_path = ovr_dir / f"{bqms_code}__product_photo{ext}"
    contents = await file.read()
    out_path.write_bytes(contents)
    logger.info("Override image saved: %s (%d bytes)", out_path, len(contents))
    return {"data": {"path": str(out_path), "size": len(contents)}}


# ---------------------------------------------------------------------------
# Batch 2C — V-round / D-N round history (event-log timeline)
# ---------------------------------------------------------------------------
@router.get("/rfq/{rfq_number}/round-history")
async def rfq_round_history(
    rfq_number: str,
    token_data: TokenData = Depends(require_role(
        "admin", "manager", "staff", "sales", "procurement", "warehouse", "accountant"
    )),
    conn: asyncpg.Connection = Depends(get_db),
):
    """Lịch sử báo giá V1→V2→V3 + chuyển trạng thái cho 1 RFQ.

    Source of truth = bqms_qt_events (append-only). Returns the full timeline
    plus the current materialized state from bqms_rfq. Guarded: if the Batch-2C
    schema is not present yet, returns an empty timeline instead of 500.
    """
    has_events = bool(await conn.fetchval(
        "SELECT to_regclass('public.bqms_qt_events') IS NOT NULL"
    ))
    if not has_events:
        return {
            "data": {
                "rfq_number": rfq_number,
                "current_state": None,
                "deadline_dt": None,
                "current_round": None,
                "events": [],
            },
            "message": "Chưa kích hoạt theo dõi V-round (migration chưa chạy)",
        }

    # Current materialized state (may be NULL if columns not yet present — guard).
    has_state_cols = bool(await conn.fetchval(
        "SELECT EXISTS (SELECT 1 FROM information_schema.columns "
        "WHERE table_name='bqms_rfq' AND column_name='qt_state')"
    ))
    cur = None
    if has_state_cols:
        cur = await conn.fetchrow(
            """
            SELECT qt_state::text AS qt_state, deadline_dt, current_round,
                   version AS samsung_round, reinvited_at, state_changed_at
              FROM bqms_rfq
             WHERE rfq_number = $1
             ORDER BY updated_at DESC NULLS LAST, id DESC
             LIMIT 1
            """,
            rfq_number,
        )

    rows = await conn.fetch(
        """
        SELECT id, bqms_code, event_type,
               from_state::text AS from_state, to_state::text AS to_state,
               round_no, deadline_dt, actor, evidence, created_at
          FROM bqms_qt_events
         WHERE rfq_number = $1
         ORDER BY created_at ASC, id ASC
        """,
        rfq_number,
    )

    def _iso(v):
        return v.isoformat() if isinstance(v, (date, datetime)) else v

    events = []
    for r in rows:
        ev = r["evidence"]
        if isinstance(ev, str):
            try:
                ev = _json.loads(ev)
            except Exception:
                ev = {}
        events.append({
            "id": int(r["id"]),
            "bqms_code": r["bqms_code"],
            "event_type": r["event_type"],
            "from_state": r["from_state"],
            "to_state": r["to_state"],
            "round_no": r["round_no"],
            "deadline_dt": _iso(r["deadline_dt"]),
            "actor": r["actor"],
            "evidence": ev,
            "created_at": _iso(r["created_at"]),
        })

    return {
        "data": {
            "rfq_number": rfq_number,
            "current_state": cur["qt_state"] if cur else None,
            "deadline_dt": _iso(cur["deadline_dt"]) if cur else None,
            "current_round": cur["current_round"] if cur else None,
            "samsung_round": cur["samsung_round"] if cur else None,
            "reinvited_at": _iso(cur["reinvited_at"]) if cur else None,
            "state_changed_at": _iso(cur["state_changed_at"]) if cur else None,
            "events": events,
        }
    }
