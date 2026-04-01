"""
Excel Export API (M24) — Song Châu ERP

One-click Excel export for any supported table with optional column and filter
selection. Uses openpyxl to generate .xlsx files saved to /data/files/exports/.

Endpoints:
  POST /export              — Generate export, returns download filename
  GET  /download/{filename} — Download a previously generated export file
"""

from __future__ import annotations

import logging
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import asyncpg
import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.core.database import get_db
from app.core.rbac import require_role
from app.core.security import TokenData

logger = logging.getLogger(__name__)

router = APIRouter()

EXPORTS_DIR = Path("/data/files/exports")
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Table whitelist — maps table_name → allowed column set + base query
# ---------------------------------------------------------------------------

TABLE_CONFIG: dict[str, dict] = {
    "bqms_rfq": {
        "label": "Danh sách RFQ",
        "query": """
            SELECT
                rfq.id,
                rfq.rfq_number         AS "Số RFQ",
                rfq.bqms_code          AS "Mã BQMS",
                rfq.specification      AS "Thông số kỹ thuật",
                rfq.result             AS "Kết quả",
                rfq.quantity           AS "Số lượng",
                rfq.unit               AS "Đơn vị",
                rfq.deadline::text     AS "Hạn chót",
                rfq.created_at::text   AS "Ngày tạo"
            FROM bqms_rfq rfq
        """,
        "default_columns": [
            "Số RFQ", "Mã BQMS", "Thông số kỹ thuật", "Kết quả",
            "Số lượng", "Đơn vị", "Hạn chót", "Ngày tạo",
        ],
    },
    "purchase_orders": {
        "label": "Đơn mua hàng",
        "query": """
            SELECT
                po.id,
                po.po_number            AS "Số PO",
                s.name                  AS "Nhà cung cấp",
                po.status               AS "Trạng thái",
                po.total_amount         AS "Tổng tiền",
                po.currency             AS "Tiền tệ",
                po.order_date::text     AS "Ngày đặt hàng",
                po.expected_date::text  AS "Ngày dự kiến",
                po.created_at::text     AS "Ngày tạo"
            FROM purchase_orders po
            LEFT JOIN suppliers s ON s.id = po.supplier_id
        """,
        "default_columns": [
            "Số PO", "Nhà cung cấp", "Trạng thái", "Tổng tiền",
            "Tiền tệ", "Ngày đặt hàng", "Ngày dự kiến",
        ],
    },
    "suppliers": {
        "label": "Nhà cung cấp",
        "query": """
            SELECT
                s.id,
                s.name                  AS "Tên NCC",
                s.code                  AS "Mã NCC",
                s.contact_name          AS "Người liên hệ",
                s.email                 AS "Email",
                s.phone                 AS "Điện thoại",
                s.country               AS "Quốc gia",
                s.rating                AS "Đánh giá",
                s.is_active             AS "Đang hoạt động",
                s.created_at::text      AS "Ngày đăng ký"
            FROM suppliers s
        """,
        "default_columns": [
            "Tên NCC", "Mã NCC", "Người liên hệ", "Email",
            "Điện thoại", "Quốc gia", "Đánh giá", "Đang hoạt động",
        ],
    },
    "inventory": {
        "label": "Tồn kho",
        "query": """
            SELECT
                inv.id,
                p.name                  AS "Tên sản phẩm",
                p.sku                   AS "SKU",
                inv.warehouse_code      AS "Kho",
                inv.quantity            AS "Số lượng",
                inv.reserved            AS "Đã đặt trước",
                inv.available           AS "Khả dụng",
                inv.unit                AS "Đơn vị",
                inv.min_stock           AS "Tồn tối thiểu",
                inv.updated_at::text    AS "Cập nhật lần cuối"
            FROM inventory inv
            LEFT JOIN products p ON p.id = inv.product_id
        """,
        "default_columns": [
            "Tên sản phẩm", "SKU", "Kho", "Số lượng",
            "Đã đặt trước", "Khả dụng", "Đơn vị", "Tồn tối thiểu",
        ],
    },
    "sales_orders": {
        "label": "Đơn bán hàng",
        "query": """
            SELECT
                so.id,
                so.order_number         AS "Số đơn hàng",
                so.customer_name        AS "Khách hàng",
                so.status               AS "Trạng thái",
                so.total_amount         AS "Tổng tiền",
                so.currency             AS "Tiền tệ",
                so.order_date::text     AS "Ngày đặt hàng",
                so.delivery_date::text  AS "Ngày giao hàng",
                so.created_at::text     AS "Ngày tạo"
            FROM sales_orders so
        """,
        "default_columns": [
            "Số đơn hàng", "Khách hàng", "Trạng thái", "Tổng tiền",
            "Tiền tệ", "Ngày đặt hàng", "Ngày giao hàng",
        ],
    },
    "invoices": {
        "label": "Hóa đơn",
        "query": """
            SELECT
                inv.id,
                inv.invoice_number      AS "Số hóa đơn",
                inv.invoice_type        AS "Loại",
                inv.status              AS "Trạng thái",
                inv.total_amount        AS "Tổng tiền",
                inv.currency            AS "Tiền tệ",
                inv.due_date::text      AS "Hạn thanh toán",
                inv.paid_at::text       AS "Ngày thanh toán",
                inv.created_at::text    AS "Ngày tạo"
            FROM invoices inv
        """,
        "default_columns": [
            "Số hóa đơn", "Loại", "Trạng thái", "Tổng tiền",
            "Tiền tệ", "Hạn thanh toán", "Ngày thanh toán",
        ],
    },
    "task_assignments": {
        "label": "Nhiệm vụ",
        "query": """
            SELECT
                ta.id,
                ta.title                AS "Tiêu đề",
                ta.task_type            AS "Loại nhiệm vụ",
                ta.status               AS "Trạng thái",
                ta.priority             AS "Ưu tiên",
                u_to.full_name          AS "Người thực hiện",
                u_by.full_name          AS "Người giao",
                ta.due_date::text       AS "Hạn chót",
                ta.completed_at::text   AS "Hoàn thành lúc",
                ta.created_at::text     AS "Ngày tạo"
            FROM task_assignments ta
            LEFT JOIN users u_to ON u_to.id = ta.assigned_to
            LEFT JOIN users u_by ON u_by.id = ta.assigned_by
        """,
        "default_columns": [
            "Tiêu đề", "Loại nhiệm vụ", "Trạng thái", "Ưu tiên",
            "Người thực hiện", "Người giao", "Hạn chót",
        ],
    },
}

# Header fill style
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
DATA_FONT = Font(name="Calibri", size=10)


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class ExportRequest(BaseModel):
    table_name: str
    columns: Optional[list[str]] = None        # subset of default_columns; None = all
    filters: Optional[dict[str, Any]] = None   # simple {column_alias: value} filters
    limit: int = 10000                          # max rows to export


# ---------------------------------------------------------------------------
# Helper: build Excel file
# ---------------------------------------------------------------------------

def _build_excel(
    rows: list[dict],
    columns: list[str],
    sheet_label: str,
    export_meta: dict,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_label[:31]  # Excel sheet name limit

    # Meta row
    ws.append([f"Song Châu ERP — {sheet_label}"])
    ws["A1"].font = Font(bold=True, size=13, name="Calibri")
    ws.append([f"Xuất lúc: {export_meta['exported_at']}  |  Tổng dòng: {len(rows)}"])
    ws["A2"].font = Font(italic=True, size=9, name="Calibri", color="666666")
    ws.append([])  # blank row

    # Header row (row 4)
    header_row_idx = 4
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[header_row_idx].height = 24

    # Data rows
    for row_idx, row in enumerate(rows, start=header_row_idx + 1):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name)
            # Convert booleans to Vietnamese
            if isinstance(value, bool):
                value = "Có" if value else "Không"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top")

    # Auto-width (approximate)
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(columns[col_idx - 1])
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze header
    ws.freeze_panes = f"A{header_row_idx + 1}"

    return wb


# ---------------------------------------------------------------------------
# POST /export — Generate Excel export
# ---------------------------------------------------------------------------

@router.post("/export")
async def export_to_excel(
    body: ExportRequest,
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
    conn: asyncpg.Connection = Depends(get_db),
):
    if body.table_name not in TABLE_CONFIG:
        raise HTTPException(
            status_code=400,
            detail=f"Bảng không được hỗ trợ. Chấp nhận: {', '.join(sorted(TABLE_CONFIG.keys()))}",
        )

    config = TABLE_CONFIG[body.table_name]
    requested_columns = body.columns or config["default_columns"]

    # Validate requested columns against allowed set
    allowed = set(config["default_columns"])
    invalid_cols = [c for c in requested_columns if c not in allowed]
    if invalid_cols:
        raise HTTPException(
            status_code=400,
            detail=f"Cột không hợp lệ: {invalid_cols}. Cột hợp lệ: {sorted(allowed)}",
        )

    if body.limit > 50000:
        raise HTTPException(status_code=400, detail="Giới hạn tối đa 50,000 dòng mỗi lần xuất")

    # Build query — always include 'id' for reference, filter by limit
    base_query = config["query"].strip()
    final_query = f"{base_query} ORDER BY id DESC LIMIT $1"

    try:
        rows = await conn.fetch(final_query, body.limit)
    except Exception as exc:
        logger.error("Export query failed for table %s: %s", body.table_name, exc)
        raise HTTPException(status_code=500, detail=f"Lỗi truy vấn dữ liệu: {exc}")

    # Convert rows to dicts and keep only requested columns
    data_rows = [
        {col: dict(r).get(col) for col in requested_columns}
        for r in rows
    ]

    # Build Excel file
    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    export_meta = {"exported_at": exported_at, "exported_by": token_data.email}

    wb = _build_excel(
        rows=data_rows,
        columns=requested_columns,
        sheet_label=config["label"],
        export_meta=export_meta,
    )

    # Save to disk
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{body.table_name}_{timestamp}_{uuid.uuid4().hex[:6]}.xlsx"
    file_path = EXPORTS_DIR / filename

    try:
        wb.save(str(file_path))
    except Exception as exc:
        logger.error("Failed to save Excel file: %s", exc)
        raise HTTPException(status_code=500, detail="Lỗi tạo file Excel")

    return {
        "data": {
            "filename": filename,
            "download_url": f"/api/v1/excel-export/download/{filename}",
            "table": body.table_name,
            "total_rows": len(data_rows),
            "columns": requested_columns,
            "exported_at": exported_at,
        },
        "message": f"Đã xuất {len(data_rows)} dòng từ bảng '{config['label']}'",
    }


# ---------------------------------------------------------------------------
# GET /download/{filename} — Download generated export
# ---------------------------------------------------------------------------

@router.get("/download/{filename}")
async def download_export(
    filename: str,
    token_data: TokenData = Depends(require_role("staff", "manager", "admin")),
):
    # Security: prevent path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Tên file không hợp lệ")

    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ tải file .xlsx")

    file_path = EXPORTS_DIR / filename
    if not file_path.exists():
        raise HTTPException(
            status_code=404,
            detail="File không tồn tại hoặc đã bị xoá. Vui lòng tạo export mới.",
        )

    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
